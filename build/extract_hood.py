# -*- coding: utf-8 -*-
"""把 Robinhood 官方 Monthly Metrics Excel 拆成本项目用的 CSV。

数据源（唯一）：investors.robinhood.com → Financials → Monthly Metrics → "Monthly Metrics Excel"。
每月中旬发布上一月，四种格式同时发（Press Release / PDF / Excel / Dashboard），本脚本只吃 Excel。
注意：HOOD 的月度数据**不走 8-K**，是按 Reg FD 挂在 IR 网站上的，盯 EDGAR 抓不到。

输出:
    data/hood.csv     —— 月度 KPI（Monthly KPIs 表）
    data/hood_q.csv   —— 季度实际收入与成交量（Quarterly GAAP P&L + Quarterly KPIs 两表）

季度那份是给「隐含收入 vs 实际收入」验证图用的：月度只有成交量没有收入，
收入按季度披露，所以费率只能是季度值 —— 这一层假设必须写进图注。

用法:
    python3 extract_hood.py <path-to-xlsx>
"""
import os
import sys

import openpyxl
import pandas as pd

D = os.path.dirname(os.path.abspath(__file__))
MON = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def grid(ws):
    """把 sheet 读成定宽二维表（openpyxl 的行长度不齐，不补齐会 IndexError）。"""
    W = ws.max_column
    return [list(r) + [None] * (W - len(r)) for r in ws.iter_rows(values_only=True)], W


def find_header(rows, W, kind):
    """年份行/期次行的位置在各 sheet 不一致（月度是 2/3，季度 P&L 是 3/4），所以扫出来而不是写死。"""
    want = MON if kind == 'M' else ['Q1', 'Q2', 'Q3', 'Q4']
    for ri in range(min(8, len(rows))):
        hits = sum(1 for i in range(W)
                   if rows[ri][i] and str(rows[ri][i]).strip()[:3] in want)
        if hits >= 4:
            return ri
    raise SystemExit(f'找不到{kind}表头')


def periods(rows, W, kind):
    pr = find_header(rows, W, kind)
    yr = rows[pr - 1]
    out, cur = [], None
    for i in range(W):
        v = yr[i]
        if v is not None and str(v).replace(',', '').strip().isdigit():
            cur = int(str(v).replace(',', '').strip())
        s = str(rows[pr][i]).strip()[:3] if rows[pr][i] else ''
        if kind == 'M' and s in MON:
            out.append((i, pd.Period(f'{cur}-{MON.index(s) + 1:02d}', 'M')))
        elif kind == 'Q' and s in ('Q1', 'Q2', 'Q3', 'Q4'):
            out.append((i, pd.Period(f'{cur}Q{s[1]}', 'Q')))
    return out


def pull(rows, cols, spec):
    """spec: {列名: excel行号}。用行号而不是行标签 —— 表里 'Equity ($B)' 这类标签重复出现
    （总成交量一次、日均一次），按标签取会静默拿错行。"""
    out = {}
    for name, ri in spec.items():
        out[name] = [rows[ri][i] if isinstance(rows[ri][i], (int, float)) else None
                     for i, _ in cols]
    return pd.DataFrame(out, index=[p for _, p in cols])


def check_labels(rows, spec, sheet):
    """行号是硬编码的，官方一改表结构就会错位。开跑前把标签打出来自检。"""
    print(f'  [{sheet}] 行号 → 标签自检')
    for name, ri in spec.items():
        lab = next((str(c).strip() for c in rows[ri][:3] if c is not None and str(c).strip()), '(空)')
        print(f'     r{ri:>2}  {name:<28} ← {lab[:46]}')


MONTHLY = {
    'funded_customers_mn': 6, 'total_platform_assets_usdbn': 9, 'net_deposits_usdbn': 10,
    'eqopt_trading_days': 13, 'crypto_trading_days': 14,
    'vol_equity_usdbn': 17, 'vol_options_mn': 18, 'vol_crypto_usdbn': 19,
    'vol_crypto_app_usdbn': 20, 'vol_crypto_bitstamp_usdbn': 21, 'vol_event_bn': 22,
    'adv_equity_usdbn': 25, 'adv_options_mn': 26, 'adv_crypto_usdmn': 27,
    'adv_crypto_app_usdmn': 28, 'adv_crypto_bitstamp_usdmn': 29, 'adv_event_mn': 30,
    'dats_equity_mn': 33, 'dats_options_mn': 34, 'dats_crypto_mn': 35,
    'margin_book_usdbn': 38, 'cash_and_deposits_usdbn': 39, 'cash_sweep_usdbn': 40,
    'seclend_total_usdmn': 43, 'seclend_net_usdmn': 44,
}

PL = {
    'rev_options_usdmn': 6, 'rev_crypto_usdmn': 7, 'rev_equities_usdmn': 8,
    'rev_event_usdmn': 9, 'rev_other_txn_usdmn': 10, 'rev_transaction_usdmn': 11,
    'rev_net_interest_usdmn': 12, 'rev_other_usdmn': 13, 'rev_total_usdmn': 14,
}

# Quarterly KPIs 表：只取成交量（余额类是「季末月」值不是季度合计，混用会错）
QVOL = {'q_vol_equity_usdbn': 20, 'q_vol_options_mn': 21,
        'q_vol_crypto_usdbn': 22, 'q_vol_event_bn': 25}


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else None
    if not src or not os.path.exists(src):
        raise SystemExit('用法: python3 extract_hood.py <hood_monthly_metrics.xlsx>')
    wb = openpyxl.load_workbook(src, data_only=True)

    ws = wb['Monthly KPIs']
    rows, W = grid(ws)
    check_labels(rows, MONTHLY, 'Monthly KPIs')
    m = pull(rows, periods(rows, W, 'M'), MONTHLY)

    pl = next(s for s in wb.sheetnames if s.startswith('Quarterly GAAP'))
    rows_q, Wq = grid(wb[pl])
    check_labels(rows_q, PL, pl)
    q = pull(rows_q, periods(rows_q, Wq, 'Q'), PL)

    # 季度成交量：由月度加总而来（官方季度表里也有，但月度加总能顺带验证两表自洽）
    rows_k, Wk = grid(wb['Quarterly KPIs'])
    check_labels(rows_k, QVOL, 'Quarterly KPIs')
    kq = pull(rows_k, periods(rows_k, Wk, 'Q'), QVOL)
    q = q.join(kq, how='outer')

    os.makedirs(os.path.join(D, 'data'), exist_ok=True)
    m.index.name = 'month'
    q.index.name = 'quarter'
    m.to_csv(os.path.join(D, 'data', 'hood.csv'))
    q.to_csv(os.path.join(D, 'data', 'hood_q.csv'))

    print(f'\n月度 {len(m)} 个月  {m.index[0]} → {m.index[-1]}   列 {len(m.columns)}')
    print(f'季度 {len(q)} 个季度 {q.index[0]} → {q.index[-1]}   列 {len(q.columns)}')

    # 自洽性检查：月度成交量加总 vs 官方季度成交量。对不上说明行号取错了。
    print('\n月度加总 vs 官方季度成交量（应当一致）')
    agg = m.groupby(m.index.asfreq('Q')).sum(min_count=3)
    for mc, qc, nm in [('vol_equity_usdbn', 'q_vol_equity_usdbn', 'Equity $bn'),
                       ('vol_options_mn', 'q_vol_options_mn', 'Options 合约 mn'),
                       ('vol_crypto_usdbn', 'q_vol_crypto_usdbn', 'Crypto $bn')]:
        j = pd.DataFrame({'月度加总': agg[mc], '官方季度': q[qc]}).dropna()
        err = (j['月度加总'] / j['官方季度'] - 1).abs().max() * 100
        print(f'  {nm:<18} 共 {len(j)} 个季度，最大偏差 {err:.2f}%  {"OK" if err < 1.0 else "← 有问题"}')


if __name__ == '__main__':
    main()
