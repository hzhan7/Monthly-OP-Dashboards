# -*- coding: utf-8 -*-
"""ICE_BRENT_IFEU 的基期常数 —— 一次性实测 + 可复算的留痕。

    用法
        python3 build/basefill/ice_brent.py            # 只抓、只算、只打印（不写表）
        python3 build/basefill/ice_brent.py --write    # 再把 ICE_BRENT_IFEU 那一行写进
                                                       # series/contract_specs.csv（幂等）

════════════════════════════════════════════════════════════════════════════
一、这个脚本存在的理由：把「口径不一致」换成「覆盖率低但零偏差」
════════════════════════════════════════════════════════════════════════════
原先 energy 池的 ICE 腿挂的是 `series/ice.csv` 的 `adv_energy_kcontracts`，
配一个叫 `ICE_ENERGY` 的**跨币种合成篮子**产品。那条路已经被否掉，理由不是"太难"，
是**分子分母不同源**：

  · `adv_energy_kcontracts` 是 ICE **全球**口径（Derivs sheet 的 TOTAL ENERGY，
    官方脚注明说 Nat Gas / Power 含 North American、NGX、UK 与 European，
    即 IFEU + ICE Endex + ICE Futures U.S. + ICE Futures Abu Dhabi + NGX 五处）；
  · 而 ICE 唯一公开的、不要 reCAPTCHA 的分产品历史表 https://www.ice.com/report/7
    **只覆盖 ICE Futures Europe 一个所**，2019-01 各能源列相加只占全球的 67.0%；
  · TTF 在 ICE Endex 上，根本不在那张表里，而且它**没有固定乘数**
    （官方页原文 "1 MW per day in contract period ... x 23, 24 or 25 hours"，
    一张月度合约 672–744 MWh 不等）。

拿 67% 子集算出来的篮子常数去套 100% 的量，等于给整条 ICE 能源线安一个
**方向与大小都不知道**的系统性偏差。图上完全看不出来 —— 柱子只是整体高一截或矮一截。

本脚本走的是另一条路：**缩小产品范围到口径一致的子集**。
Brent 这一列的官方脚注（见下）保证整列以「ICE Futures Europe 标准合约当量」计，
于是一个乘数（1,000 桶）对全列成立，不需要任何权重、不存在覆盖率缺口、零偏差。
代价是这条腿只代表 ICE 能源的三分之一 —— 这个代价必须写在页面上，
而不是靠一个不知道错多少的常数把它盖住。

════════════════════════════════════════════════════════════════════════════
二、四条实测证据（每条都在本脚本里现抓现验，不引用任何二手结论）
════════════════════════════════════════════════════════════════════════════
【① 量列的口径 —— 官方脚注原文】
  cache/ice_monthly_stats.xlsx → sheet 'Derivs (ADV, RPC, OI)' → A42：
      "Brent" includes the standard size contracts at ICE Futures Europe and as of
      November 2015 also includes mini Brent contracts on ICE Futures Singapore,
      which are converted to standard ICE Futures Europe equivalent contracts
      (mini Brent contracts are divided by 10).
  ⇒ 迷你合约**已经被官方折成 IFEU 标准当量**，所以「张数 × 1,000 桶」对全列精确成立。
    口径一致靠的是这条换算规则，不是"两个所恰好是同一个所"。
  同一张表的其余能源行，脚注逐条读过，没有一行有这个性质：
      Other Oil (3) = WTI + Midland WTI + Murban + Platts Dubai + RBOB + Heating Oil
                      + NGX Oil + Wet Freight + mini WTI  → 桶 / 加仑 / 运费三种量纲混在一行
      Nat Gas  (4) = North American + NGX + UK + European → MMBtu / therm / MWh
      Power    (5) = North American + NGX + UK + European → MWh，各国交割结构不同
      Environmentals (6) = 全部排放权 + 煤 + 铁矿石       → 吨 CO2 / 吨煤 / 干吨矿
  ⇒ **口径一致的子集 = Brent + Gasoil**，而 Gasoil 卡在价格（见 §三）⇒ 本轮只做 Brent。

【② 乘数】https://www.ice.com/products/219/Brent-Crude-Futures 现抓现验，
  Market Specifications 表原文必须同时出现 "Contract Symbol B"、
  "Contract Size 1,000 barrels"、"Currency US Dollars and cents"，缺一个就炸。

【③ 「这一列确实就是 IFEU」的独立验证】
  https://www.ice.com/report/7（ICE Futures Europe Historical Monthly Volumes，
  服务端渲染、纯标准库可读、1995 年至今 32 张年表）把 Brent 期货列 + Brent 期权列
  逐月相加，与 series/ice.csv 的 `adv_brent_kcontracts × trading_days_commod` 逐月对：
  187 个月（2011-01 – 2026-07）**没有一个月差过 0.5%**，中位 0.031%、最大 0.135%。
  残差的两个来源都不影响乘数：(a) ADV 单元格四舍五入到整千张；
  (b) 已折成 IFEU 当量的 ICE Futures Singapore 迷你合约（report/7 不含它）。
  这一条是**证伪性**的：如果 adv_brent 里混进了别的所的、别的合约规格的量，
  这个偏差会是百分之几而不是万分之几。

【④ 基期价格】美国能源信息署 EIA 官方序列 RBRTE
  "Europe Brent Spot Price FOB (Dollars per Barrel)"，
  文件 https://www.eia.gov/dnav/pet/hist_xls/RBRTEd.xls，取 2019-01 全部报价日的算术平均
  （与全表口径 `base_price_basis=avg_close` 一致：月内逐日的算术平均，不是月末收盘）。
  逐日值落库 cache/basefill/ice_brent_eia_2019_01.csv，任何人可逐位复核。

  ⚠ **必须知情的一条基差**：RBRTE 是 **Dated Brent 现货 FOB 评估价**，
    不是 ICE Brent 期货结算价 —— EIA 全站没有 Brent 期货序列
    （https://www.eia.gov/dnav/pet/pet_pri_spt_s1_d.htm 实测：原油只有
    "WTI - Cushing" 与 "Brent - Europe" 两条现货，成品油一律 $/gallon）。
    同月可比的期货口径参照：本仓 build/basefill/cme2.py 从 CME 官方 SPAN 文件解出的
    NYMEX Brent 现金结算期货（BZ，同样 1,000 桶/张）2019-01 近月月均结算价 ≈ 60.122
    美元/桶，比 EIA 现货高约 +1.2%，方向与 2019-01 初 Brent 曲线的 contango 一致。
    这 1.2% 是**已知方向、已知大小**的期货/现货基差，落在本表已经明示并接受的
    基准混用带内（avg_close vs basket_vw 实测差 0.4%~4%，见 build/check_specs.py
    模块 docstring）。它与被否掉的「67% 子集常数套 100% 全口径量」不是一回事 ——
    后者的偏差方向与大小都不可知。
    为什么不直接用 60.122：那个数是上一轮留在 series/_specs_part_cme2.csv 里的
    **四舍五入到整数**的单张名义额反推出来的，本会话没有重跑 SPAN 解码，
    照抄它等于写一个自己没测过的数。EIA 这条是本会话逐日抓下来、可复算的。

════════════════════════════════════════════════════════════════════════════
三、Gasoil 为什么不做（准确优先于覆盖）
════════════════════════════════════════════════════════════════════════════
乘数有：https://www.ice.com/products/34361119/Low-Sulphur-Gasoil-Futures 现抓现验
（本脚本一并验，原文 "Contract Symbol G" + "Contract Size 100 metric tonnes"）。
卡的是价格：Gasoil 以 **美元/公吨**（ARA 驳船）报价，而

  · EIA 全站没有这个口径 —— 现货价页自己的表头就写着
    "(Crude Oil in Dollars per Barrel, Products in Dollars per Gallon)"，
    成品油全部是美国口岸（NY Harbor / US Gulf Coast / LA），没有 Rotterdam / ARA 任何一条；
  · ICE 自己的历史结算价在 Report Center 的 reCAPTCHA 后面
    （build/basefill/ice_enx2.py 实测：report 26/27 的 metadata
     recaptchaRequired=true、criteria 接口恒 HTTP 409）。

拿 $/gallon 的美国柴油乘一个密度系数折成 $/tonne 就是编数，本仓不做。
加上 Gasoil 能把张数覆盖率从 34.8% 抬到 46.3%（2019-01），
但代价是一个来路不明的常数 —— **不做**。缺口留在 ICE_GASOIL_FUT 那行的 notes 里。

════════════════════════════════════════════════════════════════════════════
四、依赖
════════════════════════════════════════════════════════════════════════════
标准库 + openpyxl（读 cache/ice_monthly_stats.xlsx 的脚注）+ xlrd（读 EIA 的 .xls，
BIFF8 老格式，openpyxl 读不了）。xlrd **不在 requirements.txt 里**，与
build/basefill/cme2.py 同样是「一次性基期填数」这条路专用，不进每月无人值守链路 ——
所以在函数里 import，缺了不影响任何月度任务。不需要浏览器、不需要 curl_cffi。
"""

import argparse
import csv
import io
import os
import re
import sys
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SERIES = os.path.join(ROOT, 'series')
CACHE = os.path.join(ROOT, 'cache', 'basefill')

BASE_MONTH = '2019-01'
PRODUCT_ID = 'ICE_BRENT_IFEU'

UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
      '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

URL_BRENT_SPEC = 'https://www.ice.com/products/219/Brent-Crude-Futures'
URL_GASOIL_SPEC = 'https://www.ice.com/products/34361119/Low-Sulphur-Gasoil-Futures'
URL_REPORT7 = 'https://www.ice.com/report/7'
URL_EIA_BRENT = 'https://www.eia.gov/dnav/pet/hist_xls/RBRTEd.xls'
URL_EIA_SPOTLIST = 'https://www.eia.gov/dnav/pet/pet_pri_spt_s1_d.htm'

MONTHS = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
          'August', 'September', 'October', 'November', 'December']

# report/7 与本仓的偏差上界。定的是**实测上界留一档余量**，不是拍脑袋：
# 实测 187 个月 max 0.135%，取 0.5% ⇒ 真出现口径漂移（换了所、换了合约规格）
# 一定是百分之几的量级，一定拦得住；而四舍五入噪声一定拦不了。
REPORT7_TOL = 0.005


class IceBrentError(RuntimeError):
    """源站结构变了 / 实测值对不上 —— 一律炸掉，绝不静默写一个凑出来的常数。"""


# ══════════════════════════════════════════════════════════════════════
# 取数
# ══════════════════════════════════════════════════════════════════════
def _slug(url):
    return re.sub(r'[^A-Za-z0-9._-]+', '_', url)[-120:]


def _http(url, timeout=60):
    """抓一次并落盘到 cache/basefill/。落盘不是加速，是**留证据**：
    常数是一次性写死的，别人复核时必须能拿到与我同一份原始文件。"""
    os.makedirs(CACHE, exist_ok=True)
    req = urllib.request.Request(url, headers={'User-Agent': UA})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        if r.status != 200:
            raise IceBrentError('%s 返回 HTTP %s' % (url, r.status))
        body = r.read()
    with open(os.path.join(CACHE, _slug(url)), 'wb') as f:
        f.write(body)
    print('  [http] %-72s %s B' % (url[:72], format(len(body), ',')))
    return body


def _text(body):
    import html as _html
    t = _html.unescape(re.sub(r'<[^>]+>', ' ', body.decode('utf-8', 'replace')))
    return re.sub(r'\s+', ' ', t)


# ── ① 官方脚注：量列的口径 ────────────────────────────────────────────
FOOTNOTE_MUST = (
    'includes the standard size contracts at ICE Futures Europe',
    'converted to standard ICE Futures Europe equivalent contracts',
)


def read_footnotes():
    """从 cache/ice_monthly_stats.xlsx 里把 Brent / Gasoil 的脚注原文抠出来。

    不 grep 全表：同名标签在三个 section 里各出现一次（见 fetch/ice.py 口径坑 2），
    **脚注编号也是**：Derivs sheet 的 ADV / RPC / OI 三段各自从 (1) 重新编号
    （ADV 的 (1) 是 Brent 口径，RPC 的 (1) 是 "RPC is calculated by…"，
    OI 的 (1) 是 "Other Financials includes Equity Indices and FX"）。
    所以这里只取**首次出现**的那一份 —— ADV 段在最上面，正是要的那一段。
    照编号无脑覆盖会拿到 OI 段的 (1)，而它长得完全像个正常脚注。
    """
    import openpyxl
    path = os.path.join(ROOT, 'cache', 'ice_monthly_stats.xlsx')
    if not os.path.exists(path):
        raise IceBrentError(
            '缺 %s —— 先跑 python3 fetch/ice.py 把官方工作簿抓下来' % path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[[s for s in wb.sheetnames if s.lower().startswith('deriv')][0]]
    notes = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not isinstance(v, str):
            continue
        m = re.match(r'^\((\d+)\)\s*(.*)$', v.replace('\xa0', ' ').strip())
        if m and int(m.group(1)) not in notes:      # 只留首次出现，见 docstring
            notes[int(m.group(1))] = re.sub(r'\s+', ' ', m.group(2)).strip()
    if 1 not in notes:
        raise IceBrentError('Derivs sheet 里找不到脚注 (1)（Brent 的口径说明）')
    fn1 = notes[1]
    for must in FOOTNOTE_MUST:
        if must not in fn1:
            raise IceBrentError(
                '脚注 (1) 的原文变了，没找到 %r。整条口径论证建立在这句话上 —— '
                '它一改，"张数 × 1,000 桶对全列成立" 就不再有依据，必须人工重判。\n'
                '当前原文：%s' % (must, fn1))
    return notes


# ── ② 乘数：ICE 官方合约规格页 ────────────────────────────────────────
SPEC_MUST = {
    URL_BRENT_SPEC: ('Contract Symbol B', 'Contract Size 1,000 barrels',
                     'Currency US Dollars and cents'),
    URL_GASOIL_SPEC: ('Contract Symbol G', 'Contract Size 100 metric tonnes',
                      'Currency US Dollars and cents'),
}


def check_contract_specs():
    out = {}
    for url, musts in SPEC_MUST.items():
        txt = _text(_http(url))
        for must in musts:
            if must not in txt:
                raise IceBrentError('%s 上找不到 %r —— 规格页改版了，乘数必须人工重核'
                                    % (url, must))
        out[url] = musts
        print('    ✓ %s' % ' | '.join(musts))
    return out


# ── ③ report/7：证明 adv_brent 就是 IFEU 的 Brent ────────────────────
def parse_report7(body):
    """report/7 的 32 张年表 → {'YYYY-MM': {brent_fut, brent_opt, gasoil_fut, gasoil_opt}}。

    列位置随年份变（1995 年 8 列、2014 年起 19 列），所以**按表头文字认列**，
    不按位置数格子：'Brent' 在表头里恰好出现两次 —— 第一次在 Futures 段、
    第二次在 Options 段。'Gas Oil' 同理。
    """
    def strip(x):
        import html as _html
        return _html.unescape(re.sub(r'<[^>]+>', '', x)).replace('\xa0', ' ').strip()

    def num(x):
        x = x.replace(',', '').strip()
        return float(x) if re.match(r'^-?\d+(\.\d+)?$', x) else None

    parts = re.split(r'(<table.*?</table>)', body.decode('utf-8', 'replace'), flags=re.S)
    out = {}
    for i, p in enumerate(parts):
        if not p.startswith('<table'):
            continue
        yrs = re.findall(r'\b(?:19|20)\d\d\b', strip(parts[i - 1]))
        if not yrs:
            continue
        year = int(yrs[-1])
        rows = [[strip(c) for c in re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', r, re.S)]
                for r in re.findall(r'<tr.*?</tr>', p, re.S)]
        if len(rows) < 3:
            continue
        hdr = [re.sub(r'\s+', ' ', h) for h in rows[1]]
        bi = [j for j, h in enumerate(hdr) if h == 'Brent']
        gi = [j for j, h in enumerate(hdr) if h in ('Gas Oil', 'GasOil')]
        for row in rows[2:]:
            if not row or row[0] not in MONTHS:
                continue
            key = '%04d-%02d' % (year, MONTHS.index(row[0]) + 1)

            def pick(idxs, k):
                if len(idxs) <= k or idxs[k] >= len(row):
                    return None
                return num(row[idxs[k]])
            out[key] = {'brent_fut': pick(bi, 0), 'brent_opt': pick(bi, 1),
                        'gasoil_fut': pick(gi, 0), 'gasoil_opt': pick(gi, 1)}
    if len(out) < 300:
        raise IceBrentError('report/7 只解析出 %d 个月，表结构变了' % len(out))
    return out


def read_ice_csv():
    path = os.path.join(SERIES, 'ice.csv')
    if not os.path.exists(path):
        raise IceBrentError('缺 %s —— 先跑 python3 fetch/ice.py' % path)
    with open(path, newline='', encoding='utf-8') as f:
        return {r['month']: r for r in csv.DictReader(f)}


def crosscheck_report7(ifeu, ice):
    """逐月比 adv_brent × 交易日 与 report/7 的 IFEU Brent（期货+期权）。"""
    import statistics
    rows = []
    for m in sorted(ice):
        o = ifeu.get(m)
        if not o or o['brent_fut'] is None:
            continue
        r = ice[m]
        try:
            days = float(r['trading_days_commod'])
            brent = float(r['adv_brent_kcontracts'])
            gasoil = float(r['adv_gasoil_kcontracts'])
        except (TypeError, ValueError):
            continue
        ib = o['brent_fut'] + (o['brent_opt'] or 0)
        ig = (o['gasoil_fut'] or 0) + (o['gasoil_opt'] or 0)
        rows.append({
            'month': m,
            'brent_csv': brent * 1000 * days, 'brent_r7': ib,
            'brent_dev': (brent * 1000 * days - ib) / ib,
            'gasoil_csv': gasoil * 1000 * days, 'gasoil_r7': ig,
            'gasoil_dev': ((gasoil * 1000 * days - ig) / ig) if ig else None,
        })
    if len(rows) < 150:
        raise IceBrentError('只对上 %d 个月，太少，解析多半错了' % len(rows))
    bad = [x for x in rows if abs(x['brent_dev']) > REPORT7_TOL]
    if bad:
        raise IceBrentError(
            'adv_brent_kcontracts 与 ICE 官方 report/7 的 IFEU Brent 差超过 %.1f%%：%s。'
            '这条腿的全部合法性就是「该列 = IFEU 标准合约当量」，对不上就不能用它算名义额'
            % (REPORT7_TOL * 100,
               '、'.join('%s %+.2f%%' % (x['month'], x['brent_dev'] * 100) for x in bad[:6])))
    bd = sorted(abs(x['brent_dev']) for x in rows)
    gd = sorted(abs(x['gasoil_dev']) for x in rows if x['gasoil_dev'] is not None)
    wb_ = max(rows, key=lambda x: abs(x['brent_dev']))

    # ── 残差归因：把「差多少」拆成两个已知来源，而不是笼统说一句"很小" ────────
    # ADV 单元格四舍五入到整千张 ⇒ 单月最大可能影响 = 0.5 千张 × 交易日 × 1000。
    # 超出这个界的部分只能是 report/7 不覆盖的那一块，也就是脚注(1) 说的
    # ICE Futures Singapore 迷你 Brent（已被官方 ÷10 折成 IFEU 标准当量）。
    # 这是一条**证伪性**判据：若真是这个来源，那么
    #   (a) 迷你合约 2015-11 才开始并入 ⇒ 之前的月份必须**无一例外**落在四舍五入界内；
    #   (b) 超界月份的差必须**全部为正**（本仓多出来一块，report/7 少那一块）。
    # 只要有一条不成立，就说明这一列里混了别的东西，1,000 桶这个乘数就不能用。
    ROUND_STEP = 500.0                      # 千张的一半，单位：张
    cut = '2015-11'
    over = []
    for x in rows:
        # 交易日从原始 CSV 再取一次，不拿量反推 —— 反推等于用被检验的那个数去定检验的界
        x['round_bound'] = ROUND_STEP * float(ice[x['month']]['trading_days_commod'])
        x['over_round'] = abs(x['brent_csv'] - x['brent_r7']) > x['round_bound']
        if x['over_round']:
            over.append(x)
    pre = [x for x in rows if x['month'] < cut]
    pre_over = [x for x in pre if x['over_round']]
    neg_over = [x for x in over if x['brent_csv'] - x['brent_r7'] < 0]
    if pre_over:
        raise IceBrentError(
            '%s 之前有 %d 个月超出四舍五入界（最早 %s）—— 官方脚注说迷你 Brent 是 '
            '%s 才并入的，那么之前的差就无法归因，这一列里可能混了别的合约规格，'
            '1,000 桶的乘数不能用' % (cut, len(pre_over), pre_over[0]['month'], cut))
    if neg_over:
        raise IceBrentError(
            '有 %d 个月本仓的 Brent **少于** report/7 的 IFEU Brent 且超出四舍五入界'
            '（最早 %s）—— 归因不成立：report/7 只覆盖 IFEU，本仓是 IFEU + 已折算的新交所迷你，'
            '只可能多不可能少。这一列的口径要人工重判'
            % (len(neg_over), neg_over[0]['month']))

    stat = {
        'n': len(rows), 'first': rows[0]['month'], 'last': rows[-1]['month'],
        'brent_med': statistics.median(bd), 'brent_p95': bd[int(0.95 * len(bd))],
        'brent_max': abs(wb_['brent_dev']), 'brent_max_month': wb_['month'],
        'brent_max_csv': wb_['brent_csv'], 'brent_max_r7': wb_['brent_r7'],
        'gasoil_med': statistics.median(gd) if gd else None,
        'gasoil_max': max(gd) if gd else None,
        'n_within_round': len(rows) - len(over), 'n_over_round': len(over),
        'n_pre_cut': len(pre), 'cut': cut,
        'max_over': max((x['brent_csv'] - x['brent_r7'] for x in over), default=0.0),
    }
    return rows, stat


# ── ④ EIA 基期价格 ───────────────────────────────────────────────────
EIA_HEADER = 'Europe Brent Spot Price FOB (Dollars per Barrel)'


def eia_brent_base_price():
    import xlrd
    body = _http(URL_EIA_BRENT)
    wb = xlrd.open_workbook(file_contents=body)
    sh = wb.sheet_by_name('Data 1') if 'Data 1' in wb.sheet_names() \
        else wb.sheet_by_index(wb.nsheets - 1)
    hdr = sh.cell_value(2, 1)
    if EIA_HEADER not in str(hdr):
        raise IceBrentError('EIA 文件表头变了，期望含 %r，拿到 %r' % (EIA_HEADER, hdr))
    year, mon = int(BASE_MONTH[:4]), int(BASE_MONTH[5:])
    obs = []
    for r in range(3, sh.nrows):
        v = sh.cell_value(r, 0)
        if not isinstance(v, float):
            continue
        d = xlrd.xldate_as_tuple(v, wb.datemode)
        if d[0] != year or d[1] != mon:
            continue
        px = sh.cell_value(r, 1)
        if px in ('', None):
            continue
        obs.append(('%04d-%02d-%02d' % d[:3], float(px)))
    if len(obs) < 18:
        raise IceBrentError('EIA 在 %s 只有 %d 个报价日，太少' % (BASE_MONTH, len(obs)))
    if any(px <= 0 for _d, px in obs):
        raise IceBrentError('EIA 在 %s 有非正价格' % BASE_MONTH)
    avg = sum(px for _d, px in obs) / len(obs)
    out = os.path.join(CACHE, 'ice_brent_eia_2019_01.csv')
    with open(out, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(['date', 'brent_spot_usd_per_bbl', 'source'])
        for d, px in obs:
            w.writerow([d, px, URL_EIA_BRENT])
    print('    逐日值落库 %s（%d 行）' % (os.path.relpath(out, ROOT), len(obs)))
    return avg, obs


def check_eia_has_no_gasoil():
    """把「EIA 没有 ARA gasoil」变成一条每次都跑的断言，而不是一句备注。"""
    txt = _text(_http(URL_EIA_SPOTLIST))
    banner = 'Crude Oil in Dollars per Barrel, Products in Dollars per Gallon'
    if banner not in txt:
        raise IceBrentError(
            'EIA 现货价页的口径横幅变了（原文应含 %r）—— '
            '「EIA 没有 $/公吨 的 ARA gasoil」这条结论要重验' % banner)
    hits = [w for w in ('Rotterdam', 'ARA', 'Gasoil', 'Gas Oil', 'Amsterdam')
            if w in txt]
    if hits:
        raise IceBrentError(
            'EIA 现货价页上出现了 %s —— 说明现在可能有 gasoil / ARA 序列了，'
            '「Gasoil 不做」这条判断要重来' % hits)
    print('    ✓ EIA 现货价页横幅仍是 %r，且无 Rotterdam / ARA / Gasoil 任何字样' % banner)


# ══════════════════════════════════════════════════════════════════════
# 写表
# ══════════════════════════════════════════════════════════════════════
def build_row(price, obs, stat, notes, coverage):
    """组装 ICE_BRENT_IFEU 这一行。列顺序在 write_spec() 里按表头对齐，这里只给字典。"""
    ev = (
        '【2026-08-06 实测，复算脚本 build/basefill/ice_brent.py】'
        '口径一致的**子集**产品，取代原 ICE_ENERGY（全球合成篮子）。'
        '① 量列口径：ICE 官方 Monthly Statistics Tracking 工作簿 '
        "sheet 'Derivs (ADV, RPC, OI)' 脚注(1) 原文 —— " + notes[1] + ' '
        '⇒ 迷你合约已由官方折成 IFEU 标准当量，因此「张数 × 1,000 桶」对**全列**精确成立，'
        '不需要任何权重、不存在覆盖率缺口。'
        '同表其余能源行的脚注逐条读过，无一具备这个性质：'
        'Other Oil(3) 混 WTI/Murban/RBOB/Heating Oil/Wet Freight（桶+加仑+运费三种量纲）、'
        'Nat Gas(4) 与 Power(5) 混 North American/NGX/UK/European（MMBtu+therm+MWh）、'
        'Environmentals(6) 混排放权+煤+铁矿石。'
        '② 乘数：%s 2026-08-06 实测 HTTP 200，Market Specifications 表原文 '
        '"Contract Symbol B" / "Contract Size 1,000 barrels" / '
        '"Unit of Trading Any multiple of 1,000 barrels" / "Currency US Dollars and cents"。'
        '③ 「该列确实就是 IFEU」的独立验证：ICE 官方 %s（ICE Futures Europe '
        'Historical Monthly Volumes，服务端渲染，1995-2026 共 32 张年表）的 '
        'Brent 期货列 + Brent 期权列，与 series/ice.csv 的 '
        'adv_brent_kcontracts × trading_days_commod 逐月对账：%d 个月（%s – %s）'
        '中位偏差 %.3f%%、p95 %.3f%%、最大 %.3f%%（%s：本仓 %s 张 vs report/7 %s 张），'
        '**没有一个月超过 0.5%%**。'
        '残差还做了**证伪性归因**：ADV 四舍五入到整千张的单月最大影响 = 0.5 千张 × 交易日，'
        '%d/%d 个月落在这条界内；超界的 %d 个月全部满足两条判据 —— '
        '(a) **%s 之前 %d 个月无一例外落在界内**（官方脚注说 ICE Futures Singapore 的迷你 '
        'Brent 正是 %s 才并入的），(b) 超界月份的差**全部为正**（最大 %s 张），'
        '即本仓多出的正是 report/7 不覆盖、而官方已折成 IFEU 标准当量的那一块。'
        '若这一列里混进了别的合约规格，这两条判据不可能同时成立 —— '
        '脚本把它们写成硬断言，不成立就炸。两个残差来源都**不改变 1,000 桶这个乘数**。'
        '④ 基期价格：美国能源信息署 EIA 官方序列 RBRTE「%s」，文件 %s，'
        '%s 共 %d 个报价日（%s – %s）算术平均 = %.10f 美元/桶 '
        '⇒ 单张名义额 %.6f 美元。逐日值落库 cache/basefill/ice_brent_eia_2019_01.csv，可逐位复核。'
        '⑤ 基差留痕（必须知情）：RBRTE 是 **Dated Brent 现货 FOB 评估价**，不是 ICE Brent '
        '期货结算价 —— EIA 全站没有 Brent 期货序列（%s 实测：原油只有 "WTI - Cushing" 与 '
        '"Brent - Europe" 两条现货，成品油一律 $/gallon）。同月可比的期货口径参照：'
        '本仓 build/basefill/cme2.py 从 CME 官方 SPAN 解出的 NYMEX Brent 现金结算期货（BZ，'
        '同样 1,000 桶/张）2019-01 近月月均结算价 ≈ 60.122 美元/桶，比 EIA 现货高 +%.3f%%，'
        '方向与 2019-01 初 Brent 曲线的 contango 一致。这是**已知方向、已知大小**的期货/现货基差，'
        '落在本表已明示并接受的基准混用带内（avg_close vs basket_vw 实测 0.4%%~4%%）；'
        '与被否掉的「67%% 子集常数套 100%% 全口径量」不是一回事 —— 后者偏差方向与大小都不可知。'
        '不直接采用 60.122 的理由：那是上一轮留在 series/_specs_part_cme2.csv 里、'
        '四舍五入到整数的单张名义额反推出来的，本会话没有重跑 SPAN 解码，'
        '照抄等于写一个自己没测过的数。'
        '⑥ 期权按「一张期权 = 一张标的期货」折算，与 CME_ENERGY / CME_AG / CME_RATES 同一约定，'
        '所以 energy 池里 ICE 与 CME 两条腿对期权的处理是对称的。'
    ) % (URL_BRENT_SPEC, URL_REPORT7,
         stat['n'], stat['first'], stat['last'],
         stat['brent_med'] * 100, stat['brent_p95'] * 100, stat['brent_max'] * 100,
         stat['brent_max_month'], format(int(round(stat['brent_max_csv'])), ','),
         format(int(round(stat['brent_max_r7'])), ','),
         stat['n_within_round'], stat['n'], stat['n_over_round'],
         stat['cut'], stat['n_pre_cut'], stat['cut'],
         format(int(round(stat['max_over'])), ','),
         EIA_HEADER, URL_EIA_BRENT, BASE_MONTH, len(obs), obs[0][0], obs[-1][0],
         price, price * 1000, URL_EIA_SPOTLIST, (60.122 / price - 1) * 100)

    nt = (
        '⚠ **本产品只含 Brent，不是 ICE 的全部能源** —— 任何出现这条腿的图，'
        '图注必须写「ICE 只含 Brent 原油（期货+期权）」，不许让读者读成 ICE 能源全貌。'
        '实测覆盖率（张数口径，series/ice.csv）：%s Brent 占 ICE 全能源 %.1f%%'
        '（%s / %s 千张日均），187 个月中位 %.1f%%。'
        '这是**主动选择的低覆盖**：原 ICE_ENERGY 挂 adv_energy_kcontracts（全球口径，'
        '含 IFEU + Endex + IFUS + IFAD + NGX），而 ICE 唯一不要 reCAPTCHA 的分产品历史表 '
        'report/7 只覆盖 ICE Futures Europe（2019-01 各能源列相加只占全球 67.0%%），'
        'TTF 更在 Endex 上且没有固定乘数 —— 拿 67%% 的结构去套 100%% 的量，'
        '偏差方向与大小都不可知，图上完全看不出来。宁可覆盖 %.1f%% 而零偏差。'
        ' 📌 **未找到 Gasoil 的 2019-01 官方基期价格 ⇒ 本轮不含 Gasoil**（缺口详见 '
        'ICE_GASOIL_FUT 那一行）。乘数是有的（%s 2026-08-06 实测 HTTP 200，原文 '
        '"Contract Symbol G" / "Contract Size 100 metric tonnes"），缺的只有价格：'
        'Gasoil 以**美元/公吨**（ARA 驳船）报价，EIA 全站没有这个口径'
        '（现货价页横幅原文 "Crude Oil in Dollars per Barrel, Products in Dollars per Gallon"，'
        '成品油全部是美国口岸，无 Rotterdam / ARA 任何一条），ICE 自己的历史结算价在 '
        'Report Center 的 reCAPTCHA 后面（build/basefill/ice_enx2.py 实测 '
        'recaptchaRequired=true、criteria 恒 409）。拿 $/gallon 的美国柴油乘一个密度系数'
        '折成 $/tonne 就是编数。加 Gasoil 能把张数覆盖率从 %.1f%% 抬到 %.1f%%（2019-01），'
        '但代价是一个来路不明的常数 —— 按「准确优先于覆盖」不做。'
        ' ⚠ 名义额相同 ≠ 能量相同（原油与天然气热值不同），energy 池图注须写明（同 CME_ENERGY）。'
        ' ⚠ 期权与期货同记 1,000 桶：一张期权的经济敞口是 delta × 1,000 桶 < 1,000 桶。'
        '这是全仓统一约定（CME 各篮子同样如此）所以跨所比较是对称的，'
        '但绝不可把这条腿读成「实际转移的原油风险」。'
    ) % (BASE_MONTH, coverage['brent_share_base'] * 100,
         format(int(coverage['brent_k']), ','), format(int(coverage['energy_k']), ','),
         coverage['brent_share_med'] * 100, coverage['brent_share_base'] * 100,
         URL_GASOIL_SPEC,
         coverage['brent_share_base'] * 100, coverage['bg_share_base'] * 100)

    return {
        'product_id': PRODUCT_ID,
        'zh': 'ICE 布伦特原油（期货+期权，IFEU 标准合约当量）',
        'exchange': 'ICE',
        'pool': 'energy',
        'level': 'contract',
        'contract_name': 'Brent Crude Futures & Options '
                         '(ICE Futures Europe standard-size equivalents)',
        'underlying_symbol': 'B',
        'kind': 'contract',
        'ccy': 'USD',
        'multiplier': '1000',
        'mult_unit': '桶/张',
        'base_month': BASE_MONTH,
        'base_price_basis': 'avg_close',
        'base_price_local': repr(price),
        'base_notional_per_unit_local': repr(price * 1000),
        'notional_source': 'reconstructed',
        'price_id': 'BRENT',
        'spec_group': PRODUCT_ID,
        'effective_from': '',
        'effective_to': '',
        'source': ' | '.join([URL_BRENT_SPEC, URL_EIA_BRENT, URL_EIA_SPOTLIST,
                              URL_REPORT7,
                              "cache/ice_monthly_stats.xlsx sheet 'Derivs (ADV, RPC, OI)' 脚注(1)",
                              'build/basefill/ice_brent.py']),
        'evidence': ev,
        'notes': nt,
    }


def coverage_stats(ice):
    """Brent 在 ICE 全能源张数里的占比 —— 这是要写在页面上的那个"覆盖率"。"""
    import statistics
    shares_b, shares_bg = [], []
    base = None
    for m in sorted(ice):
        r = ice[m]
        try:
            b = float(r['adv_brent_kcontracts'])
            g = float(r['adv_gasoil_kcontracts'])
            e = float(r['adv_energy_kcontracts'])
        except (TypeError, ValueError):
            continue
        shares_b.append(b / e)
        shares_bg.append((b + g) / e)
        if m == BASE_MONTH:
            base = (b, g, e)
    if base is None:
        raise IceBrentError('series/ice.csv 里没有基期 %s' % BASE_MONTH)
    b, g, e = base
    return {'brent_k': b, 'gasoil_k': g, 'energy_k': e,
            'brent_share_base': b / e, 'bg_share_base': (b + g) / e,
            'brent_share_med': statistics.median(shares_b),
            'bg_share_med': statistics.median(shares_bg)}


def write_spec(row):
    """把这一行写进 series/contract_specs.csv —— **幂等**：同 id 覆盖，无则追加。

    只碰自己这一行。别的行（包括 ICE_ENERGY / ICE_STIR / ICE_MLTIR 的注解）是人工编辑，
    脚本重跑不许把它们冲掉。
    """
    path = os.path.join(SERIES, 'contract_specs.csv')
    raw = open(path, newline='', encoding='utf-8').read()
    rows = list(csv.reader(io.StringIO(raw)))
    hdr = rows[0]
    unknown = sorted(set(row) - set(hdr))
    if unknown:
        raise IceBrentError('contract_specs.csv 没有这些列：%s' % unknown)
    line = [row.get(c, '') for c in hdr]
    idx = [i for i, r in enumerate(rows) if i and r and r[0] == PRODUCT_ID]
    if idx:
        same = rows[idx[0]] == line
        rows[idx[0]] = line
        action = '未变（幂等）' if same else '覆盖'
    else:
        rows.append(line)
        action = '新增'
    buf = io.StringIO(newline='')
    csv.writer(buf, lineterminator='\n').writerows(rows)
    out = buf.getvalue()
    if out == raw:
        print('  [write] %s %s —— 文件逐字节未变' % (PRODUCT_ID, action))
        return False
    with open(path, 'w', newline='', encoding='utf-8') as f:
        f.write(out)
    print('  [write] %s %s → %s' % (PRODUCT_ID, action, os.path.relpath(path, ROOT)))
    return True


# ══════════════════════════════════════════════════════════════════════
# 邻座四行的口径定案 —— 与本行是同一个决定，所以在同一个脚本里落地
# ══════════════════════════════════════════════════════════════════════
# 幂等靠这个前缀：已经盖过章的行原样跳过。不这么做，脚本每跑一次就把同一段话
# 再抄一遍，notes 会滚成一坨（表里已经有几行是这么滚出来的）。
STAMP = '【2026-08-06 口径定案】'

_WHY_FROZEN = (
    '两条理由各自独立成立：'
    '① **官方拆不开** —— ICE 唯一公开、不要 reCAPTCHA 的分产品历史表 '
    'https://www.ice.com/report/7 把短端与中长端并成**一列** Interest Rates'
    '（2019-01 期货 37,491,346 + 期权 7,632,263 = 45,123,609 张，'
    '与 series/ice.csv 的 adv_stir + adv_mltir = 45,122,000 张只差 0.004%，'
    '确认是同一批合约）；分合约张数只在 Report Center 的 report 26/27，'
    'metadata 写着 recaptchaRequired=true、criteria 接口不带 token 恒返回 HTTP 409'
    '（build/basefill/ice_enx2.py 实测）。绕验证码是明令禁止的动作，'
    '人工在浏览器里点一次 cron 也复算不了。'
    '② **即便拆开，名义额对利率衍生品本身就是误导性单位** —— '
    '同样的名义额下 2 年期与 10 年期的 DV01 差 5 倍以上（久期 ~1.9 年 vs ~8 年）；'
    '正确的单位是 DV01 或久期加权名义额，而月度成交报表里没有久期字段'
    '（📌 未找到：CME / ICE / Eurex / MX / JPX 的月报都不含久期；'
    '检索路径 = 逐合约取到期日与票息自行算曲线，那是一整套要每月维护的数据，'
    '远超本仓无人值守的边界）。'
    '第②条与能不能拿到分合约张数**无关** ⇒ 这不是一个等着被补上的缺口。'
    '⇒ 这条腿只进增长图（定基名义额的增长率与张数增长率恒等，所以增长图上'
    '它与别人完全可比），不进水平值图与份额图，占比的分母也不含它。')

RETIRE_NOTES = {
    'ICE_ENERGY': (
        '⛔ **本行已停用**：不再被 build/pools.py 的 energy 池引用，'
        '也不再被 build/exchanges12.py 引用，由新行 **ICE_BRENT_IFEU** 取代。'
        '停用的理由不是"篮子太难"，是**分子分母不同源**：本行要配的量列 '
        'adv_energy_kcontracts 是 ICE **全球**口径（官方脚注明说 Nat Gas 与 Power 含 '
        'North American、NGX、UK、European，即 IFEU + Endex + IFUS + IFAD + NGX 五处），'
        '而 ICE 唯一公开、不要 reCAPTCHA 的分产品历史表 https://www.ice.com/report/7 '
        '**只覆盖 ICE Futures Europe** —— 2019-01 各能源列相加只占全球的 67.0%；'
        'TTF 在 ICE Endex 上根本不在那张表里，而且它**没有固定乘数**'
        '（官方页原文 "1 MW per day in contract period ... x 23, 24 or 25 hours"，'
        '一张月度合约 672–744 MWh 不等）。拿 67% 子集的品种结构算出的篮子常数'
        '去乘 100% 的量 = **方向与大小都不可知**的系统性偏差，'
        '而它在图上完全看不出来（柱子只是整体高一截）。'
        '⇒ 按「准确优先于覆盖」改走口径一致的子集 ICE_BRENT_IFEU：'
        '2019-01 只覆盖 ICE 全能源张数的 34.8%，但零偏差。'
        '本行保留不删，只为留住这段决策记录（check_specs 允许表里有没被引用的行）。'
        '**不要再去补这个常数** —— 要补也该补别的东西（比如 Gasoil 的基期价，见 '
        'ICE_GASOIL_FUT）。'),
    'ICE_STIR': (
        '⛔ **永久张数口径：基期常数永远留空，本行不再是待办。**'
        'build/pools.py 的 rates 与 eu_deriv 两个池已用 `contracts_only=True` '
        '显式声明这个状态（见该模块 docstring 六），'
        'build/check_specs.py 的覆盖率报告也把它列在「⛔ 永久张数口径」而不是「待测」。'
        + _WHY_FROZEN +
        ' 另：ICE Futures Europe 的官方合约规则 SECTION NNNN'
        '（https://www.ice.com/publicdocs/contractregs/112_SECTION_NNNN.pdf）通篇只写 '
        '"Unit of trading €2,500 * Rate Index" 与 "Contract Multiplier €2,500"，'
        '**没有任何 nominal / notional / face value 字样** —— '
        '常被引用的 EUR 1,000,000 是从每指数点欧元数反推的，反推值不许进本表。'),
    'ICE_MLTIR': (
        '⛔ **永久张数口径：基期常数永远留空，本行不再是待办。**'
        'build/pools.py 的 rates 与 eu_deriv 两个池已用 `contracts_only=True` '
        '显式声明这个状态（见该模块 docstring 六），'
        'build/check_specs.py 的覆盖率报告也把它列在「⛔ 永久张数口径」而不是「待测」。'
        + _WHY_FROZEN +
        ' 另：与 ICE_STIR 不同，中长端的**面值是有官方原文的** —— SECTION RRRR'
        '（https://www.ice.com/publicdocs/contractregs/116_SECTION_RRRR.pdf）'
        '"Unit of Trading £100,000 nominal value notional Gilt"，四个 Gilt 合约同值；'
        '缺的只是权重。但理由②与面值有没有无关，所以结论一样：不填。'),
    'ICE_BRENT_FUT': (
        '✅ 这个缺口已由新行 **ICE_BRENT_IFEU** 解决 —— 基期价格取美国能源信息署 EIA '
        '的官方序列 RBRTE，2019-01 月均 59.4095454545 美元/桶。'
        '本行保留为**官方合约规格的出处行**（https://www.ice.com/products/219/'
        'Brent-Crude-Futures，2026-08-06 复验 HTTP 200，原文 "Contract Symbol B" / '
        '"Contract Size 1,000 barrels"），不被任何池引用。'
        '⚠ 上面那段「📌 未找到 Brent 2019-01 的官方结算价」**仍然成立**：'
        'ICE 自己的历史结算价确实在 Report Center 后面。ICE_BRENT_IFEU 用的是 '
        'EIA 的 **Dated Brent 现货 FOB**，不是 ICE 期货结算价 —— '
        '两者 2019-01 的基差已量化并写进 ICE_BRENT_IFEU 的 evidence'
        '（比本仓 SPAN 解出的 CME Brent 现金结算期货 BZ 近月月均低约 1.2%，'
        '方向与当月初的 contango 一致）。'),
    'ICE_GASOIL_FUT': (
        '⛔ **本轮不做 Gasoil**：energy 池的 ICE 腿只含 Brent（见 ICE_BRENT_IFEU）。'
        '乘数是有的 —— https://www.ice.com/products/34361119/Low-Sulphur-Gasoil-Futures '
        '2026-08-06 复验 HTTP 200，原文 "Contract Symbol G" / '
        '"Contract Size 100 metric tonnes"。'
        '📌 缺的只有 2019-01 的官方基期价格，而这一条本轮**逐条实测排除**：'
        'Gasoil 以**美元/公吨**（ARA 驳船）报价，EIA 全站没有这个口径 —— '
        '现货价页 https://www.eia.gov/dnav/pet/pet_pri_spt_s1_d.htm 的口径横幅原文就是 '
        '"Crude Oil in Dollars per Barrel, Products in Dollars per Gallon"，'
        '成品油全部是美国口岸（NY Harbor / US Gulf Coast / LA），'
        '全页无 Rotterdam / ARA / Gasoil 任何字样；'
        'ICE 自己的历史结算价在 Report Center 的 reCAPTCHA 后面'
        '（build/basefill/ice_enx2.py 实测 recaptchaRequired=true、criteria 恒 409）。'
        '拿 $/gallon 的美国柴油乘一个密度系数折成 $/tonne 就是编数，本仓不做。'
        '代价已量化：加上 Gasoil 能把 ICE 腿的张数覆盖率从 2019-01 的 34.8% 抬到 46.3%'
        '（187 个月中位 33.3% → 43.1%）—— 按「准确优先于覆盖」，'
        '宁可少这 11.5pp 也不要一个来路不明的常数。'
        '检索路径留给下一个人：ICE Report Center 的 report 26/27 人工导出，'
        '**必须把导出的原始 CSV 一起提交**，否则仓库仍然复算不了。'),
}


def stamp_retired_notes():
    """给 ICE 的另外四行盖上同一个决定的章。幂等：已有 STAMP 前缀的行原样跳过。"""
    path = os.path.join(SERIES, 'contract_specs.csv')
    raw = open(path, newline='', encoding='utf-8').read()
    rows = list(csv.reader(io.StringIO(raw)))
    hdr = rows[0]
    i_id, i_notes = hdr.index('product_id'), hdr.index('notes')
    touched = []
    for r in rows[1:]:
        pid = r[i_id]
        if pid not in RETIRE_NOTES:
            continue
        if STAMP in r[i_notes]:
            print('    · %-16s 已盖章，跳过（幂等）' % pid)
            continue
        r[i_notes] = STAMP + RETIRE_NOTES[pid] + ' ‖ 以下是定案前的旧注，保留是为了'\
                     '留住当时排除过的路径： ' + r[i_notes]
        touched.append(pid)
    if not touched:
        return False
    buf = io.StringIO(newline='')
    csv.writer(buf, lineterminator='\n').writerows(rows)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        f.write(buf.getvalue())
    print('    · 盖章 %d 行：%s' % (len(touched), '、'.join(touched)))
    return True


# ══════════════════════════════════════════════════════════════════════
def main(argv):
    ap = argparse.ArgumentParser(description='ICE_BRENT_IFEU 基期常数实测')
    ap.add_argument('--write', action='store_true',
                    help='把结果写进 series/contract_specs.csv（幂等）')
    args = ap.parse_args(argv)

    print('══ 1. 官方脚注：adv_brent_kcontracts 到底是什么口径 ' + '═' * 26)
    notes = read_footnotes()
    for k in (1, 2, 3, 4, 5, 6):
        if k in notes:
            print('  (%d) %s' % (k, notes[k][:150] + ('…' if len(notes[k]) > 150 else '')))

    print('\n══ 2. 乘数：ICE 官方合约规格页现抓现验 ' + '═' * 36)
    check_contract_specs()

    print('\n══ 3. 该列 = ICE Futures Europe：用官方 report/7 独立验证 ' + '═' * 16)
    ice = read_ice_csv()
    ifeu = parse_report7(_http(URL_REPORT7))
    _rows, stat = crosscheck_report7(ifeu, ice)
    print('    Brent  %d 个月（%s – %s）|dev| 中位 %.4f%% / p95 %.4f%% / 最大 %.4f%% @%s'
          % (stat['n'], stat['first'], stat['last'], stat['brent_med'] * 100,
             stat['brent_p95'] * 100, stat['brent_max'] * 100, stat['brent_max_month']))
    print('           最大那个月：本仓 %s 张 vs report/7 %s 张'
          % (format(int(round(stat['brent_max_csv'])), ','),
             format(int(round(stat['brent_max_r7'])), ',')))
    print('    Gasoil |dev| 中位 %.4f%% / 最大 %.4f%%（本轮不做 Gasoil，只作口径旁证）'
          % (stat['gasoil_med'] * 100, stat['gasoil_max'] * 100))
    print('    ✓ 全部 %d 个月都在 %.1f%% 容差内' % (stat['n'], REPORT7_TOL * 100))
    print('    残差归因（证伪性判据，不成立就炸）：'
          '四舍五入界内 %d/%d 个月；超界 %d 个月，'
          '其中 %s 之前 %d 个月**无一超界**（迷你 Brent %s 才并入），'
          '且超界月份的差全部为正（最大 %s 张）'
          % (stat['n_within_round'], stat['n'], stat['n_over_round'],
             stat['cut'], stat['n_pre_cut'], stat['cut'],
             format(int(round(stat['max_over'])), ',')))

    print('\n══ 4. 基期价格：EIA 官方 Brent 序列 ' + '═' * 38)
    check_eia_has_no_gasoil()
    price, obs = eia_brent_base_price()
    print('    %s 共 %d 个报价日（%s – %s）' % (BASE_MONTH, len(obs), obs[0][0], obs[-1][0]))
    print('    算术平均 = %.10f 美元/桶  ⇒ 单张名义额 = %.6f 美元'
          % (price, price * 1000))
    print('    期货/现货基差参照：CME BZ 2019-01 近月月均 ≈ 60.122 ⇒ 高 %+.3f%%'
          % ((60.122 / price - 1) * 100))

    print('\n══ 5. 覆盖率：这条腿代表 ICE 能源的多少 ' + '═' * 34)
    cov = coverage_stats(ice)
    print('    %s Brent %s / 全能源 %s 千张日均 = %.1f%%（Brent+Gasoil 本可到 %.1f%%）'
          % (BASE_MONTH, format(int(cov['brent_k']), ','),
             format(int(cov['energy_k']), ','),
             cov['brent_share_base'] * 100, cov['bg_share_base'] * 100))
    print('    187 个月中位：Brent %.1f%%、Brent+Gasoil %.1f%%'
          % (cov['brent_share_med'] * 100, cov['bg_share_med'] * 100))
    print('    ⚠ 这个数字必须出现在页面上 —— 不许让读者以为是 ICE 全部能源')

    row = build_row(price, obs, stat, notes, cov)
    print('\n══ 6. 落表 ' + '═' * 62)
    if args.write:
        write_spec(row)
        stamp_retired_notes()
    else:
        print('  （未加 --write，不写表）%s = %s 美元/张'
              % (PRODUCT_ID, row['base_notional_per_unit_local']))
    return 0


if __name__ == '__main__':
    try:
        sys.exit(main(sys.argv[1:]))
    except IceBrentError as e:
        print('FAIL %s' % e)
        sys.exit(1)
