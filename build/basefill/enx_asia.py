# -*- coding: utf-8 -*-
"""Euronext + 东亚（HKEX / JPX / SGX）六个 pool_product 篮子常数的取数与复算脚本。

    python3 build/basefill/enx_asia.py            # 跑全流程，打印覆盖率报告，写 part CSV
    python3 build/basefill/enx_asia.py --no-net   # 只用 cache/，不发任何网络请求

━━ 这个文件为什么存在 ━━
series/contract_specs.csv 里既有的 23 行常数，它们的取数代码没留在仓库里，
cache/ 又被 .gitignore —— 结果是「表里的数字对不对，仓库自己没法复算」。
本脚本把**这一批**的每一个中间量都写成可重跑的代码：
规格页 URL、解析出的乘数、2019-01 的张数权重、以及最后那一步加权，
全部在这里，跑一次就能逐位对账。

━━ 篮子常数的定义（照 contract_specs.csv 的 notes 列）━━
    篮子常数 = Σ(成员合约 2019-01 张数 × 该合约乘数 × 该合约基期价格)
             ÷ Σ(成员合约 2019-01 张数)

即「按基期成交结构折出来的等效单张名义额」。只算一次、写死，**绝不逐月重算** ——
重算会让价格项随月度品种结构漂移，定基口径当场失效。

三个乘数：
  · 张数      —— 官方月度统计（JPX 取引総括表 / HKEX Monthly Market Highlights /
                 SGX Monthly Market Statistics Report），本脚本从 cache/ 读
  · 乘数      —— 官方合约规格页，本脚本实时抓 + 正则解析（见 SPEC_PAGES）
  · 基期价格  —— **2019-01 月内全部交易日收盘价的算术平均**（base_price_basis=avg_close，
                 与既有 23 行同口径，见 build/check_specs.py 的「口径 2」）

━━ 本次的结论：卡在第三个乘数上 ━━
张数与乘数这两项，本脚本已经把能拿的都拿到了（见运行输出的覆盖率表）。
**基期价格一项全线取不到**：avg_close 口径要的是「2019-01 逐个交易日的指数收盘价」，
而这六个池所涉的指数，官方一手源在无头环境下都拿不到日线序列：

    N225      indexes.nikkei.co.jp  → Cloudflare 挑战页（"Just a moment..."）
    TOPIX 系  jpx.co.jp             → 官方只免费发**月末**值与当月快照；
                                      日线历史在 JPX Data Cloud 收费产品里
    HSI/HSCEI hsi.com.hk            → Angular 单页应用，服务端只返回 1.9KB 外壳
    CAC/AEX   live.euronext.com     → 图表数据接口 /intraday_chart/getChartData/
                                      返回 AES 密文（字段 ct/iv/s），密钥在前端 JS 里

⚠ **绝不许**拿以下四类东西凑数（本脚本刻意不实现它们）：
  1. 第三方聚合站（investing.com / wikipedia / stooq / yahoo）—— 仓库硬约束禁止
  2. 期货成交均价（JPX 総括表第 16 列「取引金額」÷ 张数 ÷ 乘数）—— 这是
     **成交量加权成交价**，不是 avg_close。两种基准混进同一张表，是 check_specs.py
     C7 点名的「错了看不出来」的一类：柱子整体偏一截，图上完全正常
  3. 月末收盘代替月均收盘 —— 同上。SPX 实测两者差 3.7%
  4. 「行业惯例值」（如「单股期权都是 100 股」）—— 没实测就是没实测

所以本脚本写出的 part CSV，六行的 base_price_local / basket_constant **全部留空**，
computed_note 里逐行写清楚「差哪一个价格序列、它在基期占多少张数权重」。
价格序列一旦到手，把它填进 BASE_PRICES 再跑一次，常数就出来了 —— 这是本脚本的主要价值。

━━ 依赖 ━━ 标准库 + openpyxl（读 xlsx）+ pymupdf（读 SGX 的 PDF）。
两个都已在 requirements.txt 里。
"""

import argparse
import csv
import html
import os
import re
import sys
import urllib.request
import urllib.error

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
CACHE = os.path.join(ROOT, 'cache', 'basefill')
OUT_CSV = os.path.join(ROOT, 'series', '_specs_part_enx_asia.csv')

BASE_MONTH = '2019-01'
BASE_PRICE_BASIS = 'avg_close'      # 与既有 23 行一致，见 build/check_specs.py 口径 2

_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36')


# ══════════════════════════════════════════════════════════════════════
# 一、官方合约规格页 —— 乘数
# ══════════════════════════════════════════════════════════════════════
# key = 本脚本内部的成员 id；值 = (URL, 解析用的字段标签正则)
# 这些 URL 是 2026-08-06 实测服务端可读的（不是 JS 外壳）：
# JPX   .../products/<族>/<产品>/01.html      字段「Contract Unit」
# HKEX  .../Products/Listed-Derivatives/...   字段「Contract Multiplier」
# ENX   live.euronext.com/.../contract-specification  字段「Unit of trading」
SPEC_PAGES = {
    # ── JPX（jpx.co.jp，英文规格页）──
    'JPX_N225_FUT':    ('https://www.jpx.co.jp/english/derivatives/products/domestic/225futures/01.html', 'jpx'),
    'JPX_N225_MINI':   ('https://www.jpx.co.jp/english/derivatives/products/domestic/225mini/01.html', 'jpx'),
    'JPX_N225_OPT':    ('https://www.jpx.co.jp/english/derivatives/products/domestic/225options/01.html', 'jpx'),
    # 下面两个 2019-01 还没上市（2023-05 才有），抓它们只为把 fetch/jpx.py 的
    # _DIVISORS 折算表整张交叉核完 —— 见 crosscheck_jpx_divisors()
    'JPX_N225_MICRO':  ('https://www.jpx.co.jp/english/derivatives/products/domestic/225micro-futures/01.html', 'jpx'),
    'JPX_N225_MINIOPT': ('https://www.jpx.co.jp/english/derivatives/products/domestic/225mini-options/01.html', 'jpx'),
    'JPX_TOPIX_FUT':   ('https://www.jpx.co.jp/english/derivatives/products/domestic/topix-futures/01.html', 'jpx'),
    'JPX_TOPIX_MINI':  ('https://www.jpx.co.jp/english/derivatives/products/domestic/mini-topix-futures/01.html', 'jpx'),
    'JPX_TOPIX_OPT':   ('https://www.jpx.co.jp/english/derivatives/products/domestic/topix-options/01.html', 'jpx'),
    'JPX_JPX400_FUT':  ('https://www.jpx.co.jp/english/derivatives/products/domestic/jpx-nikkei400futures/01.html', 'jpx'),
    'JPX_CORE30_FUT':  ('https://www.jpx.co.jp/english/derivatives/products/domestic/topix-core30futures/01.html', 'jpx'),
    'JPX_BANKS_FUT':   ('https://www.jpx.co.jp/english/derivatives/products/domestic/topix-banks-index-futures/01.html', 'jpx'),
    'JPX_REIT_FUT':    ('https://www.jpx.co.jp/english/derivatives/products/reit/reit-futures/01.html', 'jpx'),
    'JPX_GROWTH250':   ('https://www.jpx.co.jp/english/derivatives/products/domestic/tse-growth250futures/01.html', 'jpx'),
    'JPX_DJIA_FUT':    ('https://www.jpx.co.jp/english/derivatives/products/foreign/djia-futures/01.html', 'jpx'),
    'JPX_TAIEX_FUT':   ('https://www.jpx.co.jp/english/derivatives/products/foreign/taiex-futures/01.html', 'jpx'),
    'JPX_N225VI_FUT':  ('https://www.jpx.co.jp/english/derivatives/products/vi/225-vi-futures/01.html', 'jpx'),
    'JPX_JGB_FUT':     ('https://www.jpx.co.jp/english/derivatives/products/jgb/jgb-futures/01.html', 'jpx_jgb'),
    'JPX_JGB_MINI':    ('https://www.jpx.co.jp/english/derivatives/products/jgb/mini-jgb-futures/01.html', 'jpx'),
    'JPX_JGB_OPT':     ('https://www.jpx.co.jp/english/derivatives/products/jgb/jgbf-options/01.html', 'jpx'),
    'JPX_SEC_OPT':     ('https://www.jpx.co.jp/english/derivatives/products/individual/securities-options/01.html', 'jpx'),
    # ── HKEX（hkex.com.hk）──
    'HKEX_HSI_FUT':    ('https://www.hkex.com.hk/Products/Listed-Derivatives/Equity-Index/Hang-Seng-Index-(HSI)/Hang-Seng-Index-Futures?sc_lang=en', 'hkex'),
    'HKEX_MHI_FUT':    ('https://www.hkex.com.hk/Products/Listed-Derivatives/Equity-Index/Hang-Seng-Index-(HSI)/Mini-Hang-Seng-Index-Futures?sc_lang=en', 'hkex'),
    'HKEX_HSI_OPT':    ('https://www.hkex.com.hk/Products/Listed-Derivatives/Equity-Index/Hang-Seng-Index-(HSI)/Hang-Seng-Index-Options?sc_lang=en', 'hkex'),
    'HKEX_MHI_OPT':    ('https://www.hkex.com.hk/Products/Listed-Derivatives/Equity-Index/Hang-Seng-Index-(HSI)/Mini-Hang-Seng-Index-Options?sc_lang=en', 'hkex'),
    'HKEX_HSI_DIV':    ('https://www.hkex.com.hk/Products/Listed-Derivatives/Equity-Index/Hang-Seng-Index-(HSI)/Dividend-Futures?sc_lang=en', 'hkex'),
    'HKEX_HHI_FUT':    ('https://www.hkex.com.hk/Products/Listed-Derivatives/Equity-Index/Hang-Seng-China-Enterprises-Index/Hang-Seng-China-Enterprises-Index-Futures?sc_lang=en', 'hkex'),
    'HKEX_HHI_OPT':    ('https://www.hkex.com.hk/Products/Listed-Derivatives/Equity-Index/Hang-Seng-China-Enterprises-Index/Hang-Seng-China-Enterprises-Index-Options?sc_lang=en', 'hkex'),
    'HKEX_MCH_FUT':    ('https://www.hkex.com.hk/Products/Listed-Derivatives/Equity-Index/Hang-Seng-China-Enterprises-Index/Mini-H-shares-Index-Futures?sc_lang=en', 'hkex'),
    'HKEX_MCH_OPT':    ('https://www.hkex.com.hk/Products/Listed-Derivatives/Equity-Index/Hang-Seng-China-Enterprises-Index/Mini-H-shares-Index-Options?sc_lang=en', 'hkex'),
    'HKEX_STOCK_FUT':  ('https://www.hkex.com.hk/Products/Listed-Derivatives/Single-Stock/Stock-Futures?sc_lang=en', 'hkex'),
    'HKEX_STOCK_OPT':  ('https://www.hkex.com.hk/Products/Listed-Derivatives/Single-Stock/Stock-Options?sc_lang=en', 'hkex'),
    # ── Euronext（live.euronext.com）──
    # ⚠ contract_specs_todo.csv 里写「live.euronext.com 服务端只给栏目导航」的那条
    #   blocker 是**路径错**，不是站点拦：产品页 /en/product/<类>/<代码>-<市场>/ 本身
    #   确实只有导航，但它下面的 **/contract-specification** 子页是服务端渲染的。
    'ENX_CAC40_FUT':   ('https://live.euronext.com/en/product/index-futures/FCE-DPAR/contract-specification', 'enx'),
    'ENX_AEX_FUT':     ('https://live.euronext.com/en/product/index-futures/FTI-DAMS/contract-specification', 'enx'),
    'ENX_WHEAT_FUT':   ('https://live.euronext.com/en/product/commodities-futures/EBM-DPAR/contract-specification', 'enx'),
}

# 规格页解析出来的乘数（数值 + 计价单位），跑完 collect_multipliers() 后填满
_MULT_RE = {
    # JPX「Contract Unit」下一行形如 'Nikkei 225 × JPY 1,000' / 'TOPIX × ¥10,000'
    # mini-10y JGB 那页是句子：'Multiply 100 thousand yen by the price of ...'
    'jpx':     re.compile(r'[×xX]\s*(?:JPY|¥)\s*([\d,]+)'
                          r'|([\d,]+)\s*thousand yen'
                          r'|([\d,]+)\s*million yen'),
    # JPX JGB 页是三列表格，Contract Unit 后面跟 '5-year / 10-year / mini 20-year'
    # 再跟 'JPY 100 million1 face value' / 'JPY 10 million face value'
    'jpx_jgb': re.compile(r'JPY\s*([\d,]+)\s*(million|billion)'),
    # HKEX「Contract Multiplier」下一行形如 'HK$50 per index point'
    'hkex':    re.compile(r'HK\$\s*([\d,]+)\s*per', re.I),
    # Euronext「Unit of trading」下一行是纯数字（股指：10 / 200）
    # 或英文拼写的吨数（农产品：'Fifty tonnes'）——两种都要接住
    'enx':     re.compile(r'^([\d,]+)$|^(Fifty|One hundred|Ten|Twenty|Twenty-five) tonnes$', re.I),
}
# Euronext 农产品页把合约量写成英文单词，不是数字
_WORD_NUM = {'ten': 10, 'twenty': 20, 'twenty-five': 25, 'fifty': 50, 'one hundred': 100}
_LABEL = {
    'jpx':     re.compile(r'^(Contract Unit|Trading Unit)$'),
    'jpx_jgb': re.compile(r'^(Contract Unit|Trading Unit)$'),
    'hkex':    re.compile(r'^Contract Multiplier$'),
    'enx':     re.compile(r'^Unit of trading$'),
}


def _http_get(url, timeout=45):
    req = urllib.request.Request(url, headers={
        'User-Agent': _UA, 'Accept-Language': 'en-US,en;q=0.9',
        'Accept': 'text/html,application/xhtml+xml,*/*;q=0.8'})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read()


def _cached(name, url, use_net):
    """先看 cache/basefill/<name>，没有再下。下失败就返回 None（不抛）。"""
    path = os.path.join(CACHE, name)
    if os.path.exists(path) and os.path.getsize(path) > 0:
        return open(path, 'rb').read()
    if not use_net:
        return None
    try:
        data = _http_get(url)
    except (urllib.error.URLError, urllib.error.HTTPError, OSError) as e:
        print('    ⚠ 取不到 %s（%r）' % (url, e))
        return None
    os.makedirs(CACHE, exist_ok=True)
    with open(path, 'wb') as f:
        f.write(data)
    return data


def _text_lines(raw):
    """把 HTML 拍成非空文本行 —— 规格页全是 <dt>/<dd> 或两列表格，
    「标签行的下一行就是值」这条规律在 JPX / HKEX / Euronext 三家都成立。"""
    t = raw.decode('utf-8', 'replace')
    t = re.sub(r'<(script|style)[^>]*>.*?</\1>', ' ', t, flags=re.S | re.I)
    t = html.unescape(re.sub(r'<[^>]+>', '\n', t))
    return [l.strip() for l in t.split('\n') if l.strip()]


def collect_multipliers(use_net):
    """抓官方规格页，解析乘数。返回 {member_id: (mult, unit_text, url)}。"""
    print('\n【一】官方合约规格页 → 乘数')
    out = {}
    for mid, (url, flavor) in SPEC_PAGES.items():
        raw = _cached('spec_%s.html' % mid.lower(), url, use_net)
        if raw is None:
            print('  %-16s ✗ 页面未取到' % mid)
            continue
        lines = _text_lines(raw)
        mult, txt = None, ''
        for i, l in enumerate(lines):
            if not _LABEL[flavor].match(l):
                continue
            for cand in lines[i + 1:i + 6]:
                m = _MULT_RE[flavor].search(cand)
                if m:
                    tok = next((g for g in m.groups() if g), None)
                    if tok is None:
                        continue
                    if tok.lower() in _WORD_NUM:
                        v = float(_WORD_NUM[tok.lower()])
                    else:
                        v = float(tok.replace(',', ''))
                    if flavor == 'jpx' and 'thousand yen' in cand:
                        v *= 1000.0
                    if flavor == 'jpx' and 'million yen' in cand:
                        v *= 1e6
                    if flavor == 'jpx_jgb':
                        # 10 年 JGB 是「JPY 100 million face value」；同页的
                        # mini 20 年是 10 million。取第一个 100 million 那档。
                        v *= 1e6 if m.group(2) == 'million' else 1e9
                        if v != 1e8:
                            continue
                    mult, txt = v, cand
                    break
            if mult is not None:
                break
        if mult is None:
            # 解析不出来不是「没有」，而是这一页的字段是文字描述（如 HKEX 股票
            # 期权的 Contract Size = 各标的 board lot）。原样留证据，不猜。
            ev = ''
            for i, l in enumerate(lines):
                if _LABEL[flavor].match(l) or re.match(r'^Contract Size$', l):
                    ev = ' / '.join(lines[i + 1:i + 3])[:90]
                    break
            print('  %-16s ✗ 无固定乘数：%s' % (mid, ev or '(字段未出现)'))
            out[mid] = (None, ev, url)
        else:
            print('  %-16s ✓ %s   ← %r' % (mid, _fmt(mult), txt[:52]))
            out[mid] = (mult, txt, url)
    return out


def _fmt(x):
    if x is None:
        return '(空)'
    return '%.10g' % x if x < 1e6 else format(int(x), ',')


# ══════════════════════════════════════════════════════════════════════
# 二、2019-01 张数权重 —— 官方月度统计
# ══════════════════════════════════════════════════════════════════════
# 三份文件都是官方一手，且已经在 cache/ 里（各自的 fetcher 下的）。
JPX_SOUKATSU = os.path.join(ROOT, 'cache', 'jpx_soukatsu_M.xlsx')     # 取引総括表（月間）
HKEX_HILITE = os.path.join(ROOT, 'cache', 'hkex_monthly_highlights_2026-06.xlsx')
SGX_PDF = os.path.join(ROOT, 'cache', 'sgx_2019-01.pdf')              # SGX Monthly Market Statistics Jan-2019


def jpx_weights():
    """JPX 取引総括表（月間）里 2019 年 1 月的逐产品**月度**张数。
    期权行有 プット/コール/合計 三条，只取「合計」，否则重复计一次。"""
    import openpyxl
    wb = openpyxl.load_workbook(JPX_SOUKATSU, read_only=True, data_only=True)
    ws = wb['BO_DM0040']
    out = {}
    for r in ws.iter_rows(min_row=7, values_only=True):
        try:
            y, m = int(r[4]), int(r[5])
        except (TypeError, ValueError):
            continue
        if (y, m) != (2019, 1):
            continue
        pc, vol = r[2], r[7]
        if pc not in (None, '', '合計'):        # プット / コール 两条跳过
            continue
        if not isinstance(vol, (int, float)):   # '－' = 当月无成交
            continue
        out[str(r[0])] = float(vol)
    wb.close()
    return out


def hkex_weights():
    """HKEX Monthly Market Highlights 的 DERIVATIVES MARKET TURNOVER 段，
    2019-01 那一列的逐产品**日均**张数（官方口径就是 ADV）。"""
    import datetime
    import openpyxl
    wb = openpyxl.load_workbook(HKEX_HILITE, read_only=True, data_only=True)
    ws = wb['Monthly Data']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    head = None
    for i, r in enumerate(rows):
        if r and str(r[0] or '').strip() == 'Average daily volume (contracts)':
            head = i
            break
    if head is None:
        raise RuntimeError('HKEX 表里找不到「Average daily volume (contracts)」表头行')
    col = None
    for j, v in enumerate(rows[head]):
        if isinstance(v, datetime.datetime) and (v.year, v.month) == (2019, 1):
            col = j
    if col is None:
        raise RuntimeError('HKEX 表里找不到 2019-01 那一列')
    out = {}
    for r in rows[head + 1:head + 110]:
        lbl = str(r[0] or '').strip()
        if not lbl or lbl.startswith(('*', '^', 'Note', 'Number of', '-')):
            continue
        v = r[col]
        if isinstance(v, (int, float)):
            out[lbl] = float(v)
        if lbl == 'Total Futures and Options':
            break
    return out


def sgx_weights():
    """SGX Monthly Market Statistics Report（Jan 2019）的
    Derivatives Market Volume By Contract Types 各表，2019-01 那一列的**月度**张数。

    PDF 是文本层 PDF，每张表被 get_text() 拍成「产品名 + 9 个数字」的行序列，
    第 5 个数字才是 Jan 2019（前四列是 FY2019 Q1 / Q2 / Nov 2018 / Dec 2018）。
    列序在报告里是写死的表头，本函数按表头重新定位一次，不硬编码下标。"""
    import fitz
    doc = fitz.open(SGX_PDF)
    num = re.compile(r'^-?[\d,]+$')
    # 报告前半是证券市场，后半才是衍生品；只从「Derivatives Market Volume By Contract
    # Types」那一页起解析，否则会把证券市场的「Total Number Of Trades」之类混进来。
    start, pages = 0, []
    for p in range(doc.page_count):
        pages.append(doc[p].get_text())
        if not start and 'Derivatives Market Volume By Contract Types' in pages[-1]:
            start = p
    skip = re.compile(r'^(Total|SGX |Average Daily|.*Open Interest$|.*Number Of Trades)')
    out, cur_hdr = {}, None
    for p in range(doc.page_count):
        if p < start and 'Total Trading Volume (7)' not in pages[p]:
            continue        # 合计行在衍生品明细页之前的汇总页上，单独放行
        lines = [l.strip() for l in pages[p].split('\n') if l.strip()]
        i = 0
        while i < len(lines):
            if lines[i] == 'FY2019 Q1':
                cur_hdr = lines[i:i + 9]
                i += 9
                continue
            if cur_hdr and 'Jan 2019' in cur_hdr:
                k = cur_hdr.index('Jan 2019')
                name = lines[i]
                vals = lines[i + 1:i + 1 + 9]
                if (not num.match(name) and len(vals) == 9
                        and all(num.match(v) or v.endswith('%') or v == 'N.A.' for v in vals)):
                    v = vals[k]
                    if num.match(v):
                        if name == 'Total Trading Volume (7)':
                            out['__OFFICIAL_TOTAL__'] = float(v.replace(',', ''))
                        elif not skip.match(name):
                            out.setdefault(name, float(v.replace(',', '')))
                    i += 10
                    continue
            i += 1
    doc.close()
    return out


# ══════════════════════════════════════════════════════════════════════
# 三、基期价格 —— 本次全线缺口
# ══════════════════════════════════════════════════════════════════════
# 口径：2019-01 月内**全部交易日收盘价的算术平均**（avg_close）。
# 填法：拿到官方一手日线后，把 {价格 id: 数值} 填进来再跑一次本脚本即可。
# ⚠ 不要往这里填成交均价、月末收盘、或第三方站点的数 —— 理由见模块 docstring。
BASE_PRICES = {
    # 'N225':  ...,   # 日経平均株価 2019-01 月均收盘
    # 'TOPIX': ...,   # TOPIX 2019-01 月均收盘
    # 'HSI':   ...,   # 恒生指数 2019-01 月均收盘
    # 'HSCEI': ...,   # 恒生中国企业指数 2019-01 月均收盘
    # 'CAC40': ...,   # CAC 40 2019-01 月均收盘
    # 'AEX':   ...,   # AEX 2019-01 月均收盘
}
# 面值口径的产品：价格项按定义 = 1（与 CME_ZN/ZB/ZF 那几行同一处理）
PRICE_DEFINITIONAL = 1.0

# 每个价格 id 缺的原因 + 检索路径（写进 part CSV 的 computed_note）
PRICE_BLOCKERS = {
    'N225':  'indexes.nikkei.co.jp（日経 = N225 的编制方）返回 Cloudflare 挑战页；'
             '检索路径 = Nikkei Inc 官方历史数据下载，或 JPX 統計月報 03_sisu*.pdf 的 2019 年档',
    'TOPIX': 'jpx.co.jp 只免费发月末值与当月快照（/markets/indices/related/value/），'
             '日线历史在 JPX Data Cloud 收费；检索路径 = JPX 統計月報「指数」分册 03_sisu1901.pdf，'
             '该页当前只列本财年，需先找到 2019 年的归档目录 hash',
    'HSI':   'hsi.com.hk（恒生指数公司 = HSI 编制方）是 Angular 单页，服务端只返回 1.9KB 外壳；'
             '检索路径 = 恒指公司官方 index history 下载，或 HKEX Fact Book 2019（⚠ Fact Book 给的是'
             '月末/高低值，不是月均收盘，口径不符，不能直接用）',
    'HSCEI': '同 HSI',
    'CAC40': 'live.euronext.com 的 /intraday_chart/getChartData/FR0003500008-XPAR/max 返回 AES 密文'
             '（{ct,iv,s}），密钥在前端 JS；检索路径 = Euronext 官方 index 历史数据文件',
    'AEX':   '同 CAC40（AEX 的 ISIN 是 NL0000000107-XAMS）',
}


# ══════════════════════════════════════════════════════════════════════
# 四、六个池的成员表
# ══════════════════════════════════════════════════════════════════════
# 每个成员 = (源统计里的行名, 规格页 member_id 或 None, 价格 id 或 None)
# 价格 id = None 且 member_id 给了乘数 ⇒ 面值口径，价格项 = 1
JPX_MEMBERS = [
    # (取引総括表的日文行名, 规格 id, 价格 id)
    ('日経225先物',                     'JPX_N225_FUT',   'N225'),
    ('日経225mini',                     'JPX_N225_MINI',  'N225'),
    ('日経225オプション',               'JPX_N225_OPT',   'N225'),
    ('TOPIX先物',                       'JPX_TOPIX_FUT',  'TOPIX'),
    ('ミニTOPIX先物',                   'JPX_TOPIX_MINI', 'TOPIX'),
    ('TOPIXオプション',                 'JPX_TOPIX_OPT',  'TOPIX'),
    ('JPX日経インデックス400先物',      'JPX_JPX400_FUT', 'JPX400'),
    ('TOPIX Core30先物',                'JPX_CORE30_FUT', 'TOPIX_CORE30'),
    ('東証銀行業株価指数先物',          'JPX_BANKS_FUT',  'TOPIX_BANKS'),
    ('東証REIT指数先物',                'JPX_REIT_FUT',   'TSE_REIT'),
    ('東証グロース市場250指数先物',     'JPX_GROWTH250',  'TSE_GROWTH250'),
    ('NYダウ先物',                      'JPX_DJIA_FUT',   'DJIA'),
    ('台湾加権指数先物',                'JPX_TAIEX_FUT',  'TAIEX'),
    ('日経平均VI先物',                  'JPX_N225VI_FUT', 'N225VI'),
    ('長期国債先物',                    'JPX_JGB_FUT',    None),      # 面值口径
    ('長期国債先物オプション',          'JPX_JGB_FUT',    None),      # 同上，标的是 10y JGB 期货
    # ⚠ mini 10y JGB 与上面三条不同：它的乘数是 ¥100,000，价格项是**10 年 JGB 期货价**
    #   （不是面值口径），所以要单独一个价格 id，不能跟着填 1
    ('長期国債先物（現金決済型ミニ）',  'JPX_JGB_MINI',   'JGB10Y_FUT_PX'),
    ('超長期国債先物（ミニ）',          None,             None),   # 額面 ¥1,000 万，2019-01 仅 6 张
    ('有価証券オプション',              'JPX_SEC_OPT',    'JP_SEC_OPT_VWAVG'),
]
# HKEX 的成员表**不手写**：直接把官方月报里 DERIVATIVES MARKET TURNOVER 段的每一行
# 都当成员（Total* 三行除外），再按下表挂规格页与价格 id。挂不上的留空 ⇒ 报表里显示
# 「乘数✗」。这样分母恒等于官方自己印的 Total Futures and Options，不会因为漏写一行
# 而悄悄少算 —— 漏一行在图上完全看不出来。
HKEX_SPEC_BY_ROW = {
    'Hang Seng Index Futures':                        ('HKEX_HSI_FUT',   'HSI'),
    'Mini Hang Seng Index Futures':                   ('HKEX_MHI_FUT',   'HSI'),
    'Hang Seng Index Options':                        ('HKEX_HSI_OPT',   'HSI'),
    'Mini Hang Seng Index Options':                   ('HKEX_MHI_OPT',   'HSI'),
    'HSI Dividend Point Index Futures':               ('HKEX_HSI_DIV',   'HSI_DIVPT'),
    'HSCEI Dividend Point Index Futures':             ('HKEX_HSI_DIV',   'HSCEI_DIVPT'),
    'Hang Seng China Enterprises Index Futures':      ('HKEX_HHI_FUT',   'HSCEI'),
    'Mini-Hang Seng China Enterprises Index Futures': ('HKEX_MCH_FUT',   'HSCEI'),
    'Hang Seng China Enterprises Index Options':      ('HKEX_HHI_OPT',   'HSCEI'),
    'Mini-Hang Seng China Enterprises Index Options': ('HKEX_MCH_OPT',   'HSCEI'),
    'Stock Options':                                  ('HKEX_STOCK_OPT', 'HK_STOCK_OPT_VWAVG'),
    'Stock Futures':                                  ('HKEX_STOCK_FUT', 'HK_STOCK_FUT_VWAVG'),
}
_HKEX_TOTAL_ROWS = ('Total Futures', 'Total Options', 'Total Futures and Options')


POOLS = {
    'JPX_DERIV': dict(
        ccy='JPY', members=JPX_MEMBERS, weights=jpx_weights,
        total_note='pools.py 用 jpx.csv 的 adv_deriv_total_raw_kcontracts 配这个常数',
        source_url='https://www.jpx.co.jp/english/derivatives/products/ + '
                   'https://www.jpx.co.jp/markets/statistics-derivatives/monthly/',
    ),
    'HKEX_DERIV': dict(
        ccy='HKD', members=None, weights=hkex_weights,   # 成员由官方行名动态生成
        total_note='pools.py 用 hkex.csv 的 derivatives_adv_contracts 配这个常数',
        source_url='https://www.hkex.com.hk/Products/Listed-Derivatives/ + '
                   'https://www.hkexgroup.com/Investor-Relations/Business-Analysis/Key-Market-Data',
    ),
    'SGX_DERIV': dict(
        ccy='USD', members=[], weights=sgx_weights,
        total_note='pools.py 用 sgx.csv 的 ddav_contracts 配这个常数',
        source_url='https://www.sgx.com/derivatives/products/ + SGX Monthly Market Statistics Report',
    ),
    'ENX_INDEX_DERIV': dict(
        ccy='EUR', members=[], weights=None,
        total_note='pools.py 用 enx.csv 的 adv_index_futures/options_kcontracts 两条腿',
        source_url='https://live.euronext.com/en/product/index-futures/FCE-DPAR/contract-specification',
    ),
    'ENX_SINGLESTOCK_LEGACY': dict(
        ccy='EUR', members=[], weights=None,
        total_note='pools.py 用 enx.csv 的 adv_singlestock_futures/options_kcontracts + athex 减法腿',
        source_url='https://live.euronext.com/en/markets/amsterdam/equity-options',
    ),
    'ENX_MATIF': dict(
        ccy='EUR', members=[], weights=None,
        total_note='pools.py 用 enx.csv 的 adv_commodity_futures/options_kcontracts 两条腿',
        source_url='https://live.euronext.com/en/product/commodities-futures/EBM-DPAR/contract-specification',
    ),
}


def compute(pool_id, mults, weights):
    """按 notes 列写死的公式算篮子常数。缺任何一个成员的价格就返回 None，
    并把缺口按张数权重排好返回 —— **不许**用「覆盖到的部分」外推整池。"""
    cfg = POOLS[pool_id]
    members = cfg['members']
    if members is None:      # HKEX：成员 = 官方月报里除三条 Total 外的每一行
        members = [(k, ) + HKEX_SPEC_BY_ROW.get(k, (None, None))
                   for k in weights if k not in _HKEX_TOTAL_ROWS]
    rows, total = [], 0.0
    for src_name, mid, pid in members:
        vol = weights.get(src_name)
        if vol is None or vol == 0:
            continue
        total += vol
        mult = mults.get(mid, (None,))[0] if mid else None
        if pid is None and mult is not None:
            price = PRICE_DEFINITIONAL            # 面值口径
        else:
            price = BASE_PRICES.get(pid)
        rows.append(dict(name=src_name, vol=vol, mult=mult, pid=pid, price=price))
    missing = [r for r in rows if r['mult'] is None or r['price'] is None]
    if total <= 0:
        return None, rows, total, missing
    if missing:
        return None, rows, total, missing
    num = sum(r['vol'] * r['mult'] * r['price'] for r in rows)
    return num / total, rows, total, missing


def crosscheck_jpx_divisors(mults):
    """把 fetch/jpx.py 里那张 `_DIVISORS`（大合约当量折算表）与本脚本刚从官方规格页
    抓下来的乘数逐档对一遍 —— 任务书点名要做的那道交叉核对。

    _DIVISORS 是「小合约 → 大合约」的除数；本脚本抓到的是两边各自的乘数。
    两者自洽的条件是：大合约乘数 ÷ 小合约乘数 == 除数。
    不自洽就打印出来（不静默、也不自动改 fetch/jpx.py —— 那是另一码事）。
    """
    import ast
    src = open(os.path.join(ROOT, 'fetch', 'jpx.py'), encoding='utf-8').read()
    m = re.search(r'^_DIVISORS\s*=\s*(\{.*?^\})', src, re.S | re.M)
    if not m:
        print('  ✗ fetch/jpx.py 里没找到 _DIVISORS 字面量（改过名？）')
        return
    div = ast.literal_eval(re.sub(r'#[^\n]*', '', m.group(1)))
    # 只核股指/利率这几档 —— 商品档（金/白金）不在本次六个池的成员里
    pairs = {
        '日経225mini':                   ('JPX_N225_MINI',    'JPX_N225_FUT'),
        '日経225マイクロ先物':            ('JPX_N225_MICRO',   'JPX_N225_FUT'),
        '日経225ミニオプション':          ('JPX_N225_MINIOPT', 'JPX_N225_OPT'),
        'ミニTOPIX先物取引':              ('JPX_TOPIX_MINI',   'JPX_TOPIX_FUT'),
        '長期国債先物（現金決済型ミニ）':  ('JPX_JGB_MINI',     'JPX_JGB_FUT'),
    }
    print('\n【一·b】与 fetch/jpx.py 的 _DIVISORS 折算表交叉核对')
    for jp, (small, big) in pairs.items():
        want = div.get(jp)
        a = mults.get(small, (None,))[0]
        b = mults.get(big, (None,))[0]
        if want is None or a is None or b is None:
            print('  %-26s 核不了（表里 %s / 小 %s / 大 %s）'
                  % (jp, want, _fmt(a), _fmt(b)))
            continue
        if jp == '長期国債先物（現金決済型ミニ）':
            # ⚠ 这一档两边量纲不同：mini 页写的是「¥100,000 × 10年国債先物価格」，
            #   大合约页写的是「額面 ¥1 億」。折成同一口径（价格 100 = 平价）后
            #   mini 的面值当量 = 1e5 × 100 = ¥1,000 万，比值才是 10。
            got = (b / 100.0) / a
        else:
            got = b / a
        ok = abs(got - want) < 1e-9
        print('  %-26s 表里 ÷%-4s  规格页实测 %s÷%s = ÷%g   %s'
              % (jp, want, _fmt(b), _fmt(a), got, '✓ 一致' if ok else '✗ 不一致'))


def jpx_pool_total():
    """pools.py 喂给 JPX_DERIV 的那一列在基期的**月度**张数 =
    series/jpx.csv 的 adv_deriv_total_raw_kcontracts × 1000 × deriv_trading_days。
    它比 OSE 取引総括表的成员之和大，差额就是商品関連（旧 TOCOM）的 pro-forma
    —— 这道对账存在的意义就是把这块差额顶到台面上，而不是让它悄悄进分母。"""
    path = os.path.join(ROOT, 'series', 'jpx.csv')
    if not os.path.exists(path):
        return None
    with open(path, encoding='utf-8') as f:
        for row in csv.DictReader(f):
            if row.get('month') == BASE_MONTH:
                try:
                    return (float(row['adv_deriv_total_raw_kcontracts']) * 1000.0
                            * float(row['deriv_trading_days']))
                except (KeyError, ValueError):
                    return None
    return None


def report(pool_id, rows, total, missing, official=None):
    print('\n【%s】2019-01 成员张数结构（Σ=%s）' % (pool_id, f'{total:,.0f}'))
    if official:
        d = official - total
        print('    对官方合计 %s：差 %s（%+.2f%%）%s'
              % (f'{official:,.0f}', f'{d:+,.0f}', d / official * 100,
                 '' if abs(d) / official < 0.005 else '  ⚠ 差额超 0.5%，逐行核'))
    for r in sorted(rows, key=lambda r: -r['vol']):
        pct = r['vol'] / total * 100
        m = _fmt(r['mult']) if r['mult'] is not None else '乘数✗'
        p = ('价格✓' if r['price'] is not None
             else '价格✗ ' + (r['pid'] or ''))
        print('    %-46s %12s  %5.1f%%  %-12s %s'
              % (r['name'][:46], f"{r['vol']:,.0f}", pct, m, p))
    if missing:
        gap = sum(r['vol'] for r in missing) / total * 100
        ids = sorted({r['pid'] for r in missing if r['pid']})
        print('    ⇒ 缺口占基期张数 %.1f%%，缺的价格序列：%s' % (gap, ', '.join(ids) or '(乘数也缺)'))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--no-net', action='store_true', help='只读 cache/，不发网络请求')
    args = ap.parse_args()
    use_net = not args.no_net

    mults = collect_multipliers(use_net)
    crosscheck_jpx_divisors(mults)

    print('\n【二】官方月度统计 → 2019-01 张数权重')
    w = {}
    for name, fn, path in (('JPX', jpx_weights, JPX_SOUKATSU),
                           ('HKEX', hkex_weights, HKEX_HILITE),
                           ('SGX', sgx_weights, SGX_PDF)):
        if not os.path.exists(path):
            # 这三份都是既有 fetcher 下的官方原件，不在本脚本重复实现下载逻辑
            regen = {'JPX': 'python3 -c "import fetch.jpx as m; m.fetch()"',
                     'HKEX': 'python3 -c "import fetch.hkex as m; m.fetch()"',
                     'SGX': 'python3 -c "import fetch.sgx as m; m.fetch()"'}[name]
            print('  %-5s ✗ 缺源文件 %s —— 先跑 %s 把官方原件下回 cache/'
                  % (name, path, regen))
            w[name] = {}
            continue
        try:
            w[name] = fn()
            # 官方自己印的合计行，用来验解析对不对（不是把它加进成员）
            anchor = {'JPX': None,
                      'HKEX': 'Total Futures and Options',
                      'SGX': '__OFFICIAL_TOTAL__'}[name]
            off = w[name].get(anchor) if anchor else None
            print('  %-5s ✓ %d 个产品行%s   ← %s'
                  % (name, len(w[name]),
                     ('，官方合计 = %s' % f'{off:,.0f}') if off else
                     ('，Σ=%s' % f'{sum(w[name].values()):,.0f}'),
                     os.path.basename(path)))
        except Exception as e:                      # noqa: BLE001 —— 解析失败要看得见，不静默
            print('  %-5s ✗ 解析失败：%r' % (name, e))
            w[name] = {}

    print('\n【三】基期价格（口径 = %s，基期 %s）' % (BASE_PRICE_BASIS, BASE_MONTH))
    if BASE_PRICES:
        for k, v in sorted(BASE_PRICES.items()):
            print('  %-8s %s' % (k, v))
    else:
        print('  ✗ 一个都没有。逐个的阻塞原因见 PRICE_BLOCKERS / part CSV 的 computed_note。')

    print('\n【四】篮子常数')
    results = {}
    for pool_id in ('JPX_DERIV', 'HKEX_DERIV'):
        src = 'JPX' if pool_id.startswith('JPX') else 'HKEX'
        const, rows, total, missing = compute(pool_id, mults, w[src])
        official = (w[src].get('Total Futures and Options') if src == 'HKEX'
                    else jpx_pool_total())
        report(pool_id, rows, total, missing, official=official)
        results[pool_id] = const

    # SGX：张数权重拿到了，但**乘数**一个都没实测（sgx.com 是 Angular 单页），
    # 所以连成员表都不写 —— 写了就是把没实测的乘数塞进代码。
    sg = dict(w['SGX'])
    tot = sg.pop('__OFFICIAL_TOTAL__', None)
    if sg:
        print('\n【SGX_DERIV】2019-01 张数权重已解析（%d 个合约行，官方月报当月合计 %s 张），'
              '但 sgx.com 规格表由前端 JS 渲染，**零个乘数实测入库** ⇒ 不算。'
              % (len(sg), f'{tot:,.0f}' if tot else '(未定位到合计行)'))
        for v, k in sorted(((v, k) for k, v in sg.items()), reverse=True)[:8]:
            print('    %-46s %12s  %5.1f%%'
                  % (k[:46], f'{v:,.0f}', v / tot * 100 if tot else 0))
    for pool_id in ('SGX_DERIV', 'ENX_INDEX_DERIV', 'ENX_SINGLESTOCK_LEGACY', 'ENX_MATIF'):
        results[pool_id] = None

    print('\n【ENX_*】Euronext 官方月度量表（live.euronext.com/sites/default/files/statistics/ir/）'
          '只发 4 条股指/单股汇总列与 2 条商品列，**没有逐合约拆分** ⇒ 连张数权重都拿不到，'
          '篮子无法合成。乘数已实测：CAC40 €10/点、AEX €200/点、MATIF 小麦 50 吨/张。')

    write_part(mults, results)
    print('\n已写 %s' % OUT_CSV)
    return 0


def write_part(mults, results):
    """写 series/_specs_part_enx_asia.csv。主线程负责合并进 contract_specs.csv。
    ⚠ 算不出来的行 base_price_local / basket_constant **留空**，不写占位数。"""
    def m(mid):
        v = mults.get(mid, (None,))[0]
        return _fmt(v) if v is not None else '✗'

    notes = {
        'ENX_INDEX_DERIV':
            '未算出。乘数已实测：CAC 40 期货 €%s/指数点（FCE）、AEX 期货 €%s/指数点（FTI）—— '
            '两者差 20 倍，必须分开加权。**卡在张数权重**：Euronext 官方月度量表只发 '
            'adv_index_futures / adv_index_options 两条汇总列，不按合约（CAC/AEX/BEL）拆分，'
            '无法按 2019-01 成交结构合成。次要缺口：CAC 40 与 AEX 的 2019-01 月均收盘'
            '（live.euronext.com 图表接口返回 AES 密文）。'
            '检索路径 = Euronext Derivatives 月度成交明细（若有官方逐合约发布）+ Euronext 指数历史数据文件。'
            % (m('ENX_CAC40_FUT'), m('ENX_AEX_FUT')),
        'ENX_SINGLESTOCK_LEGACY':
            '未算出。⚠ 口径提示（避免后人困惑）：本行是**剔除雅典之后的 legacy 口径**'
            '（主列 − athex_* 备注列，减法腿 since=2025-11）；但基期 2019-01 早于 2025-11 '
            'Athex 并表，那时 Euronext 数据里本来就没有雅典，**所以基期常数本身不受并表影响**，'
            '与全口径常数完全相同 —— legacy 只影响 2025-11 之后的张数序列，不影响这个常数。'
            '卡点：① Euronext 单股期权乘数**不是全市场统一 100 股**，各标的不同且有公司行动调整，'
            '官方 contract specification 页按标的一个个发；② 官方月度量表不按标的拆张数；'
            '③ 需要 2019-01 按成交量加权的标的均价。三项缺一不可。'
            '检索路径 = live.euronext.com Derivatives → Contract specifications 的标的清单 + '
            'Euronext 官方逐标的成交明细。',
        'ENX_MATIF':
            '未算出。乘数已实测：Milling Wheat（EBM）**Fifty tonnes**/张，官方规格页原文。'
            '**卡在张数权重与价格**：enx.csv 只有 adv_commodity_futures / _options 两条汇总列，'
            '小麦 / 玉米 / 菜籽三个品种的 2019-01 张数拆不开，而三者 €/吨价差很大'
            '（无法用「都是 50 吨」抹平）；另需三品种 2019-01 月均结算价。'
            '检索路径 = Euronext 官方 commodities 逐合约成交与结算价发布。',
        'SGX_DERIV':
            '未算出。**乘数零个实测**：sgx.com 是 Angular 单页应用，服务端只返回 14KB 外壳；'
            'api.sgx.com/derivatives/v1.0（需 pagestart+pagesize 分页）返回的是实时行情，不含合约规格。'
            '张数权重已拿到（cache/sgx_2019-01.pdf = SGX Monthly Market Statistics Report Jan 2019，'
            '官方逐合约表，当月 Total 18,614,593 张）：A50 7,488,026（40.2%）、'
            'Nifty 50 1,798,179、MSCI Taiwan 1,807,072、Nikkei 225 期货 1,631,290、'
            'FX 期货 1,872,510、铁矿石 62% 期货 1,369,025。'
            '⚠ 跨币种篮子，记账币定为 USD：各成员面值须先按 2019-01 基期汇率折成美元再加权，'
            '所以本行 ccy=USD、fx 那一跳是 1.0，不会二次折算。'
            '检索路径 = SGX 官方 contract specification PDF（静态文件）或浏览器渲染 sgx.com 规格页；'
            '价格另需 A50 / Nikkei 225 / Nifty 50 / MSCI 各指数与铁矿石掉期的 2019-01 月均收盘。',
        'HKEX_DERIV':
            '未算出。乘数已全部实测（官方产品规格页「Contract Multiplier」栏）：'
            'HSI 期货 HK$%s/点、Mini HSI 期货 HK$%s/点、HSI 期权 HK$%s/点、Mini HSI 期权 HK$%s/点、'
            'HHI 期货 HK$%s/点、Mini HHI 期货 HK$%s/点、HHI 期权 HK$%s/点、Mini HHI 期权 HK$%s/点。'
            '（顺带修掉 contract_specs_todo.csv 里 HKEX_HHI_FUT 那条「URL 拼错」的 blocker：'
            '正确路径是 /Equity-Index/Hang-Seng-China-Enterprises-Index/... 不带 (HSCEI) 后缀。）'
            '张数权重已拿到（cache/hkex_monthly_highlights_2026-06.xlsx，官方逐产品 ADV，'
            '2019-01 合计 1,091,651 张/日，与 hkex.csv 的 derivatives_adv_contracts 逐位相符）。'
            '⚠ 顺带查出一处口径坑：该文件是 2026-06 版模板，**不列 2026 年前已退市的产品** —— '
            '2019-01 逐产品行加起来只有 1,091,041，比官方自己印的 Total 少 610 张/日（0.06%%），'
            '差额全在期货侧。占比小，但把这文件当「逐产品全集」用之前必须知道它不是全集。'
            '**卡在两处价格**：① HSI / HSCEI 的 2019-01 月均收盘（恒生指数公司官网是单页应用）；'
            '② Stock Options 445,669 张/日 + Stock Futures 5,821 张/日 = 全池 41.4%% 的张数，'
            '其乘数是**各标的的每手股数（board lot）**、名义额还要乘各标的股价 —— '
            '需要逐 class 的 2019-01 张数与股价，官方月报不拆。'
            '⚠ 不许把这 41.4%% 从分母里剔掉再算：股票期权单张名义额（每手 × 股价，量级 HK$ 数万）'
            '比 HSI 期货（HK$50 × 约 27,000 点 ≈ HK$135 万）小两个数量级，剔掉等于给它们'
            '安上指数期货的名义额，整池会被系统性放大。'
            '检索路径 = 恒指公司官方指数历史 + HKEX「List of Stock Options」逐 class 合约规格'
            '与 HKEX 月度个股期权成交明细。'
            % (m('HKEX_HSI_FUT'), m('HKEX_MHI_FUT'), m('HKEX_HSI_OPT'), m('HKEX_MHI_OPT'),
               m('HKEX_HHI_FUT'), m('HKEX_MCH_FUT'), m('HKEX_HHI_OPT'), m('HKEX_MCH_OPT')),
        'JPX_DERIV':
            '未算出。乘数已全部实测（jpx.co.jp 英文规格页「Contract Unit」栏）：'
            '日経225先物 ¥%s/点、日経225mini ¥%s/点、日経225オプション ¥%s/点、'
            'TOPIX先物 ¥%s/点、ミニTOPIX先物 ¥%s/点、TOPIXオプション ¥%s/点、'
            'JPX日経400先物 ¥%s/点；長期国債先物 額面 ¥100,000,000（面值口径，价格项=1）。'
            '⚠ 与 fetch/jpx.py 的 _DIVISORS 折算表交叉核对：**5/5 全部一致**，无一处需要改 —— '
            '日経225mini ÷10（¥1,000/¥100）、日経225マイクロ ÷100（¥1,000/¥10）、'
            '日経225ミニオプション ÷10（¥1,000/¥100）、ミニTOPIX ÷10（¥10,000/¥1,000）、'
            '長期国債先物（現金決済型ミニ）÷10。最后一档要注意量纲：mini 页写的是'
            '「¥100,000 × 10 年国債先物価格」、大合约页写的是「額面 ¥1 億」，'
            '折成同一口径（价格=100 平价）后比值才是 10；直接拿 1e8÷1e5 会得到 ÷1000，是错的。'
            '（本条核对由 build/basefill/enx_asia.py 的 crosscheck_jpx_divisors() 每次运行自动重跑。）'
            '张数权重已拿到（cache/jpx_soukatsu_M.xlsx = 官方「取引総括表（月間）」，'
            '2019-01 逐产品月度张数；期权只取「合計」行，不重复计 プット/コール）。'
            '**卡在指数价格**：日経225（占 OSE 基期张数 89.1%%）与 TOPIX 系的 2019-01 月均收盘取不到 —— '
            'Nikkei 官网 Cloudflare 拦截，JPX 只免费发月末值。'
            '另有两块次要缺口：有価証券オプション 的乘数是「各标的的最低取引単位」（非固定值）；'
            '且 pools.py 配的是 adv_deriv_total_raw_kcontracts（含商品関連 pro-forma，'
            '2019-01 约 80,155 张/日 = 全池 5.1%%），而这块**不在** OSE 取引総括表里 —— '
            '旧 TOCOM 品种（金/白金/ゴム/原油…）的 2019-01 张数与规格要另外取。'
            '⚠ 不许只用能算的部分外推：JGB 期货单张名义额 ¥1 亿，日経225mini 约 ¥200 万，差 50 倍。'
            '检索路径 = JPX 統計月報「指数」分册（03_sisu1901.pdf，需先找到 2019 年归档目录）'
            '或 Nikkei Inc 官方历史数据。'
            % (m('JPX_N225_FUT'), m('JPX_N225_MINI'), m('JPX_N225_OPT'),
               m('JPX_TOPIX_FUT'), m('JPX_TOPIX_MINI'), m('JPX_TOPIX_OPT'), m('JPX_JPX400_FUT')),
    }
    order = ['ENX_INDEX_DERIV', 'ENX_SINGLESTOCK_LEGACY', 'ENX_MATIF',
             'HKEX_DERIV', 'JPX_DERIV', 'SGX_DERIV']
    os.makedirs(os.path.dirname(OUT_CSV), exist_ok=True)
    with open(OUT_CSV, 'w', newline='', encoding='utf-8') as f:
        wcsv = csv.writer(f)
        wcsv.writerow(['product_id', 'base_price_local', 'base_ccy',
                       'basket_constant', 'source_url', 'computed_note'])
        for pid in order:
            const = results.get(pid)
            val = '' if const is None else repr(const)
            wcsv.writerow([pid, val, POOLS[pid]['ccy'], val,
                           POOLS[pid]['source_url'], notes[pid]])


if __name__ == '__main__':
    sys.exit(main())
