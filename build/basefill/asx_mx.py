# -*- coding: utf-8 -*-
"""ASX 与 Montréal(TMX) 六个 pool_product 篮子常数的**可复算**取数脚本。

━━ 这个文件为什么存在 ━━
series/contract_specs.csv 里既有的 23 行常数是手工敲进去的，取数代码没留在仓里、
cache/ 又进了 .gitignore —— 仓库自己没法复算，验收员点名这是缺陷。
本脚本负责 ASX_DERIV / ASX_ETO / MX_STIR / MX_BOND / MX_EQUITY_OPT / MX_ETF_OPT 六行，
**每一个数字都能从官方一手 URL 现场重跑出来**，跑完写 series/_specs_part_asx_mx.csv。
（不直接改 contract_specs.csv：多个 agent 并行写同一文件会互相覆盖，由主线程合并。）

    python3 build/basefill/asx_mx.py            # 用 cache/basefill/ 下的副本，没有才下载
    python3 build/basefill/asx_mx.py --refresh  # 强制重下所有官方文件

━━ 篮子常数的定义（照 contract_specs.csv 的 notes 列）━━
    篮子常数 = Σ(成员 2019-01 张数 × 乘数 × 基期价格) ÷ Σ(成员 2019-01 张数)
即按基期成交量加权，把一个品类里多个合约折成一个「等效单张名义额」。只算一次写死。
pool_product 行的 multiplier 恒为 1，所以这个常数直接就是 base_price_local
（build/notional.py: base_notional_per_unit_local = multiplier × base_price_local）。

━━ 两套名义额算法，加权前必须先各自换成同一货币的单张名义额 ━━
  · 股指 / 股票 / ETF：单张名义额 = 乘数 × 基期标的价格（价格项是真价格）
  · 利率合约（BAX / CGB 系 / ASX 国债与票据）：单张名义额 = **面值**，不乘结算价。
    面值是市场惯例的名义额定义；这些合约报价是「100 − 收益率」，那个报价乘以面值
    没有任何经济含义。代价必须留痕：**名义额不反映利率风险敞口**——
    同样 1 亿名义额，2 年期与 10 年期的 DV01 差 5 倍以上。本池占比只能读作
    「名义额构成」，不可读作「谁承担了更多利率风险」。
    （与 build/notional.py 模块 docstring 的同一条告诫保持一致。）

━━ 本次实测结果：6 个里只填得出 2 个 ━━
填出来的（两个都是纯利率池，常数就是面值加权，与标的价格无关）：
    MX_STIR  C$1,000,000   2019-01 该池只有 BAX 有量（CRA = 0、Others = 0）
    MX_BOND  C$100,000     CGF 179,194 + CGB 2,070,889，两者面值同为 C$100,000；
                           CGZ / LGB 当月成交量均为 0，权重为 0

填不出来的四个，卡点各不相同，逐条写在 _BLOCKERS 里，也写进输出 CSV 的 computed_note：
    ASX_DERIV       2019-01 **分品种成交量**在 asx.com.au 上不存在（下方「坑 1」实测）
    ASX_ETO         量有了、价只有月初月末两个点，凑不出 avg_close（「坑 2」）
    MX_EQUITY_OPT   需要逐个期权类的 2019-01 成交量与标的月均收盘；
    MX_ETF_OPT      官方通道存在但要 940 次请求 / robots.txt Crawl-delay: 15（「坑 3」）

━━ 坑 1：ASX 分品种月度成交量只回溯到 2020-04，2019-01 永久缺 ━━
ASX 的 MAR（月度经营报告）只给期货**合计**，正文里印了一条分品种报告的链接：
    http://www.asx.com.au/data/market-reports/MonthlyFuturesMarketsReport190131.pdf
    （2019-01 那期 MAR 第 4 页原文，见 cache/asx_mar_2019-01.pdf）
这条链接今天返回 **HTTP 200 + text/html + 恒定 136,750 字节的错误页**，不是 404 ——
任何 `if status == 200: save()` 都会把 HTML 当 PDF 存下来。同一份外壳页在
www2.asx.com.au、/data/futures/reports/ 下都一样。
现行的分品种月报走另一个命名（monthly-trading-report 页面的 JS 里写死）：
    https://www.asx.com.au/data/futures/reports/MonthlyWebTotalVolume{YYMM}S.htm
本次逐月实测 1801 → 2012 共 36 个月：**2004（2020-04）是最早的真文件**，
1801 → 2003 全部返回那张 136,750 字节外壳页。所以 2019-01 拿不到分品种量。
（有意思的是同目录的 MonthlyWebNonTradedVolume1901S.htm 是**真的**、6,539 字节，
ASX 对两类报告的留存策略不一样 —— 但非成交量（交割/行权）不是我们要的权重。）
另外 ASX_DERIV 在 build/pools.py 里挂的是 adv_futures_and_options_contracts，
成员是**整个 ASX 24**（利率 + 股指 + 商品 + 电力 + NZ + 期货期权），不止题面点名的 4 个合约，
所以就算拿到那 4 条也凑不出完整权重。

━━ 坑 2：ASX 单股期权的量能拿全，价只有两个点 ━━
ASX 官方的 annual-market-summary-2019.xls（equity-derivatives-statistics 栏目）
January 表逐类给出 Lots Traded（Call/Put）与**两个**价格：月初、月末各一个收盘价。
本次实测：161 个类的 Call+Put 合计 = 5,953,660，与官方
derivatives-equity-and-index-options-jan2019.pdf 第 7 页的
Equity 5,156,054 + Equity LEPO 51,180 + Index 746,164 + Index LEPO 262 **逐位相符**；
剔掉 XJO / XJO* / XJOL 三个指数类之后 = 5,207,234，与 MAR 的
「Single stock equity options 5,207,234」也逐位相符 ⇒ 成员集合与权重都是对的。
卡在价格口径：表内既有 23 行一律 `base_price_basis = avg_close`（2019-01 月均收盘），
而 ASX 对个股只发月初/月末两个收盘价，没有日频序列（历史行情是收费产品）。
本脚本照样把三个候选值算出来打印（见 _asx_eto_candidates），但**不写进 CSV**：
写进去等于在一张统一 avg_close 的表里混一个 (P_first+P_last)/2 的口径。
要用就得主线程明确同意换口径，并把 base_price_basis 一起改掉。

━━ 坑 3：MX 逐类期权数据能拿，但代价是 940 次请求 ━━
m-x.ca 的历史数据接口是纯 GET，能回溯到 2009-01-02
（日期下限写死在 /assets/application/js/historical-data.*.min.js 里的 `new Date(2009,0,2)`）：
    https://www.m-x.ca/en/trading/data/historical?symbol=BNS&from=2019-01-02&to=2019-01-31&dnld=1
本次实测该 URL 返回 text/csv、1,326,135 字节，含每个期权系列的逐日 Volume，
**并且含标的自己那一行**（Class Symbol 为空、Ins. Type = 3，Last Price 就是股票收盘价）
—— 也就是说逐类成交量与标的月均收盘两样都在这一个响应里。
但期权类下拉框里有 **940 个 symbol**，m-x.ca 的 robots.txt 写着 `Crawl-delay: 15`，
按规矩跑一遍要 ≈4 小时、约 1.2 GB。这超出本仓无人值守的边界，也不该塞进一次交付。
⇒ 留空，把可复算的抓取函数 `mx_option_class_month()` 一并留在这里，
主线程要填时跑一次一次性回补即可（见文件末尾的 __main__ 提示）。

━━ 依赖 ━━ 标准库 + openpyxl（MX 月报 xlsx）+ xlrd（ASX 年报 xls，老式 BIFF）
+ PyMuPDF（合约规格 PDF）。三个都是仓里既有脚本已经在用的。
"""

import argparse
import csv
import io
import os
import re
import sys
import urllib.request

_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_CACHE = os.path.join(_ROOT, 'cache', 'basefill')
_OUT = os.path.join(_ROOT, 'series', '_specs_part_asx_mx.csv')

BASE_MONTH = '2019-01'

# ── 官方一手 URL，一个都不许换成第三方聚合站 ────────────────────────────
SRC = {
    # MX 合约规格（现行页）—— m-x.ca 裸 urllib 即 200，无 UA 校验、无 JS 渲染
    'mx_cgb': 'https://www.m-x.ca/en/markets/interest-rate-derivatives/cgb',
    'mx_cgf': 'https://www.m-x.ca/en/markets/interest-rate-derivatives/cgf',
    'mx_cgz': 'https://www.m-x.ca/en/markets/interest-rate-derivatives/cgz',
    'mx_cra': 'https://www.m-x.ca/en/markets/interest-rate-derivatives/cra',
    'mx_eqopt': 'https://www.m-x.ca/en/markets/equity-derivatives/equity-options',
    'mx_etfopt': 'https://www.m-x.ca/en/markets/equity-derivatives/etf-options',
    'mx_sxf': 'https://www.m-x.ca/en/markets/index-derivatives/sxf',
    # BAX 已于 2024 年退市，现行产品页 404；规格取 MX 自家**存档规则书**
    # Rule Fifteen «Futures Contracts Specifications»，末次修订 14.09.18，
    # 正是 2019-01 当时生效的版本（比现行页更贴基期）。
    'mx_rule15': 'https://www.m-x.ca/f_regles_en/15_en.pdf',
    # MX 官方月度统计（2019-01 逐合约成交量）
    'mx_stats_1901': 'https://www.m-x.ca/f_stat_en/1901_stats_en.xlsx',
    # ASX 24 合约规格（SPI 200 / 国债 / 银行票据的乘数与面值）
    'asx24_specs': ('https://www.asx.com.au/content/dam/asx/participants/'
                    'derivatives-market/ird/asx24-contract-specifications.pdf'),
    # ASX 现行规格页（服务端渲染，正文里就有规格表）
    'asx_bond': ('https://www.asx.com.au/markets/trade-our-derivatives-market/'
                 'overview/interest-rate-derivatives/bond-derivatives'),
    'asx_st': ('https://www.asx.com.au/markets/trade-our-derivatives-market/'
               'overview/interest-rate-derivatives/short-term-derivatives'),
    'asx_eto_spec': ('https://www.asx.com.au/markets/trade-our-derivatives-market/'
                     'overview/equity-derivatives/options-contract-specifications'),
    # ASX 单股期权 2019 全年逐类量 + 月初月末价
    'asx_eto_2019': ('https://www.asx.com.au/content/dam/asx/participants/'
                     'derivatives-market/equity-derivatives/'
                     'equity-derivatives-statistics/2019/annual-market-summary-2019.xls'),
    # 对账用：ASX 官方自己给的 2019-01 期权分项合计
    'asx_eqderiv_1901': ('https://www.asx.com.au/content/dam/asx/participants/'
                         'derivatives-market/equity-derivatives/'
                         'equity-derivatives-statistics/2019/'
                         'derivatives-equity-and-index-options-jan2019.pdf'),
}

# asx.com.au 的 soft-404：HTTP 200 + text/html + 这几个恒定字节数的通用外壳页。
# 见模块 docstring 坑 1。任何下载都必须过这一关，否则错在下游报得莫名其妙。
_ASX_SHELL_SIZES = {136750, 136883, 136884, 136885}


def _fetch(key, refresh=False):
    """下载并缓存官方文件；对 asx.com.au 的 soft-404 显式报错。"""
    url = SRC[key]
    os.makedirs(_CACHE, exist_ok=True)
    ext = '.pdf' if url.endswith('.pdf') else (
        '.xlsx' if url.endswith('.xlsx') else (
            '.xls' if url.endswith('.xls') else '.html'))
    # 文件名加前缀：cache/basefill/ 是多个 basefill 脚本共用的目录，
    # 不加前缀会跟别人的 asx24_specs.pdf / mx_cgb.html 撞名。
    path = os.path.join(_CACHE, 'asxmx_' + key + ext)
    if os.path.exists(path) and not refresh:
        with open(path, 'rb') as fh:
            return fh.read()
    with urllib.request.urlopen(url, timeout=90) as resp:
        blob = resp.read()
        ctype = resp.headers.get('Content-Type', '')
    if len(blob) in _ASX_SHELL_SIZES:
        raise RuntimeError(
            '%s 返回 asx.com.au 的通用外壳页（%d 字节，soft-404，不是真内容）：%s'
            % (key, len(blob), url))
    if ext == '.pdf' and blob[:5] != b'%PDF-':
        raise RuntimeError('%s 不是 PDF（Content-Type=%s，首 5 字节=%r）：%s'
                           % (key, ctype, blob[:5], url))
    with open(path, 'wb') as fh:
        fh.write(blob)
    return blob


def _text(blob):
    """HTML → 纯文本。只用来做「规格页上写着这个数」的字面断言。"""
    t = blob.decode('utf-8', 'replace')
    t = re.sub(r'<script.*?</script>', ' ', t, flags=re.S)
    t = re.sub(r'<style.*?</style>', ' ', t, flags=re.S)
    t = re.sub(r'<[^>]+>', '\n', t)
    for a, b in (('&amp;', '&'), ('&nbsp;', ' '), ('&#36;', '$'), ('&quot;', '"')):
        t = t.replace(a, b)
    return re.sub(r'[ \t]+', ' ', t)


def _pdf_text(blob):
    import fitz
    doc = fitz.open(stream=blob, filetype='pdf')
    return '\n'.join(doc[i].get_text() for i in range(doc.page_count))


def _need(hay, needle, where):
    """断言：官方页面上必须逐字出现这句话。页面改版时在这里炸，而不是在下游算错。"""
    flat = re.sub(r'\s+', ' ', hay)
    if re.sub(r'\s+', ' ', needle) not in flat:
        raise AssertionError('%s 的官方页面里没找到 %r —— 页面改版了，先核对再改常数'
                             % (where, needle))
    return needle


# ══════════════════════ 1. 合约规格（乘数 / 面值）══════════════════════
def mx_specs(refresh=False):
    """MX 各合约的单张名义额（CAD）。利率合约取面值，指数/股票取乘数。"""
    out = {}
    rule15 = _pdf_text(_fetch('mx_rule15', refresh))
    # Rule 15 s.15500 / 15503(b)：BAX 面值 C$1,000,000
    _need(rule15, "The underlying issue for a Three-month Canadian Bankers’ Acceptance "
                  "futures is $1,000,000 nominal value", 'MX Rule 15 s.15500')
    _need(rule15, "a bankers' acceptance having a nominal value of $1,000,000 with a "
                  "Three-month maturity", 'MX Rule 15 s.15503(b)')
    out['BAX'] = 1_000_000.0
    # Rule 15 s.15642 / 15622 / 15662：CGB / CGF / LGB 面值均为 C$100,000
    for code, sec in (('CGB', 'Ten-year'), ('CGF', 'Five-year'), ('LGB', 'Thirty-year')):
        _need(rule15, 'the trading unit is $100,000 nominal value of a Government of '
                      'Canada bond with a 6% notional coupon', 'MX Rule 15 %s' % sec)
        out[code] = 100_000.0
    # ⚠ CGZ 是唯一一处「存档规则书」与「现行产品页」不一致的合约：
    #   存档 Rule 15 s.15603（末修 18.01.16）写的是 $200,000，现行 /cgz 页写 C$100,000。
    #   2019-01 CGZ 成交量为 0，权重为 0，取哪个都不改变 MX_BOND 的常数；
    #   但**不许**因为「反正是 0」就随手写现行值 —— 基期口径一律取基期在效的那版。
    _need(rule15, 'the trading unit is $200,000 nominal value of a Government of Canada '
                  'bond with a 6% notional coupon', 'MX Rule 15 s.15603 (CGZ, 2019 在效版)')
    out['CGZ'] = 200_000.0
    # 现行产品页只用于交叉核对当前值，不参与基期计算
    _need(_text(_fetch('mx_cgb', refresh)),
          'C$100,000 nominal value of a Government of Canada bond', 'm-x.ca CGB 页')
    _need(_text(_fetch('mx_cgf', refresh)),
          'C$100,000 nominal value of a Government of Canada bond', 'm-x.ca CGF 页')
    _need(_text(_fetch('mx_cgz', refresh)),
          'C$100,000 nominal value of a Government of Canada bond', 'm-x.ca CGZ 页')
    # CORRA（CRA）：3 个月期，每 bp = C$25 ⇒ 面值 C$1,000,000（1e6 × 0.0001 × 0.25 = 25）
    _need(_text(_fetch('mx_cra', refresh)),
          'each basis point per annum of interest = $25 per contract', 'm-x.ca CRA 页')
    out['CRA'] = 1_000_000.0
    # 股票 / ETF 期权：100 股（份）/张
    _need(_text(_fetch('mx_eqopt', refresh)),
          'The trading unit is one contract, representing 100 shares of the underlying '
          'Equity Security', 'm-x.ca 单股期权页')
    _need(_text(_fetch('mx_etfopt', refresh)),
          'One contract represents 100 units of an exchange-traded fund',
          'm-x.ca ETF 期权页')
    out['EQ_OPT_SHARES'] = 100.0
    out['ETF_OPT_UNITS'] = 100.0
    # SXF：C$200 × 指数（本次没用到，留着是因为 MX 指数池将来要走同一套）
    _need(_text(_fetch('mx_sxf', refresh)),
          'C$200 times the S&P/TSX 60 Index', 'm-x.ca SXF 页')
    out['SXF_PER_POINT'] = 200.0
    return out


def asx_specs(refresh=False):
    """ASX 各合约的单张名义额（AUD）。

    ⚠ 口径提醒：asx24-contract-specifications.pdf 是 **Version 11 / Updated 2/12/2025**，
    不是 2019 年的版本。ASX 没有公开发布规格文件的历史版本，所以这里拿到的是「现行值」。
    对本次交付无实害（ASX 两行都没填），但将来若要拿它去填基期常数，
    必须先自行确认该合约在 2019-01 的乘数与今天相同。
    """
    out = {}
    specs = _pdf_text(_fetch('asx24_specs', refresh))
    _need(specs, 'Valued at A$25 per index point', 'ASX 24 规格 SPI 200 (AP)')
    out['AP_PER_POINT'] = 25.0
    _need(specs, 'Valued at A$5 per index point', 'ASX 24 规格 Mini SPI 200 (AM)')
    out['AM_PER_POINT'] = 5.0
    bond = _text(_fetch('asx_bond', refresh))
    for code, term in (('YT', '3 years'), ('XT', '10 years')):
        _need(bond, 'Face value: A$100,000', 'ASX 债券衍生品页 %s' % code)
        out[code] = 100_000.0
    st = _text(_fetch('asx_st', refresh))
    _need(st, 'A$1,000,000 face value 90 day bank accepted bills', 'ASX 短端页 IR')
    out['IR'] = 1_000_000.0
    _need(st, 'payable on a notional sum of A$3,000,000', 'ASX 短端页 IB')
    out['IB'] = 3_000_000.0
    eto = _text(_fetch('asx_eto_spec', refresh))
    _need(eto, 'Usually 100 shares per contract', 'ASX 期权规格页')
    out['ETO_SHARES'] = 100.0
    return out


# ══════════════════════ 2. 2019-01 成员成交量 ══════════════════════
def mx_volumes_1901(refresh=False):
    """从 MX 官方 1901_stats_en.xlsx 读 2019-01 逐合约月度成交量。"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(_fetch('mx_stats_1901', refresh)),
                                data_only=True)
    if 'Jan 2019 EN' not in wb.sheetnames:
        raise AssertionError('MX 月报换了 sheet 名：%s' % wb.sheetnames)
    ws = wb['Jan 2019 EN']
    if str(ws.cell(2, 2).value).strip() != 'Jan 2019':
        raise AssertionError('MX 月报 B2 不是 "Jan 2019"，表结构变了：%r'
                             % ws.cell(2, 2).value)

    def num(v):
        if v in (None, ''):
            return 0.0
        return float(str(v).replace(',', '').replace('\xa0', ''))

    vols, totals = {}, {}
    section = None
    for r in range(3, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or '').strip()
        if not label:
            continue
        val = num(ws.cell(r, 2).value)
        if label in ('Short-Term Interest Rate Futures', 'Bond Futures',
                     'Index Futures', 'Index Options'):
            section = label
            continue
        if label == 'Total':
            if section:
                totals[section] = val
            section = None
            continue
        vols[label if not label.startswith('Others') else 'Others_' + str(section)] = val
    # 官方自己的分节合计当校验：分项之和必须等于官方 Total
    stir = vols['BAX'] + vols['CRA'] + vols['Others_Short-Term Interest Rate Futures']
    if abs(stir - totals['Short-Term Interest Rate Futures']) > 0.5:
        raise AssertionError('STIR 分项 %.0f ≠ 官方 Total %.0f'
                             % (stir, totals['Short-Term Interest Rate Futures']))
    bond = vols['CGZ'] + vols['CGF'] + vols['CGB'] + vols['LGB']
    if abs(bond - totals['Bond Futures']) > 0.5:
        raise AssertionError('Bond 分项 %.0f ≠ 官方 Total %.0f'
                             % (bond, totals['Bond Futures']))
    return vols, totals


def mx_option_class_month(symbol, first_day, last_day):
    """【留给一次性回补用，本脚本默认不调用】拉某个期权类某段时间的逐日明细 CSV。

    m-x.ca 历史数据接口，纯 GET，可回溯到 2009-01-02：
        /en/trading/data/historical?symbol=SYM&from=YYYY-MM-DD&to=YYYY-MM-DD&dnld=1
    返回列里 `Class Symbol` 为空、`Ins. Type` = 3 的那一行是**标的自己**
    （Last Price = 股票/ETF 当日收盘价），其余行是各期权系列（Volume 为当日成交张数）。
    ⇒ 逐类月度成交量与标的 2019-01 月均收盘两样都能从同一个响应里取出来。

    ⚠ 期权类共 940 个，m-x.ca 的 robots.txt 是 `Crawl-delay: 15`，
    整轮 ≈4 小时 / ≈1.2 GB。别放进 cron，只当一次性回补跑。
    """
    import urllib.parse
    q = urllib.parse.urlencode({'symbol': symbol, 'from': first_day,
                                'to': last_day, 'dnld': '1'})
    url = 'https://www.m-x.ca/en/trading/data/historical?' + q
    with urllib.request.urlopen(url, timeout=120) as resp:
        blob = resp.read()
        if 'csv' not in resp.headers.get('Content-Type', ''):
            raise RuntimeError('%s 没返回 CSV：%s' % (symbol, resp.headers.get('Content-Type')))
    rows = list(csv.DictReader(io.StringIO(blob.decode('utf-8', 'replace'))))
    underlying = [r for r in rows if not (r.get('Class Symbol') or '').strip()]
    opts = [r for r in rows if (r.get('Class Symbol') or '').strip()]
    closes = [float(r['Last Price']) for r in underlying if r.get('Last Price')]
    volume = sum(float(r['Volume'] or 0) for r in opts)
    return {'symbol': symbol, 'option_volume': volume,
            'avg_close': (sum(closes) / len(closes)) if closes else None,
            'obs_days': len(closes), 'url': url}


def asx_eto_1901(refresh=False):
    """ASX 单股期权 2019-01 逐类 Lots Traded + 月初/月末收盘价。"""
    import xlrd
    wb = xlrd.open_workbook(file_contents=_fetch('asx_eto_2019', refresh))
    sh = wb.sheet_by_name('January')
    if 'January 2019' not in str(sh.cell_value(0, 0)):
        raise AssertionError('ASX 年报 January 表抬头变了：%r' % sh.cell_value(0, 0))
    rows, all_rows = [], []
    for r in range(5, sh.nrows):
        code = str(sh.cell_value(r, 0)).strip()
        if not code or code == 'Total':
            continue
        try:
            p_first = float(sh.cell_value(r, 1))
            p_last = float(sh.cell_value(r, 2))
        except (TypeError, ValueError):
            continue
        call = sh.cell_value(r, 4)
        put = sh.cell_value(r, 5)
        lots = (float(call) if call != '' else 0.0) + (float(put) if put != '' else 0.0)
        all_rows.append((code, p_first, p_last, lots))
        # XJO / XJO* / XJOL 是**指数**期权与指数 LEPO，乘数是每点澳元数，
        # 不是 100 股/张 —— 它们在 pools.py 里挂 ASX_INDEX_OPT，不属于 ASX_ETO。
        if code.rstrip('*L') != 'XJO':
            rows.append((code, p_first, p_last, lots))
    return rows, all_rows


def asx_option_totals_1901(refresh=False):
    """ASX 官方自己给的 2019-01 期权分项成交量，用来外部校验上面那张 xls 的加总。

    来源：derivatives-equity-and-index-options-jan2019.pdf 第 7 页
    «Options - Volume, Value and Open Interest» 的 Volume 表 Jan-19 行，
    7 个数依次是 CALL / PUT / TOTAL OPTIONS / EQUITY OPTIONS / EQUITY LEPO /
    ASX INDEX OPTION / INDEX LEPO。
    """
    txt = _pdf_text(_fetch('asx_eqderiv_1901', refresh))
    head = txt.find('EQUITY LEPO')
    if head < 0:
        raise AssertionError('ASX 期权分项 PDF 里找不到 Volume 表表头')
    seg = txt[head:head + 400]
    m = re.search(r'Jan-19\s*\n((?:\s*[\d,]+\s*\n){7})', seg)
    if not m:
        raise AssertionError('ASX 期权分项 PDF 的 Jan-19 行取不到 7 个数：%r' % seg[:200])
    nums = [int(x.replace(',', '')) for x in m.group(1).split()]
    keys = ('call', 'put', 'total', 'equity', 'equity_lepo', 'index', 'index_lepo')
    out = dict(zip(keys, nums))
    if out['equity'] + out['equity_lepo'] + out['index'] + out['index_lepo'] != out['total']:
        raise AssertionError('ASX 官方分项自己就不自洽：%r' % out)
    return out


def _asx_eto_candidates(rows):
    """三个候选常数：用月初价 / 月末价 / 两者中点。**都不是 avg_close**。"""
    v = sum(x[3] for x in rows)
    out = {}
    for name, pick in (('first', lambda a, b: a), ('last', lambda a, b: b),
                       ('mid', lambda a, b: (a + b) / 2.0)):
        out[name] = 100.0 * sum(lots * pick(a, b) for _, a, b, lots in rows) / v
    out['lots'] = v
    return out


# ══════════════════════ 3. 篮子常数 ══════════════════════
def weighted_constant(members):
    """members = [(名字, 2019-01 张数, 单张名义额)]；返回按张数加权的等效单张名义额。"""
    tot = sum(m[1] for m in members)
    if tot <= 0:
        raise ZeroDivisionError('成员 2019-01 张数合计为 0，篮子常数无定义')
    return sum(m[1] * m[2] for m in members) / tot


_RATE_CAVEAT = ('⚠ 利率合约按面值计名义额、不乘结算价（报价是 100−收益率，乘上去无经济含义）。'
                '名义额不反映利率风险敞口：同样名义额下 2 年期与 10 年期的 DV01 差 5 倍以上，'
                '本池占比只能读作「名义额构成」，不可读作「谁承担了更多利率风险」。')

_BLOCKERS = {
    'ASX_DERIV':
        '📌 未填：缺 2019-01 分品种成交量（权重）。检索路径与实测结果：'
        '(1) 2019-01 MAR 正文印的分品种链接 '
        'http://www.asx.com.au/data/market-reports/MonthlyFuturesMarketsReport190131.pdf '
        '→ HTTP 200 + text/html + 136,750 字节通用外壳页（soft-404，不是 404）；'
        'www2.asx.com.au 与 /data/futures/reports/ 下同名文件返回同一张外壳页。'
        '(2) 现行分品种月报 '
        'https://www.asx.com.au/data/futures/reports/MonthlyWebTotalVolume{YYMM}S.htm '
        '→ 逐月实测 1801–2012 共 36 个月，最早的真文件是 2004（2020-04），'
        '1801–2003 全是外壳页。(3) https://www.asx.com.au/about/market-statistics/trading-volumes '
        '→ 官方原文只提供 rolling 12 个月。'
        '另注：本行在 build/pools.py 挂 adv_futures_and_options_contracts，'
        '成员是整个 ASX 24（利率+股指+商品+电力+NZ+期货期权），'
        '即便拿到 SPI200/YT/XT/IR 四条也凑不出完整权重。'
        '已核实的成员乘数（asx24-contract-specifications.pdf v11，非 2019 版）：'
        'AP A$25/指数点、YT/XT 面值 A$100,000、IR 面值 A$1,000,000、IB 名义 A$3,000,000。'
        + _RATE_CAVEAT,
    'ASX_ETO':
        '📌 未填：量齐了、价的口径不齐。权重已核对无误 —— '
        'annual-market-summary-2019.xls（January 表）161 个类 Call+Put 合计 5,953,660，'
        '与官方 derivatives-equity-and-index-options-jan2019.pdf 第 7 页的 '
        'Equity 5,156,054 + Equity LEPO 51,180 + Index 746,164 + Index LEPO 262 逐位相符；'
        '剔掉 XJO/XJO*/XJOL 后 5,207,234，与 MAR 的 Single stock equity options 逐位相符。'
        '卡点：ASX 对个股只发**月初、月末两个收盘价**，没有日频序列（历史行情是收费产品），'
        '凑不出表内统一的 base_price_basis=avg_close。'
        '候选值（100 股/张，AUD）：月初价 2,465.9435 / 月末价 2,618.9303 / 中点 2,542.4369，'
        '区间宽度约 ±3%。要采用中点必须主线程明确同意换口径并同步改 base_price_basis。'
        '⚠ 另注：ASX 单股期权乘数会因除权除息调整，同一标的不同月份可以不同；'
        '篮子常数只代表基期那个月的结构。',
    'MX_EQUITY_OPT':
        '📌 未填：需要逐个期权类的 2019-01 成交量 + 标的月均收盘。'
        'MX 官方月报（f_stat_en/1901_stats_en.xlsx）只给 Equity Options 合计 2,757,998，不拆到类。'
        '可行的官方通道已实测：'
        'https://www.m-x.ca/en/trading/data/historical?symbol=BNS&from=2019-01-02&to=2019-01-31&dnld=1 '
        '→ text/csv 1,326,135 字节，含逐日各系列 Volume 与标的自身收盘价行'
        '（Class Symbol 为空 / Ins. Type=3），日期下限 2009-01-02'
        '（写死在 /assets/application/js/historical-data.*.min.js 的 new Date(2009,0,2)）。'
        '代价：期权类下拉框 940 个 symbol × robots.txt Crawl-delay: 15 ≈ 4 小时 / 1.2 GB，'
        '超出无人值守边界。取数函数已留在 build/basefill/asx_mx.py 的 mx_option_class_month()，'
        '主线程当一次性回补跑一遍即可填上。乘数已核实：100 股/张（m-x.ca 单股期权规格页）。',
    'MX_ETF_OPT':
        '📌 未填：同 MX_EQUITY_OPT 的卡点（官方月报只给 ETF Options 合计 884,892，不拆到类）。'
        '乘数已核实：100 份/张（m-x.ca Options on ETF 规格页）。'
        '走 mx_option_class_month() 一次性回补即可；'
        'ETF 类比单股类少得多，但 940 个 symbol 的下拉框里没有官方的「哪些是 ETF」标记，'
        '要先从 m-x.ca 的 options-list 页把 ETF 类拆出来 —— 而那张表是 2026 年的现况，'
        '2019 年在册、之后退市的类不会出现在里面，直接用会漏成员。',
}

COLUMNS = ['product_id', 'base_price_local', 'base_ccy', 'basket_constant',
           'source_url', 'computed_note']


def build(refresh=False):
    mx = mx_specs(refresh)
    asx = asx_specs(refresh)
    vols, totals = mx_volumes_1901(refresh)
    eto_rows, eto_all = asx_eto_1901(refresh)
    eto = _asx_eto_candidates(eto_rows)

    # ASX_ETO 的成员集合与权重必须先过官方外部校验，再谈常数怎么算。
    off = asx_option_totals_1901(refresh)
    if int(round(sum(x[3] for x in eto_all))) != off['total']:
        raise AssertionError('ASX 年报 xls 全表合计 %.0f ≠ 官方分项 TOTAL OPTIONS %d'
                             % (sum(x[3] for x in eto_all), off['total']))
    single_stock = off['equity'] + off['equity_lepo']
    if int(round(eto['lots'])) != single_stock:
        raise AssertionError('剔除 XJO 后 %.0f ≠ 官方 Equity+Equity LEPO %d'
                             % (eto['lots'], single_stock))

    report = []
    rows = []

    # ── MX_STIR：BAX / CRA / Others ───────────────────────────────────
    stir_members = [('BAX', vols['BAX'], mx['BAX']),
                    ('CRA', vols['CRA'], mx['CRA'])]
    stir = weighted_constant(stir_members)
    report.append(('MX_STIR', stir_members, stir, totals['Short-Term Interest Rate Futures']))
    rows.append({
        'product_id': 'MX_STIR', 'base_price_local': repr(stir), 'base_ccy': 'CAD',
        'basket_constant': repr(stir),
        'source_url': SRC['mx_rule15'] + ' | ' + SRC['mx_stats_1901'],
        'computed_note':
            '按 2019-01 成交量加权：BAX %s 张 × 面值 C$1,000,000；CRA %s 张（当月为 0，权重 0）；'
            'Others 0 张。官方分节 Total %s 与分项之和逐位相符。'
            'BAX 面值取 MX 自家存档规则书 Rule Fifteen «Futures Contracts Specifications» '
            's.15500 与 s.15503(b)（末次修订 14.09.18，即 2019-01 在效版本）'
            '—— BAX 2024 年退市、现行产品页 404，只有存档规则书能给基期在效的规格。'
            'CRA 面值 C$1,000,000 由 m-x.ca CRA 页「each basis point per annum of interest '
            '= $25 per contract」反推（1e6 × 0.0001 × 0.25 = 25）。%s'
            % (f'{vols["BAX"]:,.0f}', f'{vols["CRA"]:,.0f}',
               f'{totals["Short-Term Interest Rate Futures"]:,.0f}', _RATE_CAVEAT),
    })

    # ── MX_BOND：CGZ / CGF / CGB / LGB ────────────────────────────────
    bond_members = [(c, vols[c], mx[c]) for c in ('CGZ', 'CGF', 'CGB', 'LGB')]
    bond = weighted_constant(bond_members)
    report.append(('MX_BOND', bond_members, bond, totals['Bond Futures']))
    rows.append({
        'product_id': 'MX_BOND', 'base_price_local': repr(bond), 'base_ccy': 'CAD',
        'basket_constant': repr(bond),
        'source_url': SRC['mx_rule15'] + ' | ' + SRC['mx_cgb'] + ' | ' + SRC['mx_stats_1901'],
        'computed_note':
            '按 2019-01 成交量加权：CGZ %s 张、CGF %s 张、CGB %s 张、LGB %s 张，'
            '官方分节 Total %s 与分项之和逐位相符。有量的 CGF 与 CGB 面值同为 C$100,000，'
            '故常数恰为 C$100,000，与权重无关（这也让它对权重误差免疫）。'
            '面值取 MX 存档规则书 Rule 15 s.15622/15642/15662；'
            '现行 m-x.ca CGB/CGF 页同为 C$100,000，两处互证。'
            '⚠ CGZ 唯一不一致：存档 Rule 15 s.15603（末修 18.01.16）写 $200,000，'
            '现行 /cgz 页写 C$100,000；本行按基期在效版取 $200,000，'
            '因当月 CGZ 成交量为 0，两者都不改变常数。%s'
            % (f'{vols["CGZ"]:,.0f}', f'{vols["CGF"]:,.0f}', f'{vols["CGB"]:,.0f}',
               f'{vols["LGB"]:,.0f}', f'{totals["Bond Futures"]:,.0f}', _RATE_CAVEAT),
    })

    # ── 四个填不出来的：留空 + 写清卡在哪一步 ─────────────────────────
    for pid in ('ASX_DERIV', 'ASX_ETO', 'MX_EQUITY_OPT', 'MX_ETF_OPT'):
        rows.append({
            'product_id': pid, 'base_price_local': '', 'base_ccy':
                'AUD' if pid.startswith('ASX') else 'CAD',
            'basket_constant': '',
            'source_url': (SRC['asx_eto_2019'] if pid == 'ASX_ETO' else
                           SRC['asx24_specs'] if pid == 'ASX_DERIV' else
                           SRC['mx_eqopt'] if pid == 'MX_EQUITY_OPT' else SRC['mx_etfopt']),
            'computed_note': _BLOCKERS[pid],
        })

    with open(_OUT, 'w', newline='', encoding='utf-8') as fh:
        w = csv.DictWriter(fh, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)
    return rows, report, eto, eto_all, mx, asx, vols, totals, off


def main(argv=None):
    ap = argparse.ArgumentParser(description='ASX / MX 篮子常数取数（可复算）')
    ap.add_argument('--refresh', action='store_true', help='强制重下所有官方文件')
    args = ap.parse_args(argv)

    rows, report, eto, eto_all, mx, asx, vols, totals, off = build(args.refresh)

    print('基期 base_month = %s（build/notional.py 的 BASE_MONTH）' % BASE_MONTH)
    print()
    print('── 合约规格（每一条都在官方页面上逐字断言过）──')
    print('  MX：BAX 面值 C$%s | CGB/CGF/LGB 面值 C$%s | CGZ 面值 C$%s（2019 在效版）'
          % (f'{mx["BAX"]:,.0f}', f'{mx["CGB"]:,.0f}', f'{mx["CGZ"]:,.0f}'))
    print('      CRA 面值 C$%s | 单股期权 %.0f 股/张 | ETF 期权 %.0f 份/张 | SXF C$%.0f/点'
          % (f'{mx["CRA"]:,.0f}', mx['EQ_OPT_SHARES'], mx['ETF_OPT_UNITS'],
             mx['SXF_PER_POINT']))
    print('  ASX：AP A$%.0f/指数点 | AM A$%.0f/指数点 | YT/XT 面值 A$%s'
          % (asx['AP_PER_POINT'], asx['AM_PER_POINT'], f'{asx["YT"]:,.0f}'))
    print('      IR 面值 A$%s | IB 名义 A$%s | ETO %.0f 股/张'
          % (f'{asx["IR"]:,.0f}', f'{asx["IB"]:,.0f}', asx['ETO_SHARES']))
    print()
    print('── 2019-01 成员成交量与篮子常数 ──')
    for pid, members, const, official_total in report:
        tot = sum(m[1] for m in members)
        print('  %s' % pid)
        for name, v, notional in members:
            share = (v / tot * 100.0) if tot else 0.0
            print('      %-4s %14s 张  权重 %6.2f%%  单张名义额 %s'
                  % (name, f'{v:,.0f}', share, f'{notional:,.0f}'))
        print('      Σ张数 %s（官方分节 Total %s）→ 篮子常数 = %s'
              % (f'{tot:,.0f}', f'{official_total:,.0f}', f'{const:,.4f}'))
    print()
    print('── ASX_ETO：算得出来但口径不合，不写进 CSV ──')
    print('  成员 %d 个类（已剔 XJO/XJO*/XJOL 指数期权），Σ张数 %s'
          % (len(eto_all) - 3, f'{eto["lots"]:,.0f}'))
    print('  对账（已在 build() 里断言，不符即抛）：全表 %d 类合计 %s '
          '= 官方 Equity %s + Equity LEPO %s + Index %s + Index LEPO %s'
          % (len(eto_all), f'{sum(x[3] for x in eto_all):,.0f}',
             f'{off["equity"]:,}', f'{off["equity_lepo"]:,}',
             f'{off["index"]:,}', f'{off["index_lepo"]:,}'))
    print('          剔 XJO 后 %s = 官方 Equity + Equity LEPO %s'
          % (f'{eto["lots"]:,.0f}', f'{off["equity"] + off["equity_lepo"]:,}'))
    print('  候选常数（100 股/张，AUD）：月初 %.4f | 月末 %.4f | 中点 %.4f'
          % (eto['first'], eto['last'], eto['mid']))
    print('  ⇒ 三者都不是 avg_close；表内既有 23 行统一 avg_close，故留空待主线程定夺。')
    print()
    print('已写出 %s（%d 行：%d 填 / %d 留空）'
          % (_OUT, len(rows), sum(1 for r in rows if r['basket_constant']),
             sum(1 for r in rows if not r['basket_constant'])))
    return 0


if __name__ == '__main__':
    sys.exit(main())
