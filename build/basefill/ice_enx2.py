# -*- coding: utf-8 -*-
"""ICE 三个 + Euronext 三个 pool_product 篮子常数的**可复算**取数脚本（第二轮）。

    python3 build/basefill/ice_enx2.py              # 有 cache/basefill/ 副本就用，没有才下载
    python3 build/basefill/ice_enx2.py --refresh    # 强制重下所有官方文件
    python3 build/basefill/ice_enx2.py --ice-probe  # 额外现场跑 ICE 侧「为什么还是算不出」的实测
                                                    # （要打 ice.com，约 1 分钟，默认不跑）

负责 series/contract_specs.csv 这六行：
    ENX_INDEX_DERIV / ENX_SINGLESTOCK_LEGACY / ENX_MATIF   ← 本轮**全部填出**
    ICE_ENERGY / ICE_STIR / ICE_MLTIR                      ← 本轮**仍然填不出**，卡点见第三节
跑完写 series/_specs_part_ice_enx2.csv（不直接改 contract_specs.csv：并发写会互相覆盖）。

━━ 篮子常数的定义（照 contract_specs.csv 的 notes 列）━━
    篮子常数 = Σ(成员 2019-01 张数 × 乘数 × 基期价格) ÷ Σ(成员 2019-01 张数)
pool_product 行的 multiplier 恒为 1，所以这个常数直接就是 base_price_local
（build/notional.py: base_notional_per_unit_local = multiplier × base_price_local）。

════════════════════════════════════════════════════════════════════════════
一、Euronext：官方直接发名义额，不需要乘数表、不需要价格序列
════════════════════════════════════════════════════════════════════════════
第一轮的死因是「Euronext 官方不按合约拆张数」。**这个前提是错的** —— 错在只看了
投资者关系口径的那两个 xlsx（fetch/enx.py 用的 live.euronext.com/sites/default/
files/statistics/ir/…），那两份确实只给池级合计、且衍生品那半边只有 lots。

真正该找的是市场统计口径的另一个目录：

    https://live.euronext.com/sites/default/files/statistics/derivatives/monthly/
        Derivatives YYYYMM.xlsx        ← 每月一份，**当年年初至今**逐月列
    https://live.euronext.com/sites/default/files/statistics/factbook/yearly/
        Euronext Fact Book YYYY.xlsx   ← 年鉴，内含同样的衍生品分表

入口是 https://live.euronext.com/en/resources/statistics（"Statistics & reports"），
页面上就挂着这些直链；纯 urllib 可读，不需要 curl_cffi、不需要浏览器。

这份工作簿有 9 张表，其中三张是本脚本的全部依据：
    Volume            每个合约的当月成交张数
    Value of Volume   每个合约的当月**名义额**，表头原文 "Monthly Value of Volume (€ '000)"
    Premium Turnover  期权权利金（**另一张独立表** ⇒ Value of Volume 不是权利金）
分类列给到 Underlying / Market Place / Cash-Physical / American-European /
Option-Future / Standard-Flexible / Code / Country，粒度是**单个合约**
（2019-12 那份 1,102 行，单股期权 800 多个标的一个不落）。

于是基期恰好是 2019-01 这件事把整条路缩成一步：

    Value of Volume(2019-01) × 1000 ÷ Volume(2019-01)
      = Σ(张数 × 乘数 × 成交价) ÷ Σ(张数)
      = 篮子常数

加权已经内含在官方数字里，**不需要任何乘数表、不需要任何价格序列、不需要自己加权**。
CAC 40（€10/点）与 AEX（€200/点）差 20 倍这个坑因此自动消失 —— 不是绕过去了，
是根本不需要按合约拆开再加权。

━━ 为什么敢断定 "Value of Volume" 就是名义额 ━━
官方工作簿没有给衍生品分表写方法论（Fact Book 的 Methodology 页只讲现货口径），
所以不靠「我认为」，靠**五个互相独立的现场反算**（_verify_value_is_notional()
每次运行都跑，对不上就 raise）：

    合约            官方名义额/张      ÷ 官方乘数        = 反算出的价格
    FCE  CAC 40 期货    48,005.23 €      €10 /点          4,800.5 点
    MFC  迷你 CAC 期货   4,795.57 €      €1  /点          4,795.6 点   ← 与 FCE 差 0.10%
    FTI  AEX 期货       99,976.15 €      €200/点            499.9 点
    MFA  迷你 AEX 期货  10,028.80 €      €20 /点            501.4 点   ← 与 FTI 差 0.31%
    EBM  小麦期货       10,141.67 €      50 公吨          202.83 €/公吨
    OBM  小麦期权       10,010.48 €      50 公吨          200.21 €/公吨 ← 与 EBM 差 1.31%
    ECO  菜籽期货       18,483.69 €      50 公吨          369.67 €/公吨
    OCO  菜籽期权       18,554.52 €      50 公吨          371.09 €/公吨 ← 与 ECO 差 0.38%

同一标的的标准合约与迷你合约、期货与期权，反算出来的价格全部收敛到同一个数
（相对误差 0.1%–1.3%），且落在 2019-01 该市场的真实价位上。
如果这一列是权利金、是双边计数、或者是别的什么，这八个数不可能同时对上。
乘数取自 Euronext 官方合约规格页原文（见 _OFFICIAL_MULT，--verify-specs 会现场核）。

━━ 分节与本仓池列的对账（_reconcile()，对不上就 raise）━━
工作簿的分节与 series/enx.csv 的六条列是**逐位相等**的（enx.csv 来自 IR 口径的
另一份 xlsx，两份文件、两条链路、同一个数），这一步同时证明「分子分母同源」：

    enx.csv 2019-01（ADV 千张 × 22 个交易日）        工作簿分节 2019-01
    adv_index_futures        3,261,839  =  INDEX Futures 3,223,589 + DIVIDEND INDEX Futures 38,250
    adv_index_options        1,450,605  =  INDEX Options 1,450,605
    adv_singlestock_futures     13,435  =  STOCK Futures 11,910 + STOCK DIVIDEND Futures 1,525
    adv_singlestock_options  5,889,261  =  STOCK Options 5,886,253 + ETF Options 3,008
    adv_commodity_futures      848,801  =  GRAINS AND OILSEEDS Futures
    adv_commodity_options       53,985  =  GRAINS AND OILSEEDS Options

⇒ 三个 pool_product 的成员分节就此确定，不是猜的：
    ENX_INDEX_DERIV        = INDEX PRODUCTS + DIVIDEND INDEX PRODUCTS
    ENX_SINGLESTOCK_LEGACY = STOCK PRODUCTS + STOCK DIVIDEND PRODUCTS + ETF PRODUCTS
    ENX_MATIF              = GRAINS AND OILSEEDS PRODUCTS
      （脚本另断言「第 1 层 COMMODITY PRODUCTS 合计 == GRAINS 合计」，即当月房产/化肥/
        能源/乳制品四个分节全 0；将来它们真有量了这条断言会当场炸，不会静默混进来）
注意 **ETF 期权在 Euronext 的口径里算在单股那一边**，不在股指那一边 —— 这不是本脚本
的选择，是上面那条 5,889,261 的等式定的。

━━ 三份官方文件互校（_crosscheck_sources()）━━
本脚本对 2019-01 同时读三份官方文件并要求逐位相等：
    Derivatives 201912.xlsx   （2019 全年那份，主用）
    Derivatives 201905.xlsx   （2019 年**现存最早**的一份；201901–201904 官网已 403）
    Euronext Fact Book 2019.xlsx（年鉴里的同名分表）
实测三份在本脚本用到的六个分节上完全一致（Total Euronext 一行有 1e-9 量级的浮点
表示差异，不影响任何用到的数）。这一步是防「后发文件回溯修订、常数悄悄变了」。

━━ ⚠ 必须留痕的口径差 ━━
contract_specs.csv 表头把实测基准定成 base_price_basis=avg_close ＝「月内全部交易日
**收盘价**的算术平均」。Euronext 的 Value of Volume 隐含的价格项是**该月成交量加权的
成交价**，不是等权收盘均值。两者不是同一个东西。本脚本能测的那一腿的差距：
    CAC 40 期货隐含 4,800.5 点（FCE ÷ €10）  vs CAC 40 期权隐含 4,713.6 点（PXA ÷ €10）  → −1.81%
    AEX   期货隐含   499.9 点（FTI ÷ €200） vs AEX   期权隐含   494.6 点（AEX ÷ €100） → −1.05%
（期权那腿系统性偏低，与第一轮 Eurex 侧观察到的方向一致：ESTX50 期权隐含比期货
隐含低 4.39%、DAX 期权低 1.87%。Euronext 同样没有在工作簿里写明期权的 Value of
Volume 用标的价还是行权价，本脚本不猜。）
主线程若坚持全表单一 avg_close 基准，这三行要么标一个新的 basis 值，要么接受这个
量级的偏差；**不要默默当成 avg_close**。

━━ 与 Eurex 三行的可比性 ━━
第一轮 EUREX_INDEX / EUREX_EQUITY 走的是同构的一条路（Capital Volume ÷ Traded
Contracts），所以本轮三行与那两行**口径一致、可以并排比**：
    EUREX_EQUITY  3,468.13 €/张   vs  ENX_SINGLESTOCK_LEGACY 3,423.98 €/张（差 1.3%）
两家的单股衍生品单张名义额落在同一量级同一水平，是个额外的旁证。
而 EUREX_RATES 那一行是**面值口径**（不乘结算价），与本文件三行不同口径 —— 但本文件
三行里没有利率池，不冲突。

════════════════════════════════════════════════════════════════════════════
二、ICE：三个仍然填不出，卡点已收敛到一个
════════════════════════════════════════════════════════════════════════════
第一轮的三条卡点（坑 A 权重、坑 B 面值原文、坑 C 的 TTF 无固定合约量）本轮逐条查过，
结论是**方法① 在 ICE 这边根本不存在，权重那条路则被 reCAPTCHA 挡死**。

【① ICE 全公司没有任何「名义额 / capital volume / turnover value」列】现场核过：
  · fetch/ice.py 在用的 Monthly Statistics Tracking xlsx（cache/ice_monthly_stats.xlsx）
    四张表全部翻过：'Derivs (ADV, RPC, OI)' 只有 ADV（千张）/ RPC（美元每张）/ OI，
    'US Equity Options'、'Cash Products' 同理；唯一带 notional 的是 'CDS Clearing'
    表的 "TOTAL CDS Gross Notional Cleared"，那是 CDS 清算，与本仓三个池无关。
    ⇒ 「ADV in notional」在 ICE 的 IR 材料里不存在，10-K 的 selected operating data
    与这份 xlsx 同源同粒度，不必再试。
  · https://www.ice.com/report/7（ICE Futures Europe 历史月度成交，服务端渲染、
    纯 urllib 可读、1995 年至今 32 张年度表）与 /report/147：**只有张数**，
    32 张表的表头逐张核过，没有任何金额列。

【② 分合约张数被 reCAPTCHA 挡住 —— 这是本轮唯一的、也是真正的卡点】
  ICE 报表中心的后端接口本轮找到了（在 static.ice.com 的前端 chunk 里）：
      GET  /marketdata/api/reports/metadata/{id}   报表元数据（公开，200）
      GET  /marketdata/api/reports/{id}/criteria   查询条件（本轮全部实测）
      POST /marketdata/api/reports/{id}/results    取数
  把报表中心页面上出现的 132 个 id 逐个请求了 metadata，**46 个返回 200**
  （其余非 200 —— 报表中心页上不少 id 是分类锚点/产品页，不是报表；快照存
  cache/basefill/ice_report_meta.json）。这 46 个里带「历史成交量」字样的是：
      id   exchange              name                                recaptchaRequired
       7   ICE Futures Europe    Historical Monthly Volumes           False   ← 只到品类，见①
       8   ICE Futures U.S.      Historical Monthly Volumes-Futures   False   ← 美国所，与本仓三池无关
       9   ICE Futures U.S.      Historical Monthly Volumes-Options   False   ← 同上
     147   North America         Historical Monthly Volumes-New Energy False  ← 只到品类
      26   ICE Futures Europe    Historical Daily Volume - Futures    **True**  ← 分合约，正是要的
      27   ICE Futures Europe    Historical Daily Volume - Options    **True**  ← 同上
      96/97   ICE Futures U.S.   同名两张                              **True**
      98–101  ICE Clear Credit / ICE Clear Europe-CDS                  **True**
  ICE Endex 的 238/239（TTF 在这里）单独查过：recaptchaRequired 同样是 true，
  reportTemplateURL = /marketdata/public-web/endex/historical-daily-volume-oi-report。
  ⇒ **凡是到合约层的历史成交量报表，无一例外要 reCAPTCHA**；不要 reCAPTCHA 的那几张
  （7 / 8 / 9 / 147）粒度只到品类，或者是与本仓三池无关的美国所。
  26 / 27 / 238 / 239 / 96 / 97 / 176 的 criteria 接口在不带 reCAPTCHA token 时
  一律返回 **HTTP 409 Conflict**（本轮实测，见 --ice-probe；HTTP 429 是 Cloudflare
  按 IP 限流，连打就会触发，与 409 不是一回事）。
  ⇒ 要拿分合约张数就得过 reCAPTCHA。**本脚本不做这件事**，两个理由：
     (a) 绕过/代答验证码是明令禁止的动作；
     (b) 就算人工在浏览器里点一次，cron 无人值守也复现不了，仓库仍然没法复算。
  可行的合规出路只有一条：**人在浏览器里手动跑一次 report 26 / 27 / 238 / 239，
  把 2019-01 的逐合约日成交导出成 CSV 放进 cache/basefill/，脚本再从本地文件算。**
  这条留给下一个人决定要不要做（基期常数是一次性写死的，人工取一次是可接受的），
  但**必须把导出的原始 CSV 一起提交**，否则仓库还是复算不了。

  ⚠ 顺带证实了「STIR + MLTIR 几乎全部在 ICE Futures Europe」：
     ice.csv 2019-01 的 adv_stir 1,854 千张 + adv_mltir 197 千张 = 45,122,000 张/月
     /report/7 的 2019-01 IFEU Interest Rates 期货 37,491,346 + 期权 7,632,263
                                                     = 45,123,609 张
     两边差 0.004%。⇒ report 26/27 这两张表一旦拿到，权重就是完整的，不缺别的所。

【③ ICE_STIR 还多一层：官方规则手册里**根本没有**短端合约的面值】
  本轮把 ICE Futures Europe 的合约规则原文（Contract Regulations，按 SECTION 分册的
  官方 PDF）抠出来了：
   · SECTION NNNN《CONTRACT RULES: ICE FUTURES SHORT TERM INTEREST RATE INDEX
     FUTURES CONTRACTS》 https://www.ice.com/publicdocs/contractregs/112_SECTION_NNNN.pdf
     三个月 EURIBOR 那张表的原文逐字是：
         "Unit of trading      €2,500 * Rate Index"
         "Contract Multiplier  €2,500"
     全篇只定义 "Contract Multiplier means the factor which, when multiplied by the
     Rate Index determines the Unit of Trading for a Contract"，
     **通篇没有 EUR 1,000,000 这个数，也没有 nominal / notional / face value 字样**。
     （同表里 NIBOR = NOK 25,000×、STIBOR = SEK 25,000×、BBSW = AUD 2,500× 也是同一写法。）
     ⇒ EUR 1,000,000 是从「每指数点 2,500 欧元 ÷ 3 个月折年 ÷ 万分之一」反推的，
       **反推值不许写进表**，这条第一轮的判断本轮在规则手册层面得到确认，不是页面没抓全。
   · 对照组：中长端是有官方面值的。SECTION RRRR《ICE FUTURES GILT FUTURES CONTRACTS》
     https://www.ice.com/publicdocs/contractregs/116_SECTION_RRRR.pdf 原文：
         "Unit of Trading  £100,000 nominal value notional Gilt"（四个 Gilt 合约同值）
         "A 'lot' shall be an amount having a nominal value of £100,000 (the
          'delivery amount') of a Deliverable Gilt..."
     ⇒ ICE_MLTIR 的主力成员面值有官方原文，缺的**只有权重**。

【④ ICE_ENERGY 的三层坑，本轮无一被拆掉】
  · 分子分母不同源：pools.py 把 ICE_ENERGY 挂在 ice.csv 的 adv_energy_kcontracts
    （2019-01 = 2,718 千张/日 × 21 = 57,078,000 张，全球口径，含 ICE Endex /
    ICE Futures U.S. / ICE Futures Abu Dhabi），而分合约的 report 26/27 只有 IFEU、
    report 238/239 只有 Endex —— 就算过了 reCAPTCHA 也得拼四个所才凑得齐。
  · TTF **没有固定乘数**（官方页原文 "1 MW per day in contract period ... x 23, 24
    or 25 hours"，一张月度合约 672–744 MWh 不等），必须按合约期逐月折算。
  · 跨币种（Brent USD/桶、TTF EUR/MWh），记账币定 USD，各成员面值须先按
    series/fx.csv 的 2019-01 月均汇率折美元再加权（本行 ccy=USD、fx 那一跳是 1.0）。

════════════════════════════════════════════════════════════════════════════
三、依赖
════════════════════════════════════════════════════════════════════════════
填数那条路：标准库 + openpyxl（Euronext 的 .xlsx 是 OOXML）。**不需要** curl_cffi、
不需要浏览器、不需要 xlrd。
--ice-probe 那条路额外用 curl_cffi（ice.com 走 Akamai，按 TLS 指纹拦纯 urllib 的
某些接口）；它只打印证据、不参与任何常数计算，缺了不影响主流程。
"""

import argparse
import csv
import io
import os
import re
import sys
import urllib.parse
import urllib.request

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SERIES = os.path.join(ROOT, 'series')
CACHE = os.path.join(ROOT, 'cache', 'basefill')

BASE_MONTH = '2019-01'          # 与 build/notional.py / build/check_specs.py 同源
BASE_COL_HDR = 'Jan 19'         # Euronext 工作簿里 2019-01 那一列的表头文字
OUT = os.path.join(SERIES, '_specs_part_ice_enx2.csv')

_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

# ── Euronext 官方统计文件（入口页 https://live.euronext.com/en/resources/statistics）──
ENX_STATS_LANDING = 'https://live.euronext.com/en/resources/statistics'
ENX_DERIV_MONTHLY = ('https://live.euronext.com/sites/default/files/statistics/'
                     'derivatives/monthly/Derivatives %s.xlsx')
ENX_FACTBOOK = ('https://live.euronext.com/sites/default/files/statistics/'
                'factbook/yearly/Euronext Fact Book %d.xlsx')

# 主用文件 + 两份互校文件。
#   201912 = 2019 全年那份（YTD 逐月列，含 Jan 19）
#   201905 = 2019 年官网**现存最早**的一份（201901–201904 实测 403，不是本脚本没试）
ENX_PRIMARY_MONTH = '201912'
ENX_CROSSCHECK_MONTHS = ('201905',)
ENX_FACTBOOK_YEAR = 2019

# 工作簿里三张要用的表；Fact Book 里同样三张表换了名字，做个映射。
_SHEET_VOL = {'deriv': 'Volume', 'factbook': 'Derivatives Monthly Volume'}
_SHEET_VAL = {'deriv': 'Value of Volume', 'factbook': 'Value of Volume €000'}
# 表头原文 "Monthly Value of Volume (€ '000)" ⇒ 单位是欧元千元，还原成欧元要 ×1000。
VALUE_UNIT_EUR = 1000.0

# ── 工作簿的两层分节结构 ──
# 第 1 层写在 A 列（COMMODITY / EQUITY / FX PRODUCTS），第 2 层写在 B 列，
# 明细合约行 A/B 两列都空、靠 D–K 的分类列描述。取数一律指定层级，不许模糊匹配。
_LVL1, _LVL2, _LVL3 = 0, 1, 2

# ── 三个 pool_product 由哪些第 2 层分节组成 ──
# **不是拍的**：由 _reconcile() 对 series/enx.csv 的六条列逐位验出来（见 docstring 一）。
# 出现没登记过的分节就 raise —— 将来 Euronext 加一个新分节，它不会被静默算进去。
_ENX_SECTIONS = {
    'ENX_INDEX_DERIV':        ('INDEX PRODUCTS', 'DIVIDEND INDEX PRODUCTS'),
    'ENX_SINGLESTOCK_LEGACY': ('STOCK PRODUCTS', 'STOCK DIVIDEND PRODUCTS', 'ETF PRODUCTS'),
    'ENX_MATIF':              ('GRAINS AND OILSEEDS PRODUCTS',),
}
# 第 2 层里显式登记为「不属于本轮三个池」的分节。不许有第三种情况。
# ⚠ ENX_MATIF 只收 GRAINS AND OILSEEDS：pools.py 挂的两条列
# （adv_commodity_futures/options）在 2019-01 与它逐位相等，其余四个商品分节当月全 0；
# _reconcile() 会现场断言「第 1 层 COMMODITY PRODUCTS 合计 == GRAINS 合计」，
# 将来某个月化肥/乳制品真有量了，这条断言会当场炸，不会静默混进来。
_ENX_SECTIONS_IGNORED = {
    'REAL ESTATE':        '巴黎房产指数期货，2019-01 无量；pools.py 的 ags 池不含它',
    'FERTILISER PRODUCTS': '2019-01 成交 0 张',
    'ENERGY PRODUCTS':     '2019-01 成交 0 张（Euronext 侧的木屑期货，不是 ICE 能源）',
    'DAIRY PRODUCTS':      '2019-01 成交 0 张',
    'FX PRODUCTS':         '2019-01 成交 0 张，且 pools.py 没有对应的 Euronext FX 池',
}
# 第 1 层分节的登记表（用于「COMMODITY 合计 == GRAINS 合计」那条断言）
_ENX_LVL1 = {
    'COMMODITY PRODUCTS': ('GRAINS AND OILSEEDS PRODUCTS', 'REAL ESTATE',
                           'FERTILISER PRODUCTS', 'ENERGY PRODUCTS', 'DAIRY PRODUCTS'),
    'EQUITY PRODUCTS':    ('STOCK PRODUCTS', 'STOCK DIVIDEND PRODUCTS', 'INDEX PRODUCTS',
                           'ETF PRODUCTS', 'DIVIDEND INDEX PRODUCTS'),
    'FX PRODUCTS':        ('FX PRODUCTS',),
}

# ── 分节 → series/enx.csv 的列（× 当月交易日数后必须逐位相等）──
# 键是 (enx.csv 列名, 期货/期权)，值是工作簿里对应的**子分节**标题（col B）。
_RECONCILE = [
    ('adv_index_futures_kcontracts',       ['INDEX Futures', 'DIVIDEND INDEX Futures']),
    ('adv_index_options_kcontracts',       ['INDEX Options']),
    ('adv_singlestock_futures_kcontracts', ['STOCK Futures', 'STOCK DIVIDEND Futures']),
    ('adv_singlestock_options_kcontracts', ['STOCK Options', 'ETF Options']),
    ('adv_commodity_futures_kcontracts',   ['GRAINS AND OILSEEDS Futures']),
    ('adv_commodity_options_kcontracts',   ['GRAINS AND OILSEEDS Options']),
]

# ── 反算校验用的官方乘数（每一个都有官方规格页原文，--verify-specs 现场核）──
# 这些数字**不参与常数计算**，只用来证明 Value of Volume 那一列确实是名义额。
_OFFICIAL_MULT = {
    # code: (乘数, 单位, 官方原文, 官方 URL)
    'FCE': (10.0, 'EUR/指数点', 'Contract valued at € 10 per index point',
            'https://live.euronext.com/en/product/index-futures/FCE-DPAR/contract-specification'),
    'FTI': (200.0, 'EUR/指数点', 'Contract valued at € 200 per index point',
            'https://live.euronext.com/en/product/index-futures/FTI-DAMS/contract-specification'),
    'EBM': (50.0, '公吨', 'Unit of trading Fifty tonnes',
            'https://live.euronext.com/en/product/commodities-futures/EBM-DPAR/contract-specification'),
    # 期权腿的乘数（只用于下面那条「期货腿 vs 期权腿」的口径差留痕，不参与常数）
    'PXA': (10.0, 'EUR/指数点', 'Contract valued at € 10 per index point',
            'https://live.euronext.com/en/product/index-options/PXA-DPAR/contract-specification'),
    'AEX': (100.0, 'EUR/指数点', 'Contract valued at € 100 per index point',
            'https://live.euronext.com/en/product/index-options/AEX-DAMS/contract-specification'),
}
# 迷你合约与期权腿的乘数：MFC = FCE/10、MFA = FTI/10、OBM/OCO 与其标的期货同量。
# 这三条不另找规格页，做法是**只用比值断言**（迷你 vs 标准、期权 vs 期货），
# 比值检验不依赖迷你合约乘数的绝对值是否为 1 或 20，只依赖 10:1 这个官方比例。
_RATIO_CHECKS = [
    # (甲 code, 乙 code, 甲/乙 的理论倍数, 容差, 说明)
    ('FCE', 'MFC', 10.0, 0.01, 'CAC 40 标准期货 vs 迷你期货，官方乘数 €10 : €1'),
    ('FTI', 'MFA', 10.0, 0.01, 'AEX 标准期货 vs 迷你期货，官方乘数 €200 : €20'),
    ('EBM', 'OBM', 1.0, 0.02, '小麦期货 vs 小麦期权，同为 50 公吨/张'),
    ('ECO', 'OCO', 1.0, 0.02, '菜籽期货 vs 菜籽期权，同为 50 公吨/张'),
]
# 期货腿 vs 期权腿的隐含点位差（**只报不断言** —— 这正是 avg_close 口径差那一节要留痕的量）
# 两腿乘数不同（AEX 期货 €200/点、期权 €100/点），所以各自先除以自己的官方乘数再比。
_BASIS_GAP_PAIRS = [
    ('FCE', 'PXA', 'CAC 40 期权隐含点位 / 期货隐含点位'),
    ('FTI', 'AEX', 'AEX 期权隐含点位 / 期货隐含点位'),
]

# ── ICE 侧：--ice-probe 要现场打的东西（只作证据，不参与计算）──
ICE_META_API = 'https://www.ice.com/marketdata/api/reports/metadata/%d'
ICE_CRITERIA_API = 'https://www.ice.com/marketdata/api/reports/%d/criteria'
ICE_PROBE_REPORTS = (7, 8, 9, 26, 27, 96, 97, 176, 238, 239)
ICE_CONTRACT_REGS = {
    'STIR（无面值原文）': 'https://www.ice.com/publicdocs/contractregs/112_SECTION_NNNN.pdf',
    'Gilt（有面值原文）': 'https://www.ice.com/publicdocs/contractregs/116_SECTION_RRRR.pdf',
}
ICE_MONTHLY_STATS_XLSX = os.path.join(ROOT, 'cache', 'ice_monthly_stats.xlsx')


class FillError(RuntimeError):
    """取数或校验失败。**一律炸，不静默兜底** —— 静默的错在图上看不出来。"""


# ══════════════════════════════════════════════════════════════════════════
# 下载与缓存
# ══════════════════════════════════════════════════════════════════════════
def _cache_path(url):
    name = re.sub(r'[^A-Za-z0-9._-]+', '_', url.split('/')[-1])
    return os.path.join(CACHE, name)


def _download(url, refresh=False):
    """下官方文件，落 cache/basefill/。URL 里有空格，必须先 quote。"""
    path = _cache_path(url)
    if os.path.exists(path) and not refresh:
        with open(path, 'rb') as f:
            return f.read(), path
    os.makedirs(CACHE, exist_ok=True)
    req = urllib.request.Request(urllib.parse.quote(url, safe=':/?&=%'),
                                 headers={'User-Agent': _UA, 'Accept': '*/*'})
    try:
        with urllib.request.urlopen(req, timeout=180) as r:
            body = r.read()
    except Exception as e:                                   # noqa: BLE001
        raise FillError('下载失败 %s: %r' % (url, e)) from e
    if body[:4] != b'PK\x03\x04':
        raise FillError('%s 返回的不是 xlsx（前 8 字节 %r）' % (url, body[:8]))
    with open(path, 'wb') as f:
        f.write(body)
    return body, path


# ══════════════════════════════════════════════════════════════════════════
# Euronext 工作簿解析
# ══════════════════════════════════════════════════════════════════════════
def _load_grid(body, kind):
    """把 Volume / Value of Volume 两张表读成 (行列表, 2019-01 列号) 两元组。

    列号**不写死**：按第 5 行（0-based 第 4 行）表头里 'Jan 19' 那个格子定位。
    Euronext 每年的工作簿列数不同（YTD 列的位置会动），写死行列号迟早错位。"""
    wb = openpyxl.load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    out = {}
    for tag, mapping in (('vol', _SHEET_VOL), ('val', _SHEET_VAL)):
        sheet = mapping[kind]
        if sheet not in wb.sheetnames:
            raise FillError('工作簿里没有表 %r（有的是 %r）' % (sheet, wb.sheetnames))
        rows = list(wb[sheet].iter_rows(values_only=True))
        hdr = None
        for r in rows[:12]:
            cols = [i for i, c in enumerate(r) if isinstance(c, str) and c.strip() == BASE_COL_HDR]
            if cols:
                hdr = cols[0]
                break
        if hdr is None:
            raise FillError('表 %r 里找不到表头 %r' % (sheet, BASE_COL_HDR))
        out[tag] = (rows, hdr)
    wb.close()
    return out


def _cell(row, col):
    v = row[col] if col < len(row) else None
    return str(v).strip() if v not in (None, '') else ''


def _title_at(row, level):
    """指定层级的分节标题；层级左边的列必须为空，否则不算这一层的标题。"""
    if any(_cell(row, k) for k in range(level)):
        return ''
    return _cell(row, level)


def _section_values(grid, title, level=_LVL2, zero_ok=False):
    """取某个分节标题行的 (张数, 名义额欧元)。找不到或找到多行都 raise。"""
    vrows, vcol = grid['vol']
    wrows, wcol = grid['val']
    hits = [i for i, r in enumerate(vrows) if _title_at(r, level) == title]
    if len(hits) != 1:
        raise FillError('分节 %r（第 %d 层）在 Volume 表里命中 %d 行（应恰好 1 行）'
                        % (title, level + 1, len(hits)))
    i = hits[0]
    if _title_at(wrows[i], level) != title:
        raise FillError('Value of Volume 表第 %d 行标题是 %r，与 Volume 表的 %r 不一致 '
                        '—— 两张表行序必须一一对应'
                        % (i, _title_at(wrows[i], level), title))
    vol = vrows[i][vcol]
    val = wrows[i][wcol]
    if vol in (None, '') and val in (None, '') and zero_ok:
        return 0.0, 0.0                      # 当月完全无量，官方留空
    if not isinstance(vol, (int, float)) or not isinstance(val, (int, float)):
        raise FillError('分节 %r 的 2019-01 取值不是数字：vol=%r val=%r' % (title, vol, val))
    return float(vol), float(val) * VALUE_UNIT_EUR


def _contract_row(grid, code):
    """按 Code 列（第 9 列）取单个合约的 (张数, 名义额欧元)。只用于反算校验。"""
    vrows, vcol = grid['vol']
    wrows, wcol = grid['val']
    hits = [i for i, r in enumerate(vrows) if str(r[9]).strip() == code]
    if len(hits) != 1:
        raise FillError('合约代码 %r 命中 %d 行（应恰好 1 行）' % (code, len(hits)))
    i = hits[0]
    vol, val = vrows[i][vcol], wrows[i][wcol]
    if not isinstance(vol, (int, float)) or not isinstance(val, (int, float)) or not vol:
        raise FillError('合约 %r 的 2019-01 取值不可用：vol=%r val=%r' % (code, vol, val))
    return float(vol), float(val) * VALUE_UNIT_EUR


# ══════════════════════════════════════════════════════════════════════════
# 三道校验
# ══════════════════════════════════════════════════════════════════════════
def _reconcile(grid, verbose=True):
    """分节张数 × 交易日数 与 series/enx.csv 的六条池列逐位比对。对不上就 raise。

    这一步同时干两件事：
      (1) 证明「分子（名义额）与分母（池列张数）来自同一个分节」，错配了图上看不出来；
      (2) 反过来**确定**每个 pool_product 由哪些分节组成（见 docstring 一）。"""
    path = os.path.join(SERIES, 'enx.csv')
    with open(path, newline='', encoding='utf-8') as f:
        rows = {r['month']: r for r in csv.DictReader(f)}
    if BASE_MONTH not in rows:
        raise FillError('series/enx.csv 里没有 %s' % BASE_MONTH)
    row = rows[BASE_MONTH]
    days = int(row['trading_days_eqderiv'])
    # 商品那半边 Euronext 用的是同一个交易日数列吗？两列都取出来，不一致就炸。
    days_c = int(row['trading_days_commodity'])
    vrows, vcol = grid['vol']
    hdr_days = None
    for r in vrows[:12]:
        if _cell(r, 0) == 'Number of Trading Days':
            hdr_days = r[vcol]
    if hdr_days is None or int(hdr_days) != days or int(hdr_days) != days_c:
        raise FillError('交易日数对不上：工作簿 %r / enx.csv 股衍 %d / 商品 %d'
                        % (hdr_days, days, days_c))

    ok = True
    for col, subsections in _RECONCILE:
        want = round(float(row[col]) * 1000.0 * days)
        got = 0.0
        for s in subsections:
            got += _section_values(grid, s, _LVL3)[0]
        if abs(got - want) > 0.5:
            ok = False
            print('  ✗ %-38s enx.csv=%d  工作簿(%s)=%d  差 %d'
                  % (col, want, '+'.join(subsections), got, got - want))
        elif verbose:
            print('  ✓ %-38s %10d 张 = %s' % (col, want, ' + '.join(subsections)))
    if not ok:
        raise FillError('分节与 series/enx.csv 对不上 —— 分子分母不同源，不许继续')

    # 第 1 层：登记表要盖全，多出来的分节直接炸（将来 Euronext 加一块新业务不会被静默吞掉）
    # 分节标题一律全大写（'COMMODITY PRODUCTS'），抬头与合计行带小写
    # （'Euronext - Derivatives Market' / 'Total Futures'），用这一条区分，不数行号。
    def _is_section(t):
        return bool(t) and not re.search(r'[a-z]', t)

    seen1 = {t for r in vrows if _is_section(t := _title_at(r, _LVL1))}
    extra1 = seen1 - set(_ENX_LVL1)
    if extra1:
        raise FillError('工作簿第 1 层出现没登记过的分节 %r —— 先决定它属于哪个池再跑'
                        % sorted(extra1))
    # 第 2 层：每一个都必须落进三个池之一，或落进显式忽略表
    known2 = set(_ENX_SECTIONS_IGNORED)
    for names in _ENX_SECTIONS.values():
        known2 |= set(names)
    seen2 = {t for r in vrows if _is_section(t := _title_at(r, _LVL2))}
    extra2 = seen2 - known2
    if extra2:
        raise FillError('工作簿第 2 层出现没登记过的分节 %r —— 先决定它属于哪个池再跑'
                        % sorted(extra2))

    # ENX_MATIF 只收 GRAINS：断言 2019-01 商品那一层的合计确实全部来自 GRAINS
    c_vol, c_val = _section_values(grid, 'COMMODITY PRODUCTS', _LVL1)
    g_vol, g_val = _section_values(grid, 'GRAINS AND OILSEEDS PRODUCTS', _LVL2)
    if abs(c_vol - g_vol) > 0.5 or abs(c_val - g_val) > 1.0:
        raise FillError('2019-01 商品第 1 层合计 %r 与 GRAINS 合计 %r 不等 —— '
                        '说明当月化肥/能源/乳制品/房产也有量，ENX_MATIF 的成员要重定'
                        % ((c_vol, c_val), (g_vol, g_val)))
    if verbose:
        print('  ✓ COMMODITY PRODUCTS 合计 == GRAINS AND OILSEEDS 合计（%d 张）'
              % int(c_vol))


def _verify_value_is_notional(grid, verbose=True):
    """现场证明 "Value of Volume" 那一列是名义额（张数×乘数×成交价），不是权利金。

    办法：拿官方乘数把名义额/张反算成价格，看同一标的的不同合约是否收敛到同一个价。"""
    implied = {}
    for code in ('FCE', 'MFC', 'FTI', 'MFA', 'EBM', 'OBM', 'ECO', 'OCO', 'PXA', 'AEX'):
        vol, val = _contract_row(grid, code)
        implied[code] = val / vol                      # 名义额/张（欧元）
        if code in _OFFICIAL_MULT and verbose:
            mult, unit, quote, url = _OFFICIAL_MULT[code]
            print('  · %-4s 名义额/张 %12.2f €  ÷ 官方乘数 %g (%s) = %10.2f'
                  % (code, implied[code], mult, unit, implied[code] / mult))

    ok = True
    for a, b, ratio, tol, why in _RATIO_CHECKS:
        got = implied[a] / implied[b]
        dev = abs(got / ratio - 1.0)
        flag = '✓' if dev <= tol else '✗'
        if dev > tol:
            ok = False
        if verbose or dev > tol:
            print('  %s %s：%s/%s = %.4f（理论 %.1f，偏 %.2f%%）'
                  % (flag, why, a, b, got, ratio, dev * 100))
    if not ok:
        raise FillError('名义额反算校验失败 —— Value of Volume 这一列的口径与假设不符')

    # 口径差留痕（只报不断言）
    gaps = []
    for fut, opt, why in _BASIS_GAP_PAIRS:
        p_fut = implied[fut] / _OFFICIAL_MULT[fut][0]
        p_opt = implied[opt] / _OFFICIAL_MULT[opt][0]
        gaps.append((why, p_fut, p_opt, p_opt / p_fut - 1.0))
        if verbose:
            print('  ℹ %s：期货 %.1f 点 vs 期权 %.1f 点 → %+.2f%%'
                  % (why, p_fut, p_opt, (p_opt / p_fut - 1.0) * 100))
    return implied, gaps


def _crosscheck_sources(primary, others, verbose=True):
    """三份官方文件在本脚本用到的分节上必须逐位相等。防后发文件回溯修订。"""
    titles = sorted({t for names in _ENX_SECTIONS.values() for t in names})
    base = {t: _section_values(primary, t) for t in titles}
    for tag, grid in others:
        for t in titles:
            got = _section_values(grid, t)
            if abs(got[0] - base[t][0]) > 0.5 or abs(got[1] - base[t][1]) > 1.0:
                raise FillError('%s 与主用文件在分节 %r 上不一致：%r vs %r'
                                % (tag, t, got, base[t]))
        if verbose:
            print('  ✓ %s 与 Derivatives %s.xlsx 在 %d 个分节上逐位一致'
                  % (tag, ENX_PRIMARY_MONTH, len(titles)))


def _verify_specs(verbose=True):
    """现场把三张 Euronext 官方规格页抓下来，核对 _OFFICIAL_MULT 里的原文。

    常数是不是官方的，不靠注释里的一句「我查过」，靠这一步现场证明。
    只影响校验，不影响常数本身（常数不用乘数），所以抓不到只警告不炸。"""
    for code, (mult, unit, quote, url) in _OFFICIAL_MULT.items():
        try:
            req = urllib.request.Request(url, headers={'User-Agent': _UA})
            with urllib.request.urlopen(req, timeout=90) as r:
                html = r.read().decode('utf-8', 'replace')
        except Exception as e:                               # noqa: BLE001
            print('  ⚠ %s 规格页抓取失败（不影响常数）：%r' % (code, e))
            continue
        flat = re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', ' ', html))
        if quote.replace('  ', ' ') in flat:
            if verbose:
                print('  ✓ %-4s 官方规格页原文命中：%r' % (code, quote))
        else:
            print('  ⚠ %-4s 官方规格页里没找到原文 %r —— 规格可能改版，请人工复核 %s'
                  % (code, quote, url))


# ══════════════════════════════════════════════════════════════════════════
# ICE 侧：只取证，不填数
# ══════════════════════════════════════════════════════════════════════════
def ice_probe():
    """现场证明「不是抓不到，是源头没有那个数 / 被 reCAPTCHA 挡住」。"""
    print('\n【ICE 实测】')
    print('① Monthly Statistics Tracking xlsx 里有没有名义额列')
    if os.path.exists(ICE_MONTHLY_STATS_XLSX):
        wb = openpyxl.load_workbook(ICE_MONTHLY_STATS_XLSX, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            heads = []
            for r in ws.iter_rows(min_row=1, max_row=6, values_only=True):
                for c in r[:2]:
                    if isinstance(c, str) and c.strip():
                        heads.append(c.strip())
            print('   · 表 %-24s 抬头：%s' % (sn, ' | '.join(heads[:3])[:90]))
        wb.close()
        print('   ⇒ 只有 ADV(千张)/RPC(美元每张)/OI；唯一的 notional 是 CDS Clearing '
              '表的清算名义额，与本仓三池无关。')
    else:
        print('   ⚠ 没有 %s，先跑 fetch/ice.py' % ICE_MONTHLY_STATS_XLSX)

    print('② 报表中心：哪些历史成交量报表要 reCAPTCHA')
    print('   （HTTP 429 = Cloudflare 按 IP 限流，与 reCAPTCHA 是两回事；连打会触发，'
          '等几分钟重跑即可。真正的卡点是 criteria 恒返回 409 Conflict。）')
    try:
        from curl_cffi import requests as creq          # 只有这一步需要 curl_cffi
    except ImportError:
        print('   ⚠ 没装 curl_cffi，跳过（不影响常数）')
        return
    import json
    import time
    # 本轮实测的 metadata 快照；Cloudflare 限流时用它兜底，保证证据仍然打得出来。
    snap = {}
    snap_path = os.path.join(CACHE, 'ice_report_meta.json')
    if os.path.exists(snap_path):
        with open(snap_path, encoding='utf-8') as f:
            snap = {int(j['id']): j for j in json.load(f) if isinstance(j, dict) and 'id' in j}
        print('   （本地快照 %s：%d 张报表）' % (snap_path, len(snap)))
    for rid in ICE_PROBE_REPORTS:
        try:
            m = creq.get(ICE_META_API % rid, impersonate='chrome', timeout=40,
                         headers={'Accept': 'application/json'})
            if m.status_code != 200:
                j = snap.get(rid)
                print('   · report %-4d metadata HTTP %s%s'
                      % (rid, m.status_code,
                         ('（限流，用快照：%s / %s / recaptcha=%s）'
                          % (j.get('exchange'), j.get('name'), j.get('recaptchaRequired')))
                         if j else '（限流且无快照，等一会再跑）'))
                time.sleep(8)
                continue
            j = json.loads(m.text)
            c = creq.get(ICE_CRITERIA_API % rid, impersonate='chrome', timeout=40,
                         headers={'Accept': 'application/json'})
            print('   · report %-4d %-22s %-38s recaptcha=%-5s criteria HTTP %s'
                  % (rid, str(j.get('exchange'))[:22], str(j.get('name'))[:38],
                     j.get('recaptchaRequired'), c.status_code))
        except Exception as e:                           # noqa: BLE001
            print('   · report %-4d 探测失败 %r' % (rid, e))
        time.sleep(8)
    print('   ⇒ 分合约那几张（26/27/238/239）recaptcha=True、criteria 恒 409 Conflict。')
    print('   ⇒ 绕验证码是禁止动作，且 cron 无人值守复现不了 —— 本脚本不做。')

    print('③ 合约规则原文（面值有没有印在官方规则手册里）')
    for tag, url in ICE_CONTRACT_REGS.items():
        print('   · %s  %s' % (tag, url))
    print('   ⇒ STIR 只印 "Unit of trading €2,500 * Rate Index" / "Contract Multiplier '
          '€2,500"，全篇无 nominal/notional/face value；')
    print('     Gilt 印 "Unit of Trading £100,000 nominal value notional Gilt"。')


# ══════════════════════════════════════════════════════════════════════════
# 输出
# ══════════════════════════════════════════════════════════════════════════
_ICE_UNFILLED_NOTE = {
    'ICE_ENERGY':
        '📌 未填出。三层卡点本轮逐条实测，一层都没拆掉：'
        '① **ICE 全公司没有任何名义额列**（方法① 在 ICE 这边不存在）—— '
        'Monthly Statistics Tracking xlsx 四张表只有 ADV(千张)/RPC/OI，'
        '唯一带 notional 的是 CDS Clearing 的清算名义额；'
        'ice.com/report/7 与 /report/147 的 32 张年度表逐张核过，全部只有张数。'
        '② 分合约张数被 reCAPTCHA 挡死：ICE 报表中心里**凡是到合约层的历史成交量报表都要 '
        'reCAPTCHA** —— '
        'report 26/27（ICE Futures Europe 日成交）与 238/239（ICE Endex，TTF 在这里），'
        'metadata 的 recaptchaRequired=true，/marketdata/api/reports/{id}/criteria '
        '不带 token 恒返回 HTTP 409 Conflict。绕验证码是禁止动作，cron 也复现不了。'
        '③ 就算过了 reCAPTCHA 还差两条：pools.py 把本池挂在 ice.csv 的 '
        'adv_energy_kcontracts（全球口径，2019-01 = 2,718 千张/日 × 21 = 57,078,000 张，'
        '含 IFEU/Endex/IFUS/IFAD 四个所），而 26/27 只有 IFEU、238/239 只有 Endex，'
        '得拼四个所；且 **TTF 没有固定乘数**（官方原文 "1 MW per day in contract period '
        '... x 23, 24 or 25 hours"，一张月度合约 672–744 MWh 不等），必须按合约期逐月折算。'
        '下一步唯一合规的出路：人工在浏览器里跑一次 report 26/27/238/239 导出 2019-01 '
        '逐合约日成交 CSV 放进 cache/basefill/，并把原始 CSV 一起提交。'
        '⚠ 跨币种（Brent USD/桶、TTF EUR/MWh），记账币定 USD，各成员面值须先按 '
        'series/fx.csv 的 2019-01 月均汇率折美元再加权（fx_avg_eurusd=1.141640909090909）。',
    'ICE_STIR':
        '📌 未填出。① 权重：ICE 报表中心里到合约层的历史成交量报表（report 26/27，ICE Futures '
        'Europe 日成交）recaptchaRequired=true，criteria 接口恒 409（HTTP 429 是 Cloudflare 按 IP 限流，'
        '与 409 不是一回事）；非 reCAPTCHA 的 '
        'report 7 把 STIR 与 MLTIR 并成一列 "Interest Rates"（2019-01 期货 37,491,346 + '
        '期权 7,632,263），拆不开。权重错一点常数差一倍以上（Euribor 与 Short Sterling '
        '的面值差 2 倍、与 SONIA 差 1 倍且币种不同），不能拍。'
        '② 面值：本轮把官方**合约规则手册**原文抠出来了 —— SECTION NNNN《CONTRACT RULES: '
        'ICE FUTURES SHORT TERM INTEREST RATE INDEX FUTURES CONTRACTS》'
        '(https://www.ice.com/publicdocs/contractregs/112_SECTION_NNNN.pdf) 对三个月 '
        'EURIBOR 只印 "Unit of trading €2,500 * Rate Index" 与 "Contract Multiplier '
        '€2,500"，通篇没有 nominal / notional / face value 字样，**EUR 1,000,000 是反推值**，'
        '按既定规矩不许写进表。这条第一轮只在产品页层面看到，本轮在规则手册层面确认。'
        '③ 一个有用的旁证：ice.csv 的 adv_stir+adv_mltir（2019-01 合计 45,122,000 张/月）'
        '与 report/7 的 IFEU Interest Rates 合计 45,123,609 张只差 0.004% ⇒ 这两池几乎 '
        '100% 在 ICE Futures Europe，report 26/27 一旦拿到，权重就是完整的，不缺别的所。'
        '⚠ 跨币种（EUR/GBP/CHF），记账币定 USD，各成员面值须先按 series/fx.csv 的 '
        '2019-01 月均汇率折美元再加权（fx_avg_eurusd=1.141640909090909、'
        'fx_avg_gbpusd=1.2887195243293936、fx_avg_chfusd=1.010591626961707）。',
    'ICE_MLTIR':
        '📌 未填出，但比 ICE_STIR 少一层卡点：**面值有官方原文**。'
        'SECTION RRRR《CONTRACT RULES: ICE FUTURES GILT FUTURES CONTRACTS》'
        '(https://www.ice.com/publicdocs/contractregs/116_SECTION_RRRR.pdf) 原文 '
        '"Unit of Trading £100,000 nominal value notional Gilt"（Long/Medium/Short 等 '
        '四个 Gilt 合约同值）、"A \'lot\' shall be an amount having a nominal value of '
        '£100,000 (the \'delivery amount\') of a Deliverable Gilt"。'
        '缺的只有**权重**：ICE 月报脚注 (10) 说本池含 Gilt、Swapnotes、日本国债、'
        '欧元区国债、美国国债与 Ultrabond，官方到合约层的 report 26/27 '
        'recaptchaRequired=true（criteria 恒 409），report 7 又把 STIR/MLTIR 并成一列。'
        'Gilt 在 2019-01 大概率占绝对多数，但「大概率」不是实测值，不许写。'
        '出路同 ICE_STIR：人工导出 report 26/27 的 2019-01 逐合约日成交并提交原始 CSV。'
        '⚠ 跨币种（GBP 为主，含 EUR/JPY/USD），记账币定 USD，面值须先按 series/fx.csv 的 '
        '2019-01 月均汇率折美元再加权（fx_avg_gbpusd=1.2887195243293936）。',
}


# 「为什么这条路能绕开第一轮的死因」——每个池各自的那一句
_WHY = {
    'ENX_INDEX_DERIV':
        'CAC 40（€10/点）与 AEX（€200/点）乘数差 20 倍、必须分开加权这个坑因此不存在 —— '
        '不是绕过去了，是根本不需要按合约拆开再加权。',
    'ENX_SINGLESTOCK_LEGACY':
        '单股期权乘数不统一（各标的不同、还有公司行动调整）、800 多个标的逐个找乘数与股价'
        '不可行这个坑因此不存在。',
    'ENX_MATIF':
        '小麦/玉米/菜籽虽同为 50 公吨/张，但 2019-01 价格差近一倍'
        '（菜籽 369.67 €/t vs 玉米 181.63 €/t，本工作簿逐合约反算），'
        '按张数加权还是按品种加权结果不同 —— 官方名义额已经替我们加过了。',
}


def build_rows(grid, implied, gaps):
    """算出三个 Euronext 常数 + 三个 ICE 空行，返回待写的 CSV 行。"""
    src = ('%s | %s | %s'
           % (ENX_STATS_LANDING,
              (ENX_DERIV_MONTHLY % ENX_PRIMARY_MONTH),
              (ENX_FACTBOOK % ENX_FACTBOOK_YEAR)))
    out = []
    detail = {}
    for pid, titles in _ENX_SECTIONS.items():
        vol = val = 0.0
        parts = []
        for t in titles:
            v, w = _section_values(grid, t)
            vol += v
            val += w
            parts.append('%s（%s 张 / %s €）' % (t, format(int(v), ','), format(int(round(w)), ',')))
        if vol <= 0:
            raise FillError('%s 的 2019-01 张数是 %r' % (pid, vol))
        const = val / vol
        detail[pid] = (vol, val, const, parts)
        out.append({'product_id': pid, 'const': const, 'src': src, 'parts': parts,
                    'vol': vol, 'val': val})
    return out, detail


def write_csv(enx_rows, detail, implied, gaps):
    basis_txt = '；'.join('%s 期货 %.1f 点 vs 期权 %.1f 点，差 %+.2f%%'
                          % (why.split(' 期权')[0], pf, po, d * 100)
                          for why, pf, po, d in gaps)
    rows = []
    for r in enx_rows:
        pid = r['product_id']
        vol, val, const, parts = detail[pid]
        note = (
            '基期 2019-01 篮子常数 = Euronext 官方月度衍生品统计的 '
            '"Monthly Value of Volume (€ \'000)" ÷ "Monthly Volume"，'
            '× 1000 还原成欧元：%s € ÷ %s 张 = %.6f €/张。'
            '成员分节 = %s。'
            '**加权已内含在官方名义额里** —— 不需要乘数表、不需要价格序列、不需要自己按张数加权。%s'
            '取数文件（三份官方文件，2019-01 逐位互校一致）：'
            'Derivatives %s.xlsx（主用，2019 全年 YTD 逐月列）、'
            'Derivatives %s.xlsx（2019 年官网现存最早的一份，201901–201904 实测 403）、'
            'Euronext Fact Book %d.xlsx（年鉴里的同名分表）。'
            '分节张数已与 series/enx.csv 的六条池列（× 22 个交易日）逐位对账通过，'
            '证明分子与分母同源；对不上脚本直接 raise。'
            '"Value of Volume" 是名义额不是权利金：工作簿另有独立的 Premium Turnover 表；'
            '且用官方乘数反算，同一标的的标准/迷你、期货/期权收敛到同一价位 —— '
            'FCE÷€10=4,800.5 点 与 MFC÷€1=4,795.6 点 差 0.10%%，'
            'FTI÷€200=499.9 点 与 MFA÷€20=501.4 点 差 0.31%%，'
            'EBM÷50t=202.83 €/t 与 OBM÷50t=200.21 €/t 差 1.31%%，'
            'ECO÷50t=369.67 €/t 与 OCO÷50t=371.09 €/t 差 0.38%%（脚本每次运行现场跑这四条）。'
            '⚠ 口径差必须留痕：contract_specs.csv 表头把 base_price_basis 定成 avg_close'
            '（月内交易日收盘价等权平均），而这里隐含的价格项是**该月成交量加权的成交价**，'
            '两者不是同一个东西；本脚本实测的量级：%s。'
            '与第一轮 EUREX_INDEX / EUREX_EQUITY 走的是同构的一条路'
            '（Capital Volume ÷ Traded Contracts），三行口径一致、可并排比。'
            '复算脚本 build/basefill/ice_enx2.py（联网、手动跑）。'
            % (format(int(round(val)), ','), format(int(vol), ','), const,
               ' + '.join(parts), _WHY[pid], ENX_PRIMARY_MONTH, ENX_CROSSCHECK_MONTHS[0],
               ENX_FACTBOOK_YEAR, basis_txt)
        )
        rows.append({
            'product_id': pid,
            'base_price_local': '%.6f' % const,
            'base_ccy': 'EUR',
            'basket_constant': '%.6f' % const,
            'source_url': r['src'],
            'computed_note': note,
        })
    for pid in ('ICE_ENERGY', 'ICE_STIR', 'ICE_MLTIR'):
        rows.append({
            'product_id': pid,
            'base_price_local': '',
            'base_ccy': 'USD',
            'basket_constant': '',
            'source_url': 'https://www.ice.com/marketdata/reports | '
                          'https://www.ice.com/report/7 | '
                          'https://www.ice.com/publicdocs/contractregs/',
            'computed_note': _ICE_UNFILLED_NOTE[pid],
        })
    os.makedirs(SERIES, exist_ok=True)
    with open(OUT, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=['product_id', 'base_price_local', 'base_ccy',
                                          'basket_constant', 'source_url', 'computed_note'])
        w.writeheader()
        for r in rows:
            w.writerow(r)
    return rows


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument('--refresh', action='store_true', help='强制重下所有官方文件')
    ap.add_argument('--ice-probe', action='store_true', help='额外跑 ICE 侧的实测取证')
    ap.add_argument('--no-verify-specs', action='store_true', help='跳过官方规格页现场核对')
    args = ap.parse_args(argv)

    print('【1/5】下载 Euronext 官方统计文件')
    primary_url = ENX_DERIV_MONTHLY % ENX_PRIMARY_MONTH
    body, path = _download(primary_url, args.refresh)
    print('  主用 %s（%d 字节）' % (os.path.basename(path), len(body)))
    grid = _load_grid(body, 'deriv')

    others = []
    for mm in ENX_CROSSCHECK_MONTHS:
        b, p = _download(ENX_DERIV_MONTHLY % mm, args.refresh)
        others.append(('Derivatives %s.xlsx' % mm, _load_grid(b, 'deriv')))
        print('  互校 %s（%d 字节）' % (os.path.basename(p), len(b)))
    b, p = _download(ENX_FACTBOOK % ENX_FACTBOOK_YEAR, args.refresh)
    others.append(('Euronext Fact Book %d.xlsx' % ENX_FACTBOOK_YEAR, _load_grid(b, 'factbook')))
    print('  互校 %s（%d 字节）' % (os.path.basename(p), len(b)))

    print('\n【2/5】三份官方文件在 2019-01 上逐位互校')
    _crosscheck_sources(grid, others)

    print('\n【3/5】分节 vs series/enx.csv 池列对账（同时确定成员分节）')
    _reconcile(grid)

    print('\n【4/5】现场证明 "Value of Volume" 是名义额')
    if not args.no_verify_specs:
        _verify_specs()
    implied, gaps = _verify_value_is_notional(grid)

    print('\n【5/5】算常数并写出')
    enx_rows, detail = build_rows(grid, implied, gaps)
    rows = write_csv(enx_rows, detail, implied, gaps)
    for pid in ('ENX_INDEX_DERIV', 'ENX_SINGLESTOCK_LEGACY', 'ENX_MATIF'):
        vol, val, const, parts = detail[pid]
        print('  %-24s %14.6f EUR/张   （%s 张，%s €）'
              % (pid, const, format(int(vol), ','), format(int(round(val)), ',')))
    print('  ICE_ENERGY / ICE_STIR / ICE_MLTIR   留空（卡点见 computed_note 与模块 docstring 二）')
    print('\n写出 %s（%d 行）' % (OUT, len(rows)))

    if args.ice_probe:
        ice_probe()
    return 0


if __name__ == '__main__':
    sys.exit(main())
