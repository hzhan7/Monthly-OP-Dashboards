# -*- coding: utf-8 -*-
"""HOOD **季度**序列的历史回填：series/hood_q.csv 从 2023Q2 起推到 2021Q1 起（+9 期）。

用法:
    python3 build/basefill/hood_q_2021.py                      # 取数 + 核对 + 写 CSV
    python3 build/basefill/hood_q_2021.py --dry                # 只核对、只打印，不写
    python3 build/basefill/hood_q_2021.py --refresh            # 强制重下原件
    python3 build/basefill/hood_q_2021.py --cache-dir <路径>   # 原件落在哪

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
为什么 series/hood_q.csv 以前只有 13 期，以及那**不是**数据的边界
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
fetch/hood.py 每季只吃**最新那一份** Earnings Supplement，而那份文件的
「Quarterly GAAP P&L」与「Quarterly KPIs」两页都是**滚动 13 个季度**的窗口
（Q2'26 那份 = 2023Q2…2026Q2，Q1'26 那份 = 2023Q1…2026Q1）。update() 又只追加、
不回头，于是 series/hood_q.csv 的左端就永远停在「本页面第一次接上管道那一季」。

也就是说 2023Q2 是**管道的**起点，不是**数据的**起点。IR 站 Quarterly Results 页上
至今还挂着覆盖 2021Q1 的旧 Supplement，逐季逐行都印着本 CSV 要的每一列。
本脚本把那一段一次性补进来，与 build/basefill/hood_2021.py（月度那一半）同一套做法、
同一批源文件，跑完这个洞就永远关上了，所以同样放在 basefill/ 而不是塞进无人值守链路。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
天花板就是 2021Q1
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Robinhood 2021-07 才 IPO。现今 IR 站上覆盖最早的一份是随 Q1'23 财报发的
robinhood-markets-q1-2023-supplemental-information.xlsx，两页都从 **2021Q1** 起
（P&L 表头 "Q1'21"，KPI 表头 2021/Q1），再往前只有 S-1 里的年度数，不是季度粒度。
所以本脚本的 FIRST = 2021Q1，**不硬凑**。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
两个源，各管一段；取的一律是「能拿到的最早那一版」
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  ① LEGACY  robinhood-markets-q1-2023-supplemental-information.xlsx（随 Q1'23 财报发）
     P&L 与 KPI 两页都覆盖 2021Q1~2023Q1。喂 **2021Q1~2022Q4** 这 8 期。
  ② MODERN  Q1'26 Earnings Supplement.xlsx，两页都覆盖 2023Q1~2026Q1。
     只喂 **2023Q1** 这 1 期 —— 与已有行（2023Q2 起，解自现代版式）同源同尺子。
  ③ SECLEND Q4'23 Earnings Supplement.xlsx **不入库**，只当第三方证人：
     它同时覆盖 2021Q1~2023Q4，用来验「① 与 ② 之间没有断层、也没有被重述」。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
留空的两列：官方**本就没有**，不补 0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
· rev_event_usdmn —— 事件合约收入是 **Q2'26 那份** Supplement 才从 'Other' 里拆出来的
  一行（Q1'26 及更早的 P&L 只有 Options / Cryptocurrencies / Equities / Other 四行，
  且 Other 恒等于「现今的 Other + Event」，本脚本 check_other() 每次运行都重验一遍）。
  Q2'26 那份的滚动窗口从 2023Q2 起 —— 于是 2023Q1 及更早，**没有任何一份官方文件
  印过事件合约收入**。业务本身 2024-10 才上线，事后看它就是 0，但那是我们的推断不是
  公司的披露，所以留空（与 build/basefill/hood_2021.py 对 vol_event_bn 的处理同规矩）。
· q_vol_event_bn —— 2023Q1 那一期 ② 印了 0（Event Contracts (B) 那一行有），照抄；
  2021Q1~2022Q4 那 8 期 ① 根本没有这一行，留空。

留空的后果是显式的、可被引擎处理的：build/hood.py 的 _rate() 见到成交量非有限或 ≤ 0
就返回 NaN，费率图按 null 断笔；Exhibit 17（stacked_dual，平滑图型不吃 null）会由
mrwin.resolve() 自动把左端截在「四段都有值」的 2023Q2，并自己写出那句为什么。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
四道看门狗（任何一道不过就抛异常、一格都不写）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
A check_modern()  ② 与**已有的** 2023Q2~2026Q1 逐格比 —— 除 rev_event / rev_other_txn
                  这对拆分列外必须全同（见 check_other）。保证「拿 ② 填 2023Q1」
                  和「已有行的来源」是同一把尺子。
B check_other()   ② 的 'Other' ≡ 已有行的 rev_event + rev_other_txn，逐季验。
                  这是「旧版式的 Other 就是新版式 Other+Event 之和」这句话的证据。
C check_seam()    ① 与 ② 在 2023Q1 这一期重叠（11 列），必须逐位相同。
D check_witness() ③ 与 ① 在 2021Q1~2022Q4、与 ② 在 2023Q1 逐格比，报告差异。
                  只报告不阻断（后期重述本来就允许存在），但入库一律取最早那一版。
"""
import argparse
import csv
import importlib.util
import os
import re

import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SERIES_DIR = os.path.join(ROOT, 'series')
CSV_PATH = os.path.join(SERIES_DIR, 'hood_q.csv')

FIRST = pd.Period('2021Q1', 'Q')          # 天花板，见文件头
GUARD = pd.Period('2023Q2', 'Q')          # 本脚本一格都不许写到这一季及以后


class HoodQBasefillError(RuntimeError):
    pass


_F = _B = None


def _fetch():
    """借 fetch/hood.py 的 `_norm` / `_fmt_like`（标签归一化与列内格式对齐）。"""
    global _F
    if _F is None:
        spec = importlib.util.spec_from_file_location(
            'fetch_hood', os.path.join(ROOT, 'fetch', 'hood.py'))
        _F = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(_F)
    return _F


def _basefill_m():
    """借月度回填脚本的 `download_all` / `SOURCES` —— 源文件与 uuid 兜底表只维护一份。"""
    global _B
    if _B is None:
        spec = importlib.util.spec_from_file_location(
            'hood_2021', os.path.join(HERE, 'hood_2021.py'))
        _B = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(_B)
    return _B


# ─────────────────────────── 解析 ───────────────────────────
# 季度 P&L：(章节, 标签) → CSV 列名。章节是必须的 —— 'Other' 在收入段与费用段各出现过。
PL_MAP = {
    'options': 'rev_options_usdmn',
    'cryptocurrencies': 'rev_crypto_usdmn',
    'crypto': 'rev_crypto_usdmn',                  # 2023 及更早的文件用短名
    'equities': 'rev_equities_usdmn',
    'event contracts': 'rev_event_usdmn',          # Q2'26 才拆出来，旧文件没有
    'other': 'rev_other_txn_usdmn',
    'transaction-based revenues': 'rev_transaction_usdmn',
    'net interest revenues': 'rev_net_interest_usdmn',
    'other revenues': 'rev_other_usdmn',
    'total net revenues': 'rev_total_usdmn',
}
PL_SECTION = 'revenues:'

# 季度 KPI：只取「Total Trading Volumes」那一节。余额类在季度表里是季末月时点值，不是
# 季度合计，混进来就是两把尺子（fetch/hood.py 的 QVOL_SPEC 也是这么划的）。
QVOL_MAP = {
    'equity ($b)': 'q_vol_equity_usdbn',
    'options contracts (m)': 'q_vol_options_mn',
    'crypto ($b)': 'q_vol_crypto_usdbn',
    'event contracts (b)': 'q_vol_event_bn',
}
QVOL_SECTION = 'total trading volumes'

_QPAT = re.compile(r"^q\s*([1-4])\b")
_YR2 = re.compile(r"'(\d{2})")
_YR4 = re.compile(r'\b(20\d{2})\b')


def _sheet(wb, want):
    """按前缀挑 sheet：'Quarterly GAAP P&L (in millions' 这种名字被 Excel 截过，
    各文件长度不一，不能按全名比。"""
    for s in wb.sheetnames:
        if s.strip().lower().startswith(want):
            return wb[s]
    raise HoodQBasefillError(f'找不到以「{want}」开头的 sheet；有的是 {wb.sheetnames}')


def _qcols(rows, W):
    """扫出季度列 → [(列下标, Period)]。

    三种表头都要吃：
      · Q1'23 P&L      年在上一行、季标签自带年（"Q1'21" / 'Q2 2021 '）
      · Q4'23 / Q1'26  年在上一行、季标签只有 'Q1'…'Q4'
      · Q1'23 KPI      年在上一行（第 2 行）、季标签 'Q1'…'Q4'，右边还有 'Qr' / 'Yr'
    'Q/Q'、'Qr' 这类变化率列不匹配 `^q[1-4]\\b`，自然被挡掉。
    """
    pr = None
    for ri in range(min(12, len(rows))):
        hits = sum(1 for i in range(W) if _QPAT.match(str(rows[ri][i] or '').strip().lower()))
        if hits >= 4:
            pr = ri
            break
    if pr is None:
        raise HoodQBasefillError('找不到季度表头行')
    yr_row = rows[pr - 1] if pr else [None] * W
    cols, cur = [], None
    for i in range(W):
        v = yr_row[i]
        if v is not None and str(v).replace(',', '').strip().isdigit():
            n = int(str(v).replace(',', '').strip())
            if 2000 <= n <= 2099:
                cur = n
        lab = str(rows[pr][i] or '').strip()
        m = _QPAT.match(lab.lower())
        if not m:
            continue
        y = (int('20' + _YR2.search(lab).group(1)) if _YR2.search(lab)
             else int(_YR4.search(lab).group(1)) if _YR4.search(lab) else cur)
        if y is None:
            raise HoodQBasefillError(f'季度列「{lab}」定不出年份')
        cols.append((i, pd.Period(f'{y}Q{m.group(1)}', 'Q')))
    if not cols:
        raise HoodQBasefillError('表头找到了但没解析出季度')
    return cols


def _scan(ws, section, mapping):
    """→ {CSV 列名: {Period: 值}}。按 (章节, 标签) 两级定位，第一次命中为准。

    标签可能落在 A/B/C 任一列（Q1'23 那份的 A 列还躺着一套与 B 列**不对行**的旧标签），
    所以三列都当候选；但只有 `section` 那一节里的行才会被采纳，节外同名行进不来。
    """
    F = _fetch()
    W = ws.max_column
    rows = [list(r) + [None] * (W - len(r)) for r in ws.iter_rows(values_only=True)]
    cols = _qcols(rows, W)
    out, inside = {}, False
    for r in rows:
        labs = [F._norm(r[i]) for i in range(min(3, W)) if r[i] is not None and str(r[i]).strip()]
        if not labs:
            continue
        numeric = any(isinstance(r[i], (int, float)) for i, _ in cols)
        if not numeric:                                   # 章节标题行 / 全空行
            if any(l.startswith(section) for l in labs):
                inside = True
            elif any(l.endswith(':') or l.startswith('operating expenses') for l in labs):
                inside = False
            continue
        if not inside:
            continue
        for lab in labs:
            col = mapping.get(lab)
            if col and col not in out:
                out[col] = {p: (r[i] if isinstance(r[i], (int, float)) else None)
                            for i, p in cols}
                break
    return out


def read_quarters(path):
    """一份 xlsx → {CSV 列名: {Period: 值}}（P&L 与 KPI 两页合起来）。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    got = _scan(_sheet(wb, 'quarterly gaap p&l'), PL_SECTION, PL_MAP)
    got.update(_scan(_sheet(wb, 'quarterly kpis'), QVOL_SECTION, QVOL_MAP))
    return got


# 每一份都必须解出来的列（rev_event / q_vol_event 只有新版式才有，单独判）
CORE_COLS = [c for c in dict.fromkeys(list(PL_MAP.values()) + list(QVOL_MAP.values()))
             if c not in ('rev_event_usdmn', 'q_vol_event_bn')]


def require(got, name, extra=()):
    miss = [c for c in list(CORE_COLS) + list(extra) if c not in got]
    if miss:
        raise HoodQBasefillError(
            f'{name}: 这些列一行都没匹配上 {miss} —— 官方改标签了，先人工看一眼')
    return got


# ─────────────────────────── 核对 ───────────────────────────
def _csv():
    with open(CSV_PATH, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    return rows[0], [r for r in rows[1:] if r and r[0].strip()]


SPLIT_COLS = ('rev_event_usdmn', 'rev_other_txn_usdmn')


def check_modern(mod, header, body):
    """看门狗 A：② 与已有行逐格比。拆分列（Other ↔ Other+Event）单独走 check_other。"""
    have = {r[0]: dict(zip(header, r)) for r in body}
    n = bad = 0
    unexpected = []
    for key, row in sorted(have.items()):
        p = pd.Period(key, 'Q')
        for c in header[1:]:
            if c in SPLIT_COLS:
                continue
            txt = (row.get(c) or '').strip()
            v = mod.get(c, {}).get(p)
            if not txt or v is None:
                continue
            n += 1
            if abs(float(txt) - float(v)) > 1e-9:
                bad += 1
                unexpected.append((key, c, txt, v))
    print(f"\n── 看门狗 A：MODERN(Q1'26) vs 已有 series 行，{n} 格（不含拆分列）──")
    print(f'  一致 {n - bad} 格；不一致 {bad} 格')
    if unexpected:
        for it in unexpected[:12]:
            print(f"    {it[0]} {it[1]}: CSV={it[2]} Q1'26={it[3]}")
        raise HoodQBasefillError("MODERN 与已有行对不上 —— 不是同一把尺子，别写")
    return n


def check_other(mod, header, body):
    """看门狗 B：② 的 'Other' ≡ 已有行的 rev_event + rev_other_txn。

    这是「旧版式那一行 Other 里就含着后来拆出去的 Event」这句话的全部依据 ——
    验过了才敢说 2021~2023Q1 的 rev_other_txn 与今天的同列可比。
    """
    have = {r[0]: dict(zip(header, r)) for r in body}
    n = bad = 0
    diffs = []
    for key, row in sorted(have.items()):
        p = pd.Period(key, 'Q')
        v = mod.get('rev_other_txn_usdmn', {}).get(p)
        a, b = (row.get('rev_event_usdmn') or '').strip(), (row.get('rev_other_txn_usdmn') or '').strip()
        if v is None or not a or not b:
            continue
        n += 1
        if abs(float(a) + float(b) - float(v)) > 1e-9:
            bad += 1
            diffs.append((key, a, b, v))
    print(f"\n── 看门狗 B：Q1'26 的 Other ≟ 现 rev_event + rev_other_txn，{n} 季 ──")
    print(f'  恒等 {n - bad} 季；对不上 {bad} 季')
    if diffs:
        for it in diffs[:12]:
            print(f'    {it[0]}: {it[1]} + {it[2]} ≠ {it[3]}')
        raise HoodQBasefillError('Other 拆分恒等式不成立 —— 旧版式的 Other 不能当同一列用')
    if not n:
        raise HoodQBasefillError('一季都没验到 —— 拆分恒等式没有证据，别写')
    return n


def check_seam(leg, mod):
    """看门狗 C：① 与 ② 在 2023Q1 的重叠必须逐位相同。"""
    p = pd.Period('2023Q1', 'Q')
    bad, n = [], 0
    for col in CORE_COLS:
        a, b = leg.get(col, {}).get(p), mod.get(col, {}).get(p)
        if a is None or b is None:
            bad.append((col, a, b))
            continue
        n += 1
        if abs(float(a) - float(b)) > 1e-9:
            bad.append((col, a, b))
    print(f'\n── 看门狗 C：老版式 vs 新版式在 2023Q1 的重叠（{len(CORE_COLS)} 列）──')
    print(f'  逐位相同 {n - len([x for x in bad if x[1] is not None])} 列；对不上 {len(bad)} 列')
    if bad:
        for it in bad:
            print(f'    {it[0]}: 老={it[1]} 新={it[2]}')
        raise HoodQBasefillError('两种版式在 2023Q1 对不上 —— 行标签映射错了，别写')
    return n


def check_witness(leg, mod, wit, quarters):
    """看门狗 D：③（Q4'23）当第三方证人。只报告差异，入库仍取最早那一版。"""
    print("\n── 看门狗 D：Q4'23 证人 vs 入库源 ──")
    total = 0
    for col in CORE_COLS:
        d = []
        for p in quarters:
            src = (leg if p < pd.Period('2023Q1', 'Q') else mod).get(col, {}).get(p)
            w = wit.get(col, {}).get(p)
            if src is None or w is None:
                continue
            if abs(float(src) - float(w)) > 1e-9:
                d.append((str(p), src, w))
        total += len(d)
        if d:
            print(f'  {col:<26} 差 {len(d)} 季：'
                  + '、'.join(f'{m} 入库{x}→证人{y}' for m, x, y in d[:6]))
    print(f'  合计 {total} 处差异（0 = 三份文件对这段历史完全一致）')
    return total


# ─────────────────────────── 写盘 ───────────────────────────
def build_rows(leg, mod, quarters):
    rows = {}
    for p in quarters:
        src = leg if p < pd.Period('2023Q1', 'Q') else mod
        rec = {c: v[p] for c, v in src.items()
               if v.get(p) is not None and not pd.isna(v[p])}
        if rec:
            rows[p] = rec
    over = [str(p) for p in rows if p >= GUARD]
    if over:
        raise HoodQBasefillError(f'这些季度 ≥ {GUARD}，本脚本不许写：{sorted(over)[:6]}')
    under = [str(p) for p in rows if p < FIRST]
    if under:
        raise HoodQBasefillError(f'这些季度 < {FIRST}（官方季度披露的天花板）：{sorted(under)[:6]}')
    return rows


def write(rows, dry=False):
    F = _fetch()
    header, body = _csv()
    have = {r[0]: list(r) for r in body}
    cols = header[1:]
    existing = {c: [r[i + 1] for r in body] for i, c in enumerate(cols)}

    added, filled, skipped = [], 0, 0
    for p in sorted(rows):
        key = str(p)
        row = have.get(key) or ([key] + [''] * len(cols))
        if key not in have:
            added.append(key)
        for i, c in enumerate(cols):
            if c not in rows[p]:
                continue
            if (row[i + 1] or '').strip():        # 已有值永不覆盖
                skipped += 1
                continue
            row[i + 1] = F._fmt_like(existing[c], rows[p][c])
            filled += 1
        have[key] = row

    out = [have[k] for k in sorted(have, key=lambda s: pd.Period(s, 'Q'))]
    blanks = {c: sum(1 for r in out if not (r[i + 1] or '').strip())
              for i, c in enumerate(cols)}
    print('\n── 写盘 ──')
    print(f'  新建 {len(added)} 行、填 {filled} 格、跳过（已有值）{skipped} 格')
    print('  留空的列：'
          + ('、'.join(f'{c}×{n}' for c, n in sorted(blanks.items()) if n) or '（无）'))
    if dry:
        print('  --dry：**未**写入')
        for k in sorted(rows)[:2] + ['…'] + sorted(rows)[-1:]:
            print('    ' + (str(k) if k == '…' else ','.join(have[str(k)])))
        return 0
    with open(CSV_PATH, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(out)
    print(f'✓ {CSV_PATH}：现共 {len(out)} 行，{out[0][0]} → {out[-1][0]}')
    return len(added)


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry', action='store_true')
    ap.add_argument('--refresh', action='store_true')
    ap.add_argument('--cache-dir', default=os.path.join(ROOT, 'cache'))
    a = ap.parse_args(argv)

    print('── 下载 / 取缓存 ──')
    paths = _basefill_m().download_all(a.cache_dir, a.refresh)

    leg = require(read_quarters(paths['LEGACY']), "LEGACY(Q1'23)")
    mod = require(read_quarters(paths['MODERN']), "MODERN(Q1'26)", extra=('q_vol_event_bn',))
    wit = require(read_quarters(paths['SECLEND']), "WITNESS(Q4'23)")

    header, body = _csv()
    check_modern(mod, header, body)
    check_other(mod, header, body)
    check_seam(leg, mod)

    quarters = [pd.Period(f'{y}Q{q}', 'Q') for y in (2021, 2022) for q in (1, 2, 3, 4)]
    quarters.append(pd.Period('2023Q1', 'Q'))
    check_witness(leg, mod, wit, quarters)

    rows = build_rows(leg, mod, quarters)
    return write(rows, dry=a.dry)


if __name__ == '__main__':
    raise SystemExit(0 if main() is not None else 1)
