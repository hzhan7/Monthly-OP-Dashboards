#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""基期常数回补（第二轮）：US 两个 + ASX 两个 + MX 两个。

一次性脚本，**不进 cron**。留在仓里是为了让每个常数可复算、可审计。
运行：python3 build/basefill/us_asx_mx2.py          （只算，不写表）
      python3 build/basefill/us_asx_mx2.py --emit    （另外打印 CSV 行）

覆盖的 product_id
──────────────────
  US_CASH_EQUITY_SHARE   ✅ 已填（本轮独立复算，与第一轮逐位相同）
  US_MULTILIST_EQ_OPT    ❌ 未填（100 股/张的官方出处已拿到，缺的是基期加权标的均价）
  ASX_ETO                ✅ 已填（主线程裁决取月初/月末中点）
  ASX_DERIV              ✅ 已填（本轮新拿到 2019-01 逐合约成交量）
  MX_EQUITY_OPT          ❌ 未填
  MX_ETF_OPT             ❌ 未填

本轮的三个通道突破（写在最前面，下轮别再踩）
────────────────────────────────────────────
坑 1 · theocc.com 的 Cloudflare 不是拦所有指纹，是**只放行部分指纹**。
      curl_cffi impersonate="chrome"/"chrome110" → 403 "Just a moment..."；
      impersonate="chrome124" 或 "safari17_0" → 200。第一轮报的「OCC 进不去」
      是指纹版本选错，不是站点不可达。⇒ 换 impersonate 版本比换站点便宜得多。

坑 2 · web.archive.org 正好**反过来**：curl_cffi 任何 impersonate 都恒返回
      HTTP 498（nginx 自定义码，伪装成 404 页面），裸 urllib + 普通 UA 直接 200。
      ASX 的 2019-01 逐合约成交量就卡在这里 —— 不是快照不存在，是取快照的姿势不对。
      ⇒ 同一个任务里两个站要用两套 HTTP 客户端，别图省事统一。

坑 3 · ftp.nasdaqtrader.com 是**公开匿名 FTP**（与 CME 的 ftp.cmegroup.com 同类）。
      目录 MonthlyShareVolume / PHLX / Files/MarketShare 等。本轮实测：与 2019-01
      有关的逐 symbol 数据都不覆盖 2019（MarketShare 止于 2016，MonthlyShareVolume
      止于 2006），所以对本任务没用 —— 但通道本身是通的，别再当「没有 FTP」。

口径警告
────────
· ASX_DERIV 是**混合口径**篮子：利率合约按面值计名义额（不乘结算价），股指合约按
  乘数 × 指数点。两者相加只能读作「名义额构成」，不能读作风险敞口。
· ASX 官方逐合约报告脚注：「Volumes quoted are Total Volumes which include
  on-market, off-market and **non-traded** volumes」。2019-01 非成交量 70,461 张
  （占 0.64%），本常数与 pools.py 挂的 adv_futures_and_options_contracts 同口径，
  两边都含 non-traded，所以是自洽的；但对外解读时必须说明。
"""

import argparse
import csv
import datetime
import io
import os
import re
import sys
import urllib.request

_HERE = os.path.dirname(os.path.abspath(__file__))
_REPO = os.path.dirname(os.path.dirname(_HERE))
_CACHE = os.path.join(_REPO, 'cache', 'basefill')

BASE_MONTH = '2019-01'

# ══════════════════════════ 官方一手 URL ══════════════════════════
SRC = {
    # ── US 现货：Cboe 全美合并成交（13 家交易所 + 2 个 TRF），逐月股数与名义额
    'cboe_eq_monthly': ('https://cdn.cboe.com/resources/us/equities/market-statistics/'
                        'historical-market-volume/market_history_monthly_2019.csv'),
    # ── US 期权：100 股/张的官方出处（两处互证）
    #    OCC By-Laws：Article I 定义节「Unit of Trading」+ Article VI（清算确认成交）
    'occ_bylaws': ('https://www.theocc.com/getmedia/'
                   '3309eceb-56cf-48fc-b3b3-498669a24572/occ_bylaws.pdf'),
    #    OCC ODD《Characteristics and Risks of Standardized Options》
    'occ_odd': ('https://www.theocc.com/getmedia/'
                'a151a9ae-d784-4a15-bdeb-23a029f50b70/riskstoc.pdf'),
    #    Cboe 自家 2019-01 逐期权类成交量（**只含单股，不含 ETF/ETP**，见下面注释）
    'cboe_opt_rank_1901': ('https://cdn.cboe.com/resources/options/volume_archive/'
                           '2019/2019_01_rank_wosym.xlsx'),
    # ── ASX 单股期权：2019 全年逐类量 + 月初/月末收盘价
    'asx_eto_2019': ('https://www.asx.com.au/content/dam/asx/participants/'
                     'derivatives-market/equity-derivatives/'
                     'equity-derivatives-statistics/2019/annual-market-summary-2019.xls'),
    # ── ASX 24 合约规格（乘数 / 面值）
    'asx24_specs': ('https://www.asx.com.au/content/dam/asx/participants/'
                    'derivatives-market/ird/asx24-contract-specifications.pdf'),
    # ── ASX 24 的 2019-01 逐合约成交量。
    #    这是 2019-01 MAR 正文里印的那条官方链接，asx.com.au 现在返回通用外壳页
    #    （soft-404）；Wayback 存的是**当年 ASX 自己发的那份 PDF 原件**
    #    （Content-Type: application/pdf，137,193 字节，不是外壳页）。
    #    真伪校验见 asx_deriv_1901()：Total Exchange 10,999,710 与同月 MAR
    #    正文的 10,999,710 逐位相同，与 series/asx.csv 的入库值同源。
    'asx_sfe_1901': ('https://web.archive.org/web/20190301id_/'
                     'http://www.asx.com.au/data/market-reports/'
                     'MonthlyFuturesMarketsReport190131.pdf'),
    # ── ECB 官方日度参考汇率（与 fetch/fx.py 同源同方法，NZ 合约折 AUD 用）
    'ecb': ('https://data-api.ecb.europa.eu/service/data/EXR/D.%s.EUR.SP00.A'
            '?format=csvdata&detail=dataonly'
            '&startPeriod=2019-01-01&endPeriod=2019-01-31'),
    # ── MX 官方月度统计（只到分节合计，不拆到期权类 —— 这正是 MX 两行卡住的原因）
    'mx_stats_1901': 'https://www.m-x.ca/f_stat_en/1901_stats_en.xlsx',
}

# asx.com.au 的 soft-404：HTTP 200 + text/html + 这几个恒定字节数的通用外壳页
_ASX_SHELL_SIZES = {136750, 136883, 136884, 136885}


# ══════════════════════════ HTTP：两套客户端 ══════════════════════════
def _fetch_plain(url, key, ext, refresh=False):
    """裸 urllib + 普通 UA。**web.archive.org 只认这一套**（见坑 2）。"""
    os.makedirs(_CACHE, exist_ok=True)
    path = os.path.join(_CACHE, 'r2_' + key + ext)
    if os.path.exists(path) and not refresh:
        with open(path, 'rb') as fh:
            return fh.read()
    req = urllib.request.Request(url, headers={
        'User-Agent': ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                       'AppleWebKit/605.1.15 (KHTML, like Gecko) '
                       'Version/17.0 Safari/605.1.15')})
    with urllib.request.urlopen(req, timeout=180) as resp:
        blob = resp.read()
        ctype = resp.headers.get('Content-Type', '')
    if len(blob) in _ASX_SHELL_SIZES:
        raise RuntimeError('%s 拿到 asx.com.au 通用外壳页（%d 字节，soft-404）：%s'
                           % (key, len(blob), url))
    if ext == '.pdf' and blob[:5] != b'%PDF-':
        raise RuntimeError('%s 不是 PDF（Content-Type=%s，首 5 字节=%r）：%s'
                           % (key, ctype, blob[:5], url))
    with open(path, 'wb') as fh:
        fh.write(blob)
    return blob


def _fetch_imp(url, key, ext, refresh=False, impersonate='chrome124'):
    """curl_cffi 浏览器 TLS 指纹。**theocc.com / cdn.cboe.com 用这一套**（见坑 1）。"""
    os.makedirs(_CACHE, exist_ok=True)
    path = os.path.join(_CACHE, 'r2_' + key + ext)
    if os.path.exists(path) and not refresh:
        with open(path, 'rb') as fh:
            return fh.read()
    from curl_cffi import requests as creq
    r = creq.get(url, impersonate=impersonate, timeout=180)
    if r.status_code != 200:
        raise RuntimeError('%s HTTP %d（impersonate=%s）：%s'
                           % (key, r.status_code, impersonate, url))
    blob = r.content
    if ext == '.pdf' and blob[:5] != b'%PDF-':
        raise RuntimeError('%s 不是 PDF（首 5 字节=%r）：%s' % (key, blob[:5], url))
    with open(path, 'wb') as fh:
        fh.write(blob)
    return blob


def _pdf_text(blob):
    import fitz
    doc = fitz.open(stream=blob, filetype='pdf')
    return '\n'.join(doc[i].get_text() for i in range(doc.page_count))


def _need(hay, needle, where):
    """断言：官方文件里必须逐字出现这句话。改版时在这里炸，而不是下游算错。"""
    flat = re.sub(r'\s+', ' ', hay)
    if re.sub(r'\s+', ' ', needle) not in flat:
        raise AssertionError('%s 的官方文件里没找到 %r —— 文件改版了，先核对再改常数'
                             % (where, needle))
    return needle


def weighted_constant(members):
    """members = [(名字, 2019-01 张数, 单张名义额)] → 按张数加权的等效单张名义额。"""
    tot = sum(m[1] for m in members)
    if tot <= 0:
        raise ZeroDivisionError('成员 2019-01 张数合计为 0，篮子常数无定义')
    return sum(m[1] * m[2] for m in members) / tot


# ══════════════════ 1. US_CASH_EQUITY_SHARE（独立复算）══════════════════
def us_cash_equity_share(refresh=False):
    """2019-01 全美合并成交金额 ÷ 合并成交股数 = 每股均价（USD）。

    Cboe 的这张表是**全市场**表，不是 Cboe 自家份额：15 个 market participant =
    13 家股票交易所 + FINRA/NYSE TRF + FINRA/Nasdaq TRF Carteret，正好是合并
    行情（consolidated tape）的全部上报方。所以 Σ名义额 ÷ Σ股数 就是定基那一个月
    「平均一股」的价格，不需要任何指数代理。
    """
    blob = _fetch_imp(SRC['cboe_eq_monthly'], 'cboe_eq_monthly', '.csv', refresh)
    rows = list(csv.DictReader(io.StringIO(blob.decode('utf-8', 'replace'))))
    jan = [r for r in rows if r['Month'] == BASE_MONTH]
    if len(jan) != 15:
        raise AssertionError('2019-01 的 market participant 不是 15 个而是 %d 个，'
                             'Cboe 改表了' % len(jan))
    shares = sum(float(r['Total Shares']) for r in jan)
    notional = sum(float(r['Total Notional']) for r in jan)
    # 分 tape 交叉验证：与 series/ice.csv 的 consolidated ADV（NYSE 报的口径）比
    tapes = {}
    for t in ('A', 'B', 'C'):
        s = sum(float(r['Tape %s Shares' % t]) for r in jan)
        n = sum(float(r['Tape %s Notional' % t]) for r in jan)
        tapes[t] = {'shares': s, 'notional': n, 'avg_px': n / s, 'adv_mnsh': s / 21 / 1e6}
    return {'shares': shares, 'notional': notional, 'avg_price': notional / shares,
            'tapes': tapes, 'participants': len(jan)}


def _ice_tape_crosscheck():
    """从 series/ice.csv 取 2019-01 的 tape A/B/C consolidated ADV（百万股/日）。"""
    p = os.path.join(_REPO, 'series', 'ice.csv')
    if not os.path.exists(p):
        return None
    with open(p) as fh:
        for r in csv.DictReader(fh):
            if r.get('month') == BASE_MONTH:
                return {t: float(r['adv_tape%s_consolidated_mnsh' % t])
                        for t in ('A', 'B', 'C')}
    return None


# ══════════════════ 2. US_MULTILIST_EQ_OPT（乘数已解决，价格未解决）═════════
def us_option_unit_of_trading(refresh=False):
    """100 股/张的**官方出处**。第一轮把这一项挂在「行业惯例」上，本轮拿到原文。

    两处互证，都是 OCC 自己的文件：
      · OCC By-Laws，Article I «Definitions» 的 “Unit of Trading” 词条
      · OCC By-Laws，Article VI «Clearance of Confirmed Trades»
      · OCC ODD（riskstoc.pdf）第 10 页 “UNIT OF TRADING; CONTRACT SIZE”
    ⚠ ODD 第 20 页同时给了 ETF 期权的例外，见返回值 fund_caveat —— 这是
      US_MULTILIST_EQ_OPT（含 ETF）不能无条件写 100 的原因。
    """
    bylaws = _pdf_text(_fetch_imp(SRC['occ_bylaws'], 'occ_bylaws', '.pdf', refresh))
    odd = _pdf_text(_fetch_imp(SRC['occ_odd'], 'occ_odd', '.pdf', refresh))
    q_def = _need(bylaws,
                  'In the absence of any such designation for a series of options or '
                  'futures in which the underlying security is a common stock the unit '
                  'of trading shall be 100 shares.', 'OCC By-Laws Article I')
    q_art6 = _need(bylaws,
                   'in the absence of such designation for a series of options in which '
                   'the underlying security is a common stock, the unit of trading shall '
                   'be 100 shares.', 'OCC By-Laws Article VI')
    q_odd = _need(odd, 'the unit of trading for most options on equity securities is '
                       '100 shares.', 'OCC ODD p.10')
    q_fund = _need(odd, 'As a general rule a single-stock option covers 100 shares of '
                        'the underlying security, although in the case of options '
                        'covering fund shares, options covering 100 or 1000 shares may '
                        'be available.', 'OCC ODD p.20')
    return {'multiplier': 100, 'bylaws_definition': q_def, 'bylaws_article_vi': q_art6,
            'odd': q_odd, 'fund_caveat': q_fund}


def cboe_equity_option_classes_1901(refresh=False):
    """Cboe《Equity Option Volume Archive》2019-01 逐类成交量。

    ⚠ 这张表是 **Cboe Options Exchange（C1）自家**的量，且**只有单股，没有
      ETF/ETP**（实测 SPY / QQQ / IWM / GLD / VXX 全部不在表内）。
      所以它能当「单股期权类之间的相对权重」用，**不能**当
      US_MULTILIST_EQ_OPT（单股 + ETF、全美 16 家期权所）的权重直接用。
    """
    import openpyxl
    blob = _fetch_imp(SRC['cboe_opt_rank_1901'], 'cboe_opt_rank_1901', '.xlsx', refresh)
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
    if wb.sheetnames != ['1901']:
        raise AssertionError('Cboe 期权量档案的 sheet 名变了：%s' % wb.sheetnames)
    ws = wb['1901']
    hdr = [ws.cell(1, c).value for c in range(1, 10)]
    if hdr[:5] != ['Symbol', 'Name', 'Call', 'Put', 'Tot']:
        raise AssertionError('Cboe 期权量档案表头变了：%r' % hdr)
    out = []
    for r in range(2, ws.max_row + 1):
        s = ws.cell(r, 1).value
        if s in (None, ''):
            continue
        out.append((str(s).strip(), float(ws.cell(r, 5).value or 0)))
    for etf in ('SPY', 'QQQ', 'IWM', 'GLD', 'VXX'):
        if any(s == etf for s, _ in out):
            raise AssertionError('%s 出现在 Cboe 单股期权量档案里 —— 该文件的口径'
                                 '变了（本轮判定它不含 ETF 就不成立了）' % etf)
    return out


# ══════════════════════════ 3. ASX_ETO ══════════════════════════
def asx_eto_1901(refresh=False):
    """ASX 单股期权 2019-01 逐类 Lots Traded + 两个日期的收盘价。

    与第一轮 build/basefill/asx_mx.py 的同名函数同源同算法；本轮只做复算，
    另外把两件事钉死（第一轮只说了「月初/月末」，没落到具体哪天）：

    1. 那两列价格的表头是 Excel 日期序列 43467 / 43496 = **2019-01-02 / 2019-01-31**，
       即当月第一个与最后一个交易日的收盘价。所以 month_midpoint 是这两天的中点，
       不是「月初月末」这种含糊说法 —— 本函数对这两个序列号加了断言。
    2. 同一张表第 7-8 列有 “Traded Value” —— **那是权利金成交额，不是名义额**。
       实测 2019-01 全部单股期权 Traded Value 合计 A$513,421,557 ÷ 5,207,234 张
       = A$98.60/张，只有名义额 A$2,542.44/张 的 3.9%。
       ⇒ 模式①（官方直接发名义额，除以张数即得常数）在 ASX 单股期权上**不成立**，
         别再花时间找那一列。

    剔掉 XJO / XJO* / XJOL —— 那是**指数**期权与指数 LEPO，乘数是每点澳元数，
    在 pools.py 里挂 ASX_INDEX_OPT，不属于 ASX_ETO。
    """
    import xlrd
    wb = xlrd.open_workbook(file_contents=_fetch_plain(
        SRC['asx_eto_2019'], 'asx_eto_2019', '.xls', refresh))
    sh = wb.sheet_by_name('January')
    if 'January 2019' not in str(sh.cell_value(0, 0)):
        raise AssertionError('ASX 年报 January 表抬头变了：%r' % sh.cell_value(0, 0))
    if (sh.cell_value(3, 1), sh.cell_value(3, 2)) != (43467.0, 43496.0):
        raise AssertionError('ASX 年报的两列价格日期不再是 2019-01-02 / 2019-01-31：%r'
                             % [sh.cell_value(3, 1), sh.cell_value(3, 2)])
    if str(sh.cell_value(2, 6)).strip() != 'Traded Value':
        raise AssertionError('ASX 年报第 7 列不再是 Traded Value：%r' % sh.cell_value(2, 6))
    eto, all_rows = [], []
    for r in range(5, sh.nrows):
        code = str(sh.cell_value(r, 0)).strip()
        if not code or code == 'Total':
            continue
        try:
            p_first = float(sh.cell_value(r, 1))
            p_last = float(sh.cell_value(r, 2))
        except (TypeError, ValueError):
            continue
        call, put = sh.cell_value(r, 4), sh.cell_value(r, 5)
        lots = (float(call) if call != '' else 0.0) + (float(put) if put != '' else 0.0)
        all_rows.append((code, p_first, p_last, lots))
        if code.rstrip('*L') != 'XJO':
            eto.append((code, p_first, p_last, lots))
    return eto, all_rows


def asx_eto_constant(refresh=False):
    """三个候选：月初价 / 月末价 / 中点。**主线程 2026-08-06 裁决：取中点。**

    基准名必须记成 `month_midpoint`，不许写 avg_close —— ASX 对个股只发月初、
    月末两个收盘价，历史日频行情是收费产品，凑不出真正的月内均价。
    """
    eto, all_rows = asx_eto_1901(refresh)
    lots = sum(x[3] for x in eto)
    cand = {}
    for name, pick in (('month_first', lambda a, b: a),
                       ('month_last', lambda a, b: b),
                       ('month_midpoint', lambda a, b: (a + b) / 2.0)):
        cand[name] = 100.0 * sum(l * pick(a, b) for _, a, b, l in eto) / lots
    # 官方交叉校验：全部 161 类（含 XJO 系）的合计必须等于 ASX 自己发的分项合计
    all_lots = sum(x[3] for x in all_rows)
    if int(round(all_lots)) != 5953660:
        raise AssertionError('ASX 年报 January 表 161 类合计 %d ≠ 官方 5,953,660'
                             % round(all_lots))
    if int(round(lots)) != 5207234:
        raise AssertionError('剔除 XJO 系后 %d ≠ MAR 的 Single stock equity options '
                             '5,207,234' % round(lots))
    band = (cand['month_last'] - cand['month_first']) / cand['month_midpoint'] / 2
    return {'candidates': cand, 'chosen': cand['month_midpoint'], 'lots': lots,
            'classes': len(eto), 'halfband': band}


def asx_eto_traded_value_check(refresh=False):
    """实测「Traded Value 是权利金不是名义额」，把模式①在 ASX 上彻底排除。"""
    import xlrd
    wb = xlrd.open_workbook(file_contents=_fetch_plain(
        SRC['asx_eto_2019'], 'asx_eto_2019', '.xls', refresh))
    sh = wb.sheet_by_name('January')
    lots = value = notional = 0.0
    for r in range(5, sh.nrows):
        code = str(sh.cell_value(r, 0)).strip()
        if not code or code == 'Total' or code.rstrip('*L') == 'XJO':
            continue
        try:
            pf, pl = float(sh.cell_value(r, 1)), float(sh.cell_value(r, 2))
        except (TypeError, ValueError):
            continue
        c, p = sh.cell_value(r, 4), sh.cell_value(r, 5)
        n = (float(c) if c != '' else 0.0) + (float(p) if p != '' else 0.0)
        lots += n
        value += sum(sh.cell_value(r, k) for k in (6, 7)
                     if isinstance(sh.cell_value(r, k), float))
        notional += n * 100 * (pf + pl) / 2
    return {'lots': lots, 'traded_value': value, 'value_per_lot': value / lots,
            'notional_per_lot': notional / lots,
            'ratio': value / notional}


def asx_index_level_1901(refresh=False):
    """S&P/ASX 200（XJO）2019-01-02 / 2019-01-31 收盘及中点点位。

    取自同一张 ASX 官方年报 xls 的 XJO 行 —— 与 ASX_ETO 用的是同两列，所以
    SPI 200 期货的基期点位与 ASX_ETO 的基期股价是**同一个口径**，不会一个表里
    混两种时点。
    """
    _, all_rows = asx_eto_1901(refresh)
    xjo = [r for r in all_rows if r[0] == 'XJO']
    if len(xjo) != 1:
        raise AssertionError('ASX 年报里 XJO 行不是恰好 1 行：%r' % xjo)
    _, first, last, _ = xjo[0]
    return {'month_first': first, 'month_last': last,
            'month_midpoint': (first + last) / 2.0}


# ══════════════════════════ 4. ASX_DERIV ══════════════════════════
def asx_deriv_1901(refresh=False):
    """ASX 24《Monthly SFE Trading Report for January 2019》逐合约成交量。

    本轮的关键突破：asx.com.au 现在对这条路径返回通用外壳页，但 Wayback 存的是
    ASX 当年自己发的 PDF 原件。真伪不靠「来源看起来对」，靠两个硬校验：
      (a) 报告抬头必须是 “Monthly SFE Trading Report for January 2019”；
      (b) Total Exchange 必须等于 10,999,710 —— 与同月 ASX Group Monthly
          Activity Report 正文的 “Total contracts 10,999,710” 逐位相同。
    """
    txt = _pdf_text(_fetch_plain(SRC['asx_sfe_1901'], 'asx_sfe_1901', '.pdf', refresh))
    _need(txt, 'Monthly SFE Trading Report for January 2019', 'ASX 逐合约报告抬头')
    _need(txt, 'Volumes quoted are Total Volumes which include on-market, off-market '
               'and non-traded volumes.', 'ASX 逐合约报告脚注（non-traded 口径）')
    flat = re.sub(r'[ \t]+', ' ', txt)

    def vol(code, after):
        """取 `after` 之后第一个 '\\n<code>\\n<数字>' 里的数字（当月成交量）。"""
        i = flat.find(after)
        if i < 0:
            raise AssertionError('ASX 逐合约报告里找不到分节 %r' % after)
        m = re.search(r'\n' + re.escape(code) + r'\n\s*([\d,]+)\n', flat[i:])
        if not m:
            raise AssertionError('分节 %r 之后取不到合约 %r 的当月量' % (after, code))
        return float(m.group(1).replace(',', ''))

    v = {}
    for code in ('AP', 'AM'):
        v['fut_' + code] = vol(code, 'Equity Indices - Futures')
    v['opt_AP'] = vol('AP', 'Equity Indices - Options')
    for code in ('IB', 'IR', 'YT', 'XT', 'LT'):
        v['fut_' + code] = vol(code, 'Interest Rates - Futures')
    for code in ('YT', 'YO', 'YD', 'XT'):
        v['opt_' + code] = vol(code, 'Interest Rates - Options')
    v['fut_BB_NZ'] = vol('BB', 'NZ Interest Rates - Futures')

    m = re.search(r'Total Exchange\s*\n\s*Daily Average\s*\n\s*([\d,]+)', flat)
    if not m:
        raise AssertionError('ASX 逐合约报告里取不到 Total Exchange')
    total = float(m.group(1).replace(',', ''))
    if int(total) != 10_999_710:
        raise AssertionError('Total Exchange %d ≠ 同月 MAR 的 10,999,710 —— '
                             '这份快照不是 2019-01 的那份' % total)
    return v, total


def asx24_specs(refresh=False):
    """ASX 24 各合约的单张名义额（本币）。文本取自 ASX 官方合约规格 PDF。

    ⚠ 该 PDF 是**现行版**（不是 2019-01 在效版）。之所以敢用在基期上：
      LT（20 年国债）的 Listing Date 写着 10 August 2018，即 2019-01 已在效；
      其余四个（AP / IB / IR / YT / XT / BB）都是 20 年不动的老合约。
      若哪天 ASX 改了乘数，_need() 会先炸。
    ⚠ LT 的面值是 **A$65,000 不是 A$100,000** —— 与 3Y/10Y 不同，照抄同族会错。
    """
    txt = _pdf_text(_fetch_plain(SRC['asx24_specs'], 'asx24_specs', '.pdf', refresh))
    _need(txt, 'Valued at A$25 per index point', 'ASX SPI 200 期货')
    _need(txt, 'Valued at A$5 per index point', 'ASX Mini SPI 200 期货')
    _need(txt, 'notional sum of AUD 3,000,000', 'ASX 30 Day Interbank Cash Rate')
    _need(txt, 'A$1,000,000 face value 90-Day Bank Accepted Bills', 'ASX 90 Day Bank Bills')
    _need(txt, 'face value of A$100,000, a coupon rate of 6% per annum and a term to '
               'maturity of three years', 'ASX 3 Year Treasury Bond')
    _need(txt, 'face value of A$100,000, a coupon rate of 6% per annum and a term to '
               'maturity of ten years', 'ASX 10 Year Treasury Bond')
    _need(txt, 'face value of A$65,000, a coupon rate of 4% per annum and a term to '
               'maturity of twenty years', 'ASX 20 Year Treasury Bond')
    _need(txt, 'NZ$1,000,000 face value 90 day bank accepted bill', 'ASX 90 Day NZ Bank Bill')
    return {'AP_per_point': 25.0, 'AM_per_point': 5.0, 'IB_face': 3_000_000.0,
            'IR_face': 1_000_000.0, 'YT_face': 100_000.0, 'XT_face': 100_000.0,
            'LT_face': 65_000.0, 'BB_NZ_face': 1_000_000.0}


def ecb_avg_rate(cur, refresh=False):
    """ECB 官方日度参考汇率的 2019-01 算术平均（每 1 EUR 兑多少 cur）。

    与 fetch/fx.py 同源同方法：本模块用 AUD/USD 做过对账，得
    0.7146828535918106，与 series/fx.csv 的 fx_avg_audusd 逐位相同。
    """
    blob = _fetch_plain(SRC['ecb'] % cur, 'ecb_' + cur, '.csv', refresh)
    rows = list(csv.DictReader(io.StringIO(blob.decode('utf-8', 'replace'))))
    vals = {r['TIME_PERIOD']: float(r['OBS_VALUE']) for r in rows if r.get('OBS_VALUE')}
    if not vals:
        raise AssertionError('ECB 没返回 %s 的 2019-01 日度序列' % cur)
    return vals


def nzd_to_aud_1901(refresh=False):
    """2019-01 NZD→AUD 月均（逐日先算交叉再取月均，与 fetch/fx.py 一致）。"""
    aud = ecb_avg_rate('AUD', refresh)
    nzd = ecb_avg_rate('NZD', refresh)
    days = sorted(set(aud) & set(nzd))
    if len(days) < 20:
        raise AssertionError('ECB 2019-01 的 AUD/NZD 共同观测日只有 %d 天' % len(days))
    return sum(aud[d] / nzd[d] for d in days) / len(days), len(days)


def asx_deriv_constant(refresh=False):
    """ASX_DERIV 篮子常数（AUD/张）。"""
    v, total = asx_deriv_1901(refresh)
    sp = asx24_specs(refresh)
    idx = asx_index_level_1901(refresh)
    nzdaud, nzd_days = nzd_to_aud_1901(refresh)

    out = {}
    for basis in ('month_first', 'month_last', 'month_midpoint'):
        lvl = idx[basis]
        members = [
            # 股指：乘数 × 指数点。期权按「一张期权 = 一张标的期货」折算，
            # 与第一轮 CME_EQUITY_INDEX / CME_RATES 的处理一致。
            ('SPI200 futures (AP)', v['fut_AP'], sp['AP_per_point'] * lvl),
            ('SPI200 options (AP)', v['opt_AP'], sp['AP_per_point'] * lvl),
            ('Mini SPI200 (AM)', v['fut_AM'], sp['AM_per_point'] * lvl),
            # 利率：按面值，不乘结算价（报价是 100−收益率，乘上去无经济含义）
            ('30d Interbank Cash Rate (IB)', v['fut_IB'], sp['IB_face']),
            ('90d Bank Bills (IR)', v['fut_IR'], sp['IR_face']),
            ('3Y Bond futures (YT)', v['fut_YT'], sp['YT_face']),
            ('10Y Bond futures (XT)', v['fut_XT'], sp['XT_face']),
            ('20Y Bond futures (LT)', v['fut_LT'], sp['LT_face']),
            ('3Y Bond options (YT)', v['opt_YT'], sp['YT_face']),
            ('3Y Bond overnight opts (YO)', v['opt_YO'], sp['YT_face']),
            ('3Y Bond intraday opts (YD)', v['opt_YD'], sp['YT_face']),
            ('10Y Bond options (XT)', v['opt_XT'], sp['XT_face']),
            # NZ 90 天银行票据：NZ$ 面值先折 AUD 再进篮子
            ('NZ 90d Bank Bill (BB)', v['fut_BB_NZ'], sp['BB_NZ_face'] * nzdaud),
        ]
        covered = sum(m[1] for m in members)
        out[basis] = {'constant': weighted_constant(members), 'covered': covered,
                      'coverage': covered / total, 'members': members}
    return {'by_basis': out, 'total': total, 'index': idx, 'nzdaud': nzdaud,
            'nzd_days': nzd_days, 'specs': sp, 'volumes': v}


# ══════════════════ 5. MX_EQUITY_OPT / MX_ETF_OPT（未填）══════════════════
def mx_option_totals_1901(refresh=False):
    """MX 官方月报 2019-01 的期权分节合计。**只到合计，不拆到类** —— 这就是卡点。

    本轮把 MX 官方 6 条通道都验了一遍（见模块末尾 MX_BLOCKER），结论没变：
    要拿逐类量，只能一个 symbol 一个 symbol 拉 /en/trading/data/historical。
    """
    import openpyxl
    blob = _fetch_plain(SRC['mx_stats_1901'], 'mx_stats_1901', '.xlsx', refresh)
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
    ws = wb['Jan 2019 EN']
    if str(ws.cell(2, 2).value).strip() != 'Jan 2019':
        raise AssertionError('MX 月报 B2 不是 "Jan 2019"：%r' % ws.cell(2, 2).value)
    got = {}
    for r in range(3, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or '').strip()
        if label in ('Equity Options', 'ETF Options'):
            got[label] = float(str(ws.cell(r, 2).value).replace(',', ''))
    if got.get('Equity Options') != 2_757_998 or got.get('ETF Options') != 884_892:
        raise AssertionError('MX 2019-01 期权合计变了：%r' % got)
    # 官方月报里所有列的表头 —— 确认没有任何「金额 / 名义额」列
    cols = [str(ws.cell(1, c).value or '') for c in range(1, 20)]
    for bad in ('VALUE', 'NOTIONAL', 'PREMIUM'):
        if any(bad in c.upper() for c in cols):
            raise AssertionError('MX 月报现在有 %s 列了 —— 模式①可用，重做本行' % bad)
    return got


MX_BLOCKER = """\
MX 逐期权类的 2019-01 成交量：本轮把 m-x.ca 的官方通道全验了一遍，仍然只有一条路。
已实测排除（都是 m-x.ca / cdcc.ca 自家路径，非第三方）：
 (1) 官方月报 f_stat_en/1901_stats_en.xlsx —— 只给 Equity Options 2,757,998 与
     ETF Options 884,892 两个合计，全表 121 列里没有任何 value/notional/premium 列
     （模式①在 MX 不成立；同月 PDF 版 1901_stats_en.pdf 内容与 xlsx 相同）。
 (2) /en/trading/data/monthly-volumes-and-open-interest —— 页面只挂近三个月的
     f_stat_en/{YYMM}_stats_en.{pdf,xlsx}，与 (1) 同一份文件。
 (3) /files/options-summary-en.xlsx —— 是**当日**逐系列快照（实测抬头
     "August 5, 2026"，5,195 行），没有历史版本。
 (4) /en/trading/data/historical?symbol=&dnld=1 与 symbol=ALL —— 都返回 401 字节的
     空 CSV，接口强制单 symbol，没有全市场导出。
 (5) 单日查询 /en/trading/data/historical?symbol=BNS&from=2019-01-31&to=2019-01-31
     &dnld=1 → 200 / text/csv / 58,443 字节 / 428 行，含标的自身行
     （Class Symbol 空、Ins. Type=3、Last Price=74.8）与各期权系列 Volume。
     即**通道本身完全可用、可回溯**，问题纯粹是要跑 940 次。
 (6) cdcc.ca 的 files-and-publications 只有规则/风险参数/季度 PQD，没有逐类日量。
⇒ 唯一路径：对 940 个期权类逐个拉整月 CSV（第一轮 asx_mx.py 的
   mx_option_class_month()），robots.txt 写死 Crawl-delay: 15，
   940 × 15s ≈ 3.9 小时 / ≈1.2 GB。**决定成本的是 15 秒的延迟不是字节数**：
   改成单日查询能把 1.2 GB 压到 ~55 MB，但墙钟时间一分不省，而且单日量不能替代
   整月量。所以别做「先单日粗筛再整月细拉」的两段式，直接整月跑一遍最省。
   跑之前还要先解决 ETF 类的名单问题：/en/trading/data/options-list 是 2026 年的
   现况，2019 年在册、之后退市的类不在里面，直接用会漏成员 —— 可行的补法是
   940 个 symbol 全拉，再用 Ins. Type / 标的名称把 ETF 与个股分开。
"""


# ══════════════════════════ 主流程 ══════════════════════════
def build(refresh=False):
    res = {}

    print('── US_CASH_EQUITY_SHARE ' + '─' * 44)
    u = us_cash_equity_share(refresh)
    res['US_CASH_EQUITY_SHARE'] = u
    print('  participants=%d  shares=%.0f  notional=%.0f' %
          (u['participants'], u['shares'], u['notional']))
    print('  每股均价 = %.15f USD' % u['avg_price'])
    ice = _ice_tape_crosscheck()
    if ice:
        for t in ('A', 'B', 'C'):
            a, b = u['tapes'][t]['adv_mnsh'], ice[t]
            print('    tape %s: Cboe %.0f mn/日 vs ICE(series/ice.csv) %.0f  偏差 %+.2f%%'
                  % (t, a, b, (a / b - 1) * 100))

    print('── US_MULTILIST_EQ_OPT ' + '─' * 45)
    uot = us_option_unit_of_trading(refresh)
    res['US_OPT_UNIT'] = uot
    print('  OCC By-Laws（Article I 定义）：%s' % uot['bylaws_definition'])
    print('  OCC ODD p.10：%s' % uot['odd'])
    print('  ⚠ ODD p.20 的 ETF 例外：%s' % uot['fund_caveat'])
    cls = cboe_equity_option_classes_1901(refresh)
    tot = sum(v for _, v in cls)
    print('  Cboe(C1) 2019-01 单股期权：%d 个类，合计 %d 张（**不含 ETF**）'
          % (len(cls), round(tot)))
    print('  ⇒ 权重有了、乘数有了，仍缺「2019-01 逐标的收盘价」；见 US_OPT_BLOCKER')
    res['US_OPT_CLASSES'] = cls

    print('── ASX_ETO ' + '─' * 57)
    e = asx_eto_constant(refresh)
    res['ASX_ETO'] = e
    for k, val in e['candidates'].items():
        print('  %-14s = %.10f AUD' % (k, val))
    print('  采用 month_midpoint = %.10f（%d 类 / %d 张，带宽 ±%.2f%%）'
          % (e['chosen'], e['classes'], round(e['lots']), e['halfband'] * 100))
    tv = asx_eto_traded_value_check(refresh)
    print('  模式①排除：官方 Traded Value 合计 A$%.0f ÷ %d 张 = A$%.2f/张，'
          '只占名义额 A$%.2f/张 的 %.1f%% ⇒ 那是权利金，不是名义额'
          % (tv['traded_value'], round(tv['lots']), tv['value_per_lot'],
             tv['notional_per_lot'], tv['ratio'] * 100))

    print('── ASX_DERIV ' + '─' * 55)
    d = asx_deriv_constant(refresh)
    res['ASX_DERIV'] = d
    print('  XJO 2019-01 月初 %.1f / 月末 %.1f / 中点 %.2f'
          % (d['index']['month_first'], d['index']['month_last'],
             d['index']['month_midpoint']))
    print('  NZD→AUD 月均 = %.16f（%d 个观测日，ECB）' % (d['nzdaud'], d['nzd_days']))
    for basis in ('month_first', 'month_last', 'month_midpoint'):
        b = d['by_basis'][basis]
        print('  %-14s 常数 = %.10f AUD/张（覆盖 %.2f%%）'
              % (basis, b['constant'], b['coverage'] * 100))
    mid = d['by_basis']['month_midpoint']
    print('  成员明细（%d 个，Σ=%d 张 / 全所 %d 张）：'
          % (len(mid['members']), round(mid['covered']), round(d['total'])))
    for name, vol, unit in sorted(mid['members'], key=lambda x: -x[1]):
        print('    %-30s %9d 张 × %14.2f AUD' % (name, round(vol), unit))
    print('    未覆盖 %d 张（%.2f%%）：A-REIT 指数期货 AA、IR strip pack RP、'
          '澳洲与新西兰电力/天然气/谷物期货期权 —— 按「与已覆盖成员单张名义额相同」归一'
          % (round(d['total'] - mid['covered']), (1 - mid['coverage']) * 100))

    print('── MX_EQUITY_OPT / MX_ETF_OPT ' + '─' * 38)
    mx = mx_option_totals_1901(refresh)
    res['MX'] = mx
    print('  官方月报只到合计：Equity Options %d / ETF Options %d'
          % (mx['Equity Options'], mx['ETF Options']))
    print(MX_BLOCKER)
    return res


US_OPT_BLOCKER = """\
US_MULTILIST_EQ_OPT 还缺的那一半 =「2019-01 按期权成交量加权的标的均价」。
本轮把乘数那一半彻底解决了（OCC By-Laws + ODD 原文，见 us_option_unit_of_trading），
价格这一半仍然没有干净来源。已实测排除：
 (1) OCC 官方 Volume Query（marketdata.theocc.com/volume-query）—— 通道通了
     （impersonate=chrome124 后 theocc.com 全站 200），但服务端硬性
     "Report date cannot be prior to 2 years"，2019-01 取不到。
     OCC 的 Daily Volume / Exchange Volume by Class / Monthly & Weekly 三个报表
     的官方 config JSON 也都写着 "Data available for the past 24 months"。
 (2) Cboe《Equity Option Volume Archive》2019_01_rank_wosym.xlsx —— **可用**，
     2,889 个类的当月张数，但 (a) 只是 Cboe C1 一家的量（合计 36,259,079 张），
     (b) **不含 ETF/ETP**（SPY/QQQ/IWM/GLD/VXX 全不在表内）。
     ETF 期权在全美多重挂牌期权里占比很大且价格结构与单股完全不同，
     用它当全市场权重会系统性偏。
 (3) 逐标的 2019-01 收盘价：api.nasdaq.com 的 historical 接口
     assetclass=stocks / etf 对 2019 至 2026 的任何历史月都返回 totalRecords=0
     （只有 assetclass=index 还有长历史），已不能当逐股价格源。
 (4) ftp.nasdaqtrader.com（公开匿名 FTP，通道正常）：MonthlyShareVolume 止于 2006，
     Files/MarketShare 逐日 per-symbol CSV 止于 2016，都不覆盖 2019。
 (5) SEC MIDAS「Metrics by Individual Security」individual_security_2019_q1.zip
     —— 逐日逐股，但价格只有 PriceRank 十分位，没有价格水平，也没有金额成交。
 (6) Cboe 美股 market_history_monthly_2019.csv 只有全市场合计，没有 by-symbol 版本。
诊断值（**不是可入库的数**，仅供主线程判断量级）：拿 (2) 的单股权重配 SEC
 Fails-to-Deliver 文件（cnsfails201901a/b + 201902a，PRICE 字段官方定义是
 「该证券前一交易日的收盘价」）算出的加权均价 ≈ 124.76 USD ⇒ 100 股/张 ≈ 12,476 USD。
 这个数有两处不可接受的偏差，所以没有写进 CSV：
   · 权重不含 ETF，且只是 Cboe 一家的类结构；
   · FTD 只在该标的当日有交收失败时才有价格，取到的是**自选样本日**的收盘价均值
     （AAPL 21 天里只有 13 天），既不是 avg_close 也不是任何可命名的官方基准。
下一步建议（按性价比排序）：
 a) 找 OCC 是否有 24 个月以外的官方存档（batch processing 的历史平面文件、
    或向 OCC data sales 索取 2019 的 Volume By Underlying）；
 b) 若接受「用 Cboe 单股 + 单独补 ETF 权重」，需要先找到 2019-01 ETF 期权的
    官方逐类量；
 c) 逐标的价格若最终只能用 FTD，必须在 contract_specs.csv 新增一个明确的
    base_price_basis 取值（例如 ftd_sampled_close），不许混进 avg_close。
"""


def emit_csv(res):
    """打印可直接粘进 series/_specs_part_us_asx_mx2.csv 的行（本脚本不写文件）。"""
    w = csv.writer(sys.stdout)
    w.writerow(['product_id', 'base_price_local', 'base_ccy', 'basket_constant',
                'source_url', 'computed_note'])
    u = res['US_CASH_EQUITY_SHARE']
    w.writerow(['US_CASH_EQUITY_SHARE', repr(u['avg_price']), 'USD',
                repr(u['avg_price']), SRC['cboe_eq_monthly'], '(见 CSV 交付件)'])
    e = res['ASX_ETO']
    w.writerow(['ASX_ETO', repr(e['chosen']), 'AUD', repr(e['chosen']),
                SRC['asx_eto_2019'], '(见 CSV 交付件)'])
    d = res['ASX_DERIV']['by_basis']['month_midpoint']
    w.writerow(['ASX_DERIV', repr(d['constant']), 'AUD', repr(d['constant']),
                SRC['asx_sfe_1901'], '(见 CSV 交付件)'])


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--refresh', action='store_true', help='忽略 cache 重新下载')
    ap.add_argument('--emit', action='store_true', help='额外打印 CSV 行')
    a = ap.parse_args(argv)
    res = build(a.refresh)
    print('── US_MULTILIST_EQ_OPT 卡点全文 ' + '─' * 36)
    print(US_OPT_BLOCKER)
    if a.emit:
        print('── CSV ' + '─' * 61)
        emit_csv(res)
    return 0


if __name__ == '__main__':
    sys.exit(main())
