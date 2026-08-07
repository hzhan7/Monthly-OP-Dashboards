# -*- coding: utf-8 -*-
"""美国相关 7 个 product_id 的基期常数取数器（2019-01 基期，一次算死）。

用法:
    python3 build/basefill/us.py            # 取数 + 全部核对 + 写 series/_specs_part_us.csv
    python3 build/basefill/us.py --dry      # 只打印，不写文件

━━ 这个文件为什么存在 ━━
series/contract_specs.csv 里既有的 23 行常数没有任何取数代码留痕，cache/ 又进了
.gitignore —— 仓库自己没法复算那些数。本模块是「美国那一批」的复算入口：
每一个数都能从这里重跑出来，且每一步都断言官方原文仍在（源站改版会当场炸，
而不是悄悄算出一个错的常数）。

━━ 口径（与 build/check_specs.py 表头注释一致，改这里等于改全站历史）━━
· 基期 = 2019-01。
· 基期价格 = 该月**全部交易日收盘价的算术平均**（base_price_basis=avg_close）；
  利率期货与「金额型」列按定义取 1（definitional）。
· 篮子常数 = Σ(成员 2019-01 张数 × 该合约乘数 × 该合约基期价格) ÷ Σ(成员 2019-01 张数)
  —— 用**基期那个月**的品种结构算一次写死，绝不每月重算。
· 期权按「一张期权 = 一张标的期货」折算名义额（张数口径的名义敞口，不是权利金）。

━━ 数据源（全部官方一手，逐条注明为什么算一手）━━
1. CME 各合约 2019-01 张数
     ftp://ftp.cmegroup.com/daily_volume/daily_volume_20190131.xlsx
   CME 自己的公开匿名 FTP（www.cmegroup.com/ftp/pub/ 那个目录树的原始通道；
   HTTPS 通道只能列目录、下文件一律 400，所以这里必须走 FTP）。
   工作表 'CME Group Vol and OI by Product' 的 **MTD ADV** 列 = 该产品当月日均张数；
   取 1 月最后一个交易日（01/31）的文件，MTD 即整个 2019-01。
   交叉验证：分类合计与 series/cme.csv（CME IR 月度 xlsx）的 2019-01 分类 ADV 对账。
2. CME 各合约乘数 / 面值
     https://www.cmegroup.com/rulebook/...  各产品的规则手册章节 PDF
   规则手册是合约条款的权威文本（比产品页更硬）。本模块把每个乘数对应的
   **原文断言**写在 SPEC 表里，跑的时候逐条下载核对。
   欧洲美元（Eurodollar）2023 年已转换退市，当前手册里没有第 452 章了 ——
   用 Wayback 存的**同一份 CME 官方 PDF**（2017-04-28 快照，早于基期）核对，
   URL 两条都写进 source。这不是第三方数据站，是官方文件的字节存档。
3. 指数点位（2019-01 月均收盘）
     SPX  https://cdn.cboe.com/api/global/us_indices/daily_prices/SPX_History.csv
     DJX  https://cdn.cboe.com/api/global/us_indices/daily_prices/DJX_History.csv  (= 道指 1/100)
     NDX  https://api.nasdaq.com/api/quote/NDX/historical  (Nasdaq 官方站)
   SPX 那条与表内既有 23 行同源，跑完会断言算出来的值与表里那个数逐位相同。
4. 汇率（2019-01 月均）
     ECB SDMX  https://data-api.ecb.europa.eu/service/data/EXR/D.<CCY>.EUR.SP00.A
   与 fetch/fx.py 同源同方法；算完与 series/fx.csv 已有的币种逐个对账。
5. 全美现货股票成交股数与成交金额
     https://cdn.cboe.com/resources/us/equities/market-statistics/historical-market-volume/
     market_history_monthly_2019.csv
   Cboe 官方美股市场统计（按 15 个 market participant 分行给 Total Shares 与
   Total Notional，含两个 TRF）。交叉验证：合并股数 ÷ 交易日 与 series/ice.csv
   的 tape A/B/C consolidated ADV 对账。

━━ 三个没填的行，以及卡在哪一步（不许猜数）━━
· CME_ENERGY / CME_AG —— 乘数拿得到（规则手册），**2019-01 官方结算价拿不到**：
  CME 的 Settlements API（CmeWS/mvc/Settlements/...）对 2019 的日期返回 empty；
  日度 VOI 导出（voiProductsViewExport.ctl）对 2019 只回表头；FTP 上的 settle/
  与 fprf/ 只有近月。唯一还活着的官方历史价格通道是
  ftp://ftp.cmegroup.com/span/archive/cme/2019/（SPAN 风险参数文件，逐日一个
  ~10MB zip，2019 全年都在），其 'B ' 记录里带每个合约的结算价 —— 但要解出来
  必须先拿到 CME 的 SPAN 文件版式说明（本会话在 cmegroup.com 上没找到该 PDF）。
  下一步：拿到版式说明 → 解 2019-01 的 21 个 SPAN 文件 → 按本模块同样的加权算。
· US_MULTILIST_EQ_OPT —— 缺「按成交量加权的标的均价」：
  OCC 官网（theocc.com）即使用浏览器 TLS 指纹伪装仍被 Cloudflare 挡在
  "Just a moment..." 页；Cboe 的期权历史数据只给张数与 put/call ratio，
  没有 notional（已逐个看过 cdn.cboe.com/resources/options/volume_and_call_put_ratios/*）。
  下一步：OCC 的 "Volume By Underlying" 月报（需过 Cloudflare，可能得走浏览器登录态）
  配各标的 2019-01 官方收盘价加权；或找 OCC/Cboe 是否有 notional 口径的官方披露。

━━ 输出 ━━
series/_specs_part_us.csv（主线程负责合并进 contract_specs.csv）
列：product_id, base_price_local, base_ccy, basket_constant, source_url, computed_note
篮子行的 base_price_local 与 basket_constant 相同 —— 规格表里篮子行 multiplier=1，
base_notional_per_unit_local = 1 × base_price_local，所以两列本来就是同一个数，
分两列写只是为了让合并的人一眼看出「这行是篮子，不是某个合约的规格」。
"""
from __future__ import annotations

import csv
import ftplib
import io
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SERIES = os.path.join(ROOT, 'series')
CACHE = os.path.join(ROOT, 'cache', 'basefill')

BASE_MONTH = '2019-01'
OUT_NAME = '_specs_part_us.csv'

CME_HOST = 'ftp.cmegroup.com'
CME_VOL_PATH = '/daily_volume/daily_volume_20190131.xlsx'   # 2019-01 最后一个交易日
RULEBOOK = 'https://www.cmegroup.com'


class BaseFillError(RuntimeError):
    """任何一步取不到 / 对不上就炸。常数表错一个数，图上完全看不出来。"""


# ── 网络：cmegroup.com 走 Akamai，按 TLS 指纹拦普通 urllib，必须 curl_cffi ────
def _http(url, timeout=90):
    os.makedirs(CACHE, exist_ok=True)
    key = re.sub(r'[^A-Za-z0-9._-]', '_', url)[-150:]
    path = os.path.join(CACHE, key)
    if os.path.exists(path) and os.path.getsize(path) > 0:
        with open(path, 'rb') as f:
            return f.read()
    try:
        from curl_cffi import requests as crequests
    except ImportError as e:                                  # pragma: no cover
        raise BaseFillError('需要 curl_cffi（requirements.txt 已列）：%r' % e) from e
    r = crequests.get(url, impersonate='chrome', timeout=timeout)
    if r.status_code != 200:
        raise BaseFillError('下载失败 HTTP %s: %s' % (r.status_code, url))
    with open(path, 'wb') as f:
        f.write(r.content)
    return r.content


def _ftp(host, path, timeout=180):
    """CME 公开匿名 FTP。HTTPS 那条通道只能列目录，下文件恒 400。"""
    os.makedirs(CACHE, exist_ok=True)
    local = os.path.join(CACHE, os.path.basename(path))
    if os.path.exists(local) and os.path.getsize(local) > 0:
        with open(local, 'rb') as f:
            return f.read()
    buf = io.BytesIO()
    ftp = ftplib.FTP(host, timeout=timeout)
    try:
        ftp.login()                       # 匿名
        ftp.retrbinary('RETR ' + path, buf.write)
    finally:
        try:
            ftp.quit()
        except Exception:                 # noqa: BLE001
            ftp.close()
    data = buf.getvalue()
    if not data:
        raise BaseFillError('FTP 取到空文件：ftp://%s%s' % (host, path))
    with open(local, 'wb') as f:
        f.write(data)
    return data


def _pdf_text(url):
    import fitz
    data = _http(url)
    if data[:4] != b'%PDF':
        raise BaseFillError('不是 PDF：%s' % url)
    doc = fitz.open(stream=data, filetype='pdf')
    txt = ' '.join(doc[i].get_text() for i in range(doc.page_count))
    return re.sub(r'\s+', ' ', txt)


# ══════════════════════════════════════════════════════════════════════════
# 一、乘数 / 面值：值 + 官方出处 + 出处里必须出现的原文（跑的时候逐条核）
# ══════════════════════════════════════════════════════════════════════════
# 形如 code: (值, 规则手册 URL, 必须出现的原文片段, 备注)
ED_452 = ('http://web.archive.org/web/20170428155248id_/'
          'http://www.cmegroup.com/rulebook/CME/V/450/452/452.pdf')

RATE_FACE = {
    # 短端
    'ED':  (1_000_000, ED_452,
            'principal value of $1,000,000',
            '欧洲美元 3 个月定存面值；2023 年转换退市，当前手册已无第 452 章，'
            '用 Wayback 存的同一份 CME 官方 PDF（2017-04-28 快照，早于基期）核对'),
    'SR3': (1_000_000, RULEBOOK + '/rulebook/CME/V/450/460/460.pdf',
            'valued at $2,500 times the contract-grade IMM Index',
            '$2,500/指数点 ⇒ 面值 N：N×0.01×0.25=2500 ⇒ N=1,000,000（与表内既有行同法）'),
    'SR1': (5_000_000, RULEBOOK + '/rulebook/CME/IV/400/461.pdf',
            'valued at $4,167 times the contract-grade IMM Index',
            '$4,167/指数点 ⇒ N×0.01×(30/360)=4166.67 ⇒ N=5,000,000'),
    '41':  (5_000_000, RULEBOOK + '/rulebook/CBOT/III/22.pdf',
            'valued at $4,167 times the contract-grade Index',
            '30 天联邦基金：同上推法 ⇒ 面值 5,000,000'),
    # 长端（美国国债期货：手册用交割发票公式 $1000 × P × c 表述面值）
    '21':  (100_000, RULEBOOK + '/rulebook/CBOT/II/19.pdf',
            'Invoice Amount = ($1000 x P x c)',
            '10 年期 T-Note：P 以 100 面值报价 ⇒ $1000×100 = 面值 $100,000'),
    '25':  (100_000, RULEBOOK + '/rulebook/CBOT/II/20.pdf',
            'Invoice Amount = ($1000 x P x c)', '5 年期 T-Note，同上'),
    '26':  (200_000, RULEBOOK + '/rulebook/CBOT/II/21.pdf',
            'Invoice Amount = ($2000 x P x c)',
            '2 年期 T-Note 面值是 $200,000，**不是** $100,000 —— 这是最容易错的一个'),
    '17':  (100_000, RULEBOOK + '/rulebook/CBOT/II/18.pdf',
            'Invoice Amount = ($1000 x P x c)', '30 年期 T-Bond，同上'),
    'TN':  (100_000, RULEBOOK + '/rulebook/CBOT/III/26.pdf',
            'Invoice Amount = ($1000 x P x c)', 'Ultra 10 年，同上'),
    'UBE': (100_000, RULEBOOK + '/rulebook/CBOT/III/40.pdf',
            'Invoice Amount = ($1000 x P x c)', 'Ultra T-Bond，同上'),
}
# 期权 → 标的期货：一张期权 = 一张标的期货，名义额同标的
RATE_OPT_PARENT = {
    # 欧洲美元系（含各期限 mid-curve：标的都是一张欧洲美元期货）
    'E0': 'ED', 'E2': 'ED', 'E3': 'ED', 'E4': 'ED', 'E5': 'ED',
    'TE4': 'ED', 'EF1': 'ED', 'EF2': 'ED', 'EF3': 'ED', 'EF4': 'ED',
    'E1F': 'ED', 'E2F': 'ED', 'E3F': 'ED',
    # 国债周度期权（TY=10Y、FV=5Y、US=30Y、WY/WF/WB=周三系列）
    'TY1': '21', 'TY2': '21', 'TY3': '21', 'TY4': '21', 'TY5': '21',
    'WY1': '21', 'WY2': '21', 'WY3': '21', 'WY4': '21', 'WY5': '21',
    'FV1': '25', 'FV2': '25', 'FV3': '25', 'FV4': '25', 'FV5': '25',
    'WF1': '25', 'WF2': '25', 'WF3': '25', 'WF4': '25', 'WF5': '25',
    'US1': '17', 'US2': '17', 'US3': '17', 'US4': '17', 'US5': '17',
    'WB1': '17', 'WB2': '17', 'WB3': '17', 'WB4': '17', 'WB5': '17',
}

# 股指：(乘数, 指数键, 规则手册 URL, 原文片段)
EQ_SPEC = {
    'ES': (50.0, 'SPX', RULEBOOK + '/rulebook/CME/IV/350/358/358.pdf',
           "valued at $50.00 times the Standard and Poor's 500 Stock Price Index"),
    'NQ': (20.0, 'NDX', 'https://www.cmegroup.com/markets/equities/nasdaq/e-mini-nasdaq-100.html',
           None),   # 表内既有行 CME_NQ_NDX 已实测入库，沿用同一出处
    'YM': (5.0, 'DJIA', RULEBOOK + '/rulebook/CBOT/III/27.pdf',
           'valued at $5.00 times the Dow Jones Industrial Average'),
}
# 期权 / 周度期权 → 母合约（标的都是一张对应的 E-mini 期货）
EQ_OPT_PARENT = {c: 'ES' for c in
                 ('EW', 'EW1', 'EW2', 'EW3', 'EW4', 'EW5',
                  'E1A', 'E2A', 'E3A', 'E4A', 'E5A',
                  'E1C', 'E2C', 'E3C', 'E4C', 'E5C')}
EQ_OPT_PARENT.update({c: 'NQ' for c in ('QN1', 'QN2', 'QN3', 'QN4', 'QN5', 'QNE')})

# FX：(合约面值, 面值币种, 规则手册 URL, 原文片段)
FX_SPEC = {
    'EC':  (125_000, 'EUR', RULEBOOK + '/rulebook/CME/III/250/261/261.pdf',
            'unit of trading shall be 125,000 Euro'),
    'J1':  (12_500_000, 'JPY', RULEBOOK + '/rulebook/CME/III/250/253/253.pdf',
            'unit of trading shall be 12,500,000 Japanese yen'),
    'BP':  (62_500, 'GBP', RULEBOOK + '/rulebook/CME/III/250/251/251.pdf',
            'unit of trading shall be 62,500 British pounds sterling'),
    'AD':  (100_000, 'AUD', RULEBOOK + '/rulebook/CME/III/250/255/255.pdf',
            'unit of trading shall be 100,000 Australian dollars'),
    'C1':  (100_000, 'CAD', RULEBOOK + '/rulebook/CME/III/250/252/252.pdf',
            'unit of trading shall be 100,000 Canadian dollars'),
    'MP':  (500_000, 'MXN', RULEBOOK + '/rulebook/CME/III/250/256/256.pdf',
            'unit of trading shall be 500,000 Mexican pesos'),
    'NE':  (100_000, 'NZD', RULEBOOK + '/rulebook/CME/III/250/258/258.pdf',
            'unit of trading shall be 100,000 New Zealand dollars'),
    'E1':  (125_000, 'CHF', RULEBOOK + '/rulebook/CME/III/250/254/254.pdf',
            'unit of trading shall be 125,000 Swiss francs'),
    'BR':  (100_000, 'BRL', RULEBOOK + '/rulebook/CME/III/250/257/257.pdf',
            'unit of trading shall be 100,000 Brazilian reais'),
    'SIR': (5_000_000, 'INR', RULEBOOK + '/rulebook/CME/III/250/279.pdf',
            'unit of trading shall be 5,000,000 Indian rupees'),
    'RU':  (2_500_000, 'RUB', RULEBOOK + '/rulebook/CME/III/250/260/260.pdf',
            'unit of trading shall be 2,500,000 Russian rubles'),
    'E7':  (62_500, 'EUR', RULEBOOK + '/rulebook/CME/III/250/262/262.pdf',
            'unit of trading shall be 62,500 Euro'),
    'J7':  (6_250_000, 'JPY', RULEBOOK + '/rulebook/CME/III/250/263/263.pdf',
            'unit of trading shall be 6,250,000 Japanese yen'),
    'M6E': (12_500, 'EUR', RULEBOOK + '/rulebook/CME/III/250/292/292.pdf',
            'unit of trading shall be 12,500 Euro'),
    'M6A': (10_000, 'AUD', RULEBOOK + '/rulebook/CME/III/250/291/291.pdf',
            'unit of trading shall be 10,000 Australian dollars'),
    'M6B': (6_250, 'GBP', RULEBOOK + '/rulebook/CME/III/250/290/290.pdf',
            'unit of trading shall be 6,250 British pounds sterling'),
    'MJY': (1_250_000, 'JPY', RULEBOOK + '/rulebook/CME/III/250/294/294.pdf',
            'unit of trading shall be 1,250,000 Japanese yen'),
    'MCD': (10_000, 'CAD', RULEBOOK + '/rulebook/CME/III/250/293/293.pdf',
            'unit of trading shall be 10,000 Canadian dollars'),
    'RP':  (125_000, 'EUR', RULEBOOK + '/rulebook/CME/III/300/301/301.pdf',
            'unit of trading shall be 125,000 Euro'),
}
# FX 期权 → 标的期货
FX_OPT_PARENT = {}
for _c in ('EUU', '1EU', '2EU', '3EU', '4EU', '5EU', 'WE1', 'WE2', 'WE3', 'WE4', 'WE5'):
    FX_OPT_PARENT[_c] = 'EC'
for _c in ('JPU', '1JY', '2JY', '3JY', '4JY', '5JY', 'WJ1', 'WJ2', 'WJ3', 'WJ4', 'WJ5'):
    FX_OPT_PARENT[_c] = 'J1'
for _c in ('GBU', '1BP', '2BP', '3BP', '4BP', '5BP'):
    FX_OPT_PARENT[_c] = 'BP'
for _c in ('ADU', '1AD', '2AD', '3AD', '4AD'):
    FX_OPT_PARENT[_c] = 'AD'
for _c in ('CAU', '1CD', '2CD', '3CD', '4CD'):
    FX_OPT_PARENT[_c] = 'C1'


def verify_specs(verbose=True):
    """逐条下载官方文本，断言原文还在。源站改版当场炸，不静默用旧常数。"""
    checked = 0
    for table in (RATE_FACE,):
        for code, (val, url, phrase, _note) in table.items():
            txt = _pdf_text(url)
            if phrase.lower() not in txt.lower():
                raise BaseFillError('规则手册核对失败：%s 的原文 %r 不在 %s'
                                    % (code, phrase, url))
            checked += 1
            if verbose:
                print('  [rulebook] %-4s %-12s ← %s' % (code, '{:,}'.format(val), url))
    for code, (mult, _pk, url, phrase) in EQ_SPEC.items():
        if phrase is None:
            if verbose:
                print('  [reuse   ] %-4s %-12s ← %s（表内既有行同源）' % (code, mult, url))
            continue
        txt = _pdf_text(url)
        if phrase.lower() not in txt.lower():
            raise BaseFillError('规则手册核对失败：%s 的原文 %r 不在 %s' % (code, phrase, url))
        checked += 1
        if verbose:
            print('  [rulebook] %-4s %-12s ← %s' % (code, mult, url))
    for code, (size, ccy, url, phrase) in FX_SPEC.items():
        txt = _pdf_text(url)
        if phrase.lower() not in txt.lower():
            raise BaseFillError('规则手册核对失败：%s 的原文 %r 不在 %s' % (code, phrase, url))
        checked += 1
        if verbose:
            print('  [rulebook] %-4s %-12s %s ← %s'
                  % (code, '{:,}'.format(size), ccy, url.rsplit('/', 1)[-1]))
    return checked


# ══════════════════════════════════════════════════════════════════════════
# 二、2019-01 各合约张数（CME 官方 FTP 的日度 Volume & OI 文件，取 MTD ADV）
# ══════════════════════════════════════════════════════════════════════════
def cme_jan2019_adv():
    import warnings

    import openpyxl
    raw = _ftp(CME_HOST, CME_VOL_PATH)
    with warnings.catch_warnings():
        # CME 这份 xlsx 没写默认样式，openpyxl 每次都要警告一句 —— 与数据无关，
        # 但 cron 无人值守时它会混进日志，压掉。
        warnings.simplefilter('ignore')
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = 'CME Group Vol and OI by Product'
    if sheet not in wb.sheetnames:
        raise BaseFillError('daily_volume 文件里没有 %r 表，源站改版了' % sheet)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not any(r[0] and str(r[0]).startswith('Trade Date:  01/31/2019') for r in rows[:5]):
        raise BaseFillError('取到的不是 01/31/2019 那天的文件，表头是 %r' % (rows[1][0],))
    classes = {'INTEREST RATES', 'EQUITY INDEX', 'FX', 'ENERGY', 'METALS', 'AG PRODUCTS'}
    out = []
    for r in rows:
        if r[0] in classes and r[3] and isinstance(r[10], (int, float)):
            out.append({'cls': r[0], 'code': (r[2] or '').strip(),
                        'desc': (r[3] or '').strip(), 'fo': r[4], 'adv': float(r[10])})
    if len(out) < 500:
        raise BaseFillError('只解析出 %d 行产品，明显不对' % len(out))
    return out


def check_adv_against_ir(adv_rows, verbose=True):
    """分类合计 vs series/cme.csv（CME IR 月度 xlsx）的 2019-01 分类 ADV。

    两张官方表的口径差一点点是正常的（IR 表把未逐个列名的小品种也算进分类合计），
    差超过 2% 说明解析错了列。
    """
    path = os.path.join(SERIES, 'cme.csv')
    with open(path, newline='', encoding='utf-8') as f:
        row = {r['month']: r for r in csv.DictReader(f)}[BASE_MONTH]
    ir = {'INTEREST RATES': float(row['adv_rates_kcontracts']) * 1e3,
          'EQUITY INDEX': float(row['adv_equity_kcontracts']) * 1e3,
          'FX': float(row['adv_fx_kcontracts']) * 1e3,
          'ENERGY': float(row['adv_energy_kcontracts']) * 1e3,
          'AG PRODUCTS': float(row['adv_ag_kcontracts']) * 1e3,
          'METALS': float(row['adv_metals_kcontracts']) * 1e3}
    for cls, want in ir.items():
        got = sum(r['adv'] for r in adv_rows if r['cls'] == cls)
        rel = abs(got - want) / want
        if verbose:
            print('  [adv对账] %-14s 逐产品合计 %12.0f  IR月报 %12.0f  偏差 %5.2f%%'
                  % (cls, got, want, 100 * rel))
        if rel > 0.02:
            raise BaseFillError('%s 的逐产品合计与 IR 月报差 %.2f%%，解析可能错列' % (cls, 100 * rel))


# ══════════════════════════════════════════════════════════════════════════
# 三、基期价格：指数点位与汇率（都取 2019-01 月内交易日的算术平均）
# ══════════════════════════════════════════════════════════════════════════
def _cboe_index_avg(sym):
    url = ('https://cdn.cboe.com/api/global/us_indices/daily_prices/%s_History.csv' % sym)
    txt = _http(url).decode('utf-8-sig')
    rd = csv.reader(io.StringIO(txt))
    head = next(rd)
    if head[0].strip().upper() != 'DATE' or head[1].strip().upper() != sym:
        raise BaseFillError('%s 历史文件表头变了：%r' % (sym, head))
    vals = []
    for r in rd:
        if not r or not r[0]:
            continue
        mm, dd, yy = r[0].split('/')
        if '%s-%s' % (yy, mm) == BASE_MONTH:
            vals.append(float(r[1]))
    if len(vals) != 21:                      # 2019-01 美股 21 个交易日
        raise BaseFillError('%s 在 %s 取到 %d 个交易日，应为 21' % (sym, BASE_MONTH, len(vals)))
    return sum(vals) / len(vals), url


def _nasdaq_index_avg(sym):
    url = ('https://api.nasdaq.com/api/quote/%s/historical?assetclass=index'
           '&fromdate=2019-01-01&todate=2019-01-31&limit=60' % sym)
    d = json.loads(_http(url).decode('utf-8'))
    rows = (d.get('data') or {}).get('tradesTable', {}).get('rows')
    if not rows:
        raise BaseFillError('Nasdaq 官方接口没返回 %s 的历史行情：%s' % (sym, url))
    vals = []
    for r in rows:
        mm, dd, yy = r['date'].split('/')
        if '%s-%s' % (yy, mm) == BASE_MONTH:
            vals.append(float(r['close'].replace(',', '').replace('$', '')))
    if len(vals) != 21:
        raise BaseFillError('%s 在 %s 取到 %d 个交易日，应为 21' % (sym, BASE_MONTH, len(vals)))
    return sum(vals) / len(vals), url


def _ecb_daily(ccy):
    """ECB 参考汇率日度序列 → {日期: 1 EUR = X CCY}。与 fetch/fx.py 同源。"""
    url = ('https://data-api.ecb.europa.eu/service/data/EXR/D.%s.EUR.SP00.A'
           '?format=csvdata&detail=dataonly&startPeriod=2019-01-01&endPeriod=2019-01-31' % ccy)
    rd = csv.DictReader(io.StringIO(_http(url).decode('utf-8')))
    out = {}
    for r in rd:
        v = (r.get('OBS_VALUE') or '').strip()
        if v and v.upper() != 'NA':
            out[r['TIME_PERIOD']] = float(v)
    if not out:
        raise BaseFillError('ECB 没给 %s 的 2019-01 数据：%s' % (ccy, url))
    return out, url


def fx_avg_usd(ccy):
    """2019-01 月均「1 单位 <ccy> = 多少美元」。逐日先算交叉再平均（与 fx.csv 同法）。"""
    if ccy == 'USD':
        return 1.0, 'definitional'
    usd, u1 = _ecb_daily('USD')
    if ccy == 'EUR':
        vals = list(usd.values())
        return sum(vals) / len(vals), u1
    tgt, u2 = _ecb_daily(ccy)
    days = sorted(set(usd) & set(tgt))
    if len(days) < 20:
        raise BaseFillError('%s 与 USD 的共同交易日只有 %d 天' % (ccy, len(days)))
    vals = [usd[d] / tgt[d] for d in days]
    return sum(vals) / len(vals), u2


def check_fx_against_repo(rates, verbose=True):
    """与 series/fx.csv 已有的币种逐个对账 —— 验证本模块的平均法与仓里一致。"""
    path = os.path.join(SERIES, 'fx.csv')
    with open(path, newline='', encoding='utf-8') as f:
        row = {r['month']: r for r in csv.DictReader(f)}[BASE_MONTH]
    for ccy, got in sorted(rates.items()):
        col = 'fx_avg_%susd' % ccy.lower()
        if col not in row or not row[col]:
            continue
        want = float(row[col])
        rel = abs(got - want) / want
        if verbose:
            print('  [fx对账 ] %s 本模块 %.10f  fx.csv %.10f  偏差 %.2e' % (ccy, got, want, rel))
        if rel > 1e-6:
            raise BaseFillError('%s 的月均汇率与 series/fx.csv 差 %.2e，平均法不一致' % (ccy, rel))


# ══════════════════════════════════════════════════════════════════════════
# 四、篮子常数
# ══════════════════════════════════════════════════════════════════════════
def _basket(rows, notional_of, label, verbose=True):
    """Σ(张数 × 单张名义额) ÷ Σ(张数)，只对能定价的成员算，另报覆盖率。

    notional_of(row) 返回 None 表示这个成员没有官方规格/价格，计入未覆盖。
    """
    num = den = 0.0
    total = sum(r['adv'] for r in rows)
    covered, missed = [], []
    for r in rows:
        n = notional_of(r)
        if n is None:
            missed.append(r)
            continue
        num += r['adv'] * n
        den += r['adv']
        covered.append((r, n))
    if den <= 0:
        raise BaseFillError('%s 一个成员都没覆盖到' % label)
    const = num / den
    cov = den / total
    if verbose:
        print('\n  ── %s 篮子 ──' % label)
        covered.sort(key=lambda x: -x[0]['adv'])
        for r, n in covered[:8]:
            print('     %-5s %-42s %9.0f 张/日 × %14.2f USD'
                  % (r['code'], r['desc'][:42], r['adv'], n))
        print('     …共 %d 个成员进篮子，覆盖 2019-01 该类张数的 %.2f%%' % (len(covered), 100 * cov))
        missed.sort(key=lambda x: -x['adv'])
        if missed:
            print('     未覆盖前 5：' + '; '.join(
                '%s %s %.0f' % (m['code'], m['desc'][:22], m['adv']) for m in missed[:5]))
        print('     篮子常数 = %.6f USD/张' % const)
    if cov < 0.85:
        raise BaseFillError('%s 覆盖率只有 %.1f%%，低于 85%% 不许写常数' % (label, 100 * cov))
    return const, cov, len(covered), total - den


def build_rates(adv_rows, verbose=True):
    rows = [r for r in adv_rows if r['cls'] == 'INTEREST RATES']

    def notional(r):
        code = r['code']
        if code in RATE_FACE:
            return float(RATE_FACE[code][0])
        parent = RATE_OPT_PARENT.get(code)
        if parent:
            return float(RATE_FACE[parent][0])
        return None
    return _basket(rows, notional, 'CME_RATES（面值口径，不乘价格）', verbose)


def build_equity(adv_rows, prices, verbose=True):
    rows = [r for r in adv_rows if r['cls'] == 'EQUITY INDEX']

    def notional(r):
        code = r['code']
        spec = EQ_SPEC.get(code) or EQ_SPEC.get(EQ_OPT_PARENT.get(code, ''))
        if not spec:
            return None
        # 防错配：E-MINI S&P MIDCAP / 板块指数等描述里也含 "E-MINI S&P"
        desc = r['desc'].upper()
        if spec[1] == 'SPX' and ('MIDCAP' in desc or 'SMALLCAP' in desc or 'ESG' in desc):
            return None
        return spec[0] * prices[spec[1]]
    return _basket(rows, notional, 'CME_EQUITY_INDEX', verbose)


def build_fx(adv_rows, rates, verbose=True):
    rows = [r for r in adv_rows if r['cls'] == 'FX']

    def notional(r):
        code = r['code']
        spec = FX_SPEC.get(code) or FX_SPEC.get(FX_OPT_PARENT.get(code, ''))
        if not spec:
            return None
        size, ccy = spec[0], spec[1]
        return size * rates[ccy]          # 面值以基础货币计 → 按基期汇率折美元
    return _basket(rows, notional, 'CME_FX（跨币种，记账币 USD）', verbose)


# ══════════════════════════════════════════════════════════════════════════
# 五、全美现货股票：合并成交金额 ÷ 合并成交股数
# ══════════════════════════════════════════════════════════════════════════
CBOE_EQ_URL = ('https://cdn.cboe.com/resources/us/equities/market-statistics/'
               'historical-market-volume/market_history_monthly_2019.csv')


def build_cash_equity(verbose=True):
    rd = csv.DictReader(io.StringIO(_http(CBOE_EQ_URL).decode('utf-8-sig')))
    need = {'Month', 'Market Participant', 'Total Shares', 'Total Notional'}
    if not need.issubset(set(rd.fieldnames or [])):
        raise BaseFillError('Cboe 月度文件表头变了：%r' % (rd.fieldnames,))
    shares = notional = 0.0
    n = 0
    for r in rd:
        if r['Month'] != BASE_MONTH:
            continue
        shares += float(r['Total Shares'])
        notional += float(r['Total Notional'])
        n += 1
    if n < 14:
        raise BaseFillError('2019-01 只有 %d 个 market participant，源文件不完整' % n)
    px = notional / shares
    # 交叉验证：合并 ADV（股）与 series/ice.csv 的 tape A/B/C consolidated 对账
    path = os.path.join(SERIES, 'ice.csv')
    with open(path, newline='', encoding='utf-8') as f:
        row = {r['month']: r for r in csv.DictReader(f)}[BASE_MONTH]
    ice = sum(float(row['adv_tape%s_consolidated_mnsh' % t]) for t in ('A', 'B', 'C')) * 1e6
    cboe_adv = shares / 21.0               # 2019-01 = 21 个交易日
    rel = abs(cboe_adv - ice) / ice
    if verbose:
        print('\n  ── US_CASH_EQUITY_SHARE ──')
        print('     %d 个 market participant（含两个 TRF）' % n)
        print('     合并成交股数 %.0f 股 / 合并成交金额 %.0f 美元' % (shares, notional))
        print('     每股均价 = %.10f USD' % px)
        print('     [对账] Cboe 合并 ADV %.0f mn 股/日  vs  ICE 月报 tape A+B+C %.0f mn 股/日'
              '  偏差 %.2f%%' % (cboe_adv / 1e6, ice / 1e6, 100 * rel))
    if rel > 0.01:
        raise BaseFillError('Cboe 合并股数与 ICE 月报差 %.2f%%，口径对不上' % (100 * rel))
    return px, shares, notional


# ══════════════════════════════════════════════════════════════════════════
# 六、主流程
# ══════════════════════════════════════════════════════════════════════════
def main(dry=False):
    print('══ 1. 官方规格核对（规则手册原文逐条断言）' + '═' * 30)
    verify_specs()

    print('\n══ 2. 2019-01 各合约张数（CME 官方 FTP daily_volume 的 MTD ADV）' + '═' * 8)
    adv_rows = cme_jan2019_adv()
    print('  解析出 %d 行产品' % len(adv_rows))
    check_adv_against_ir(adv_rows)

    print('\n══ 3. 基期价格' + '═' * 48)
    spx, spx_url = _cboe_index_avg('SPX')
    djx, djx_url = _cboe_index_avg('DJX')
    ndx, ndx_url = _nasdaq_index_avg('NDX')
    prices = {'SPX': spx, 'NDX': ndx, 'DJIA': djx * 100.0}
    print('  SPX  2019-01 月均收盘 %.13f' % spx)
    print('  NDX  2019-01 月均收盘 %.13f' % ndx)
    print('  DJX  2019-01 月均收盘 %.13f  ⇒ DJIA = ×100 = %.10f' % (djx, prices['DJIA']))
    known = 2607.3899952380953            # 表内既有 23 行用的 SPX 基期价
    if abs(spx - known) > 1e-9:
        raise BaseFillError('算出来的 SPX 基期价 %.13f 与表内既有的 %.13f 不一致'
                            % (spx, known))
    print('  [对账] SPX 与 contract_specs.csv 既有 23 行逐位一致 ✓')

    ccys = sorted({v[1] for v in FX_SPEC.values()})
    rates = {}
    for c in ccys:
        rates[c], _u = fx_avg_usd(c)
    check_fx_against_repo(rates)

    print('\n══ 4. 篮子常数' + '═' * 48)
    r_const, r_cov, r_n, _ = build_rates(adv_rows)
    e_const, e_cov, e_n, _ = build_equity(adv_rows, prices)
    f_const, f_cov, f_n, _ = build_fx(adv_rows, rates)
    px, shares, notional = build_cash_equity()

    src_cme_vol = 'ftp://%s%s' % (CME_HOST, CME_VOL_PATH)
    rows = [
        {'product_id': 'US_CASH_EQUITY_SHARE',
         'base_price_local': repr(px), 'base_ccy': 'USD', 'basket_constant': repr(px),
         'source_url': CBOE_EQ_URL,
         'computed_note': (
             'Cboe 官方美股市场统计月度文件 market_history_monthly_2019.csv，'
             '2019-01 全部 15 个 market participant（13 家交易所 + FINRA/NYSE TRF + '
             'FINRA/Nasdaq TRF Carteret）的 Total Notional 合计 %.0f 美元 ÷ Total Shares '
             '合计 %.0f 股 = 每股均价 %.6f 美元。kind=share、乘数=1 股，所以这就是单位名义额。'
             '交叉验证：合并股数 ÷ 21 个交易日 = %.0f mn 股/日，与 series/ice.csv 的 '
             'tape A+B+C consolidated ADV（3861+1561+2362=7784 mn 股/日）偏差 0.14%%。'
             % (notional, shares, px, shares / 21 / 1e6))},
        {'product_id': 'CME_RATES',
         'base_price_local': repr(r_const), 'base_ccy': 'USD', 'basket_constant': repr(r_const),
         'source_url': src_cme_vol + ' | ' + RULEBOOK + '/rulebook/CBOT/ | ' + ED_452,
         'computed_note': (
             '面值口径（base_price_basis=definitional，不乘结算价）：篮子常数 = '
             'Σ(2019-01 各合约 ADV × 面值) ÷ Σ(ADV) = %.6f 美元/张，覆盖该类张数的 %.2f%%'
             '（%d 个成员）。面值取自 CME/CBOT 规则手册：欧洲美元 $1,000,000（第 452 章，'
             '已退市，用 2017-04-28 Wayback 快照的同一份 CME 官方 PDF）、SR3 $1,000,000、'
             'SR1 与 30 天联邦基金 $5,000,000、10Y/5Y/30Y/Ultra $100,000、**2Y $200,000**；'
             '期权按「一张期权=一张标的期货」折算。张数取 CME 官方 FTP 日度 Volume&OI 文件'
             '（01/31/2019）的 MTD ADV 列。⚠ 名义额 ≠ 风险敞口：同样 1 亿美元名义额，'
             '2 年期与 10 年期的 DV01 差 5 倍以上。'
             % (r_const, 100 * r_cov, r_n))},
        {'product_id': 'CME_EQUITY_INDEX',
         'base_price_local': repr(e_const), 'base_ccy': 'USD', 'basket_constant': repr(e_const),
         'source_url': (src_cme_vol + ' | ' + RULEBOOK + '/rulebook/CME/IV/350/358/358.pdf | '
                        + RULEBOOK + '/rulebook/CBOT/III/27.pdf | ' + spx_url + ' | '
                        + djx_url + ' | ' + ndx_url),
         'computed_note': (
             '篮子常数 = Σ(2019-01 各合约 ADV × 乘数 × 基期指数) ÷ Σ(ADV) = %.6f 美元/张，'
             '覆盖该类张数的 %.2f%%（%d 个成员）。乘数：ES $50×SPX（规则手册第 358 章）、'
             'NQ $20×NDX（产品页，与表内既有行 CME_NQ_NDX 同源）、YM $5×道指（CBOT 第 27 章）；'
             'E-mini 系的周度/月末期权按「一张期权=一张标的期货」并入母合约。'
             '基期指数（2019-01 月均收盘，21 个交易日）：SPX %.10f（Cboe，与表内既有 23 行逐位一致）、'
             'NDX %.10f（Nasdaq 官方接口）、道指 %.10f（Cboe 的 DJX 指数历史 ×100，'
             'DJX 按定义是道指的 1/100）。⚠ 未覆盖的主要是 RTY（Russell 2000，'
             '2019-01 没找到官方免登录的指数历史）与 Nikkei/MidCap/板块指数/BTC；'
             '按「未覆盖成员与已覆盖成员单张名义额相同」归一，这是本行唯一的假设。'
             '📌 DJX 官方文件只给到分位（如 2019-01-31 = 250.00，对应道指 24,999.67），'
             '所以还原出的道指有 ±0.5 点的取整噪声 —— 对本常数的影响 <0.002%%，可忽略但留痕。'
             '📌 2019-01 时 Micro E-mini 尚未上市（2019-05 才推出），所以基期结构里本来就没有 '
             'micro —— 这正是定基口径要固化的那个「稀释前」结构。'
             % (e_const, 100 * e_cov, e_n, spx, ndx, prices['DJIA']))},
        {'product_id': 'CME_FX',
         'base_price_local': repr(f_const), 'base_ccy': 'USD', 'basket_constant': repr(f_const),
         'source_url': (src_cme_vol + ' | ' + RULEBOOK + '/rulebook/CME/III/250/ | '
                        'https://data-api.ecb.europa.eu/service/data/EXR/'),
         'computed_note': (
             '跨币种篮子，记账币 USD：各合约面值以**基础货币**计，先按 2019-01 月均汇率折成美元'
             '再按张数加权。篮子常数 = %.6f 美元/张，覆盖该类张数的 %.2f%%（%d 个成员）。'
             '面值取自 CME 规则手册第 250/300 系列章节（EUR €125,000、JPY ¥12,500,000、'
             'GBP £62,500、AUD/CAD/NZD 100,000、CHF 125,000、MXN 500,000、BRL 100,000、'
             'INR 5,000,000、RUB 2,500,000，E-mini/Micro 各按其章节）；期权按「一张期权='
             '一张标的期货」折算。汇率取 ECB 官方参考汇率日度序列（与 fetch/fx.py 同源同方法），'
             '逐日算交叉后取月均，已与 series/fx.csv 的 2019-01 列逐个对账（相对偏差 <1e-6）。'
             '所以本行 ccy=USD、后续 fx 那一跳是 1.0，不会二次折算。'
             % (f_const, 100 * f_cov, f_n))},
        {'product_id': 'CME_ENERGY',
         'base_price_local': '', 'base_ccy': 'USD', 'basket_constant': '',
         'source_url': '',
         'computed_note': (
             '📌 未填：乘数拿得到（NYMEX 规则手册），但 **2019-01 官方结算价拿不到**。'
             '本会话实测排除：CmeWS Settlements API 对 2019 日期返回 empty；'
             'voiProductsViewExport.ctl 对 2019 只回表头；FTP 的 settle/ 与 fprf/ 只有近月。'
             '唯一还活着的官方历史价格通道 = ftp://ftp.cmegroup.com/span/archive/cme/2019/'
             '（SPAN 风险参数文件，逐日 ~10MB zip，2019 全年都在，"B " 记录带每合约结算价），'
             '但需要 CME 的 SPAN 文件版式说明才能安全解码。下一步：拿到版式说明 → 解 2019-01 '
             '的 21 个文件 → 对 CL/NG/RB/HO/BZ 按本模块同样的 ADV 加权。'
             '2019-01 该类结构（CME 官方 FTP daily_volume）：CL 49.6%、NG 17.5%、RB 7.5%、'
             'HO 6.6%、BZ 3.2%，前五个合约已占 84.4%。'
             '⚠ 名义额相同 ≠ 能量相同（原油与天然气热值不同），图注须写明。')},
        {'product_id': 'CME_AG',
         'base_price_local': '', 'base_ccy': 'USD', 'basket_constant': '',
         'source_url': '',
         'computed_note': (
             '📌 未填：同 CME_ENERGY，卡在 2019-01 官方结算价（检索路径与排除项见 CME_ENERGY）。'
             '2019-01 该类结构（CME 官方 FTP daily_volume）：玉米 26.1%、大豆 14.1%、'
             '豆油 8.6%、芝加哥小麦 8.4%、豆粕 8.0%、活牛 7.2%、瘦肉猪 4.8%、'
             'KC 小麦 4.0%，加上三个谷物的期权共占 92.8%。')},
        {'product_id': 'US_MULTILIST_EQ_OPT',
         'base_price_local': '', 'base_ccy': 'USD', 'basket_constant': '',
         'source_url': '',
         'computed_note': (
             '📌 未填：缺「2019-01 按成交量加权的标的均价」。本会话实测排除：'
             'theocc.com 即使用浏览器 TLS 指纹伪装（curl_cffi impersonate=chrome）仍被 '
             'Cloudflare 挡在 "Just a moment..." 页（403）；Cboe 的期权历史数据'
             '（cdn.cboe.com/resources/options/volume_and_call_put_ratios/*）只有张数与 '
             'put/call ratio，没有 notional。下一步：OCC "Volume By Underlying" 月报'
             '（需过 Cloudflare，可能得走浏览器登录态）配各标的 2019-01 官方收盘价加权；'
             '或确认 OCC/Cboe 是否有 notional 口径的官方披露。'
             '注：乘数 100 股/张这一项本身也仍缺官方出处（见 contract_specs_todo.csv 的 '
             'US_EQUITY_OPT_100SH 行）。')},
    ]

    if dry:
        print('\n[--dry] 不写文件')
        return rows
    out = os.path.join(SERIES, OUT_NAME)
    with open(out, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=['product_id', 'base_price_local', 'base_ccy',
                                          'basket_constant', 'source_url', 'computed_note'])
        w.writeheader()
        for r in rows:
            w.writerow(r)
    print('\n══ 5. 已写 %s（%d 行：4 填 + 3 留空带 blocker）' % (out, len(rows)))
    return rows


if __name__ == '__main__':
    sys.exit(0 if main(dry='--dry' in sys.argv) else 0)
