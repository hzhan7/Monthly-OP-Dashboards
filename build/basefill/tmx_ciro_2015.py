# -*- coding: utf-8 -*-
"""TMX 加拿大现货 12 列的**历史回填**：series/tmx.csv 从 2021-08 起推到 2015-01 起（+79 期）。

用法:
    python3 build/basefill/tmx_ciro_2015.py                      # 取数 + 核对 + 写 CSV
    python3 build/basefill/tmx_ciro_2015.py --dry                # 只打印，不写文件
    python3 build/basefill/tmx_ciro_2015.py --refresh            # 强制重下工作簿
    python3 build/basefill/tmx_ciro_2015.py --cache-dir <路径>   # 原始档案落在哪

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
这个文件为什么存在，以及为什么**不**把它塞进 fetch/tmx.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
fetch/tmx.py 的现货腿只吃 TMX IR 的 CTS 新闻稿正文，而**正文里的表格 2021-08 才开始有**
（该模块口径坑 7；2026-08-18 复核 feed 140 期，无表 80 期 / 有表 60 期，边界干净无交叉）。
更早那 80 期正文只有一段摘要 + 一条 `/resource/en/<id>` 链接，表格落在 tmx.com 域下，
对本网络整段 CloudFront 403。所以那不是「抓取器窗口没开」，是**那条链路没有历史入口**。

历史另有入口，而且是**另一家机构**：CIRO（原 IIROC，TSX / TSXV / Alpha 的市场监管服务
提供方）发布『Report of Marketshare by Marketplace (Historical 2015–Present)』，
一份 xlsx 覆盖 2015-01 至今、逐月无洞，列里恰好有 Toronto Stock Exchange /
TSX Venture Exchange / Alpha / Alpha-X / Alpha-DRK。

    落地页 https://www.ciro.ca/markets/reports-statistics-and-other-information/
               reports-market-share-marketplace
    直链   https://www.ciro.ca/media/<id>/download?attachment
           （2026-08-18 是 8821；**id 会漂，必须每次从落地页现抓**）

不塞进 fetch/tmx.py 的三条理由，缺一条都不足以单独成立：
  1. **另一个源、另一种版式**：监管方的市场份额报表，不是 TMX 自己的新闻稿。
  2. **只会用这一次**：2015-01~2021-07 这个洞补完就永远关上了 —— 2021-08 起 CTS 正文
     自己有表，CIRO 再往右填一格都是错的（见下面「方向只能往左」）。
  3. **口径不同**，混进无人值守链路迟早出事：CIRO 的历史报每月更新、含最新月，
     真把它焊进 update()，某个月 CTS 迟到时它就会抢先把 CIRO 的数写进当月，
     一列里悄悄混两把尺子，而且没有任何人会发现。
照 build/basefill/cboe_2016.py 与 build/basefill/spgi_history.py 的先例放在这里。
⚠ 与 mtk 那次不同：mtk 是**同一个源、同一种版式**、只是窗口没开，那次的正确解是改
  update() 让抓取器自己长回去；这里是**另一个机构的另一套表**，一次性脚本才是对的。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
口径：CIRO 与 TMX 自报不是同一把尺子 —— 方向只能往左，且要打断点
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
60 个重叠月（2021-08~2026-07）逐月比过（本脚本每次运行都会重算一遍并打出来，
数字不写死）。TMX 自报 ÷ CIRO 的中位数：笔数三列与 tsxv 两列 ≈ 1.00000（多数月逐位相同），
而 tsx_volume_shares 只有 0.98683、tsx_value_cad 1.00162、alpha_value_cad 1.00249 ——
**量偏低、额偏高，方向相反**，所以不是「含不含大宗对敲」那种可加减的一块。
（2021-08 实测：TMX 自报 TSX 6,324,849,035 股，夹在 CIRO 的 Non-Cross 6,185,693,468
与 All Trade 6,429,009,165 之间；解出来的「对敲计入比例」在 60 个月里 0.09~0.73 乱跳。）

⇒ 本脚本的三条硬规矩：
  · **只写 2021-08 之前的月份**（`CTS_FROM`），一格都不越界；
  · 走 fetch/tmx.py 的 `_merge`，**已有值永不覆盖**（哪怕是自己上次写的）；
  · 接缝的口径台阶交给 build/specs/tmx.py 画断点线：实测 tsx_volume_shares −1.62%、
    tmx_all_volume_shares −0.98% 值得标，其余 10 列 |台阶| ≤0.17% 不标
    （标了等于说「这两个月不可比」，而它们其实可比）。

**为什么不整段改用 CIRO**（那样就没有台阶了）：那等于把 TMX 官方新闻稿印出来的数换成
第三方重算值，还要把官方直接披露的 `tmx_all_*` 降级成我们自己的加总。本仓的规矩是
「入库值是当期官方公告原值」，一条 1.3% 的台阶 + 一条断点线，比 60 个月的静默替换便宜。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
版式与取行
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
· 6 个 sheet：Value Traded / Volume Traded / Number of Trades 及各自的 `% of …` 版本。
  只读前三个，百分比版一律不碰。
· 表头在**第 3 行**（第 1 行空、第 2 行是大标题），但本脚本**不写死行号**：
  逐行扫到第一列等于 'Month' 的那一行为止。
· 行按 `Trade Type and Listing Market` × 上市地拆成 11 种，只取
  **`All Trade All Listing Total`** —— 它才是「该市场当月全部成交」。
  Non-Cross / Intentional Cross 是它的两个子集，取错就是漏掉大宗对敲。
· 月份格式 '2015-Jan'，转成 '2015-01'。
· `All Traded Marketplaces` 那一列是**全加拿大**（含 CSE / Nasdaq CXC / MATCHNow / NEO
  等 22 个市场），**不是 TMX 集团**，绝不能拿它当 tmx_all_*。
  回补段的 `tmx_all_* = TSX + TSXV + Alpha`（Alpha-X / Alpha-DRK 在 CIRO 里同样
  2023-11 才有值，这一段天然不参与），加总方是我们 —— 图注里已写明。
· 数值：成交股数与笔数是整数；成交额是浮点（源里就带到 3 位小数，个别 4 位），
  **原样入库不四舍五入** —— 这样任何人拿 CIRO 那份 xlsx 都能逐位对上。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
不做的事
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
· **alphax_drk_* 三列不补。** TMX 与 CIRO 两个**互相独立**的源都恰好 2023-11 才有
  Alpha-X / Alpha-DRK（本脚本每次运行都会复验这条断言）⇒ 这两个盘口此前不单独披露，
  是真空缺不是解析漏。
· **发布日不记。** CIRO 那份 xlsx 只在文件名里带一个 `202607`，没有逐月发布日；
  拿 CTS 新闻稿的 PressReleaseDate 去套是另一份文件的日期。
  series/source_dates.csv 里这 79 个月**留白**，本脚本一行都不写。
· **2014 及更早补不了。** CIRO 官方页面写明 2007–2014 的报表需要人工索取
  （"please contact"），所以现货 12 列的真实地板就是 2015-01，比 MX 的 2002-01 晚 13 年。
"""
import argparse
import csv
import importlib.util
import os
import re
import statistics
import sys
import time
import urllib.error
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SERIES_DIR = os.path.join(ROOT, 'series')

LANDING = ('https://www.ciro.ca/markets/reports-statistics-and-other-information/'
           'reports-market-share-marketplace')
#: 落地页锚文本里必须含它才是「历史报」（另一条 media 链接是「当月报」，只有一个月）。
HIST_MARK = 'Historical'
FNAME = 'ciro_marketshare_historical.xlsx'
MIN_BYTES = 500_000          # 实测 1,438,433 字节；小于这个数不像是那份全历史工作簿
#: WAF 会间歇 403（实测同一分钟内 200/403 交替）。**单次 403 绝不能判成「文件没了」**。
RETRIES = 8

#: 换源边界：这个月**及其之后**归 fetch/tmx.py 的 CTS 解析器，本脚本一格都不写。
#: 与 fetch/tmx.py 的 SPOT_START 是同一个月，改一处要改两处（本脚本启动时会核对）。
CTS_FROM = '2021-08'

SHEETS = {'Volume Traded': 'volume',
          'Value Traded': 'value',
          'Number of Trades': 'transactions'}
ROW_WANTED = 'All Trade All Listing Total'
MONTH_RE = re.compile(r'^(\d{4})-([A-Za-z]{3})$')
_MON = {m: i for i, m in enumerate(
    ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
     'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'], 1)}

#: CSV 列 → (CIRO 列名, 指标)。tmx_all_* 不在这里，它由 GROUP 加总（见 docstring）。
VENUE = {'tsx': 'Toronto Stock Exchange',
         'tsxv': 'TSX Venture Exchange',
         'alpha': 'Alpha'}
METRIC = {'volume_shares': 'volume', 'value_cad': 'value', 'transactions': 'transactions'}
#: tmx_all_* 的加数。Alpha-X / Alpha-DRK 2023-11 才有值 → 回补段自然只有前三家。
GROUP = ['Toronto Stock Exchange', 'TSX Venture Exchange', 'Alpha', 'Alpha-X', 'Alpha-DRK']
#: 这两列在 CIRO 与 TMX 两个独立源里都恰好从这个月起才有值 —— 本脚本每次复验。
ALPHAX_FIRST = '2023-11'

#: 「同一套数」的 6 列：重叠月中位比值必须贴着 1，否则说明 CIRO 换了口径。
SAME_CALIBER = ['tsx_transactions', 'tsxv_volume_shares', 'tsxv_value_cad',
                'tsxv_transactions', 'alpha_volume_shares', 'alpha_transactions']
SAME_TOL = 1e-3
#: 「有系统性差」的 5 列：中位比值必须落在这个带内。跑出带外 = 源变了，停下来看。
DRIFT_BAND = {'tsx_volume_shares': (0.97, 1.00),
              'tsx_value_cad': (0.99, 1.02),
              'alpha_value_cad': (0.99, 1.02),
              'tmx_all_volume_shares': (0.98, 1.01),
              'tmx_all_value_cad': (0.99, 1.02)}
#: 断点线要不要画的门槛（接缝台阶的绝对值，%）。见 docstring「口径」一节。
BREAK_PCT = 0.5


class CiroError(RuntimeError):
    pass


def _load_fetch():
    """按路径加载 fetch/tmx.py —— CSV 列定义、`_merge`、写盘全部复用它，不另写一份。

    复用而不是复制：CSV 的两条铁律（已有值永不覆盖 / 未触碰的格子字节级不变）
    只该有一份实现，否则总有一天两份会分叉。
    """
    spec = importlib.util.spec_from_file_location(
        'fetch_tmx', os.path.join(ROOT, 'fetch', 'tmx.py'))
    mod = importlib.util.module_from_spec(spec)
    sys.modules['fetch_tmx'] = mod
    spec.loader.exec_module(mod)
    return mod


F = _load_fetch()

if F.SPOT_START != CTS_FROM:                # 两处常量必须一致，见 CTS_FROM 那段注释
    raise CiroError('fetch/tmx.py 的 SPOT_START=%s 与本脚本的 CTS_FROM=%s 不一致 —— '
                    '换源边界只能有一个' % (F.SPOT_START, CTS_FROM))


# ─────────────────────────────── 取工作簿 ───────────────────────────────
def _get(url, timeout=180):
    """带指数退避的 GET。WAF 的 403 是**间歇**的，重试到底再判失败。"""
    last = None
    for i in range(RETRIES):
        req = urllib.request.Request(url, headers={
            'User-Agent': ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                           'AppleWebKit/537.36 (KHTML, like Gecko) '
                           'Chrome/126.0.0.0 Safari/537.36'),
            'Accept': '*/*',
            'Accept-Language': 'en-CA,en;q=0.9',
        })
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read(), r.headers
        except urllib.error.HTTPError as e:
            last = 'HTTP %s' % e.code
            if e.code not in (403, 429, 502, 503):
                break
        except Exception as e:                              # noqa: BLE001
            last = repr(e)
        time.sleep(1.5 * (i + 1))
    raise CiroError('%s 重试 %d 次仍失败（最后一次：%s）—— '
                    '这**不等于**文件没了，先手工开一下落地页' % (url, RETRIES, last))


def discover():
    """从落地页现抓「历史报」直链。media id 会漂，绝不写死。"""
    html, _h = _get(LANDING)
    html = html.decode('utf-8', 'replace')
    hits = []
    for m in re.finditer(r'<a[^>]+href="([^"]*?/media/\d+/download[^"]*)"[^>]*>(.*?)</a>',
                         html, re.S | re.I):
        text = re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', '', m.group(2))).strip()
        if HIST_MARK.lower() in text.lower():
            hits.append((m.group(1), text))
    if len(hits) != 1:
        raise CiroError('落地页上锚文本含 %r 的 media 链接有 %d 条（期待恰好 1 条）：%s'
                        % (HIST_MARK, len(hits), [t for _u, t in hits]))
    url, text = hits[0]
    if url.startswith('/'):
        url = 'https://www.ciro.ca' + url
    print('· 历史报直链 %s\n  锚文本「%s」' % (url, text))
    return url


def download(cache_dir, refresh=False):
    os.makedirs(cache_dir, exist_ok=True)
    path = os.path.join(cache_dir, FNAME)
    if os.path.exists(path) and not refresh and os.path.getsize(path) >= MIN_BYTES:
        print('· 复用已下载的 %s（%d 字节）' % (path, os.path.getsize(path)))
        return path
    blob, headers = _get(discover())
    if len(blob) < MIN_BYTES:
        raise CiroError('只回了 %d 字节（<%d），不像那份全历史工作簿'
                        % (len(blob), MIN_BYTES))
    with open(path, 'wb') as f:
        f.write(blob)
    print('· 下载 %d 字节 → %s\n  Content-Disposition: %s'
          % (len(blob), path, headers.get('Content-Disposition')))
    return path


# ─────────────────────────────── 解析 ───────────────────────────────────
def parse(path):
    """→ {'YYYY-MM': {CIRO 列名|指标: 值}}。结构不符一律抛，不猜。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        raise CiroError('工作簿里缺这些 sheet：%s（现有 %s）' % (missing, wb.sheetnames))

    out = {}
    for sheet, metric in SHEETS.items():
        ws = wb[sheet]
        col = None
        seen_labels = set()
        for row in ws.iter_rows(values_only=True):
            if col is None:
                # 表头行不写死行号：扫到第一列是 'Month' 的那一行为止
                if row and str(row[0]).strip() == 'Month':
                    hdr = [str(c).strip() if c is not None else '' for c in row]
                    col = {n: i for i, n in enumerate(hdr)}
                    absent = [v for v in GROUP if v not in col]
                    if absent:
                        raise CiroError('%s 的表头里没有这些市场：%s' % (sheet, absent))
                continue
            if not row or row[0] is None:
                continue
            label = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            seen_labels.add(label)
            if label != ROW_WANTED:
                continue
            mm = MONTH_RE.match(str(row[0]).strip())
            if not mm:
                raise CiroError('%s 里认不出月份格式：%r' % (sheet, row[0]))
            month = '%s-%02d' % (mm.group(1), _MON[mm.group(2)])
            rec = out.setdefault(month, {})
            for ven in GROUP:
                v = row[col[ven]]
                if v in (None, ''):
                    continue
                rec[ven + '|' + metric] = float(str(v).replace(',', ''))
        if col is None:
            raise CiroError('%s 里找不到表头行（第一列 = "Month"）' % sheet)
        if ROW_WANTED not in seen_labels:
            raise CiroError('%s 里一行 %r 都没有（现有行类型 %s）'
                            % (sheet, ROW_WANTED, sorted(seen_labels)[:6]))

    months = sorted(out)
    want = [months[0]] + F._month_range(months[0], months[-1])
    holes = [m for m in want if m not in out]
    if holes:
        raise CiroError('CIRO 月份序列中间有洞：%s' % holes[:6])
    # 每个月三张表都要齐 —— 少一张就会写出一行「有量没额」的残行
    for m in months:
        for ven in ('Toronto Stock Exchange', 'TSX Venture Exchange', 'Alpha'):
            for metric in SHEETS.values():
                if ven + '|' + metric not in out[m]:
                    raise CiroError('%s 缺 %s 的 %s' % (m, ven, metric))
    # Alpha-X / Alpha-DRK 的起点断言（见 docstring「不做的事」）
    for ven in ('Alpha-X', 'Alpha-DRK'):
        got = sorted(m for m in months if ven + '|volume' in out[m])
        if not got or got[0] != ALPHAX_FIRST:
            raise CiroError('%s 在 CIRO 里从 %s 起才有值，与已知的 %s 不符'
                            % (ven, got[0] if got else '(无)', ALPHAX_FIRST))
    print('· 解析出 %d 个月（%s → %s），逐月无洞；Alpha-X / Alpha-DRK 均自 %s 起'
          % (len(months), months[0], months[-1], ALPHAX_FIRST))
    return out


def to_rows(data):
    """CIRO 记录 → {'YYYY-MM': {CSV 列: 值}}，只含 CTS_FROM 之前的月份。"""
    rows = {}
    for month in sorted(data):
        if month >= CTS_FROM:
            continue
        rec, d = {}, data[month]
        for pre, ven in VENUE.items():
            for suf, metric in METRIC.items():
                rec['%s_%s' % (pre, suf)] = d[ven + '|' + metric]
        for suf, metric in METRIC.items():
            rec['tmx_all_' + suf] = sum(d[v + '|' + metric] for v in GROUP
                                        if v + '|' + metric in d)
        rows[month] = rec
    return rows


# ─────────────────────────────── 核对 ───────────────────────────────────
def _csv_rows():
    path = os.path.join(SERIES_DIR, 'tmx.csv')
    with open(path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    header = rows[0]
    return path, header, {r[0]: dict(zip(header, r)) for r in rows[1:] if r and r[0].strip()}


def _ciro_col(data, month, col):
    """CSV 列名 → 该月的 CIRO 值（含 tmx_all_* 的加总）。取不到返回 None。"""
    d = data.get(month)
    if not d:
        return None
    if col.startswith('tmx_all_'):
        metric = METRIC[col[len('tmx_all_'):]]
        parts = [d.get(v + '|' + metric) for v in GROUP]
        if any(parts[i] is None for i in range(3)):
            return None
        return sum(p for p in parts if p is not None)
    pre, suf = col.split('_', 1)
    ven, metric = VENUE.get(pre), METRIC.get(suf)
    if ven is None or metric is None:
        return None
    return d.get(ven + '|' + metric)


def check_overlap(data, have):
    """重叠月逐列对表。**这是本脚本的口径看门狗**，不是装饰。"""
    cols = ([f'{p}_{s}' for p in VENUE for s in METRIC]
            + ['tmx_all_' + s for s in METRIC])
    print('\n── 重叠月对表（TMX 自报 ÷ CIRO）%s 起 ──' % CTS_FROM)
    print('%-24s %4s %5s %9s %9s %9s' % ('列', 'n', '逐位同', '中位', '最小', '最大'))
    meds = {}
    for col in cols:
        rr = []
        for month, row in sorted(have.items()):
            if month < CTS_FROM or not (row.get(col) or '').strip():
                continue
            c = _ciro_col(data, month, col)
            if not c:
                continue
            rr.append((month, float(row[col]), c))
        if not rr:
            raise CiroError('%s 一个重叠月都没有 —— 无从判断口径' % col)
        ratios = [a / b for _m, a, b in rr]
        ident = sum(1 for _m, a, b in rr if a == b)
        meds[col] = statistics.median(ratios)
        lo = min(rr, key=lambda t: t[1] / t[2])
        hi = max(rr, key=lambda t: t[1] / t[2])
        print('%-24s %4d %5d %9.5f %9.5f %9.5f   (min@%s max@%s)'
              % (col, len(rr), ident, meds[col], lo[1] / lo[2], hi[1] / hi[2],
                 lo[0], hi[0]))

    bad = [(c, meds[c]) for c in SAME_CALIBER if abs(meds[c] - 1.0) > SAME_TOL]
    if bad:
        raise CiroError('本该同口径的列偏了：%s（容差 %s）—— CIRO 改口径了，先查清再写'
                        % (bad, SAME_TOL))
    out = [(c, meds[c], b) for c, b in DRIFT_BAND.items() if not b[0] <= meds[c] <= b[1]]
    if out:
        raise CiroError('系统性差的列跑出已知带宽：%s —— 停下来看，别写' % out)
    print('· %d 条同口径列全部贴着 1.0（容差 %s）；%d 条有差的列都在已知带内'
          % (len(SAME_CALIBER), SAME_TOL, len(DRIFT_BAND)))
    return meds


def check_seam(data, rows, have):
    """接缝逐列体检：CIRO 的最后一个月 vs 既有数据的第一个月。**只报告，不修改**。

    三个环比要分开看，否则会把口径台阶当成业务变化（或者反过来）：
      · 拼接 m/m —— 读者在图上真正看到的那个环比（左 CIRO、右 TMX）；
      · 真   m/m —— 两个月都取 CIRO，即**同一把尺子**量出来的环比；
      · 口径台阶 —— 前两者之差，纯粹是换源造出来的，一分钱业务都不含。
    """
    left = max(rows)                     # 2021-07
    right = CTS_FROM                     # 2021-08
    print('\n── 接缝 %s（CIRO） → %s（TMX 自报）──' % (left, right))
    print('%-24s %20s %20s %9s %9s %9s'
          % ('列', 'CIRO ' + left, 'TMX ' + right, '拼接m/m', '真m/m', '口径台阶'))
    flagged = []
    for col in sorted(rows[left]):
        a = rows[left][col]
        b = float(have[right][col])
        c = _ciro_col(data, right, col)
        if not c:
            raise CiroError('%s 在 %s 没有 CIRO 值，接缝无从分解' % (col, right))
        step = (b / c - 1) * 100
        print('%-24s %20.3f %20.3f %+8.2f%% %+8.2f%% %+8.2f%%'
              % (col, a, b, (b / a - 1) * 100, (c / a - 1) * 100, step))
        if abs(step) >= BREAK_PCT:
            flagged.append((col, step))
    print('· 台阶 ≥%.1f%% 的列（build/specs/tmx.py 应给它们画 %s 断点线）：%s'
          % (BREAK_PCT, right,
             '、'.join('%s %+.2f%%' % (c, s) for c, s in flagged) or '（无）'))
    return flagged


def write(rows, dry=False):
    csv_path = os.path.join(SERIES_DIR, 'tmx.csv')
    header, body, have = F._load_csv(csv_path, F.COLUMNS, 'month')
    idx = {n: i for i, n in enumerate(header)}
    over = [m for m in rows if m >= CTS_FROM]
    if over:                             # 双保险：to_rows 已经拦过一次
        raise CiroError('这些月份 ≥ %s，本脚本不许写：%s' % (CTS_FROM, over[:6]))
    would = sum(1 for m in rows for c in rows[m]
                if m not in have or not have[m][idx[c]].strip())
    if dry:
        print('\n--dry：会新建 %d 行、填 %d 格，**未**写入'
              % (len([m for m in rows if m not in have]), would))
        for m in sorted(rows)[:2] + ['…'] + sorted(rows)[-1:]:
            if m == '…':
                print('    …')
                continue
            print('    %s  %s' % (m, {k: round(v, 3) for k, v in sorted(rows[m].items())}))
        return 0
    added = [m for m in sorted(rows) if F._merge(header, body, have, m, rows[m])]
    F._write_csv(csv_path, header, body)
    print('\n✓ 新建 %d 行、填 %d 格 → %s（现共 %d 行，%s → %s）'
          % (len(added), would, csv_path, len(body), body[0][0], body[-1][0]))
    return len(added)


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry', action='store_true')
    ap.add_argument('--refresh', action='store_true')
    ap.add_argument('--cache-dir', default=os.path.join(ROOT, 'cache', 'tmx'),
                    help='原始 xlsx 落在哪（默认 <repo>/cache/tmx）')
    a = ap.parse_args(argv)

    path = download(a.cache_dir, a.refresh)
    data = parse(path)
    _p, _hdr, have = _csv_rows()
    check_overlap(data, have)
    rows = to_rows(data)
    check_seam(data, rows, have)
    write(rows, a.dry)
    return 0


if __name__ == '__main__':
    sys.exit(main())
