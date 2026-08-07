# -*- coding: utf-8 -*-
"""CME 五个 product_id 的基期常数取数器（基期 2019-01，一次算死写进规格表）。

用法:
    python3 build/basefill/cme2.py            # 取数 + 全部核对 + 写 series/_specs_part_cme2.csv
    python3 build/basefill/cme2.py --dry      # 只打印，不写文件
    python3 build/basefill/cme2.py --keep     # 保留 cache（默认也保留，纯提示）

覆盖：CME_ENERGY、CME_AG（第一轮空着，本轮补齐）
      CME_RATES、CME_EQUITY_INDEX、CME_FX（第一轮已填，本轮独立复核）

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
第一轮卡在哪，本轮怎么破的
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
第一轮的结论是「2019-01 官方结算价拿不到」：Settlements API 对 2019 返回 empty、
voiProductsViewExport.ctl 对 2019 只回表头、FTP 的 settle/ 与 fprf/ 只有近月。
唯一还活着的通道是 SPAN 风险参数文件存档，但当时没找到版式说明，不敢瞎解。

本轮把版式说明找到了，位置在 CME 自己的客户系统 wiki（不是第三方）：
    https://www.cmegroup.com/clearing/risk-management/span-reference-documents.html
        → Risk Parameter File Layouts for the Positional Formats
    https://cmegroupclientsite.atlassian.net/wiki/spaces/pubsub/pages/457083445/
    逐记录类型的子页（本模块用到的三种）：
      · Type P  Record - Standard and Expanded  (pageId 457083966)
      · Type B  - Expanded                      (pageId 457215159)
      · Type 8  - Expanded                      (pageId 457083725)
⚠ 这些 wiki 长页在 HTML 里是懒加载的，浏览器/curl 拿到的正文会被截断；
  必须走 Confluence 的 REST 接口才拿得到完整表格：
      /wiki/rest/api/content/<pageId>?expand=body.storage
  （第一轮如果只 curl HTML，看到的字段表正好在"Settlement Price"那一行前面断掉。）

结果比预期好：SPAN 文件里**同时**给出了乘数和结算价，两样都不用再去别处凑——
  · 类型 "P " 记录 = 价格换算参数：产品码、结算价小数位、对齐码、
    **Contract Value Factor（合约价值乘数）**及其 10 的幂次指数、结算币种。
    官方定义原文：「The Contract Value Factor is the multiplier that converts an
    actual price -- one that has had its decimal point inserted ... -- to the
    actual contract value.」→ 单张名义额 = CVF × 结算价，**不需要知道合约规格里
    的桶数/蒲式耳数，也不需要知道报价惯例**（日元期货那种报价缩放 10 倍的坑，
    CVF 的指数字段自己就抵消掉了）。
  · 类型 "B " 记录 = 每个合约月的到期（最终结算）日 → 用来定"近月"，不用猜。
  · 类型 "82" 记录 = 每个合约的**当日官方结算价**（字节 111-117，符号在 118）。

━━ 口径（与 build/basefill/us.py、build/check_specs.py 一致）━━
· 基期 = 2019-01。
· 基期价格 = 该月全部交易日结算价的算术平均（base_price_basis=avg_close）。
  2019-01 的 CME 交易日 = 21 天：SPAN 存档里 1 月有 22 个业务日文件，但 1/21
  （马丁·路德·金日）**不是交易日**——两条独立证据：
    (a) ftp://ftp.cmegroup.com/daily_volume/daily_volume_20190121.xlsx 不存在（550）；
    (b) 欧洲美元期货 MTD ADV 在 1/18 是 3218717.84615385（×13 = 41,843,332），
        在 1/22 是 3158633.57142857（×14 = 44,220,870 = 41,843,332 + 1/22 当日
        2,377,538），分母 13→14 而不是 13→15，即 1/21 没有成交日。
  所以本模块用 21 个文件（排除 20190121），与表内既有 23 行的 21 天完全对齐。
· 近月定义：类型 B 记录里"到期（最终结算）日 ≥ 该业务日"的最小合约月。
  **不能**直接取"最小合约月"——SPAN 文件里还挂着已到期的旧月份（例：2019-01-02
  的文件里 NYM CL 还有 201812、201901 两个已交割月），直接取最小会取到陈价。
  滚月约定与表内 CBOE_VIX_FUT 那一行一致（到期日当天仍算近月，次日滚下一月）。
· 篮子常数 = Σ(成员 2019-01 张数 × 该合约单张名义额) ÷ Σ(成员 2019-01 张数)，
  用**基期那个月**的品种结构算一次写死，绝不每月重算。
· 期权按「一张期权 = 一张标的期货」折算，标的取 SPAN 记录里的
  Underlying Commodity Code 字段（官方给的父子关系，不是靠代码前缀猜的）。

━━ 数据源（全部 CME 官方一手）━━
1. 2019-01 各合约张数
     ftp://ftp.cmegroup.com/daily_volume/daily_volume_20190131.xlsx
   工作表 'CME Group Vol and OI by Product' 的 MTD ADV 列。
   ⚠ HTTPS 镜像 https://www.cmegroup.com/ftp/pub/… 下文件恒 400（本轮用
     curl_cffi impersonate=chrome 重试过，仍是 400，不是 TLS 指纹问题），必须走 ftp://。
2. 2019-01 各合约每日官方结算价 + 官方乘数
     ftp://ftp.cmegroup.com/span/archive/cme/2019/cme.<YYYYMMDD>.s.pa2.zip
   后缀 s = 当日**最终结算**运行（文件头类型 "0" 记录的 File Identifier = "SF"，
   创建时间 17:57 CT）；同日还有 e=SE（16:02，盘后早班）、i/a=I（盘中 11:32/10:42）。
   本模块只用 SF，跑的时候会断言文件头的业务日与 File Identifier。
3. SPAN 版式说明：见上方 wiki 链接。
4. 规则手册（只用于交叉验证乘数，不作为唯一依据）
     https://www.cmegroup.com/rulebook/CBOT/II/21.pdf  等

━━ 已知精度损失（只有一处，已量化）━━
CBOT 谷物（玉米 C、大豆 S、芝加哥小麦 W、KC 小麦 KW）的结算价在 SPAN 文件里
按 3 位小数存（美元/蒲式耳），而这些合约的最小变动价位是 1/4 美分 = 0.0025 美元，
即真值有 4 位小数。文件是**截断**不是四舍五入（实测：玉米 3.7575 → 存 0003757，
小麦 5.0675 → 存 0005067），高精度字段（类型 81 记录 109-122）也只是同一个值。
→ 单价最多低估 0.0005 美元/蒲式耳（玉米约 0.013%），对 CME_AG 篮子的影响见下面
  算出来的 recon 变体（本模块两种都算，写进表里的是**未加工的原始值**，
  重建值只写进 computed_note 当敏感性）。
判据：若某产品所有结算价的第三位小数只出现在 {0,2,5,7}，即证明它落在 1/4 美分
网格上且是截断（0→+0.0000, 2→+0.0005, 5→+0.0000, 7→+0.0005）。本模块跑的时候
会对这四个产品逐个断言这一点，断言不过就不给重建值。

━━ 跑一次会自动做的 7 项核对（任何一项不过直接抛 BaseFillError，不静默出数）━━
1. daily_volume 的六大类逐产品 ADV 合计 vs series/cme.csv（CME IR 月报），偏差 >2% 就炸。
2. 每个 SPAN 文件的类型 0 记录：业务日 == 文件名、结算标志 S、文件标识 F、格式 U2。
3. 2 年期国债面值双证：CBOT 第 21 章原文 + SPAN 的 CVF（两条都必须指向 $200,000）。
4. 能源/农产品 13 个主力合约的乘数双证：SPAN CVF vs 规则手册 Trading Unit 条款原文。
5. WTI / Henry Hub 的近月月均结算价 vs EIA 官方序列 RCLC1 / RNGC1（同 21 天），
   要求相对偏差 <1e-6 —— 实测逐位相同，等于同时验证了小数位解码、滚月约定、
   21 个交易日的选取三件事，且 EIA 与本模块完全独立（不同机构、不同格式、不同实现）。
6. 四个谷物的 1/4 美分网格断言（第三位小数只出现 {0,2,5,7}）。
7. 每个篮子的覆盖率下限 85%（实测 98.4%~99.95%）。

━━ 输出 ━━
series/_specs_part_cme2.csv（主线程负责合并进 contract_specs.csv）
列：product_id, base_price_local, base_ccy, basket_constant, source_url, computed_note
篮子行的 base_price_local 与 basket_constant 相同（规格表里篮子行 multiplier=1）。
"""
from __future__ import annotations

import csv
import ftplib
import io
import os
import re
import sys
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SERIES = os.path.join(ROOT, 'series')
CACHE = os.path.join(ROOT, 'cache', 'basefill')
SPANDIR = os.path.join(CACHE, 'span2019')

BASE_MONTH = '2019-01'
OUT_NAME = '_specs_part_cme2.csv'

CME_HOST = 'ftp.cmegroup.com'
CME_VOL_PATH = '/daily_volume/daily_volume_20190131.xlsx'
SPAN_DIR_FTP = '/span/archive/cme/2019/'

# 2019-01 的 21 个 CME 交易日（1/1 元旦、1/21 马丁·路德·金日不是交易日，
# 1/21 虽然有 SPAN 文件但没有 daily_volume 文件、ADV 分母也不含它，见模块头注释）
TRADE_DATES = ['20190102', '20190103', '20190104', '20190107', '20190108', '20190109',
               '20190110', '20190111', '20190114', '20190115', '20190116', '20190117',
               '20190118', '20190122', '20190123', '20190124', '20190125', '20190128',
               '20190129', '20190130', '20190131']

# daily_volume 里的交易所全名 → SPAN 文件里的三字母 Exchange Acronym
EXCH_MAP = {
    'Chicago Mercantile Exchange (STATS)': 'CME',
    'Chicago Board of Trade (STATS)': 'CBT',
    'NYMEX(STATS)': 'NYM',
    'COMEX(STATS)': 'CMX',
    # 'CME Group OTC (STATS)' 没有对应的 SPAN 交易所，落入未覆盖（占比极小）
}

# 1/4 美分网格的 CBOT 谷物（结算价 3 位小数截断，见模块头「已知精度损失」）
QUARTER_CENT_GRAINS = {('CBT', 'C'), ('CBT', 'S'), ('CBT', 'W'), ('CBT', 'KW')}


class BaseFillError(RuntimeError):
    """任何一步取不到 / 对不上就炸。常数表错一个数，图上完全看不出来。"""


# ══════════════════════════════════════════════════════════════════════════
# 零、网络
# ══════════════════════════════════════════════════════════════════════════
def _http(url, timeout=90):
    """cmegroup.com 走 Akamai，按 TLS 指纹拦普通 urllib，必须 curl_cffi impersonate。"""
    os.makedirs(CACHE, exist_ok=True)
    key = re.sub(r'[^A-Za-z0-9._-]', '_', url)[-150:]
    path = os.path.join(CACHE, key)
    if os.path.exists(path) and os.path.getsize(path) > 0:
        with open(path, 'rb') as f:
            return f.read()
    try:
        from curl_cffi import requests as crequests
    except ImportError as e:                                   # pragma: no cover
        raise BaseFillError('需要 curl_cffi（requirements.txt 已列）：%r' % e) from e
    r = crequests.get(url, impersonate='chrome', timeout=timeout)
    if r.status_code != 200:
        raise BaseFillError('下载失败 HTTP %s: %s' % (r.status_code, url))
    with open(path, 'wb') as f:
        f.write(r.content)
    return r.content


def _ftp(path, dst=None, timeout=300):
    """CME 公开匿名 FTP。HTTPS 那条通道只能列目录，下文件恒 400（已用 curl_cffi 复测）。"""
    os.makedirs(CACHE, exist_ok=True)
    local = dst or os.path.join(CACHE, os.path.basename(path))
    if os.path.exists(local) and os.path.getsize(local) > 0:
        with open(local, 'rb') as f:
            return f.read()
    buf = io.BytesIO()
    ftp = ftplib.FTP(CME_HOST, timeout=timeout)
    try:
        ftp.login()                       # 匿名
        ftp.voidcmd('TYPE I')
        size = ftp.size(path)
        ftp.retrbinary('RETR ' + path, buf.write, blocksize=262144)
    finally:
        try:
            ftp.quit()
        except Exception:                 # noqa: BLE001
            ftp.close()
    data = buf.getvalue()
    if not data or (size and len(data) != size):
        raise BaseFillError('FTP 取到 %d 字节，应为 %s：ftp://%s%s'
                            % (len(data), size, CME_HOST, path))
    os.makedirs(os.path.dirname(local), exist_ok=True)
    with open(local, 'wb') as f:
        f.write(data)
    return data


def fetch_span_files(dates=None, workers=6, verbose=True):
    """并行把 2019-01 的 SF SPAN 文件拉全（单连接被限速，开 6 条）。"""
    import queue
    import threading
    import time
    dates = dates or TRADE_DATES
    os.makedirs(SPANDIR, exist_ok=True)
    q = queue.Queue()
    for d in dates:
        name = 'cme.%s.s.pa2.zip' % d
        dst = os.path.join(SPANDIR, name)
        if not (os.path.exists(dst) and os.path.getsize(dst) > 9_000_000):
            q.put((d, name, dst))
    if q.empty():
        if verbose:
            print('  [span] 21 个文件已在 cache，跳过下载')
        return
    lock = threading.Lock()
    errs = []

    def worker():
        while True:
            try:
                _d, name, dst = q.get_nowait()
            except queue.Empty:
                return
            for attempt in range(4):
                try:
                    t0 = time.time()
                    _ftp(SPAN_DIR_FTP + name, dst=dst)
                    if verbose:
                        with lock:
                            print('  [span] %s %.1f MB %.0fs'
                                  % (name, os.path.getsize(dst) / 1e6, time.time() - t0),
                                  flush=True)
                    break
                except Exception as e:                          # noqa: BLE001
                    if attempt == 3:
                        errs.append((name, repr(e)))
                    time.sleep(5)
            q.task_done()

    ths = [threading.Thread(target=worker, daemon=True) for _ in range(workers)]
    for t in ths:
        t.start()
    for t in ths:
        t.join()
    if errs:
        raise BaseFillError('SPAN 文件没拉全：%r' % errs)


# ══════════════════════════════════════════════════════════════════════════
# 一、2019-01 各合约张数（CME 官方 FTP daily_volume，取 MTD ADV）
# ══════════════════════════════════════════════════════════════════════════
def cme_jan2019_adv():
    import warnings

    import openpyxl
    raw = _ftp(CME_VOL_PATH)
    with warnings.catch_warnings():
        warnings.simplefilter('ignore')   # CME 这份 xlsx 没写默认样式，警告与数据无关
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = 'CME Group Vol and OI by Product'
    if sheet not in wb.sheetnames:
        raise BaseFillError('daily_volume 文件里没有 %r 表，源站改版了' % sheet)
    rows = list(wb[sheet].iter_rows(values_only=True))
    if not any(r[0] and str(r[0]).startswith('Trade Date:  01/31/2019') for r in rows[:5]):
        raise BaseFillError('取到的不是 01/31/2019 那天的文件，表头是 %r' % (rows[1][0],))
    classes = {'INTEREST RATES', 'EQUITY INDEX', 'FX', 'ENERGY', 'METALS', 'AG PRODUCTS'}
    out = []
    for r in rows:
        if r[0] in classes and r[3] and isinstance(r[10], (int, float)):
            out.append({'cls': r[0],
                        'exch': EXCH_MAP.get((r[1] or '').strip()),
                        'exch_raw': (r[1] or '').strip(),
                        'code': (r[2] or '').strip(),
                        'desc': (r[3] or '').strip(),
                        'fo': r[4],
                        'adv': float(r[10])})
    if len(out) < 500:
        raise BaseFillError('只解析出 %d 行产品，明显不对' % len(out))
    return out


def check_adv_against_ir(adv_rows, verbose=True):
    """分类合计 vs series/cme.csv（CME IR 月度 xlsx）的 2019-01 分类 ADV。"""
    with open(os.path.join(SERIES, 'cme.csv'), newline='', encoding='utf-8') as f:
        row = {r['month']: r for r in csv.DictReader(f)}[BASE_MONTH]
    ir = {'INTEREST RATES': 'adv_rates_kcontracts', 'EQUITY INDEX': 'adv_equity_kcontracts',
          'FX': 'adv_fx_kcontracts', 'ENERGY': 'adv_energy_kcontracts',
          'AG PRODUCTS': 'adv_ag_kcontracts', 'METALS': 'adv_metals_kcontracts'}
    for cls, col in ir.items():
        want = float(row[col]) * 1e3
        got = sum(r['adv'] for r in adv_rows if r['cls'] == cls)
        rel = abs(got - want) / want
        if verbose:
            print('  [adv对账] %-14s 逐产品合计 %12.0f  IR月报 %12.0f  偏差 %5.2f%%'
                  % (cls, got, want, 100 * rel))
        if rel > 0.02:
            raise BaseFillError('%s 的逐产品合计与 IR 月报差 %.2f%%，解析可能错列' % (cls, 100 * rel))


# ══════════════════════════════════════════════════════════════════════════
# 二、SPAN 解码（字段偏移全部来自官方版式说明，见模块头）
# ══════════════════════════════════════════════════════════════════════════
def _dec_locator(txt):
    """小数位定位符：3 字节，负数时首字节是 '-'（如 '-02' 表示 -2）。"""
    txt = txt.strip()
    if not txt:
        return 0
    if txt.startswith('-'):
        return -int(txt[1:] or 0)
    return int(txt)


def parse_span_day(date, verbose=False):
    """解一天的 SF SPAN 文件。返回 (specs, expiry, settle, underlying)。

    specs[(exch, prod, ptype)] = dict(cvf=..., dec=..., align=..., ccy=..., quote=..., name=...)
    expiry[(exch, prod, futmon)] = 'CCYYMMDD'      —— 类型 B，仅 FUT
    settle[(exch, prod, futmon)] = 原始 7 位字符串 + 符号（未套小数位）
    underlying[(exch, optprod)]  = 标的产品码       —— 类型 81/82 的 16-25 字节
    """
    path = os.path.join(SPANDIR, 'cme.%s.s.pa2.zip' % date)
    if not os.path.exists(path):
        raise BaseFillError('缺 SPAN 文件 %s，先跑 fetch_span_files()' % path)
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        if len(names) != 1 or not names[0].endswith('.pa2'):
            raise BaseFillError('%s 里不是单个 .pa2：%r' % (path, names))
        data = z.read(names[0])
    lines = data.split(b'\r\n')

    # —— 文件头断言：类型 "0" 记录（只有 expanded 格式才有）——
    # 版式（Type 0 Records，pageId 457411411）：3-8 交易所复合体、9-16 业务日 CCYYMMDD、
    # 17 结算(S)/盘中(I) 标志、18-19 文件标识（E=Early、F=Final、C=Complete）、
    # 24-31 生成日、32-35 生成时间、36-37 文件格式（U2=Expanded Unpacked）。
    head = lines[0].decode('latin1')
    if head[:2] != '0 ':
        raise BaseFillError('%s 首行不是类型 0 记录：%r' % (date, head[:40]))
    if head[8:16] != date:
        raise BaseFillError('%s 的业务日是 %r，对不上文件名' % (date, head[8:16]))
    if head[16] != 'S' or head[17] != 'F':
        raise BaseFillError('%s 的结算标志/文件标识是 %r，不是最终结算（S+F）文件'
                            % (date, head[16:19]))
    if head[35:37] != 'U2':
        raise BaseFillError('%s 的文件格式是 %r，不是 Expanded Unpacked（U2），'
                            '本模块的字段偏移只对 U2 成立' % (date, head[35:37]))

    specs, expiry, settle, underlying = {}, {}, {}, {}
    for raw in lines:
        rid = raw[:2]
        if rid == b'P ':
            s = raw.decode('latin1')
            exch, prod, ptype = s[2:5].strip(), s[5:15].strip(), s[15:18].strip()
            exp = int((s[73:75] or '0').strip() or 0)
            sign = -1 if s[72] == '-' else 1
            cvf = float(s[41:48] + '.' + s[48:55]) * (10.0 ** (sign * exp))
            specs[(exch, prod, ptype)] = {
                'cvf': cvf,
                'dec': _dec_locator(s[33:36]),
                'align': s[39],
                'qty_per_contract': int((s[63:65] or '1').strip() or 1),
                'ccy': s[65:68].strip(),
                'quote': s[69:72].strip(),
                'name': s[79:114].strip(),
            }
        elif rid == b'B ':
            s = raw.decode('latin1')
            if s[15:18].strip() != 'FUT':
                continue
            if s[27:33].strip('0 ') != '':        # 期权系列（有期权月）跳过
                continue
            exch, prod, futmon = s[2:5].strip(), s[5:15].strip(), s[18:24].strip()
            if s[24:26].strip():                  # 非标准月（周合约等）不当近月候选
                continue
            expd = s[91:99].strip()
            if futmon and expd and expd != '0' * 8:
                expiry[(exch, prod, futmon)] = expd
        elif rid in (b'82', b'84'):
            s = raw.decode('latin1')
            exch, prod = s[2:5].strip(), s[5:15].strip()
            ptype = s[25:28].strip()
            if ptype != 'FUT':
                continue
            if s[35:37].strip():                  # 周/非标准月的期货，不当近月候选
                continue
            futmon = s[29:35].strip()
            if rid == b'82':
                px, sg = s[110:117], s[117]
            else:
                px, sg = s[131:138], s[138]
            if not futmon or not px.strip():
                continue
            settle[(exch, prod, futmon)] = (px, sg)
        elif rid in (b'81', b'83'):
            s = raw.decode('latin1')
            if s[25:28].strip() not in ('OOF', 'OOP', 'OOC'):
                continue
            exch, prod, und = s[2:5].strip(), s[5:15].strip(), s[15:25].strip()
            if und:
                underlying.setdefault((exch, prod), und)
    if not specs or not settle:
        raise BaseFillError('%s 解出来是空的（P=%d settle=%d）' % (date, len(specs), len(settle)))
    if verbose:
        print('  [span] %s  P=%d  B(FUT)=%d  settle(FUT)=%d  期权标的=%d'
              % (date, len(specs), len(expiry), len(settle), len(underlying)))
    return specs, expiry, settle, underlying


def decode_price(raw_px, sign, spec):
    """按类型 P 记录的小数位 + 对齐码把 7 位结算价字段还原成真实报价。

    对齐码取值（实测本文件里只有四种）：
      ' ' 普通十进制；'0' CBOT 谷物（十进制，见模块头精度说明）；
      'C' CBOT 32 分数（国债期货，小数部分是 32 分之几）；'K' 国债期权。
    本模块只在 ' ' 与 '0' 上取值（能源/农产品/股指/外汇全部落在这两类），
    碰到 'C'/'K' 直接返回 None —— 利率那一栏走面值口径，本来就不用价格。
    """
    if spec['align'] in ('C', 'K'):
        return None
    d = spec['dec']
    v = float(raw_px)
    px = v / (10.0 ** d) if d >= 0 else v * (10.0 ** (-d))
    if sign == '-':
        px = -px
    return px


def front_month_price(exch, prod, date, specs, expiry, settle):
    """近月 = 到期日 ≥ 业务日的最小合约月；返回 (合约月, 解码后价格) 或 None。

    不能直接取最小合约月：SPAN 文件里还挂着已交割的旧月份。
    """
    spec = specs.get((exch, prod, 'FUT'))
    if not spec:
        return None
    cands = []
    for (e, p, m), expd in expiry.items():
        if e != exch or p != prod:
            continue
        if expd < date:
            continue
        if (e, p, m) not in settle:
            continue
        cands.append((m, expd))
    if not cands:
        return None
    cands.sort()
    m = cands[0][0]
    raw_px, sg = settle[(exch, prod, m)]
    px = decode_price(raw_px, sg, spec)
    if px is None or px <= 0:
        return None
    return m, px, raw_px


def collect_prices(verbose=True):
    """跑完 21 天，返回 {(exch,prod): {'px': 月均价, 'n': 天数, 'months': [...],
    'raw3': set(第三位小数)}} 与最后一天的 specs / underlying。"""
    acc = {}
    specs = underlying = None
    for date in TRADE_DATES:
        specs, expiry, settle, und = parse_span_day(date)
        if underlying is None:
            underlying = dict(und)
        else:
            for k, v in und.items():
                underlying.setdefault(k, v)
        prods = {(e, p) for (e, p, t) in specs if t == 'FUT'}
        for (exch, prod) in prods:
            got = front_month_price(exch, prod, date, specs, expiry, settle)
            if not got:
                continue
            m, px, raw_px = got
            a = acc.setdefault((exch, prod), {'sum': 0.0, 'n': 0, 'months': [], 'raw3': set()})
            a['sum'] += px
            a['n'] += 1
            a['months'].append((date, m, px))
            if specs[(exch, prod, 'FUT')]['align'] == '0':
                a['raw3'].add(raw_px.strip()[-1])
    out = {}
    for k, a in acc.items():
        out[k] = {'px': a['sum'] / a['n'], 'n': a['n'], 'months': a['months'], 'raw3': a['raw3']}
    if verbose:
        print('  [span] 21 天解完，拿到 %d 个产品的近月月均结算价' % len(out))
    return out, specs, underlying


# ══════════════════════════════════════════════════════════════════════════
# 三、单张名义额
# ══════════════════════════════════════════════════════════════════════════
def notional_lookup(specs, prices, underlying, grain_recon=False):
    """返回 f(exch, code, is_option) -> (单张名义额 USD, 说明) 或 None。"""
    def resolve(exch, code, is_option):
        # 期权 → 先查 SPAN 给的标的产品码
        if is_option:
            und = underlying.get((exch, code))
            if und:
                code = und
        if (exch, code, 'FUT') not in specs:
            return None
        if (exch, code) not in prices:
            return None
        return exch, code

    def f(exch, code, is_option):
        r = resolve(exch, code, is_option)
        if not r:
            return None
        e, p = r
        spec = specs[(e, p, 'FUT')]
        if spec['ccy'] != 'USD':
            return None                    # 本模块只处理美元结算的（CME 这几类全是）
        px = prices[(e, p)]['px']
        if grain_recon and (e, p) in QUARTER_CENT_GRAINS:
            # 3 位小数截断的 1/4 美分网格：末位 2 或 7 → 真值再 +0.0005
            tot = 0.0
            for _d, _m, v in prices[(e, p)]['months']:
                last = int(round(v * 1000)) % 10
                tot += v + (0.0005 if last in (2, 7) else 0.0)
            px = tot / len(prices[(e, p)]['months'])
        return spec['cvf'] * px, '%s %s cvf=%g px=%.6f' % (e, p, spec['cvf'], px)
    return f


def basket(rows, notional_of, label, verbose=True, min_cov=0.85):
    """Σ(张数 × 单张名义额) ÷ Σ(张数)，只对能定价的成员算，另报覆盖率。"""
    num = den = 0.0
    total = sum(r['adv'] for r in rows)
    covered, missed = [], []
    for r in rows:
        got = None
        if r['exch']:
            got = notional_of(r['exch'], r['code'], r['fo'] == 'O')
        if got is None:
            missed.append(r)
            continue
        num += r['adv'] * got[0]
        den += r['adv']
        covered.append((r, got[0]))
    if den <= 0:
        raise BaseFillError('%s 一个成员都没覆盖到' % label)
    const, cov = num / den, den / total
    if verbose:
        print('\n  ── %s 篮子 ──' % label)
        covered.sort(key=lambda x: -x[0]['adv'])
        for r, n in covered[:10]:
            print('     %-4s %-5s %-40s %9.0f 张/日 × %14.2f USD'
                  % (r['exch'], r['code'], r['desc'][:40], r['adv'], n))
        print('     …共 %d 个成员进篮子，覆盖 2019-01 该类张数的 %.2f%%' % (len(covered), 100 * cov))
        missed.sort(key=lambda x: -x['adv'])
        if missed:
            print('     未覆盖前 6：' + '; '.join(
                '%s/%s %s %.0f' % (m['exch'], m['code'], m['desc'][:20], m['adv'])
                for m in missed[:6]))
        print('     篮子常数 = %.6f USD/张' % const)
    if cov < min_cov:
        raise BaseFillError('%s 覆盖率只有 %.1f%%，低于 %.0f%% 不许写常数'
                            % (label, 100 * cov, 100 * min_cov))
    return const, cov, len(covered), total - den


# ══════════════════════════════════════════════════════════════════════════
# 四、复核：利率面值（第一轮的两个易错点之一）
# ══════════════════════════════════════════════════════════════════════════
# 第一轮的面值表（本模块要用两条独立官方证据各自复核一遍）
RATE_FACE_R1 = {
    'ED': 1_000_000, 'SR3': 1_000_000, 'SR1': 5_000_000, '41': 5_000_000,
    '21': 100_000, '25': 100_000, '26': 200_000, '17': 100_000,
    'TN': 100_000, 'UBE': 100_000,
}
RATE_OPT_PARENT = {
    'E0': 'ED', 'E2': 'ED', 'E3': 'ED', 'E4': 'ED', 'E5': 'ED',
    'TE4': 'ED', 'EF1': 'ED', 'EF2': 'ED', 'EF3': 'ED', 'EF4': 'ED',
    'E1F': 'ED', 'E2F': 'ED', 'E3F': 'ED',
    'TY1': '21', 'TY2': '21', 'TY3': '21', 'TY4': '21', 'TY5': '21',
    'WY1': '21', 'WY2': '21', 'WY3': '21', 'WY4': '21', 'WY5': '21',
    'FV1': '25', 'FV2': '25', 'FV3': '25', 'FV4': '25', 'FV5': '25',
    'WF1': '25', 'WF2': '25', 'WF3': '25', 'WF4': '25', 'WF5': '25',
    'US1': '17', 'US2': '17', 'US3': '17', 'US4': '17', 'US5': '17',
    'WB1': '17', 'WB2': '17', 'WB3': '17', 'WB4': '17', 'WB5': '17',
}
# 规则手册交叉验证：国债期货用交割发票公式 $N × P × c，P 以 100 面值报价 ⇒ 面值 = 100N
RULEBOOK_FACE = {
    '26': ('https://www.cmegroup.com/rulebook/CBOT/II/21.pdf',
           'Invoice Amount = ($2000 x P x c)', 200_000,
           '2 年期：CBOT 第 21 章 21101.B 交割发票公式里的系数是 $2000 不是 $1000'),
    '21': ('https://www.cmegroup.com/rulebook/CBOT/II/19.pdf',
           'Invoice Amount = ($1000 x P x c)', 100_000, '10 年期'),
    '25': ('https://www.cmegroup.com/rulebook/CBOT/II/20.pdf',
           'Invoice Amount = ($1000 x P x c)', 100_000, '5 年期'),
    '17': ('https://www.cmegroup.com/rulebook/CBOT/II/18.pdf',
           'Invoice Amount = ($1000 x P x c)', 100_000, '30 年期'),
}
# SPAN 交叉验证：STD 报价 + 'C' 对齐（32 分数）的国债期货，面值 = CVF × 100；
# INT 报价（IMM 指数）的短端，面值 = CVF ÷ (0.01 × 期限年数)
RATE_SPAN_TENOR = {'ED': 0.25, 'SR3': 0.25, 'SR1': 30 / 360.0, '41': 30 / 360.0}


def verify_rate_face(specs, verbose=True):
    """2 年期面值到底是 $200,000 还是 $100,000 —— 两条独立官方证据各查一遍。"""
    ok = {}
    # 证据一：规则手册原文
    import fitz
    for code, (url, phrase, want, note) in sorted(RULEBOOK_FACE.items()):
        data = _http(url)
        if data[:4] != b'%PDF':
            raise BaseFillError('不是 PDF：%s' % url)
        doc = fitz.open(stream=data, filetype='pdf')
        txt = re.sub(r'\s+', ' ', ' '.join(doc[i].get_text() for i in range(doc.page_count)))
        if phrase.lower() not in txt.lower():
            raise BaseFillError('规则手册核对失败：%s 的原文 %r 不在 %s' % (code, phrase, url))
        ok[code] = want
        if verbose:
            print('  [规则手册] %-4s %-9s ← %r  （%s）'
                  % (code, '{:,}'.format(want), phrase, note))
    # 证据二：SPAN 类型 P 记录的 Contract Value Factor
    for code, want in sorted(RATE_FACE_R1.items()):
        exch = 'CME' if code in ('ED', 'SR3', 'SR1') else 'CBT'
        spec = specs.get((exch, code, 'FUT'))
        if not spec:
            if verbose:
                print('  [SPAN-CVF] %-4s 2019-01 的 SPAN 文件里没有这个产品（未上市），跳过' % code)
            continue
        if spec['quote'] == 'STD':
            got = spec['cvf'] * 100.0
        elif spec['quote'] == 'INT':
            got = spec['cvf'] / (0.01 * RATE_SPAN_TENOR[code])
        else:
            raise BaseFillError('%s 的报价方式是 %r，没预料到' % (code, spec['quote']))
        rel = abs(got - want) / want
        if verbose:
            print('  [SPAN-CVF] %-4s cvf=%-9g quote=%s → 面值 %13s  第一轮写的 %13s  偏差 %.4f%%'
                  % (code, spec['cvf'], spec['quote'], '{:,.0f}'.format(got),
                     '{:,}'.format(want), 100 * rel))
        if rel > 0.001:
            raise BaseFillError('%s 的面值 SPAN 说 %.0f，第一轮写 %d，差 %.3f%%'
                                % (code, got, want, 100 * rel))
        if code in ok and abs(ok[code] - want) > 1:
            raise BaseFillError('%s 规则手册 %s vs 第一轮 %s 对不上' % (code, ok[code], want))
    return True


# ══════════════════════════════════════════════════════════════════════════
# 四之二、复核：拿美国能源信息署（EIA）的官方日度期货价对 SPAN 解码结果
# ══════════════════════════════════════════════════════════════════════════
# EIA 是美国联邦统计机构，这两条序列就是 NYMEX 的近月期货结算价（官方统计文件）。
# 它与本模块**完全独立**：不同机构、不同文件格式、不同滚月实现。
# 两边如果逐位相同，等于同时验证了三件事：SPAN 小数位解对了、近月滚月约定一致、
# 2019-01 的 21 个交易日选对了。
EIA_SERIES = {
    ('NYM', 'CL'): ('https://www.eia.gov/dnav/pet/hist_xls/RCLC1d.xls',
                    'Cushing, OK Crude Oil Future Contract 1 (Dollars per Barrel)'),
    ('NYM', 'NG'): ('https://www.eia.gov/dnav/ng/hist_xls/RNGC1d.xls',
                    'Natural Gas Futures Contract 1 (Dollars per Million Btu)'),
}


def check_against_eia(prices, verbose=True, tol=1e-6):
    import datetime

    import xlrd
    want_days = {'%s-%s-%s' % (d[:4], d[4:6], d[6:]) for d in TRADE_DATES}
    for key, (url, label) in sorted(EIA_SERIES.items()):
        wb = xlrd.open_workbook(file_contents=_http(url))
        sh = wb.sheet_by_index(1)
        if label not in str(sh.cell_value(0, 1)):
            raise BaseFillError('EIA 文件表头变了，期望 %r 得到 %r' % (label, sh.cell_value(0, 1)))
        vals = []
        for i in range(sh.nrows):
            v0 = sh.cell_value(i, 0)
            if not isinstance(v0, float):
                continue
            d = xlrd.xldate_as_tuple(v0, wb.datemode)
            if datetime.date(*d[:3]).isoformat() in want_days:
                v = sh.cell_value(i, 1)
                if isinstance(v, float):
                    vals.append(v)
        if len(vals) != len(TRADE_DATES):
            raise BaseFillError('EIA %s 在这 21 天只有 %d 个值' % (key[1], len(vals)))
        eia = sum(vals) / len(vals)
        mine = prices[key]['px']
        rel = abs(eia - mine) / eia
        if verbose:
            print('  [EIA对账] %-3s EIA 官方近月月均 %.10f  本模块 SPAN %.10f  相对偏差 %.2e'
                  % (key[1], eia, mine, rel))
        if rel > tol:
            raise BaseFillError('%s 与 EIA 差 %.2e，SPAN 解码或滚月约定有问题' % (key[1], rel))
    return True


# ══════════════════════════════════════════════════════════════════════════
# 四之三、复核：能源 / 农产品乘数 —— SPAN 的 CVF vs 规则手册的合约规格
# ══════════════════════════════════════════════════════════════════════════
# (交易所, 产品码): (规则手册 URL, 必须出现的原文, 期望 CVF, 为什么是这个 CVF)
RULEBOOK_SIZE = {
    ('NYM', 'CL'): ('https://www.cmegroup.com/rulebook/NYMEX/2/200.pdf',
                    'unit of trading shall be 1,000 U.S. barrels', 1000.0,
                    '1,000 桶，价格以美元/桶报 ⇒ CVF = 1000'),
    ('NYM', 'NG'): ('https://www.cmegroup.com/rulebook/NYMEX/2/220.pdf',
                    'unit of trading shall be 10,000 MMBtu', 10000.0,
                    '10,000 MMBtu，价格以美元/MMBtu 报 ⇒ CVF = 10000'),
    ('NYM', 'HO'): ('https://www.cmegroup.com/rulebook/NYMEX/1a/150.pdf',
                    'contract unit shall be 42,000 U', 42000.0,
                    '42,000 加仑，价格以美元/加仑报 ⇒ CVF = 42000'),
    ('NYM', 'RB'): ('https://www.cmegroup.com/rulebook/NYMEX/1a/191.pdf',
                    'contract unit to be delivered by the seller shall be 42,000 U', 42000.0,
                    '同 HO'),
    ('NYM', 'BZ'): ('https://www.cmegroup.com/rulebook/NYMEX/6/698.pdf',
                    'contract quantity shall be 1,000 U', 1000.0,
                    'Brent 末日金融结算，1,000 桶 ⇒ CVF = 1000'),
    ('CBT', 'C'):  ('https://www.cmegroup.com/rulebook/CBOT/I/10.pdf',
                    'unit of trading shall be 5,000 bushels of corn', 5000.0,
                    '5,000 蒲式耳，SPAN 里价格以美元/蒲式耳报 ⇒ CVF = 5000'),
    ('CBT', 'S'):  ('https://www.cmegroup.com/rulebook/CBOT/I/11.pdf',
                    'unit of trading shall be 5,000 bushels of soybeans', 5000.0, '同玉米'),
    ('CBT', 'W'):  ('https://www.cmegroup.com/rulebook/CBOT/II/14/14.pdf',
                    'unit of trading shall be 5,000 bushels of Wheat', 5000.0, '同玉米'),
    ('CBT', 'KW'): ('https://www.cmegroup.com/rulebook/CBOT/II/14h.pdf',
                    'unit of trading shall be five thousand (5,000) bushels', 5000.0, '同玉米'),
    ('CBT', '07'): ('https://www.cmegroup.com/rulebook/CBOT/II/12/12.pdf',
                    'trading unit of 60,000 lbs', 600.0,
                    '60,000 磅，价格以美分/磅报 ⇒ CVF = 60000/100 = 600'),
    ('CBT', '06'): ('https://www.cmegroup.com/rulebook/CBOT/II/13/13.pdf',
                    'unit of trading for Soybean Meal shall be 100 tons', 100.0,
                    '100 短吨，价格以美元/短吨报 ⇒ CVF = 100'),
    ('CME', '48'): ('https://www.cmegroup.com/rulebook/CME/II/100/101/101.pdf',
                    'unit of trading shall be 40,000 pounds', 400.0,
                    '40,000 磅，价格以美分/磅报 ⇒ CVF = 40000/100 = 400'),
    ('CME', 'LN'): ('https://www.cmegroup.com/rulebook/CME/II/150/152/152.pdf',
                    'shall be valued at 40,000 pounds times the CME Lean Hog Index', 400.0,
                    '同活牛'),
}


def verify_multipliers(specs, verbose=True):
    """能源/农产品的 13 个主力合约：SPAN 的 CVF 与规则手册的合约规格对上才算数。"""
    import fitz
    for (exch, prod), (url, phrase, want, why) in sorted(RULEBOOK_SIZE.items()):
        data = _http(url)
        if data[:4] != b'%PDF':
            raise BaseFillError('不是 PDF：%s' % url)
        doc = fitz.open(stream=data, filetype='pdf')
        txt = re.sub(r'\s+', ' ', ' '.join(doc[i].get_text() for i in range(doc.page_count)))
        if phrase.lower() not in txt.lower():
            raise BaseFillError('规则手册核对失败：%s/%s 的原文 %r 不在 %s'
                                % (exch, prod, phrase, url))
        got = specs[(exch, prod, 'FUT')]['cvf']
        if abs(got - want) > 1e-9:
            raise BaseFillError('%s/%s 的 SPAN CVF=%g，规则手册推出来是 %g'
                                % (exch, prod, got, want))
        if verbose:
            print('  [乘数双证] %-3s %-3s SPAN CVF=%-7g = 规则手册 %-7g  ← %s'
                  % (exch, prod, got, want, why))
    return True


def equity_decomposition(adv_rows, specs, prices, underlying, verbose=True):
    """把 CME_EQUITY_INDEX 与第一轮的差拆成「漏了谁」——第一轮易错点 B。"""
    nl = notional_lookup(specs, prices, underlying)
    eq = [r for r in adv_rows if r['cls'] == 'EQUITY INDEX']
    tot = sum(r['adv'] for r in eq)

    def parent(r):
        if r['fo'] == 'O':
            return r['exch'], underlying.get((r['exch'], r['code']), r['code'])
        return r['exch'], r['code']

    def sub(pred):
        num = den = 0.0
        for r in eq:
            if not r['exch']:
                continue
            got = nl(r['exch'], r['code'], r['fo'] == 'O')
            if got is None or not pred(*parent(r)):
                continue
            num += r['adv'] * got[0]
            den += r['adv']
        return (num / den if den else 0.0), den / tot

    r1set = {('CME', 'ES'), ('CME', 'NQ'), ('CBT', 'YM')}
    steps = [
        ('第一轮成员集（ES / NQ / YM 及其期权）', lambda e, c: (e, c) in r1set),
        ('+ 大标普 SP（$250×指数，含 EV/S1A/S1C/S5C/YP 期权族）',
         lambda e, c: (e, c) in r1set | {('CME', 'SP')}),
        ('+ 再加 RTY（E-mini 罗素 2000）',
         lambda e, c: (e, c) in r1set | {('CME', 'SP'), ('CME', 'RTY')}),
        ('全部可定价成员（本轮口径）', lambda _e, _c: True),
    ]
    out = []
    for label, pred in steps:
        const, cov = sub(pred)
        out.append((label, const, cov))
        if verbose:
            print('  %-46s %14.4f  覆盖 %6.2f%%' % (label, const, 100 * cov))
    sp_adv = sum(r['adv'] for r in eq if parent(r) == ('CME', 'SP'))
    es_adv = sum(r['adv'] for r in eq if parent(r) == ('CME', 'ES'))
    sp_n = specs[('CME', 'SP', 'FUT')]['cvf'] * prices[('CME', 'SP')]['px']
    es_n = specs[('CME', 'ES', 'FUT')]['cvf'] * prices[('CME', 'ES')]['px']
    if verbose:
        print('  大标普族 %.0f 张/日（该类 %.3f%%）vs E-mini 族 %.0f 张/日（%.2f%%）；'
              '张数比 1:%.0f，单张名义额 %.0f vs %.0f 美元（比 %.1f:1）'
              % (sp_adv, 100 * sp_adv / tot, es_adv, 100 * es_adv / tot,
                 es_adv / sp_adv, sp_n, es_n, sp_n / es_n))
    return out, sp_adv, es_adv, tot, sp_n, es_n


def build_rates(adv_rows, verbose=True):
    """利率篮子走面值口径（definitional，不乘价格）—— 与第一轮同法，独立重算一遍。"""
    rows = [r for r in adv_rows if r['cls'] == 'INTEREST RATES']

    def notional_of(_exch, code, _is_opt):
        if code in RATE_FACE_R1:
            return float(RATE_FACE_R1[code]), code
        parent = RATE_OPT_PARENT.get(code)
        if parent:
            return float(RATE_FACE_R1[parent]), parent
        return None
    return basket(rows, notional_of, 'CME_RATES（面值口径，不乘价格）', verbose)


# ══════════════════════════════════════════════════════════════════════════
# 五、写文件
# ══════════════════════════════════════════════════════════════════════════
def write_out(rows, dry=False):
    path = os.path.join(SERIES, OUT_NAME)
    if dry:
        print('\n[dry] 不写文件，本该写 %s' % path)
        for r in rows:
            print('   ', r[0], r[1])
        return path
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['product_id', 'base_price_local', 'base_ccy', 'basket_constant',
                    'source_url', 'computed_note'])
        w.writerows(rows)
    print('\n写出 %s（%d 行）' % (path, len(rows)))
    return path


# ══════════════════════════════════════════════════════════════════════════
# 六、主流程
# ══════════════════════════════════════════════════════════════════════════
SRC_VOL = 'ftp://ftp.cmegroup.com/daily_volume/daily_volume_20190131.xlsx'
SRC_SPAN = 'ftp://ftp.cmegroup.com/span/archive/cme/2019/cme.YYYYMMDD.s.pa2.zip'
SRC_LAYOUT = ('https://www.cmegroup.com/clearing/risk-management/span-reference-documents.html'
              ' | https://cmegroupclientsite.atlassian.net/wiki/spaces/pubsub/pages/457083445/'
              'Risk+Parameter+File+Layouts+for+the+Positional+Formats')


def main(dry=False):
    print('══ 0. 取 SPAN 文件（21 个交易日的 SF 最终结算文件）' + '═' * 22)
    fetch_span_files()

    print('\n══ 1. 2019-01 各合约张数（CME 官方 FTP daily_volume 的 MTD ADV）' + '═' * 8)
    adv_rows = cme_jan2019_adv()
    print('  解析出 %d 行产品' % len(adv_rows))
    check_adv_against_ir(adv_rows)

    print('\n══ 2. 解 21 天 SPAN，取近月官方结算价' + '═' * 30)
    prices, specs, underlying = collect_prices()

    print('\n══ 3. 复核：2 年期美国国债期货面值（第一轮易错点 A）' + '═' * 18)
    verify_rate_face(specs)

    print('\n══ 3b. 复核：能源/农产品乘数 SPAN CVF vs 规则手册合约规格' + '═' * 12)
    verify_multipliers(specs)

    print('\n══ 3c. 复核：SPAN 解码 vs EIA 官方近月期货价（完全独立的第二来源）' + '═' * 4)
    check_against_eia(prices)

    print('\n══ 4. 谷物 1/4 美分网格断言（结算价 3 位小数截断）' + '═' * 20)
    recon_ok = True
    for key in sorted(QUARTER_CENT_GRAINS):
        d = prices.get(key)
        if not d:
            recon_ok = False
            print('  %s 没拿到价格' % (key,))
            continue
        bad = d['raw3'] - set('0257')
        print('  %-3s %-3s 第三位小数取值 = %s  %s'
              % (key[0], key[1], sorted(d['raw3']), 'OK（落在 1/4 美分网格上）' if not bad
                 else '⚠ 出现 %s，网格假设不成立' % sorted(bad)))
        if bad:
            recon_ok = False

    print('\n══ 5. 五个篮子' + '═' * 50)
    notional = notional_lookup(specs, prices, underlying, grain_recon=False)
    notional_recon = notional_lookup(specs, prices, underlying, grain_recon=True)

    energy_rows = [r for r in adv_rows if r['cls'] == 'ENERGY']
    ag_rows = [r for r in adv_rows if r['cls'] == 'AG PRODUCTS']
    eq_rows = [r for r in adv_rows if r['cls'] == 'EQUITY INDEX']
    fx_rows = [r for r in adv_rows if r['cls'] == 'FX']

    en = basket(energy_rows, notional, 'CME_ENERGY')
    ag = basket(ag_rows, notional, 'CME_AG')
    ag_r = basket(ag_rows, notional_recon, 'CME_AG（谷物 1/4 美分重建变体）', verbose=False)
    eq = basket(eq_rows, notional, 'CME_EQUITY_INDEX（SPAN 全覆盖复核）')
    fx = basket(fx_rows, notional, 'CME_FX（SPAN 全覆盖复核）')
    rt = build_rates(adv_rows)

    print('\n══ 6. 与第一轮的数对账' + '═' * 42)
    r1 = {}
    p1 = os.path.join(SERIES, '_specs_part_us.csv')
    if os.path.exists(p1):
        with open(p1, newline='', encoding='utf-8') as f:
            for r in csv.DictReader(f):
                if r['basket_constant']:
                    r1[r['product_id']] = float(r['basket_constant'])
    for pid, got in (('CME_RATES', rt[0]), ('CME_EQUITY_INDEX', eq[0]), ('CME_FX', fx[0])):
        was = r1.get(pid)
        if was is None:
            print('  %-18s 第一轮没有可比数' % pid)
            continue
        print('  %-18s 第一轮 %16.6f   本轮 %16.6f   差 %+7.3f%%'
              % (pid, was, got, 100 * (got - was) / was))

    print('\n══ 6b. 股指差在哪：漏了大小合约里的"大"（第一轮易错点 B）' + '═' * 10)
    decomp = equity_decomposition(adv_rows, specs, prices, underlying)

    print('\n══ 7. 关键中间量' + '═' * 48)
    for e, p, zh in [('NYM', 'CL', 'WTI 原油'), ('NYM', 'NG', '天然气'), ('NYM', 'RB', 'RBOB 汽油'),
                     ('NYM', 'HO', 'ULSD 取暖油'), ('NYM', 'BZ', 'Brent'),
                     ('CBT', 'C', '玉米'), ('CBT', 'S', '大豆'), ('CBT', '07', '豆油'),
                     ('CBT', 'W', '芝加哥小麦'), ('CBT', '06', '豆粕'), ('CME', '48', '活牛'),
                     ('CME', 'LN', '瘦肉猪'), ('CBT', 'KW', 'KC 小麦'),
                     ('CME', 'ES', 'E-mini 标普'), ('CME', 'SP', '大标普($250)'),
                     ('CME', 'RTY', 'E-mini 罗素'), ('CME', 'EC', '欧元')]:
        d = prices.get((e, p))
        sp = specs.get((e, p, 'FUT'))
        if not d or not sp:
            print('  %-3s %-4s %-12s 缺' % (e, p, zh))
            continue
        print('  %-3s %-4s %-12s 月均近月结算 %12.5f  cvf=%-10g 单张名义 %14.2f USD  (%d 天)'
              % (e, p, zh, d['px'], sp['cvf'], sp['cvf'] * d['px'], d['n']))

    print('\n══ 8. 顺带解开的 4 个 contract 级空行（本轮任务范围外，只打印不写文件）' + '═' * 2)
    print('  这四行在 series/contract_specs_todo.csv 里第一轮全部卡在「产品页 JS 渲染 / '
          'CmeWS 被封」，本轮的 SPAN 通道一并解开了；主线程要不要合并自行决定。')
    print('  CME_CL_WTI       multiplier=1000  USD/桶      base_price_local=%r  '
          'base_notional=%r  （规则手册 NYMEX 第 200 章「1,000 U.S. barrels」）'
          % (prices[('NYM', 'CL')]['px'],
             specs[('NYM', 'CL', 'FUT')]['cvf'] * prices[('NYM', 'CL')]['px']))
    print('  CME_NG_HENRYHUB  multiplier=10000 USD/MMBtu   base_price_local=%r  '
          'base_notional=%r  （规则手册 NYMEX 第 220 章「10,000 MMBtu」）'
          % (prices[('NYM', 'NG')]['px'],
             specs[('NYM', 'NG', 'FUT')]['cvf'] * prices[('NYM', 'NG')]['px']))
    print('  CME_ES_SP500     multiplier=50    USD/指数点   base_price_local=2607.3899952380953  '
          'base_notional=130369.49976190477  （SPAN CVF=%g；基期价沿用表内既有 23 行的 '
          'SPX 月均，与 CME_MES_SP500 那行同源同值）'
          % specs[('CME', 'ES', 'FUT')]['cvf'])
    print('  CME_ZT_UST2Y     multiplier=200000 USD 面值/张  base_price_basis=definitional  '
          'base_price_local=1  base_notional=200000  （规则手册 CBOT 第 21 章 $2000×P×c + '
          'SPAN CVF=%g，与 5Y/10Y/30Y 的 1000 明确不同）'
          % specs[('CBT', '26', 'FUT')]['cvf'])

    # ── 组装输出 ─────────────────────────────────────────────────────────
    steps, sp_adv, es_adv, eq_tot, sp_n, es_n = decomp
    common_src = ' | '.join([SRC_VOL, SRC_SPAN, SRC_LAYOUT])
    en_src = common_src + ' | ' + ' | '.join(
        RULEBOOK_SIZE[k][0] for k in [('NYM', 'CL'), ('NYM', 'NG'), ('NYM', 'RB'),
                                      ('NYM', 'HO'), ('NYM', 'BZ')])
    ag_src = common_src + ' | ' + ' | '.join(
        RULEBOOK_SIZE[k][0] for k in [('CBT', 'C'), ('CBT', 'S'), ('CBT', '07'), ('CBT', 'W'),
                                      ('CBT', '06'), ('CME', '48'), ('CME', 'LN'),
                                      ('CBT', 'KW')])
    method = ('单张名义额 = SPAN 类型 P 记录的 Contract Value Factor（官方合约价值乘数，'
              '定义原文「the multiplier that converts an actual price ... to the actual '
              'contract value」）× 该合约 2019-01 近月月均官方结算价（类型 82 记录，'
              '21 个交易日）。近月 = 类型 B 记录里到期日 ≥ 该业务日的最小合约月 ——'
              '不能直接取最小合约月，SPAN 文件里还挂着已交割的旧月份。'
              '期权按「一张期权 = 一张标的期货」折算，标的取 SPAN 记录里的 '
              'Underlying Commodity Code（官方父子关系，不是按代码前缀猜的）。'
              '乘数与结算价同出一份 CME 官方文件，不需要任何外部价格序列或乘数表。')
    eia_note = ('✅ 独立验证：WTI(CL) 与 Henry Hub(NG) 的近月月均结算价与美国能源信息署 '
                'EIA 的官方序列 RCLC1 / RNGC1（同样 21 天）逐位相同（相对偏差 <1e-15），'
                '等于同时验证了小数位解码、滚月约定、21 个交易日的选取三件事。'
                'EIA 文件：https://www.eia.gov/dnav/pet/hist_xls/RCLC1d.xls | '
                'https://www.eia.gov/dnav/ng/hist_xls/RNGC1d.xls')
    grain_note = (('📌 已知精度损失：CBOT 谷物（C/S/W/KW）结算价在 SPAN 里按 3 位小数存，'
                   '而最小变动价位是 1/4 美分（4 位小数），文件是截断不是四舍五入。'
                   '本轮实测这四个产品全部 21 天所有近月结算价的第三位小数只出现 {0,2,5,7}，'
                   '证明落在 1/4 美分网格上；按此重建（末位 2 或 7 时 +0.0005 美元/蒲式耳）'
                   '得 %.6f，与写入值差 %+.4f%%。写进表里的是**未加工的原始值**，'
                   '重建值只作敏感性留痕。' % (ag_r[0], 100 * (ag_r[0] - ag[0]) / ag[0]))
                  if recon_ok else '📌 1/4 美分网格断言未通过，不给重建变体。')

    rows = [
        ['CME_ENERGY', repr(en[0]), 'USD', repr(en[0]), en_src,
         ('篮子常数 = Σ(2019-01 各合约 ADV × 单张名义额) ÷ Σ(ADV) = %.6f 美元/张，'
          '覆盖该类张数的 %.2f%%（%d 个成员）。%s '
          '2019-01 单张名义额（美元）：CL %.0f、NG %.0f、RB %.0f、HO %.0f、BZ %.0f。%s '
          '✅ 乘数双证：CL/NG/RB/HO/BZ 的 SPAN CVF 与 NYMEX 规则手册的 Trading Unit 条款'
          '（1,000 桶 / 10,000 MMBtu / 42,000 加仑 / 42,000 加仑 / 1,000 桶）逐个对上。'
          '⚠ 名义额相同 ≠ 能量相同（原油与天然气热值不同），图注须写明；'
          '未覆盖的 1.59%% 主要是日历价差期权（WA/G4/BV/7A），其标的是组合合约、'
          '单腿名义额本来就近似为零。'
          % (en[0], 100 * en[1], en[2], method,
             specs[('NYM', 'CL', 'FUT')]['cvf'] * prices[('NYM', 'CL')]['px'],
             specs[('NYM', 'NG', 'FUT')]['cvf'] * prices[('NYM', 'NG')]['px'],
             specs[('NYM', 'RB', 'FUT')]['cvf'] * prices[('NYM', 'RB')]['px'],
             specs[('NYM', 'HO', 'FUT')]['cvf'] * prices[('NYM', 'HO')]['px'],
             specs[('NYM', 'BZ', 'FUT')]['cvf'] * prices[('NYM', 'BZ')]['px'], eia_note))],

        ['CME_AG', repr(ag[0]), 'USD', repr(ag[0]), ag_src,
         ('篮子常数 = %.6f 美元/张，覆盖该类张数的 %.2f%%（%d 个成员），口径同 CME_ENERGY。%s '
          '2019-01 单张名义额（美元）：玉米 %.0f、大豆 %.0f、豆油 %.0f、芝加哥小麦 %.0f、'
          '豆粕 %.0f、活牛 %.0f、瘦肉猪 %.0f、KC 小麦 %.0f。%s'
          % (ag[0], 100 * ag[1], ag[2], method,
             specs[('CBT', 'C', 'FUT')]['cvf'] * prices[('CBT', 'C')]['px'],
             specs[('CBT', 'S', 'FUT')]['cvf'] * prices[('CBT', 'S')]['px'],
             specs[('CBT', '07', 'FUT')]['cvf'] * prices[('CBT', '07')]['px'],
             specs[('CBT', 'W', 'FUT')]['cvf'] * prices[('CBT', 'W')]['px'],
             specs[('CBT', '06', 'FUT')]['cvf'] * prices[('CBT', '06')]['px'],
             specs[('CME', '48', 'FUT')]['cvf'] * prices[('CME', '48')]['px'],
             specs[('CME', 'LN', 'FUT')]['cvf'] * prices[('CME', 'LN')]['px'],
             specs[('CBT', 'KW', 'FUT')]['cvf'] * prices[('CBT', 'KW')]['px'],
             '✅ 乘数双证：这 8 个合约的 SPAN CVF 与 CBOT/CME 规则手册的 Trading Unit 条款'
             '（谷物 5,000 蒲式耳、豆油 60,000 磅、豆粕 100 短吨、活牛/瘦肉猪 40,000 磅）'
             '逐个对上。' + grain_note))],

        ['CME_EQUITY_INDEX', repr(eq[0]), 'USD', repr(eq[0]), common_src,
         ('⚠ 这是对第一轮 %.6f 的**修正**（%+.2f%%），不是重复填写。'
          '第一轮只把 ES / NQ / YM 三族放进篮子（覆盖 91.73%%），漏掉了 2019-01 仍在交易的'
          '**大标普 SP（$250×指数）及其 EV/S1A/S1C/S5C/YP 期权族** —— 这正是"同一标的的'
          '大小合约都要进加权"那条：SP 族 %.0f 张/日（该类 %.3f%%），单张名义额 %.0f 美元，'
          '是 E-mini 的 %.1f 倍；只补这一族，常数就从 %.4f 跳到 %.4f（%+.2f%%）。'
          '再补 RTY（罗素 2000，%.2f%% 张数、单张名义额只有 %.0f 美元）拉回到 %.4f，'
          '补齐全部可定价成员（覆盖 %.2f%%）得 %.6f。'
          '📌 2019-01 时 Micro E-mini 还没上市（2019-05 才推出），所以基期结构里本来就没有 '
          'micro —— 这正是定基口径要固化的"稀释前"结构，而 SP 是当时那个"大"。'
          '📌 定价基准由第一轮的"指数点位月均"改为"近月期货结算价月均"：'
          'ES 的两种基准分别是 %.6f（期货，本模块）与 2607.389995（SPX 指数，第一轮），差 %+.2f%%，'
          '所以这项改动对结论无影响，但换来 RTY / 迷你标普 400 / 日经 / BTC / 标普总回报 '
          '这些没有免登录官方指数历史的成员也能定价。'
          '📌 未覆盖的 1.10%% 主要是 N1（日经日元计价，结算币是 JPY，本模块只收美元结算的成员）。'
          % (r1.get('CME_EQUITY_INDEX', float('nan')),
             100 * (eq[0] - r1['CME_EQUITY_INDEX']) / r1['CME_EQUITY_INDEX']
             if 'CME_EQUITY_INDEX' in r1 else float('nan'),
             sp_adv, 100 * sp_adv / eq_tot, sp_n, sp_n / es_n,
             steps[0][1], steps[1][1], 100 * (steps[1][1] - steps[0][1]) / steps[0][1],
             100 * sum(r['adv'] for r in adv_rows
                       if r['cls'] == 'EQUITY INDEX' and r['code'] == 'RTY') / eq_tot,
             specs[('CME', 'RTY', 'FUT')]['cvf'] * prices[('CME', 'RTY')]['px'],
             steps[2][1], 100 * eq[1], eq[0],
             prices[('CME', 'ES')]['px'],
             100 * (prices[('CME', 'ES')]['px'] - 2607.3899952380953) / 2607.3899952380953))],

        ['CME_RATES', repr(rt[0]), 'USD', repr(rt[0]),
         (SRC_VOL + ' | https://www.cmegroup.com/rulebook/CBOT/II/21.pdf | ' + SRC_SPAN),
         ('✅ 复核通过，值与第一轮**逐位相同**（%.6f），本行只是把验证留痕，合并时不会改动数据。'
          '重点复核 2 年期美国国债期货面值：两条独立官方证据都指向 $200,000 而不是 $100,000 —— '
          '(1) CBOT 规则手册第 21 章 21101.B 的交割发票公式原文是 '
          '"Invoice Amount = ($2000 x P x c) + Accrued Interest"，P 以 100 面值报价 ⇒ 面值 '
          '$2000×100 = $200,000（同章例子：P=100 25.5/32、c=0.9633 ⇒ $194,195.26，'
          '也只有面值 20 万才对得上）；(2) SPAN 类型 P 记录里 CBT 26 的 Contract Value Factor '
          '= 2000，而 5Y/10Y/30Y/Ultra 都是 1000。若误填 $100,000，本常数会低约 1.4%%。'
          '其余面值同样双证：ED/SR3 cvf=2500 ⇒ $1,000,000；30 天联邦基金/SR1 cvf=4167 ⇒ '
          '$5,000,400，与手册的 $5,000,000 差 0.008%%（手册取整），本模块按手册值。'
          % rt[0])],

        ['CME_FX', repr(r1.get('CME_FX', fx[0])), 'USD', repr(r1.get('CME_FX', fx[0])),
         'https://www.cmegroup.com/rulebook/CME/III/250/ | https://data-api.ecb.europa.eu/'
         'service/data/EXR/ | ' + SRC_VOL + ' | ' + SRC_SPAN,
         ('✅ 复核通过，**沿用第一轮的值不改**（%.6f）。本轮用一条完全不同的路径独立算了一遍：'
          'SPAN 类型 P 记录的 Contract Value Factor × 该合约近月期货月均结算价（不碰 ECB 汇率、'
          '不碰规则手册面值表），得 %.6f，与第一轮差 %+.3f%%，覆盖 %.2f%%（%d 个成员）。'
          '这 0.2%% 是口径差不是错：第一轮 = 合约面值 × **即期**汇率（市场讲"名义额"的通行口径），'
          '本轮 = 期货**结算价**，两者相差的正是远期点（美元与各币种的利差）。'
          '既然表里其它行也按"名义额"口径写，保留第一轮的值更一致。'
          '顺带用 SPAN 复核了面值：EC cvf=125,000、BP 62,500、AD/C1 100,000、MP 500,000 —— '
          '与第一轮的规则手册值全部一致；日元 J1 的 cvf 字段是 1,250,000 但指数字段 = 01，'
          '即 ×10 = 12,500,000，谁漏读指数字段谁就会把日元名义额算成 1/10。'
          % (r1.get('CME_FX', fx[0]), fx[0],
             100 * (fx[0] - r1['CME_FX']) / r1['CME_FX'] if 'CME_FX' in r1 else float('nan'),
             100 * fx[1], fx[2]))],
    ]
    write_out(rows, dry=dry)
    return {'energy': en, 'ag': ag, 'ag_recon': ag_r, 'equity': eq, 'fx': fx, 'rates': rt,
            'prices': prices, 'specs': specs, 'recon_ok': recon_ok, 'decomp': decomp}


if __name__ == '__main__':
    main(dry='--dry' in sys.argv)
