# -*- coding: utf-8 -*-
"""HOOD 月度序列的**历史回填**：series/hood.csv 从 2023-04 起推到 2021-01 起（+27 期）。

用法:
    python3 build/basefill/hood_2021.py                      # 取数 + 核对 + 写 CSV
    python3 build/basefill/hood_2021.py --dry                # 只核对、只打印，不写
    python3 build/basefill/hood_2021.py --refresh            # 强制重下三份原件
    python3 build/basefill/hood_2021.py --cache-dir <路径>   # 原件落在哪

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
这个文件为什么存在，以及为什么**不**塞进 fetch/hood.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
fetch/hood.py 的 update() 只往右走（它自己的 docstring 写明了理由）。往左走是另一件事：
IR 的 Monthly Metrics 索引页**最早只排到 2023-01**，再往前的月份不在那张页面上，
只能去 Quarterly Results 页翻**当年的 Earnings Supplement**——那是另一张页、另一种版式、
行数少一半、章节名和行标签全不一样。而且这个洞补完就永远关上了（2021-01 是官方月度
披露的天花板，见下），把它焊进无人值守链路只会让每月例行任务多背一套永远不会再用的
分支。照 build/basefill/cboe_2016.py、tmx_ciro_2015.py、spgi_history.py 的先例放这里。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
天花板就是 2021-01，不是 2016
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
8-K 0001783879-22-000088（2022-04-28，Item 7.01 Reg FD）原文：

    "Robinhood is going to start reporting certain limited purpose statistical and
     operational results on a monthly basis. The first report will cover the month of
     March 2022 (and each of the preceding 12 calendar months)…"

即**月度披露 2022-04 才开始**，首期回填到 2021-03。2021-01 / 2021-02 两个月是后来的
Earnings Supplement 才补上的（本脚本用的 Q1'23 那份就有）。再往前只存在于 S-1 与 10-Q，
且不是月度粒度 —— 所以本页的历史到 2021-01 为止，**不硬凑**。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
三个源，各管一段；取的一律是「**能拿到的最早那一版**」
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  ① LEGACY  robinhood-markets-q1-2023-supplemental-information.xlsx（随 Q1'23 财报发）
     覆盖 2021-01~2023-03，是**现今仍挂在 IR 站上、覆盖 2021 年的最早一份**。
     喂 2021-01~2022-12 的 12 列。
  ② SECLEND Q4'23 Earnings Supplement.xlsx（随 Q4'23 财报发）
     覆盖 2021-01~2023-12。它比 ① 多一行 Total Securities Lending Revenue ($M)，
     2022-05 起有数（更早印 NA —— 出借业务 2022-05 才上线）。只喂
     seclend_total_usdmn 的 2022-05~2022-12 这 8 格，别的列一格不碰。
  ③ MODERN  Q1'26 Earnings Supplement.xlsx，覆盖 2023-01~2026-03，25 列俱全。
     喂 2023-01~2023-03。选它而不选 2023 年当期那三份月度表，是为了跟**已有行同源**：
     series/hood.csv 里 2023-04 起的行本来就是从现代版式的 Supplement 解出来的
     （2023 年当期的月度表只有 12 列，没有 Bitstamp / Event / ADV / Cash and Deposits）。

「取最早那一版」不是洁癖，是仓规「不许拿后期重述值盖历史」的执行：本脚本每次运行都会
把 ① 与 Q1'25 那份逐月比一遍（`check_restatements`），实测**只有 Net Deposits 有 4 个月
被后来改过 0.1**（2021-06 1.0→0.9、2021-09 1.0→0.9、2022-09 1.3→1.2、2022-12 1.6→1.5），
其余 11 列逐月逐位相同。入库取的是 ① 的原值。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
留空的列：官方**本就没有**，不补 0、不换算
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
2021-01~2022-12 这 24 个月，25 列里只有 12（2022-05 起 13）列有官方数字。留空的：

  · vol_crypto_bitstamp_usdbn / adv_crypto_bitstamp_usdmn / vol_crypto_app_usdbn /
    adv_crypto_app_usdmn —— Bitstamp 2025-06 才并表，那之前官方**不拆**加密成交量。
    注意：拿 vol_crypto 整数照抄进 app 列 = 断言「当时 app 就是全部」，虽然事后看是真的，
    但那是**我们**的断言不是公司的披露，所以留空。
  · vol_event_bn / adv_event_mn —— Event Contracts 2024-10 才上线。
  · seclend_total_usdmn（2021-01~2022-04）/ seclend_net_usdmn（全段）——
    出借收入 2022-05 才有数；Net 那一行更晚，2021~2022 的任何一版文件里都没有。
  · cash_and_deposits_usdbn —— 那一行 2023 年才进表（2021~2022 的表只有 Cash Sweep）。
  · crypto_trading_days —— 「Crypto and Prediction Markets Trading Days」那一行同样是
    后来才有的。它等于自然日数，**但那是我们算的，不是官方印的**，所以不填。
  · adv_equity_usdbn / adv_options_mn / adv_crypto_usdmn —— 老版式没有 ADV 那一节。
    ADV = 当月成交量 ÷ 交易日，两个分量本脚本都写进去了，但**换算出来的数不是披露值**，
    仓规明令不许换算，所以这三列留空，由图自己把左端裁到 2023-01（mrwin.resolve）。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
⚠ dats_* 三列装的是 **DARTs**，不是 DATs —— 唯一一处口径不完全同名的填充
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Q2'26 Earnings Supplement（2026-07-29）之前，那一节叫 Daily Average **Revenue** Trades
(DARTs)，之后改叫 Daily Average Trades (DATs)。这不只是改名，数字跟着变了：
同一个 2025-01，Q1'26 印 equity 2.6，Q2'26 印 3.3（+27%）。

但**重述只到 2025-01 为止**：2023-04~2024-12 这 21 个月，两种口径逐月逐位完全相同
（本脚本 `check_darts()` 每次运行都重算一遍并打印，不写死结论）。也就是说
「不产生收入的交易」是 2025 年才变得可观的东西，在本脚本要填的 2021-01~2023-03 那一段，
公司自己的 DATs 序列就等于当时印出来的 DARTs。据此把 DARTs 填进 dats_*，
并在 build/hood.py 的图注里写明这一段是 DARTs 口径。

所以 fetch/hood.py 的 MONTHLY_SPEC **故意不给 DARTs 做别名**（那边有一段注释说明），
免得哪天官方又发一份老版式文件、update() 把窄口径悄悄写进宽口径的列。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
接缝检验：老版式 vs 新版式在 2023-01~2023-03 三个月上重叠
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
① 覆盖到 2023-03，③ 从 2023-01 起 —— 两种版式有 3 个月 x 12 列的重叠，
`check_seam()` 逐格比，**必须逐位相同**，不同就直接报错停下（不是打印警告）。
实测 36 格全同，所以「老表的 Total AUC / NCFA / DARTs 就是新表的 Total Platform Assets /
Funded Customers / DATs 那三行」这句话不是推测，是量出来的。

另一条看门狗 `check_modern()`：③ 与**已有的** 2023-04~2026-03 逐格比。除 dats_equity_mn /
dats_crypto_mn 在 2025-01 起的重述外必须全同 —— 这道闸保证「拿 ③ 去填 2023-01~03」
和「已有行的来源」确实是同一把尺子。
"""
import argparse
import csv
import html as _html
import importlib.util
import os
import re
import sys

import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SERIES_DIR = os.path.join(ROOT, 'series')
CSV_PATH = os.path.join(SERIES_DIR, 'hood.csv')

QR_URL = 'https://investors.robinhood.com/financials/quarterly-results/default.aspx'
FIRST = pd.Period('2021-01', 'M')          # 天花板，见文件头
GUARD = pd.Period('2023-04', 'M')          # 本脚本一格都不许写到这个月及以后
SECLEND_FROM = pd.Period('2022-05', 'M')   # 出借业务上线月


class HoodBasefillError(RuntimeError):
    pass


_F = None


def _fetch():
    """借 fetch/hood.py 的下载通道与解析零件（Akamai 按 TLS 指纹拦人，curl/urllib 必超时）。

    只加载一次并缓存：本脚本要用它的 `fetch_bytes` / `download` / `_norm` / `_fmt_like`
    与生产版的 `parse_sheet`，重复 exec 会白跑几遍模块级代码。
    """
    global _F
    if _F is None:
        p = os.path.join(ROOT, 'fetch', 'hood.py')
        spec = importlib.util.spec_from_file_location('fetch_hood', p)
        _F = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(_F)
    return _F


# ─────────────────────────── 源 ───────────────────────────
# (键, 落盘期号, 官方文件名, 兜底 static-file 路径)
# 文件名是从 Quarterly Results 页现抓的匹配依据；uuid 只是抓不到页面时的兜底，
# **优先走页面**（uuid 会漂，而文件名十年不变）。
SOURCES = {
    'LEGACY': ('2023-03', 'robinhood-markets-q1-2023-supplemental-information.xlsx',
               '/static-files/b1c9df36-ae4a-4247-8229-be9f5732496a'),
    'SECLEND': ('2023-12', "Q4-23-Earnings-Supplement.xlsx",
                '/static-files/4ed71da8-9569-4901-b5a6-257b44dc89be'),
    'MODERN': ('2026-03', "Q1'26 Earnings Supplement.xlsx",
               '/static-files/7eb0f912-2e68-4b70-8189-bf221327d686'),
    # 只用于「后期重述了哪几格」的对照，不入库
    'RESTATE': ('2025-03', "Q1'25 Earnings Supplement.xlsx",
                '/static-files/5b9ed8ff-bf09-429c-85b7-cd2fbaacf1ee'),
}


def _index_links(F):
    """Quarterly Results 页 → {文件名: 绝对 URL}。抓不到就返回空字典，走 uuid 兜底。"""
    try:
        page = F.fetch_bytes(QR_URL).decode('utf-8', 'replace')
    except Exception as e:                                       # noqa: BLE001
        print(f'  [index] Quarterly Results 页抓不到（{type(e).__name__}），改用兜底 uuid')
        return {}
    out = {}
    for href, title in re.findall(r'<a href="(/static-files/[^"]+)"[^>]*title="([^"]*)"', page):
        out.setdefault(_html.unescape(title), F.BASE + href)
    return out


def download_all(cache_dir, refresh=False):
    F = _fetch()
    links = _index_links(F)
    paths = {}
    for key, (period, title, fallback) in SOURCES.items():
        url = links.get(title) or (F.BASE + fallback)
        if title not in links and links:
            print(f'  [index] 页面上没有「{title}」，用兜底 uuid（官方可能换了文件名）')
        entry = {'period': period, 'title': title, 'url': url}
        path = os.path.join(cache_dir, F._safe_name(title, period))
        if refresh and os.path.exists(path):
            os.unlink(path)
        paths[key] = F.download(entry, cache_dir)
        print(f'  [{key:<7}] {os.path.basename(paths[key])}  '
              f'{os.path.getsize(paths[key]):,} bytes')
    return paths


# ─────────────────────────── 解析 ───────────────────────────
MON = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def _read_labels(path):
    """→ {归一化行标签: {Period: 数值或 None}}。

    老版式里行标签**不重复**（没有 ADV 那一节，'Equity ($B)' 只出现一次），所以这里按
    标签直取即可 —— 不像 fetch/hood.py 要靠 (章节, 标签) 两级定位。为防哪天真的重了，
    重名一律留第一次命中并在下面 `_pick()` 里对拿不到的列大声报错。
    """
    F = _fetch()
    rows, cols = _grid(path)
    out = {}
    for r in rows:
        lab = F._norm(next((c for c in r[:3] if c is not None and str(c).strip()), None))
        if not lab:
            continue
        vals = {p: (r[i] if isinstance(r[i], (int, float)) else None) for i, p in cols}
        if all(v is None for v in vals.values()):
            continue                                    # 章节标题行 / 全 NA 行
        out.setdefault(lab, vals)
    if not out:
        raise HoodBasefillError(f'{os.path.basename(path)}: 一行数据都没解析出来')
    return out


# 老版式（2021~2024 的 Earnings Supplement）→ CSV 列名。
# 左边这些标签在新版式里全部改过名，对应关系由 check_seam() 在 3 个重叠月上逐格验过。
LEGACY_MAP = {
    'net cumulative funded accounts (ncfa)': 'funded_customers_mn',   # → Funded Customers
    'total auc': 'total_platform_assets_usdbn',                       # → Total Platform Assets
    'net deposits': 'net_deposits_usdbn',
    'trading days (equities and options)': 'eqopt_trading_days',
    'equity ($b)': 'vol_equity_usdbn',
    'options contracts (m)': 'vol_options_mn',
    'crypto ($b)': 'vol_crypto_usdbn',
    'equity darts': 'dats_equity_mn',        # ⚠ DARTs，见文件头
    'option darts': 'dats_options_mn',
    'crypto darts': 'dats_crypto_mn',
    'margin book': 'margin_book_usdbn',
    'cash sweep': 'cash_sweep_usdbn',        # 更晚的文件改叫 'Total Cash Sweep'
}
LEGACY_ALIASES = {'cash sweep': ('cash sweep', 'total cash sweep'),
                  'net cumulative funded accounts (ncfa)':
                      ('net cumulative funded accounts (ncfa)', 'funded customers')}
SECLEND_LABEL = 'total securities lending revenue ($m)'

# 新版式（Q1'26 Supplement）行标签 → CSV 列名。25 列俱全。
# dats_* 走 DARTs 那三行 —— 那份文件里那一节还叫 DARTs（见文件头）。
MODERN_MAP = {
    'funded customers': 'funded_customers_mn',
    'total platform assets': 'total_platform_assets_usdbn',
    'net deposits': 'net_deposits_usdbn',
    'equities and options trading days': 'eqopt_trading_days',
    'crypto and prediction markets trading days': 'crypto_trading_days',
    'equity ($b)': 'vol_equity_usdbn',
    'options contracts (m)': 'vol_options_mn',
    'crypto ($b)': 'vol_crypto_usdbn',
    'robinhood app ($b)': 'vol_crypto_app_usdbn',
    'bitstamp ($b)': 'vol_crypto_bitstamp_usdbn',
    'event contracts (b)': 'vol_event_bn',
    'equity': 'adv_equity_usdbn',                 # ⚠ 见下面 MODERN_ADV 的说明
    'options contracts (m)#adv': 'adv_options_mn',
    'crypto ($m)': 'adv_crypto_usdmn',
    'robinhood app': 'adv_crypto_app_usdmn',
    'bitstamp': 'adv_crypto_bitstamp_usdmn',
    'event contracts (m)': 'adv_event_mn',
    'margin book': 'margin_book_usdbn',
    'cash and deposits': 'cash_and_deposits_usdbn',
    'cash sweep': 'cash_sweep_usdbn',
    'total securities lending revenue': 'seclend_total_usdmn',
    'securities lending, net': 'seclend_net_usdmn',
}


DARTS_SECTION = 'daily average revenue trades'
DARTS_MAP = {'equity': 'dats_equity_mn', 'options': 'dats_options_mn',
             'crypto': 'dats_crypto_mn'}


def _grid(path, sheet='Monthly KPIs'):
    """→ (rows, 期次列)。两个解析函数共用，省得把同一段表头扫描写两遍。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    name = sheet if sheet in wb.sheetnames else next(
        s for s in wb.sheetnames if s.strip().lower() in ('monthly kpis', 'monthly metrics'))
    ws = wb[name]
    W = ws.max_column
    rows = [list(r) + [None] * (W - len(r)) for r in ws.iter_rows(values_only=True)]
    pr = None
    for ri in range(min(12, len(rows))):
        if sum(1 for i in range(W)
               if rows[ri][i] and str(rows[ri][i]).strip()[:3] in MON) >= 4:
            pr = ri
            break
    if pr is None:
        raise HoodBasefillError(f'{os.path.basename(path)}: 找不到月份表头行')
    yr, cols, cur = rows[pr - 1], [], None
    for i in range(W):
        v = yr[i]
        if v is not None and str(v).replace(',', '').strip().isdigit():
            cur = int(str(v).replace(',', '').strip())
        s = str(rows[pr][i]).strip()[:3] if rows[pr][i] else ''
        if s in MON:
            cols.append((i, pd.Period(f'{cur}-{MON.index(s) + 1:02d}', 'M')))
    if not cols:
        raise HoodBasefillError(f'{os.path.basename(path)}: 表头找到了但没解析出期次')
    return rows, cols


def _modern_frame(path):
    """新版式那份走**生产解析器**（fetch/hood.py 的 (章节, 标签) 两级定位），除了 DARTs 三行。

    为什么不像老版式那样按标签直取：ADV 那一节的 'Equity ($B)' / 'Options Contracts (M)'
    与总量那一节同名，按标签直取必然拿错行 —— 生产解析器带章节正是为了这个。
    而 DARTs 那三行生产 spec 里**故意没有**（见文件头），所以在这里单独、显式地取，
    并且只在 `DARTS_SECTION` 这个章节名下面取，不碰任何别的行。
    """
    F = _fetch()
    wb = openpyxl.load_workbook(path, data_only=True)
    # 生产 spec 里没有 dats_*（故意的），拿它去解会因「缺列」直接报错 —— 先摘掉那三列再解。
    spec = {k: v for k, v in F.MONTHLY_SPEC.items() if v not in set(DARTS_MAP.values())}
    df = F.parse_sheet(wb['Monthly KPIs'], spec, 'M')

    rows, cols = _grid(path)
    hit, inside = {}, False
    for r in rows:
        lab = F._norm(next((c for c in r[:3] if c is not None and str(c).strip()), None))
        if not lab:
            continue
        numeric = any(isinstance(r[i], (int, float)) for i, _ in cols)
        if not numeric:
            inside = lab.startswith(DARTS_SECTION)
            continue
        if inside and lab in DARTS_MAP and DARTS_MAP[lab] not in hit:
            hit[DARTS_MAP[lab]] = [r[i] if isinstance(r[i], (int, float)) else None
                                   for i, _ in cols]
    missing = [c for c in DARTS_MAP.values() if c not in hit]
    if missing:
        raise HoodBasefillError(
            f'{os.path.basename(path)}: 「{DARTS_SECTION}」章节下取不到 {missing} —— '
            '这一份不是预期中的老版 Supplement，先人工看一眼')
    return df.join(pd.DataFrame(hit, index=[p for _, p in cols]), how='outer')


def _legacy_rows(labels, months):
    """老版式 → {Period: {列: 值}}。拿不到任何一列就报错（不静默少列）。"""
    got = {}
    for lab, col in LEGACY_MAP.items():
        src = next((labels[a] for a in LEGACY_ALIASES.get(lab, (lab,)) if a in labels), None)
        if src is None:
            raise HoodBasefillError(f'老版式里找不到「{lab}」→ {col}；官方改标签了，先人工看一眼')
        got[col] = src
    out = {}
    for p in months:
        rec = {c: v[p] for c, v in got.items() if v.get(p) is not None}
        if rec:
            out[p] = rec
    return out


# ─────────────────────────── 核对 ───────────────────────────
def _csv():
    with open(CSV_PATH, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    header = rows[0]
    body = [r for r in rows[1:] if r and r[0].strip()]
    return header, body


def check_modern(mod, header, body):
    """③ 与已有行逐格比。除 DARTs→DATs 那次重述外必须全同。"""
    have = {r[0]: dict(zip(header, r)) for r in body}
    n = bad = 0
    allowed, unexpected = [], []
    for key, row in sorted(have.items()):
        p = pd.Period(key, 'M')
        if p not in mod.index:
            continue
        for c in header[1:]:
            txt = (row.get(c) or '').strip()
            if not txt:
                continue
            n += 1
            v = mod.loc[p, c]
            if v is None or pd.isna(v) or abs(float(txt) - float(v)) > 1e-9:
                bad += 1
                item = (key, c, txt, v)
                (allowed if c in ('dats_equity_mn', 'dats_crypto_mn')
                 and p >= pd.Period('2025-01', 'M') else unexpected).append(item)
    print(f'\n── 看门狗 A：MODERN(Q1\'26) vs 已有 series 行，{n} 格 ──')
    print(f'  一致 {n - bad} 格；DARTs→DATs 重述 {len(allowed)} 格；'
          f'**意料之外的不一致 {len(unexpected)} 格**')
    if allowed:
        print(f'    重述样例 {allowed[0][0]} {allowed[0][1]}: 老版 {allowed[0][3]} → 现 {allowed[0][2]}'
              f'（共 {len({a[0] for a in allowed})} 个月，'
              f'列 {sorted({a[1] for a in allowed})}）')
    if unexpected:
        for it in unexpected[:12]:
            print(f'    {it[0]} {it[1]}: CSV={it[2]} Q1\'26={it[3]}')
        raise HoodBasefillError('MODERN 与已有行对不上（上面这些格）—— 这一份不是同一把尺子，别写')
    return len(allowed)


def check_darts(mod, header, body):
    """DARTs 与 DATs 在 2025-01 之前是否逐月相同 —— 这是「用 DARTs 填 dats_*」的全部依据。"""
    have = {r[0]: dict(zip(header, r)) for r in body}
    same = diff = 0
    last_same = None
    for key, row in sorted(have.items()):
        p = pd.Period(key, 'M')
        if p not in mod.index or p >= pd.Period('2025-01', 'M'):
            continue
        for c in ('dats_equity_mn', 'dats_options_mn', 'dats_crypto_mn'):
            txt = (row.get(c) or '').strip()
            if not txt:
                continue
            if abs(float(txt) - float(mod.loc[p, c])) <= 1e-9:
                same += 1
                last_same = key
            else:
                diff += 1
    print(f'\n── 看门狗 B：DARTs(Q1\'26) vs DATs(现) 在 2025-01 之前 ──')
    print(f'  逐位相同 {same} 格、不同 {diff} 格（重叠到 {last_same}）')
    if diff:
        raise HoodBasefillError('2025-01 之前 DARTs 与 DATs 就已经对不上 —— '
                                '「那一段两种口径等价」不成立，dats_* 三列应改为留空')
    return same


def check_seam(leg, mod):
    """老/新两种版式在 2023-01~2023-03 的重叠：12 列 x 3 个月，必须逐位相同。"""
    months = [pd.Period(f'2023-{m:02d}', 'M') for m in (1, 2, 3)]
    print('\n── 看门狗 C：老版式 vs 新版式在 2023-01~03 的重叠（%d 列 x %d 月）──'
          % (len(LEGACY_MAP), len(months)))
    bad, n = [], 0
    for p in months:
        for col in sorted(set(LEGACY_MAP.values())):
            a = leg.get(p, {}).get(col)
            b = mod.loc[p, col] if p in mod.index else None
            if a is None or b is None or pd.isna(b):
                bad.append((str(p), col, a, b))
                continue
            n += 1
            if abs(float(a) - float(b)) > 1e-9:
                bad.append((str(p), col, a, b))
    print(f'  逐位相同 {n} 格；对不上 {len(bad)} 格')
    if bad:
        for it in bad[:12]:
            print(f'    {it[0]} {it[1]}: 老={it[2]} 新={it[3]}')
        raise HoodBasefillError('两种版式在重叠月对不上 —— 行标签映射错了，别写')
    return n


def check_restatements(leg_labels, later, months):
    """① vs 后来的版本：哪些格被改过。**只报告**，入库一律取 ① 的原值。

    `later` 是 {版本名: 标签表}。要两份而不是一份：Q1'25 那份的窗口 2022-01 才开始，
    只拿它比会漏掉 2021 年被改过的月份（实测 2021-06 / 2021-09 各改过 0.1）。
    """
    print('\n── 看门狗 D：Q1\'23 原值 vs 后期版本（2021-01~2022-12）──')
    total = 0
    for lab, col in LEGACY_MAP.items():
        a = next((leg_labels[x] for x in LEGACY_ALIASES.get(lab, (lab,)) if x in leg_labels), None)
        if a is None:
            continue
        for vname, tbl in later.items():
            b = next((tbl[x] for x in LEGACY_ALIASES.get(lab, (lab,)) if x in tbl), None)
            if b is None:
                continue
            d = [(str(p), a[p], b[p]) for p in months
                 if a.get(p) is not None and b.get(p) is not None and a[p] != b[p]]
            total += len(d)
            if d:
                print(f'  {col:<30} vs {vname:<7} 改过 {len(d)} 个月：'
                      + '、'.join(f'{m} {x}→{y}' for m, x, y in d[:6]))
    print(f'  合计 {total} 处后期改动；本脚本入库取 Q1\'23 的原值'
          f'（仓规：不许拿后期重述值盖历史）')
    return total


# ─────────────────────────── 写盘 ───────────────────────────
def build_rows(leg, mod, seclend):
    """→ {Period: {列: 值}}，只含 < GUARD 的月份。"""
    rows = {}
    for p, rec in leg.items():
        rows.setdefault(p, {}).update(rec)
    for p in [pd.Period(f'2023-{m:02d}', 'M') for m in (1, 2, 3)]:
        rec = {c: mod.loc[p, c] for c in mod.columns
               if mod.loc[p, c] is not None and not pd.isna(mod.loc[p, c])}
        rows.setdefault(p, {}).update(rec)
    for p, v in seclend.items():
        rows.setdefault(p, {})['seclend_total_usdmn'] = v
    over = [str(p) for p in rows if p >= GUARD]
    if over:
        raise HoodBasefillError(f'这些月份 ≥ {GUARD}，本脚本不许写：{sorted(over)[:6]}')
    under = [str(p) for p in rows if p < FIRST]
    if under:
        raise HoodBasefillError(f'这些月份 < {FIRST}（官方月度披露的天花板）：{sorted(under)[:6]}')
    return rows


def write(rows, dry=False):
    F = _fetch()
    header, body = _csv()
    have = {r[0]: list(r) for r in body}
    cols = header[1:]
    existing = {c: [r[i + 1] for r in body] for i, c in enumerate(cols)}

    added, filled, skipped = [], 0, 0
    for p in sorted(rows):
        key = str(p)
        row = have.get(key) or ([key] + [''] * len(cols))
        if key not in have:
            added.append(key)
        for i, c in enumerate(cols):
            if c not in rows[p]:
                continue
            if (row[i + 1] or '').strip():        # 已有值永不覆盖
                skipped += 1
                continue
            row[i + 1] = F._fmt_like(existing[c], rows[p][c])
            filled += 1
        have[key] = row

    out = [have[k] for k in sorted(have)]
    blanks = {c: sum(1 for r in out if not (r[i + 1] or '').strip()) for i, c in enumerate(cols)}
    print('\n── 写盘 ──')
    print(f'  新建 {len(added)} 行、填 {filled} 格、跳过（已有值）{skipped} 格')
    print('  留空的列：'
          + ('、'.join(f'{c}×{n}' for c, n in sorted(blanks.items()) if n) or '（无）'))
    if dry:
        print('  --dry：**未**写入')
        for k in sorted(rows)[:2] + ['…'] + sorted(rows)[-1:]:
            print('    ' + (str(k) if k == '…' else ','.join(have[str(k)])))
        return 0
    with open(CSV_PATH, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(out)
    print(f'✓ {CSV_PATH}：现共 {len(out)} 行，{out[0][0]} → {out[-1][0]}')
    return len(added)


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry', action='store_true')
    ap.add_argument('--refresh', action='store_true')
    ap.add_argument('--cache-dir', default=os.path.join(ROOT, 'cache'))
    a = ap.parse_args(argv)

    print('── 下载 / 取缓存 ──')
    paths = download_all(a.cache_dir, a.refresh)

    mod = _modern_frame(paths['MODERN'])
    leg_labels = _read_labels(paths['LEGACY'])
    sec_labels = _read_labels(paths['SECLEND'])
    rst_labels = _read_labels(paths['RESTATE'])

    months = [pd.Period(f'{y}-{m:02d}', 'M') for y in (2021, 2022) for m in range(1, 13)]
    leg_all = _legacy_rows(leg_labels, months + [pd.Period(f'2023-{m:02d}', 'M')
                                                 for m in (1, 2, 3)])
    header, body = _csv()

    check_modern(mod, header, body)
    check_darts(mod, header, body)
    check_seam(leg_all, mod)
    check_restatements(leg_labels, {"Q4'23": sec_labels, "Q1'25": rst_labels}, months)

    if SECLEND_LABEL not in sec_labels:
        raise HoodBasefillError(f'SECLEND 源里找不到「{SECLEND_LABEL}」')
    sl = {p: v for p, v in sec_labels[SECLEND_LABEL].items()
          if p in months and v is not None}
    early = sorted(p for p in sl if p < SECLEND_FROM)
    if early:
        raise HoodBasefillError(f'{SECLEND_FROM} 之前竟然有出借收入数字：{early} —— 先查清口径')
    print(f'\n── 出借收入 ──\n  {min(sl)} 起共 {len(sl)} 个月有数'
          f'（{SECLEND_FROM} 之前官方印 NA，本脚本留空不补 0）')

    leg = {p: rec for p, rec in leg_all.items() if p in months}
    rows = build_rows(leg, mod, sl)
    return write(rows, a.dry)


if __name__ == '__main__':
    sys.exit(0 if main() is not None else 1)
