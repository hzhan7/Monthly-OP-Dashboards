# -*- coding: utf-8 -*-
"""CBOE 的**历史回填**：把 series/cboe.csv 从 2017-01 起推到 2016-01 起（+12 期）。

用法:
    python3 build/basefill/cboe_2016.py            # 取数 + 核对 + 写 CSV
    python3 build/basefill/cboe_2016.py --dry      # 只打印，不写文件
    python3 build/basefill/cboe_2016.py --refresh  # 强制重下工作簿

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
这个文件为什么存在，以及为什么**不**把它塞进 fetch/cboe.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
fetch/cboe.py 只认一套文件名模板（`{Month}-{Year}-Monthly-Volume-Statistics-Xlsx-.xlsx`）
+ 列表页，而列表页**永远只挂当前最新一期**。按那套模板往回扫 2022-01…2026-07 共 48 个月，
只有 2024-08 及之后返回 200，更早一律 403（Akamai 拿 403 当 404 用，见该模块 docstring）。
所以「往回补」在那条链路上做不到，不是代码写得不好，是那条链路就没有历史入口。

历史另有入口：2020 年官网改版时，旧的年度合订工作簿被整批迁到同一个 CDN 目录下、
换了一套文件名，至今 200（Last-Modified 全部停在 2020-10-12，是一批迁过来就没人管的
静态残留）。本脚本要的那一份是：

    https://cdn.cboe.com/resources/investor_relations/revenue_per_contract/
        monthly-volume-stats-worksheet-2016-combined.xlsx     （34,698 字节）

它与 fetch/cboe.py 每月抓的那份**不是同一种版式**（详见下面「版式差异」），
而且这一份只会被用这一次 —— 2016 年的数字不会再变。把一个只跑一次的 legacy 分支
焊进无人值守的抓取器里，等于让它每月都带着一段永远走不到的代码，
下一个人读 fetch/cboe.py 时还要先搞清楚那段是干嘛的。所以照 spgi_history.py 的先例
放在 build/basefill/ 下，一次性跑完就完事。
⚠ 与 mtk 那次不同：mtk 的历史是**同一个源、同一种版式**，只是抓取器的窗口没开，
  所以那次是改 update() 让它自己长得回去；这里是**另一个源、另一种版式**，
  一次性脚本才是对的。别把两种情形套用同一个结论。

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
版式差异（相对 fetch/cboe.py 认的那一套）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
· sheet 名是 `2016 (updated)`，不是月份名。
· 标签列在 **D 列**，表头行是**第 6 行**（`Period | Jan-16 … Dec-16 | 1Q16 … | Year`）。
· 两个行标签换过名，本脚本按别名映射（**只映射这两条，其余逐字相同**）：
      'U.S. Equities - ADV (matched shares, billions)'
          → 现名 'U.S. Equities - Exchange - ADV (matched shares, billions)'
      'U.S. Equities - per 100 touched shares'
          → 现名 'U.S. Equities - Exchange - per 100 touched shares'
· RPC 那一段的口径与今天**一致**：段标题是 `Rolling Three-Month Average RPC/Net Capture²`
  —— 即入库的 rpc_* 一直都是**滚动三月均**，不是当月值。这一点两代相同，不构成断点。
  （自检：本册 1Q16 列的 RPC 恰等于 Mar-16 列，符合「季末的滚动三月均 = 该季均值」。）

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
口径：2016 与 2017 一样，都是 Bats **pro-forma**，不是新断点
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
本册 D3 抬头与脚注 ¹ 原文：
    'Combined Information for CBOE Holdings Monthly Volume & Revenue Per Contract/
     Net Revenue Capture Report - 2016¹'
    'For informational purposes, the operating statistics for these periods are
     presented on a combined basis to reflect information pertaining to Bats Global
     Markets, Inc., which was acquired by CBOE Holdings, Inc. on February 28, 2017.'
也就是说 2016 与 2017 是**同一种 pro-forma 口径**，红色断点线的位置（首个 2018 月左缘）
不动；要改的是 build/cboe.py 里写死的年份文案（`PF_YEAR = 2017` 那一族），
否则会变成「图上画了两年 pro-forma、图注只认一年」。

⚠ **2015 及更早不要拼。** 那些是 CBOE 单体口径：Total Options ADV 只有 combined 的 73%、
RPC 高 32%，且 U.S. Equities / European Equities / Global FX 三条线整列不存在
（那是 Bats 带来的业务）。接上去等于在 2015/2016 之间造一个比并购本身还大的假跳变。
"""
import argparse
import csv
import os
import sys
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(os.path.dirname(HERE))
SERIES = os.path.join(ROOT, 'series')
CACHE = os.path.join(ROOT, 'cache', 'cboe_hist')

URL = ('https://cdn.cboe.com/resources/investor_relations/revenue_per_contract/'
       'monthly-volume-stats-worksheet-2016-combined.xlsx')
FNAME = 'monthly-volume-stats-worksheet-2016-combined.xlsx'
MIN_BYTES = 20_000
SHEET_MARK = '2016'                 # sheet 名里必须含它（实测 '2016 (updated)'）
HDR_ROW, LAB_COL = 6, 4             # 表头行 / 标签列（D）
MONTHS = [f'2016-{m:02d}' for m in range(1, 13)]

#: 目标列 → 本册里的行标签。顺序与 series/cboe.csv 的表头一致（除 month）。
#: xsp / minivix 不在这里：XSP 2019-01 才单列、Mini VIX 2020-08 才上市，
#: 本册的 'ADV for Select Index Products' 段只有 SPX / VIX options / VIX futures 三行。
ROWS = [
    ('adv_us_options_kcontracts',         'Total Options',                                  'adv'),
    ('rpc_us_options_usd',                'Total Options',                                  'rpc'),
    ('adv_futures_kcontracts',            'Futures - ADV (contracts, thousands)',           'adv'),
    ('rpc_futures_usd',                   'Futures - per contract',                         'rpc'),
    ('adv_us_equities_matched_shares_bn', 'U.S. Equities - ADV (matched shares, billions)', 'adv'),
    ('rpc_us_equities_usd_per100shares',  'U.S. Equities - per 100 touched shares',         'rpc'),
    ('adv_eu_equities_adnv_eurbn',        'European Equities - ADNV (€ billions)',          'adv'),
    ('rpc_eu_equities_bps',               'European Equities - per matched notional value (bps)', 'rpc'),
    ('adv_fx_adnv_usdbn',                 'Global FX - ADNV ($ billions)',                  'adv'),
    ('rpc_fx_usd_per_usdmn',              'Global FX - per one million dollars traded',     'rpc'),
    ('adv_multilist_options_kcontracts',  'Multiply-listed options (Equities & ETPs)',      'adv'),
    ('rpc_multilist_options_usd',         'Multiply-listed options (Equities & ETPs)',      'rpc'),
    ('adv_index_options_kcontracts',      'Index options',                                  'adv'),
    ('rpc_index_options_usd',             'Index options',                                  'rpc'),
    ('adv_spx_options_kcontracts',        'SPX options',                                    'idx'),
    ('adv_vix_options_kcontracts',        'VIX options',                                    'idx'),
    ('adv_vix_futures_kcontracts',        'VIX futures',                                    'idx'),
]
EMPTY_COLS = ['adv_xsp_options_kcontracts', 'adv_minivix_futures_kcontracts']

#: 段落锚点。**必须按段取行**：'Multiply-listed options (Equities & ETPs)' /
#: 'Index options' / 'Total Options' 这三个标签在本册里各出现 **三次**
#: （ADV 段、Market Share 段、RPC 段），只按标签找会取到市占率那一段的小数。
SECTIONS = {
    'adv': 'ADV/ADNV by Business Segment',
    'shr': 'Market Share by Business Segment',
    'rpc': 'Rolling Three-Month Average RPC',      # 前缀匹配（原文带脚注 ²）
    'idx': 'ADV for Select Index Products',
}


class Cboe2016Error(RuntimeError):
    pass


def _download(refresh=False):
    os.makedirs(CACHE, exist_ok=True)
    path = os.path.join(CACHE, FNAME)
    if os.path.exists(path) and not refresh and os.path.getsize(path) >= MIN_BYTES:
        return path
    req = urllib.request.Request(URL, headers={'User-Agent': 'Mozilla/5.0'})
    blob = urllib.request.urlopen(req, timeout=90).read()
    if len(blob) < MIN_BYTES:
        raise Cboe2016Error(f'{URL} 只回了 {len(blob)} 字节，不像工作簿')
    with open(path, 'wb') as f:
        f.write(blob)
    return path


def _lab(v):
    """标签归一化：去脚注上标数字与尾部 * / 空白。与 fetch/cboe.py 的 _lab 同精神。"""
    s = str(v or '').strip()
    while s and (s[-1] in '*¹²³' or s[-1].isdigit()):
        s = s[:-1].rstrip()
    return s


def parse(path):
    """→ {'YYYY-MM': {列: 值}}，共 12 个月。任何结构不符一律抛，不猜。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    names = [s for s in wb.sheetnames if SHEET_MARK in s]
    if len(names) != 1:
        raise Cboe2016Error(f'sheet 名里含 {SHEET_MARK!r} 的有 {names}，期待恰好一个')
    ws = wb[names[0]]

    # 表头：D6='Period'，E6.. 是 Jan-16…Dec-16
    if _lab(ws.cell(HDR_ROW, LAB_COL).value) != 'Period':
        raise Cboe2016Error(f'第 {HDR_ROW} 行 {LAB_COL} 列不是 "Period"，版式变了')
    col_of = {}
    for c in range(LAB_COL + 1, ws.max_column + 1):
        h = str(ws.cell(HDR_ROW, c).value or '').strip()
        if len(h) == 6 and h[3] == '-':                       # 'Jan-16'
            mon = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                   'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}.get(h[:3])
            if mon:
                col_of[f'20{h[4:]}-{mon:02d}'] = c
    missing = [m for m in MONTHS if m not in col_of]
    if missing:
        raise Cboe2016Error(f'表头里缺这些月份：{missing}')

    # 段落边界
    sec_at = {}
    for r in range(HDR_ROW, ws.max_row + 1):
        s = _lab(ws.cell(r, LAB_COL).value)
        for key, anchor in SECTIONS.items():
            if s.startswith(anchor):
                sec_at[key] = r
    if set(sec_at) != set(SECTIONS):
        raise Cboe2016Error(f'段落锚点只认出 {sorted(sec_at)}，期待 {sorted(SECTIONS)}')
    order = sorted(sec_at.items(), key=lambda kv: kv[1])
    bounds = {}
    for i, (key, r0) in enumerate(order):
        r1 = order[i + 1][1] if i + 1 < len(order) else ws.max_row + 1
        bounds[key] = (r0, r1)

    out = {m: {} for m in MONTHS}
    for col, label, sec in ROWS:
        r0, r1 = bounds[sec]
        hits = [r for r in range(r0 + 1, r1) if _lab(ws.cell(r, LAB_COL).value) == label]
        if len(hits) != 1:
            raise Cboe2016Error(
                f'{col}: 段 {sec} 内标签 {label!r} 命中 {len(hits)} 行（期待 1）')
        for m in MONTHS:
            v = ws.cell(hits[0], col_of[m]).value
            if not isinstance(v, (int, float)):
                raise Cboe2016Error(f'{col} {m} 取到 {v!r}，不是数')
            out[m][col] = float(v)
    return out


def _read_csv():
    path = os.path.join(SERIES, 'cboe.csv')
    with open(path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    return path, rows[0], [r for r in rows[1:] if r and r[0].strip()]


def check(data, header, body):
    """接缝体检。**只报告，不修改** —— 判断留给人。"""
    have = {r[0]: dict(zip(header, r)) for r in body}
    print(f'· 解析出 {len(data)} 个月 × {len(ROWS)} 列')
    # 1) 与既有 2017-01 的平滑度
    jan17 = have.get('2017-01')
    if not jan17:
        raise Cboe2016Error('series/cboe.csv 里没有 2017-01，接缝无从核对')
    print(f'{"列":42s} {"Dec-16":>14s} {"Jan-17":>14s} {"m/m":>9s}')
    worst = []
    for col, _, _ in ROWS:
        a, b = data['2016-12'][col], jan17.get(col, '')
        if b in ('', None):
            continue
        b = float(b)
        pct = (b / a - 1) * 100 if a else float('nan')
        worst.append((abs(pct), col, pct))
        print(f'{col:42s} {a:14.6f} {b:14.6f} {pct:+8.1f}%')
    worst.sort(reverse=True)
    print(f'· 接缝最大单月变动：{worst[0][1]} {worst[0][2]:+.1f}%')
    # 2) 自检：本册 RPC 是滚动三月均 ⇒ 3 月那格应等于 1Q16 那格（在源里，不在这里核，
    #    但可以核 12 个月都在合理量级、无 0、无负）
    for m in MONTHS:
        for col, _, _ in ROWS:
            v = data[m][col]
            if not (v > 0):
                raise Cboe2016Error(f'{m} {col} = {v}，非正数')
    print('· 12 × 17 = 204 格全部为正，无空、无零')


def write(data, dry=False):
    path, header, body = _read_csv()
    have = {r[0] for r in body}
    dup = [m for m in MONTHS if m in have]
    if dup:
        raise Cboe2016Error(f'这些月份已在库里，拒绝覆盖（本脚本只前插）：{dup}')
    idx = {c: i for i, c in enumerate(header)}
    new = []
    for m in MONTHS:
        row = [''] * len(header)
        row[0] = m
        for col, _, _ in ROWS:
            row[idx[col]] = f'{data[m][col]:.6f}'
        new.append(row)
    if dry:
        print('\n--dry：以下 12 行**未**写入')
        for r in new[:2] + ['…'] + new[-1:]:
            print('   ', r if isinstance(r, str) else ','.join(r)[:120])
        return 0
    body = new + body
    body.sort(key=lambda r: r[0])
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(body)
    print(f'✓ 已前插 {len(new)} 行 → {path}（现共 {len(body)} 行，'
          f'{body[0][0]} → {body[-1][0]}）')
    return len(new)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry', action='store_true')
    ap.add_argument('--refresh', action='store_true')
    a = ap.parse_args()
    path = _download(a.refresh)
    print(f'· 工作簿 {path}（{os.path.getsize(path):,} 字节）')
    data = parse(path)
    _, header, body = _read_csv()
    check(data, header, body)
    write(data, a.dry)
    return 0


if __name__ == '__main__':
    sys.exit(main())
