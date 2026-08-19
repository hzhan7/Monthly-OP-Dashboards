# -*- coding: utf-8 -*-
"""Cboe Global Markets (CBOE) 月度成交量与 RPC —— 无人值守抓取。

━━ 数据源 ━━
列表页 : https://www.cboe.com/us/options/market_statistics/monthly_volume_rpc_reports/
        页面上永远只挂**当前最新一期**的 xlsx / pdf，历史期不列出来。
下载直链: https://cdn.cboe.com/resources/investor_relations/revenue_per_contract/
          {Month}-{YYYY}-Monthly-Volume-Statistics-Xlsx-.xlsx
          例：June-2026-Monthly-Volume-Statistics-Xlsx-.xlsx

为什么先解析列表页、而不是直接按模板拼 URL：
  模板本身是稳定的（2025-12 / 2026-01 / 2026-05 / 2026-06 实测全部 200），
  但「本月到底发了没有」只有列表页知道 —— 未发布的月份直链返回 403 而不是 404，
  光靠拼 URL 无法区分「还没发」和「命名规则变了」。所以：
  列表页给出 latest，模板只在需要回补中间月份时兜底使用。
cdn.cboe.com 与 www.cboe.com 都不拦 urllib / curl（无 Cloudflare / PerimeterX 挑战），
普通 UA 即可，无需浏览器登录态 —— 满足无人值守。

━━ 发布节奏 ━━
次月第 3 个美股交易日（工作日 + 交易所假期顺延）。实测：
  2025-12 数据 → 2026-01-06 发布（1/1 假期，1/2、1/5、1/6）
  2026-01 数据 → 2026-02-04 发布
  2026-05 数据 → 2026-06-03 发布
  2026-06 数据 → 2026-07-06 发布（7/3 因 7/4 落在周六而休市）
xlsx 内 A2 单元格写着 "Updated on <date>"，是判断这一期新鲜度的权威字段；
它同时也是页面抬头「官方发布于」那半句的唯一合法来源，入库时记进
series/source_dates.csv（见 _record_source_dates）。

━━ 口径坑（按踩坑概率排序）━━
1. **RPC 滞后一个月，而且是三个月滚动平均**。同一份文件里，ADV 填到本月，
   RPC 只填到上月；且 RPC 不是单月数，是 rolling 3M。所以：
   · 新月份入库时 RPC 天然为空，这**不是解析失败**，不能因此抛异常；
   · 下个月的文件会把这一格补上 —— 本模块因此额外做「只填空、不覆盖」的回补，
     否则每个月的 RPC 会永久留白，build_cboe.py 的 ADV x RPC 图直接废掉。
2. **同名标签在三个 section 里各出现一次**。"Total Options" 同时是 ADV(row15)、
   市占率(row30)、RPC(row40) 的行名。所以查行必须先定位 section，
   绝不能全表 grep 标签名 —— 抓错了会把 0.30（市占率）当成 ADV 写进去。
3. **表头既有月份列也有季度列**（Jan-26 … Dec-26, 1Q26, 2Q26, …, Year TD）。
   只认 %b-%y 形状的列，否则 1Q26 会被当成一个月。
4. **期货 ADV 有定义断点**：2025Q2 起 Digital futures 并入 CFE，之前的
   adv_futures_kcontracts 不含数字资产期货。跨 2025Q2 的同比不可直接比。
5. **官方明说会重述**（"subject to change and revisions"）。因此本模块对已入库
   的历史值一律不覆盖，只在空格上回补；真要重刷历史请手工重跑全量。
6. 工作簿每年一张 sheet（名字就是 "2025" / "2026"），跨年那一期
   （比如 2027-01 发的 2026-12 报告）里只有上一年的 sheet，
   所以按「名字是 4 位数字」找 sheet，不要写死年份。
7. xsp / mini-VIX 两列是后来才有的（XSP 自 2019-01，Mini VIX 自 2020-08），
   2017-2018 的历史行天然为空；本模块只处理新月份，不受影响。
8. **12 月的 RPC 是个结构性窟窿**：月度 RPC 滞后一月 → 12 月的 RPC 本该由次年
   1 月那期给出，但 1 月那期只带新一年的 sheet，上一年 12 月那格再也不会出现。
   实测 2026-01 那期整张 RPC 区一格未填。解法：从同一期的 4Qxx 列取
   （季度 RPC ≡ 季末月的三月滚动 RPC，已验证），见 _fill_december_rpc。
9. **精度：本模块写的是 xlsx 里的完整浮点值，既有历史行大多是 6 位小数**。
   官方把 ADV 存成「总量 / 交易日数」的原始商（21 个交易日 → 22977.300428571423），
   Excel 只是显示成 6 位；series/cboe.csv 的历史行当初是照显示值抄的 6 位数，
   唯独 xsp / mini-VIX 两行留了完整精度。实测最后 3 个月：解析值 round 到 6 位后
   45/50 个单元格与 CSV 完全相同，最大相对偏差 2.46e-05（0.0025%，
   rpc_us_equities 0.017734 vs 0.0177344363）—— 纯粹是四舍五入痕迹，
   不是口径或解析错误。新月份保留完整精度（无损），图上按 0-3 位小数格式化，
   看不出任何差别；历史行一律不改写，所以接缝处不会出现值被"改动"的假象。

━━ 依赖 ━━ openpyxl（读 xlsx）。不依赖 pandas，避免 to_csv 重排既有行的格式。
"""

import csv
import os
import re
import urllib.request
from datetime import datetime

import openpyxl

LISTING_URL = ('https://www.cboe.com/us/options/market_statistics/'
               'monthly_volume_rpc_reports/')
CDN_BASE = ('https://cdn.cboe.com/resources/investor_relations/'
            'revenue_per_contract/')
FNAME_TMPL = '{month}-{year}-Monthly-Volume-Statistics-Xlsx-.xlsx'

# 用真实浏览器 UA：cdn.cboe.com 目前不校验，但 Akamai 策略随时可能收紧，
# 带上常规 UA 是零成本的保险。
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

# ── 表结构：(csv 列名, section, 该 section 内的行标签) ─────────────────────
# section 用工作簿里的大标题定位，见「口径坑 2」。
S_ADV = 'ADV/ADNV by Business Segment'
S_RPC = 'Rolling Three-Month Average RPC'
S_IDX = 'ADV for Select Index Products'

COLUMN_SPEC = [
    # --- ADV（当月实数，随文件一起到最新月）---
    ('adv_us_options_kcontracts',          S_ADV, 'Total Options'),
    ('adv_futures_kcontracts',             S_ADV, 'Futures - ADV (contracts, thousands)'),  # 原文带脚注 *
    ('adv_us_equities_matched_shares_bn',  S_ADV, 'U.S. Equities - Exchange - ADV (matched shares, billions)'),
    ('adv_eu_equities_adnv_eurbn',         S_ADV, 'European Equities - ADNV (€ billions)'),
    ('adv_fx_adnv_usdbn',                  S_ADV, 'Global FX - ADNV ($ billions)'),
    ('adv_multilist_options_kcontracts',   S_ADV, 'Multiply-listed options (Equities & ETPs)'),
    ('adv_index_options_kcontracts',       S_ADV, 'Index options'),
    ('adv_spx_options_kcontracts',         S_IDX, 'SPX options'),
    ('adv_vix_options_kcontracts',         S_IDX, 'VIX options'),
    ('adv_vix_futures_kcontracts',         S_IDX, 'VIX futures'),
    ('adv_xsp_options_kcontracts',         S_IDX, 'XSP options'),
    ('adv_minivix_futures_kcontracts',     S_IDX, 'Mini VIX futures'),
    # --- RPC（滚动三月均，滞后一月）---
    ('rpc_us_options_usd',                 S_RPC, 'Total Options'),
    ('rpc_futures_usd',                    S_RPC, 'Futures - per contract'),
    ('rpc_us_equities_usd_per100shares',   S_RPC, 'U.S. Equities - Exchange - per 100 touched shares'),
    ('rpc_eu_equities_bps',                S_RPC, 'European Equities - per matched notional value (bps)'),
    ('rpc_fx_usd_per_usdmn',               S_RPC, 'Global FX - per one million dollars traded'),
    ('rpc_multilist_options_usd',          S_RPC, 'Multiply-listed options (Equities & ETPs)'),
    ('rpc_index_options_usd',              S_RPC, 'Index options'),
]

ADV_COLS = [c for c, s, _ in COLUMN_SPEC if s != S_RPC]
RPC_COLS = [c for c, s, _ in COLUMN_SPEC if s == S_RPC]

# section 锚点用前缀匹配：官方标题带脚注上标（"…Net Capture1"），全等会失配。
_SECTION_ANCHORS = [S_ADV, 'Market Share by Business Segment', S_RPC, S_IDX,
                    'FX Rates']
# 注：COLUMN_SPEC 里的 label 一律写成 _lab() 归一化后的形态（无尾部 * 和脚注数字）

_MONTH_HDR = re.compile(r'^[A-Z][a-z]{2}-\d{2}$')
_QTR_HDR = re.compile(r'^([1-4])Q(\d{2})$')
_UPDATED_ON = re.compile(r'^Updated\s+on\s+(.+?)\s*$', re.I)
_XLSX_LINK = re.compile(
    r'https://cdn\.cboe\.com/resources/investor_relations/revenue_per_contract/'
    r'([A-Z][a-z]+)-(\d{4})-Monthly-Volume-Statistics-Xlsx-\.xlsx')


class CboeFetchError(RuntimeError):
    """源站结构变化 / 下载失败 / 解析结果不完整。一律炸掉，绝不静默写 NaN。"""


# ── 网络 ────────────────────────────────────────────────────────────────
def _http_get(url, timeout=60):
    req = urllib.request.Request(url, headers={
        'User-Agent': _UA,
        'Accept': '*/*',
        'Accept-Language': 'en-US,en;q=0.9',
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.read()
    except Exception as e:                                # noqa: BLE001
        raise CboeFetchError('下载失败 %s: %r' % (url, e)) from e


def _discover_latest(cache_dir):
    """从列表页拿最新一期的 (url, 'YYYY-MM' 形式的报告月, 本地文件名)。

    列表页只挂最新一期，所以这里同时也就拿到了「官方当前最新月」的答案 ——
    但真正写进 CSV 的最新月仍以工作簿内容为准（见 latest_month）。
    """
    html = _http_get(LISTING_URL).decode('utf-8', 'replace')
    # 存一份原始页面，源站改版时可以事后取证
    _write_bytes(os.path.join(cache_dir, 'cboe_listing.html'),
                 html.encode('utf-8'))
    hits = _XLSX_LINK.findall(html)
    if not hits:
        raise CboeFetchError(
            '列表页上找不到 Monthly-Volume-Statistics xlsx 直链，源站可能改版：'
            + LISTING_URL)
    # 页面上可能挂着补充历史文件，取报告月最大的那个
    best = max(hits, key=lambda h: _mon_key(h[0], h[1]))
    month_name, year = best
    fname = FNAME_TMPL.format(month=month_name, year=year)
    return CDN_BASE + fname, _mon_key(month_name, year), fname


def _mon_key(month_name, year):
    return '%s-%02d' % (year, datetime.strptime(month_name, '%B').month)


def _fetch_report(cache_dir, month_name=None, year=None):
    """下载某一期 xlsx，返回 (本地路径, 这一期**自报**的报告月 'YYYY-MM')。

    第二个返回值是本模块唯一一条**独立于工作簿解析器**的月份判据：
      · 取最新一期时，它来自列表页挂出的文件名（官方说「当前最新是这一期」）；
      · 回补指定月时，它就是我们点名要的那一期。
    两种来源都与文件内容同批发布，不存在「新闻稿先出、文件后更新」那种时间差，
    所以拿它和 _validate() 解析出来的最新月对不上时，只可能是解析出了问题
    （见 _crosscheck_report_month）。曾经这个值被写成 `_mon` 就地丢弃 —— 判据
    明明已经拿在手里却没用，等于把 msci_silent_parse_miss 那类静默漏抓的门敞着。
    """
    os.makedirs(cache_dir, exist_ok=True)
    if month_name is None:
        url, mon, fname = _discover_latest(cache_dir)
    else:
        fname = FNAME_TMPL.format(month=month_name, year=year)
        url = CDN_BASE + fname
        mon = _mon_key(month_name, year)
    path = os.path.join(cache_dir, fname)
    _write_bytes(path, _http_get(url))
    return path, mon


def _crosscheck_report_month(expected, newest, path, where):
    """这一期自报的报告月 vs 工作簿里解析出来的最新 ADV 月，不等就炸。

    防的是这一类：官方改了表头写法（比如给 'Jul-26' 加了脚注、或换成 'Jul 26'），
    _month_columns 认不出那一列于是静默丢掉，_validate 拿剩下的月份算出 newest =
    上个月，fetch 干干净净报 NOCHANGE —— 没有 FAIL、没有红点、streaks 也不动，
    整条链上没有任何人会发现少了一个月。MSCI 2026-07 就是这么漏的
    （见 resolved_issues.msci_silent_parse_miss）。

    这里刻意 raise 而不是 print warn：warn 之后状态仍是 NOCHANGE，等于没有护栏。
    """
    if expected != newest:
        raise CboeFetchError(
            '%s：这一期自报的报告月是 %s，但工作簿里解析出来的最新 ADV 月是 %s。'
            '文件 %s。二者同批发布、不该不一致 —— 最可能是月份表头写法变了导致'
            '整整一列被静默丢弃（_MONTH_HDR 只认 %%b-%%y 形状），也可能是官方换了'
            '报告月的定义。拒绝写入，请人工看一眼工作簿第 4 行的表头。'
            % (where, expected, newest, os.path.basename(path)))


def _write_bytes(path, data):
    with open(path, 'wb') as f:
        f.write(data)


# ── 解析 ────────────────────────────────────────────────────────────────
def _norm(v):
    return re.sub(r'\s+', ' ', str(v)).strip() if v is not None else ''


def _lab(v):
    """行标签归一化：砍掉尾部脚注记号。

    官方在部分行名后面挂脚注 —— 'Futures - ADV (contracts, thousands)*'
    （数字期货 2025Q2 并入 CFE 的说明）、'…Net Capture1'。这些记号历史上加过也
    去过，用全等匹配会在某个月突然全表失配。约定的行名里没有一个以数字结尾，
    所以直接右剥 '*' 和数字是安全的。
    """
    return _norm(v).rstrip(' *¹²³0123456789')


def _locate_rows(ws):
    """返回 {(section, label): row}。section 之间靠大标题切段，见「口径坑 2」。"""
    out = {}
    section = None
    for r in range(1, ws.max_row + 1):
        lab = _lab(ws.cell(r, 1).value)
        if not lab:
            continue
        anchor = next((a for a in _SECTION_ANCHORS if lab.startswith(a)), None)
        if anchor:
            section = anchor
            continue
        if section:
            out.setdefault((section, lab), r)
    return out


def _month_columns(ws):
    """{'YYYY-MM': col}。只认 Jan-26 这种形状，季度列 / Year TD 一概丢掉。"""
    out = {}
    for c in range(2, ws.max_column + 1):
        h = _norm(ws.cell(4, c).value)
        if _MONTH_HDR.match(h):
            d = datetime.strptime(h, '%b-%y')
            out['%04d-%02d' % (d.year, d.month)] = c
    return out


def _quarter_columns(ws):
    """{'YYYY-Q4': col}。只取 1Q26 这种季度列，用于补 12 月 RPC，见 _fill_december_rpc。"""
    out = {}
    for c in range(2, ws.max_column + 1):
        m = _QTR_HDR.match(_norm(ws.cell(4, c).value))
        if m:
            out['%04d-Q%s' % (2000 + int(m.group(2)), m.group(1))] = c
    return out


def _fill_december_rpc(ws, rows, data, sheet_year):
    """用 4Qxx 列补 12 月的 RPC —— 否则这一格永远补不上。

    结构性窟窿：月度 RPC 滞后一个月，所以 12 月的 RPC 要等 1 月的报告；
    但 1 月的报告只带新一年的 sheet（2027 年 1 月发的文件里只有 '2027'），
    上一年 12 月那一格再也不会出现在任何一期文件里。实测 2026-01 那期：
    整张 RPC 区一格未填，且没有 2025 的 sheet —— Dec-2025 的 RPC 无处可取。

    补法的依据（已实测）：季度列的 RPC 就是该季末月的三月滚动 RPC。
      · Mar-25/1Q25、Jun-25/2Q25、Sep-25/3Q25 七个 RPC 行两两相等，
        差异 ≤2e-15（纯浮点噪声）。
      · 季度列不受一个月滞后影响：2026-06 那期里 Jun-26 月度格是空的，
        但 2Q26 列已是完整的 Apr-Jun 口径 —— 用 Year TD 按合约数加权反推，
        (YTD_Jan-Jun x C_Jan-Jun - 1Q26 x C_Q1) / C_Apr-Jun = 0.31690231541,
        与官方 2Q26 完全一致，说明 6 月的费用数据其实已经在里面了。

    季度列确实含季末那个月，不是 QTD —— 用 Dec-2025 那期做过闭合检验：
      按合约数把 1Q-4Q 加权（Q4 记作 Oct+Nov+Dec）得到的 Year TD = 0.29651938，
      与官方 Year TD 相对偏差 3.7e-16；若把 Q4 当成只有 Oct+Nov，偏差 6.0e-03。
      差 13 个数量级，结论没有歧义。

    只对 12 月这么干，不对 3/6/9 月这么干：3/6/9 月的月度格下个月自然会填上同一个数
    （已验证两者相等到 2e-15），没必要多走一条代码路径去猜；只有 12 月是真的
    没有别的来源。代码路径越窄，出错面越小。

    ⚠ 已知不一致：series 里现存的 2025-12 那行 RPC 与本函数算出的 4Q25 值差
    2e-7 ~ 2.4e-2（rpc_us_equities 0.017661 vs 0.018093 差 2.45%）。上面的闭合检验
    说明 4Q25 才是官方口径的 12 月滚动 RPC，那一行大概率是当初从别处（季报/10-K）
    抄的。本模块不改历史，所以这条不一致会留在 2025-12；以后新的 12 月按官方口径写。
    """
    key = '%s-12' % sheet_year
    if key not in data or all(data[key][c] is not None for c in RPC_COLS):
        return
    qcol = _quarter_columns(ws).get('%s-Q4' % sheet_year)
    if qcol is None:
        return
    for name, sec, lab in COLUMN_SPEC:
        if sec == S_RPC and data[key][name] is None:
            data[key][name] = _cell_num(ws, rows[(sec, lab)], qcol)


def _cell_num(ws, row, col):
    v = ws.cell(row, col).value
    if v is None or (isinstance(v, str) and v.strip() in ('', '-', 'n/a', 'N/A')):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        raise CboeFetchError(
            '单元格 R%dC%d 不是数字：%r' % (row, col, v))


def parse_workbook(path):
    """解析一份官方 xlsx，返回 {'YYYY-MM': {csv列名: float|None}}（按月升序）。

    任何一个约定的行标签找不到 —— 说明官方改了表结构 —— 直接抛异常，
    宁可整月不更新，也不要写出一列悄悄全空的 CSV。
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheets = [s for s in wb.sheetnames if s.isdigit() and len(s) == 4]
    if not sheets:
        raise CboeFetchError('%s 里没有年份 sheet（拿到 %r）'
                             % (os.path.basename(path), wb.sheetnames))

    data = {}
    for sn in sorted(sheets):
        ws = wb[sn]
        rows = _locate_rows(ws)
        missing = [(sec, lab) for _c, sec, lab in COLUMN_SPEC
                   if (sec, lab) not in rows]
        if missing:
            raise CboeFetchError(
                'sheet %s 缺行标签，官方表结构可能已变：%s' % (sn, missing))
        cols = _month_columns(ws)
        if not cols:
            raise CboeFetchError('sheet %s 第 4 行找不到 %%b-%%y 形状的月份表头' % sn)
        for mon, c in cols.items():
            rec = {}
            for name, sec, lab in COLUMN_SPEC:
                rec[name] = _cell_num(ws, rows[(sec, lab)], c)
            # 整月全空 = 未来月份的占位列，跳过
            if all(v is None for v in rec.values()):
                continue
            data[mon] = rec
        _fill_december_rpc(ws, rows, data, sn)
    if not data:
        raise CboeFetchError('%s 解析后没有任何有数的月份' % os.path.basename(path))
    return dict(sorted(data.items()))


def _validate(data):
    """ADV 必须齐；RPC 只允许在「最新 ADV 月」上为空（官方一个月滞后）。"""
    adv_months = sorted(m for m, r in data.items()
                        if r['adv_us_options_kcontracts'] is not None)
    if not adv_months:
        raise CboeFetchError('解析结果里没有任何一个月有 Total Options ADV')
    newest = adv_months[-1]
    for mon in adv_months:
        rec = data[mon]
        bad = [c for c in ADV_COLS if rec[c] is None]
        # XSP 自 2019-01、Mini VIX 自 2020-08 才有，早期历史行天然为空
        if mon < '2019-01':
            bad = [c for c in bad if c != 'adv_xsp_options_kcontracts']
        if mon < '2020-08':
            bad = [c for c in bad if c != 'adv_minivix_futures_kcontracts']
        if bad:
            raise CboeFetchError('%s 缺 ADV 列 %s —— 解析异常，拒绝写入' % (mon, bad))
        if mon != newest:
            miss = [c for c in RPC_COLS if rec[c] is None]
            if miss:
                raise CboeFetchError(
                    '%s（非最新月）缺 RPC 列 %s —— 官方 RPC 只滞后一个月，'
                    '这里为空说明解析错行或口径变了' % (mon, miss))
    return newest


# ── 发布日 ───────────────────────────────────────────────────────────────
def _source_dates():
    """按路径加载仓库根的 source_dates.py。

    不能裸 import：本模块被 monthly_run 用 spec_from_file_location 加载，
    那时 sys.path 上既没有 fetch/ 也没有仓库根。
    """
    import importlib.util
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(root, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _updated_on(path):
    """取这一期工作簿**自己写的**发布日，返回 ('YYYY-MM-DD', 出处文字)；取不到返回 (None, None)。

    页面抬头「官方发布于 X」是一句关于外部世界的事实断言，只能由源头自己说出口 ——
    这里就是年份 sheet 的 A2「Updated on August 5, 2026」（模块 docstring 里
    称它为「判断这一期新鲜度的权威字段」）。构建日、文件 mtime、按「次月第 3 个交易日」
    推算出来的那天，看上去都一样体面，但都是我们编的，一律不许顶替。

    月名两种写法都真实出现过 —— 2026-07 那期是 'August 5, 2026'，2026-03 那期是
    'Apr 6, 2026' —— 所以 %B / %b 都要试。官方哪天换成别的格式，宁可返回 None
    让这半句缺席，也不要猜。
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            # 只扫表头区：正文里若有别的 "Updated on" 字样（脚注之类）不该抢答
            for r in range(1, min(6, ws.max_row) + 1):
                raw = _norm(ws.cell(r, 1).value)
                m = _UPDATED_ON.match(raw)
                if not m:
                    continue
                for fmt in ('%B %d, %Y', '%b %d, %Y'):
                    try:
                        d = datetime.strptime(m.group(1), fmt)
                    except ValueError:
                        continue
                    return d.strftime('%Y-%m-%d'), (
                        '%s 工作表 "%s" 第 %d 行 A 列 "%s"'
                        % (os.path.basename(path), sn, r, raw))
                return None, None       # 认得出这行、读不懂日期 → 宁缺勿猜
    finally:
        wb.close()
    return None, None


def _record_source_dates(series_dir, pub, ingested):
    """把各期自述的发布日记进 series/source_dates.csv。

    pub: {'YYYY-MM': (xlsx 路径, 该期自述发布日, 出处文字)}；ingested: 已在 CSV 里的月份集合。

    只对已经真的落进 series/cboe.csv 的月份作证 —— 解析炸了、写盘没成，台账上就不该
    多出一行说「这个月官方发过了」。
    已有记录一律不覆盖：官方明说会重述（口径坑 5），重述版工作簿的 Updated on 是改稿日
    不是首发日，覆盖会把「7 月数据 8/5 发布」悄悄改成某个更晚的日子，而页面照印不误。
    """
    sd = _source_dates()
    for mon in sorted(pub):
        _path, day, evidence = pub[mon]
        if mon not in ingested or not day:
            continue
        if sd.lookup(series_dir, 'cboe', mon):
            continue
        sd.record(series_dir, 'cboe', mon, day, evidence)


# ── 对外接口 ─────────────────────────────────────────────────────────────
def latest_month(cache_dir):
    """官方源当前最新月 'YYYY-MM'（以 ADV 口径为准，RPC 比它再滞后一个月）。

    抓不到 / 解析不出来一律抛 CboeFetchError，不返回 None 掩盖故障。
    （签名里的 -> str | None 是仓库统一约定；本实现只在成功时返回字符串。）
    """
    path, official = _fetch_report(cache_dir)
    data = parse_workbook(path)
    newest = _validate(data)
    _crosscheck_report_month(official, newest, path, 'latest_month')
    return newest


def _fmt(v):
    """写回 CSV 用最短往返表示，和既有行的风格一致（repr(float)）。"""
    return '' if v is None else repr(float(v))


def update(series_dir, cache_dir):
    """把新月份追加进 series/cboe.csv，返回新增月份列表（升序）。

    幂等保证：
      · 已存在的月份不重复追加；
      · 已经有值的单元格**永不覆盖**（官方会重述历史，重述不由本模块自动吞进来）；
      · 只对既有行里**原本为空**的 RPC 单元格做回补 —— 不做的话，每个月因一个月
        滞后而留白的 RPC 会永久为空，build_cboe.py 的「ADV x RPC」图就没数据了。
        回补不计入返回值（它不是新月份）。
    """
    csv_path = os.path.join(series_dir, 'cboe.csv')
    with open(csv_path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
    idx = {name: i for i, name in enumerate(header)}

    unknown = [c for c, _s, _l in COLUMN_SPEC if c not in idx]
    if unknown:
        raise CboeFetchError('series/cboe.csv 里没有这些列：%s' % unknown)

    path, official = _fetch_report(cache_dir)
    data = parse_workbook(path)
    newest = _validate(data)
    _crosscheck_report_month(official, newest, path, 'update 主抓')

    # 一期文件只为**它自己的最新月**的发布日作证。同一张表里更早的那些月是更早那几期
    # 发出来的，这份文件的 Updated on 与它们无关 —— 顺手给它们都盖上这个日期，
    # 等于凭空发明「4 月的数据是 8 月 5 日发的」。
    pub = {newest: (path,) + _updated_on(path)}

    have = {r[0]: r for r in body}

    # 跨年那一期只带上一年的 sheet（口径坑 6）：若 CSV 末月与本期之间还有窟窿，
    # 按模板把中间月份的报告一并拉下来补上。
    if have:
        last_csv = max(have)
        need = [m for m in _month_range(last_csv, newest) if m not in data]
        for mon in need:
            y, mm = int(mon[:4]), int(mon[5:])
            extra, extra_official = _fetch_report(
                cache_dir, month_name=datetime(y, mm, 1).strftime('%B'), year=y)
            more = parse_workbook(extra)
            more_newest = _validate(more)
            _crosscheck_report_month(extra_official, more_newest, extra,
                                     'update 回补 %s' % mon)
            pub.setdefault(more_newest, (extra,) + _updated_on(extra))
            for k, v in more.items():
                data.setdefault(k, v)

    added = []
    for mon in sorted(data):
        rec = data[mon]
        if rec['adv_us_options_kcontracts'] is None:
            continue                       # 只有 RPC 没有 ADV = 未来月，不建行
        if mon in have:
            row = have[mon]
            for name, _s, _l in COLUMN_SPEC:      # 只填空，不覆盖
                if not row[idx[name]].strip() and rec[name] is not None:
                    row[idx[name]] = _fmt(rec[name])
            continue
        row = [''] * len(header)
        row[0] = mon
        for name, _s, _l in COLUMN_SPEC:
            row[idx[name]] = _fmt(rec[name])
        have[mon] = row
        body.append(row)
        added.append(mon)

    # 无条件落盘：即便没有新月份，上面的 RPC 回补也可能改了既有行。
    # 未被触碰的单元格是原样字符串搬运，所以「什么都没变」时文件字节级不变。
    body.sort(key=lambda r: r[0])
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(body)

    # 记发布日放在落盘之后：写盘失败就没有「这个月官方发过了」这条断言。
    _record_source_dates(series_dir, pub, set(have))
    return sorted(added)


def _month_range(start, end):
    """('2026-06','2026-09') -> ['2026-07','2026-08','2026-09']（不含 start）。"""
    y, m = int(start[:4]), int(start[5:])
    out = []
    while True:
        m += 1
        if m == 13:
            y, m = y + 1, 1
        cur = '%04d-%02d' % (y, m)
        if cur > end:
            return out
        out.append(cur)


if __name__ == '__main__':
    _here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    print('latest:', latest_month(os.path.join(_here, 'cache')))
    print('added :', update(os.path.join(_here, 'series'),
                            os.path.join(_here, 'cache')))
