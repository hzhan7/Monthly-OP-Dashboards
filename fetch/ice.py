# -*- coding: utf-8 -*-
"""Intercontinental Exchange (ICE) 月度经营统计 —— 无人值守抓取。

ICE 在本仓的位置和别家不一样：它是**唯一一家把行业分母带进来的公司**。
`adv_tapeA/B/C_consolidated_mnsh`（Tape A/B/C 全市场合并成交量）与
`adv_us_equity_options_industry_kcontracts`（全美股票/ETF 期权行业总量）是全仓仅有的两个
官方口径的「全行业总量」，有了它们，NYSE 与 Cboe 的份额第一次能放在**同一个分母**下比，
而不是各报各的。没有 ICE，北美页只能算「仓内几家之和里的占比」—— 那个数会随着仓内
成员增减而变，不是市场份额。所以本模块的两个 consolidated / industry 列，
重要性高于 ICE 自己的量。

━━ 数据源 ━━
指针 : https://ir.theice.com/feed/ContentAsset.svc/GetContentAssetList
       (LanguageId=1&contentAssetTypeList=Supplemental Information&year=-1&pageSize=500…)
       取 Title == "Monthly Statistics Tracking" 那条的 FilePath。
主源 : 该 FilePath 指向的 xlsx（2026-08-06 实测
       s2.q4cdn.com/154085107/files/doc_downloads/2026/07/2011-2026-Monthly-Stats-July-2026_vF.xlsx，
       377,178 B，sha256 b091b514d1…）—— **一份文件带 2011-01 至今的全序列**，
       四个 sheet、55 个字段、187 个月，没有断档。首次建库一次写满，之后每月追加一行。
发布日: https://ir.theice.com/feed/PressRelease.svc/GetPressReleaseList
       取标题形如 "Intercontinental Exchange Reports <Month> [YYYY] Statistics" 的那条，
       字段 PressReleaseDate 就是 ICE 自己对外宣告的发布日（见「发布节奏」）。
       同一条目的 DocumentPath 是新闻稿全文 PDF，本模块拿它做 y/y 回校（见 _crosscheck_release）。

**为什么必须走 feed 而不是解析 HTML 页，也不是猜 CDN 路径**（两个理由都不是"被拦"）：

  1. HTML 页里根本没有 xlsx 直链。2026-08-06 实测
     `ir.theice.com/ir-resources/supplemental-information` 与
     `.../investor-resources/supplemental-information/default.aspx` 纯 urllib **都是 200**
     （Cloudflare 在前面，但没下挑战），把返回的 64KB HTML 全文 grep
     `Monthly-Stats[^"']*\\.xlsx` —— **0 个命中**。那张列表是前端 JS 拿同一个
     ContentAsset feed 客户端渲染出来的。
     ⚠ 不要把这一段读成「ICE 的 HTML 页 = Cloudflare 403」。那是错的。侦察阶段短时间打了
     几十次触发了限速，被误记成结构性拦截；照那个结论写下去，下一个人会去搬
     curl_cffi / Chrome MCP，把一条 3 秒的纯标准库链换成需要浏览器的东西，直接毁掉无人值守。
  2. CDN 路径规则不自洽，猜不得。June-2026 与 July-2026 两期**都**在 `/2026/07/` 下，
     而 May-2026 在 `/2026/06/` 下（实测 `/2026/07/…May-2026_vF.xlsx` = 404）；
     后缀 2020 年是 `_v1`、2026 年是 `_vF`。feed 里的 FilePath 是 ICE 自己维护的
     「当前最新」指针，跟着它走才是无人值守。

**不要写死 apiKey。** 页面 JS 里有一个 `Q4ApiKey`，但本模块全程不带 key，
两个 feed 与 CDN 三个请求 2026-08-06 实测全部 200。带上反而埋一个「key 轮换 → 静默失败」的雷。

整条链只有两个域：`ir.theice.com`（ICE 的 IR 站）与 `s2.q4cdn.com/154085107/`
（ICE 自己的 IR CDN，154085107 是 ICE 的 Q4 站点号）。没有任何第三方聚合站。

━━ 发布节奏 ━━
**次月第 3 个美股交易日**（假期顺延），与 Cboe 同节奏。

不是抄来的：本模块作者逐年扫了 PressRelease feed 2015-2026，匹配到 **124 期**月度统计
新闻稿（2015-09 数据月起，之前 ICE 的 IR feed 里没有这类稿），发布日的「几号」分布是

    3 号 44 期 | 4 号 15 期 | 5 号 46 期 | 6 号 19 期 | **最晚 6 号，无一例外**

所以 `build/roster.py` 的 LAG 填 **(6, 6)** —— 常规月与季末月同节奏，ICE 没有季末月延后
（季末月标题变成 "Reports June and Second Quarter 2026 Statistics"，日子不变）。
Cboe 同样是「次月第 3 个交易日」却配 (4,4)：那是按 32 期样本定的；ICE 这里配 6 是因为
124 期样本里实测触到过 6 号 19 次，配 4 会让红点每年误报好几回。
`monthly_run.EARLY = 5` ⇒ 闸门在月末次日就开，代价是每月多打几个空请求。

发布日的权威来源只有一个：**新闻稿的 PressReleaseDate**。三条候选实测对比过：

  · 新闻稿 PressReleaseDate  —— ICE 自己对外宣告的日子，唯一合法。
  · CDN 的 HTTP Last-Modified —— 2026-07 期 "Wed, 05 Aug 2026 12:30:32 GMT"，与新闻稿同日，
    可作互证；单独用它不够格（那是文件上传时刻，不是 ICE 的公开宣告）。
  · 工作簿 docProps/core.xml 的 dcterms:modified —— 内部存盘时刻，**会早一天**
    （2020-07 那期存于 08-04、发布于 08-05）。只能当旁证。

⚠ 工作簿里**没有**任何自述发布日字符串。整份 sharedStrings 只有 181 条，逐条扫过，
唯一带日期的是口径脚注 "Cash Equities ADV includes CHX volumes as of 7/18/2018"。
Cboe 那种 A2 "Updated on August 5, 2026" 在 ICE 这里不存在，别再找了。

━━ 口径坑（按踩坑概率排序）━━

1. **月度单元格全部四舍五入到整千张 / 整百万股，所以恒等式一律用容差、不要用相等。**
   10,026 个月度单元格里只有 2,429 个非整数，且全部集中在 RPC 与份额那 13 列。
   六个各自取整的加数，和的误差自然能到 ±2。实测（187 个月）：

       TOTAL ENERGY  = 6 子项之和   严格相等只有  85/187，max|差| = 2
       TOTAL AG      = sugar+other  严格相等     140/187，max|差| = 1
       TOTAL COMM    = energy+ag    严格相等     138/187，max|差| = 1
       TOTAL FIN     = 4 子项之和   严格相等     109/187，max|差| = 2
       OI COMM / OI FIN 同理        严格相等 134/187、139/187，max|差| = 1

   写成 `== ` 会在 187 个月里失败 102 次 —— 每月体检天天红，人两周就学会无视它。
   所以 `_validate()` 里这几条一律 `abs(diff) <= 3`（6 项和与 4 项和给 3 最稳）。
   页面侧的推论：**月度 ADV 按 0 位小数格式化，别标小数位**，那是假精度；
   要精确季度数就去读工作簿的 `1Q26`/`2Q26` 季度列（那些是全精度浮点，
   例 2Q26 Financials ADV = 4996.292370711726），不要拿月度值加权反推。

2. **同名标签在三个 section 里各出现一次，全表 grep 必然抓错段。**
   `TOTAL COMMODITIES` / `TOTAL FINANCIALS` / `Energy` / `Agriculture & Metals`
   在 Derivs sheet 的 ADV、RPC、OI 三段里各有一行 —— grep 到 RPC 那行会把 1.92
   当成 ADV 写进去，而 1.92 看上去完全像个正常数字。`Cash Products` 更狠：
   `Handled Volume` / `Matched Volume` / `Share of Total Matched Consolidated Volume`
   在 Tape A/B/C 三个小节里**各出现三次**。
   本模块的解法见 `_blocks()`：**日期表头行自己就是分段符**。表里凡是「一行里有 ≥3 个
   datetime 单元格」的，就是一个新数据块的表头；块与块之间不会串。三个 sheet 的表头行
   位置各不相同（Derivs 5/58/69、US Equity Options 5/15、Cash Products **4**/34、
   CDS Clearing 5），硬编行号迟早会炸，用这个办法一行都不用写死。
   块的身份再用它上方 8 行内的标题文字认（'MONTHLY ADV' / 'RATE PER CONTRACT' / …），
   同名标签的歧义在块内用「Tape A/B/C」小节头消解。

3. **表头里月份列与季度列交替出现，只认 datetime 单元格。**
   实际长这样：`2011-01-01, 2011-02-01, 2011-03-01, '1Q11', 2011-04-01, …` ——
   月份是真 datetime，季度是字符串 `'1Q11'`。按 `isinstance(v, datetime)` 过滤最稳，
   千万别按位置数格子。另外 `CDS Clearing` 的日期是**月末**（2013-01-31）而其余三张是
   **月初**，所以只取 (year, month)；`CDS Clearing` 的数据还**从第 2 列开始**（其余从第 3 列），
   这一条在按表头行现扫列号的写法下自动消失，不需要特判。

4. **OI 是「千张」不是「张」，与 `series/cme.csv` 差 1000 倍。**
   ICE：`oi_energy_kcontracts` 2026-07 = 68,093（千张）；
   CME：`oi_energy_contracts`  2026-07 = 11,042,384（裸张）。
   列名后缀写对是最后一道防线，横截面页最容易在这里翻车。
   同理现货是**百万股**（`_mnsh`），`cboe.csv` 的 `adv_us_equities_matched_shares_bn` 是
   **十亿股**，跨家比要 ÷1000。

5. **没有 TOTAL OI 行 —— "Total OI" 要自己加：`oi_commodities + oi_financials`。**
   这不是猜的：2026-07 期算出 72748+49078 = 121,826，2025-07 期 67322+35541 = 102,863，
   +18.4%，与官方新闻稿 "open interest (OI) up 18% y/y" 吻合；2026-06 期同法算出 +19.7%
   对稿里的 20%。反过来证明了加总方法正确。`_crosscheck_release()` 每月自动重跑这一条。

6. **`TOTAL FUTURES & OPTIONS ≠ COMMODITIES + FINANCIALS`，不要当校验条件。**
   实测 187 个月里 121 个月严格相等，其余偏差 0-0.53%（2011-2013 最大，max|差| 32 千张，
   2013 之后基本收敛到 ±1）。
   ⚠ **别写「因为商品与金融各自用不同交易日数归一」那个解释 —— 已被证伪**：
   偏差最大的 2011-08（差 32）、2011-10（差 27）、2012-08（差 16）里，
   `trading_days_commod` 与 `trading_days_rates` 恰恰**相等**。真实原因未查明
   （可能是 2011-2013 的追溯重刷只重刷了小类没重刷合计）。结论仍是「不当硬校验」，
   但写一个错的因果，下一个人会去"修"一个修不好的东西。

7. **单股（`adv_single_stock_kcontracts`）不进任何合计、不进任何池。**
   脚注 (13) 原文：Single Stock Equity 已从 Total Financials 小计里剔除，
   理由是"收入封顶、与量无相关性"，因此 ADV / RPC / OI 全部剔除。
   它 2011-01 有 905 千张、2026-07 只剩 69 —— 误加进 Total Financials 会凭空造出一段
   "2011-2013 金融衍生品塌方"。入库是为了留档，画图只能单独一张并注明"官方口径外"。

8. **`adv_fx_credit_kcontracts` 的行标签写着 "& CREDIT"，但口径里没有信用。别把标签当定义。**
   同一张表的脚注 (12) 原文只说：「"TOTAL FX" includes futures and options for the
   U.S. Dollar Index and foreign exchange」—— 一个字没提 credit。
   独立验证：另一份官方文件 `2015-2021-Historical-ADV_OI_vF.xlsx` 里对应的是
   `FX & Other Financials` + `USDX` = `Total FX & Other:`，2019-06 = 38,947 张 = 38.9 千张，
   与本表同月的 39 精确相符 —— 信用不在里面。
   列名保留 `fx_credit`（**忠于官方行标签**，改名会让人对不上原表），但页面上不能写成
   "含信用"。跨家比也只能与 `cme.csv` 的 `adv_fx_kcontracts` 比增速：
   ICE 22-69 千张 vs CME 811 千张（2026-07）是量级差，并排画绝对量没有意义。

9. **新闻稿的产品分组名与本表的行名同名不同物，不要拿新闻稿逐行对账。**
   2026-07 稿写 "Other Crude & Refined products ADV up 10% y/y"，本表 `Other Oil` 行是
   838 vs 998（−16%）。原因：稿里用的是**合约级**小类（明细文件里
   `Other Crude & Refined Products` 2015-01 只有 153.4 千张），本表的 `Other Oil` 是
   **产品组**（同月 394 千张 = Total Oil − Brent − Gasoil）。
   `_crosscheck_release()` 因此只对下面这批**可对账**的口径回校：
   Total / Total Energy / Total Oil / Brent / Gasoil / Total Natural Gas /
   Total Ag & Metals / Sugar / Total Financials / Total Interest Rates /
   Total Equity Indices / NYSE Cash Equities / NYSE Equity Options。
   稿里的 TTF / Asia gas / Cocoa / Coffee / Cotton / Euribor / SONIA / Gilts / MSCI /
   North American Gas / Other Crude & Refined / Total Environmentals 在本表**没有对应行**，
   一律不认（Environmentals 有行但基数小到 54-63 千张，四舍五入就能差 1.3pp，也不认）。

10. **RPC 是滚动三月均，而且 ICE 的 RPC 不滞后**（与 Cboe 相反）。
    脚注 (1) 原文：RPC = transaction revenues ÷ contract volume。
    最新月的 RPC 与 ADV 一起给出（2026-07 期 `rpc_energy_usd` = 1.89 已填）。
    ⇒ 任何 ICE vs Cboe 的 RPC 并排图，**ICE 那条线每个月都会比 Cboe 多伸出一格**
    （`series/cboe.csv` 最后一行的 rpc_* 全空是正常的）。绘图层必须截齐或加注，
    否则每月看板一刷新都像是 Cboe 抓挂了。
    另外 RPC 是三月滚动，不能与单月量相乘去算单月收入；要画 "ADV × RPC = 收入代理"
    必须在图上标明。

11. **官方做过两次追溯性重刷，所以本仓 CSV 与 ICE 历年季报/10-K 原文可能对不上。**
    表内自述：(a) NGX 的量与收入追溯并入 **2011 年起**的 Other Oil / Nat Gas / Power /
    Total Energy / Total Commodities / Total F&O 及 Energy、Commodities RPC；
    (b) 2013 年起的 Power ADV、Energy RPC、Energy OI 按新的电力量折算法重算；
    (c) Russell 合约 2016-12 规格减半，量、OI、RPC 全部追溯调整。
    本仓从 2011 一次性全量摄入当前文件，不受影响；对不上时以当前文件为准。
    历史极稳（跨 6 年的两期文件对比只有 8 格不同 / 6,210 格，6 格是四舍五入痕迹），
    所以**已有值永不覆盖、只填空**是安全的；官方与本仓不一致的格子写进
    `cache/ice_restatements.csv` 供人工判断，绝不自动吞。
    唯一一次实质重述是 CDS：2026-06 期里 2026-01 的 Non-Client=291 / Client=1330 /
    Total=1720（291+1330=1621，差 99），2026-07 期把 Non-Client 改成 391。
    `_validate()` 的 CDS foot check（容差 2）当期就能抓住它。

12. **2011-01 ~ 2013-10 的 NYSE 数据是「追溯并入」的形式数，不是 ICE 当时的业务。**
    Derivs sheet 底部原文：「For comparison purposes, we include NYSE ADV, RPC and OI in
    all periods covered in this spreadsheet」。ICE 是 **2013 年 11 月**才完成 NYSE Euronext
    收购的。所以 `share_nyse_us_cash_matched` 从 2011-01 的 26.9% 一路跌到 2026-07 的 19.1%，
    **前 34 个月讲的是被收购前 NYSE Euronext 的份额**。数据本身没错（ICE 就是这么披露的），
    但画这两条份额线时页面必须标注「2013-11 前为追溯并入口径」。

13. **`latest_month()` 以「最后一个 TOTAL FUTURES & OPTIONS 非空的月」为准，不信文件名。**
    文件名里的 "July-2026" 是**数据月**不是发布月。当前表没有未来占位列
    （187 个月的 TOTAL F&O 全部非空，实测），但代码写成能容忍占位列的形式 ——
    HKEX 就是先把下个月的列开出来再填数的。

14. **`adv_us_equity_options_industry_kcontracts` 是否含指数期权，ICE 从未书面说明。**
    工作簿里 `Total Equity Options` 这一行**没有任何脚注**。「不含指数期权」是交叉推出来的
    合理读法：2026-07 用它算 NYSE 份额 = 13614/64394 = 21.1%（与 ICE 自报的 0.211 一致），
    而 10-K 自报 2025 年 NYSE 份额 18.9%、本表反算 18.92% —— 若分母含指数期权，
    这两个数会对不上。**页面上不能写成「ICE 官方定义为不含指数期权」**，
    只能写「经与 Cboe multilist 及 ICE 10-K 交叉验证，该分母与多重上市股票/ETF 期权口径一致」。

15. **交易日有三行，不是两行；而且三行会不一样。**
    Derivs sheet 给 `Commodities & Other Financials` 与 `Interest Rates & Single Stock
    Equities` 两行（187 个月里只有 69 个月相同 —— 欧洲利率市场的假期表与商品不同）。
    `US Equity Options` 与 `Cash Products` 两张表还各自有一行 Trading Days，
    实测这两张彼此永远相同，但与 Derivs 的商品行在 **3 个月**上不同：
    2012-10（23 vs 21，飓风 Sandy 关了美股两天，ICE Futures Europe 照开）、
    2018-12（20 vs 19，老布什国葬日休市）、2025-01（21 vs 20，卡特国葬日休市）。
    所以本模块入库三列：`trading_days_commod` / `trading_days_rates` /
    `trading_days_us_equities`。把美股 ADV 还原成当月总量时用第三列，用前两列会在那 3 个月
    多算一天。
    （`US Equity Options` 那张表也有一行 Trading Days，实测 187/187 个月与 `Cash Products`
    的那行**完全相同**，所以只入库一次，不重复存同一个事实。）

16. **📌 `CDS Clearing` 表底部还有两行，本模块**故意不入库**，不是漏了。**
    检索路径：`cache/ice_monthly_stats.xlsx` → sheet `CDS Clearing` → 第 14-20 行
    （在三行 Gross Notional 之下、被两段脚注隔开，`_blocks()` 把它们归在同一个块里，
    所以枚举块内行时看得见）。两行是：
      · r17 `CDS Clearing Revenue`（$ in Millions）
      · r20 `Gross Notional Value of CDS`（$ in Trillions）
    三条不入库的理由，任何一条单独成立都够：
      (a) **已死**。两行都只有 2016-09 .. 2019-06 共 34 个月有值，此后 7 年一格未填
          （163 个月份列里 129 个是空的）。入库等于在表里挂两列永久空白，
          还得给 `_validate()` 的「缺列一律失败」再开两个豁免口子 —— 护栏开的口子越多越不值钱。
      (b) **口径自相矛盾，无法确定**。块标题 r14 原文写着
          `CDS Transaction and Clearing Revenue (Quarterly)`，明说是**季度**；
          但实测每个月的值都不同（2016-09/10/11 = 52/43/62，不是同一个季度值重复三遍），
          数据形状是**月度**。标题与数据打架，ICE 没有任何脚注解释，
          无从判断到底是「月度收入放在月列」还是「季度收入错位填进了月列」。
          按仓库纪律，口径不确定的数字一律不入库。
      (c) 真要用 CDS 收入，正确的源是 ICE 的季度 `Key-Metrics-Q*.xlsx` 与 10-K 分部收入，
          那里有明确口径与完整历史，不该从一张停更七年的月度表里刨。
    ⇒ 下一个人若「发现」了这两行，请先读这一条再动手。

17. **📌 未找到：TTF 没有独立行 —— 想要「ICE 的 TTF 成交量」，本表给不了。**
    检索路径：整份工作簿四张 sheet 逐格扫过 `TTF` / `Title Transfer` / `Dutch`
    三个模式，**0 命中**；181 条 sharedStrings 里也 0 命中。
    TTF 被折进 `Nat Gas` 一行 —— 脚注 (4) 原文：
    `"Natural Gas" includes North American, NGX, UK and European Natural Gas`，
    欧洲天然气（TTF 就在其中）与北美、NGX、英国四块合并披露，官方不拆。
    官方新闻稿里确实单独点评过 TTF 的同比，但那是**合约级**口径，与本表的产品组行
    同名不同物（见口径坑 9），拿稿里的百分比去反推一个 TTF 绝对量是在编数。
    ⇒ 需要 TTF 单品的话只能另找源（ICE 的合约级明细文件
      `2015-2021-Historical-ADV_OI_vF.xlsx` 有，但 2022-02 已冻结不再更新），
      不要试图从 `adv_natgas_kcontracts` 里拆。

━━ series/ice.csv 里放什么、不放什么 ━━
放**官方原样给出的行**，不放派生列。两条推论：

  · **张数原样保留**（`adv_*_kcontracts`）。本批次主口径是「定基名义额」
    （张数 × 乘数 × 基期价格，基期锁 2019-01），但那是 build 层用一张外部维护的
    乘数/基期价表算出来的东西 —— 源数据必须存张数，否则乘数表一改，历史就再也复原不了。
    张数另有两个不可替代的用途：(a) 与官方新闻稿逐位对账、证明解析没错（本模块每月自动做）；
    (b) 将来接收入模型（ADV × 交易日 × RPC ≈ 交易收入，实测 6/6 个季度在 ±1% 内）。
  · **`adv_nyse_us_cash_matched_mnsh` = tapeA+B+C matched 不入库**，因为官方没有这一行。
    它是与 `cboe.csv` 的 `adv_us_equities_matched_shares_bn` 口径一致的那个数，
    build 层一行就能派生：
        (adv_nyse_tapeA_matched_mnsh + adv_nyse_tapeB_matched_mnsh
         + adv_nyse_tapeC_matched_mnsh) / 1000     # → 十亿股，与 cboe 同量纲
    分母同理用三个 `adv_tape*_consolidated_mnsh` 之和。
    **不要用 handled** —— handled 含路由到别家成交的量，Cboe 不披露对应口径。

━━ 依赖 ━━
openpyxl（读 xlsx）、pymupdf（读新闻稿 PDF 做 y/y 回校；取不到 PDF 只警告不阻断）。
不依赖 pandas —— 避免 to_csv 重排既有行的格式。
"""

import csv
import datetime as _dt
import json
import os
import re
import urllib.parse
import urllib.request

import openpyxl

TICKER = 'ice'

FEED_BASE = 'https://ir.theice.com/feed'
ASSET_FEED = FEED_BASE + '/ContentAsset.svc/GetContentAssetList'
PR_FEED = FEED_BASE + '/PressRelease.svc/GetPressReleaseList'

# 用常规桌面 UA。q4cdn 与 feed 目前都不校验，但 Cloudflare 策略随时可能收紧，
# 带上 UA 是零成本的保险。**不带 apiKey**，理由见模块 docstring。
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

# feed 的 pageSize 是硬截断，不是分页提示 —— 实测 2013 年用 250 只回 250 条、
# 用 600 回 392 条，2022 年 250→274。回补 source_dates 时用小值会静默丢掉整年的稿。
_PAGE = 600

WORKBOOK_TITLE = 'Monthly Statistics Tracking'
# 兜底指针：同一个 URL 在 feed 里挂了两条，另一条标题就是 "Monthly Statistics"
# （Type = Additional Resources）。ICE 哪天改了主标题，这条天然接上。
WORKBOOK_TITLE_FALLBACK = re.compile(r'monthly\s+stat', re.I)

MONTH_NAMES = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
               'August', 'September', 'October', 'November', 'December']


class IceFetchError(RuntimeError):
    """源站结构变化 / 下载失败 / 解析结果不完整 / 内部恒等式不成立。

    一律炸掉。宁可整月不更新（线上留着自己的旧数据），也绝不静默写空列或 NaN。
    """


# ══════════════════════════════════════════════════════════════════════
# 表结构
# ══════════════════════════════════════════════════════════════════════
# 块（block）不按 sheet 名+行号写死，按「日期表头行」切分后再用标题文字认身份，
# 见模块 docstring 口径坑 2。这里的 key 是块的身份标识。
B_DERIV_ADV = ('Derivs', 'MONTHLY ADV')
B_DERIV_RPC = ('Derivs', 'RATE PER CONTRACT')
B_DERIV_OI = ('Derivs', 'DERIVATIVES OI')
B_OPT_ADV = ('Options', 'MONTHLY ADV')
B_OPT_RPC = ('Options', 'RATE PER CONTRACT')
B_CASH_ADV = ('Cash', 'MONTHLY ADV')
B_CASH_RPC = ('Cash', 'PER 100 SHARES')
B_CDS = ('CDS', 'GROSS NOTIONAL CLEARED')

# sheet 也按内容认，不按全名 —— ICE 的 sheet 名带括号说明
# （'Derivs (ADV, RPC, OI)'），加一个字就失配。
SHEET_KEYS = {
    'Derivs': re.compile(r'^deriv', re.I),
    'Options': re.compile(r'equity\s*options', re.I),
    'Cash': re.compile(r'cash\s*products', re.I),
    'CDS': re.compile(r'^cds', re.I),
}

# (csv 列名, 块, 小节关键词或 None, 可接受的行标签集合)
# 行标签为空元组 = 「该块里唯一那行有数的」—— RPC 两块的数据行在原表里**没有行名**
# （US Equity Options r16、Cash Products r35 的 A/B 列都是空的），只能这样定位。
COLUMN_SPEC = [
    # ── 交易日 ───────────────────────────────────────────────────────
    ('trading_days_commod', B_DERIV_ADV, None, ('COMMODITIES & OTHER FINANCIALS',)),
    ('trading_days_rates', B_DERIV_ADV, None, ('INTEREST RATES & SINGLE STOCK EQUITIES',)),
    ('trading_days_us_equities', B_CASH_ADV, None, ('TRADING DAYS',)),

    # ── 衍生品 ADV（千张）───────────────────────────────────────────
    ('adv_brent_kcontracts', B_DERIV_ADV, None, ('BRENT',)),
    ('adv_gasoil_kcontracts', B_DERIV_ADV, None, ('GASOIL',)),
    ('adv_otheroil_kcontracts', B_DERIV_ADV, None, ('OTHER OIL',)),
    ('adv_natgas_kcontracts', B_DERIV_ADV, None, ('NAT GAS', 'NATURAL GAS')),
    ('adv_power_kcontracts', B_DERIV_ADV, None, ('POWER',)),
    # 行名漂移：2020 年那期叫 "Emissions & Other"，2026 年这期叫 "Environmentals & Other"
    # （而两期的脚注正文都写 "Emissions & Other"）。别名集合是为这一行准备的。
    ('adv_environmentals_kcontracts', B_DERIV_ADV, None,
     ('ENVIRONMENTALS & OTHER', 'EMISSIONS & OTHER')),
    ('adv_energy_kcontracts', B_DERIV_ADV, None, ('TOTAL ENERGY',)),
    ('adv_sugar_kcontracts', B_DERIV_ADV, None, ('SUGAR',)),
    ('adv_otherags_metals_kcontracts', B_DERIV_ADV, None, ('OTHER AGS & METALS',)),
    ('adv_ag_metals_kcontracts', B_DERIV_ADV, None, ('TOTAL AGRICULTURE & METALS',)),
    ('adv_commodities_kcontracts', B_DERIV_ADV, None, ('TOTAL COMMODITIES',)),
    ('adv_stir_kcontracts', B_DERIV_ADV, None, ('SHORT-TERM INTEREST RATES',)),
    ('adv_mltir_kcontracts', B_DERIV_ADV, None,
     ('MEDIUM & L-T INTEREST RATES', 'MEDIUM & LT INTEREST RATES')),
    ('adv_equity_index_kcontracts', B_DERIV_ADV, None, ('EQUITY INDICES',)),
    ('adv_fx_credit_kcontracts', B_DERIV_ADV, None, ('FX & CREDIT', 'FX')),
    ('adv_financials_kcontracts', B_DERIV_ADV, None, ('TOTAL FINANCIALS',)),
    ('adv_futures_options_kcontracts', B_DERIV_ADV, None, ('TOTAL FUTURES & OPTIONS',)),
    ('adv_single_stock_kcontracts', B_DERIV_ADV, None, ('SINGLE STOCK EQUITIES',)),

    # ── 衍生品 RPC（滚动三月均，美元/张）────────────────────────────
    ('rpc_energy_usd', B_DERIV_RPC, None, ('ENERGY',)),
    ('rpc_ag_metals_usd', B_DERIV_RPC, None, ('AGRICULTURE & METALS',)),
    ('rpc_commodities_usd', B_DERIV_RPC, None, ('TOTAL COMMODITIES',)),
    ('rpc_rates_usd', B_DERIV_RPC, None, ('INTEREST RATES',)),
    ('rpc_other_financials_usd', B_DERIV_RPC, None, ('OTHER FINANCIALS',)),
    ('rpc_financials_usd', B_DERIV_RPC, None, ('TOTAL FINANCIALS',)),

    # ── 衍生品 OI（月末净未平仓，千张）──────────────────────────────
    ('oi_energy_kcontracts', B_DERIV_OI, None, ('ENERGY',)),
    ('oi_ag_metals_kcontracts', B_DERIV_OI, None, ('AGRICULTURE & METALS',)),
    ('oi_commodities_kcontracts', B_DERIV_OI, None, ('TOTAL COMMODITIES',)),
    ('oi_rates_kcontracts', B_DERIV_OI, None, ('INTEREST RATES',)),
    ('oi_other_financials_kcontracts', B_DERIV_OI, None, ('OTHER FINANCIALS',)),
    ('oi_financials_kcontracts', B_DERIV_OI, None, ('TOTAL FINANCIALS',)),

    # ── 美股期权（千张）──────────────────────────────────────────────
    ('adv_nyse_equity_options_kcontracts', B_OPT_ADV, None, ('US EQUITY OPTIONS',)),
    # ⭐ 行业分母之一：全美股票/ETF 期权总量
    ('adv_us_equity_options_industry_kcontracts', B_OPT_ADV, None, ('TOTAL EQUITY OPTIONS',)),
    ('share_nyse_equity_options', B_OPT_ADV, None, ('NYSE SHARE OF GROUP TOTAL',)),
    ('rpc_nyse_equity_options_usd', B_OPT_RPC, None, ()),

    # ── 美股现货（百万股）────────────────────────────────────────────
    # 三个 tape 的行标签完全同名，只能靠小节头（"…(Tape A) Issues"）区分。
    ('adv_nyse_tapeA_handled_mnsh', B_CASH_ADV, 'TAPE A', ('HANDLED VOLUME',)),
    ('adv_nyse_tapeA_matched_mnsh', B_CASH_ADV, 'TAPE A', ('MATCHED VOLUME',)),
    # ⭐ 行业分母之二：Tape A/B/C 各自的全市场合并成交量
    ('adv_tapeA_consolidated_mnsh', B_CASH_ADV, 'TAPE A',
     ('TOTAL NYSE LISTED CONSOLIDATED VOLUME',)),
    ('share_nyse_tapeA_matched', B_CASH_ADV, 'TAPE A',
     ('SHARE OF TOTAL MATCHED CONSOLIDATED VOLUME',)),
    ('adv_nyse_tapeB_handled_mnsh', B_CASH_ADV, 'TAPE B', ('HANDLED VOLUME',)),
    ('adv_nyse_tapeB_matched_mnsh', B_CASH_ADV, 'TAPE B', ('MATCHED VOLUME',)),
    ('adv_tapeB_consolidated_mnsh', B_CASH_ADV, 'TAPE B',
     ('TOTAL NYSE ARCA AND AMERICAN LISTED CONSOLIDATED VOLUME',)),
    ('share_nyse_tapeB_matched', B_CASH_ADV, 'TAPE B',
     ('SHARE OF TOTAL MATCHED CONSOLIDATED VOLUME',)),
    ('adv_nyse_tapeC_handled_mnsh', B_CASH_ADV, 'TAPE C', ('HANDLED VOLUME',)),
    ('adv_nyse_tapeC_matched_mnsh', B_CASH_ADV, 'TAPE C', ('MATCHED VOLUME',)),
    ('adv_tapeC_consolidated_mnsh', B_CASH_ADV, 'TAPE C',
     ('TOTAL NASDAQ LISTED CONSOLIDATED VOLUME',)),
    ('share_nyse_tapeC_matched', B_CASH_ADV, 'TAPE C',
     ('SHARE OF TOTAL MATCHED CONSOLIDATED VOLUME',)),
    # 自校验锚点：ICE 自报的全美 matched 份额。入库不是为了画图，是为了每月拿
    # (A+B+C matched) / (A+B+C consolidated) 去撞它 —— 撞得上就说明 12 个 tape 单元格
    # 一格没错行。见 _validate()。
    ('share_nyse_us_cash_matched', B_CASH_ADV, None,
     ('TOTAL U.S. CASH MARKET SHARE MATCHED',)),
    ('adv_nyse_us_cash_handled_mnsh', B_CASH_ADV, None,
     ('TOTAL U.S. CASH PRODUCTS HANDLED',)),
    ('rpc_nyse_us_cash_usd_per100sh', B_CASH_RPC, None, ()),

    # ── CDS 清算（十亿美元，当月总额不是 ADV，2013-01 起）────────────
    ('cds_nonclient_notional_usdbn', B_CDS, None, ('NON-CLIENT',)),
    ('cds_client_notional_usdbn', B_CDS, None, ('CLIENT',)),
    ('cds_total_notional_usdbn', B_CDS, None, ('TOTAL CDS GROSS NOTIONAL CLEARED',)),
]

COLUMNS = [c for c, _b, _g, _l in COLUMN_SPEC]

# CDS 段 2013-01 才开始，比其余字段晚两年。这不是解析失败，是官方就没有。
CDS_COLS = [c for c in COLUMNS if c.startswith('cds_')]
CDS_FIRST_MONTH = '2013-01'

# 「解析成功」的锚：这一列有值才算这个月真的有数据（见 latest_month / 占位列判定）。
ANCHOR_COL = 'adv_futures_options_kcontracts'


# ══════════════════════════════════════════════════════════════════════
# 网络
# ══════════════════════════════════════════════════════════════════════
def _http_get(url, timeout=90):
    req = urllib.request.Request(url, headers={
        'User-Agent': _UA,
        'Accept': '*/*',
        'Accept-Language': 'en-US,en;q=0.9',
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.read(), dict(r.headers)
    except Exception as e:                                # noqa: BLE001
        raise IceFetchError('下载失败 %s: %r' % (url, e)) from e


def _get_json(url, result_key):
    """Q4 的 .svc 接口返回 UTF-8 BOM + {"<Method>Result": [...]}。"""
    raw, _hdr = _http_get(url)
    try:
        obj = json.loads(raw.decode('utf-8-sig'))
    except Exception as e:                                # noqa: BLE001
        raise IceFetchError('%s 返回的不是 JSON（前 200 字节 %r）'
                            % (url, raw[:200])) from e
    if result_key not in obj:
        raise IceFetchError('%s 的响应里没有 %s，接口可能改版：%s'
                            % (url, result_key, sorted(obj)[:8]))
    return obj[result_key]


def _asset_feed():
    q = urllib.parse.urlencode({
        'LanguageId': 1,
        # 这个参数 ICE 端**并没有真的过滤**（实测同样返回 Governance Documents 等其他
        # 类型），带上只是照官网 JS 的原样。真正的筛选在下面按 Title 做。
        'contentAssetTypeList': 'Supplemental Information',
        'year': -1, 'pageSize': _PAGE, 'includeTags': 'true',
        'excludeSelection': 1, 'tagList': '',
    })
    return _get_json(ASSET_FEED + '?' + q, 'GetContentAssetListResult')


def _discover_workbook(cache_dir):
    """返回 (xlsx 直链, feed 条目)。找不到就炸，不猜 URL（见 docstring）。"""
    items = _asset_feed()
    os.makedirs(cache_dir, exist_ok=True)
    # 存一份 feed 原文：源站改版时这是唯一能事后取证的东西
    with open(os.path.join(cache_dir, 'ice_asset_feed.json'), 'w',
              encoding='utf-8') as f:
        json.dump(items, f, ensure_ascii=False, indent=1)

    def _xlsx(it):
        return (it.get('FileType') or '').upper() == 'XLSX' and it.get('FilePath')

    hit = [it for it in items
           if _xlsx(it) and (it.get('Title') or '').strip() == WORKBOOK_TITLE]
    if not hit:
        # 兜底指针：同一 URL 的第二条（Title = "Monthly Statistics"）
        hit = [it for it in items
               if _xlsx(it) and WORKBOOK_TITLE_FALLBACK.search(it.get('Title') or '')
               and 'HISTORICAL' not in it['FilePath'].upper()]
    if not hit:
        raise IceFetchError(
            'ContentAsset feed 里找不到 %r（也没有 "Monthly Stat…" 的兜底条目）。'
            'feed 共 %d 条，标题样例：%s'
            % (WORKBOOK_TITLE, len(items),
               [it.get('Title') for it in items if _xlsx(it)][:6]))
    if len({it['FilePath'] for it in hit}) > 1:
        raise IceFetchError('feed 里有多条互不相同的月度统计 xlsx，无法判断用哪一份：%s'
                            % sorted({it['FilePath'] for it in hit}))
    return hit[0]['FilePath'], hit[0]


_WB_URL_MONTH = re.compile(r'Monthly-Stats-([A-Za-z]+)-(\d{4})', re.I)


def _url_month(url):
    """从工作簿 URL 里取它自报的数据月 'YYYY-MM'；认不出返回 None。

    实测形如 `2011-2026-Monthly-Stats-July-2026_vF.xlsx`。后缀历年变过
    （2020 是 `_v1`、2026 是 `_vF`），但 `Monthly-Stats-<Month>-<Year>` 这一段
    2015-2026 一直稳定。模块 docstring 已实测确认这里的 July-2026 是**数据月**
    而不是发布月，所以它可以和解析结果直接比。
    """
    m = _WB_URL_MONTH.search(url or '')
    if not m:
        return None
    try:
        return '%s-%02d' % (m.group(2), MONTH_NAMES.index(m.group(1).title()) + 1)
    except ValueError:
        return None


def _crosscheck_workbook_month(url, newest):
    """工作簿自报的数据月 vs 解析出来的最新月，不等就炸。

    这是本模块唯一一条**独立于解析器**的月份判据。防的是这一类：ICE 改了行标签
    或缩进（本表大量用 \xa0 缩进、标签带 (n) 脚注），_lab / 行匹配认不出于是整行
    静默丢掉，_validate 拿剩下的算出 newest = 上个月，fetch 干净报 NOCHANGE ——
    没有 FAIL、没有红点、streaks 也不动。MSCI 2026-07 就是这么漏的
    （见 resolved_issues.msci_silent_parse_miss）。

    与新闻稿判据的区别、以及为什么这条能 raise 而那条只能 warn：
    URL 和文件内容是**同一个 artifact 同批发布**的，不存在时间差；新闻稿则可能
    先于 CDN 上的文件更新，拿它 raise 会在发布日误杀。

    注意与 latest_month() 的「不信文件名」不冲突：真值仍然只从内容取，
    文件名只做对账。文件名认不出时只 warn 不 raise —— 官方改个命名不该让
    一个本来好好的源直接停摆，但必须让人看见护栏掉了。
    """
    declared = _url_month(url)
    if declared is None:
        print('[ice] ⚠ 护栏失效：工作簿 URL 认不出数据月（命名可能变了），'
              '这一轮没有独立于解析器的月份判据。URL=%s' % url)
        return
    if declared != newest:
        raise IceFetchError(
            '工作簿自报数据月 %s，但解析出来的最新月是 %s（URL=%s）。二者同批发布、'
            '不该不一致 —— 最可能是行标签/缩进变了导致整行被静默丢弃。拒绝写入，'
            '请人工看一眼工作簿最后一列。' % (declared, newest, url))


def _crosscheck_release_month(releases, newest):
    """新闻稿 feed 里最新的数据月 vs 解析出来的最新月 —— **只告警，不阻断**。

    ICE 的月度统计新闻稿和工作簿理论上同批出，但两者挂在不同的 feed / CDN 上，
    发布日当天存在「稿先出、文件后更新」的窗口。所以这条只能提示，不能 raise，
    否则每个发布日都可能误杀。真正能 raise 的判据是 _crosscheck_workbook_month。
    """
    if not releases:
        return
    latest_pr = max(releases)
    if latest_pr > newest:
        print('[ice] ⚠ 新闻稿已出到 %s，但本轮工作簿只解析到 %s。若非发布日当天的'
              '时间差，就是工作簿还没更新或解析漏了月份 —— 明天这条 cron 会再试；'
              '连续两天还这样就要人工看。' % (latest_pr, newest))


def _download_workbook(cache_dir):
    """下载当前最新一期 xlsx，返回 (本地路径, url, HTTP Last-Modified 或 None)。"""
    url, _item = _discover_workbook(cache_dir)
    body, hdr = _http_get(url)
    if body[:2] != b'PK':
        raise IceFetchError('%s 返回的不是 xlsx（前 8 字节 %r）' % (url, body[:8]))
    os.makedirs(cache_dir, exist_ok=True)
    path = os.path.join(cache_dir, 'ice_monthly_stats.xlsx')
    with open(path, 'wb') as f:
        f.write(body)
    return path, url, hdr.get('Last-Modified')


# ══════════════════════════════════════════════════════════════════════
# 解析
# ══════════════════════════════════════════════════════════════════════
_FOOTNOTE = re.compile(r'\s*\(\d{1,2}\)\s*$')


def _txt(v):
    """单元格 → 规整文本。ICE 的标签里大量用 \\xa0（不换行空格）缩进。"""
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v).replace('\xa0', ' ')).strip()


def _lab(cells):
    """行标签归一化：取**最右边那个非空的标签单元格**，剥掉脚注记号，转大写。

    原表的排版是「A 列写大类或合计名、B 列写细项名」：
      A='   Energy'  B='Brent (1)'                 → 这一行是 Brent
      A='TOTAL ENERGY' B=空                        → 这一行是 Total Energy
      A='TOTAL FX & CREDIT' B='FX & Credit (12) '  → 这一行是 FX & Credit
    所以 B 有值就用 B。

    ⚠ 「标签单元格是哪几列」不能写死成 A/B —— `CDS Clearing` 的月份**从第 2 列开始**
    （其余三张从第 3 列），把 B 当标签会读到 630 这种数字当行名。所以调用方按
    「第 1 列 .. 最小月份列 − 1」现算，见 _blocks。

    脚注记号 "(1)" / "(12)" 与 CDS 的 "*" / "**" 历史上加过也去过，一律右剥。
    """
    lab = ''
    for v in cells:
        t = _txt(v)
        if t:
            lab = t
    return _FOOTNOTE.sub('', lab).rstrip('* ').strip().upper()


def _month_columns(ws, header_row):
    """{'YYYY-MM': col}。只认 datetime 单元格 —— 季度列是字符串 '1Q11'，会被丢掉。

    CDS 那张表的日期是月末（2013-01-31）、其余是月初，所以只取 (year, month)。
    """
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, _dt.datetime):
            out['%04d-%02d' % (v.year, v.month)] = c
    return out


def _header_rows(ws):
    """返回这张 sheet 里所有「日期表头行」的行号（升序）。

    判据是「这一行有 ≥3 个 datetime 单元格」。四张 sheet 的表头行位置各不相同
    （Derivs 5/58/69、US Equity Options 5/15、Cash Products 4/34、CDS 5），
    现扫出来就不必写死，官方插一行也不会炸。
    """
    rows = []
    for r in range(1, ws.max_row + 1):
        n = sum(1 for c in range(1, ws.max_column + 1)
                if isinstance(ws.cell(r, c).value, _dt.datetime))
        if n >= 3:
            rows.append(r)
    return rows


def _blocks(ws):
    """把一张 sheet 切成若干数据块，返回 [{'key':…, 'cols':…, 'rows':[(group,label,row)]}]。

    块 = 一个日期表头行 + 它下面直到下一个表头行之前的所有行。块的身份用它上方 8 行内
    （含表头行自己）的标题文字认，见模块 docstring 口径坑 2。

    块内每一行分两类：
      · **数据行** —— 在本块月份列里至少有一个数字。可被 COLUMN_SPEC 定位。
      · **小节头** —— 有标签但一个数字都没有（"NYSE Listed (Tape A) Issues"、
        "COMMODITIES"、以及底部那一堆脚注）。它只用来给后面的数据行打 group 标记，
        Cash Products 里三组同名的 Handled/Matched/Share 全靠它区分。
    """
    hrows = _header_rows(ws)
    out = []
    for i, h in enumerate(hrows):
        end = hrows[i + 1] - 1 if i + 1 < len(hrows) else ws.max_row
        cols = _month_columns(ws, h)
        if not cols:
            continue
        # 标签列 = 第 1 列 .. 最小月份列 − 1。CDS 那张表只有第 1 列，其余三张有 1-2 列。
        lab_cols = list(range(1, min(cols.values())))
        title = ' | '.join(
            _txt(ws.cell(r, c).value)
            for r in range(max(1, h - 8), h + 1) for c in lab_cols
            if _txt(ws.cell(r, c).value)).upper()
        rows, group = [], ''
        for r in range(h + 1, end + 1):
            lab = _lab([ws.cell(r, c).value for c in lab_cols])
            has_num = any(isinstance(ws.cell(r, c).value, (int, float))
                          for c in cols.values())
            # 有数就是数据行 —— 哪怕它没有行名。两个 RPC 块的数据行正是这样
            # （US Equity Options r16、Cash Products r35 的 A/B 列全空），
            # 先前按「无标签就跳过」写会把它们整行漏掉，而漏掉的表现是「少了一列」，
            # 不是报错。
            if has_num:
                rows.append((group, lab, r))
            elif lab:
                group = _txt(ws.cell(r, 1).value).upper() or group
        out.append({'title': title, 'cols': cols, 'rows': rows})
    return out


def _find_block(blocks, key_text, where):
    hit = [b for b in blocks if key_text in b['title']]
    if len(hit) != 1:
        raise IceFetchError(
            '%s 里按标题关键词 %r 找到 %d 个数据块（应当恰好 1 个），官方表结构可能已变。'
            '现有块标题：%s'
            % (where, key_text, len(hit), [b['title'][:70] for b in blocks]))
    return hit[0]


def _find_row(block, group, labels, name, where):
    """在块内定位唯一一行。找不到或找到多行都炸 —— 这正是「缺列一律失败」那条护栏。

    labels 为空 = 「该块里唯一那行有数的」。US Equity Options 与 Cash Products 的两个
    RPC 块，数据行在原表里 A/B 两列都是空的（没有行名可认），只能这么定位；
    因此「块内不止一行有数」在那两个块里就等于官方改版，同样要炸。
    """
    cand = [r for g, lab, r in block['rows']
            if (not group or group in g) and (not labels or lab in labels)]
    if len(cand) != 1:
        raise IceFetchError(
            '%s：列 %s 在块 %r 内按 group=%r labels=%r 匹配到 %d 行（应当恰好 1 行）。'
            '块内现有行：%s'
            % (where, name, block['title'][:50], group, labels, len(cand),
               [(g, l) for g, l, _r in block['rows']][:24]))
    return cand[0]


def parse_workbook(path):
    """解析一份官方 xlsx，返回 {'YYYY-MM': {csv列名: float|None}}（按月升序）。

    任何一个约定的行定位不到 —— 说明官方改了表结构 —— 直接抛 IceFetchError。
    宁可整月不更新，也不要写出一列悄悄全空的 CSV。
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        sheets, ws_of = {}, {}
        for key, pat in SHEET_KEYS.items():
            hit = [s for s in wb.sheetnames if pat.search(s)]
            if len(hit) != 1:
                raise IceFetchError(
                    '%s 里按 %s 匹配 sheet 得到 %r（应当恰好 1 个）；实际 sheet：%r'
                    % (os.path.basename(path), key, hit, wb.sheetnames))
            ws_of[key] = wb[hit[0]]
            sheets[key] = _blocks(ws_of[key])

        where = os.path.basename(path)
        loc = {}          # csv 列名 → (worksheet, 行号, {'YYYY-MM': 列号})
        for name, (sheet_key, block_key), group, labels in COLUMN_SPEC:
            blk = _find_block(sheets[sheet_key], block_key,
                              '%s / sheet %s' % (where, sheet_key))
            row = _find_row(blk, group, labels, name,
                            '%s / sheet %s' % (where, sheet_key))
            loc[name] = (ws_of[sheet_key], row, blk['cols'])

        # 月份集合取并集：CDS 段 2013-01 才开始，取交集会把 2011-2012 整整两年扔掉
        months = set()
        for name in COLUMNS:
            months |= set(loc[name][2])
        data = {}
        for mon in sorted(months):
            rec = {}
            for name in COLUMNS:
                ws, row, cols = loc[name]
                rec[name] = (_cell_num(ws, row, cols[mon], name)
                             if mon in cols else None)
            data[mon] = rec
    finally:
        wb.close()

    if not data:
        raise IceFetchError('%s 解析后没有任何月份' % os.path.basename(path))
    # 未来占位列：整月只有表头没有数（HKEX 那种），丢掉
    data = {m: r for m, r in data.items() if r[ANCHOR_COL] is not None}
    if not data:
        raise IceFetchError('%s 解析后没有任何一个月有 %s'
                            % (os.path.basename(path), ANCHOR_COL))
    return dict(sorted(data.items()))


def _cell_num(ws, row, col, name):
    v = ws.cell(row, col).value
    if v is None or (isinstance(v, str) and v.strip() in ('', '-', '–', 'n/a', 'N/A')):
        return None
    if isinstance(v, bool) or not isinstance(v, (int, float)):
        raise IceFetchError('列 %s 的单元格 R%dC%d 不是数字：%r' % (name, row, col, v))
    return float(v)


# ══════════════════════════════════════════════════════════════════════
# 校验
# ══════════════════════════════════════════════════════════════════════
# 加数四舍五入到整千张 → 和的误差可达 ±2，见口径坑 1。给 3 留一格余量。
FOOT_TOL = 3.0
# CDS 三行是整十亿美元，两项和，实测 max|差| = 1。容差 2 既不误报、又能抓住
# 2026-06 期那个 291/1330/1720 的真错值（差 99）。
CDS_TOL = 2.0
# 份额列 = 官方自己给的商。实测 187 个月里自算与自报最大差 0.092pp。
# 给 0.15pp（0.0015）—— 真错行会差几十 pp，这个阈值足够灵敏。
SHARE_TOL = 0.0015


def _validate(data):
    """完整性 + 内部恒等式体检，返回最新月。任何一条不过就炸。

    这些恒等式全部用**容差**而不是相等，理由见模块 docstring 口径坑 1 ——
    月度单元格四舍五入到整千张，严格相等只在 85/187 个月成立，写成 `==`
    会每月产生 100 多条假告警，而天天报的警等于没有警。
    """
    months = sorted(data)
    if not months:
        raise IceFetchError('解析结果为空')

    # ── 1. 缺列一律失败 ──────────────────────────────────────────────
    for mon in months:
        rec = data[mon]
        bad = [c for c in COLUMNS if rec[c] is None]
        if mon < CDS_FIRST_MONTH:
            # CDS 段 2013-01 才开始，早于此的空是官方就没有，不是解析失败
            bad = [c for c in bad if c not in CDS_COLS]
        if bad:
            raise IceFetchError('%s 缺列 %s —— 解析异常或官方改版，拒绝写入' % (mon, bad))

    def near(mon, lhs, rhs, tol, what):
        a, b = data[mon][lhs], sum(data[mon][c] for c in rhs)
        if abs(a - b) > tol:
            raise IceFetchError(
                '%s 恒等式不成立：%s = %.4f，但 %s 之和 = %.4f（差 %.4f > 容差 %s）—— %s'
                % (mon, lhs, a, '+'.join(rhs), b, a - b, tol, what))

    for mon in months:
        # ── 2. ADV 小计闭合 ──────────────────────────────────────────
        near(mon, 'adv_energy_kcontracts',
             ['adv_brent_kcontracts', 'adv_gasoil_kcontracts', 'adv_otheroil_kcontracts',
              'adv_natgas_kcontracts', 'adv_power_kcontracts',
              'adv_environmentals_kcontracts'], FOOT_TOL, '能源六子项没有加成 TOTAL ENERGY')
        near(mon, 'adv_ag_metals_kcontracts',
             ['adv_sugar_kcontracts', 'adv_otherags_metals_kcontracts'], FOOT_TOL,
             '糖 + 其他农金没有加成 TOTAL AG & METALS')
        near(mon, 'adv_commodities_kcontracts',
             ['adv_energy_kcontracts', 'adv_ag_metals_kcontracts'], FOOT_TOL,
             '能源 + 农金没有加成 TOTAL COMMODITIES')
        near(mon, 'adv_financials_kcontracts',
             ['adv_stir_kcontracts', 'adv_mltir_kcontracts',
              'adv_equity_index_kcontracts', 'adv_fx_credit_kcontracts'], FOOT_TOL,
             '金融四子项没有加成 TOTAL FINANCIALS（注意单股不在里面，脚注 13）')
        # ⚠ 这里**故意不校验** TOTAL F&O = COMMODITIES + FINANCIALS。
        #    实测 187 个月里 66 个月不等，max|差| 32 千张（2011-08），原因未查明，
        #    见口径坑 6。当硬校验会天天误杀。

        # ── 3. OI 小计闭合 ───────────────────────────────────────────
        near(mon, 'oi_commodities_kcontracts',
             ['oi_energy_kcontracts', 'oi_ag_metals_kcontracts'], FOOT_TOL,
             'OI 能源 + 农金没有加成 OI TOTAL COMMODITIES')
        near(mon, 'oi_financials_kcontracts',
             ['oi_rates_kcontracts', 'oi_other_financials_kcontracts'], FOOT_TOL,
             'OI 利率 + 其他金融没有加成 OI TOTAL FINANCIALS')

        # ── 4. CDS foot check（唯一抓到过官方真错值的一条）───────────
        if mon >= CDS_FIRST_MONTH:
            near(mon, 'cds_total_notional_usdbn',
                 ['cds_nonclient_notional_usdbn', 'cds_client_notional_usdbn'], CDS_TOL,
                 'CDS 非客户 + 客户没有加成 TOTAL（2026-06 期就在这里露过马脚）')

        rec = data[mon]
        # ── 5. 份额自洽：本模块解析出来的分子分母，要能撞上官方自报的份额 ──
        #     这是证明「12 个 tape 单元格一格没错行」的唯一硬证据。
        for tag in ('A', 'B', 'C'):
            num = rec['adv_nyse_tape%s_matched_mnsh' % tag]
            den = rec['adv_tape%s_consolidated_mnsh' % tag]
            rep = rec['share_nyse_tape%s_matched' % tag]
            if den and abs(num / den - rep) > SHARE_TOL:
                raise IceFetchError(
                    '%s Tape %s 份额对不上：matched %.0f ÷ consolidated %.0f = %.5f，'
                    '官方自报 %.5f（差 %.3fpp）—— 多半是错行了'
                    % (mon, tag, num, den, num / den, rep, (num / den - rep) * 100))
        num = sum(rec['adv_nyse_tape%s_matched_mnsh' % t] for t in 'ABC')
        den = sum(rec['adv_tape%s_consolidated_mnsh' % t] for t in 'ABC')
        rep = rec['share_nyse_us_cash_matched']
        if den and abs(num / den - rep) > SHARE_TOL:
            raise IceFetchError(
                '%s 全美 matched 份额对不上：(A+B+C matched) %.0f ÷ (A+B+C consolidated) '
                '%.0f = %.5f，官方自报 %.5f（差 %.3fpp）'
                % (mon, num, den, num / den, rep, (num / den - rep) * 100))
        # ── 6. 期权份额自洽（行业分母是不是真的行业分母，就靠这一条）──
        num = rec['adv_nyse_equity_options_kcontracts']
        den = rec['adv_us_equity_options_industry_kcontracts']
        rep = rec['share_nyse_equity_options']
        if den and abs(num / den - rep) > SHARE_TOL:
            raise IceFetchError(
                '%s NYSE 期权份额对不上：%.0f ÷ 行业 %.0f = %.5f，官方自报 %.5f'
                % (mon, num, den, num / den, rep))
        # ── 7. handled ≥ matched（handled 含路由出去成交的量，恒不小于 matched）──
        for tag in ('A', 'B', 'C'):
            h, m = (rec['adv_nyse_tape%s_handled_mnsh' % tag],
                    rec['adv_nyse_tape%s_matched_mnsh' % tag])
            if h + 0.5 < m:
                raise IceFetchError(
                    '%s Tape %s 的 handled %.0f < matched %.0f —— 两行读反了'
                    % (mon, tag, h, m))
    return months[-1]


# ══════════════════════════════════════════════════════════════════════
# 官方新闻稿：发布日 + y/y 回校
# ══════════════════════════════════════════════════════════════════════
# 标题三种写法都真实出现过：
#   "Reports July 2026 Statistics"（现行）
#   "Reports June and Second Quarter 2026 Statistics"（季末月）
#   "Reports ICE & NYSE September Statistics"（2015-2016 期）
_PR_TITLE = re.compile(
    r'Reports\s+(?:ICE\s*&\s*NYSE\s+)?(%s)\b[^;:]*?Statistics' % '|'.join(MONTH_NAMES),
    re.I)


def _stat_releases(year=-1):
    """{'YYYY-MM'(数据月): {'date':'YYYY-MM-DD', 'headline':…, 'pdf':…}}。

    year=-1 只回最近 250 条（其中月度统计稿约 16 期），够生产用；
    回补历史要逐年调（year=2015…2026），见 backfill_source_dates。
    """
    q = urllib.parse.urlencode({
        'LanguageId': 1, 'bodyType': 0,
        # 1 = 按年过滤；3 = 不限年（配 year=-1 用）
        'pressReleaseDateFilter': 1 if year != -1 else 3,
        'year': year, 'pageSize': _PAGE,
        'includeTags': 'false', 'tagList': '', 'excludeSelection': 1,
    })
    out = {}
    for it in _get_json(PR_FEED + '?' + q, 'GetPressReleaseListResult'):
        head = (it.get('Headline') or '').strip()
        m = _PR_TITLE.search(head)
        if not m:
            continue
        raw = (it.get('PressReleaseDate') or '')[:10]      # MM/DD/YYYY
        if not re.match(r'^\d{2}/\d{2}/\d{4}$', raw):
            continue
        pm, pd, py = (int(x) for x in raw.split('/'))
        dmon = MONTH_NAMES.index(m.group(1).title()) + 1
        # 12 月的数据稿在次年 1 月发；其余都在同年次月发
        dyear = py - 1 if dmon == 12 and pm == 1 else py
        out['%04d-%02d' % (dyear, dmon)] = {
            'date': '%04d-%02d-%02d' % (py, pm, pd),
            'headline': head, 'monthname': m.group(1).title(),
            'pdf': it.get('DocumentPath') or '',
        }
    return out


# 新闻稿里能与本表逐条对上的口径。值 = (ADV 对应的算式, OI 对应的算式)；None = 稿里
# 提到了但本表没有对应行，不认。算式用列名元组表示"这几列相加"。见口径坑 9。
_PR_METRICS = {
    'total': (('adv_futures_options_kcontracts',),
              ('oi_commodities_kcontracts', 'oi_financials_kcontracts')),
    'total energy': (('adv_energy_kcontracts',), ('oi_energy_kcontracts',)),
    'total oil': (('adv_brent_kcontracts', 'adv_gasoil_kcontracts',
                   'adv_otheroil_kcontracts'), None),
    'brent': (('adv_brent_kcontracts',), None),
    'gasoil': (('adv_gasoil_kcontracts',), None),
    'total natural gas': (('adv_natgas_kcontracts',), None),
    'total agriculture & metals': (('adv_ag_metals_kcontracts',),
                                   ('oi_ag_metals_kcontracts',)),
    'sugar': (('adv_sugar_kcontracts',), None),
    'total financials': (('adv_financials_kcontracts',), ('oi_financials_kcontracts',)),
    'total interest rates': (('adv_stir_kcontracts', 'adv_mltir_kcontracts'),
                             ('oi_rates_kcontracts',)),
    'total equity indices': (('adv_equity_index_kcontracts',), None),
    'nyse cash equities': (('adv_nyse_us_cash_handled_mnsh',), None),
    'nyse equity options': (('adv_nyse_equity_options_kcontracts',), None),
}

_PR_CLAIM = re.compile(
    r'(?P<lab>[A-Za-z0-9&\. ]+?)\s+'
    r'(?P<kind>ADV|average daily volume \(ADV\)|open interest \(OI\)|OI)\s+'
    r'(?P<dir>up|down)\s+(?P<pct>\d+(?:\.\d+)?)%(?P<tail>[^;\n]{0,40})', re.I)

# 稿方把 y/y 四舍五入到整数百分点，本表的月度值又四舍五入到整千张，两头一叠
# 天然有 ~1pp 的余量。实测 2023-01..2026-07 共 44 期、312 条断言：
# 容差 1.5pp 时唯一越界的是 2024-08 的 "NYSE Cash Equities ADV up 13%"
# （本表 handled 算出 11.4%、matched 也是 11.4%，两种口径都对不上，判定为稿方口径或笔误），
# 放到 2.0pp 后 312/312 全过。真正的错行会差成百上千 pp，2.0 完全够灵敏。
PR_TOL_PP = 2.0
# 少于这么多条可对账断言就不算一次有效体检（12 月那期是 "December, Fourth Quarter and
# Full Year" 混排，只能取到 0-2 条），打印说明后跳过，不当失败。
PR_MIN_CLAIMS = 3


def _release_section(text, monthname):
    """从新闻稿正文里切出「本月」那一段，返回 (段落文本, 是否需要月份限定词)。

    结构实测（2023-2026 共 44 期）：
      · 常规月/季末月 —— "July highlights include:" … 下一个 "… highlights include:"
        或 "About Intercontinental Exchange" 为止。季末月在后面另起
        "Second quarter highlights include:"，**必须切掉** —— 不切的话
        2026-06 那期会拿 2Q 的 22%/24%/8% 去撞月度值，凭空炸出 6 条假不一致。
      · 12 月那期 —— 标题是 "December, Fourth Quarter and Full Year"，正文只有一个
        笼统的 "Highlights include:"，而且**同一句里混着三个口径**
        （"ADV up 19% y/y in December, up 7% y/y in 4Q25 and record 2025 ADV up 14%"），
        更要命的是有的行直接省掉月度值（"Gasoil OI up 20% y/y; ADV up 7% y/y in 4Q25"）。
        这种情况返回 need_qualifier=True，只认写明了 "in December" 的断言。
    """
    m = re.search(r'^%s highlights include:?\s*$' % monthname, text, re.I | re.M)
    if m:
        rest = text[m.end():]
        e = re.search(r'^(?:.*highlights include:?|About Intercontinental Exchange)\s*$',
                      rest, re.I | re.M)
        return (rest[:e.start()] if e else rest), False
    m = re.search(r'^Highlights include:?\s*$', text, re.I | re.M)
    if m:
        rest = text[m.end():]
        e = re.search(r'^About Intercontinental Exchange\s*$', rest, re.I | re.M)
        return (rest[:e.start()] if e else rest), True
    return None, None


def _crosscheck_release(data, month, release, cache_dir):
    """拿官方新闻稿的 y/y 百分比回校解析结果。返回 (对上的条数, 说明文字)。

    为什么值得多下一个 PDF：内部恒等式只能证明「表内自洽」，证明不了「我读的是不是
    这一行」。新闻稿是**外部**证据 —— 它说 Brent ADV up 40% y/y，我用解析出来的
    1948/1394 算出 39.7%，这才排除了错行、错段、错单位。实测 44 期 312 条断言全过。

    为什么它不是硬闸门（拿不到 PDF、或稿里一条可对账断言都没有时只警告不阻断）：
    这道检查依赖 ICE 的**行文措辞**，不依赖数据。哪天文案改版，数据完全正确也会取不到
    断言 —— 那时应该照常入库并把话说清楚，而不是让整家停更。反过来，**取到了断言却对
    不上**是强信号，直接炸。
    """
    if not release or not release.get('pdf'):
        return 0, '没有找到 %s 的官方新闻稿 PDF，本月跳过 y/y 回校' % month
    try:
        import fitz                                    # pymupdf，延迟 import
    except ImportError:
        return 0, '未安装 pymupdf，跳过 y/y 回校（requirements.txt: pymupdf==1.28.0）'
    try:
        pdf, _hdr = _http_get(release['pdf'])
        os.makedirs(cache_dir, exist_ok=True)
        with open(os.path.join(cache_dir, 'ice_release_%s.pdf' % month), 'wb') as f:
            f.write(pdf)
        with fitz.open(stream=pdf, filetype='pdf') as doc:
            text = '\n'.join(p.get_text() for p in doc)
    except Exception as e:                             # noqa: BLE001
        return 0, '新闻稿 PDF 取用失败（%r），本月跳过 y/y 回校' % (e,)

    sec, need_qual = _release_section(text, release['monthname'])
    if sec is None:
        return 0, '新闻稿正文里找不到 "%s highlights include:" 段，跳过 y/y 回校' \
                  % release['monthname']

    y, mo = int(month[:4]), int(month[5:])
    prev = '%04d-%02d' % (y - 1, mo)
    if prev not in data:
        return 0, '缺 %s，无法算同比，跳过 y/y 回校' % prev

    def val(mon, cols):
        vs = [data[mon][c] for c in cols]
        return None if any(v is None for v in vs) else sum(vs)

    checked, bad = [], []
    for c in _PR_CLAIM.finditer(sec):
        lab = re.sub(r'^(?:and\s+)?(?:record\s+)?', '',
                     re.sub(r'\s+', ' ', c.group('lab')).strip().lower())
        if lab == 'total average daily volume (adv)':
            lab = 'total'
        if lab not in _PR_METRICS:
            continue
        tail = c.group('tail') or ''
        qualified = re.match(r'\s*y/y\s+in\s+%s\b' % release['monthname'],
                             tail, re.I) is not None
        if need_qual and not qualified:
            continue
        # 常规段落里若断言自带 "in 4Q26" / "in 2026" 这种限定词，那是季度/年度数，不认
        if not need_qual and re.match(r'\s*y/y\s+in\s+(?:[1-4]Q|20\d\d)', tail, re.I):
            continue
        kind = c.group('kind').lower()
        is_oi = 'oi' == kind or 'interest' in kind
        cols = _PR_METRICS[lab][1 if is_oi else 0]
        if not cols:
            continue
        cur, pre = val(month, cols), val(prev, cols)
        if cur is None or not pre:
            continue
        claim = float(c.group('pct')) * (1 if c.group('dir').lower() == 'up' else -1)
        calc = (cur / pre - 1) * 100.0
        checked.append((lab, 'OI' if is_oi else 'ADV', claim, calc))
        if abs(calc - claim) > PR_TOL_PP:
            bad.append('%s %s：稿称 %+.0f%%，本模块解析值算出 %+.1f%%（%.4g → %.4g）'
                       % (lab, 'OI' if is_oi else 'ADV', claim, calc, pre, cur))
    if bad:
        raise IceFetchError(
            '%s 与官方新闻稿「%s」对不上 %d 条（容差 %.1fpp）：%s'
            % (month, release['headline'], len(bad), PR_TOL_PP, ' ; '.join(bad)))
    if len(checked) < PR_MIN_CLAIMS:
        return len(checked), ('新闻稿「%s」里只取到 %d 条可对账断言（<%d），'
                              '不足以构成一次有效回校，跳过'
                              % (release['headline'], len(checked), PR_MIN_CLAIMS))
    return len(checked), '与官方新闻稿「%s」逐条对账 %d/%d 一致（容差 %.1fpp）' % (
        release['headline'], len(checked), len(checked), PR_TOL_PP)


# ══════════════════════════════════════════════════════════════════════
# 发布日台账
# ══════════════════════════════════════════════════════════════════════
def _source_dates():
    """按路径加载仓库根的 source_dates.py。

    不能裸 import：本模块被 monthly_run 用 spec_from_file_location 加载，
    那时 sys.path 上既没有 fetch/ 也没有仓库根。
    """
    import importlib.util
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(root, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _record_source_date(series_dir, month, release, last_modified):
    """给「这一档文件自己带来的那个最新月」记一笔官方发布日，返回 (是否新记了一条, 说明)。

    只记最新月：同一个 xlsx 里躺着 2011 年以来的全部历史，但这一档的上线只能证明
    **最新月**是这天发的。顺手给 2011-03 也盖上今天的日期，等于凭空发明
    「2011 年 3 月的数据是 2026 年 8 月 5 日发的」，而页面会照印不误。

    已有记录一律不覆盖：ICE 会原地重传（口径坑 11 的追溯重刷），重传后的 Last-Modified
    更晚，覆盖等于把当初那次真发布的日期改错。

    三条出路各自返回一句说明，而不是一律 None —— 「已经记过」与「feed 里没有稿」
    是两件完全不同的事，调用方只有 None 可看时无法分辨，会把前者也喊成警告。
    """
    sd = _source_dates()
    if sd.lookup(series_dir, TICKER, month):
        return False, '发布日台账里已有 %s，不覆盖' % month
    if not release:
        return False, ('警告：feed 里没有 %s 的月度统计新闻稿，本月不记发布日'
                       '（页面抬头会省掉「官方发布于」那半句）' % month)
    evidence = ('新闻稿「%s」PressReleaseDate=%s（ir.theice.com/feed/PressRelease.svc）'
                % (release['headline'], release['date']))
    if last_modified:
        # 互证而非替代：Last-Modified 是文件上传时刻，只有与 ICE 自己宣告的日子同日
        # 才有意义；不同日就只写新闻稿那一条，别把两个日子混成一句。
        evidence += '；xlsx 直链 HTTP Last-Modified "%s" %s' % (
            last_modified,
            '与之同日，互证' if _lm_date(last_modified) == release['date']
            else '（不同日，仅记录）')
    sd.record(series_dir, TICKER, month, release['date'], evidence)
    return True, '%s 的官方发布日 %s 已记入 source_dates.csv' % (month, release['date'])


def _lm_date(last_modified):
    """'Wed, 05 Aug 2026 12:30:32 GMT' → '2026-08-05'；认不出返回 None。"""
    try:
        return _dt.datetime.strptime(
            last_modified.strip(), '%a, %d %b %Y %H:%M:%S %Z').strftime('%Y-%m-%d')
    except (ValueError, AttributeError):
        return None


def backfill_source_dates(series_dir, first_year=2015, last_year=None):
    """把历史各月的官方发布日一次性补进 series/source_dates.csv（**手工调用，cron 不跑**）。

    用途：核对 build/roster.py 那张 LAG 表准不准。README 说这张台账「随时可以回头核对」，
    对 ICE 而言「随时」可以一直回到 2015-09 —— feed 里最早的月度统计稿就在那。
    2011-2014 的稿 ICE 的 IR feed 里根本没有（逐年扫过，"statistic" 字样 0 命中），
    所以那 56 个月只能没有发布日，不许拿次月第 3 个交易日推一个填进去。

    默认不进 update()：monthly_run 每月该做的事是记一条，不是重写 124 条。
    命令行：python3 fetch/ice.py source-dates
    """
    last_year = last_year or _dt.date.today().year
    sd = _source_dates()
    got, added = {}, []
    for year in range(first_year, last_year + 1):
        got.update(_stat_releases(year))
    for mon in sorted(got):
        if sd.lookup(series_dir, TICKER, mon):
            continue
        rel = got[mon]
        sd.record(series_dir, TICKER, mon, rel['date'],
                  '新闻稿「%s」PressReleaseDate=%s（ir.theice.com/feed/PressRelease.svc，'
                  '按 year=%s 逐年拉取）' % (rel['headline'], rel['date'], mon[:4]))
        added.append(mon)
    return added


# ══════════════════════════════════════════════════════════════════════
# 对外接口
# ══════════════════════════════════════════════════════════════════════
def latest_month(cache_dir):
    """官方源当前最新月 'YYYY-MM'。

    以**表里最后一个 TOTAL FUTURES & OPTIONS 非空的月**为准，不信文件名 ——
    文件名里的 "July-2026" 是数据月不是发布月，而 HKEX 那种「先把下个月的列开出来再填数」
    的做法 ICE 哪天也可能学。抓不到 / 解析不出来一律抛 IceFetchError，不返回 None 掩盖故障。
    """
    path, url, _lm = _download_workbook(cache_dir)
    newest = _validate(parse_workbook(path))
    # 真值仍然只从内容取（见上）；文件名只用来对账，不等就炸。
    _crosscheck_workbook_month(url, newest)
    return newest


def _fmt(v):
    """写回 CSV。整数写整数（官方月度格本来就是整千张/整百万股），其余用最短往返表示。

    不无脑 repr(float)：那会把 2568 写成 '2568.0'，与 cme.csv / hkex.csv 的风格不一致，
    也让人拿 CSV 与官方原表逐位对照时多一层心智负担。
    """
    if v is None:
        return ''
    f = float(v)
    return str(int(f)) if f.is_integer() and abs(f) < 1e15 else repr(f)


def update(series_dir, cache_dir):
    """把新月份写进 series/ice.csv，返回新增月份列表（升序）。

    幂等保证：
      · 已存在的月份不重复追加；
      · 已经有值的单元格**永不覆盖** —— 官方做过追溯重刷（口径坑 11），重刷不由
        无人值守任务自动吞进来；官方与本仓不一致的格子写进 cache/ice_restatements.csv
        供人工判断；
      · 只在既有行**原本为空**的格子上回补（正常情况下不会有：ICE 一次给全 55 列）；
      · 什么都没变时，未被触碰的单元格是原样字符串搬运 ⇒ 文件字节级不变。

    首次调用时 series/ice.csv 不存在 —— 本模块会按 COLUMNS 建表并一次写满全历史
    （2011-01 起 187 个月）。这不是"顺手做掉"：ICE 的单一 xlsx 本来就带全序列，
    分两步（先手工 bootstrap 再增量）反而会出现「bootstrap 脚本与 fetch 解析器两套代码」
    的经典漂移，README 讲的 cost/ibkr 搬迁就是在治这个病。
    """
    csv_path = os.path.join(series_dir, 'ice.csv')
    if os.path.exists(csv_path):
        with open(csv_path, newline='', encoding='utf-8') as f:
            rows = list(csv.reader(f))
        header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
        if header != ['month'] + COLUMNS:
            missing = [c for c in COLUMNS if c not in header]
            raise IceFetchError(
                'series/ice.csv 的列名与本模块不符；缺 %s，多 %s'
                % (missing, [c for c in header[1:] if c not in COLUMNS]))
    else:
        header, body = ['month'] + COLUMNS, []

    idx = {name: i for i, name in enumerate(header)}
    path, url, last_modified = _download_workbook(cache_dir)
    data = parse_workbook(path)
    newest = _validate(data)

    # 独立于解析器的月份判据：工作簿自报的数据月必须等于解析出来的最新月。
    _crosscheck_workbook_month(url, newest)

    # 外部证据：拿官方新闻稿的 y/y 回校最新月。取不到不阻断，但一定把话说清楚。
    releases = _stat_releases(-1)
    _crosscheck_release_month(releases, newest)
    n_ok, note = _crosscheck_release(data, newest, releases.get(newest), cache_dir)
    print('[ice] %s' % note)

    have = {r[0]: r for r in body}
    added, filled, restated = [], [], []
    for mon in sorted(data):
        rec = data[mon]
        if mon in have:
            row = have[mon]
            for name in COLUMNS:
                if rec[name] is None:
                    continue
                new = _fmt(rec[name])
                if not row[idx[name]].strip():
                    row[idx[name]] = new
                    filled.append((mon, name, new))
                elif row[idx[name]] != new:
                    restated.append((mon, name, row[idx[name]], new))
            continue
        row = [''] * len(header)
        row[0] = mon
        for name in COLUMNS:
            row[idx[name]] = _fmt(rec[name])
        have[mon] = row
        body.append(row)
        added.append(mon)

    # 官方与本仓不一致的格子一律落盘、绝不自动覆盖 —— 是口径重述还是解析出错，
    # 只有人能判断（照 fetch/hkex.py 的做法）。
    if restated:
        os.makedirs(cache_dir, exist_ok=True)
        rp = os.path.join(cache_dir, 'ice_restatements.csv')
        with open(rp, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f, lineterminator='\n')
            w.writerow(['month', 'column', 'in_series_csv', 'in_official_xlsx'])
            w.writerows(restated)
        print('[ice] 官方源与 series 有 %d 处不一致，已写 %s（本模块不覆盖，请人工判断）'
              % (len(restated), rp))

    # 落盘只在真有变化时做 —— 什么都没变时连写都不写，文件 mtime 都不动。
    if added or filled:
        body.sort(key=lambda r: r[0])
        tmp = csv_path + '.tmp'
        with open(tmp, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f, lineterminator='\n')
            w.writerow(header)
            w.writerows(body)
        os.replace(tmp, csv_path)      # 原子替换：中途挂掉不会留下半张表
        if filled:
            print('[ice] 补空 %d 格：%s' % (len(filled), filled[:12]))
        print('[ice] 源文件 %s（Last-Modified %s）'
              % (url.rsplit('/', 1)[-1], last_modified))

    # 记发布日放在落盘之后：写盘失败就不该留下「这个月官方发过了」这条断言。
    # 但**不设 `newest in added` 这个条件**，也不放在上面的 early return 后面：
    # 上一次若跑到「ice.csv 已落盘、发布日还没记」之间被打断（cron 被 kill、机器休眠
    # 都会），重跑时 newest 已在表里、进不了 added，那一条就**永久**缺席 ——
    # 而它的表现只是页面少半句「官方发布于」，不报错，没人会发现。
    # sd.record 按 (ticker, month) 去重，所以无条件调用是幂等的：已经记过就原样返回，
    # 不重写、不覆盖。fetch/cboe.py 的 _record_source_dates 也是这么写的。
    wrote, note = _record_source_date(series_dir, newest,
                                      releases.get(newest), last_modified)
    if wrote or note.startswith('警告'):
        print('[ice] %s' % note)
    return sorted(added)


if __name__ == '__main__':
    import sys
    _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _series, _cache = os.path.join(_root, 'series'), os.path.join(_root, 'cache')
    if len(sys.argv) > 1 and sys.argv[1] == 'source-dates':
        print('source_dates 回补:', backfill_source_dates(_series))
    elif len(sys.argv) > 1 and sys.argv[1] == 'latest':
        print('latest:', latest_month(_cache))
    else:
        _added = update(_series, _cache)
        print('added : %d 个月 %s' % (len(_added),
                                     (_added[:3] + ['…'] + _added[-3:])
                                     if len(_added) > 6 else _added))
