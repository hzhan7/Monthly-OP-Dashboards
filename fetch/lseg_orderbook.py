# -*- coding: utf-8 -*-
"""LSEG Capital Markets 现货腿 —— LSE 与 Turquoise 电子订单簿月度成交（无人值守抓取）。

本模块只负责 LSEG 四条腿里的 **Capital Markets / 现货订单簿** 这一条：
London Stock Exchange 主板订单簿 + Turquoise（Integrated 与暗池）成交额、笔数、
日均、交易日数，以及官方自己公布的两个份额（LSE 在英国 Lit 订单簿的份额、
Turquoise 在泛欧的份额）。衍生品、清算（LCH）、Tradeweb、数据业务都不在这里。

━━ 数据源（两个，互为独立核对）━━

**主源：LSEG Monthly Market Report（每月一个 PDF）**
  标题固定为 `LSEG market report <Month> <Year>`，第 1 页表名
  "LSEG - Electronic Order Book Trading"。本模块只读第 1 页最上面那张 **MTD 表**。
  文件落在 https://docs.londonstockexchange.com/sites/default/files/reports/ 。

**副源（只做核对，不入库）：`Order book trading` 工作簿**
  同一个文件服务器上的一个 xlsx，三个 sheet（Daily / Monthly / Yearly Order Book
  Trading）。Monthly sheet 从 **1997-10 到当月**共 346 行（2026-08-07 实测），
  给的是 LSE 订单簿的成交笔数、成交额（**精确到便士**，不是 PDF 那样四舍五入到 £m）、
  交易日数、日均笔数、日均成交额。每月逐笔核对 PDF 的 LSE 那三个数，对不上就抛异常。

⚠️ **两个文件的 URL 都不能写死，必须每次从官方检索接口里查出来**，理由见下面「文件名坑」。

━━ 怎么找到文件（**不许猜文件名**）━━

londonstockexchange.com 是 Angular SPA，`/reports?tab=...` 那个页面 curl 下来只有空壳，
文件链接是运行时才注入的。但站内检索有一个**可以直接 curl 的 JSON 接口**：

    GET https://api.londonstockexchange.com/api/v1/pages
        ?path=search&parameters=<双重 urlencode 的 "q=...&tab=documents&size=..&page=..">

`parameters` 是 **双重编码**：内层先把 `q=LSEG market report June 2026&tab=documents`
整串 urlencode 一次，再整体 urlencode 一次（浏览器发的就是 `q%253D...%2526tab%253D...`）。
少编一层不会报错，只会**静音返回错误的结果集**，所以 `_search()` 里那两层 quote 谁都别删。

返回体里 `components[type=="search"].content[0].value.pagesdocuments` 是文档命中列表，
每条形如：

    {"url": "https://docs.londonstockexchange.com/sites/default/files/reports/
             LSEG%20market%20report%20June%202026.pdf",
     "world": "documents", "title": "LSEG market report June 2026",
     "lastupdate": "2026-07-30T21:04:07"}

本模块按 title 取链接，**绝不按文件名拼 URL**：先全等匹配，全等落空再退一步做
「标题里同时含 'market report' 与该月标签、且 url 以 .pdf 结尾」的宽松匹配
（见 `_find_month_report`）。退这一步的理由见下面「文件名坑 2」与「护栏」两节 ——
CMS 标题历史上就漂过，而标题一漂，全等匹配返回的是 None，那条路是**静默**的。

📌 已经踩到的两个文件名坑（正是「不许猜」的理由）：
  1. **同名重传会被 Drupal 加后缀**：`Order book trading_1558.xlsx`（数字每传一次 +1）、
     `LSEG market report January 2026_1.pdf`、`March 2022_1`、`September 2022_0`、
     `December 2022_0`。不带后缀的 `Order book trading.xlsx` **确实存在且返回 200**，
     但 Last-Modified 停在 **2020-09-15**（实测）—— 这是最恶心的一种坑：
     猜出来的 URL 不报错，只是永远给你六年前的数据。
  2. **少数条目的 title 里带扩展名**：2022-06 那期在 CMS 里的标题是
     `LSEG market report June 2022.pdf`（其余月份都不带）。所以匹配时要把 title 末尾的
     `.pdf` 去掉再比。

反面教材记在这里免得下次重来：docs.londonstockexchange.com 的目录页返回 403，
Angular 的 `/api/v1/components/refresh` 用页面里那些 block_content id 去拉 tab 模块
**一律返回 `[]`**（页面级 component 如 hero 却能拉到，说明 id 不通用），
web.archive.org 在本机是黑名单。别再往这三条路上走。

━━ 抓取方式与依赖 ━━
`urllib.request` 裸奔即可：api.londonstockexchange.com 与 docs.londonstockexchange.com
都是 CloudFront，实测无 Cloudflare/Akamai 挑战、无 JS 渲染、无登录墙、无 JA3 指纹拦截，
**满足无人值守**。PDF 解析用 PyMuPDF（`fitz`，仓里已有），xlsx 用 openpyxl（仓里已有）。

━━ 实测发布节奏 ━━
**闸门样本 = 2021-01 起**，逐期读 PDF 内嵌的 /CreationDate（Excel 导出时间戳），
算相对**数据月月末**的日历天数。复算：`python3 fetch/lseg_orderbook.py --cadence`
（只读 cache/lseg_orderbook/mmr_*.pdf，不联网）。下面是 2026-08-19 重算：

    2021-01 起 67 期（含新到的 2026-07）：最早 +1 天，最晚 +51 天，中位 +4 天
    季末月与非季末月**没有系统差异**（中位 4.5 vs 4.0），
    所以 LAG / EARLY_BY 两个位置写同一个值即可。

⚠️ **节奏在 2024 年明显变慢，定闸门只能照近两年，不能照全样本中位数：**

    2021 年 12 期  +1..+6  天，中位 +2
    2022 年 12 期  +1..+5  天，中位 +2
    2023 年 12 期  +1..+6  天，中位 +2.5
    2024 年 12 期  +5..+20 天，中位 +9.5
    2025 年 12 期  +2..+15 天，中位 +10
    2026 年  7 期  +2..+51 天，中位 +19（2026-07 = +4，8 月 4 日发）

    近 30 期（2024-01 起）实测最晚 = **2026-03 的数据 → 2026-04-24（+24 天）**，
    实测最早 = +2 天（2025-06 → 07-02、2026-06 → 07-02）。

⚠️ **2016-01…2020-12 那新回补的 60 期不进闸门样本** —— 与 fetch/lseg_primary.py 里
   AIM 2017-01…2018-04 那 16 期同一个坑：CreationDate 被重传污染。逐年中位看着更快
   （2017 中位 +3、2020 中位 +2.5），但样本里藏着 **2019-04 的 +1855 天**
   （CreationDate 2024-05-28，五年后重导的文件）与 2016-09 的 +49 天。
   一段既有 +1 又有 +1855 的分布描述的是重传时刻，不是当年的发布节奏。
   ⇒ **回补没有、也不该改动 LAG / EARLY_BY。**

📌 那个 +51 天（2026-01 数据，CreationDate 2026-03-23）**只能当上界读**：
   这一期的文件名是 `..._1.pdf`，是重传件，CreationDate 记的是重新导出的时间而不是首发时间。
   但同日（2026-03-23）生成的还有 2026-02 那期（文件名无后缀，不是重传），
   说明 2026 年初确实积压了两个月一起补发 —— 这不是纯粹的重传假象。

给 docs/CRON_WIRING.md §2 的建议（按本仓判据：LAG 照实测最晚、开闸日=实测最早）：

    LAG['lseg'] = (26, 26)          # 近 30 期最晚 +24，留 2 天余量
    EARLY_BY['lseg'] = (25, 25)     # 26 − 25 = 次月第 1 天开闸 = 实测最早发布日

⚠ EARLY_BY 必须写成**元组**：monthly_run.py 取值处是
`EARLY_BY.get(t, (EARLY, EARLY))[1 if qe else 0]`，写成裸整数会在下标那步 TypeError，
崩掉的是**整轮** monthly_run，不只是这一家。

📌 另一个必须知道的滞后：**副源 xlsx 自己也会落后。** 2026-08-07 实测，
`Order book trading` 工作簿的 Last-Modified 停在 2026-07-30 21:04 GMT、数据只到 7 月 30 日；
同一天 2026-07 那期月报**还没发**（检索接口查无此条）。所以「LSEG 的月度数据慢」
是常态，别把 NOCHANGE 当成抓取坏了。

━━ 护栏：这一路「读不到」时必须出声 ━━

检索接口按标题找不到月报时返回的是 None 而不是异常，`fetch_rows` 把该月记进
`missing` 就继续往下走。而稳态下 `skip` 覆盖全部已入库月份 ⇒ rows 为空 ⇒ 函数走
「全部已入库，无新月份」那一支干净返回 []，`fetch/lseg.py` 那边 after−before 为空、
只打一行 `orderbook ok N 个月`。**标题模板一改，这一路就会这样连年安静空转**：
没有异常、没有缺列、没有断档，而 orderbook 又是 build/specs 里的慢腿（页面自带
「最新月留空是正常的」那句说明、不点红点），growing 的右边缘缺口反而被解释掉了。
README 判据一句话：连续失败十天和成功十天在日志里长得一样，就缺一道护栏。

三道，照 fetch/msci.py 的形状，缺一不可：

  ① **宽松匹配**（`_find_month_report`）：只挡**已知**的那种漂法（改词、带扩展名）。
     选错文件不会静默 —— `parse_report` 核 PDF 自己印的抬头月份、`_sanity` 核
     日均×交易日恒等式、`_crosscheck` 核副源工作簿，三道都在下游等着。
  ② **逾期哨兵**（`_guard_overdue_missing`）：`missing` 无条件打印，且其中任何一个月
     离数据月月末超过 `_MAX_PUBLISH_LAG_DAYS` 就抛。这道才是挡住**下一种没见过的**
     漂法的那道 —— ① 只认识见过的变体。别只留 ①。
  ③ **独立外部判据**（②里那段 judge，形状同 `fetch/cboe.py::_crosscheck_report_month`
     与 `fetch/ice.py::_crosscheck_workbook_month`）：过了阈值之后再去问副源工作簿
     「这个月它算不算完整月」，把「官方还没发」和「我们找不到」分开写进错误消息。
     ⚠ 副源**只能作证、不能当触发器**：它在次月头几天就把上个月记成完整月（口径坑 6），
     而月报 PDF 的 2026 年中位滞后是 +19 天 —— 拿它当触发器等于每个月误报三个星期，
     而每月假一次的警报，人很快就学会无视了。

抛出来的后果是**降级不是 FAIL**：`fetch/lseg.py` 的 `_update` 逐路 catch，所以这里抛
只让 orderbook 这一路当天不前进（日志多一行 `⚠ orderbook 这一路本轮失败` 与
「本轮降级的路：orderbook」），另外三路照常合流发布。这正是要的形状：出声，
但不牵连别人。

━━ 口径坑（按踩坑概率排序）━━

**1. 只读第一张表（MTD），YTD 那张表是坏的。**
   2026-06 那期的 YTD 区块里 Turquoise Integrated 的 £m 印成 28,264（跟 MTD 一模一样）、
   €m 印成 204,771、同比 −83% —— 官方自己排版错了。本模块只解析第 1 页
   "Average Daily" 之前 + "Trading days" 表，YTD 区块从不碰。谁要加 YTD 列先去核对这一期。

**2. 行标签改过三轮名字，同一条序列换过三个标题。**
   · LSE 现货：`UK order book`（→2020-12）→ `LSE Order Book`（2021-01→）
     2021-01 改名的同时，报告里 `Italian order book` 那一行消失 —— 那是 Borsa Italiana
     卖给 Euronext（2021-04 交割）前后的口径切换。本模块从来只取 UK/LSE 那一行，
     所以**这条序列本身不含意大利**，2021 年前后不存在口径断点。
   · Turquoise 暗池：`Turquoise MidPoint` → `Turquoise Plato™`（2017 起）→
     `Turquoise Dark`（2026-01 起，实测 2026-06 那期已是 Dark）。三个名字是同一条腿，
     统一写进 `turquoise_dark_*`。
   · 份额行：`UK Lit Orderbook trading`（→2020-12）→ `LSE Lit Orderbook trading in UK`。
   标签里的 `™` 和结尾空格都要 strip（"Turquoise Dark " 在 Average Daily 区块里带尾空格）。

**3. 起点 2016-01。**（2026-08-18 从 2021-01 前推 60 个月）
   上一版这里写的是「2021-01 起，因为更早的 PDF 少列、且 2021-01 才出现
   `LSE Lit Orderbook trading in UK` 这个标签」。**后半句站不住**：份额行的旧标签
   `UK Lit Orderbook trading` 早就在 `UK_SHARE_LABELS` 里了，2016-2020 那 60 期
   17 列一列不缺。实测 2016-01..2020-12 逐期解析 **60/60 通过**，
   其中 2017-01..2020-12 共 48 期是**零代码改动**直接过的。
   前半句（「更早的 PDF 还额外印 Italian / Derivatives / MTS / EuroTLX 等行」）
   属实，但那些是**多出来的行**，本模块按标签取行、从不按行号，多几行不影响。

   为了收下 2016 全年，只加了两处兼容，都写在各自函数的 docstring 里：
     · `_take_days()`：交易日那一行 2016-12 及以前只有 2 列（无 YTD），2017-01 起 4 列。
     · `_OFFICIAL_INCONSISTENT`：2016-03 官方自己的日均与交易日对不上（放行恒等式，
       不放松通用容差；入库仍是官方原值）。
   另外双源比对改成「绝对容差与相对容差取宽者」，理由见 `_XCHECK_REL`，
   并把 2016-07 登记进 `_RESTATED`（官方对那一个月做过实质下修）。

   **再往前（2013-01..2015-12）没做。** 那 36 期还要处理官方错字 `Turquouse`、
   空格千分位、以及 2013 年的旧抬头字符串 —— 收益（3 年）与风险（三处只在旧档
   出现一次的特判）不成比例，而 2016-01 已经是全站统一的起点。

**4. 成交额单位是「官方印的 £m」，不做换算。**
   PDF 只印到百万英镑整数（146,827 = £146.827bn）。xlsx 有精确到便士的同一个数
   （146,827,153,894.25）。本模块**入库 PDF 的 £m**、拿 xlsx 做核对（容差见
   `_crosscheck()` 里那段实测记录）。反过来（入库 xlsx 精确值）会让 LSE 那几列和
   Turquoise 那几列**精度口径不一致**，画在同一张图上没人说得清差在哪。
   两源逐月对得上。**2026-08-19 用 cache 里的 127 份 PDF + 副源工作簿离线重算**
   （127 期里 126 期参与比对：2016-07 是官方真重述月，按 `_RESTATED` 跳过）：
       笔数        最大差 313 笔     @2019-01（20,025,721 vs 20,026,034）
       成交额      最大差 5.41 £m    @2016-02（113,828 vs 113,833.41）
       日均成交额  最大差 0.71 £m    @2016-01（相对 1.37e-4，全表最大相对偏差）
       日均笔数    最大差 14.27      @2019-01
       交易日      最大差 0 天
   ⚠ 上一版这里写的「66 个月、最大差 8 笔 / 1.30 £m」只是 2021-2026 那一段。
     偏差随年头变大不是解析变差，是副源持续重述、PDF 是发布当时的快照 —— 见 `_XCHECK_REL`。

**5. 交易日数 LSE 与 Turquoise 不一样，各存一列。**
   Turquoise 跟的是泛欧日历（2026-06：LSE 22 天 / Turquoise 22 天，但 2019-12 是
   UK 20 / Turquoise 20 / Italy 18，差异真实存在）。日均成交额=月成交额÷各自的交易日数，
   别拿 LSE 的天数去除 Turquoise 的成交额。

**6. 副源 xlsx 的最后一个月经常是残月，核对时要跳过。**
   2026-08-07 实测：xlsx 的 Daily sheet 最新一天是 2026-07-30，Monthly sheet 里
   `Jul-26` 只有 22 个交易日（7 月实际 23 个），即**当月未走完就已入表**。
   `_xlsx_monthly()` 因此只承认「Daily sheet 里已经出现过更晚月份」的那些月，
   剩下的直接丢掉 —— 拿残月去核对 PDF 会得到一个假的不一致。

**7. 份额是百分数，入库存百分点数值（69.2 表示 69.2%），不是 0.692。**
"""
import csv
import datetime
import json
import os
import re
import time
import urllib.parse
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
CACHE = os.path.join(ROOT, 'cache', 'lseg_orderbook')
SERIES = os.path.join(ROOT, 'series')

SEARCH_API = 'https://api.londonstockexchange.com/api/v1/pages'
UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
      '(KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36')
REFERER = 'https://www.londonstockexchange.com/search'

# 起点见 docstring 口径坑 3。终点由官方发到哪算哪。
START_MONTH = '2016-01'

MONTH_NAMES = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
               'August', 'September', 'October', 'November', 'December']

# ── 行标签 → 列前缀。全部先经 _norm_label() 归一化（去 ™、去尾空格、小写、压空格）。
LSE_LABELS = ('lse order book', 'uk order book')
TQ_INT_LABELS = ('turquoise integrated',)
TQ_DARK_LABELS = ('turquoise dark', 'turquoise plato', 'turquoise midpoint')
UK_SHARE_LABELS = ('lse lit orderbook trading in uk', 'uk lit orderbook trading')
TQ_SHARE_LABELS = ('turquoise total pan european trading',)
LSE_DAY_LABELS = ('lse', 'uk')
TQ_DAY_LABELS = ('turquoise',)

COLUMNS = [
    'lse_orderbook_value_gbp_m',
    'lse_orderbook_trades_count',
    'lse_orderbook_adv_gbp_m',
    'lse_orderbook_avg_daily_trades_count',
    'lse_trading_days_count',
    'turquoise_integrated_value_gbp_m',
    'turquoise_integrated_trades_count',
    'turquoise_integrated_adv_gbp_m',
    'turquoise_integrated_avg_daily_trades_count',
    'turquoise_dark_value_gbp_m',
    'turquoise_dark_trades_count',
    'turquoise_dark_adv_gbp_m',
    'turquoise_dark_avg_daily_trades_count',
    'turquoise_trading_days_count',
    'lse_lit_uk_share_pct',
    'turquoise_paneuropean_share_pct',
    'gbp_eur_rate',
]


class LsegOrderbookFetchError(RuntimeError):
    """本模块所有失败路径统一抛它，调度器只需 catch 一种。"""


# ──────────────────────────────────────────────────────────────── HTTP 与检索

def _http_get(url, timeout=90, tries=3, pause=1.5):
    """带退避的 GET。docs/api 两个域都是 CloudFront，偶发 5xx 重试即可。"""
    last = None
    for i in range(tries):
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': UA, 'Referer': REFERER,
                'Accept': 'application/json,text/html,*/*'})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read()
        except Exception as e:                      # noqa: BLE001 —— 什么错都重试
            last = e
            if i + 1 < tries:
                time.sleep(pause * (i + 1))
    raise LsegOrderbookFetchError(f'GET 失败 {url}: {last}')


def _search(q, size=20, page=0, tab='documents'):
    """站内检索。返回 pagesdocuments 列表（可能为空）。

    `parameters` 必须双重 urlencode，见 docstring「怎么找到文件」。
    """
    inner = urllib.parse.urlencode({'q': q, 'tab': tab,
                                    'size': str(size), 'page': str(page)})
    once = urllib.parse.quote(inner, safe='')
    twice = urllib.parse.quote(once, safe='')
    url = f'{SEARCH_API}?path=search&parameters={twice}'
    raw = _http_get(url)
    try:
        doc = json.loads(raw)
    except Exception as e:
        raise LsegOrderbookFetchError(f'检索接口返回的不是 JSON（q={q!r}）: {e}') from e
    comps = [c for c in (doc.get('components') or []) if c.get('type') == 'search']
    if not comps:
        raise LsegOrderbookFetchError(
            f'检索接口没有 search 组件（q={q!r}）—— 接口改版了，别再往下解析')
    content = comps[0].get('content') or []
    if not content:
        return []
    return (content[0].get('value') or {}).get('pagesdocuments') or []


def _title_key(title):
    """CMS 里少数条目的 title 带扩展名（2022-06 那期），比对前去掉。"""
    t = (title or '').strip()
    for ext in ('.pdf', '.xlsx', '.xls'):
        if t.lower().endswith(ext):
            t = t[:-len(ext)]
    return re.sub(r'\s+', ' ', t).strip().lower()


def _find_doc(title):
    """按标题精确匹配拿一条文档记录，找不到返回 None（不抛）。"""
    for hit in _search(title, size=30):
        if _title_key(hit.get('title')) == _title_key(title):
            if not (hit.get('url') or '').startswith('http'):
                continue                            # 页面命中不是文档，跳过
            return hit
    return None


def _find_month_report(month):
    """取某个月的 Monthly Market Report 文档记录；找不到返回 None（不抛）。

    先全等匹配，落空再退一步做宽松匹配。为什么要退这一步：CMS 标题会漂 ——
    2022-06 那期的标题带 `.pdf`（`_title_key` 已经吃掉），而**下一次漂成什么样
    没人知道**（改成 "LSEG Monthly Market Report August 2026"、挂个 "(revised)"、
    把空格换成不间断空格，都会让全等匹配返回 None）。全等落空是**静默**的，
    正是本模块要拦的那种失败形状，所以宽松匹配的意义是把「标题小改」从
    「安静空转一年」拉回「照常抓到」。

    宽松判据是三条**与**起来：标题含 'market report'、标题含该月标签
    （'august 2026'）、url 以 .pdf 结尾。2026-08-26 实测全站只有
    `LSEG market report <Month> <Year>` 这一族 PDF 用到 "market report" 这两个词
    （拿 'market report June 2026' 与 'AIM market report June 2026' 各拉 30 条命中
    逐条核过，没有第二族），所以三条与起来在今天是唯一的。
    **命中多于一份不同 url 时宁可抛、绝不猜**：那说明将来真出了同名的第二族文档，
    该由人决定取哪一份。

    退这一步之后仍然安全，靠的是下游三道：`parse_report` 核 PDF 自己印的抬头月份
    （下错月直接抛）、`_sanity` 核恒等式、`_crosscheck` 核副源。而「宽松匹配也认不出」
    这一种由 `_guard_overdue_missing` 兜，不在这里兜 —— 正则只认识见过的变体。
    """
    title = f'LSEG market report {_month_label(month)}'
    hits = _search(title, size=30)
    key = _title_key(title)
    for hit in hits:
        # 第一轮与 `_find_doc` 完全同义：全等命中时行为一个字节都不变。
        if _title_key(hit.get('title')) == key and (hit.get('url') or '').startswith('http'):
            return hit
    label = _month_label(month).lower()
    loose = {}
    for hit in hits:
        url = hit.get('url') or ''
        if not url.startswith('http') or not url.lower().endswith('.pdf'):
            continue                            # 页面命中不是文档；非 PDF 不是月报
        tk = _title_key(hit.get('title'))
        if 'market report' in tk and label in tk:
            loose.setdefault(url, hit)
    if len(loose) > 1:
        raise LsegOrderbookFetchError(
            f'{month}: 标题全等匹配落空，宽松匹配又命中 {len(loose)} 份不同文档 '
            f'{sorted(loose)} —— 拒绝猜，请人工确认哪一份才是 Monthly Market Report')
    if loose:
        hit = next(iter(loose.values()))
        print(f'[lseg_orderbook] ⚠ {month}: CMS 标题已经不是 {title!r}，'
              f'宽松匹配取到 {hit.get("title")!r} —— 全等匹配这一路已经失效，'
              f'核对无误后把新写法补进 _find_month_report 的判据')
        return hit
    return None


#: 副源工作簿的缓存寿命（小时）。**只有这一份文件开缓存过期，mmr_*.pdf 一律不开。**
#: 工作簿是「同一个文件名持续更新」的，而 `_download` 一命中缓存就直接返回 ——
#: 2026-08-26 实测 cache/lseg_orderbook/order_book_trading.xlsx 的 mtime 停在
#: 2026-08-07、里面最新一天是 2026-07-30，也就是说这道「独立外部判据」自己先冻了
#: 大半个月。冻住的后果不是报错：`_crosscheck` 对**新月份**永远查不到对应行，转而
#: 打印一句「副源 xlsx 尚无这些月份，未做双源核对」—— 一道从不生效的护栏，在日志里
#: 和一道生效的护栏长得一模一样，正是第四类失败。24 小时 = 每天最多重下一次。
_XLSX_MAX_AGE_HOURS = 24.0


def _download(url, path, min_bytes=20000, max_age_hours=None):
    """下到 cache/。判有效而不只判存在 —— 0 字节残骸会让后续每轮都以为「已经有了」。

    `max_age_hours` 默认 None = 关，**不要给 mmr_*.pdf 打开它**：那 127 份是每月一份的
    不可变快照，重下只会拉到重传件（文件名带 `_1`），而重传件的 /CreationDate 记的是
    重新导出那天，会污染 `--cadence` 那张节奏表（文件头「+1855 天」那条就是这么来的），
    而 LAG / EARLY_BY 又是照那张表定的。只有副源工作簿要开它，见 `_XLSX_MAX_AGE_HOURS`。
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    stale = False
    if os.path.exists(path) and os.path.getsize(path) >= min_bytes:
        if max_age_hours is None:
            return path
        if (time.time() - os.path.getmtime(path)) / 3600.0 < max_age_hours:
            return path
        stale = True     # 手上有一份能用的旧档：重下失败时不能先把它删了（见下）
    if os.path.exists(path) and not stale:
        # 只删「残骸」。过期但完整的那份留到新文件真的下回来为止 —— 先删后下会让一次
        # 网络抖动把副源判据整个抹掉，而抹掉之后 `_crosscheck` 只会安静地少核几个月。
        os.remove(path)
    data = _http_get(url, timeout=120)
    if len(data) < min_bytes:
        raise LsegOrderbookFetchError(
            f'下到的文件只有 {len(data)} 字节，疑似拦截页: {url}')
    with open(path, 'wb') as f:
        f.write(data)
    return path


# ──────────────────────────────────────────────────────────────── 月份小工具

def _month_label(month):
    y, m = month.split('-')
    return f'{MONTH_NAMES[int(m) - 1]} {y}'


def _months(start, end):
    y0, m0 = (int(x) for x in start.split('-'))
    y1, m1 = (int(x) for x in end.split('-'))
    out = []
    while (y0, m0) <= (y1, m1):
        out.append(f'{y0}-{m0:02d}')
        m0 += 1
        if m0 == 13:
            y0, m0 = y0 + 1, 1
    return out


def _prev_month(month):
    y, m = (int(x) for x in month.split('-'))
    return f'{y - 1}-12' if m == 1 else f'{y}-{m - 1:02d}'


def _today_month():
    d = datetime.date.today()
    return f'{d.year}-{d.month:02d}'


def _month_end(month):
    """'YYYY-MM' → 该月最后一天的 date。

    逾期哨兵按「离**数据月月末**几天」算，和文件头那张实测节奏表、和 `cadence()`
    是同一个口径 —— 阈值是从那张表读出来的，两边口径必须一致，否则阈值就不是它
    看起来的那个意思了。
    """
    y, m = (int(x) for x in month.split('-'))
    return datetime.date(y + (m == 12), (m % 12) + 1, 1) - datetime.timedelta(days=1)


# ──────────────────────────────────────────────────────────────── PDF 解析

_NUM = re.compile(r'^-?[\d,]+(?:\.\d+)?%?$')


def _norm_label(s):
    s = (s or '').replace('™', '').replace('\xa0', ' ')
    return re.sub(r'\s+', ' ', s).strip().lower()


def _num(tok):
    """'146,827' → 146827.0；'69.2%' → 69.2；不是数就返回 None。"""
    t = (tok or '').strip().replace('\xa0', '')
    if not _NUM.match(t):
        return None
    return float(t.rstrip('%').replace(',', ''))


def _take(lines, i, n, month, where):
    """从 lines[i+1] 起取 n 个数值行，中间不允许夹非数值行。"""
    out = []
    j = i + 1
    while j < len(lines) and len(out) < n:
        v = _num(lines[j])
        if v is None:
            raise LsegOrderbookFetchError(
                f'{month} {where}: 第 {len(out) + 1} 个数位置上是 {lines[j]!r}，'
                f'不是数字 —— 排版变了，拒绝猜')
        out.append(v)
        j += 1
    if len(out) < n:
        raise LsegOrderbookFetchError(f'{month} {where}: 只取到 {len(out)}/{n} 个数')
    return out


def _take_days(lines, i, month, where):
    """交易日那一行：取到第一个非数字为止，返回 MTD（第一个数）。

    为什么不能沿用 `_take(..., 4, ...)`：**这一行的列数按年份变**。
      · 2017-01 起：`Trading days … MTD | YTD | Prev. Yr MTD | Prev. Yr YTD`
        ⇒ `UK   22   125   22   125`，**4 个数**
      · 2016-12 及以前：表头只有 `Trading days`，没有 YTD 那两列
        ⇒ `UK   22   22`，**2 个数**，第 3 行位置上是下一个标签 `Italy`
    写死 4 会在 2016 全年 12 期上抛「第 3 个数位置上是 'Italy'」—— 那不是排版坏了，
    是那一代版式本来就只印两列。两代都只用第一个数（MTD），所以这里放宽的是**列数**，
    不是**取哪一列**：仍然严格要求首格是数字、且至少两格（MTD + 去年同期 MTD）。
    """
    out = []
    j = i + 1
    while j < len(lines) and len(out) < 4:
        v = _num(lines[j])
        if v is None:
            break
        out.append(v)
        j += 1
    if len(out) not in (2, 4):
        raise LsegOrderbookFetchError(
            f'{month} {where}: 取到 {len(out)} 个数（{out}），既不是老版式的 2 列、'
            f'也不是新版式的 4 列 —— 排版变了，拒绝猜')
    return out[0]


def _find_line(lines, labels, lo, hi, month, where):
    for i in range(lo, min(hi, len(lines))):
        if _norm_label(lines[i]) in labels:
            return i
    raise LsegOrderbookFetchError(
        f'{month} {where}: 在第 1 页找不到标签 {labels}（标签又改名了？）')


def parse_report(path, month):
    """解析 Monthly Market Report 第 1 页的 MTD 表，返回 {列: 值}。

    只碰第一张表：`Average Daily` 之前是月合计区块，之后到 `Exchange Rate` 是日均区块，
    再往后是汇率 / 份额 / 交易日。`Trading days` 之后还有一张 YTD 表，本函数不看。
    """
    import fitz
    doc = fitz.open(path)
    if doc.page_count < 1:
        raise LsegOrderbookFetchError(f'{month}: PDF 没有页面 {path}')
    text = doc[0].get_text()
    lines = [ln.rstrip() for ln in text.split('\n')]
    norm = [_norm_label(ln) for ln in lines]

    head = next((i for i, s in enumerate(norm)
                 if s.startswith('lseg - electronic order book trading')), None)
    if head is None:
        raise TypeError('not-an-eob-report')       # 由调用方翻译成更友好的错误

    want = _norm_label(_month_label(month))
    if want not in norm[head:head + 4]:
        raise LsegOrderbookFetchError(
            f'{month}: PDF 抬头写的不是 {want!r}（前几行 {lines[head:head + 4]}）'
            f' —— 下到了别的月份，拒绝入库')

    def idx(pred, what):
        for i in range(head, len(norm)):
            if pred(norm[i]):
                return i
        raise LsegOrderbookFetchError(f'{month}: 第 1 页找不到 {what} 区块')

    i_avg = idx(lambda s: s == 'average daily', '"Average Daily"')
    i_fx = idx(lambda s: s.startswith('exchange rate'), '"Exchange Rate"')
    i_share = idx(lambda s: s == 'share of trading', '"Share of trading"')
    i_days = idx(lambda s: s == 'trading days', '"Trading days"')
    if not head < i_avg < i_fx <= i_share < i_days:
        raise LsegOrderbookFetchError(
            f'{month}: 第 1 页区块顺序反常 '
            f'(head={head} avg={i_avg} fx={i_fx} share={i_share} days={i_days})')

    rec = {}
    # 月合计区块：每行 9 个数 = [笔数, £m, €m] 当月 / [同上] 去年同月 / [同比 ×3]，只取前 3。
    for labels, pfx in ((LSE_LABELS, 'lse_orderbook'),
                        (TQ_INT_LABELS, 'turquoise_integrated'),
                        (TQ_DARK_LABELS, 'turquoise_dark')):
        i = _find_line(lines, labels, head, i_avg, month, f'月合计 {pfx}')
        trades, gbp, _eur = _take(lines, i, 3, month, f'月合计 {pfx}')
        rec[f'{pfx}_trades_count'] = int(round(trades))
        rec[f'{pfx}_value_gbp_m'] = gbp
    # 日均区块
    for labels, pfx in ((LSE_LABELS, 'lse_orderbook'),
                        (TQ_INT_LABELS, 'turquoise_integrated'),
                        (TQ_DARK_LABELS, 'turquoise_dark')):
        i = _find_line(lines, labels, i_avg, i_fx, month, f'日均 {pfx}')
        trades, gbp, _eur = _take(lines, i, 3, month, f'日均 {pfx}')
        rec[f'{pfx}_adv_gbp_m'] = gbp
        rec[f'{pfx}_avg_daily_trades_count'] = int(round(trades))

    # 汇率：Exchange Rate (GBP/EUR) 后面两个数 = 当月 / 去年同月
    rec['gbp_eur_rate'] = _take(lines, i_fx, 2, month, '汇率')[0]

    # 份额：标签后面两个百分数 = 当月 / 去年同月
    i = _find_line(lines, UK_SHARE_LABELS, i_share, i_days, month, '份额 UK Lit')
    rec['lse_lit_uk_share_pct'] = _take(lines, i, 2, month, '份额 UK Lit')[0]
    i = _find_line(lines, TQ_SHARE_LABELS, i_share, i_days, month, '份额 Turquoise')
    rec['turquoise_paneuropean_share_pct'] = _take(lines, i, 2, month, '份额 Turquoise')[0]

    # 交易日：标签后面 4 个数 = MTD / YTD / 去年 MTD / 去年 YTD，只要 MTD。
    # 窗口给到 30 行：2021 年那批还多印一行 CurveGlobal（2022-01 停业前），
    # Turquoise 会被挤到第 14 行往后。窗口再宽也不会误命中 YTD 表 ——
    # 那张表里的标签是 "Turquoise Integrated"，归一化后不等于 "turquoise"。
    i = _find_line(lines, LSE_DAY_LABELS, i_days, i_days + 30, month, '交易日 LSE')
    rec['lse_trading_days_count'] = int(round(_take_days(lines, i, month, '交易日 LSE')))
    i = _find_line(lines, TQ_DAY_LABELS, i_days, i_days + 30, month, '交易日 Turquoise')
    rec['turquoise_trading_days_count'] = int(
        round(_take_days(lines, i, month, '交易日 Turquoise')))

    missing = [c for c in COLUMNS if rec.get(c) is None]
    if missing:
        # 缺列一律失败，绝不静默写空 —— 空值会一路画成 null 点上线而全程无报错。
        raise LsegOrderbookFetchError(f'{month} 解析缺列 {missing}')
    _sanity(month, rec)
    return rec


#: 「官方自己不自洽」的白名单：(月份, 列前缀) → 一句为什么。
#:
#: ⚠️ 这里放行的是**恒等式**，不是数据 —— 入库的仍然是官方原值，一个字节不改
#: （本仓的规矩是「入库值必须是当期官方公告原值」）。放行只是承认「这一期官方印的
#: 日均、月合计、交易日三个数彼此对不上」，而那是官方的事实，不是我们解析错了。
#:
#: **不要改成放松 `_sanity` 的通用容差。** 那条 1.5% 的容差是用来抓「把某一行的数
#: 接到了别的行」的 —— 接错行造成的偏差动辄几十个百分点，而 2016-03 这一期偏差 9.5%，
#: 把容差抬到能放行它，就等于把这条护栏关掉。逐期白名单 + 写下实测数字，
#: 才能让下一个人一眼看出「这是已知的、被核过的一期」而不是「护栏松了」。
_OFFICIAL_INCONSISTENT = {
    ('2016-03', 'turquoise_integrated'):
        '官方 2016-03 那期第 1 页：Turquoise Integrated 月合计 £92,288m、日均 £4,013m，'
        '两者相除 = 23.0 天，而同一页「Trading days / Turquoise」印的是 21 天。'
        'Turquoise MidPoint 同期 11,013 ÷ 479 = 23.0，也是 23 —— 即**两条 Turquoise 腿的'
        '日均分母都用了 23**，只有印出来的天数是 21。同页 UK 自洽（102,097 ÷ 4,862 = 21.0），'
        '所以不是整页错位，是 Turquoise 那两行的分母与打印的天数不是同一个数。',
    ('2016-03', 'turquoise_dark'):
        '同上（Turquoise MidPoint，即今天的 turquoise_dark）：11,013 ÷ 479 = 23.0 天 vs 印的 21 天。',
}


def _sanity(month, rec):
    """结构性自检：解析没报错但把数字接错行时，只有恒等式能发现。"""
    for pfx in ('lse_orderbook', 'turquoise_integrated', 'turquoise_dark'):
        if (month, pfx) in _OFFICIAL_INCONSISTENT:
            print(f'[lseg_orderbook] {month} {pfx}: 跳过日均×交易日恒等式 —— '
                  f'{_OFFICIAL_INCONSISTENT[(month, pfx)][:60]}…')
            continue
        days = rec['lse_trading_days_count' if pfx == 'lse_orderbook'
                   else 'turquoise_trading_days_count']
        tot, adv = rec[f'{pfx}_value_gbp_m'], rec[f'{pfx}_adv_gbp_m']
        if adv <= 0 or tot <= 0 or days <= 0:
            raise LsegOrderbookFetchError(f'{month} {pfx}: 非正数 tot={tot} adv={adv} d={days}')
        # 官方自己四舍五入到 £m，日均×天数与月合计允许 1.5% 的漂移。
        if abs(adv * days - tot) > max(1.5, 0.015 * tot):
            raise LsegOrderbookFetchError(
                f'{month} {pfx}: 日均×交易日({adv}×{days}={adv * days:.0f}) '
                f'与月合计({tot}) 对不上 —— 多半是把某一行的数接到了别的行')
    for col in ('lse_lit_uk_share_pct', 'turquoise_paneuropean_share_pct'):
        if not 0 < rec[col] <= 100:
            raise LsegOrderbookFetchError(f'{month} {col}={rec[col]} 不像百分点')
    if not 0.9 < rec['gbp_eur_rate'] < 1.5:
        raise LsegOrderbookFetchError(f'{month} gbp_eur_rate={rec["gbp_eur_rate"]} 不像汇率')


# ─────────────────────────────────────────────── 副源 xlsx（只做核对，不入库）

def _xlsx_monthly():
    """返回 {'YYYY-MM': (trades, value_gbp, days, adv_trades, adv_gbp)}，只含**完整月**。

    完整月的判据不看日历、不查假期表：Daily sheet 里出现过比该月更晚的交易日，
    该月就一定走完了。见 docstring 口径坑 6。

    ⚠ 这份工作簿必须**带过期时间**下载（`_XLSX_MAX_AGE_HOURS`），不能吃永久缓存：
    它是本模块唯一一个独立于 PDF 解析器的判据，冻在半个月前等于这道判据对最新月
    永远失效，而失效时它只会打印一句「副源尚无这些月份」，不会报错。
    """
    import openpyxl
    hit = _find_doc('Order book trading')
    if hit is None:
        raise LsegOrderbookFetchError('检索接口找不到 "Order book trading" 工作簿')
    path = _download(hit['url'], os.path.join(CACHE, 'order_book_trading.xlsx'),
                     min_bytes=100000, max_age_hours=_XLSX_MAX_AGE_HOURS)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for need in ('Daily Order Book Trading', 'Monthly Order Book Trading'):
        if need not in wb.sheetnames:
            raise LsegOrderbookFetchError(f'工作簿少 sheet {need!r}: {wb.sheetnames}')

    newest_day = None
    for row in wb['Daily Order Book Trading'].iter_rows(values_only=True):
        v = row[1] if len(row) > 1 else None
        if isinstance(v, datetime.datetime):
            if newest_day is None or v > newest_day:
                newest_day = v
    if newest_day is None:
        raise LsegOrderbookFetchError('Daily sheet 里一个交易日期都没解析出来')
    cutoff = f'{newest_day.year}-{newest_day.month:02d}'   # 该月本身算未完成

    out = {}
    for row in wb['Monthly Order Book Trading'].iter_rows(values_only=True):
        lab = row[1] if len(row) > 1 else None
        if not isinstance(lab, str):
            continue
        m = re.match(r'^([A-Za-z]{3})-(\d{2})$', lab.strip())
        if not m:
            continue
        mon = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
               'jul', 'aug', 'sep', 'oct', 'nov', 'dec'].index(m.group(1).lower()) + 1
        # 两位年份：这份表从 Oct-97 起头，'97'/'98'/'99' 必须还原成 19xx。
        # 直接拼 '20'+yy 会得到 2097-10 —— 不报错，只是悄悄多出一堆未来月份。
        yy = int(m.group(2))
        key = f'{1900 + yy if yy >= 70 else 2000 + yy}-{mon:02d}'
        if key >= cutoff:
            continue
        vals = [row[i] if len(row) > i else None for i in (2, 3, 4, 5, 6)]
        if any(v is None for v in vals):
            continue
        out[key] = tuple(float(v) for v in vals)
    wb.close()
    if not out:
        raise LsegOrderbookFetchError('工作簿 Monthly sheet 解析不出任何完整月')
    return out


#: 双源比对的**相对**容差。绝对容差（下面 pairs 里那几个数）是 2021-01 起 66 个月
#: 实测定的，窗口前推到 2016-01 之后它们不够用了 —— 但不是因为解析变差，而是因为
#: **副源 xlsx 是持续重述的、月报 PDF 是发布当时的快照**，年头越久累积的重述越多：
#:     2021-2026 段：笔数最大差 8 笔
#:     2016-2020 段：笔数最大差 313 笔（2019-01），成交额最大差 **5.41 £m（2016-02）**
#:       —— 2026-08-19 离线重算修正：此处原写「4.8 £m（2018-06）」，2018-06 的 4.81 £m
#:       其实是第二大，最大的是 2016-02 的 5.41 £m。结论与阈值都不受影响。
#: 换成相对口径看，两段其实是同一个量级：**排除 2016-07 那一个真重述月之后**，
#: 126 个月的四项最大相对偏差是 1.37e-4（2016-01 的 adv_gbp_m —— 它是 PDF 四舍五入到
#: 整数 £m 的小数，相对误差天然偏大，本来就由绝对容差 1.0 £m 兜着）。
#: 取 5e-4 = 那个最坏值的 3.6 倍余量，同时比 2016-07 的 5.6e-3 紧 11 倍 ——
#: 也就是说这个阈值**刚好把「重述漂移」与「真重述」分开**，而接错行是百分级偏差，
#: 照样一撞就响。
_XCHECK_REL = 5e-4

#: 官方**真重述**过的月份：月报 PDF（当期快照）与今天的 xlsx 差得远超重述漂移。
#: 入库的仍是 PDF 原值 —— 本仓的规矩是「入库值必须是当期官方公告原值」，
#: 而且 PDF 自己是自洽的（113,143 ÷ 21 = 5,387.8，与它印的日均 5,388 相符）。
_RESTATED = {
    '2016-07':
        'LSEG 事后把 2016-07 下修了：月报 PDF（当期）笔数 21,870,812 / 成交额 £113,143m，'
        '今天的 Order book trading 工作簿是 21,846,419 / £112,522.4m —— '
        '差 24,393 笔（1.1e-3）与 £620.6m（5.5e-3），比相邻月份（2016-06 差 £3.6m、'
        '2016-08 差 £4.3m）大两个数量级，而且方向相反（这里 PDF 高于 xlsx，'
        '其余月份一律 PDF 低于 xlsx）。两边各自都自洽，交易日同为 21 天，'
        '所以不是解析错位，是官方对这一个月做过一次实质修订。',
}


def _crosscheck(rows, xlsx):
    """逐月拿 xlsx 精确值核对 PDF 的 LSE 三个数。xlsx 没有的月份跳过并说明。"""
    skipped = []
    for r in rows:
        ref = xlsx.get(r['month'])
        if ref is None:
            skipped.append(r['month'])
            continue
        if r['month'] in _RESTATED:
            print(f'[lseg_orderbook] {r["month"]} 跳过双源比对 —— {_RESTATED[r["month"]][:70]}…')
            continue
        trades, value_gbp, days, adv_trades, adv_gbp = ref
        # 下面这几个**绝对**容差是 2021-01..2026-06 那 66 个月实测定出来的（不是拍的）：
        # 那一段观测到的最大偏差是笔数 8 笔（2026-03，21,124,329 vs 21,124,326）、
        # 成交额 1.30 £m、日均笔数 0.5、日均成交额 0.49 £m、交易日 0，阈值留了 2-3 倍余量。
        # ⚠ 窗口 2026-08-18 前推到 2016-01 之后，**绝对容差单独已经不够用**
        #   （2019-01 笔数差 313 > 25）—— 兜住老月份的是下面那行 `max(tol, _XCHECK_REL*|want|)`
        #   里的相对项。别看到「313 > 25」就去调大这里的 25：那会同时放松近端月份的判据。
        # 接错行会差好几个数量级，绝对与相对哪一条都拦得住。
        pairs = [
            ('trades', r['lse_orderbook_trades_count'], trades, 25),
            ('value_gbp_m', r['lse_orderbook_value_gbp_m'], value_gbp / 1e6, 2.0),
            ('trading_days', r['lse_trading_days_count'], days, 0.5),
            ('avg_daily_trades', r['lse_orderbook_avg_daily_trades_count'], adv_trades, 1.5),
            ('adv_gbp_m', r['lse_orderbook_adv_gbp_m'], adv_gbp / 1e6, 1.0),
        ]
        for name, got, want, tol in pairs:
            # 绝对容差与相对容差取宽者：绝对那个管小数（adv 只印到整数 £m、交易日是整数），
            # 相对那个管大数（笔数上千万，重述漂移按绝对值看会越走越大）。见 _XCHECK_REL。
            lim = max(tol, _XCHECK_REL * abs(want))
            if abs(got - want) > lim:
                raise LsegOrderbookFetchError(
                    f'{r["month"]} 双源对不上 {name}: 月报 PDF={got} vs '
                    f'Order book trading.xlsx={want:.4f}'
                    f'（差 {abs(got - want):.4f} = {abs(got - want) / abs(want):.2e}，'
                    f'容差 {lim:.4f}）')
    if skipped:
        print(f'[lseg_orderbook] 副源 xlsx 尚无这些月份，未做双源核对: {skipped}')
    return len(rows) - len(skipped)


# ─────────────────────────────────────────────────── 逾期哨兵（护栏②③）

#: 逾期阈值：某个月的月报离**数据月月末**超过这么多天还没在检索接口里出现，就不再
#: 当「官方还没发」，而当「我们找不到」抛出来。
#:
#: 取值只能从本模块自己量过的节奏来（文件头「实测发布节奏」），不许拍脑袋：
#:     2021-01 起 67 期：中位 +4 天，**最晚 +51 天**（2026-01 那期。它是重传件，
#:                      但同日还发了 2026-02 那期，说明 2026 年初确实积压了两个月
#:                      一起补 —— 所以这个 51 当上界读是站得住的）
#:     2024-01 起 30 期：**最晚 +24 天**（2026-03 的数据 → 2026-04-24）
#: 90 天 ≈ 全样本最坏值的 1.8 倍、近两年最坏值的 3.75 倍，还跨满一个季度。
#: 也就是说它只可能在「官方真的停发」或「我们真的找不到」时响，不会在一次正常的
#: 晚发上响 —— 这条护栏宁可迟三个月才响，也不能每月假响一次：
#: 每月假一次的警报，人很快就学会无视，然后连真的那次也一起无视了。
#:
#: ⚠ **它哪天响了，不要往上调它。** 官方永久停发这份月报，要照 monthly_run 里 JPX
#: cmdty_proforma 那套处置（写下停发结论 + 把该月移出预期窗口），调阈值只是把哨兵
#: 关掉，而关掉之后这一路又回到「安静空转」那个状态。
_MAX_PUBLISH_LAG_DAYS = 90


def _guard_overdue_missing(missing, use_judge=True, today=None):
    """`missing` 里有月份逾期未见 → 抛。这一路唯一能把「还没发」和「找不到」分开的判据。

    为什么必须**抛**而不是 print 一句 warn：稳态下 `skip` 覆盖全部已入库月份，检索
    接口一旦按标题找不到新月份，`fetch_rows` 的 rows 就是空的，函数走「全部已入库」
    那一支干净返回，`fetch/lseg.py` 那边 after−before 为空、只打一行 `orderbook ok`。
    warn 治不了这个 —— 打完 warn 状态仍然是「正常」，连续失败计数、红点、断档哨兵
    照样一个都不动。要让「坏了十天」和「稳了十天」在日志里长得不一样，只有抛。

    判据顺序刻意是「先看时间，再问副源」，不能反过来：副源在次月头几天就把上个月
    记成完整月，而月报 PDF 的 2026 年中位滞后是 +19 天。拿副源当触发器 = 每个月
    误报三个星期。所以副源在这里只**作证**（把错误消息分成两种），不参与「要不要抛」。
    """
    if not missing:
        return
    today = today or datetime.date.today()
    overdue = [(m, (today - _month_end(m)).days) for m in sorted(missing)]
    overdue = [(m, d) for m, d in overdue if d > _MAX_PUBLISH_LAG_DAYS]
    if not overdue:
        return                                  # 还在正常晚发的窗口里，闭嘴

    # 外部判据（形状同 fetch/cboe.py::_crosscheck_report_month、
    # fetch/ice.py::_crosscheck_workbook_month）：问一个**不经过 PDF 解析器**的源。
    judge = '（本轮没问副源：crosscheck 关着）'
    if use_judge:
        try:
            done = _xlsx_monthly()
        except Exception as e:                  # noqa: BLE001 —— 作证失败不许盖住主错
            # 副源自己取不到时仍然要抛上面那件事，只是消息里说清楚证人没到场，
            # 免得下一个人以为「没提副源 = 副源说没事」。
            judge = f'（副源工作簿本轮也取不到：{type(e).__name__}: {e}）'
        else:
            ready = [m for m, _d in overdue if m in done]
            if ready:
                judge = (f'而副源 `Order book trading` 工作簿已经把 {ready} 记成完整月 —— '
                         f'数据本身早就出来了，是月报 PDF 这一路找不到')
            else:
                judge = ('副源 `Order book trading` 工作簿也还没把这些月记成完整月 —— '
                         '更像官方整体停发，而不是标题写法变了')

    raise LsegOrderbookFetchError(
        '检索接口里这些月份的月报逾期未见（阈值 {} 天，本模块实测最慢一期 +51 天）：'
        '{}。{}。按标题找不到时取链接那一步是返回 None 不抛的，所以不抛在这里，'
        '这一路就会一直安静空转。先手工搜一次 "LSEG market report <Month> <Year>" '
        '看标题写法变没变（变了就把新写法补进 `_find_month_report` 的宽松判据）；'
        '若确认是官方永久停发，照 JPX cmdty_proforma 那套写下结论并收窄窗口，'
        '**不要调 _MAX_PUBLISH_LAG_DAYS**。'.format(
            _MAX_PUBLISH_LAG_DAYS,
            ', '.join(f'{m}（月末后 {d} 天）' for m, d in overdue),
            judge))


# ──────────────────────────────────────────────────────────────── 对外接口

def fetch_rows(start=START_MONTH, end=None, verbose=True, crosscheck=True, skip=()):
    """返回 [{'month': 'YYYY-MM', <17 列>}, ...]，按月份升序。

    end 默认取「上个月」—— 当月的月报当然还没出。逐月按标题去检索接口要链接
    （`_find_month_report`：先全等、后宽松），检索不到的记进 `missing`。

    `missing` 里的月份**只在还没超期时**算「这一期还没发」：超过
    `_MAX_PUBLISH_LAG_DAYS` 的由 `_guard_overdue_missing` 抛出来。为什么不能一律
    当「还没发」静静跳过，见文件头「护栏」那一节 —— 那正是本模块踩过的第四类失败。

    `skip` = 已经入库、不必再抓的月份集合。**窗口从 2021-01 前推到 2016-01 之后
    这个参数是必需的，不是优化**：本函数对窗口里的每个月都发一次检索请求 + 下一份
    PDF + 解析一遍（还各 sleep 0.25s），66 个月时每轮约 1 分钟，127 个月时要三倍多，
    而 `write_csv()` 转头就把已有月份全部丢弃（「只填空不覆盖」）—— 那些请求从头到尾
    没有任何产出。
    ⚠️ 跳过**不会**漏掉官方重述：本模块本来就不做重述比对（`write_csv` 只填空、
    从不与已有值比较），所以重抓旧月份在今天这套代码里一个字节的作用都没有。
    哪天要加重述体检，那是另一件事，得显式地做（照 fetch/mtk.py 的 drift 那套写），
    不能靠「碰巧每轮都重抓一遍」来实现 —— 那种依赖没人看得出来，删掉也不会报错。
    """
    end = end or _prev_month(_today_month())
    skip = set(skip)
    rows, missing = [], []
    for month in _months(start, end):
        if month in skip:
            continue
        hit = _find_month_report(month)
        if hit is None:
            missing.append(month)
            continue
        fn = os.path.join(CACHE, f'mmr_{month}.pdf')
        _download(hit['url'], fn)
        try:
            rec = parse_report(fn, month)
        except TypeError:
            raise LsegOrderbookFetchError(
                f'{month}: {os.path.basename(hit["url"])} 第 1 页不是 '
                f'"LSEG - Electronic Order Book Trading" 表 —— 官方换排版了')
        rec['month'] = month
        rec['_src_url'] = hit['url']
        rows.append(rec)
        if verbose:
            print(f'[lseg_orderbook] {month} ok  LSE £{rec["lse_orderbook_value_gbp_m"]:,.0f}m '
                  f'/ {rec["lse_orderbook_trades_count"]:,} trades')
        time.sleep(0.25)                            # 对官方站客气一点

    # missing 无条件先打印、哨兵无条件先跑。**顺序是护栏的一部分**：这两句原先排在
    # 下面 `if not rows: return []` 的后面，而稳态下 rows 恰恰总是空的（skip 覆盖
    # 全部已入库月份）—— 于是「某个月检索不到月报」这件事一次都没被打印过，函数还
    # 顺口报了一句「全部已入库」（在有 missing 的日子里，那句话是假的）。
    if missing:
        print(f'[lseg_orderbook] 检索接口没有这些月份的月报: {missing}')
    _guard_overdue_missing(missing, use_judge=crosscheck)

    if not rows and not skip:
        raise LsegOrderbookFetchError(f'{start}..{end} 一期月报都没抓到')
    if not rows:
        # 全窗口都已入库 = 稳态下最常见的一轮，不是故障。
        # （加 skip 之前这里是无条件抛 —— 直接搬过来会让「本月还没发」变成整家 FAIL。）
        # ⚠ 只有 missing 为空时「全部已入库」才是真话；有 missing 时如实说，
        #   别再让一句假的正常话盖住上面那行。
        if missing:
            print(f'[lseg_orderbook] {start}..{end} 本轮无新月份入库；'
                  f'其中 {len(missing)} 期检索不到，但都还在 '
                  f'{_MAX_PUBLISH_LAG_DAYS} 天的正常晚发窗口内')
        else:
            print(f'[lseg_orderbook] {start}..{end} 全部已入库，无新月份')
        return []
    if crosscheck:
        n = _crosscheck(rows, _xlsx_monthly())
        print(f'[lseg_orderbook] 双源核对通过 {n}/{len(rows)} 个月')
    rows.sort(key=lambda r: r['month'])
    return rows


def _fmt(v):
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def write_csv(rows, series_dir=SERIES):
    """写 series/lseg_part_orderbook.csv（首列 month，升序）。只填空不覆盖。"""
    path = os.path.join(series_dir, 'lseg_part_orderbook.csv')
    have = {}
    if os.path.exists(path):
        with open(path, newline='', encoding='utf-8') as f:
            rd = list(csv.reader(f))
        if rd:
            head = rd[0]
            if head != ['month'] + COLUMNS:
                raise LsegOrderbookFetchError(
                    f'已有 CSV 的列与本模块不一致，拒绝覆盖:\n  旧 {head}\n  新 {["month"] + COLUMNS}')
            have = {r[0]: r for r in rd[1:] if r}
    for r in rows:
        if r['month'] in have:
            continue                                # 幂等：已有月份不重写
        have[r['month']] = [r['month']] + [_fmt(r[c]) for c in COLUMNS]
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(['month'] + COLUMNS)
        for k in sorted(have):
            w.writerow(have[k])
    return path


def cadence(rows=None):
    """实测发布节奏：读每期 PDF 内嵌 CreationDate，算相对数据月月末的日历天数。"""
    import fitz
    out = []
    for fn in sorted(os.listdir(CACHE)) if os.path.isdir(CACHE) else []:
        m = re.match(r'^mmr_(\d{4})-(\d{2})\.pdf$', fn)
        if not m:
            continue
        month = f'{m.group(1)}-{m.group(2)}'
        raw = (fitz.open(os.path.join(CACHE, fn)).metadata or {}).get('creationDate') or ''
        d = re.match(r"D:(\d{4})(\d{2})(\d{2})", raw)
        if not d:
            continue
        made = datetime.date(int(d.group(1)), int(d.group(2)), int(d.group(3)))
        y, mo = int(m.group(1)), int(m.group(2))
        eom = (datetime.date(y + (mo == 12), (mo % 12) + 1, 1) - datetime.timedelta(days=1))
        out.append((month, made.isoformat(), (made - eom).days))
    return out


def main():
    import sys
    if '--cadence' in sys.argv:
        for month, made, lag in cadence():
            print(f'{month}\t{made}\t+{lag}d')
        return
    rows = fetch_rows()
    path = write_csv(rows)
    print(f'[lseg_orderbook] {len(rows)} 个月 '
          f'{rows[0]["month"]}..{rows[-1]["month"]} → {path}')


if __name__ == '__main__':
    main()
