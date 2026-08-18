# -*- coding: utf-8 -*-
"""HKEX 香港交易所 00388 —— 月度市场统计抓取模块（无人值守）。

═══ 数据源 ═══
本模块有**两条互不相干的官方链路**，各管一批列，谁挂了另一条照跑：

  A. hkexgroup.com 的「Monthly HK Market Highlight Data」xlsx
     → adt_hkdbn / mktcap_hkdtn / new_listings / ipo_funds_hkdbn /
       derivatives_adv_contracts / southbound_adt_hkdbn（下面「数据源」整节讲的都是它）

  B. hkex.com.hk 的 Monthly Bulletin 报表 JSON（逐日档案 + 每月 Stock market highlights）
     → trading_days_cash / vol_shares_* / trades_* / adv_shares_mn / adt_trades
       （见文件后半「═══ 成交股数与成交笔数 ═══」那一节）

链路 A 的真值文件：HKEX 官方「Monthly HK Market Highlight Data」Excel
    落地页  https://www.hkexgroup.com/Investor-Relations/Business-Analysis/Key-Market-Data?sc_lang=en
    直链模板 https://www.hkexgroup.com/-/media/HKEX-Group-Site/Ir/Monthly-market-highlights/
             HKEX-HK-market-monthly-highlight-statistics-(Updated-to-{Mon}-{YYYY}).xlsx
             例：...(Updated-to-Jun-2026).xlsx

为什么不写死直链、而是先去落地页抓 href：
    文件名里带月份，写死等于每月都要改代码；而 HKEX 偶尔会跳月或改大小写。
    落地页上那一条「Monthly HK Market Highlight Data Download (Updated to X)」链接
    是官方自己维护的「当前最新」指针，跟着它走才是真正无人值守。
    落地页挂了才退回按月份倒推试直链（_probe_direct），两条腿都断才抛异常。

反爬情况（2026-08-05 实测）：
    hkexgroup.com 与其 /-/media/ 静态资源都没有 Cloudflare / Akamai 交互式挑战。
    python urllib + 普通桌面 UA 即可 200，curl 亦可。不需要浏览器、不需要登录态。
    唯一坑：路径大小写敏感，且带 `?sc_lang=en` 的页面路径写错会返回 404 而不是 302。

═══ 发布节奏 ═══
    次月上旬发布上一个月的数据。实测 Last-Modified：
        2025-10 → 11/11、2025-11 → 12/08、2025-12 → 01/08、2026-01 → 02/09、
        2026-02 → 03/17（异常晚，疑似重传）、2026-03 → 04/09、
        2026-04 → 05/08、2026-05 → 06/04、2026-06 → 07/07
    ⇒ 定时任务排在**次月 10 日之后**跑最稳；10 号之前跑很可能扑空，
      这不是故障，latest_month() 会如实返回上上个月。

═══ 发布日（页面抬头「官方发布于」那半句）═══
    HKEX **哪儿都不写发布日**：工作簿里没有 "Updated on …" 之类的字符串（整表扫过），
    落地页那条链接只写 "(Updated to Jun 2026)" —— 那是数据月，不是发布日；
    也没有配套的新闻稿。所以本家只能用上面那张表的口径：**xlsx 直链的 HTTP
    Last-Modified**，它本来就是本模块记录发布节奏的实测依据。
    另有一条独立佐证写在文件肚子里：工作簿 docProps/core.xml 的 dcterms:modified
    （HKEX 自己的 Excel 存盘时刻，Company = "Hong Kong Exchanges and Clearing"）。
    2026-06 档两者相差 11 秒（存盘 07-07 10:24:36Z、上线 07-07 10:24:47 GMT），
    说明存盘即上传，用 Last-Modified 当发布日是站得住的。
    ⇒ 只在**首次摄入某个月**时记一笔（见 update()），事后不覆盖：HKEX 会原地重传，
      重传会把 Last-Modified 推后，拿它盖掉当初那次真发布的日期就是把事实改错。

═══ 口径坑（都是真踩过的，不是想象的）═══
1. 表是**横排**的：第 1 列是指标名，第 2 列起每列一个月（2018-01 起，逐月累加，不滚动）。
   行号会因为 HKEX 新上产品而整体下移（例如 2025-11 新增 Hang Seng Biotech Index Futures），
   所以本模块一律**按指标名定位行**，绝不按行号硬编码。

2. `Equities - IPO Total*` 这个标签在文件里**出现两次**（Main Board 段 + GEM 段），
   必须靠上方的段落标题 `FUND RAISED AMOUNT BY TYPES (MAIN BOARD)/(GEM)` 区分。
   series 里的 ipo_funds_hkdbn = 两段之和（GEM 近年恒为 0，但别省）。

3. **IPO 募资额是"暂定数"，会被下个月的文件上修**（表里原文脚注：
   "* Provisional figures for latest month"）。实测幅度：
       2025-11  40,455.8 →(Dec 档) 41,832.2   +3.4%
       2026-01  39,179.8 →(Feb 档) 42,298     +8.0%
       2026-05  13,357.3 →(Jun 档) 14,431.4   +8.0%
   ⇒ 本月刚发布的 ipo_funds_hkdbn 天生偏低，不要拿它做同比结论。
   ⇒ 本模块**默认不覆盖 series 里已有的非空数值**（见 update(allow_restate)），
     免得历史被官方重述悄悄改掉；差异会写进 cache/hkex_restatements.csv 供人工判断。

4. series 里有三段**故意的留白**，别手贱回填：
       new_listings   2019-01~2024-05 空（当年逐月简报没这项）
       ipo_funds      2019-01~2023-12 空（同上）
       southbound     2022-01~2025-06 空（官方停发了 42 个月，
                      build_hkex.py 的 multi_line 图注专门讲了这个 gap）
   今天这份 xlsx 其实把 2018-01 起的这些历史都补全了，但一次性回填 =
   把两代口径混进同一条序列、并改掉看板叙事，属于人工决策。
   ⇒ update() 只对「adt 为空的行」补空 + 追加新月，adt 已有的历史行一个字节都不碰。
   ⇒ 2026-08-18 新增的历史回补（START_MONTH，见文末「历史深度」）**不破这条**：
     回补行里 new_listings / ipo_funds_hkdbn 一律留空（BACKFILL_HOLD），
     否则会在这两列上造出中间空洞、让 build/hkex.py 的完整性体检硬失败。

5. mktcap 是月末时点数（$Bil.→ /1000 得 HK$tn）；ADT / 南向 ADT / 衍生品 ADV 是月内日均。
   南向 ADT 官方口径含买卖双边（表内注 "ADT for Stock Connect includes buy and sell trades"）。

6. series/hkex.csv 是 **CRLF** 换行、无引号、7 列定宽小数
   （adt 3 位 / mktcap 4 位 / new_listings 整数 / ipo 3 位 / deriv 整数 / 南向 3 位）。
   本模块逐行改写文本、不经 pandas.to_csv，就是为了保证没动过的行**逐字节不变**。

7. 已知遗留：series 里 2026-07 那一行的 ipo/deriv/southbound 有值而 adt/mktcap/new_listings 为空，
   且 2025-12、2026-06 的 ipo 高于官方 xlsx 任何一版。说明前人对最新月用过一个
   比 xlsx 更快的**未知来源**。本模块不假装知道那是什么，也不去动它，
   只在 7 月档 xlsx 出来后把空格补上，并把冲突记进 cache/hkex_restatements.csv。


═══ 成交股数与成交笔数（链路 B）═══
上面那份 Monthly Market Highlights xlsx 的现货段**只有三行**（ADT、市值、新上市），
成交股数与成交笔数根本不在里面。它们在 HKEX 的**另一份刊物 Monthly Bulletin**（月报），
栏目原文就叫 `Turnover volume (mil shares)` 与 `No. of deals`，
两行都同时给「当月合计」和「- Daily average」。主板与 GEM 分册发布。

本模块用的两个端点（都在 https://www.hkex.com.hk/eng/stat/smstat/mthbull/ 下）：

  B1【历史主力】Securities Statistics Archive —— Trading value, volume and number of deals
     索引 rpt_data_statistics_archive_trading_data.json      （主板）
          rpt_data_statistics_archive_trading_data_gem.json  （GEM）
     索引里每条指向一册 5 年期的**逐日**表，列是
         Year/Month/Day │ │ Total trading value (HKD) │ Total trading volume (Shares) │ Number of deals
     主板最老一册标签 1986-1989、GEM 1999-2003 —— 也就是逐日档案能回到 1980 年代。
     逐日 → 逐月只做**求和与计数**，不做任何口径改写。

  B2【最新月 + 交叉核对】Monthly Bulletin —— Stock market highlights
     rpt_Stock_market_highlights_{YYMM}.json       （主板）
     rpt_Stock_market_highlights_GEM_{YYMM}.json   （GEM）
     只挂**最近 13 个月**（更早的月份 404，实测 2412/2312/2212/2505 全 404），
     所以它当不了历史来源，但它有两个不可替代的用处：
       ① 比逐日档案快。实测 2026-06：Bulletin 07-02 14:02 GMT 就上线了，
          逐日档案 07-02 14:59 GMT 同日跟上；而 2026-07 的 Bulletin 08-03 14:04 GMT
          已经上线，逐日档案到 08-07 仍停在 06-30（带 no-cache 复核过，不是 CDN 缓存）。
          ⇒ 逐日档案对最新月**会落后整整一个月**，最新月必须靠 B2。
       ② 给 B1 当闸门。逐日档案的最后一个月有可能是「月中快照」（合计天生偏小），
          从档案本身看不出来。所以本模块规定：**档案的最后一个月必须被同月的
          Bulletin 逐位确认才收**，确认不了就整月不要（见 _trading_stats）。

  B3【历史回补专用】Securities Statistics Archive —— Market capitalisation
     索引 rpt_data_statistics_archive_market_cap.json      （主板，最老一册 1986-1994）
          rpt_data_statistics_archive_market_cap_gem.json  （GEM，最老一册 1999-2003）
     两列的**逐日**表：Year/Month/Day │ Total market capitalisation (HKD)。
     2026-08-18 新接。它只有一个用途：链路 A 的 xlsx 最早只到 2018-01，
     而 series 要铺到 2016-01，那 24 个月的 mktcap_hkdtn 只能从这里取
     （取每月**最后一个交易日**主板 + GEM 之和 ÷ 1e12，见下面「历史深度」的闭合④）。
     xlsx 覆盖得到的月份一律以 xlsx 为准，本端点不参与。

  为什么不用 hkexgroup.com 那份 xlsx 顺手带出来：它压根没有这两行，试过了。

── 三重闭合（2026-08-07 本机实测，逐位相同，不是四舍五入到差不多）──
  ① 逐日档案汇总（主板，2026-06，21 个交易日）
        成交额 6,698,704,112,075 HKD → 6,698,704 HK$mil，日均 318,986 HK$mil
        成交股数 7,370,724,218,440 股 → 7,370,724 mil sh，日均 350,987 mil sh
        成交笔数 100,533,281 笔，日均 4,787,299 笔
     Monthly Bulletin 2606 主板同月原文：6,698,704 / 318,986、7,370,724 / 350,987、
        100,533,281 / 4,787,299 —— **六个数逐位相同**。
  ② 主板日均成交额 318.986 + GEM 日均成交额 0.090 = 319.076 HK$bn
     = series/hkex.csv 2026-06 的 adt_hkdbn **319.076**，逐位相同。
  ③ 把 ② 推广到全序列：series 里 90 个有 adt_hkdbn 的月（2019-01~2026-06），
     用 B1 逐日档案（主板+GEM）重算日均成交额，**90/90 全部在 3 位小数上逐位相同**。
     ⇒ 新列与既有 adt_hkdbn 是**同一套逐日底稿**，口径天然一致，不是"看着差不多"。

── 口径 ──
  · 「主板+GEM」：合计口径与 adt_hkdbn 完全对齐（adt_hkdbn 就是两板之和，见闭合 ②）。
  · 交易日：以主板逐日档案的行数为准。GEM 与主板同一个交易日历，2019-01~2026-06
    实测 90 个月两边行数**从无差异**；万一将来差了会打印警告并仍以主板为准
    （GEM 只占成交额的 0.03%，用主板日历不会改变任何结论）。
  · 半日市：档案在日期后打 `*`（圣诞前夕 / 除夕 / 年三十，以及 2020-08-19、2021-06-28、
    2022-08-25、2023-10-09 这类台风黑雨缩短交易时段的日子）。HKEX 自己的日均**照算一整天**——
    这是闭合 ③ 90/90 反推出来的事实，不是假设：含半日市的月份（如 2025-01、2024-02）
    如果按 0.5 天折算，日均就对不上官方数了。所以本模块也不折算。
  · 单位：档案发的是**股**，Bulletin 发的是**百万股**。本模块统一按 Bulletin 的
    `mil shares` 存（档案值 ÷ 1e6），与既有列把 HK$mil ÷1000 存成 HK$bn 是同一类换算——
    只换单位刻度，不换口径、不引入任何外部参数。
  · 「笔数」= HKEX 原文 `No. of deals`（成交宗数）。列名用 trades 是跟仓里
    asx.csv 的 trades_cash_total / adt_cash_trades、enx.csv 的 adv_cash_trades_k 对齐。

── 发布节奏与滞后 ──
  Bulletin：次月上旬，实测 2026-06 档 07-02、2026-07 档 08-03（Last-Modified, GMT）。
            比链路 A 的 xlsx 还早 ~5 天（2026-06 的 xlsx 是 07-07）。
  逐日档案：正常也是次月上旬同日跟上（2026-06 档 07-02 14:59 GMT），
            但 2026-07 这一轮明显掉队（到 08-07 未更新）。这正是要留 B2 兜底的原因。

── 历史深度：START_MONTH = 2016-01（2026-08-18 打开，此前只补到序列现有首月）──
  以前这里写的是「绝不新增早于序列首月的行」，理由是「往前加行会改变 build 侧的窗口，
  而那些文件不归本模块管」。build/hkex.py 的窗口 2026-08-18 已经改成 WIN_FROM='2016-01'
  的现算式（不再是写死的 `.iloc[-25:]`），那条理由就不成立了 —— 于是把闸门开到 START_MONTH。

  实现：update() 里的 _backfill_rows() 只在**序列首月之前**造行，且只造到 START_MONTH；
  序列**内部**的空档仍然一格不填（那是口径留白，见坑 4）。所以这条回补路径是幂等的：
  序列一旦从 2016-01 起，下个月再跑 _backfill_rows() 直接空转、连档案都不多下一份。

  ── 回补行的取数分工（谁给哪一列）──
      2018-01 起：链路 A 的 xlsx 本来就有，adt / mktcap / deriv / southbound 一律取 xlsx 原值。
      2016-01~2017-12：xlsx 够不到。
          adt_hkdbn    ← B1 逐日档案（主板+GEM 当月成交额 ÷ 主板交易日数 ÷ 1e9）
          mktcap_hkdtn ← B3 市值逐日档案（当月最后一个交易日主板+GEM 之和 ÷ 1e12）
          derivatives_adv_contracts / southbound_adt_hkdbn ← **留空**（见下）
      全部回补行：TRADING_COLUMNS 那 7 列照旧由 B1 逐日档案算，与 2019 年以后同一套代码。

  ── 为什么 2016-01~2017-12 的衍生品与南向留空（不是漏了）──
      衍生品：机器可读源全部很浅 —— MonthlyStatistics_FnO.json 13 个月、
        DerivativesMarketHighlights 12 个月、Monthly Bulletin 13 期、日报存档 2 个月，
        猜名探测的 archive 端点全 404（2026-08-18 实测）。逐月数字只存在于 HKEX Fact Book
        PDF 的 ~50 张分品种月表里，没有月度合计表；试解析在 2018 年 12/12 个月都比官方
        低 6.6%~8.7%，**未闭合**。未闭合就入库 = 在图上放一条系统性偏低、肉眼看不出的线。
      南向：月度 js 只挂 13 个月、日度 js 只挂 7 个月。Fact Book 有逐月表且配方已在 2018 年
        逐位验证过，但那是 PDF 手抄，且 2016-01~2016-11 只有沪股通（深港通 2016-12-05 才开市），
        是**真实的口径断点**，不该混进同一条列里而不作声明。
      ⇒ 两列都从 2018-01 起。**前导空格不是中间空洞**：build/hkex.py 的完整性体检
        （GAPPY_OK 那一段）只拦「首末之间缺月」，前导空格照过；两张图各自从自己的首个
        有值月开始画。

  ── 为什么 new_listings / ipo_funds_hkdbn 在回补行里一律留空（BACKFILL_HOLD）──
      今天这份 xlsx 对 2018-01 起的这两列**全部有值**，技术上填得进去。但 series 里
      2019-01~2024-05（新股）与 2019-01~2023-12（IPO）是坑 4 说的**故意留白**。
      只回填 2018 而不回填那两段，就会造出「2018 有值 → 2019-2024 空 → 2024 起又有值」的
      中间空洞，build/hkex.py 的 GAPPY_OK 体检会直接 SystemExit，整页停更。
      要么两段一起填（= 改看板叙事，人工决策），要么一格不填。本模块选后者：
      回补行的这两列**永远留空**，于是空白段从 2016-01 一路连到 2024-05/2023-12，
      仍然是一段前导空格，体检照过。

  ── 闭合④：2016-2017 的 adt / mktcap 是不是同一个口径（逐位，不是差不多）──
      2026-08-18 本机实测，拿 xlsx 有值的 103 个月（2018-01~2026-07）做对照组：
        · ADT：档案重算 vs xlsx，**102/103 在 3 位小数上逐位相同**；唯一一处 2020-11
          档案 161.287 / xlsx 161.286，差 1 个末位单位（四舍五入噪声，未超
          _materially_differs 的容差）。
        · 市值：档案月末（主板+GEM）vs xlsx，**103/103 在 4 位小数上逐位相同**。
          只用主板会系统性低 0.19~0.26 HK$tn —— 两板必须相加，这是实测出来的。
      光有对照组还不够：对照组证明的是「档案 = xlsx」，而 2016-2017 恰恰**没有 xlsx**。
      所以再拿一份**独立于以上两条链路**的官方出版物逐月核 —— HKEX Fact Book PDF
      （cache/hkex_factbook_{2016,2017,2018}.pdf，直链
      https://www.hkex.com.hk/-/media/HKEX-Market/Market-Data/Statistics/
      Consolidated-Reports/HKEX-Fact-Book/HKEX-Fact-Book-{年}/FB_{年}.pdf）：
        · 「Trading value and volume」（主板，第 24 页）与「Trading value and volume for GEM」
          给逐月的交易日数 / 成交额 / 成交股数 / 成交笔数（含官方自己印的日均）；
        · 「Total market capitalisation by HSICS」与 GEM 对应表的 Total 行给逐月**月末**市值。
      2026-08-18 逐位比对结果：
        · 2016 + 2017 共 24 个月 × 10 项（两板的交易日/额/量/笔数 + 两板月末市值）
          = **240 项全对**；再把它们换算成 series 真正入库的 9 个字段（含 adt_hkdbn 到
          3 位小数、mktcap_hkdtn 到 4 位小数）+ 官方日均 + 4 个应留空的列，
          共 **360 项全对**。
        · 2018 那 12 行（adt/mktcap 来自 xlsx、7 个交易列来自逐日档案）对 FB2018
          **108 项全对** ⇒ xlsx、逐日档案、Fact Book 三方一致。
      顺带证实了两条口径：**半日市照算一整天**（FB 印的日均 = 合计 ÷ 交易日数，
      2016-2017 含多个半日市月，逐位对得上，不折算），以及 mktcap 取**月末最后一个交易日**
      （FB 表下注 "Month-end figures"，与档案取的日期逐月相同）。
      ⇒ 2016-2017 的这两列与 2018 年以后是同一套底稿、同一个口径，不是"看着差不多"。

  ── 还能往前到哪 ──
      主板逐日档案回到 1986-04、GEM 回到 1999-11-25，所以「主板+GEM」同口径的最早完整月是
      1999-12；市值档案主板回到 1986、GEM 回到 1999。START_MONTH 定在 2016-01 是与全站
      其余页面（build/single.py 的 WIN_FROM、cboe / cme / msci 的 WIN0）对齐的产物，
      不是源的上限。真要更早，改 START_MONTH 一个常数即可，但衍生品与南向的前导空格会
      跟着变长 22 年，那时该重新讨论这一页还该不该保留那两张图。
"""

import csv
import datetime as _dt
import email.utils
import io
import json
import os
import re
import urllib.error
import urllib.parse
import urllib.request
import zipfile

# ── 常量 ──────────────────────────────────────────────────────────────────
LANDING_URLS = (
    'https://www.hkexgroup.com/Investor-Relations/Business-Analysis/Key-Market-Data?sc_lang=en',
    'https://www.hkexgroup.com/?sc_lang=en',   # 首页也挂同一条链接，作为备份锚点
)
MEDIA_BASE = ('https://www.hkexgroup.com/-/media/HKEX-Group-Site/Ir/Monthly-market-highlights/'
              'HKEX-HK-market-monthly-highlight-statistics-(Updated-to-%s-%d).xlsx')
_HREF_RE = re.compile(
    rb'HKEX-HK-market-monthly-highlight-statistics-\(Updated-to-([A-Za-z]{3})-(\d{4})\)\.xlsx')

# 普通桌面 UA 就够；不要用 curl 默认 UA，HKEX 的 CDN 对它偶发 403
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36')
_HEADERS = {'User-Agent': _UA, 'Accept-Language': 'en-US,en;q=0.9'}

_MON = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# 发布方在香港，发布日按它自己的日历算。实测最近 10 档的上线时刻落在 09:43–19:32 HKT
# （01:43–11:32 GMT），换不换时区都是同一天 —— 转成港时只是把口径写明确，不是为了改日期。
_HK = _dt.timezone(_dt.timedelta(hours=8))

SHEET = 'Monthly Data'

# series/hkex.csv 的列顺序与小数位。改这里等于改 CSV 格式，别改。
#
# HIGHLIGHT_COLUMNS = 链路 A（hkexgroup.com 的 xlsx）供数的 6 列，历来就有。
# TRADING_COLUMNS   = 链路 B（hkex.com.hk 的 Monthly Bulletin 报表 JSON）供数的 7 列，
#                     2026-08 新增，一律追加在末尾 —— 旧列的位置一格都不动。
HIGHLIGHT_COLUMNS = ['adt_hkdbn', 'mktcap_hkdtn', 'new_listings', 'ipo_funds_hkdbn',
                     'derivatives_adv_contracts', 'southbound_adt_hkdbn']
TRADING_COLUMNS = ['trading_days_cash',        # 现货交易日数（主板日历；半日市算整天）
                   'vol_shares_mb_mn',         # 当月成交股数合计 · 主板（百万股）
                   'vol_shares_gem_mn',        # 当月成交股数合计 · GEM（百万股）
                   'trades_mb_total',          # 当月成交笔数合计 · 主板（No. of deals）
                   'trades_gem_total',         # 当月成交笔数合计 · GEM
                   'adv_shares_mn',            # 日均成交股数 · 主板+GEM（百万股）
                   'adt_trades']               # 日均成交笔数 · 主板+GEM
COLUMNS = HIGHLIGHT_COLUMNS + TRADING_COLUMNS

#: 历史回补的目标起点。update() 只在**序列首月之前**造行、且不早于这个月；
#: 序列内部的空档一格不填（那是口径留白，见坑 4）。改这一个常数就能往前铺得更早，
#: 源的实际上限是 1999-12（主板+GEM 同口径的最早完整月），细节见模块 docstring「历史深度」。
START_MONTH = '2016-01'

#: 回补行里**永远留空**的列。技术上 xlsx 从 2018-01 起就有值，但 series 在这两列上有
#: 2019-01~2024-05 / 2019-01~2023-12 两段故意留白（坑 4）；只填 2018 会造出中间空洞，
#: build/hkex.py 的完整性体检会直接 SystemExit。要填就得连那两段一起填 = 改看板叙事，
#: 属人工决策，不由无人值守任务顺手做掉。
BACKFILL_HOLD = ('new_listings', 'ipo_funds_hkdbn')

# 新 7 列全是整数位：股数已经以「百万股」为单位（末位 = 100 万股，占日均的 3e-6），
# 笔数与交易日本来就是计数。整数位同时让「逐日档案算出来的值」与
# 「Bulletin 印出来的值」落在同一个刻度上，两条链路互为替代时不会因为精度打架。
_DECIMALS = {'adt_hkdbn': 3, 'mktcap_hkdtn': 4, 'new_listings': 0, 'ipo_funds_hkdbn': 3,
             'derivatives_adv_contracts': 0, 'southbound_adt_hkdbn': 3,
             'trading_days_cash': 0, 'vol_shares_mb_mn': 0, 'vol_shares_gem_mn': 0,
             'trades_mb_total': 0, 'trades_gem_total': 0,
             'adv_shares_mn': 0, 'adt_trades': 0}

# 指标名 → (段落限定, 精确/前缀匹配的标签)。段落为 None 表示全表唯一。
# 用 tuple 是因为 IPO 那一行标签重名，必须靠段落标题夹住。
_ROWSPEC = {
    'mktcap_bil':   (None, 'Total market capitalisation ($Bil.)'),
    'newlist_mb':   (None, 'No. of newly listed companies (Main Board)'),
    'newlist_gem':  (None, 'No. of newly listed companies (GEM)'),
    'ipo_mb_mil':   ('FUND RAISED AMOUNT BY TYPES  (MAIN BOARD)', 'Equities - IPO Total'),
    'ipo_gem_mil':  ('FUND RAISED AMOUNT BY TYPES  (GEM)', 'Equities - IPO Total'),
    'adt_mil':      (None, 'Average daily turnover by value'),
    'southbound_mil': (None, 'Total Southbound average daily turnover by value'),
    'deriv_adv':    (None, 'Total Futures and Options'),
}


class HkexFetchError(RuntimeError):
    """抓取或解析失败。调度器看到它就该报警，而不是当成'本月没数据'。"""


# ── 网络 ──────────────────────────────────────────────────────────────────
def _get(url, timeout=90, want_headers=False):
    """want_headers=True 时连响应头一起给出去 —— 发布日只能从 Last-Modified 拿
    （HKEX 的文件与页面都不写发布日，见模块 docstring），而响应头一旦出了这个
    with 块就没了，事后补一次 HEAD 拿到的可能已经是下一次重传的时刻。"""
    req = urllib.request.Request(url, headers=_HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        blob = r.read()
        return (blob, r.headers.get('Last-Modified')) if want_headers else blob


def _discover_url():
    """从官方落地页读出「当前最新」那一份 xlsx 的直链。

    返回 (url, 'YYYY-MM')。落地页给的月份只是文件名上的声明，
    真实最新月以解析结果为准（见 _parse）。
    """
    errs = []
    for page in LANDING_URLS:
        try:
            html = _get(page, timeout=60)
        except Exception as e:                      # noqa: BLE001
            errs.append('%s -> %r' % (page, e))
            continue
        hits = set(_HREF_RE.findall(html))
        if not hits:
            errs.append('%s -> 页面 200 但没匹配到 xlsx 链接（版式可能改了）' % page)
            continue
        # 页面上可能同时挂着旧档，取月份最大的那个
        best = max(hits, key=lambda t: (int(t[1]), _MON.index(t[0].decode().capitalize()) + 1))
        mon, year = best[0].decode().capitalize(), int(best[1])
        return MEDIA_BASE % (mon, year), '%04d-%02d' % (year, _MON.index(mon) + 1)
    raise HkexFetchError('落地页发现失败：\n  ' + '\n  '.join(errs))


def _probe_direct(max_back=6):
    """落地页不可用时的退路：按 (Mon-YYYY) 从上个月往前倒推试直链。

    为什么从"上个月"起：当月数据永远要等次月上旬才发，当月档必然 404。
    """
    today = _dt.date.today()
    y, m = today.year, today.month
    tried = []
    for _ in range(max_back):
        m -= 1
        if m == 0:
            m, y = 12, y - 1
        url = MEDIA_BASE % (_MON[m - 1], y)
        try:
            req = urllib.request.Request(url, headers=_HEADERS, method='HEAD')
            with urllib.request.urlopen(req, timeout=45) as r:
                if r.status == 200:
                    return url, '%04d-%02d' % (y, m)
        except Exception as e:                      # noqa: BLE001
            tried.append('%s-%d %r' % (_MON[m - 1], y, e))
    raise HkexFetchError('直链倒推也失败：' + '; '.join(tried))


def _download(cache_dir):
    """把最新一档 xlsx 落到 cache_dir，返回 (本地路径, 文件名声明的月份, Last-Modified 原文)。

    每次都重新下载：文件才 ~160KB，而 HKEX 会**原地重传**同名文件
    （2026-02 档的 Last-Modified 就晚到了 3/17），缓存命中反而会漏掉重述。
    """
    os.makedirs(cache_dir, exist_ok=True)
    try:
        url, claimed = _discover_url()
    except HkexFetchError as e:
        url, claimed = _probe_direct()
        print('[hkex] 落地页发现失败(%s)，改用直链倒推：%s' % (e, url))
    blob, last_modified = _get(url, want_headers=True)
    if not blob.startswith(b'PK'):                   # xlsx 必须是 zip；返回 HTML 说明被挡或路径错
        raise HkexFetchError('下载到的不是 xlsx（前 16 字节 %r），URL=%s' % (blob[:16], url))
    path = os.path.join(cache_dir, 'hkex_monthly_highlights_%s.xlsx' % claimed)
    with open(path, 'wb') as f:
        f.write(blob)
    return path, claimed, last_modified


# ── 发布日 ────────────────────────────────────────────────────────────────
_SAVED_RE = re.compile(rb'<dcterms:modified[^>]*>'
                       rb'(\d{4}-\d{2}-\d{2})T(\d{2}:\d{2}:\d{2})Z?</dcterms:modified>')


def _doc_saved_at(xlsx_path):
    """工作簿自己记的最后存盘时刻（docProps/core.xml 的 dcterms:modified，UTC）。

    这是 HKEX 那台 Excel 写进文件肚子里的时间戳，不是我们下载的时间，也不受 CDN 影响，
    所以它能给 Last-Modified 当独立佐证。取不到就 None —— 它只是佐证，不该让抓取失败。
    """
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            hit = _SAVED_RE.search(z.read('docProps/core.xml'))
    except (OSError, KeyError, zipfile.BadZipFile):
        return None
    if not hit:
        return None
    return _dt.datetime.strptime(hit.group(1).decode() + ' ' + hit.group(2).decode(),
                                 '%Y-%m-%d %H:%M:%S').replace(tzinfo=_dt.timezone.utc)


def _published_on(xlsx_path, last_modified):
    """这一档 xlsx 的发布日，返回 ("YYYY-MM-DD", 出处描述)；判断不出来就 (None, None)。

    HKEX 不写发布日（工作簿、落地页都只写数据月），所以这里用的是 HTTP Last-Modified ——
    本模块 docstring 的「发布节奏」表一直就是按它记的。出处描述里把原始头和佐证都写全，
    将来有人怀疑某个日期，不用重新做一遍考古。
    """
    saved = _doc_saved_at(xlsx_path)
    online = None
    if last_modified:
        try:
            online = email.utils.parsedate_to_datetime(last_modified)
        except (TypeError, ValueError):
            online = None
    # 头缺失时退回工作簿存盘时刻：存盘早于上线，宁可偏早也不编一个。
    ref, src = (online, 'online') if online else (saved, 'saved')
    if ref is None:
        return None, None
    iso = ref.astimezone(_HK).date().isoformat()

    if src == 'online':
        ev = ('xlsx 直链 HTTP Last-Modified: %s（= %s HKT）'
              % (last_modified, online.astimezone(_HK).strftime('%Y-%m-%d %H:%M')))
        if saved:
            same = saved.astimezone(_HK).date().isoformat() == iso
            ev += ('；工作簿 docProps/core.xml dcterms:modified=%s %s'
                   % (saved.strftime('%Y-%m-%dT%H:%M:%SZ'),
                      '同日，互为印证' if same else '不同日（疑似原地重传），以上线时刻为准'))
    else:
        ev = ('响应无 Last-Modified，退回工作簿 docProps/core.xml dcterms:modified=%s（= %s HKT）'
              % (saved.strftime('%Y-%m-%dT%H:%M:%SZ'), saved.astimezone(_HK).strftime('%Y-%m-%d %H:%M')))
    return iso, ev + '。HKEX 的工作簿与落地页都不写发布日，落地页只写数据月 "Updated to …"'


def _source_dates():
    """按路径加载仓库根的 source_dates.py（发布日台账）。

    不能裸 import：本模块是被 monthly_run 用 spec_from_file_location 加载的，
    那时 sys.path 上既没有 fetch/ 也没有仓库根。
    """
    import importlib.util
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(root, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _record_source_date(series_dir, month, xlsx_path, last_modified):
    """把该月的发布日记进台账。失败只告警，不抛。

    调用点在 series 落盘之后：数据已经进库了，再为「抬头那半句话」把整月抓取判成失败，
    赔本。写不进去的后果是页面少半句，不是数据错。
    """
    iso, ev = _published_on(xlsx_path, last_modified)
    if not iso:
        print('[hkex] %s 拿不到发布日（无 Last-Modified 也无 docProps 时间戳），台账留空' % month)
        return
    try:
        _source_dates().record(series_dir, 'hkex', month, iso, ev)
    except Exception as e:                          # noqa: BLE001
        print('[hkex] 发布日台账没写成（%r），series 数据不受影响' % (e,))


# ── 解析 ──────────────────────────────────────────────────────────────────
def _find_rows(ws):
    """按标签定位行号；缺任何一个必需标签直接抛异常。

    这是最重要的一道闸：HKEX 一旦改版式（改标签、拆表、换 sheet），
    宁可让任务红着停掉，也不能让 update() 写出一串 NaN。
    """
    labels = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        labels.append(v.replace('\xa0', ' ').strip() if isinstance(v, str) else None)

    def section_of(r):
        """向上找最近一个全大写的段落标题。"""
        for i in range(r - 1, 0, -1):
            s = labels[i - 1]
            if s and s.isupper() and len(s) > 8:
                return s
        return None

    out = {}
    missing = []
    for key, (sec, lab) in _ROWSPEC.items():
        hit = None
        for r in range(1, len(labels) + 1):
            s = labels[r - 1]
            if not s or not s.startswith(lab):
                continue
            if sec is not None and section_of(r) != sec:
                continue
            hit = r
            break
        if hit is None:
            missing.append('%s（段落=%s，标签=%s）' % (key, sec, lab))
        out[key] = hit
    if missing:
        raise HkexFetchError('xlsx 版式变了，找不到这些指标行：' + '；'.join(missing))
    return out


def _num(v):
    """'-' / None / 空 一律当缺失；HKEX 用 '-' 表示该产品当月不存在。"""
    if v is None or v == '' or v == '-':
        return None
    if isinstance(v, str):
        v = v.replace(',', '').strip()
        if v in ('', '-', 'N/A'):
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse(path):
    """xlsx → {'YYYY-MM': {列名: 数值 or None}}，只保留 series 的 6 列口径。"""
    import openpyxl                                   # 延迟 import，纯读 CSV 的调用方不用装

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    if SHEET not in wb.sheetnames:
        raise HkexFetchError('xlsx 里没有 sheet %r（现有：%s）' % (SHEET, wb.sheetnames))
    ws = wb[SHEET]

    # 月份表头：第 2 列起是日期。别写死第 2 行，HKEX 加过说明行。
    hdr_row = None
    for r in range(1, 9):
        if isinstance(ws.cell(r, 2).value, _dt.datetime):
            hdr_row = r
            break
    if hdr_row is None:
        raise HkexFetchError('找不到月份表头行（前 8 行第 2 列都不是日期）')
    cols = {}
    for c in range(2, ws.max_column + 1):
        v = ws.cell(hdr_row, c).value
        if isinstance(v, _dt.datetime):
            cols['%04d-%02d' % (v.year, v.month)] = c
    if not cols:
        raise HkexFetchError('月份表头行没有任何日期列')

    rows = _find_rows(ws)
    data = {}
    for month, c in cols.items():
        g = {k: _num(ws.cell(r, c).value) for k, r in rows.items()}

        def add(a, b):
            """两段相加；两边都缺才算缺（GEM 段常年 0，但偶尔留空）。"""
            if g[a] is None and g[b] is None:
                return None
            return (g[a] or 0.0) + (g[b] or 0.0)

        data[month] = {
            'adt_hkdbn': None if g['adt_mil'] is None else g['adt_mil'] / 1000.0,
            'mktcap_hkdtn': None if g['mktcap_bil'] is None else g['mktcap_bil'] / 1000.0,
            'new_listings': add('newlist_mb', 'newlist_gem'),
            'ipo_funds_hkdbn': (lambda v: None if v is None else v / 1000.0)(
                add('ipo_mb_mil', 'ipo_gem_mil')),
            'derivatives_adv_contracts': g['deriv_adv'],
            'southbound_adt_hkdbn': (None if g['southbound_mil'] is None
                                     else g['southbound_mil'] / 1000.0),
        }
    return data


def _materially_differs(col, old_s, new_s):
    """末位 1 个单位以内的差算舍入噪声，不算重述。"""
    try:
        a, b = float(old_s), float(new_s)
    except ValueError:
        return True
    return abs(a - b) > 10.0 ** (-_DECIMALS[col]) + 1e-12


def _fmt(col, val):
    if val is None:
        return ''
    d = _DECIMALS[col]
    return ('%d' % round(val)) if d == 0 else ('%.*f' % (d, val))


# ── 月份算术（"YYYY-MM" ↔ 序号）─────────────────────────────────────────────
# update() 里本来有一个同名的闭包 month_key；回补逻辑在函数外也要用，所以提到模块层，
# 闭包改成引用它 —— 两份实现各自演化过一次就够写一天的对账。
def _mkey(m):
    """'2016-01' → 24193（自公元 0 年 1 月起的月序号，只用来比大小与加减）。"""
    y, mm = m.split('-')
    return int(y) * 12 + int(mm)


def _mstr(k):
    """_mkey 的逆。"""
    return '%04d-%02d' % ((k - 1) // 12, (k - 1) % 12 + 1)


# ══ 链路 B：成交股数 / 成交笔数 ═══════════════════════════════════════════
# 端点、口径、闭合证据全部写在模块 docstring 的「═══ 成交股数与成交笔数 ═══」一节，
# 这里只放代码。改动这一段前先把那节读完，尤其是「档案最后一个月必须被 Bulletin 确认」
# 那条闸门 —— 它是唯一能挡住「月中快照被当成整月合计入库」的东西。

_BULL_BASE = 'https://www.hkex.com.hk/eng/stat/smstat/mthbull/'
_ARCHIVE_INDEX = {
    'mb':  _BULL_BASE + 'rpt_data_statistics_archive_trading_data.json',
    'gem': _BULL_BASE + 'rpt_data_statistics_archive_trading_data_gem.json',
}
_HIGHLIGHT_URL = {
    'mb':  _BULL_BASE + 'rpt_Stock_market_highlights_%s.json',
    'gem': _BULL_BASE + 'rpt_Stock_market_highlights_GEM_%s.json',
}
# B3：市值逐日档案。只在历史回补时用（xlsx 最早 2018-01，series 要到 2016-01），
# 端点与闭合证据见模块 docstring 的 B3 与「闭合④」。
_MKTCAP_INDEX = {
    'mb':  _BULL_BASE + 'rpt_data_statistics_archive_market_cap.json',
    'gem': _BULL_BASE + 'rpt_data_statistics_archive_market_cap_gem.json',
}
_MKTCAP_COL = 'Total market capitalisation'
_BOARDS = ('mb', 'gem')
_BOARD_ZH = {'mb': '主板', 'gem': 'GEM'}

_ARCH_DATE_RE = re.compile(r'^(\d{4})/(\d{2})/(\d{2})$')
_BUCKET_RE = re.compile(r'_(\d{4})_(\d{4})\.json$')
_MON_FULL = ['January', 'February', 'March', 'April', 'May', 'June',
             'July', 'August', 'September', 'October', 'November', 'December']

# 逐日表的三个数值列。按**表头文字**认列，不按列号 —— 表头之间还夹着一个空列
# （放半日市那颗 `*`），HKEX 将来挪一下列序，认文字的写法不会静默错位。
_ARCH_COLS = (('value', 'Total trading value'),
              ('volume', 'Total trading volume'),
              ('deals', 'Number of deals'))

# Stock market highlights 里要的三行。用 startswith 匹配：HKEX 在标签尾部挂了
# `<br/>- Daily average`，而且 "No. of deals " 后面还有个尾随空格。
_HL_ROWS = (('value', 'Turnover value (HK$mil)'),
            ('volume', 'Turnover volume (mil shares)'),
            ('deals', 'No. of deals'))


def _tidy(s):
    """报表 JSON 的单元格文本 → 纯文本。`<br/>` 转成换行（Bulletin 靠它把
    「当月合计」和「日均」塞进同一格），其余标签与 &nbsp; 一律抹掉。"""
    s = re.sub(r'<\s*br\s*/?\s*>', '\n', s or '')
    s = re.sub(r'<[^>]+>', ' ', s)
    s = s.replace('\xa0', ' ').replace('&nbsp;', ' ')
    return '\n'.join(re.sub(r'\s+', ' ', ln).strip() for ln in s.split('\n')).strip()


def _grid(table, key):
    """report JSON 的 header/body 单元格列表 → 二维文本表（行内按 col 对齐补空）。"""
    cells = table.get(key) or []
    if not cells:
        return []
    ncol = max(c['col'] for c in cells) + 1
    rows = {}
    for c in cells:
        rows.setdefault(c['row'], {})[c['col']] = c.get('text', '')
    return [[rows[r].get(i, '') for i in range(ncol)] for r in sorted(rows)]


def _json_get(url, cache_dir=None, cache_name=None):
    """GET 一个 HKEX 报表 JSON。顺手把原文落到 cache/（gitignore 的），
    出了争议能拿原始档对账，不用重新做一遍考古。"""
    blob = _get(url, timeout=120)
    if cache_dir and cache_name:
        try:
            os.makedirs(cache_dir, exist_ok=True)
            with open(os.path.join(cache_dir, cache_name), 'wb') as f:
                f.write(blob)
        except OSError:
            pass                                  # 缓存写不进去不该让抓取失败
    try:
        return json.loads(blob.decode('utf-8-sig'))
    except ValueError as e:
        raise HkexFetchError('%s 返回的不是 JSON（%s；前 80 字节 %r）' % (url, e, blob[:80]))


def _int(txt):
    """'100,533,281' → 100533281。逐日档案里 21,390 个数值格实测**全是整数**
    （2026-08-07 全量扫过），所以这里不接受小数：真出现小数说明版式变了，该响。"""
    s = _tidy(txt).replace(',', '')
    if not re.match(r'^\d+$', s):
        raise HkexFetchError('逐日档案里出现非整数数值 %r' % txt)
    return int(s)


def _pick_buckets(index_url, since_year, cache_dir, index_name, what):
    """档案索引 → [(lo, hi, 分册 URL)]，只留年份区间覆盖到 since_year 及以后的分册。

    只下载用得上的分册：每册 ~600KB，全下 9 册纯属浪费；
    但也绝不按「当前年份」猜册名 —— 册名与区间一律从官方索引读。
    成交档案（B1）与市值档案（B3）的索引是同一种结构，所以共用这一段。
    """
    idx = _json_get(index_url, cache_dir, index_name)
    if not isinstance(idx, list) or not idx:
        raise HkexFetchError('%s索引不是非空数组（版式可能改了）：%s' % (what, index_url))
    picked, skipped = [], []
    for it in idx:
        url = (it or {}).get('url') or ''
        m = _BUCKET_RE.search(url)
        if not m:
            skipped.append(url)
            continue
        lo, hi = int(m.group(1)), int(m.group(2))
        if hi >= since_year:
            picked.append((lo, hi, urllib.parse.urljoin(_BULL_BASE, url)))
    if not picked:
        raise HkexFetchError('%s索引里没有覆盖 %d 年及以后的分册（索引 %d 条，'
                             '认不出册名的 %d 条）' % (what, since_year, len(idx), len(skipped)))
    return sorted(picked)


def _archive_daily(board, since_year, cache_dir):
    """某个板块的逐日档案 → {'YYYY-MM': [交易日数, 成交额HKD, 成交股数股, 成交笔数]}。"""
    picked = _pick_buckets(_ARCHIVE_INDEX[board], since_year, cache_dir,
                           'hkex_archive_index_%s.json' % board,
                           '%s逐日档案' % _BOARD_ZH[board])

    out = {}
    for lo, hi, url in sorted(picked):
        doc = _json_get(url, cache_dir, 'hkex_archive_%s_%d_%d.json' % (board, lo, hi))
        tables = doc.get('tables') or []
        if not tables:
            raise HkexFetchError('%s 逐日档案 %d-%d 没有 tables' % (_BOARD_ZH[board], lo, hi))
        for tbl in tables:
            head = [_tidy(x) for x in (_grid(tbl, 'header') or [[]])[0]]
            colof = {}
            for key, want in _ARCH_COLS:
                hit = [i for i, h in enumerate(head) if h.startswith(want)]
                if len(hit) != 1:
                    raise HkexFetchError(
                        '%s 逐日档案 %d-%d 的表头认不出「%s」（表头=%r）—— 版式变了，'
                        '宁可停也不能猜列号' % (_BOARD_ZH[board], lo, hi, want, head))
                colof[key] = hit[0]
            for row in _grid(tbl, 'body'):
                day = _tidy(row[0]) if row else ''
                if not day:
                    continue                      # 纯空行：忽略
                if not _ARCH_DATE_RE.match(day):
                    raise HkexFetchError('%s 逐日档案 %d-%d 出现非日期行 %r —— 可能是新加的'
                                         '小计/脚注行，必须人工确认后再改解析'
                                         % (_BOARD_ZH[board], lo, hi, day))
                if len(row) <= max(colof.values()):
                    # 表头有 5 列而这一行只给了 3 格：不能让它变成 IndexError，
                    # 那种堆栈看不出是"数据源缺格"还是"我写错了"。
                    raise HkexFetchError('%s 逐日档案 %d-%d 的 %s 行只有 %d 格，'
                                         '表头要到第 %d 格' % (_BOARD_ZH[board], lo, hi, day,
                                                              len(row), max(colof.values()) + 1))
                ym = day[:4] + '-' + day[5:7]
                a = out.setdefault(ym, [0, 0, 0, 0])
                a[0] += 1
                a[1] += _int(row[colof['value']])
                a[2] += _int(row[colof['volume']])
                a[3] += _int(row[colof['deals']])
    if not out:
        raise HkexFetchError('%s 逐日档案一条日线都没解析出来' % _BOARD_ZH[board])
    return out


def _archive_mktcap(board, since_year, cache_dir):
    """某个板块的**市值**逐日档案 → {'YYYY-MM': (最后一个交易日, 该日市值 HKD)}。

    只取每月最后一个交易日：市值是**时点数**，月内求和／求平均都不是任何东西
    （series 的 mktcap_hkdtn 一直是月末口径，坑 5 写明了）。
    「最后一个交易日」按档案里出现过的日期字符串取最大值 —— 档案本身只列交易日，
    不需要另外一份交易日历。

    与成交档案不同，这张表只有两列（日期 + 市值），也不打半日市那颗 `*`
    （2026-08-18 全量扫过 2015-2026 共 2,847 行主板 / 3,094 行 GEM，无一例外）。
    """
    picked = _pick_buckets(_MKTCAP_INDEX[board], since_year, cache_dir,
                           'hkex_mktcap_index_%s.json' % board,
                           '%s市值逐日档案' % _BOARD_ZH[board])
    out = {}
    for lo, hi, url in picked:
        doc = _json_get(url, cache_dir, 'hkex_mktcap_%s_%d_%d.json' % (board, lo, hi))
        tables = doc.get('tables') or []
        if not tables:
            raise HkexFetchError('%s 市值档案 %d-%d 没有 tables' % (_BOARD_ZH[board], lo, hi))
        for tbl in tables:
            head = [_tidy(x) for x in (_grid(tbl, 'header') or [[]])[0]]
            hit = [i for i, h in enumerate(head) if h.startswith(_MKTCAP_COL)]
            if len(hit) != 1:
                raise HkexFetchError('%s 市值档案 %d-%d 的表头认不出「%s」（表头=%r）—— '
                                     '版式变了，宁可停也不能猜列号'
                                     % (_BOARD_ZH[board], lo, hi, _MKTCAP_COL, head))
            col = hit[0]
            for row in _grid(tbl, 'body'):
                day = _tidy(row[0]) if row else ''
                if not day:
                    continue
                if not _ARCH_DATE_RE.match(day):
                    raise HkexFetchError('%s 市值档案 %d-%d 出现非日期行 %r —— 可能是新加的'
                                         '小计/脚注行，必须人工确认后再改解析'
                                         % (_BOARD_ZH[board], lo, hi, day))
                if len(row) <= col:
                    raise HkexFetchError('%s 市值档案 %d-%d 的 %s 行只有 %d 格，表头要到第 %d 格'
                                         % (_BOARD_ZH[board], lo, hi, day, len(row), col + 1))
                ym = day[:4] + '-' + day[5:7]
                prev = out.get(ym)
                if prev is None or day > prev[0]:
                    out[ym] = (day, _int(row[col]))
    if not out:
        raise HkexFetchError('%s 市值档案一条日线都没解析出来' % _BOARD_ZH[board])
    return out


def _highlights(board, month, cache_dir):
    """Monthly Bulletin 的 Stock market highlights → 该月三行的 (合计, 日均)。

    返回 {'value': (合计, 日均), 'volume': (...), 'deals': (...)}；
    该月没挂（HTTP 404）返回 None —— 只挂最近 13 个月是官方的正常状态，不是故障。
    """
    y, m = int(month[:4]), int(month[5:7])
    url = _HIGHLIGHT_URL[board] % ('%02d%02d' % (y % 100, m))
    try:
        doc = _json_get(url, cache_dir, 'hkex_highlights_%s_%s.json' % (board, month))
    except urllib.error.HTTPError as e:
        if e.code in (403, 404):
            return None
        raise HkexFetchError('%s %s Stock market highlights 取不到：%r' % (_BOARD_ZH[board], month, e))
    tables = doc.get('tables') or []
    if not tables:
        raise HkexFetchError('%s %s Stock market highlights 没有 tables' % (_BOARD_ZH[board], month))
    tbl = tables[0]

    # 认列：表头里写的是 "June 2026" / "June 2025"（本月 vs 去年同月）。
    # 必须按月份名认，不能默认第 1 列 —— 认错一列就是把去年同月当本月入库。
    head = [_tidy(x) for x in (_grid(tbl, 'header') or [[]])[0]]
    want = '%s %d' % (_MON_FULL[m - 1], y)
    hit = [i for i, h in enumerate(head) if h == want]
    if len(hit) != 1:
        raise HkexFetchError('%s %s 的 highlights 表头里找不到唯一的「%s」列（表头=%r）'
                             % (_BOARD_ZH[board], month, want, head))
    col = hit[0]

    body = _grid(tbl, 'body')
    out = {}
    for key, lab in _HL_ROWS:
        row = None
        for r in body:
            if _tidy(r[0]).startswith(lab):
                row = r
                break
        if row is None:
            raise HkexFetchError('%s %s 的 highlights 里找不到「%s」行'
                                 % (_BOARD_ZH[board], month, lab))
        if len(row) <= col:
            raise HkexFetchError('%s %s 的「%s」行只有 %d 格，取不到第 %d 格（%s）'
                                 % (_BOARD_ZH[board], month, lab, len(row), col + 1, want))
        parts = [p for p in _tidy(row[col]).split('\n') if p]
        if len(parts) != 2:
            raise HkexFetchError('%s %s 的「%s」格拆不出「合计 / 日均」两行（原文 %r）'
                                 % (_BOARD_ZH[board], month, lab, row[col]))
        out[key] = (_int(parts[0]), _int(parts[1]))
    return out


def _archive_row(mb, gem, month):
    """某月两板的逐日汇总 → 该月 7 个新列的值。"""
    days_mb, days_gem = mb[0], gem[0]
    if days_mb <= 0:
        raise HkexFetchError('%s 主板逐日档案 0 个交易日' % month)
    if days_gem != days_mb:
        # 同一家交易所、同一套交易日历，正常永远相等（2019-01~2026-06 实测 90/90 相等）。
        # 真差了也不该让整月落空 —— 交易日以主板为准，GEM 只占成交额 0.03%。
        print('[hkex] %s 主板 %d 个交易日、GEM %d 个，取主板；GEM 合计仍按原样入库'
              % (month, days_mb, days_gem))
    return {
        'trading_days_cash': days_mb,
        'vol_shares_mb_mn': mb[2] / 1e6,
        'vol_shares_gem_mn': gem[2] / 1e6,
        'trades_mb_total': mb[3],
        'trades_gem_total': gem[3],
        # 合计口径 = 主板+GEM，与 adt_hkdbn 完全一致（见 docstring 闭合 ②③）。
        # 先相加再取整，不是两个取整值相加 —— 后者会在末位上多出 ±1 的抖动。
        'adv_shares_mn': (mb[2] + gem[2]) / 1e6 / days_mb,
        'adt_trades': (mb[3] + gem[3]) / float(days_mb),
    }


def _confirm_with_bulletin(month, row, cache_dir):
    """拿 Monthly Bulletin 逐位核对某月的档案汇总。返回 (是否确认, 说明)。

    这是挡「月中快照」的唯一闸门：逐日档案自己不声明「这个月发完了没有」，
    只有 Bulletin 印出当月合计才等于官方宣布这个月收口了。
    """
    got = {}
    for b in _BOARDS:
        hl = _highlights(b, month, cache_dir)
        if hl is None:
            return False, 'Monthly Bulletin 还没挂 %s（只挂最近 13 个月，或该月尚未发布）' % month
        got[b] = hl
    checks = [
        ('主板当月成交股数(mil sh)', got['mb']['volume'][0], round(row['vol_shares_mb_mn'])),
        ('GEM当月成交股数(mil sh)', got['gem']['volume'][0], round(row['vol_shares_gem_mn'])),
        ('主板当月成交笔数', got['mb']['deals'][0], row['trades_mb_total']),
        ('GEM当月成交笔数', got['gem']['deals'][0], row['trades_gem_total']),
        ('主板日均成交股数(mil sh)', got['mb']['volume'][1],
         round(row['vol_shares_mb_mn'] / row['trading_days_cash'])),
        ('主板日均成交笔数', got['mb']['deals'][1],
         round(row['trades_mb_total'] / float(row['trading_days_cash']))),
    ]
    bad = ['%s：Bulletin %d vs 档案 %d' % (n, a, b) for n, a, b in checks if abs(a - b) > 1]
    if bad:
        return False, '与 Monthly Bulletin 对不上（%s）' % '；'.join(bad)
    return True, 'Monthly Bulletin %s 逐位确认（6 项，容差 ±1 个末位）' % month


def _bulletin_row(month, cache_dir):
    """只有 Bulletin、还没进逐日档案的月份 → 该月 7 个新列的值；没挂就 None。

    交易日不在 Bulletin 里印，但「当月合计 ÷ 日均」必然等于交易日数，
    而且主板与 GEM 两套数必须给出同一个整数 —— 对不上就说明我读错了格，宁可不要。
    """
    hl = {}
    for b in _BOARDS:
        one = _highlights(b, month, cache_dir)
        if one is None:
            return None
        hl[b] = one

    days = set()
    for b in _BOARDS:
        for key in ('value', 'volume', 'deals'):
            tot, avg = hl[b][key]
            if avg <= 0:
                continue
            days.add(int(round(tot / float(avg))))
    if len(days) != 1:
        print('[hkex] %s 的 Bulletin 反推交易日不唯一（%s），该月不入库'
              % (month, sorted(days)))
        return None
    d = days.pop()
    if d <= 0:
        return None
    return {
        'trading_days_cash': d,
        'vol_shares_mb_mn': hl['mb']['volume'][0],
        'vol_shares_gem_mn': hl['gem']['volume'][0],
        'trades_mb_total': hl['mb']['deals'][0],
        'trades_gem_total': hl['gem']['deals'][0],
        'adv_shares_mn': (hl['mb']['volume'][0] + hl['gem']['volume'][0]) / float(d),
        'adt_trades': (hl['mb']['deals'][0] + hl['gem']['deals'][0]) / float(d),
    }


def _since_year(months):
    """要下载的档案分册下限年份。

    取「序列现有首月」与 START_MONTH 中更早的那个 —— 回补还没发生时序列首月更晚，
    档案得先能覆盖到 START_MONTH，_backfill_rows 才有东西可用；
    回补做完之后两者相等，这里自然不再多下分册。
    """
    return min(int(min(months, key=_mkey)[:4]), int(START_MONTH[:4]))


def _cash_archives(since_year, cache_dir):
    """两个板块的成交逐日档案，一次下载。

    回补（_backfill_rows）与 7 个交易列（_trading_stats）用的是同一批分册，
    各下一遍等于把 ~2MB 白拉两次，也会让两处对「档案里有哪些月」的看法可能不一致。
    """
    return {b: _archive_daily(b, since_year, cache_dir) for b in _BOARDS}


def _backfill_rows(existing_months, xlsx, cache_dir, arch):
    """序列首月之前的历史行 → {'YYYY-MM': {列名: 值}}；没什么可补就返回 {}。

    规则（每一条都有代价，别顺手放宽）：
      · 只在**序列首月之前**造行，且不早于 START_MONTH。序列**内部**的空档一格不填 ——
        那是坑 4 说的口径留白，填了会把两代口径混进同一条序列。
      · 回补段必须是**紧贴序列首月往前的连续段**：从首月往前一个月一个月走，
        走到第一个「四套档案里缺任何一套」的月就停。只往右让、不往左借 ——
        中间断一个月就会在 adt / 交易日这些逐月必发的列上造出中间空洞，
        build/hkex.py 的完整性体检会 SystemExit，整页停更。
      · adt / mktcap：xlsx 覆盖得到的月份一律取 xlsx 原值（那是官方公告值）；
        xlsx 够不到的月份才由逐日档案重算（口径闭合证据见 docstring 闭合④）。
      · deriv / southbound：只取 xlsx 有的；xlsx 够不到就留空（前导空格，不是中间空洞）。
      · BACKFILL_HOLD 的两列：永远留空。
      · TRADING_COLUMNS 那 7 列不在这里出数 —— 行造出来之后由 _trading_stats 统一填，
        与 2019 年以后的月份走同一段代码，不另开一套。
    """
    if not existing_months:
        return {}
    first = min(existing_months, key=_mkey)
    if _mkey(START_MONTH) >= _mkey(first):
        return {}                                  # 序列已经够长，空转（不多下一份档案）

    # xlsx 够不到的月份要用市值逐日档案；够得到就不必下（省 ~1MB）。
    need_mktcap = _mkey(START_MONTH) < _mkey(min(xlsx) if xlsx else first)
    mcap = ({b: _archive_mktcap(b, int(START_MONTH[:4]), cache_dir) for b in _BOARDS}
            if need_mktcap else {b: {} for b in _BOARDS})

    def derivable(month):
        if month in xlsx and xlsx[month]['adt_hkdbn'] is not None \
                and xlsx[month]['mktcap_hkdtn'] is not None:
            # adt / mktcap 有 xlsx，但 7 个交易列仍要逐日档案 —— 缺了就是中间空洞
            return month in arch['mb'] and month in arch['gem']
        return all(month in d for d in (arch['mb'], arch['gem'], mcap['mb'], mcap['gem']))

    months, k = [], _mkey(first) - 1
    while k >= _mkey(START_MONTH) and derivable(_mstr(k)):
        months.append(_mstr(k))
        k -= 1
    if not months:
        print('[hkex] 回补：%s 之前一个月都补不出来（档案覆盖不到 %s），序列起点不变'
              % (first, _mstr(_mkey(first) - 1)))
        return {}
    if k >= _mkey(START_MONTH):
        print('[hkex] 回补：只补到 %s（再往前 %s 的档案不全），比 START_MONTH=%s 晚'
              % (months[-1], _mstr(k), START_MONTH))

    out = {}
    for month in sorted(months, key=_mkey):
        rec = xlsx.get(month) or {}
        row = {}
        for col in HIGHLIGHT_COLUMNS:
            if col in BACKFILL_HOLD:
                continue
            v = rec.get(col)
            if v is not None:
                row[col] = v
        if 'adt_hkdbn' not in row:
            mb, gem = arch['mb'][month], arch['gem'][month]
            if mb[0] <= 0:
                raise HkexFetchError('%s 主板逐日档案 0 个交易日，回补不了 adt_hkdbn' % month)
            row['adt_hkdbn'] = (mb[1] + gem[1]) / 1e9 / mb[0]
        if 'mktcap_hkdtn' not in row:
            d_mb, v_mb = mcap['mb'][month]
            d_gem, v_gem = mcap['gem'][month]
            if d_mb != d_gem:
                # 同一套交易日历，月末必然同一天（2015-2026 实测 139 个月无一例外）。
                # 真差了就是我读错了册或版式变了，宁可停 —— 两板取不同日的市值相加没有意义。
                raise HkexFetchError('%s 主板月末 %s 与 GEM 月末 %s 不是同一天，'
                                     '两板市值不能相加' % (month, d_mb, d_gem))
            row['mktcap_hkdtn'] = (v_mb + v_gem) / 1e12
        out[month] = row
    print('[hkex] 回补 %d 行：%s → %s（%s 起的 adt/mktcap 由逐日档案重算；'
          '%s 两列按 BACKFILL_HOLD 留空）'
          % (len(out), min(out, key=_mkey), max(out, key=_mkey),
             START_MONTH, '/'.join(BACKFILL_HOLD)))
    return out


def _trading_stats(months, cache_dir, arch=None):
    """series 已有的月份 → {'YYYY-MM': {新列名: 值}}。

    months 是 series/hkex.csv 现有的全部月份（**含本轮刚回补出来的历史行**）；
    本函数只给这些月出数，一个新行都不造 —— 造行是 _backfill_rows 的职责，
    两处各造各的会让「回补段必须连续」那条规则失效。

    优先级：逐日档案 > Monthly Bulletin。
      · 档案是逐日底稿，精度最高、口径与 adt_hkdbn 同源；
      · 档案的**最后一个月**必须过 Bulletin 确认，过不了就丢掉（可能是月中快照）；
      · 档案还没覆盖到的月份（实测会落后一整个月）交给 Bulletin。
    """
    if not months:
        return {}
    if arch is None:
        arch = _cash_archives(_since_year(months), cache_dir)
    common = sorted(set(arch['mb']) & set(arch['gem']))
    if not common:
        raise HkexFetchError('主板与 GEM 逐日档案没有共同月份（索引或版式可能变了）')

    out = {}
    last_arch = common[-1]
    for month in common:
        if month not in months:
            continue                              # 早于/晚于 series 现有行：不造行
        row = _archive_row(arch['mb'][month], arch['gem'][month], month)
        if month == last_arch:
            ok, why = _confirm_with_bulletin(month, row, cache_dir)
            if not ok:
                print('[hkex] 逐日档案最新月 %s 未通过确认（%s），本轮不写该月' % (month, why))
                continue
            print('[hkex] 逐日档案最新月 %s：%s' % (month, why))
        out[month] = row

    # 档案还没跟上的月份（只可能在档案末月之后）交给 Bulletin
    for month in sorted(m for m in months if m > last_arch):
        row = _bulletin_row(month, cache_dir)
        if row is None:
            continue
        out[month] = row
        print('[hkex] %s 逐日档案尚未覆盖，改用 Monthly Bulletin 当月合计' % month)
    return out


# ── 对外接口 ──────────────────────────────────────────────────────────────
def latest_month(cache_dir):
    """官方源当前最新月 "YYYY-MM"。抓不到 / 解析不出就抛 HkexFetchError。

    以**表里最后一个 ADT 非空的月**为准，而不是信文件名：
    HKEX 有时先把下个月的列开出来再填数，信文件名会得到一个空壳月。
    """
    path, claimed, _ = _download(cache_dir)
    data = _parse(path)
    filled = sorted(m for m, v in data.items() if v['adt_hkdbn'] is not None)
    if not filled:
        raise HkexFetchError('解析成功但没有任何月份有 ADT，文件疑似空壳：%s' % path)
    last = filled[-1]
    if last != claimed:
        print('[hkex] 注意：文件名声称 %s，实际最后有数的月是 %s' % (claimed, last))
    return last


def update(series_dir, cache_dir, allow_restate=False):
    """把官方源里 series 还没有的月份写进 series/hkex.csv，返回新增月份列表。

    「新增」= 该月的 adt_hkdbn 从"没有"变成"有"。这样定义是因为 build_hkex.py
    用 adt_hkdbn 挑 LATEST，一行只填了衍生品的残缺月对看板等于不存在。
    两种情况都算：整行追加、以及给已存在但 adt 为空的行补上数。

    幂等：已有且非空的单元格默认原样保留（allow_restate=True 才覆盖），
    重复跑第二遍返回 []。

    任何一个 HIGHLIGHT_COLUMNS 里的列在官方源里取不到值 → 抛异常，绝不写 NaN/空。

    链路 B（TRADING_COLUMNS，成交股数/成交笔数）在下面单独走一遍：它给**所有已有行**
    补空格，不受「adt 已有的行不动」那条规矩约束 —— 那条规矩管的是历史留白，
    而这 7 列是 2026-08 才新增的，全序列的空格都不是留白，是还没抓。

    历史回补（2026-08-18 新增）：序列**首月之前**的行由 _backfill_rows 造，最早到
    START_MONTH，回补段必须连续。它跑在 xlsx 那一轮之前，于是回补行在那一轮里被当成
    「已有行」，BACKFILL_HOLD 的两列不会被顺手填上。序列内部的空档仍然一格不填。
    序列一旦从 START_MONTH 起，这条路径下次跑就空转，连市值档案都不多下一份。

    序列落盘之后，顺手把这一档带来的那个最新月的发布日记进 series/source_dates.csv
    （页面抬头「官方发布于」用它），细节见下面那段注释与模块 docstring 的「发布日」节。
    """
    csv_path = os.path.join(series_dir, 'hkex.csv')
    if not os.path.exists(csv_path):
        raise HkexFetchError('找不到 %s；本模块只负责增量，不负责从零建序列' % csv_path)

    path, _, last_modified = _download(cache_dir)
    data = _parse(path)

    raw = open(csv_path, 'rb').read().decode('utf-8')
    nl = '\r\n' if '\r\n' in raw else '\n'
    lines = raw.replace('\r\n', '\n').rstrip('\n').split('\n')
    header = lines[0].split(',')

    body = [ln.split(',') for ln in lines[1:]]
    migrated = False
    if header == ['month'] + HIGHLIGHT_COLUMNS:
        # 一次性表头迁移：新 7 列追加在末尾，旧 7 个字段（month + 6 列）**原位不动**，
        # 每行补 7 个空格。之后由链路 B 逐格填上；填不上的就保持空。
        header = ['month'] + COLUMNS
        for f in body:
            f.extend([''] * len(TRADING_COLUMNS))
        migrated = True
    elif header != ['month'] + COLUMNS:
        raise HkexFetchError('series/hkex.csv 列名与预期不符：%s' % header)
    ncol = len(header)
    for f in body:
        if len(f) != ncol:
            raise HkexFetchError('series/hkex.csv 的 %s 行有 %d 个字段，表头是 %d 个'
                                 % (f[0], len(f), ncol))

    idx = {f[0]: i for i, f in enumerate(body)}
    last_csv_month = body[-1][0]

    added, filled_cells, restatements = [], [], []
    month_key = _mkey

    # ── 历史回补：只在序列首月之前造行，且不早于 START_MONTH ──
    # 放在 xlsx 那一轮**之前**跑，是为了让回补出来的行在下面被当成「已有行」处理：
    # 那一段对 had_adt=True 的行只做重述记账、不动任何空格，于是 BACKFILL_HOLD 的两列
    # 不会被 xlsx 顺手填上（填了就是中间空洞 → build 硬失败）。
    # 成交逐日档案在这里下载一次，下面的 TRADING_COLUMNS 直接复用同一份。
    arch = _cash_archives(_since_year(list(idx) or [START_MONTH]), cache_dir)
    for month, rec in sorted(_backfill_rows(list(idx), data, cache_dir, arch).items(),
                             key=lambda kv: month_key(kv[0])):
        body.append([month] + [_fmt(c, rec.get(c)) for c in HIGHLIGHT_COLUMNS]
                    + [''] * len(TRADING_COLUMNS))
        idx[month] = len(body) - 1
        added.append(month)

    for month in sorted(data, key=month_key):
        rec = data[month]
        if month in idx:
            # ── 已有行 ──
            # 只给「adt 为空」的行补空格。这类行是前人用比 xlsx 更快的来源先写了半行
            # （如 2026-07），等官方文件到位就该补齐。
            # 反过来，adt 已有的历史行里的空格是**故意留白**：new_listings / ipo 早期
            # 官方简报没发、southbound 有 2022-01~2025-06 的 42 个月停发窗口，
            # build_hkex.py 的图注专门讲了这个 gap。现在的 xlsx 虽然把这些历史都补全了，
            # 但一次性回填 = 把两代口径混进同一条序列、并改掉看板叙事，
            # 属于人工决策，不该由无人值守任务顺手做掉。
            row = body[idx[month]]
            had_adt = bool(row[1])
            for j, col in enumerate(HIGHLIGHT_COLUMNS, start=1):
                new = rec[col]
                if new is None:
                    continue
                s = _fmt(col, new)
                if not row[j]:
                    if had_adt:
                        continue                     # 历史留白，不动
                    row[j] = s
                    filled_cells.append((month, col, s))
                elif row[j] != s and _materially_differs(col, row[j], s):
                    # 只记"真差异"：末位 1 个单位的差是历史序列当年四舍五入留下的，
                    # 全记下来会把真正的官方重述（IPO 那种 5%）淹没掉
                    restatements.append((month, col, row[j], s))
                    if allow_restate:
                        row[j] = s
            if not had_adt and row[1]:
                added.append(month)
            continue

        # ── 新行：只接受"当月 6 列全齐"的月份 ──
        if month_key(month) <= month_key(last_csv_month):
            # 走到这里说明该月落在序列**内部**却没有对应行 —— 那是口径留白，不回填
            # （见 docstring 坑 4）。序列首月**之前**的历史行不走这条路径：
            # 它们已经由上面的 _backfill_rows 造好并进了 idx，在上面那一支处理完了。
            continue
        if rec['adt_hkdbn'] is None:
            continue                    # 官方把列开出来了但还没填数，等下个月
        missing = [c for c in HIGHLIGHT_COLUMNS if rec[c] is None]
        if missing:
            raise HkexFetchError(
                '%s 解析结果缺列 %s —— 拒绝写入残缺行。'
                '要么官方版式变了，要么该月确实只发了一半，请人工确认。' % (month, missing))
        # 新行先只填链路 A 的 6 列；链路 B 的 7 列紧接着在下面那一段填。
        body.append([month] + [_fmt(c, rec[c]) for c in HIGHLIGHT_COLUMNS]
                    + [''] * len(TRADING_COLUMNS))
        idx[month] = len(body) - 1
        added.append(month)

    # ── 链路 B：成交股数 / 成交笔数（TRADING_COLUMNS）──
    # 与上面那一段的两点不同，都是有意的：
    #   · 不看 had_adt。这 7 列是新增列，全序列的空格都不是「故意留白」，是还没抓。
    #   · 不造新行。只给 idx 里已经存在的月份补格（含本轮 _backfill_rows 刚造出来的历史行）；
    #     造行是 _backfill_rows 的独家职责，理由见模块 docstring「历史深度」一节。
    # 相同的是：非空格子一律不覆盖（allow_restate=True 才覆盖），所以第二遍跑一格不动。
    tcol_at = {c: j for j, c in enumerate(header) if c in set(TRADING_COLUMNS)}
    stats = _trading_stats(sorted(idx, key=month_key), cache_dir, arch=arch)
    for month in sorted(stats, key=month_key):
        row = body[idx[month]]
        for col, j in tcol_at.items():
            val = stats[month].get(col)
            if val is None:
                continue
            s = _fmt(col, val)
            if not row[j]:
                row[j] = s
                filled_cells.append((month, col, s))
            elif row[j] != s and _materially_differs(col, row[j], s):
                restatements.append((month, col, row[j], s))
                if allow_restate:
                    row[j] = s

    # 重述记录始终落盘，方便人工判断是口径变化还是解析出错
    if restatements:
        os.makedirs(cache_dir, exist_ok=True)
        rp = os.path.join(cache_dir, 'hkex_restatements.csv')
        with open(rp, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(['month', 'column', 'in_series_csv', 'in_official_source'])
            w.writerows(restatements)
        print('[hkex] 官方源与 series 有 %d 处不一致，已写 %s（allow_restate=%s）'
              % (len(restatements), rp, allow_restate))

    if not (migrated or added or filled_cells or (restatements and allow_restate)):
        return []

    body.sort(key=lambda f: month_key(f[0]))
    out = nl.join([','.join(header)] + [','.join(f) for f in body]) + nl
    tmp = csv_path + '.tmp'
    with open(tmp, 'wb') as f:
        f.write(out.encode('utf-8'))
    os.replace(tmp, csv_path)           # 原子替换，中途挂掉不会留半截文件

    # ── 发布日台账：只给「这一档 xlsx 自己带来的那个最新月」记一笔 ──
    # 文件里同时躺着 2018 年以来的全部历史，而这一档的上线时刻只能证明**最新月**是这天发的；
    # 本轮顺带补空的旧月是更早那些档发的，把今天的日期安到它们头上就是造假。
    # 也只在该月**首次入库**时记（in added）：HKEX 会原地重传，重传后的 Last-Modified
    # 更晚，事后覆盖等于把当初那次真发布的日期改错。
    src_months = [mth for mth, v in data.items() if v['adt_hkdbn'] is not None]
    src_last = max(src_months, key=month_key) if src_months else None
    if src_last and src_last in added:
        _record_source_date(series_dir, src_last, path, last_modified)

    if filled_cells:
        print('[hkex] 补空 %d 格：%s' % (len(filled_cells), filled_cells[:12]))
    return added


if __name__ == '__main__':
    import sys
    _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _cache = os.path.join(_root, 'cache')
    if len(sys.argv) > 1 and sys.argv[1] == 'update':
        print('added:', update(os.path.join(_root, 'series'), _cache))
    else:
        print('latest_month:', latest_month(_cache))
