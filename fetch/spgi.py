# -*- coding: utf-8 -*-
"""S&P Global (SPGI) 月度经营指标抓取模块。

════════ 数据源 ════════
发现入口（首选，无人值守可用）：
    https://investor.spglobal.com/feed/FinancialReport.svc/GetFinancialReportList
      ?apiKey=BF185719B0464B3CB809D23926182246&reportTypes=&year=-1&excludeSelection=1
    这是 Q4 Inc. 托管的 IR 站点自带的公开 JSON feed，apiKey 硬编码在
    investor.spglobal.com/quarterly-earnings/ 页面的 Q4Settings 里，不是密钥、无需登录。
    为什么必须走 feed 而不是解析页面 HTML：月度 xlsx 的链接是 JS 渲染的，
    静态 HTML 里根本没有 .xlsx 字样。

实际下载：feed 里 Documents[].DocumentPath 指向的 s29.q4cdn.com CDN 直链，形如
    https://s29.q4cdn.com/690959130/files/doc_financials/2026/q2/
        S-P-Global-Monthly-Metrics-as-of-June-2026-Published-7-15-2026.xlsx

════════ 反爬：为什么用 urllib 而不是 curl ════════
investor.spglobal.com 和 s29.q4cdn.com 都挂在 Cloudflare 后面。实测：
    · curl（默认 HTTP/2）→ 对 CDN 直接 "error in the HTTP2 framing layer"，对主站 403；
    · curl --http1.1 + 浏览器 UA → 200，可用但脆；
    · python urllib + 浏览器 UA → 200，主站 feed 和 CDN 都通。
所以本模块只用 urllib，绝不 shell out 到 curl。不需要浏览器登录态、不需要过验证码。

════════ 发布节奏 ════════
每月 15 日前后发布上一个月的数据（如 6 月数据 → 7/15 发布）；
季末月份的那一份**有时**会推迟到季报日随季报一起发，但这条规律并不稳。
工作簿右下角 "Published on M/D/YYYY" 是权威发布日，逐季实测（都是季末月）：

    2025-06 → 2025-07-31（月末后第 31 天）
    2025-12 → 2026-02-10（月末后第 41 天）
    2026-06 → 2026-07-15（月末后第 14 天）← 和常规月一样快，没等季报

跨度 27 天。所以**别按「季末月一定晚」来设下载闸门**：monthly_run.py 的
EARLY_BY['spgi'] 照最早的第 13-14 天开闸，宁可空跑几周也不能漏掉快的那一季
（roster 的 LAG[1]=46 照最慢的定，那是给红点用的，两者刻意不同）。

⚠️ feed 的一个结构性坑：月度 xlsx 不是独立条目，而是**挂在当季那条 Quarterly 记录下面
的附件**，季内每发一版就把上一版换掉。所以 feed 任何时刻只能看到「最新的那一份」，
看不到季内历史版本。这对本模块无所谓（我们只要最新一份，它含当年全部已披露月份），
但意味着补历史月份不能靠 feed，只能靠 CDN 直链猜文件名（见 _guess_cdn_urls）。

════════ 口径坑（决定了本模块能产出什么、不能产出什么）════════
1. Billed Issuance 官方**只披露同比百分比，从不给绝对面值**。所以
   billed_issuance_index 是我们自己按同比链式构造的指数（BASE_YEAR 同月 = 100），
   不是公司披露值。链式规则：index[y,m] = index[y-1,m] * (1 + yoy_fraction)，
   BASE_YEAR（=2022）各月的基数固定为 100 —— 那是同比链第一环 '23 v. '22 的分母年，
   它自己没有可用的同比数据。
2. SPDJI ADV 给绝对值，但每份 xlsx 只有「当年 + 上年」两列。
   序列起点是 2022-01：2022 年的绝对值由 2023 年值除以官方 '23 v. '22 同比反算
   （adv_derived=1 标记这一点）——是披露数据的算术推导，不是估计。
   **再往前不是拿不到，是从来没有过**：公司在 2023-02-09 的 Q4/FY2022 财报 8-K
   （SEC accession 0000064040-23-000055）"Upcoming Disclosures" 一节里预先宣布，
   这两条月度指标"beginning with results in 2023"才开始披露；工作簿也确实创刊于
   2023 Q1（as of March 2023）。2011-2022 各季在 Q4 feed 里只有财报稿 / slides /
   proxy / 年报，没有任何月度颗粒度的附件（那几份 2019-2021 的 "Supplemental
   Information" PDF 是**IHS Markit 的**文件，并购后被 IR 站吞并了文档历史才挂在
   同一个 CDN 上，两条序列它一条都没有）。取数与逐条实测见
   build/basefill/spgi_history.py。
   （历史回填靠那个脚本，不靠本模块：本模块每次只下载**最新一份** xlsx，
   正常跑一年序列也不会往回长一个月。）
3. 2025 年 12 月起 ADV 定义**剔除 event contracts，且官方不重述更早月份**。
   也就是说 2025-11 与 2025-12 之间存在一处口径断点，画图时要标注（build_spgi.py 已标）。
4. 百分比在 xlsx 里存的是小数（0.03 = +3%），CSV 里存的是百分数（3）。
   小数 ×100 会引入浮点噪声（0.29*100 == 28.999999999999996），
   spgi.csv 历史行是用 "%.15g" 清洗过的，spgi_clean.csv 历史行没清洗。
   本模块**分别复刻这两种写法**（见 _g15 / _repr），否则同一个月在两个文件里
   会出现肉眼可见的位数差异。已验证：按此规则重算，两个 CSV 的全部历史行
   与磁盘上的字节完全一致。
5. 每月 xlsx 是**全年重发**（不是增量），实测 2026-03 版与 2026-06 版对 1-3 月的
   数值完全一致，没有重述。但重述是可能的，所以 update() 会把工作簿重算的历史月份
   与 CSV 已有值对一遍，发现不一致时**告警但不改写历史**（改写历史必须人工决定）。

════════ 三道护栏：防「读不到不抛错」════════
本模块最危险的坏法不是抓取失败，是**官方改一个表头写法，我们把整整一列静默丢掉**。
那一列丢了不会报错：当年每个月都缺字段，_complete_months 只对夹在已披露区间里的
缺字段月份抛错（整年一个完整月份都不剩时它的 year_done 是空的，一声不响），于是
done 塌回上一年 12 月、update() 干净返回 []、日志和「这个月官方还没发」一模一样。
按 README「第四类：不出声的失败」的判据 —— 连续失败十天和成功十天在日志里长得
一样吗 —— 答案是「一样」，所以三道网一起上，方向不同、盲区不重叠：

  (a) 表头侧　_crosscheck_year_columns：Indices 表的同比年份集合必须等于 ADV
      年份集合。最早响，不依赖库里已有什么，parse() 里就响。
  (b) 反侧　　_crosscheck_stored_months：已入库且落在本期工作簿覆盖区间内的月份，
      必须重新解析得出来。看得见 (a) 看不见的坏法（Ratings 侧丢列、某格变横杠）。
  (c) 外部判据 _crosscheck_asof_month：文件名自报的 as-of 月 vs 表内最后一个完整
      月份。唯一一条**不经过 openpyxl** 的判据，形状照 fetch/cboe.py 的
      _crosscheck_report_month 与 fetch/ice.py 的 _crosscheck_workbook_month。

_adv_columns 顺带认了两位年写法（"'26 ADV …"），但那只是把**已经想到的**那一种
变体接住；接住下一种没想到的靠 (a)(b)(c)。别只留容错、删掉护栏。

════════ 对外接口 ════════
    latest_month(cache_dir) -> "YYYY-MM" | None
    update(series_dir, cache_dir) -> ["YYYY-MM", ...]   # 新增的月份，幂等
    published_on(xlsx_path) -> ("YYYY-MM-DD", 出处) | (None, None)

依赖：openpyxl（读 xlsx）。其余全是标准库，刻意不引 pandas —— 本模块的输出要和
历史 CSV 逐字节一致，pandas 的浮点格式化会自作主张。
"""

import calendar
import csv
import datetime as _dt
import json
import os
import re
import sys
import urllib.error
import urllib.request

import openpyxl

# ---------------------------------------------------------------- 常量

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

Q4_API_KEY = 'BF185719B0464B3CB809D23926182246'
Q4_FEED = ('https://investor.spglobal.com/feed/FinancialReport.svc/'
           'GetFinancialReportList?apiKey={key}&exchange=&symbol=&reportTypes='
           '&year=-1&excludeSelection=1&includeTags=true&pageSize=500')
CDN_BASE = 'https://s29.q4cdn.com/690959130/files/doc_financials'

# 走浏览器 UA 是硬要求：Cloudflare 对 python-urllib/3.x 这种 UA 直接 403。
UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
      '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

# billed_issuance_index 链式基期年：2023 是链上第一环（'23 v. '22），基数年就是 2022。
# build/basefill/spgi_history.py::BASE_YEAR 与本常数是同一个判断，改一个必须改另一个
# —— 不一致时增量追加会拿 100 当错误年份的基数，把整条链接歪且不报错。
BASE_YEAR = 2022

# as-of 月「只缺它自己」这一种形状的容忍期（天），见 _crosscheck_asof_month 的豁免段。
# 取值依据是本文件开头那张实测发布节奏表：as-of 月月末之后第 14 / 31 / 41 天各出过一次，
# 最慢的一次是 41 天。补发修正版走的是同一条节奏，所以 41 天是「还可能自愈」的上界；
# 留出约一个月的富余定在 75 天 —— 到那时下一期工作簿（M+1 那份，月末后 45-72 天）
# 本该早就把这一份顶掉了，还维持原样就不是「等官方补发」，而是没人看见的静默停更。
# 别把它调小到 41 附近：那等于把一次正常的季末迟发判成故障。
ASOF_BLANK_GRACE_DAYS = 75

SHEET_RATINGS = 'Ratings'          # 子串匹配，防官方改全名
SHEET_INDICES = 'Dow Jones'
MONTHS = list(calendar.month_name)[1:]      # January..December

# series/spgi.csv 的列（顺序即写盘顺序）
COLS_RAW = ['month', 'billed_issuance_yoy_pct',
            'spdji_adv_mn_contracts', 'spdji_adv_yoy_pct']
# series/spgi_clean.csv 的列
COLS_CLEAN = ['month', 'spdji_adv_mn', 'spdji_adv_yoy', 'billed_issuance_yoy',
              'adv_derived', 'billed_issuance_index']

_MM_RE = re.compile(r'monthly-metrics-as-of-([a-z]+)-(\d{4})', re.I)
# 工作簿正文下方那行 "Published on 7/15/2026"。日期是美式 M/D/YYYY —— 文件名里的
# "-Published-7-15-2026" 与它同源同序，两处对得上就说明这个读法没错。
_PUB_RE = re.compile(r'Published\s+on\s+(\d{1,2})/(\d{1,2})/(20\d{2})', re.I)


class SpgiFetchError(RuntimeError):
    """抓取 / 解析 / 口径校验失败。调用方应当让它冒出去，不要 fallback 成空数据。"""


# ---------------------------------------------------------------- 小工具

def _g15(x):
    """spgi.csv 的数值写法：15 位有效数字，把 ×100 引入的浮点尾巴抹掉。

    为什么不是 round(x, n)：round 到固定小数位会把 6.6892991348713515 这种
    本来就有意义的长尾截短；%.15g 只抹掉 28.999999999999996 → 29 这类噪声。
    """
    return float('%.15g' % x)


def _fmt15(x):
    return '%.15g' % x


def _repr(x):
    """spgi_clean.csv 的数值写法：Python 原生 repr，保留全部浮点位。"""
    return repr(float(x))


def _source_dates():
    """按路径加载仓库根的 source_dates.py（发布日台账）。

    不能裸 import：本模块是被 monthly_run 用 spec_from_file_location 加载的，
    那时 sys.path 上既没有 fetch/ 也没有仓库根。
    """
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(ROOT, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _http_get(url, timeout=120):
    req = urllib.request.Request(url, headers={
        'User-Agent': UA,
        'Accept': '*/*',
        'Accept-Language': 'en-US,en;q=0.9',
    })
    with urllib.request.urlopen(req, timeout=timeout) as fh:
        return fh.read()


def _ym(period):
    return '%04d-%02d' % period


def _parse_ym(s):
    y, m = s.split('-')
    return int(y), int(m)


def _prev_year(period):
    return (period[0] - 1, period[1])


# ---------------------------------------------------------------- 发现源文件

def _discover_via_feed():
    """从 Q4 feed 里挑出「as-of 月份最新」的一份 Monthly Metrics xlsx。

    返回 (url, asof_period)。asof 月份从文件名里取，不从 ReportDate 取——
    ReportDate 是季报的日期，和数据月份对不上（实测 Q2-2026 那条 ReportDate
    是 07/28/2026，但附件是 as-of June 2026）。
    """
    raw = _http_get(Q4_FEED.format(key=Q4_API_KEY), timeout=60)
    try:
        items = json.loads(raw)['GetFinancialReportListResult']
    except (ValueError, KeyError) as exc:
        raise SpgiFetchError('Q4 feed 返回的不是预期 JSON: %r' % (raw[:200],)) from exc

    found = []
    for item in items or []:
        for doc in item.get('Documents') or []:
            path = doc.get('DocumentPath') or ''
            if not path.lower().endswith(('.xlsx', '.xls')):
                continue
            hit = _MM_RE.search(os.path.basename(path))
            if not hit:
                continue
            name, year = hit.group(1).capitalize(), int(hit.group(2))
            if name not in MONTHS:
                continue
            found.append(((year, MONTHS.index(name) + 1), path))
    if not found:
        raise SpgiFetchError('Q4 feed 里没有任何 Monthly Metrics xlsx；'
                             'feed 结构可能变了，需人工检查')
    found.sort()
    return found[-1][1], found[-1][0]


def _guess_cdn_urls(period):
    """兜底：直接猜 CDN 上某个 as-of 月份的文件名。

    只在 feed 挂掉、或需要补 feed 已经换掉的季内历史版本时用。
    发布日在次月，最近几年固定 15 号，但季末月份会推到季报日，所以扫 8..28。
    大小写两种都试（2024/q1 那份是全小写，其余是首字母大写）。
    """
    y, m = period
    q = 'q%d' % ((m - 1) // 3 + 1)
    pub_y, pub_m = (y + 1, 1) if m == 12 else (y, m + 1)
    stem = 'S-P-Global-Monthly-Metrics-as-of-%s-%d-Published' % (MONTHS[m - 1], y)
    for day in list(range(8, 29)):
        for name in (stem, stem.lower()):
            yield '%s/%d/%s/%s-%d-%d-%d.xlsx' % (
                CDN_BASE, y, q, name, pub_m, day, pub_y)


def _resolve_source(cache_dir, refresh=False):
    """定位并下载最新一份 xlsx，返回 (本地路径, asof_period, 源 url)。

    refresh=True 时**无视本地缓存重下一遍**。默认 False，行为与从前一致。
    为什么需要这个开关：本地文件按 as-of 月做键，命中就永不重下 —— 而官方在同一个
    as-of 月内会补发修正版（缓存里 7 月那份自述 "Published on 8/17/2026"，
    6 月那份是 7/15/2026，都晚于常规的 15 号），补发时文件名里的 Published 日期
    变了、URL 也就变了。所以「重下」必须重新走一次发现（feed 每次都重新请求，
    这个函数本来就是先发现后落盘），不能拿旧 url 再 GET 一次原地打转。
    调用点见 _load_workbook_months：只有在 as-of 月没能完整解析出来时才会用到它。
    """
    try:
        url, period = _discover_via_feed()
    except (urllib.error.URLError, SpgiFetchError) as exc:
        # feed 不可用时，从「上个月」往回猜 6 个月，找到第一个能下的。
        sys.stderr.write('[spgi] feed 不可用（%s），改用 CDN 文件名猜测兜底\n' % exc)
        today = _dt.date.today()
        url = period = None
        probe = (today.year, today.month)
        for _ in range(6):
            probe = (probe[0] - 1, 12) if probe[1] == 1 else (probe[0], probe[1] - 1)
            for cand in _guess_cdn_urls(probe):
                try:
                    _http_get(cand, timeout=20)
                except urllib.error.URLError:
                    continue
                url, period = cand, probe
                break
            if url:
                break
        if not url:
            raise SpgiFetchError(
                'Q4 feed 与 CDN 猜名两条路都失败，判定为 blocked。'
                '人工兜底：浏览器打开 https://investor.spglobal.com/quarterly-earnings/ '
                '展开当季 Quarterly Earnings 卡片，下载 "Monthly Metrics" xlsx，'
                '手动放到 cache/ 下再跑 update()。')

    os.makedirs(cache_dir, exist_ok=True)
    local = os.path.join(cache_dir, 'spgi_monthly_metrics_%s.xlsx' % _ym(period))
    if refresh or not (os.path.exists(local) and os.path.getsize(local) > 4096):
        blob = _http_get(url)
        if not blob.startswith(b'PK'):
            raise SpgiFetchError('下载到的不是 xlsx（前 4 字节 %r），可能被拦截页顶替了'
                                 % blob[:4])
        with open(local, 'wb') as fh:
            fh.write(blob)
    return local, period, url


# ---------------------------------------------------------------- 解析 xlsx

def _sheet(wb, needle):
    for name in wb.sheetnames:
        if needle.lower() in name.lower():
            return wb[name]
    raise SpgiFetchError('工作簿里找不到含 %r 的 sheet，实际有 %r' % (needle, wb.sheetnames))


def _header_row(ws):
    """找到表头行号：第一行里出现 "% Change" 的那一行。"""
    for row in range(1, min(ws.max_row, 15) + 1):
        for col in range(1, min(ws.max_column, 15) + 1):
            v = ws.cell(row, col).value
            if isinstance(v, str) and '% change' in v.lower():
                return row
    # %% 是必须的：字面量里那个 "% Change" 会被 % 格式化当成转换说明符，
    # 写成单个 % 时这一行抛的是 TypeError 而不是 SpgiFetchError —— 护栏自己的
    # 报错路径挂掉，人看到的就不是「表头找不到」而是一句莫名其妙的格式化错误。
    raise SpgiFetchError('sheet %r 里找不到 "%% Change" 表头行' % ws.title)


def _yoy_columns(ws, hrow):
    """把 "'26 v. '25 % Change" 这类表头解析成 {年份: 列号}。

    年份用两位数写的，补成 20xx。硬编码列号是不行的——Ratings sheet 与
    Indices sheet 的列位不同，且官方历史上挪过列。
    """
    out = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(hrow, col).value
        if not isinstance(v, str) or '% change' not in v.lower():
            continue
        hit = re.search(r"'(\d{2})\s*v\.?\s*'(\d{2})", v)
        if not hit:
            raise SpgiFetchError('看不懂的同比表头 %r（sheet %r）' % (v, ws.title))
        out[2000 + int(hit.group(1))] = col
    if not out:
        raise SpgiFetchError('sheet %r 表头行没有同比列' % ws.title)
    return out


# ADV 表头里的**两位年**写法（"'26 ADV (in millions of contracts)"）。
# 同一行的同比表头本来就是两位年（"'26 v. '25 % Change"），官方哪天把 ADV 表头
# 也改成同一套写法，只认四位年的那条 re.search(r'(20\d{2})') 就会把当年那一整列
# 静默丢掉 —— 后果不是报错，是当年所有月份都缺 adv 字段、被 _complete_months
# 当「还没披露」丢掉，done 塌回上一年 12 月，update() 干净地报 NOCHANGE。
# 必须锚在撇号上（直角撇 ' 与弯撇 ’ 都要），并且后面不许再跟数字：
# 不锚的话 "…contracts) 26" 这种尾巴、四位年的后两位都会被认成年份，
# 那是把一列数安到错误的年份上，比丢掉更坏。
_ADV_YEAR2_RE = re.compile(r"['’](\d{2})(?!\d)")


def _adv_columns(ws, hrow, default_year=None):
    """把 "2026 ADV (in millions of contracts)" 解析成 {年份: 列号}。

    2023 年那四份（本工作簿系列的头四期）表头里**不写年份**，只有一句
    "ADV (in millions of contracts)" —— 因为当时表里只有一年，不需要区分。
    那种表头拿 default_year 兜底（调用方从月份行标签 "Jan 2023" 里取到的年份）。
    没有 default_year 又读不出年份时照旧报错：宁可炸也不能把一列数安到猜的年份上。

    三种写法按「越具体越优先」的顺序试：四位年 → 两位年 → 无年份 + default_year。
    ⚠ 加宽写法只是把**已经想到的**那一种变体接住了，接不住下一种没想到的；
    真正兜底的是 _crosscheck_year_columns（同一张表的同比年份集合与 ADV 年份集合
    必须一致）。两条要一起在，别只留这一条 —— 参见 fetch/msci.py 口径坑 6 的原话。
    """
    out = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(hrow, col).value
        if isinstance(v, str) and 'adv' in v.lower():
            hit = re.search(r'(20\d{2})', v) or _ADV_YEAR2_RE.search(v)
            if hit:
                got = int(hit.group(1))
                out[got if got > 99 else 2000 + got] = col
            elif default_year is not None and 'million' in v.lower():
                # 无年份表头只允许出现一次：出现两列就说明这份表里有多年数据却都不标年，
                # 那时 default_year 会把两列覆盖成同一年，属于静默错配，必须炸。
                if default_year in out:
                    raise SpgiFetchError(
                        'Indices sheet 有多列不带年份的 ADV 表头，无法判定各属哪一年')
                out[default_year] = col
    if not out:
        raise SpgiFetchError('Indices sheet 表头行没有 ADV 绝对值列')
    return out


def _crosscheck_year_columns(ws_title, yoy_years, adv_years, path):
    """Indices 表：同比列认出的年份集合，必须和 ADV 绝对值列认出的年份集合一样。

    这是**表头这一侧**的对账，三道网里最早响的那一道（parse() 里就响，
    还没轮到 CSV、也不依赖库里已经有什么）。防的是：官方改了 ADV 表头的年份写法
    （四位年 → 两位年、或者干脆不写年），_adv_columns 认不出于是**整整一列**被
    静默丢掉。那一列丢了不会报错 —— 当年每个月都缺 adv 字段，而 _complete_months
    只对**夹在已披露区间里**的缺字段月份抛错，整年一个完整月份都不剩时它的
    year_done 是空的，一句话都不说。结果是 done 塌回上一年 12 月，
    update() 干干净净返回 []，日志和「这个月官方还没发」长得一模一样。

    判据成立的理由：这两组列出自**同一张表的同一行表头**、同一次发布，官方每年
    开一列 ADV 就同时开一列同比、滚掉一年也是两列一起滚。cache/basefill/spgi/ 里
    2023-03 到 2026-07 的 15 份，加上 cache/ 里当前那 2 份（as-of 月与其中两份重复），
    共 17 份文件、跨三种版式（2023 年的单年截断版、2024-2025 年的三年版、
    2026 年的两年版）逐份验过，无一例外都对称。

    ⚠ 刻意**只在 Indices 表内部对账，不跨表和 Ratings 比**。跨表也是 17 份全对称，
    但 Ratings（评级）与 SPDJI（指数）是两条业务线各自出的数，哪天一边多留一年
    历史都算正常源行为 —— 拿它 raise 会把一个健康的源变成天天 FAIL。
    Ratings 那一侧丢列由 _crosscheck_stored_months 与 _crosscheck_asof_month 兜。
    """
    if set(yoy_years) != set(adv_years):
        raise SpgiFetchError(
            'sheet %r 的表头对不上：同比列认出的年份是 %r，ADV 绝对值列认出的是 %r。'
            '这两组列出自同一行表头、同一次发布，不该不一致 —— 最可能是某一侧的年份'
            '写法变了（四位年 ↔ 两位年、或表头措辞改了）导致整整一列被静默丢弃。'
            '拒绝写入，请人工打开 %s 看一眼表头行。'
            % (ws_title, sorted(yoy_years), sorted(adv_years), os.path.basename(path)))


# 月份行标签的两种写法：现行版式是裸月名（"January"），2023 年那四份带年份且
# 月名缩写不统一（"Jan 2023" / "June 2023" / "September 2023" 混排）。
# 英文月名的前三个字母互不重复，所以按前缀匹配是唯一的。
_MONTH_LABEL_RE = re.compile(r'^([A-Za-z]{3,9})\.?\s*(20\d{2})?$')


def _month_label(text):
    """月份行标签 → (月份序号 1..12, 年份或 None)；不是月份标签返回 (None, None)。"""
    hit = _MONTH_LABEL_RE.match((text or '').strip())
    if not hit:
        return None, None
    word = hit.group(1).lower()
    got = [i + 1 for i, name in enumerate(MONTHS) if name.lower().startswith(word)]
    if len(got) != 1:
        return None, None
    return got[0], (int(hit.group(2)) if hit.group(2) else None)


def _month_rows(ws, hrow):
    """表头之后的月份行，返回 ({月份序号: 行号}, 标签里的年份或 None)。

    版式约束（这是这个函数存在的理由 —— 它挡的是官方改版后静默读错行）：
    月份必须**从 1 月起逐月连续**。现行版式恒为 12 行（未披露的月份是空值行，
    行还在）；2023 年那四份是当季截断的（3 / 6 / 9 / 12 行），所以不能一律要求 12。
    改成「从 1 月起连续」既覆盖了这两种真实版式，又保留了原判据的作用：
    官方哪天挪行或漏行，认出来的就不再是 1..N 的连续段，照样炸。
    """
    out, years = {}, set()
    for row in range(hrow + 1, min(ws.max_row, hrow + 30) + 1):
        v = ws.cell(row, 1).value
        if not isinstance(v, str):
            continue
        mno, yr = _month_label(v)
        if mno is None:
            continue
        if mno in out:
            raise SpgiFetchError('sheet %r 里 %s 出现了两次（行 %d 与 %d）'
                                 % (ws.title, MONTHS[mno - 1], out[mno], row))
        out[mno] = row
        if yr:
            years.add(yr)
    if not out:
        raise SpgiFetchError('sheet %r 里一个月份行都没认出来' % ws.title)
    if sorted(out) != list(range(1, len(out) + 1)):
        raise SpgiFetchError('sheet %r 的月份行不是从 1 月起逐月连续，认出的是 %r'
                             % (ws.title, sorted(out)))
    if len(years) > 1:
        raise SpgiFetchError('sheet %r 的月份标签跨了多个年份 %r' % (ws.title, sorted(years)))
    return out, (years.pop() if years else None)


def _num(cell):
    v = cell.value
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, str):
        raise SpgiFetchError('期望数字，读到字符串 %r' % v)
    return float(v)


def parse(xlsx_path):
    """解析一份 xlsx，返回 {(year, month): {字段: 原始小数/绝对值}}。

    注意同比字段这里保持 xlsx 原样的**小数**（0.03 = +3%），不在这里 ×100，
    以便 index 链式计算用未经清洗的原值（历史 CSV 就是这么算的）。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=False)

    ws_r = _sheet(wb, SHEET_RATINGS)
    hr = _header_row(ws_r)
    r_yoy = _yoy_columns(ws_r, hr)
    r_rows, _ = _month_rows(ws_r, hr)

    ws_i = _sheet(wb, SHEET_INDICES)
    hi = _header_row(ws_i)
    i_yoy = _yoy_columns(ws_i, hi)
    # 先读月份行：2023 年那四份的 ADV 表头不带年份，年份只能从行标签 "Jan 2023" 里取。
    i_rows, i_label_year = _month_rows(ws_i, hi)
    i_adv = _adv_columns(ws_i, hi, default_year=i_label_year)
    _crosscheck_year_columns(ws_i.title, i_yoy, i_adv, xlsx_path)

    data = {}
    for year, col in sorted(r_yoy.items()):
        for m, row in r_rows.items():
            v = _num(ws_r.cell(row, col))
            if v is not None:
                data.setdefault((year, m), {})['billed_yoy'] = v
    for year, col in sorted(i_yoy.items()):
        for m, row in i_rows.items():
            v = _num(ws_i.cell(row, col))
            if v is not None:
                data.setdefault((year, m), {})['adv_yoy'] = v
    for year, col in sorted(i_adv.items()):
        for m, row in i_rows.items():
            v = _num(ws_i.cell(row, col))
            if v is not None:
                data.setdefault((year, m), {})['adv'] = v
    if not data:
        raise SpgiFetchError('xlsx 解析出 0 条月度数据: %s' % xlsx_path)
    return data


def published_on(xlsx_path):
    """工作簿自述的发布日，返回 ("YYYY-MM-DD", 出处描述)；没写就 (None, None)。

    两个 sheet 的正文下方各有一行 "Published on 7/15/2026"（当前版式落在 A22）。
    这是**源头自己说的**发布日，页面抬头「官方发布于」只认这一种来源 ——
    文件 mtime 是我们下载它的时间，构建日更是与官方无关，都不能顶替。

    整表扫字符串而不是钉死 A22：那行的行号跟在 Definition/Source 两段文字后面，
    官方改一次措辞就会上下浮动，钉死单元格的失败方式是**静默读到 None**。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        hits = {}
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    hit = _PUB_RE.search(cell.value)
                    if hit:
                        iso = '%s-%02d-%02d' % (hit.group(3), int(hit.group(1)),
                                                int(hit.group(2)))
                        hits.setdefault(iso, []).append(
                            '%s!%s %r' % (ws.title, cell.coordinate, cell.value.strip()))
    finally:
        wb.close()
    if not hits:
        return None, None
    # 两个 sheet 的页脚本该同期刷新。真不一致时取最晚的那个：整份工作簿是一次发布事件，
    # 更可能是官方漏刷了其中一张的页脚，而不是它真的早发了一版。但这属于官方版式异常，
    # 必须吵出来让人看一眼。
    iso = max(hits)
    if len(hits) > 1:
        sys.stderr.write('[spgi] ⚠ 工作簿里的 Published on 有多个日期（%s），取最晚的 %s；'
                         '请人工确认官方是不是漏刷了某个 sheet 的页脚\n'
                         % (', '.join(sorted(hits)), iso))
    return iso, '; '.join(hits[iso])


NEED = ('billed_yoy', 'adv', 'adv_yoy')


def _complete_months(data):
    """三个字段齐全的月份 —— 只有这种月份才允许入库。

    还要顺手抓一种致命情况：某个月**在已披露区间之内**却缺列。
    正常的 xlsx 是「已披露的月份三列全有，未披露的月份三列全空」，
    出现「有 ADV 没有 Billed Issuance」只能是官方改版或我们的列定位错了，
    这时候写库就会静默产生 NaN —— 所以直接抛异常，让人来看。
    """
    complete, partial = [], []
    for p, d in sorted(data.items()):
        (complete if all(k in d for k in NEED) else partial).append(p)
    for p in partial:
        year_done = [q for q in complete if q[0] == p[0]]
        if year_done and p < max(year_done):          # 夹在已披露区间里
            raise SpgiFetchError(
                '%s 处在已披露区间内却缺字段（拿到 %r，需要 %r）；'
                '不做静默跳过，也不写 NaN，请人工核对 xlsx 版式是否变了'
                % (_ym(p), sorted(data[p]), list(NEED)))
    return complete


def _crosscheck_stored_months(done, data, asof, stored, path):
    """已经入库的月份，在本期工作簿里必须重新出现 —— 少一个就炸。

    反侧那道网（照抄 fetch/msci.py update() 的哨兵②）。上面 _crosscheck_year_columns
    是从「表头上有什么」这一侧查的，看不见 Ratings 表丢列、也看不见「某一格变成
    横杠」这类坏法；这一道从「我们已知该有什么」的反侧查，两边盲区不重叠。

    判据靠得住的理由：官方每月**全年重发**（模块 docstring 口径坑 5），已入库的
    月份下个月还会原样再来一遍 —— 实测 2026-03 版与 2026-06 版对 1-3 月逐位相同。
    所以一个已入库的月份在本期工作簿里「解析不出来」，几乎只能是我们把行/列丢了。
    ⚠ 与「数值变了」处置不同：数值变了是官方重述，只告警不改历史（见 update()
    里那段循环）；整月消失不是源的正常行为，直接抛。

    ── 两条圈定范围的条件，缺一条这道护栏就变成天天误杀 ───────────────────
    (1) 只查落在本期工作簿**覆盖区间**里的月份。工作簿只带「当年 + 上年（早年那几份
        是三年）」，2026 年那两份已经把 2023 / 2024 滚掉了，而 series/spgi.csv 从
        2023-01 起 —— 不圈范围的话每一次跑都会为 2023 年的行报错。
        区间左端取 min(data)：滚掉一年是官方每年 1 月的正常动作，不能写死年份。
    (2) 只查 <= as-of 月的月份。feed 挂掉时 _resolve_source 会往回猜最多 6 个月，
        解析到的可能是一份旧工作簿；2023 年那四份更是按当季截断的（3/6/9/12 行）。
        比 as-of 还新的库内月份本来就不该在里面，不是丢了。

    已知的残余盲区（刻意留着，不要为它加检查）：整区间**最左端那一年**整体丢列时，
    min(data) 会跟着右移，那一年就被条件 (1) 划到范围外。代价有限 —— 那些月份早已
    入库，页面不会因此停更，丢的只是重述体检的基线；而收紧它就必然要写死年份，
    每年 1 月官方滚掉一年时就会误杀一次。
    """
    have = set(done)
    floor = min(data)
    lost = [ym for ym in sorted(stored)
            if floor <= _parse_ym(ym) <= asof and _parse_ym(ym) not in have]
    if lost:
        raise SpgiFetchError(
            '%d 个已入库的月份在本期工作簿里没能完整解析出来：%s（工作簿 %s，'
            'as-of=%s，覆盖到 %s 起）。官方每月全年重发，已入库的月份不该消失 —— '
            '最可能是某一列/某一行的写法变了被静默丢弃。本次不写入，请人工核对版式。'
            % (len(lost), ', '.join(lost[:6]), os.path.basename(path),
               _ym(asof), _ym(floor)))


def _prev_month(period):
    return (period[0] - 1, 12) if period[1] == 1 else (period[0], period[1] - 1)


def _crosscheck_asof_month(asof, done, data, path, where):
    """文件名自报的 as-of 月 vs 表里最后一个完整月份，对不上就炸。

    这是本模块**唯一一条独立于解析器**的月份判据：as-of 月是 _MM_RE 从 CDN
    文件名（feed 的 DocumentPath）里取的，一个字节都不经过 openpyxl，所以解析器
    认错行、认丢列，它不会跟着错。形状照抄 fetch/cboe.py 的 _crosscheck_report_month
    与 fetch/ice.py 的 _crosscheck_workbook_month，连「刻意 raise 而不是 warn」
    这一条也照抄 —— 那边的原话是：warn 之后状态仍是 NOCHANGE，等于没有护栏。
    本模块原来正是只在这里 sys.stderr.write 了一句提醒。

    防的是这一类：官方改了某个表头（ADV 的年份写法、Ratings 那列的 "% Change"
    措辞），整整一列被静默丢掉，当年所有月份都缺字段，_complete_months 把它们
    当「还没披露」丢掉（它只对夹在已披露区间里的缺字段月份抛错，整年不剩时
    year_done 是空的，一声不响），done 塌回上一年 12 月，update() 干净返回 []。
    没有 FAIL、没有红点、断档也抓不到 —— README「第四类：不出声的失败」那句判据
    「连续失败十天和成功十天在日志里长得一样吗」，在这里答案是「一样」。

    ⚠ 必须在 update() 里那句 `if not done: return []` **之前**调用。整份表一个完整
    月份都不剩时（同比列与 ADV 列同时改写法就会这样），那句 return 会先跑掉，
    连上面那行提醒都不会打 —— 那是最安静的一种坏法，护栏站在它后面等于不存在。

    ── 唯一的豁免，以及为什么非留不可 ─────────────────────────────────────
    官方确实可能把某一期发成半成品：Ratings 那几格填了、SPDJI 的 ADV 还空着
    （本模块原先那句提醒写的「多半是官方某一列漏填了」说的就是这个）。那种情况下
    done[-1] 恰好停在 as-of 的**前一个月**，别的什么都不缺。对这一种形状只喊不炸：
    否则一次几天内自愈的源侧抖动，会让 spgi 天天 FAIL 到下一份工作簿发布为止，
    而「把一个本来好好的源变成天天 FAIL」比它要治的陈旧更贵。
    判据写成「done[-1] 正好是 as-of 的前一个月」而不是「差得不多」：整列丢失会把
    **同年更早的月份**一起带走，done[-1] 会退得更远，落不进这个豁免。

    豁免带时限 ASOF_BLANK_GRACE_DAYS：源侧抖动会自愈，版式变更不会。过了这个期限
    还是同一副样子，就不再是「等官方补发」，而是没人看见的静默停更，该炸。
    ⚠ 这个时限只有配上 _load_workbook_months 那次强制重下才成立：本地文件按 as-of
    月做键、命中就永不重下，不重下的话官方补发的修正版我们根本看不到，容忍期一到
    必然误杀一次本来已经自愈了的抖动。两处要一起在。
    """
    if done and done[-1] == asof:
        return

    prev = _prev_month(asof)
    only_asof_missing = (bool(done) and done[-1] == prev
                         and asof in data and any(k in data[asof] for k in NEED))
    if only_asof_missing:
        # as-of 月月末之后过了多少天。用月末而不是月初：发布节奏表就是按月末算的。
        end = _dt.date(asof[0], asof[1], calendar.monthrange(asof[0], asof[1])[1])
        aged = (_dt.date.today() - end).days
        if aged <= ASOF_BLANK_GRACE_DAYS:
            sys.stderr.write(
                '[spgi] ⚠ %s：文件名 as-of=%s，但表内最后完整月份=%s，只差这一个月，'
                '且 %s 那一行认得出（拿到 %r）—— 判为官方漏填了某一格，先只提醒。'
                '月末后已 %d 天，超过 %d 天仍如此就会改判为故障并 FAIL。\n'
                % (where, _ym(asof), _ym(done[-1]), _ym(asof),
                   sorted(data[asof]), aged, ASOF_BLANK_GRACE_DAYS))
            return
        raise SpgiFetchError(
            '%s：as-of=%s 的这一格已经空了 %d 天（超过 %d 天的容忍期），表内最后完整'
            '月份仍是 %s。源侧漏填几天内会补发修正版，拖这么久说明不是漏填 —— '
            '要么版式变了没人发现，要么官方停更了。工作簿 %s，请人工看一眼。'
            % (where, _ym(asof), aged, ASOF_BLANK_GRACE_DAYS, _ym(done[-1]),
               os.path.basename(path)))

    raise SpgiFetchError(
        '%s：文件名自报 as-of=%s，但表内最后一个完整月份是 %s（工作簿 %s）。'
        '二者同批发布、不该不一致，而且缺的不止 as-of 这一个月 —— 最可能是某一整列'
        '（同比列或 ADV 列）的表头写法变了被静默丢弃。拒绝写入，请人工看一眼表头行。'
        % (where, _ym(asof), _ym(done[-1]) if done else '一个都没有',
           os.path.basename(path)))


# ---------------------------------------------------------------- CSV 读写

def _read_csv(path, cols):
    if not os.path.exists(path):
        raise SpgiFetchError('缺少序列文件 %s；本模块只负责追加，不负责从零建库' % path)
    with open(path, newline='', encoding='utf-8') as fh:
        rdr = csv.reader(fh)
        header = next(rdr)
        if header != cols:
            raise SpgiFetchError('%s 列名变了：磁盘 %r，期望 %r' % (path, header, cols))
        return header, [r for r in rdr if r and r[0].strip()]


def _append_csv(path, lines):
    """追加新行。

    两个坑：
    · 换行符必须沿用原文件的 —— 实测 spgi.csv 是 CRLF、spgi_clean.csv 是 LF，
      同一个仓库里两种都有。写错会让 git diff 整段飘红。
    · 原文件结尾若没有换行符，先补一个，否则新行会粘在最后一行后面。
    """
    with open(path, 'rb') as fh:
        blob = fh.read()
    eol = b'\r\n' if b'\r\n' in blob else b'\n'
    with open(path, 'ab') as fh:
        if blob and not blob.endswith(b'\n'):
            fh.write(eol)
        fh.write(eol.join(l.encode('utf-8') for l in lines) + eol)


# ---------------------------------------------------------------- 对外接口

def _load_workbook_months(cache_dir):
    """下载 + 解析 + 挑出完整月份，返回 (路径, asof, url, data, done)。

    as-of 月没能完整解析出来时，**强制重下一次再判**。这一步是
    _crosscheck_asof_month 敢 raise 的前提，不是可有可无的优化：
    本地文件按 as-of 月做键、命中就永不重下（见 _resolve_source），而官方会在
    同一个 as-of 月内补发修正版。没有这一次重下，官方发的第一版里哪怕只是漏填
    一格，我们也会抱着那份残缺的本地副本一直判到容忍期满 —— 一次两三天就自愈的
    源侧抖动会被判成一个月的 FAIL。重下只在「已经不对劲」时发生，正常月份不多花
    一个请求。

    重下本身失败（网络抖动、feed 临时挂）不能反过来拖垮这一轮：喊一声，
    拿第一次的结果继续往下判，护栏该响还是会响。
    """
    path, asof, url = _resolve_source(cache_dir)
    data = parse(path)
    done = _complete_months(data)
    if not done or done[-1] != asof:
        try:
            path, asof, url = _resolve_source(cache_dir, refresh=True)
            data = parse(path)
            done = _complete_months(data)
        except (urllib.error.URLError, OSError) as exc:
            sys.stderr.write('[spgi] ⚠ as-of 月不完整，想重下一次确认却失败了（%s）；'
                             '改用本地已有的那一份继续判\n' % exc)
    return path, asof, url, data, done


def latest_month(cache_dir):
    """官方源当前最新、且三个字段齐全的月份 "YYYY-MM"。

    抓不到 / 解析不出来一律抛 SpgiFetchError（不返回 None 掩盖故障）。
    ⚠ 「一个完整月份都没有」也算故障，同样抛：文件名自报 as-of 是某个月，表里却
    连那个月都不完整，这两件事同批发布、不该不一致（判定与豁免见
    _crosscheck_asof_month）。从前这里返回 None，把「官方发了空表」和「我们把整列
    解析丢了」混成同一个返回值 —— 而后者恰恰是本模块最安静的坏法。
    """
    path, asof, _url, data, done = _load_workbook_months(cache_dir)
    _crosscheck_asof_month(asof, done, data, path, 'latest_month')
    return _ym(done[-1]) if done else None


def update(series_dir, cache_dir):
    """把新月份追加到 series/spgi.csv 与 series/spgi_clean.csv，返回新增月份列表。

    幂等：已存在的月份跳过。任一列算不出来就抛异常，绝不写 NaN。

    ── series/spgi.csv 不可删（孤儿盘点请勿标记为可删）────────────────────────
    两个 CSV 是本函数**同一次调用里一起写的**，中间不存在「先出 raw、再清洗成
    clean」这个环节 —— 所以 spgi.csv 不是 spgi_clean.csv 的上游中间产物，
    删掉它不会「由清洗步骤重新生成」。

    没有任何 build/*.py 读 spgi.csv（画图只读 spgi_clean.csv），按「谁读它」
    盘点必然把它判成断链孤儿。但它在本模块里承担两个不可替代的职责，都是**读**：

    1. 去重台账。下面 have_raw 取自 spgi.csv，与 clean_by_month 一起构成
       「这个月是否已入库」的判据。文件没了，_read_csv 直接抛
       SpgiFetchError（见该函数：本模块只追加、不从零建库），SPGI 抓取当场失败。
    2. 官方重述体检的**唯一基线**。下面那段循环逐行拿本期 xlsx 重算 spgi.csv
       的历史月份，不一致就告警（不改历史，改写历史必须人工决定）。
       这件事 spgi_clean.csv 顶不了：它的数值是 repr 未清洗写法，且 2022 年的
       绝对值是拿 2023 年值除以官方同比反算出来的（见模块 docstring 坑 3/4），
       拿它当基线会满屏假告警。基线丢了，官方悄悄重述历史就再没人发现。

    结论：spgi.csv = 本模块的私有台账 + 重述基线，只被 fetch 读回、不进 build。
    要动它之前先把上面两条职责搬走，否则删除 = SPGI 抓取炸掉 + 永久丢失重述告警。
    """
    path, asof, url, data, done = _load_workbook_months(cache_dir)

    raw_path = os.path.join(series_dir, 'spgi.csv')
    clean_path = os.path.join(series_dir, 'spgi_clean.csv')
    # 两个 CSV 刻意提到护栏**之前**读：_crosscheck_stored_months 要拿库内已有月份
    # 当判据，而两道护栏又都必须站在 `if not done: return []` 之前（理由见它们各自的
    # docstring —— 整份表全军覆没时那句 return 会先跑掉，站在它后面的护栏等于不存在）。
    _, raw_rows = _read_csv(raw_path, COLS_RAW)
    _, clean_rows = _read_csv(clean_path, COLS_CLEAN)

    # ── 两道护栏，查的方向相反、盲区不重叠，都刻意 raise 而不是 warn ─────────
    # 前者从「库里已知该有什么」的反侧查，后者拿文件名这个独立于解析器的外部判据查。
    _crosscheck_stored_months(done, data, asof, [r[0] for r in raw_rows], path)
    _crosscheck_asof_month(asof, done, data, path, 'update')

    if not done:            # 上一行已经把这种情况判成故障，留着只是兜底
        return []

    have_raw = {r[0] for r in raw_rows}
    clean_by_month = {r[0]: r for r in clean_rows}

    # ---- 先做重述体检：用工作簿重算已有月份，和 CSV 比。不一致只告警，不改历史。
    for r in raw_rows:
        p = _parse_ym(r[0])
        if p not in data or not all(k in data[p] for k in ('billed_yoy', 'adv', 'adv_yoy')):
            continue
        recalc = (_fmt15(_g15(data[p]['billed_yoy'] * 100)),
                  _fmt15(_g15(data[p]['adv'])),
                  _fmt15(_g15(data[p]['adv_yoy'] * 100)))
        if tuple(r[1:4]) != recalc:
            sys.stderr.write('[spgi] ⚠ %s 官方数值与库内不一致（库内 %r → 本期 xlsx %r），'
                             '疑似重述；本模块不改历史，请人工确认\n'
                             % (r[0], tuple(r[1:4]), recalc))

    new_raw, new_clean, added = [], [], []
    for p in done:
        key = _ym(p)
        if key in have_raw and key in clean_by_month:
            continue
        d = data[p]

        # billed_issuance_index 需要上一年同月的指数；BASE_YEAR 是约定基数 100。
        prev = _prev_year(p)
        prev_key = _ym(prev)
        if prev[0] <= BASE_YEAR:
            base = 100.0
        else:
            prev_row = clean_by_month.get(prev_key)
            base = None
            if prev_row and prev_row[COLS_CLEAN.index('billed_issuance_index')].strip():
                base = float(prev_row[COLS_CLEAN.index('billed_issuance_index')])
            else:
                for line in new_clean:                    # 可能是本轮刚追加的
                    if line.startswith(prev_key + ','):
                        base = float(line.split(',')[COLS_CLEAN.index('billed_issuance_index')])
                        break
            if base is None:
                raise SpgiFetchError(
                    '%s 算不出 billed_issuance_index：缺少上年同月 %s 的指数值。'
                    '同比链断了就整条序列失真，宁可报错。' % (key, prev_key))

        if key not in have_raw:
            new_raw.append(','.join([key,
                                     _fmt15(_g15(d['billed_yoy'] * 100)),
                                     _fmt15(_g15(d['adv'])),
                                     _fmt15(_g15(d['adv_yoy'] * 100))]))
        if key not in clean_by_month:
            new_clean.append(','.join([key,
                                       _repr(d['adv']),
                                       _repr(d['adv_yoy'] * 100),
                                       _repr(d['billed_yoy'] * 100),
                                       '0',
                                       _repr(base * (1 + d['billed_yoy']))]))
        added.append(key)

    if new_raw:
        _append_csv(raw_path, new_raw)
    if new_clean:
        _append_csv(clean_path, new_clean)

    # ── 发布日入台账（页面抬头「官方发布于」的唯一来源）────────────────────────
    # 只记 as-of 那一个月：xlsx 是全年重发，同一份文件里更早的月份是历次旧版各自发的，
    # 把本期的 "Published on" 安到它们头上，等于给旧月份印一个偏晚的假日期。
    # 补历史月份的发布日只能各自去找当期那份 xlsx（见 _guess_cdn_urls）。
    #
    # 条件是「as-of 月确实躺在两个 CSV 里」而不是「本轮新增了它」：数据落库与台账落库
    # 是两件事，某次跑只成了前一件（xlsx 页脚缺行、台账写失败）时，该月再也进不了
    # added 分支，那半句话就永久缺席。每次跑都对 as-of 月重申一遍，record 同键覆盖。
    asof_key = _ym(asof)
    if asof_key in added or (asof_key in have_raw and asof_key in clean_by_month):
        pub, where = published_on(path)
        if pub:
            try:
                _source_dates().record(series_dir, 'spgi', asof_key, pub,
                                       '%s（源文件 %s）' % (where, os.path.basename(url)))
            except Exception as exc:                        # noqa: BLE001
                # 台账写不进去不该把已经落库的月份回滚成 FAIL —— 数值是对的，
                # 代价只是页面抬头少半句话。响一声，让人回头补。
                sys.stderr.write('[spgi] ⚠ %s 发布日 %s 写台账失败：%s\n'
                                 % (asof_key, pub, exc))
        else:
            sys.stderr.write('[spgi] ⚠ %s 的 xlsx 里找不到 "Published on M/D/YYYY"，'
                             '本页不印「官方发布于」——宁可缺这半句，也不拿下载时间冒充\n'
                             % asof_key)

    if added:
        sys.stderr.write('[spgi] 新增 %d 个月：%s（源 %s）\n'
                         % (len(added), ', '.join(added), url))
    return added


if __name__ == '__main__':
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    cache = os.path.join(root, 'cache')
    print('latest_month =', latest_month(cache))
    print('update       =', update(os.path.join(root, 'series'), cache))
