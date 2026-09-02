# -*- coding: utf-8 -*-
# 本文件的模块 docstring 用 raw 串：口径坑 5 里要原样引用正则 `\d{4}`，
# 普通串会被 Python 3.12 判成非法转义并告警。
r"""TMX Group (TSX:X) 月度经营指标 —— 无人值守抓取。

本模块把**两条互相独立的官方序列**写进同一张 series/tmx.csv：
  · 加拿大现货（TSX / TSXV / Alpha / Alpha-X & DRK）—— 来自 TMX IR 的月度新闻稿，2021-08 起；
  · Montréal Exchange 衍生品（ADV / 月末 OI / 分产品）—— 来自 m-x.ca 月度 xlsx，2002-01 起。
两条腿的起点差 19 年、每月到货时间也差几天，**任何时候都可能出现「MX 已有 7 月、现货还停在
6 月」的正常状态**（写这份代码的 2026-08-06 就正是如此）。所以本模块从不用一个标量代表
「最新月」，CSV 里 MX 那半边比现货多出一行也不是解析失败。

另有一张季度表 series/tmx_box_q.csv（BOX 期权），原因见文末「BOX：为什么只有季度」。

━━ 数据源 ━━

源 1（现货，主源）TMX Group Consolidated Trading Statistics 新闻稿，走 IR 站的 Q4 feed：
    https://investors.tmx.com/feed/PressRelease.svc/GetPressReleaseList
        ?LanguageId=1&bodyType=1&pressReleaseDateFilter=3&categoryId=
        &pageSize=<N>&pageNumber=0&tagList=trading-statistics&includeTags=true
        &year=-1&excludeSelection=1
  · `bodyType=1` 是关键：0 只给标题，1 才把整篇新闻稿 HTML 放进 `Body` 字段，表格就在里面。
  · `tagList=trading-statistics` 是官方自己打的分类标签，2014 至今每年都命中，
    比按标题 grep 稳（标题历史上出现过 "(Revised)" 等后缀）。
  · **必须显式设桌面浏览器 User-Agent**，否则 403。urllib 默认 UA（Python-urllib/3.x）与
    curl 默认 UA 都被 Cloudflare 拒掉 —— 侦察稿写的「纯 urllib 200」不准确，复核实测才发现。
    带上 UA 之后不需要 curl_cffi、不需要浏览器、不需要登录态，cron 可跑。
  · 同域的 HTML 落地页 investors.tmx.com/English/News--Events/... 走另一套 Cloudflare 规则，
    普通客户端 403，**不要去解析落地页**，那会平白引入一个 curl_cffi 依赖。
  · 每条目的 `PressReleaseDate` 就是官方自述的发布日，精确到日 —— 这是 TMX 相对 CME/HKEX
    的优势，source_dates.csv 直接用它，不必靠 Last-Modified 猜（见 _record_source_dates）。

源 2（MX 衍生品，主源）Montréal Exchange 月度统计 xlsx：
    落地页 https://www.m-x.ca/en/trading/data/monthly-volumes-and-open-interest
    直链   https://www.m-x.ca/f_stat_en/{YY}{MM}_stats_en.xlsx   （2026 年 7 月 -> 2607）
  · **裸 urllib 即 200**，无 UA 要求 —— 和源 1 相反。两条源的要求不同，不要一刀切。
  · 未发布的月份返回干净的 **404**（body 是 HTML），已发布的返回
    application/vnd...spreadsheetml.sheet，可以放心用状态码当「发了没有」的判据
    （不像 Cboe 那样 403/404 混淆）。**但 404 之外的错误一律不能当成「没发」** ——
    本次回补 295 个月，2209 那一份撞上一次 HTTP 502，重试即成功；把 502 当 404 会静默丢一个月。
  · 落地页永远只挂最近 3 期直链，但模板对任意历史月都成立：最新月由落地页给（官方自己
    维护的「当前最新」指针），历史回补直接按模板拼。
  · 每份文件**只覆盖它自己那一个月**，所以回补 N 个月就要下 N 个文件
    （不像 CME/Cboe 一份带全历史）。

源 3（BOX，季度）TMX 季度 MD&A PDF，走 IR 站的财报 feed 发现：
    https://investors.tmx.com/feed/FinancialReport.svc/GetFinancialReportList
        ?LanguageId=1&bodyType=0&year=-1&pageSize=100&pageNumber=0&excludeSelection=1
  · 文件名毫无规律（`TMX-MDA-Q1-2025-EN.pdf` / `TMX-Group-Limited-YE-2025-MD-A_EN_Final.pdf`），
    **必须靠 feed 发现，不能拼 URL**。
  · 2024Q4 起 MD&A 单独作为一份文件挂在 `DocumentCategory == 'mdna'`；更早的季度里 MD&A
    是包在 `'tenq'`（Report To Shareholders）那份大 PDF 里的，所以 mdna 缺席时回落到 tenq。
  · PDF 有完整文字层，pymupdf 直接抽得出，**不需要 OCR**。

源 4（两条指数列的历史，一次性回补）TMX Money 的行情 GraphQL：
    POST https://app-money.tmx.com/graphql
        operationName=getTimeSeriesData，variables={symbol:'^TSX'|'^JX', freq:'month', …}
  · 本机实测 200、**无需登录**，`^TSX`（S&P/TSX Composite）与 `^JX`（S&P/TSX Venture
    Composite）各回 298 条、2001-12 起，`close` 就是月末收盘点位。
  · 与 CSV 里 CTS 解析出来的 120 格重叠值**逐位相同**（差 <0.005），所以这条源不是
    「另一套数」，只是同一套数更长的那一段。
  · **它没有官方契约文档**（前端自用端点，query 形状可能随改版而变），所以它
    **不进 update()**：只由 `backfill_index()` 手工调用，且每次都拿重叠月做断言，
    对不上就抛异常而不是静默写空。日常 cron 的指数列仍然只来自 CTS 表格。
    这与 BOX 的 `backfill_box()` 是同一条摆放原则：一次性、易变的路子不摆进无人值守链路。
    重跑（幂等，只填空不覆盖）：
        python3 -c "import sys;sys.path.insert(0,'fetch');import tmx;\
                    tmx.backfill_index('series','cache/tmx')"

源 5（现货 12 列 2015-01~2021-07，一次性回补，**不在本模块**）CIRO
    『Report of Marketshare by Marketplace (Historical 2015–Present)』xlsx。
  · 落地页 https://www.ciro.ca/markets/reports-statistics-and-other-information/
        reports-market-share-marketplace（WAF 会间歇 403，要退避重试）
  · 代码在 `build/basefill/tmx_ciro_2015.py`，**不写进本模块**：那是监管方 CIRO 的
    另一个源、另一种版式，而且 2015-01~2021-07 这个洞补完就永远关上了（2021-08 起
    CTS 正文自己有表）。口径差与接缝台阶见口径坑 16。

拿不到的：`www.tmx.com` / `www.tsx.com` 对本网络整体返回 CloudFront 403（curl / urllib /
nscurl / curl_cffi / 本机真实 Chrome 全部 403，只有搜索引擎 UA 能过）。2014-12~2021-07 的
**TMX 自报**现货明细只存在于 `tmx.com/en/resource/<id>` 里，那 80 个月至今**线上没有
正当通道**（2026-08-18 复核：feed 条目里的 `DocumentPath` 指向 s21.q4cdn.com 且可下载，
但 2021-08 及更早那些 PDF 正文只有百分比叙述、**一张统计表都没有**，表格落在 PDF 内嵌的
`tmxgroup2019ir.q4web.com/resource/en/<id>` 上，该域整段 404；换成 www.tmx.com /
www.tsx.com 的同号 resource 仍是 403，Wayback 也没存成序列）。
冒充 Googlebot 能绕开，但那是对一条明确针对本出口 IP 的封禁规则做规避、且冒用搜索引擎身份，
**不写进生产代码**。
⇒ 这 80 个月改由**监管方**补：CIRO 的同口径月报回到 2015-01（源 5），已入库。
   它不是 TMX 自报值的替代品，是另一把尺子 —— 差多少、在哪一列，口径坑 16 逐列量过。

━━ 发布节奏 ━━

现货 CTS 新闻稿：数据月结束后第 2–8 天，中位数第 5 天；139 期里最坏一次是第 14 天
（2025-04 那期）。实测日分布：第2日×3 第3日×27 第4日×23 第5日×25 第6日×45 第7日×13
第8日×2 第14日×1。⇒ 闸门次月 2 日开，到 15 日仍拿不到才算异常。

MX xlsx：次月第 1–4 个工作日，几乎都在 18:01 GMT 整点批量上传。2026 年实测：
    2026-01 -> 02-02   2026-02 -> 03-02   2026-03 -> 04-01   2026-04 -> 05-01
    2026-05 -> 06-01   2026-06 -> 07-02   2026-07 -> 08-04
⚠ **Last-Modified 不能当权威发布日**：历史档被整批重传过，2019-01 那份的 Last-Modified
是 2022-07-23（一个 2019 年的文件带 2022 年时间戳），2025-06/2025-08/2025-12 也各被重传过一次。
xlsx 里也没有任何「发布于」字段（Cover Page 只写月份，docProps 的 dcterms:modified 同样是
重传时刻）。所以 **source_dates.csv 里 tmx 的发布日一律取自 CTS 新闻稿的 PressReleaseDate**，
MX 那半边宁可缺席也不猜 —— MX 比现货早到的那几天，页面抬头那半句就该留白。

━━ CSV 每一列的确切口径（下游换算全靠这一节）━━

**通则：本模块只忠实入库官方披露的原始值，不做任何单位换算、不做日均折算、不做币种换算。**
张数就是张数、金额就是金额、月总量就是月总量。定基名义额那套换算是 build 层的事。

series/tmx.csv，一行一个自然月，`month` 为 `YYYY-MM`。

【现货：来自 CTS 新闻稿的「当月」栏，2021-08 起才有】
口径统一说明 —— 全部是**当月总量（流量），不是日均**；**加拿大境内交易所内成交**；
Volume 的单位是**股**（不是张、不是手），Value 的单位是**加元 C$**（不是美元），
Transactions 的单位是**笔**。官方不说明是否双边计数，但四家相加恒等于合计
（见口径坑 5），说明同一笔交易在这套口径里只被计一次，即**单边计数**。
  tmx_all_volume_shares       全部 TMX 股票市场当月成交股数（股）= TSX+TSXV+Alpha+Alpha-X/DRK
  tmx_all_value_cad           同上口径当月成交额（加元）
  tmx_all_transactions        同上口径当月成交笔数（笔）
  tsx_volume_shares           Toronto Stock Exchange 当月成交股数（股）
  tsx_value_cad               TSX 当月成交额（加元）
  tsx_transactions            TSX 当月成交笔数（笔）
  tsx_composite_close         S&P/TSX Composite 月末收盘点位（指数点，**存量**不是流量；
                              官方印在同一张表里，不是我们外接的行情）
  tsxv_volume_shares          TSX Venture Exchange 当月成交股数（股，**含 NEX**，表下脚注
                              `*Includes NEX`，全期一致）
  tsxv_value_cad              TSXV 当月成交额（加元）
  tsxv_transactions           TSXV 当月成交笔数（笔）
  tsxv_composite_close        S&P/TSX Venture Composite 月末收盘点位（指数点，存量）
  alpha_volume_shares         TSX Alpha Exchange 当月成交股数（股，**不含** Alpha-X / Alpha DRK）
  alpha_value_cad             Alpha 当月成交额（加元）
  alpha_transactions          Alpha 当月成交笔数（笔）
  alphax_drk_volume_shares    Alpha-X + Alpha DRK 合计当月成交股数（股）。官方 2023-11 起
                              才单列，之前是真·天然为空（不是解析失败）
  alphax_drk_value_cad        Alpha-X + Alpha DRK 当月成交额（加元）
  alphax_drk_transactions     Alpha-X + Alpha DRK 当月成交笔数（笔）

【MX 衍生品：来自 m-x.ca xlsx，2002-01 起】
口径统一说明 —— 单位一律是**合约张数（contracts）**，不是千张、不是名义金额。
交易所成交量按**单边**计（一张合约成交计 1，不把买卖两边各计一次），与 CME/Cboe 同制。
`_adv_` 是**官方 `EN ADV` 表直接给的日均**（已按该产品自己的交易日历除过），
`_volume_` 是**当月总量（流量）**，`_oi_` 是**月末未平仓（存量，时点数）**。
⚠ 与仓库其它家的单位差异：`cme.csv` 存的是**千张**（`adv_*_kcontracts`），
`hkex.csv` 存的是**张**（`derivatives_adv_contracts`）。本表与 hkex 同制，
进横截面时对 cme 要除 1000，对 hkex 直接可比 —— 换算在 build 层做。
  mx_volume_contracts             MX 全所当月成交合约数（张，xlsx 的 GRAND TOTAL 行）
  mx_adv_contracts                MX 全所日均成交（张/日，GRAND TOTAL）
  mx_oi_contracts                 MX 全所月末未平仓（张，GRAND TOTAL）
  mx_adv_futures_contracts        期货合计 ADV（张/日，Total Futures 行）
  mx_adv_options_contracts        期权合计 ADV（张/日，Total Options 行）
  mx_adv_stir_futures_contracts   短端利率期货合计 ADV（张/日，BAX+CRA+COA/Others 的 Total）
  mx_adv_bax_contracts            BAX（三月期加拿大银行承兑汇票期货）ADV（张/日）
  mx_adv_cra_contracts            CRA（三月期 CORRA 期货）ADV（张/日）
  mx_adv_bond_futures_contracts   国债期货合计 ADV（张/日，CGZ+CGF+CGB+LGB 的 Total）
  mx_adv_cgb_contracts            CGB（10 年期加拿大国债期货，MX 旗舰合约）ADV（张/日）
  mx_adv_cgf_contracts            CGF（5 年期）ADV（张/日）
  mx_adv_cgz_contracts            CGZ（2 年期）ADV（张/日）
  mx_adv_index_futures_contracts  股指期货合计 ADV（张/日，Index Futures 的 Total，**纯期货**）
  mx_adv_sxf_contracts            SXF（S&P/TSX 60 标准期货）ADV（张/日）
  mx_adv_index_options_contracts  股指期权合计 ADV（张/日，SXJ+SXO+SXV；2016 年后长期为 0，
                                  那是真的没成交，不是解析失败）
  mx_adv_equity_options_contracts 个股期权 ADV（张/日）
  mx_adv_etf_options_contracts    ETF 期权 ADV（张/日）
  mx_adv_share_futures_contracts  个股期货 ADV（张/日）
  mx_oi_bax_contracts             BAX 月末未平仓（张）
  mx_oi_cra_contracts             CRA 月末未平仓（张）
  mx_oi_stir_futures_contracts    短端利率期货合计月末未平仓（张）
  mx_oi_bond_futures_contracts    国债期货合计月末未平仓（张，CGZ+CGF+CGB+LGB 的 Total，
                                  四条之和逐月精确等于它，296 个月实测零残差）
  mx_oi_cgb_contracts             CGB 月末未平仓（张）
  mx_oi_cgf_contracts             CGF 月末未平仓（张）
  mx_oi_cgz_contracts             CGZ 月末未平仓（张）
  mx_oi_sxf_contracts             SXF 月末未平仓（张）
  mx_oi_equity_options_contracts  个股期权月末未平仓（张）
  mx_oi_etf_options_contracts     ETF 期权月末未平仓（张）

【交易日数：这两列是**我们推出来的**，官方没有这一列】
  trading_days_rates    当月**利率/债券类**产品的交易日数，= round(CGB 月度量 / CGB ADV)
  trading_days_equity   当月**股票类**产品的交易日数，= round(个股期权月度量 / 个股期权 ADV)
两列存在的理由与不可互换的理由见口径坑 1；推法与除零处理见口径坑 2。
现货那几家官方只给月度总量，**日均要 build 层用 trading_days_equity 现算**
（官方新闻稿里印的 Daily Average 只保留到 0.1 million，别抄，见口径坑 9）。

series/tmx_box_q.csv，一行一个自然季，`quarter` 为 `YYYY-Qn`。
  box_volume_mncontracts        BOX 当季成交合约数（**百万张**，官方就印这个单位；
                                是**当季总量**不是 ADV —— TMX 从不披露 BOX 的日均）
  box_equity_options_share_pct  BOX 在全美股票期权成交中的市占率（%，官方只给整数）
  box_revenue_cadmn             BOX 当季收入（百万加元）
  box_revenue_usdmn             BOX 当季收入（百万美元，官方自己按下面那个汇率折的）
  usdcad_avg                    当季 USD-CAD 平均汇率（1 美元兑多少加元）

━━ 口径坑（按踩坑概率排序）━━

1. **MX 一个月里跑两套交易日历，`GRAND TOTAL` 的 ADV ≠ 总量 ÷ 单一日数。**
   利率/债券类产品跟债市日历、股票类产品跟股市日历，每年 **9 月和 11 月**必然分叉
   （真相与和解日 9/30、国殇日 11/11 债市休市而股市照开）。实测 2025-09：CGB/CGF/CGZ/CRA
   全部 20.000 天，而 equity options / ETF options / share futures / SXF 全部 21.000 天，
   `GRAND TOTAL` 被逼出 **20.499** 这种非整数。2025-11 同样是 19 vs 20。
   ⇒ ADV 一律**直接读 `EN ADV` 表**，绝不用总量自己除；交易日数存两列，不共用一列。

2. **`trading_days_*` 是反推的，两个坑。**
   (1) 停用合约会 **0/0**：2026 年的 BAX 是 vol=0 且 adv=0，拿它当基准直接除零。
       所以基准合约按「当月成交量足够大」的候选表逐个降级（利率侧 CGB→债券期货合计→
       STIR 期货合计；股票侧 个股期权→ETF 期权→SXF），全都不够活跃才写空。
   (2) **绝不能用 `GRAND TOTAL` 反推** —— 它跨两套日历，2025-09 会推出 20.499 这种非整数。
   推出来的比值必须在整数附近（容差 0.05），否则说明基准合约本身跨了日历，宁可报错。

3. **CTS 每张月度表的下半部分紧跟一个 `Daily Averages` 子块，行标签在同一张 `<table>` 里
   各出现两次。** 这是本模块最危险的一个坑，因为它**静默且结果貌似合理**：
   2026-06 的 all/transactions 若取到第二次出现的 `Transactions`，会解析成 **1,540,000**
   （日均笔数），而正确值是 **33,879,991**（当月总笔数），差 22 倍，两个数在图上都不离谱。
   ⇒ 解析每张表时遇到 `Daily Averages` 行标签立刻**停止本表取数**（本模块的做法），
   同名行只认第一次出现。

4. **小节标题不能用正则抽 —— 必须用真 HTML parser 按文档顺序遍历。**
   59 期里有 6 期（2022-03/04/07/08/10、2023-06）的 `All TMX Equities Marketplaces`
   标题嵌在 `<div class="q4default">` 内层的 `<div role="heading">` 里：
       <div class="q4default"><p>…正文…</p>
         <div class="wcag-arialevel-3" role="heading" aria-level="3">
           <b><u>All TMX Equities Marketplaces</u></b><b>&nbsp;*</b></div>
   任何 `<div[^>]*>(.*?)</div>` 式的非贪婪正则都会从外层匹到内层的 `</div>`，
   把整个小节吞掉 → 该期 All TMX 整段丢失。本模块用 bs4+lxml 遍历，
   并且**只走非 table 的文本节点**（见 _text_before_tables），天然不受嵌套影响。

5. **剥完标签会冒出词内空格，标题和表头都会中招。**
   官方把文字拆进不同 inline 标签，`TSX Venture Exchange` 剥完变成 `TSX Venture Ex change *`。
   更阴的是**表头单元格里的月份也会被拆**：2022-03 与 2022-05 两期里，TSX Alpha 那张表的
   表头是 `Ma rch 2022` / `Ma y 2022`（"Ma" 和其余部分在不同的 <strong> 里）。
   我第一版按 `^[A-Z][a-z]+ \d{4}$` 匹配表头，这两期的 Alpha 表被整张判成「不是月度表」
   而丢弃 —— 而 Alpha 一丢，恒等式立刻不成立。**这一条两份侦察稿都没写到。**
   ⇒ 一切标题 / 行标签 / 表头匹配前先把空白**全删**（`_squash`）再比，
   而且**认不出小节的表格一律抛异常**，绝不让它继承上一节的归属。

6. **CTS 里 YTD 表和月度表长得一模一样**，只有表头第一列不同
   （`June 2026 / May 2026 / June 2025` vs `2026 / 2025 / % Change`），而且每个小节后面
   都紧跟一张 YTD 表。抓错就是把年初至今当成当月。⇒ 只认表头第一个数据列形如
   `<月名><4 位年>` 的表，其余（YTD 表、正文末尾那张排版换过三种的脚注表）一概丢掉。

7. **现货表格 2021-08 才进新闻稿正文。** 2021-07 及更早的 `Body` 只有一段摘要加一条
   `/resource/en/<id>` 链接，`<table>` 计数为 0。2026-08-18 复核：feed 共 140 期，
   有表 60 期（2021-09-08 → 2026-08-07 发布）、无表 80 期（2015-01-06 → 2021-08-05），
   边界干净无交叉。「这一期没有表格」是**明确可识别的状态**（本模块直接跳过该期），
   不是解析失败，更不能解析出一堆空值混进 CSV。
   ⚠ 但「可识别」这三个字得靠判据兑现，不能靠一句 continue 蒙混：跳过的条件是
   **标题自报的数据月早于 SPOT_START**，不是「正文里没找到表」。二者当年是同一个
   分支，于是「官方改了表头写法」和「这一期本来就没表」在日志里长得一模一样 ——
   现在由 _crosscheck_headline_month 把它们分开，后者放行，前者抛。
   ⇒ 所以 `SPOT_START` 不是保守设置，是 feed 本身的上限；2015-01~2021-07 由 CIRO 补
   （源 5 / 口径坑 16），**不是**由本模块补。

8. **MX xlsx 里 `Total` 这个行名在 5 个 section 各出现一次**（STIR 期货 / STIR 期权 /
   债券期货 / 股指期货 / 股指期权），且**行号逐年漂移**：`GRAND TOTAL` 在 2015/2019 档是
   第 54 行、2022-09 档是第 45 行、2026-07 档是第 56 行（不是单调的，中间新增又删过
   CDR / CEFs / BCS 等产品行）。⇒ 必须先按 section 标题定位再取 `Total`，行号一个都不能写死。
   （列块起点倒是稳定的：2015/2019/2026 三档的 `MONTHLY VOLUME`/`YEAR-TO-DATE VOLUME`/
   `MONTH END OPEN INTEREST`/`TRANSACTIONS` 都在第 2/7/11/16 列 —— 侦察稿说「列位置逐年变」
   是错的，但「按标题找列块」本身仍是对的写法，本模块照做。）

9. **新闻稿印的 Daily Average 只到 0.1 million，别抄。** 官方印 "490.7 million"，
   而 10,796,096,148 / 22 = 490.7316 million。抄那个印出来的数会在图上留下量化台阶。
   ⇒ 只存官方给的总量，日均在 build 层用 trading_days_equity 现算。

10. **m-x.ca 的 xlsx 是活文件、会被重传订正；CTS 新闻稿是冻结快照。两者会分叉。**
    59 个重叠月实测：月度合约数 55/59 逐位相同，OI 45/59 逐位相同。
    最大分歧 2022-09：xlsx 12,681,668 vs CTS 11,923,849，差 757,819（6.36%），
    而且**是 CTS 错**：2210 档 xlsx 的「上月」对照列同样写 12,681,668（两份 xlsx 互证），
    CTS 自己的 YTD（112,033,825）比 xlsx 逐月相加（112,791,644）正好少这 757,819。
    那一期标题还带着 "(Revised)" —— 官方修订反而修出了这个口子。
    ⇒ **MX 的数一律以 m-x.ca xlsx 为准**，CTS 里的 MX 两行不入库，只在 crosscheck() 里当
    交叉校验用。CSV 里没有任何一列来自 CTS 的 MX 小节。

11. **BAX → CRA 是结构性断点，不可连比。** 2015-01：BAX 成交 2,729,283 张、CRA 为 0；
    2026-07：BAX 为 0、CRA 成交 4,221,855 张。短端利率基准从 BA 迁到 CORRA，合约整体换代。
    `mx_adv_stir_futures_contracts` 这条合计线跨越这段是连续的，但
    **`mx_adv_bax_contracts` 与 `mx_adv_cra_contracts` 在断点两侧不能画成一条线**。

12. **老档 sheet 名带月份**：2015-01 档叫 `Jan 2015 EN` / `Jan 2015 EN ADV`，
    2026 档就叫 `EN` / `EN ADV`。按后缀匹配，不要写死。

13. **不要给 2026-08 画并表断点。** 2026-08-02 完成的是收购 **Cboe Australia**，
    **不是 Cboe Canada** —— 后者尚未交割，Q2/26 MD&A 原文说两者分别交割、澳洲预计 Q3/26。
    Cboe Australia 完全不碰加拿大现货成交量。真正该打断点的是 Cboe Canada 的交割月，
    等 feed 里出现 "TMX Group Completes Acquisition of Cboe Canada" 再说（BREAKS 常量现为空）。

14. **BOX 的同一个季度在不同报告里印的精度不同。** Q2/26 那份 MD&A 把八个季度都印成一位
    小数（234.7 / 259.3 / 247.1 …），而 Q1/26、Q3/25、YE-2025 那几份印的是整数
    （259 / 247 / 236 …）。同一个数、不同印法。⇒ BOX 回补必须**从新往旧**处理，
    让最新（最精确、且已含官方订正）的那份先落地，「只填空不覆盖」自然锁住它。

15. **BOX 表的引子句和汇率行标签都变过。** 引子句在 YE-2024/YE-2025 里少了 "share" 一词
    （"the equity option market over the last eight quarters"），锚点只能取到 "market" 为止；
    季度数在早期是 5 个而不是 8 个（2023Q1 那份写 "since consolidation"）。
    汇率行 2023 年前叫 `Average CAD-USD FX rate`、之后叫 `Average USD-CAD FX rate`，
    **数字含义没变**（都是 1 美元兑多少加元，用 27.7 CAD / 1.35 = 20.5 USD 验过）。

16. **现货 12 列在 2021-08 换源：左边 CIRO、右边 TMX 自报。两把尺子不完全一样。**
    2015-01~2021-07 由 `build/basefill/tmx_ciro_2015.py` 从 CIRO 的
    "All Trade All Listing Total" 行写入，2021-08 起仍是 CTS 新闻稿（本模块）。
    60 个重叠月（2021-08~2026-07）逐月对过，比值 = TMX 自报 ÷ CIRO 的中位数：

        tsxv_transactions        1.00000（59/60 逐位相同）   ─┐
        alpha_volume_shares      1.00000（49/60 逐位相同）    │ 同一套数，
        alpha_transactions       1.00000（49/60 逐位相同）    │ 换源无台阶
        tsx_transactions         1.00000（20/60 逐位相同）    │
        tsxv_value_cad           1.00000                      │
        tsxv_volume_shares       0.99989                     ─┘
        tsx_value_cad            1.00162  ┐
        alpha_value_cad          1.00249  │ 有系统性水平差，60 个月里
        tmx_all_value_cad        1.00165  │ **一次逐位相同都没有**
        tmx_all_volume_shares    0.99131  │
        tsx_volume_shares        0.98683  ┘（TMX 自报比 CIRO 低约 1.3%）

    **不是「含不含 intentional cross」**：2021-08 TMX 自报 TSX 成交 6,324,849,035 股，
    夹在 CIRO 的 Non-Cross 6,185,693,468 与 All Trade 6,429,009,165 之间；把它写成
    「Non-Cross + f × 大宗对敲」去解 f，60 个月的 f 在 0.09~0.73 之间乱跳（中位 0.54），
    成交额那边的 f 甚至常年 >1。⇒ 是两家各自的**统计口径**不同，不是可加减的一块。
    2022-01 六列比值同时探底（tsx_volume 0.96715 / tsx_value 0.98143 /
    tsx_transactions 0.98752 / tsxv_volume 0.99783），像是某一方对该月做过重述。

    实际接缝台阶（CIRO 2021-07 → TMX 2021-08，把口径差从真实环比里剥出来）：
        tsx_volume_shares −1.62%、tmx_all_volume_shares −0.98%、
        tsx_value_cad +0.15%、tmx_all_value_cad +0.15%、alpha_value_cad +0.17%，
        其余 7 列 |台阶| ≤ 0.11%（笔数三列与 tsxv_value 恰为 0.00%）。
    ⇒ build/specs/tmx.py 只给**台阶 ≥0.5% 的那两列**画 2021-08 断点线；其余列画了
      等于说假话（它们跨这个月是可比的）。

    ⚠ **换源的方向只能是「往左补」**。CIRO 的历史报每月更新、含最新月，理论上能整段
    覆盖 12 列 —— **别这么做**：那等于把 TMX 官方新闻稿印出来的数换成第三方重算值，
    还要把 `tmx_all_*` 从官方披露列降级成我们自己的加总。本仓的做法是
    「已有值永不覆盖」（`_merge`），CIRO 只填 2021-08 之前的空格。

17. **CIRO 没有「TMX 合计」列，回补段的 `tmx_all_*` 是本仓加总出来的。**
    CIRO 的 `All Traded Marketplaces` 是全加拿大（含 CSE / Nasdaq CXC / MATCHNow / NEO…），
    不是 TMX 集团。所以回补段 `tmx_all_* = TSX + TSXV + Alpha`
    （Alpha-X / Alpha-DRK 在 CIRO 里同样 2023-11 才有值，2015~2021 段天然不参与）。
    这与 2021-08 起 TMX 自报的恒等式（`_cts_to_row` 每期都验）是同一条式子，
    但**加总方是我们**，写进图注时要说清楚。

━━ 依赖 ━━ openpyxl（读 xlsx）、beautifulsoup4 + lxml（解析新闻稿 HTML，见口径坑 4）、
pymupdf（抽 MD&A 文字层，只有 BOX 那条路用）。不依赖 pandas。

━━ BOX：为什么只有季度 ━━
起点线索点名要 **BOX 期权 ADV**（BOX 属北美多挂牌期权池、与美股期权同规格 100 股/张，
跨所张数直接可比）。**月度 ADV 拿不到，这是本模块最大的一个「没做到」。**
检索路径：(a) TMX IR 新闻稿 feed 全量 139 期，CTS 月报里只有 TSX/TSXV/Alpha/MX 五个小节，
没有 BOX；(b) TMX 财报 feed 全部 68 期，BOX 只出现在季度 MD&A 的「最近八个季度」表里，
给的是**当季总成交张数（百万张）+ 市占率 + 收入**，没有 ADV 也没有月度拆分；
(c) BOX 自己的站点不发月度统计。
⇒ 📌 **未找到 BOX 月度 ADV** —— 能拿到的最细粒度就是 tmx_box_q.csv 那四列季度数。
要折算 ADV 必须自己乘除交易日数，那是 build 层的事，本模块不替它做。
另注：2026-07-30 宣布 BOX 与 MEMX 合并成 MEMX Group，MD&A 说预计 **2027 下半年**交割
（不是"马上"），所以 BOX 还会在 TMX 季度披露里出现约六个季度，但不值得为它建长序列。
"""

import csv
import json
import os
import re
import urllib.error
import urllib.request
from datetime import datetime

import openpyxl
from bs4 import BeautifulSoup, NavigableString, Tag

# ── 源 1：现货新闻稿 feed ─────────────────────────────────────────────────
FEED_URL = ('https://investors.tmx.com/feed/PressRelease.svc/GetPressReleaseList'
            '?LanguageId=1&bodyType=1&pressReleaseDateFilter=3&categoryId='
            '&pageSize={n}&pageNumber=0&tagList=trading-statistics'
            '&includeTags=true&year=-1&excludeSelection=1')
# 现货表格最早出现的月份（更早的新闻稿正文里没有 <table>，见口径坑 7）。
# **这是 feed 本身的上限，不是保守设置**：2026-08-18 复核 140 期，无表的 80 期与有表的
# 60 期边界干净无交叉。2015-01~2021-07 由 build/basefill/tmx_ciro_2015.py 从 CIRO 补
# （源 5 / 口径坑 16），本模块一行都不往那边伸手 —— 换源的边界就钉在这个常量上。
SPOT_START = '2021-08'

# ── 源 2：MX 月度 xlsx ────────────────────────────────────────────────────
MX_LANDING = 'https://www.m-x.ca/en/trading/data/monthly-volumes-and-open-interest'
MX_TMPL = 'https://www.m-x.ca/f_stat_en/{yy}{mm}_stats_en.xlsx'
# 2002-01 是真实起点：0201 返回真 xlsx，0101 干净 404（实测）
MX_START = '2002-01'

# ── 源 3：BOX 季度 MD&A ───────────────────────────────────────────────────
FINREP_URL = ('https://investors.tmx.com/feed/FinancialReport.svc/GetFinancialReportList'
              '?LanguageId=1&bodyType=0&year=-1&pageSize=100&pageNumber=0'
              '&excludeSelection=1&reportSubTypeList=&reportTypeList=')

# ── 源 4：TMX Money 指数月线（只给 backfill_index()，不进 update()，见源 4 那段）──
MONEY_URL = 'https://app-money.tmx.com/graphql'
# 只要 dateTime 与 close 两个字段：多要一个字段就多一分随前端改版而失效的机会。
MONEY_QUERY = (
    'query getTimeSeriesData($symbol: String!, $freq: String, $interval: Int, '
    '$start: String, $end: String) {\n'
    '  getTimeSeriesData(symbol: $symbol, freq: $freq, interval: $interval, '
    'start: $start, end: $end) {\n'
    '    dateTime\n    close\n  }\n}')
# (行情代码, CSV 列)。^TSX = S&P/TSX Composite，^JX = S&P/TSX Venture Composite。
MONEY_SPEC = (('^TSX', 'tsx_composite_close'), ('^JX', 'tsxv_composite_close'))
# 两条指数的真实源底（实测各回 298 条，最老 2001-12）。
INDEX_START = '2001-12'
# 回补前必须先在重叠月上验一遍：少于这么多个月能对照就不敢写（说明 CSV 或端点不对劲）。
INDEX_MIN_OVERLAP = 24
# 指数点位官方就印两位小数，重叠月实测差 0（<0.005）。超过这个容差 = 不是同一条序列。
INDEX_TOL = 0.005

# investors.tmx.com 认 UA：默认 UA 一律 403（见 docstring 源 1）。
# m-x.ca 不认 UA，但带着也没坏处，统一用同一个。
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

# 结构性断点：等 Cboe Canada 真正交割再填（见口径坑 13）。
# 留一个空常量而不是留一句注释，是为了让「该不该画断点」有个可被 grep 到的落点。
BREAKS = {}

# ══ CSV 列定义 ══════════════════════════════════════════════════════════
# 现货：(csv 列名, CTS 小节 key, 行标签 key)
SPOT_SPEC = [
    ('tmx_all_volume_shares',     'all',    'volume'),
    ('tmx_all_value_cad',         'all',    'value'),
    ('tmx_all_transactions',      'all',    'transactions'),
    ('tsx_volume_shares',         'tsx',    'volume'),
    ('tsx_value_cad',             'tsx',    'value'),
    ('tsx_transactions',          'tsx',    'transactions'),
    ('tsx_composite_close',       'tsx',    'composite'),
    ('tsxv_volume_shares',        'tsxv',   'volume'),
    ('tsxv_value_cad',            'tsxv',   'value'),
    ('tsxv_transactions',         'tsxv',   'transactions'),
    ('tsxv_composite_close',      'tsxv',   'composite'),
    ('alpha_volume_shares',       'alpha',  'volume'),
    ('alpha_value_cad',           'alpha',  'value'),
    ('alpha_transactions',        'alpha',  'transactions'),
    ('alphax_drk_volume_shares',  'alphax', 'volume'),
    ('alphax_drk_value_cad',      'alphax', 'value'),
    ('alphax_drk_transactions',   'alphax', 'transactions'),
]
# Alpha-X / Alpha DRK 官方 2023-11 才单列，之前天然缺席（口径坑，不是解析失败）
SPOT_OPTIONAL = {c for c, sec, _ in SPOT_SPEC if sec == 'alphax'}

# MX xlsx 里的行定位。section=None 表示「该标签在整张表里唯一」，
# 非 None 表示必须先定位到该小节再取行（'Total' 在 5 个小节各出现一次，见口径坑 8）。
S_STIR_FUT = 'Short-Term Interest Rate Futures'
S_BOND_FUT = 'Bond Futures'
S_IDX_FUT = 'Index Futures'
S_IDX_OPT = 'Index Options'
# 小节标题白名单：只有精确等于这几个字符串的行才当小节标题。
# 「A 列有字、B 列往后没数字」这种启发式判据更通用，但一个全空的产品行就会被误判成小节，
# 而误判的后果是后面的 Total 全部错位 —— 白名单笨一点，但错不了。
MX_SECTIONS = [S_STIR_FUT, 'Short-Term Interest Rate Options', S_BOND_FUT,
               S_IDX_FUT, S_IDX_OPT]

# (csv 列名, 取哪个块 vol/adv/oi, section, 行标签)
MX_SPEC = [
    ('mx_volume_contracts',             'vol', None,       'GRAND TOTAL'),
    ('mx_adv_contracts',                'adv', None,       'GRAND TOTAL'),
    ('mx_oi_contracts',                 'oi',  None,       'GRAND TOTAL'),
    ('mx_adv_futures_contracts',        'adv', None,       'Total Futures'),
    ('mx_adv_options_contracts',        'adv', None,       'Total Options'),
    ('mx_adv_stir_futures_contracts',   'adv', S_STIR_FUT, 'Total'),
    ('mx_adv_bax_contracts',            'adv', None,       'BAX'),
    ('mx_adv_cra_contracts',            'adv', None,       'CRA'),
    ('mx_adv_bond_futures_contracts',   'adv', S_BOND_FUT, 'Total'),
    ('mx_adv_cgb_contracts',            'adv', None,       'CGB'),
    ('mx_adv_cgf_contracts',            'adv', None,       'CGF'),
    ('mx_adv_cgz_contracts',            'adv', None,       'CGZ'),
    ('mx_adv_index_futures_contracts',  'adv', S_IDX_FUT,  'Total'),
    ('mx_adv_sxf_contracts',            'adv', None,       'SXF'),
    ('mx_adv_index_options_contracts',  'adv', S_IDX_OPT,  'Total'),
    ('mx_adv_equity_options_contracts', 'adv', None,       'Equity Options'),
    ('mx_adv_etf_options_contracts',    'adv', None,       'ETF Options'),
    ('mx_adv_share_futures_contracts',  'adv', None,       'Share Futures'),
    ('mx_oi_bax_contracts',             'oi',  None,       'BAX'),
    ('mx_oi_cra_contracts',             'oi',  None,       'CRA'),
    ('mx_oi_stir_futures_contracts',    'oi',  S_STIR_FUT, 'Total'),
    ('mx_oi_bond_futures_contracts',    'oi',  S_BOND_FUT, 'Total'),
    # 国债期货三档 OI 与 adv 那三条同源同行：`MONTH END OPEN INTEREST` 是横跨所有产品行的
    # **列块**（第 11 列，见口径坑 8），CGF / CGZ 的格子一直都在，只是本仓此前没登记 ——
    # 于是 /tmx/ 的「国债期货月末未平仓」那一组只有「合计 + CGB」两条列，两者之差被读成「其余」：
    # 2026-07 实测 43.16%、2024 年起中位 41.05%。那 43% 是**本仓的管道边界**，不是官方的
    # 披露边界；补上这两条之后它落回 0.02%。
    # 恒等式（296 个月逐月实测）：Bond Futures 小节的 Total ≡ CGB + CGF + CGZ + LGB，
    # **精确成立、零残差**（2025-06 那个残档走 2507 的上月列，LGB=412 同样对上）。
    # 所以三档之外只剩 LGB 一个合约：残差占比中位 0.0000%、最大 0.3859%（2022-05 的
    # 2,752 张 / 713,178 张），2026-08 是 387 张 / 1,630,704 张 = 0.0237%。
    # LGB 不单列：它 296 个月里有 228 个月 OI 恰为 0，单列出来是一条贴着零轴的死线。
    # 注意小节里 Total 之**后**还有一行 `Bond Options - OGB`，它不在 Total 里（期权不是期货），
    # 别把它算进残差。
    ('mx_oi_cgb_contracts',             'oi',  None,       'CGB'),
    ('mx_oi_cgf_contracts',             'oi',  None,       'CGF'),
    ('mx_oi_cgz_contracts',             'oi',  None,       'CGZ'),
    ('mx_oi_sxf_contracts',             'oi',  None,       'SXF'),
    ('mx_oi_equity_options_contracts',  'oi',  None,       'Equity Options'),
    ('mx_oi_etf_options_contracts',     'oi',  None,       'ETF Options'),
]

# 反推交易日数的基准合约候选表（口径坑 2）。按活跃度降序，第一个够活跃的就用。
# 利率侧与股票侧必须分开，因为它们跑两套日历（口径坑 1）。
DAYS_BASIS = {
    'trading_days_rates':  [(None, 'CGB'), (S_BOND_FUT, 'Total'), (S_STIR_FUT, 'Total')],
    'trading_days_equity': [(None, 'Equity Options'), (None, 'ETF Options'), (None, 'SXF')],
}
# 基准合约当月总量低于这个数就不信它反推出来的日数：ADV 是官方四舍五入到整数的，
# 量太小时 ±0.5 的舍入误差会把比值推离整数。10 万张时误差量级 1e-4，绰绰有余。
DAYS_MIN_VOLUME = 100_000
DAYS_TOL = 0.05

# 每个列块的列序都是 [本月, 去年同月, %CHG, 上月, %CHG]，所以「上月」永远在块首 +3。
# 2015 / 2019 / 2026 三档实测一致。用它修 MX_BROKEN 里那些残档（见下）。
MX_PRIOR_OFFSET = 3
# **官方自己传坏了的档**：2025-06 那份 xlsx 里，所有产品明细行（BAX/CRA/CGZ/CGF/CGB/SXF…）
# 整批消失，只剩各小节的 Total 和 GRAND TOTAL；sheet 名也自相矛盾（'EN' 配 'Jun 2025 EN ADV'）。
# 它正是侦察稿标出「Last-Modified 被推到 2025-07-31」的那份 —— 那次原地重传把文件传坏了。
# 全量 295 份里**只有这一份**有这个毛病（逐份扫过）。
# 修法：改用**下一个月那份文件的「上月」对照列**，仍是 m-x.ca 的一手数据，
# 且用两份文件都有的 GRAND TOTAL 三个数做等值验证后才采信（见 fetch_mx_month）。
MX_BROKEN = {'2025-06'}

COLUMNS = (['month'] + [c for c, _s, _l in SPOT_SPEC]
           + [c for c, _b, _s, _l in MX_SPEC]
           + ['trading_days_rates', 'trading_days_equity'])

BOX_COLUMNS = ['quarter', 'box_volume_mncontracts', 'box_equity_options_share_pct',
               'box_revenue_cadmn', 'box_revenue_usdmn', 'usdcad_avg']

_MONTHS = {m: i for i, m in enumerate(
    ['January', 'February', 'March', 'April', 'May', 'June', 'July',
     'August', 'September', 'October', 'November', 'December'], 1)}
# 月名 -> 月号的**容错查表**（键一律小写）：全名 + 3 字母缩写 + 'Sept'。
# 官方今天在表头和标题里印的都是全名（60 期表头、140 期标题实测无一例外），
# 加缩写不是迁就今天的写法，是让「Jul 2026」这类改法**不再是静默的**：认得出就走
# _crosscheck_headline_month 的对账，认不出就整期无声消失。宽到能读、严在对账 ——
# 同 fetch/msci.py 那段注释的意思，兜住下一次改版的是对账不是正则，正则只认识见过的变体。
_MONTH_LOOKUP = {k: n for m, n in _MONTHS.items()
                 for k in (m.lower(), m[:3].lower())}
_MONTH_LOOKUP['sept'] = 9
# 表头单元格里的月份可能被 inline 标签拆开（口径坑 5），所以先 _squash 再按无空格形态匹配。
# 尾部**刻意不锚定**：这份稿子到处挂脚注符（表下的 `*Includes NEX`、小节标题的
# `All TMX Equities Marketplaces *`），脚注哪天迁进表头变成 `July2026*`，`$` 锚就会把
# 整期现货静默挡在门外。前缀仍然咬死「字母紧跟 4 位年」，所以 YTD 表进不来 ——
# 它表头第一个数据列是 `2026`，压根没有前导字母（口径坑 6）。
_MON_HDR = re.compile(r'^([A-Za-z]+)[.\-–—]?(\d{4})')
# 标题里的「月名 + 年份」。标题是纯英文散文，与正文那张 HTML 表格是两条互不相干的东西，
# 这正是它能当**外部判据**的理由（同 cboe 的自报报告月、ice 的工作簿 URL 月）。
_HEADLINE_MON = re.compile(
    r'\b(%s)\.?\s+(\d{4})\b' % '|'.join(sorted(_MONTH_LOOKUP, key=len, reverse=True)),
    re.IGNORECASE)
# 月度 CTS 正稿的标题指纹（_squash 后小写），140 期无一例外。
# 只用来把「tagList 上挂了一条临时通知」和「正稿的表格没了」分开，不参与取数：
# 认不出只会让护栏对那一条降级成告警，一个数都不会变。
_CTS_TITLE = 'consolidatedtradingstatistics'
# 引子句在 PDF 文字层里可能被换行切开，所以词间用 \s+ 而不是写死空格；
# 锚点只到 "market" 为止 —— YE-2024 / YE-2025 那两份少了 "share" 一词（口径坑 15）。
_BOX_ANCHOR = re.compile(r'The\s+following\s+table\s+summarizes\s+the\s+BOX\s+volume'
                         r'\s+and\s+the\s+equity\s+option\s+market')
_QTR = re.compile(r'^Q\s*([1-4])\s*/\s*(\d{2})$')


class TmxFetchError(RuntimeError):
    """源站结构变化 / 下载失败 / 解析结果不完整。一律炸掉，绝不静默写 NaN。"""


# ── 网络 ────────────────────────────────────────────────────────────────
def _http_get(url, timeout=120):
    req = urllib.request.Request(url, headers={
        'User-Agent': _UA,
        'Accept': '*/*',
        'Accept-Language': 'en-CA,en;q=0.9',
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.read()
    except urllib.error.HTTPError as e:
        raise TmxFetchError('下载失败 %s: HTTP %s' % (url, e.code)) from e
    except Exception as e:                                # noqa: BLE001
        raise TmxFetchError('下载失败 %s: %r' % (url, e)) from e


def _http_get_or_404(url, timeout=120):
    """返回 bytes；**只有干净的 404** 返回 None（= 这个月官方还没发）。

    502/503/超时这类瞬时故障必须继续抛 —— 本次回补 295 个月就撞上一次 502
    （2209 那份，重试即成功）。把它当成 404 会让那个月静默消失在 CSV 里，
    而且往后每次跑都「正常地」跳过它。
    """
    try:
        return _http_get(url, timeout)
    except TmxFetchError as e:
        if 'HTTP 404' in str(e):
            return None
        raise


# ── 通用小工具 ───────────────────────────────────────────────────────────
def _squash(s):
    """删掉**全部**空白。官方把词拆进不同 inline 标签，剥完会冒出词内空格

    （`TSX Venture Ex change`、表头 `Ma rch 2022`），见口径坑 5。
    一切标题 / 行标签 / 表头的匹配都要先过这一道。
    """
    return re.sub(r'\s+', '', s or '')


def _norm(v):
    return re.sub(r'\s+', ' ', str(v)).strip() if v is not None else ''


def _month_num(name):
    """月名 -> 1..12，认不出返回 None。全名 / 3-4 字母缩写 / 任意大小写都认。

    返回 None 而不是抛，是因为两个调用方要的处置不一样：表头那边「不是月份」是家常便饭
    （YTD 表、脚注表都得安静丢掉），标题那边读不出则会让护栏自己失效并喊一声。
    """
    return _MONTH_LOOKUP.get((name or '').lower())


def _num(s):
    """'$ 478,909,112,683' -> 478909112683.0；空 / '-' / 'n/a' / 读不懂 -> None。

    宽松版，只给「这一格本来就可能不是数」的场合用（CTS 表里混着文字行）。
    MX xlsx 的取数一律走 _num_strict。
    """
    if s is None:
        return None
    t = str(s).replace('\xa0', '').replace(',', '').replace('$', '').strip()
    if t in ('', '-', '--', 'n/a', 'N/A'):
        return None
    try:
        return float(t)
    except ValueError:
        return None


def _num_strict(v, where):
    """同 _num，但「非空却读不懂」直接炸。

    xlsx 里数字是带千分位的**字符串**（'4,221,855'），只有 0 是真 int。
    宽松解析在这里是危险的：官方哪天把某一格改成 'n.a.' 之外的别的写法，
    静默返回 None 就是悄悄写出一列空值，正是仓库明令禁止的那种失败。
    """
    if v is None:
        return None
    out = _num(v)
    if out is None and str(v).strip() not in ('', '-', '--', 'n/a', 'N/A'):
        raise TmxFetchError('%s 不是数字：%r' % (where, v))
    return out


def _month_range(start, end):
    """('2026-06','2026-09') -> ['2026-07','2026-08','2026-09']（不含 start）。"""
    y, m = int(start[:4]), int(start[5:7])
    out = []
    while True:
        m += 1
        if m == 13:
            y, m = y + 1, 1
        cur = '%04d-%02d' % (y, m)
        if cur > end:
            return out
        out.append(cur)


def _yymm(month):
    return month[2:4], month[5:7]


# ══════════════════════ 现货：CTS 新闻稿解析 ════════════════════════════
# CTS 小节标题 -> 内部 key。全部写成 _squash 后的小写形态。
# 注意 Montréal / Montreal 两种拼法都真实出现过（同一份稿子里也会混用）。
_CTS_SECTIONS = [
    ('alltmxequitiesmarketplaces', 'all'),
    ('torontostockexchange', 'tsx'),
    ('tsxventureexchange', 'tsxv'),
    ('tsxalphaexchange', 'alpha'),
    ('alpha-xandalphadrk', 'alphax'),
    ('alpha-x&alphadrk', 'alphax'),
    ('montréalexchange', 'mx'),
    ('montrealexchange', 'mx'),
]
# 行标签 -> 内部 key（_squash 后小写；用「以…开头」匹配，因为官方给指数行挂了脚注符号 ^）
_CTS_ROWS = [
    ('volume', 'volume'),
    ('value', 'value'),
    ('transactions', 'transactions'),
    ('s&p/tsxcompositeindexclose', 'composite'),
    ('s&p/tsxventurecompositeindexclose', 'composite'),
    ('derivativesvolume(contracts)', 'mx_volume'),
    ('openinterest(contracts)', 'mx_oi'),
]


def _text_before_tables(node):
    """按文档顺序产出 ('t', 文本) 与 ('T', <table>)，**不进入 table 内部**。

    这是口径坑 4 的解法。用正则去截 `<div>…</div>` 会在那 6 期嵌套标题上从外层 div
    匹到内层 `</div>`，把整个小节吞掉；而自顶向下递归、遇到 table 就整棵剪掉，
    既拿到了「上一张表之后、这张表之前」的全部纯文本，又完全不受 div 嵌套影响。
    """
    for ch in node.children:
        if isinstance(ch, NavigableString):
            yield 't', str(ch)
        elif isinstance(ch, Tag):
            if ch.name == 'table':
                yield 'T', ch
            else:
                yield from _text_before_tables(ch)


def _section_of(context_text):
    """从「这张表之前的那段文字」里认出小节归属；认不出返回 None。

    取**最靠右**的那个命中：正文开头那段导语会把五个市场的名字全列一遍
    （"…Toronto Stock Exchange, TSX Venture Exchange, TSX Alpha Exchange (Alpha),
    including Alpha-X & Alpha DRK, and Montréal Exchange (MX)."），
    紧跟着才是真正的小节标题 `All TMX Equities Marketplaces`。取最左会全归到 TSX。
    """
    s = _squash(context_text).lower()
    best, pos = None, -1
    for key, name in _CTS_SECTIONS:
        p = s.rfind(key)
        if p > pos:
            pos, best = p, name
    return best if pos >= 0 else None


def _parse_cts_table(table):
    """解析一张 CTS 表 -> (数据月 'YYYY-MM', {行 key: float})；不是月度表返回 (None, None)。

    两道判据：
      · 表头第 2 格必须形如 `<月名><年>`（_squash 之后）—— 把 YTD 表（`2026 / 2025 /
        % Change`）和正文末尾那张脚注表一起挡在外面，见口径坑 6；
      · 遇到 `Daily Averages` 行立刻停 —— 月度块和日均块在同一张 <table> 里，
        `Volume` / `Value` / `Transactions` 各出现两次，见口径坑 3。
    """
    rows = table.find_all('tr')
    if not rows:
        return None, None
    head = [_squash(c.get_text(' ', strip=True))
            for c in rows[0].find_all(['td', 'th'])]
    if len(head) < 2:
        return None, None
    m = _MON_HDR.match(head[1])
    num = _month_num(m.group(1)) if m else None
    if num is None:
        return None, None
    month = '%s-%02d' % (m.group(2), num)

    out = {}
    for tr in rows[1:]:
        cells = tr.find_all(['td', 'th'])
        if not cells:
            continue
        lab = _squash(cells[0].get_text(' ', strip=True)).lower()
        if lab.startswith('dailyaverages'):
            break                       # 口径坑 3：本表的取数到此为止
        if not lab or len(cells) < 2:
            continue
        key = next((k for pre, k in _CTS_ROWS if lab.startswith(pre)), None)
        if key is None:
            continue
        out.setdefault(key, _num(cells[1].get_text(' ', strip=True)))
    return month, out


def parse_cts_release(body_html, headline=''):
    """解析一期 CTS 新闻稿正文 -> (数据月, {(小节, 行): 值})；没有表格返回 (None, None)。

    没有表格是 2021-07 及更早那 80 期的**正常**状态（口径坑 7），不是故障，所以返回
    (None, None) 让调用方跳过；但「有表格却认不出小节」是故障，必须炸。
    """
    soup = BeautifulSoup(body_html, 'lxml')
    root = soup.body or soup
    buf, pairs = [], []
    for kind, x in _text_before_tables(root):
        if kind == 't':
            buf.append(x)
        else:
            pairs.append((''.join(buf), x))
            buf = []
    if not pairs:
        return None, None

    rec, months = {}, set()
    for ctx, table in pairs:
        month, vals = _parse_cts_table(table)
        if month is None:
            continue
        sec = _section_of(ctx)
        if sec is None:
            raise TmxFetchError(
                '%s：有一张月度表认不出小节归属（表头 %s），拒绝让它继承上一节 —— '
                '官方多半改了标题写法' % (headline or '?', month))
        months.add(month)
        for k, v in vals.items():
            rec.setdefault((sec, k), v)
    if not rec:
        return None, None
    if len(months) != 1:
        raise TmxFetchError('%s：一期稿子里出现多个数据月 %s' % (headline, sorted(months)))
    return months.pop(), rec


def _cts_to_row(month, rec, headline):
    """把小节字典摊成 CSV 列，并跑官方恒等式自检。"""
    row = {}
    for name, sec, key in SPOT_SPEC:
        row[name] = rec.get((sec, key))
    missing = [n for n, _s, _k in SPOT_SPEC
               if row[n] is None and n not in SPOT_OPTIONAL]
    if missing:
        raise TmxFetchError('%s %s 缺列 %s —— 解析异常，拒绝写入'
                            % (headline, month, missing))
    # 恒等式 All = TSX + TSXV + Alpha + Alpha-X/DRK，59/59 个月成立、最大差 0。
    # 它是本模块最有力的自检：口径坑 5 那两期若解析漏了 Alpha，这里立刻炸而不是
    # 悄悄少一个市场；把 Alpha-X 漏算或重复计也一样会炸。
    for kind, cols in (
            ('volume', ('tmx_all_volume_shares', 'tsx_volume_shares',
                        'tsxv_volume_shares', 'alpha_volume_shares',
                        'alphax_drk_volume_shares')),
            ('value', ('tmx_all_value_cad', 'tsx_value_cad',
                       'tsxv_value_cad', 'alpha_value_cad',
                       'alphax_drk_value_cad')),
            ('transactions', ('tmx_all_transactions', 'tsx_transactions',
                              'tsxv_transactions', 'alpha_transactions',
                              'alphax_drk_transactions'))):
        total = row[cols[0]]
        parts = sum(row[c] or 0 for c in cols[1:])
        if abs(total - parts) > 0.5:
            raise TmxFetchError(
                '%s %s 恒等式不成立：All %s = %r，四家相加 = %r，差 %r'
                % (headline, month, kind, total, parts, total - parts))
    return row


def _headline_month(headline):
    """从新闻稿标题里读出它自报的数据月 'YYYY-MM'；读不出返回 None（宁缺勿猜）。

    140 期标题实测全部形如 `TMX Group Consolidated Trading Statistics - July 2026`
    （连字符 `-` 与 en dash `–` 混用过，尾巴上挂过 `(Revised)` / `(revised)` /
    `(Corrected)`），每一条**恰好**出现一处「月名 + 年份」，且与正文表格解析出来的
    数据月 60/60 一致、与无表那 80 期的边界 2021-07 / 2021-08 严丝合缝。

    刻意要求「恰好一处」：哪天标题写成 "July 2026 (revised October 2026)"，
    两处月份就说不清哪个是数据月 —— 这时返回 None 让护栏自己失效，也绝不猜一个月份
    去和解析结果对账。猜错的代价不是漏抓一次，是把一期健康的稿子判成故障、天天 FAIL。
    """
    hits = _HEADLINE_MON.findall(headline or '')
    if len(hits) != 1:
        return None
    num = _month_num(hits[0][0])
    return None if num is None else '%s-%02d' % (hits[0][1], num)


def _crosscheck_headline_month(headline, parsed):
    """标题自报的数据月 vs 正文表格解析出来的数据月，对不上就炸。返回标题月（或 None）。

    本模块的「独立于解析器的外部判据」，同 fetch/cboe.py 的 _crosscheck_report_month、
    fetch/ice.py 的 _crosscheck_workbook_month。

    防的是 README「第四类：不出声的失败」里那一种：官方把表头 `July 2026` 改成
    `July 2026*` / `JULY 2026` / `July 2026 vs 2025`，或者干脆把 <table> 换成 div 网格，
    于是 _parse_cts_table 返回 (None, None)、parse_cts_release 跟着返回 (None, None)、
    fetch_cts 一个 continue 把整期丢掉 —— 现货 17 列就此冻结。而这件事在日志里
    **连续失败十天和成功十天长得一模一样**：MX 那条腿照常前进、data_through 照常跳月、
    首页红点本来就是按 MX 判的、build/specs/tmx.py 又明写着现货列最新月留空是正常状态。
    没有这道对账，这种坏法在本模块里没有任何人会发现。

    刻意 raise 而不是 print warn：warn 之后总状态仍是 UPDATED / NOCHANGE，等于没有护栏
    （理由同 cboe._crosscheck_report_month）。也**不需要** ice 那种「先等一天」的时间阈值 ——
    标题和表格在同一个 Body 字段里同批发出，不存在 CDN 文件晚于新闻稿的时间差。

    三条豁免，每一条都对着真实存在的合法输入：
      · 标题读不出月份 -> 判据本身没了，喊一声护栏失效然后放行（同 ice 的做法）；
      · 标题月份早于 SPOT_START -> 2021-07 及更早那 80 期正文本来就没有表格（口径坑 7），
        这是**按设计跳过**，不是故障。2026-08-18 复核过边界干净无交叉；
      · 标题不像月度 CTS 正稿 -> tagList=trading-statistics 哪天被官方拿去挂一条临时通知
        （11 年 140 期一次都没有过，但契约上没禁止），那条通知本来就不该有表格。
        这一支同样只喊一声：为一条通知让 28 家里的一家天天 FAIL，代价比它挡住的风险大。
        真出事时还有 update() 里的 _guard_spot_not_vanished 从反侧兜着。

    ⚠ **这道护栏一旦响，不会自愈。** 如果官方真把正文表格永久撤回 2021-07 那种
    「只给一条 /resource 链接」的版式，它就会天天 FAIL，重试多少次都一样，
    要人来决定怎么办（改 SPOT_START 的边界，还是另找一条现货源），代码自己决定不了。
    这是明知故犯：另一头是 17 条现货列在绿点后面悄悄冻住，没人会发现。
    """
    hm = _headline_month(headline)
    if hm is None:
        print('[tmx] ⚠ 护栏失效：这一期标题读不出数据月，本期没有独立于解析器的月份判据。'
              '标题=%r' % (headline,))
        return None
    if parsed is None:
        if hm < SPOT_START:
            return hm               # 口径坑 7：正文本来就没有表格
        if _CTS_TITLE not in _squash(headline).lower():
            print('[tmx] ⚠ %s 这条 trading-statistics 条目的标题不像月度 CTS 正稿、正文也没有'
                  '月度表格，按临时通知放过。标题=%r' % (hm, headline))
            return hm
        raise TmxFetchError(
            '标题自报数据月 %s（不早于 %s，正文本该带月度表格），但一张月度表都没解析出来。'
            '标题=%r。最可能是表头写法变了（_MON_HDR 只认「月名紧跟 4 位年」）或表格离开了'
            '正文（回到 2021-07 那种只给 /resource 链接的版式）。拒绝静默跳过 —— 跳过等于'
            '现货 17 列冻结而日志一片正常；请对照 cache/tmx_cts_feed.json 里这一期的 Body。'
            % (hm, SPOT_START, headline))
    if hm != parsed:
        raise TmxFetchError(
            '标题自报数据月 %s，正文表格解析出来却是 %s（标题=%r）。两者同批发出、不该不一致 —— '
            '最可能是表头月份写法变了导致取到了别的表（YTD 表？去年同月列？）。拒绝写入。'
            % (hm, parsed, headline))
    return hm


def fetch_cts(cache_dir, page_size, window_out=None):
    """拉 CTS feed，返回 {'YYYY-MM': (行字典, 官方发布日, 标题, MX 对账值)}。

    第 4 项是新闻稿 Montréal Exchange 小节里的成交合约数与 OI。
    **它不入库**（口径坑 10：那两行 2022-09 被官方"修订"成了错的），
    只交给 crosscheck() 当第二条独立证据链用。

    每一条都过 _crosscheck_headline_month：不拿标题自报的月份对一遍账，
    「正文表格没解析出来」和「这一期本来就没有表格」就永远是同一个 continue。

    window_out 传一个 list 进来，就把本轮 feed 窗口里各期**标题自报的数据月**填进去
    （读不出的那几期不填）。它是给 update() 的哨兵用的：哨兵得先知道窗口伸到哪个月，
    才谈得上「已入库的那个月这一轮该不该还在」。不传就是纯读，行为不变。
    """
    os.makedirs(cache_dir, exist_ok=True)
    raw = _http_get(FEED_URL.format(n=page_size))
    with open(os.path.join(cache_dir, 'tmx_cts_feed.json'), 'wb') as f:
        f.write(raw)                    # 存原始 feed：源站改版时可以事后取证
    try:
        items = json.loads(raw)['GetPressReleaseListResult']
    except Exception as e:              # noqa: BLE001
        raise TmxFetchError('CTS feed 不是预期的 JSON 结构：%r' % (e,)) from e
    if not items:
        raise TmxFetchError('CTS feed 返回 0 条 —— tagList=trading-statistics 可能被改名')

    out = {}
    for it in items:
        headline = it.get('Headline', '')
        month, rec = parse_cts_release(it.get('Body') or '', headline)
        hm = _crosscheck_headline_month(headline, month)
        if window_out is not None and hm:
            window_out.append(hm)
        if month is None:
            continue                    # 2021-07 及更早：正文没表格，正常（口径坑 7）
        row = _cts_to_row(month, rec, headline)
        day = _release_date(it)
        mx_ref = {'volume': rec.get(('mx', 'mx_volume')),
                  'oi': rec.get(('mx', 'mx_oi'))}
        out.setdefault(month, (row, day, headline, mx_ref))
    return out


def _release_date(item):
    """feed 的 PressReleaseDate 字段 -> 'YYYY-MM-DD'；读不懂返回 None（宁缺勿猜）。"""
    raw = (item.get('PressReleaseDate') or '').strip()
    for fmt in ('%m/%d/%Y %H:%M:%S', '%m/%d/%Y'):
        try:
            return datetime.strptime(raw, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


# ══════════════════════ MX：月度 xlsx 解析 ══════════════════════════════
def _mx_sheet(wb, suffix):
    """按后缀取 sheet：老档叫 'Jan 2015 EN'，新档就叫 'EN'（口径坑 12）。

    要求恰好命中一张：命中 0 张说明官方改了 sheet 命名，命中多张说明后缀判据不够，
    两种都得炸 —— 随便挑一张的话，'EN' 与 'EN ADV' 一旦串了，
    整张表的月度总量会被写成日均，而数字看上去完全正常。
    """
    hits = [s for s in wb.sheetnames if s == suffix or s.endswith(' ' + suffix)]
    if len(hits) != 1:
        raise TmxFetchError('工作簿里 %r 表命中 %d 张（sheet 列表 %r）'
                            % (suffix, len(hits), wb.sheetnames))
    return wb[hits[0]]


def _mx_locate_rows(ws):
    """返回 ({(section, label): row}, {label: 行数})。section 见口径坑 8。"""
    by_key, count = {}, {}
    section = None
    for r in range(1, ws.max_row + 1):
        lab = _norm(ws.cell(r, 1).value)
        if not lab:
            continue
        if lab in MX_SECTIONS:
            section = lab
            continue
        by_key.setdefault((section, lab), r)
        by_key.setdefault((None, lab), r)
        count[lab] = count.get(lab, 0) + 1
    return by_key, count


def _mx_blocks(ws):
    """第 1 行的大标题 -> 该列块的起始列。列块内第一列就是「本月」。"""
    out = {}
    for c in range(1, ws.max_column + 1):
        t = _norm(ws.cell(1, c).value)
        if t:
            out.setdefault(t, c)
    return out


def _mx_block_month(ws, col):
    """列块起始列的第 2 行写着该列代表哪个月（'Jul 2026'）-> 'YYYY-MM'。"""
    raw = _norm(ws.cell(2, col).value).replace('\xa0', ' ')
    for fmt in ('%b %Y', '%B %Y'):
        try:
            d = datetime.strptime(raw, fmt)
            return '%04d-%02d' % (d.year, d.month)
        except ValueError:
            continue
    raise TmxFetchError('列块表头读不懂月份：%r' % raw)


def parse_mx_workbook(path, which='current', spec=None):
    """解析一份 m-x.ca 月度 xlsx -> (数据月, {csv 列名: float|None}, 原始取数字典)。

    which='current' 取每个列块的第 1 列（这份文件自己那个月）；
    which='prior'   取第 4 列（该文件里的「上月」对照列，见 MX_PRIOR_OFFSET），
                    只给 fetch_mx_month 修 2025-06 那种残档时用。
    spec 默认取全部 MX_SPEC；传子集是给残档验证用的 —— 残档缺产品明细行，
    但三个 GRAND TOTAL 还在，只读那三行就能验证替代来源是不是同一套数。

    第三个返回值是 {(块, section, label): 值} 的全量取数，给 trading_days 反推和
    crosscheck() 用 —— 它们要的中间量（CGB 月度总量之类）不进 CSV，但必须可复算。

    约定的行标签找不到 = 官方改了表结构，直接抛异常。宁可整月不更新，
    也不要写出一列悄悄全空的 CSV。
    """
    off = 0 if which == 'current' else MX_PRIOR_OFFSET
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        ws_en, ws_adv = _mx_sheet(wb, 'EN'), _mx_sheet(wb, 'EN ADV')
        rows_en, cnt_en = _mx_locate_rows(ws_en)
        rows_adv, cnt_adv = _mx_locate_rows(ws_adv)
        blk_en, blk_adv = _mx_blocks(ws_en), _mx_blocks(ws_adv)

        for title, blocks in (('MONTHLY VOLUME', blk_en),
                              ('MONTH END OPEN INTEREST', blk_en),
                              ('AVERAGE DAILY VOLUME', blk_adv)):
            if title not in blocks:
                raise TmxFetchError('%s 第 1 行找不到列块标题 %r'
                                    % (os.path.basename(path), title))
        col = {'vol': blk_en['MONTHLY VOLUME'] + off,
               'oi': blk_en['MONTH END OPEN INTEREST'] + off,
               'adv': blk_adv['AVERAGE DAILY VOLUME'] + off}
        month = _mx_block_month(ws_en, col['vol'])
        for k in ('oi', 'adv'):
            ws = ws_adv if k == 'adv' else ws_en
            got = _mx_block_month(ws, col[k])
            if got != month:
                raise TmxFetchError('%s 的 %s 列块月份 %s 与 MONTHLY VOLUME 的 %s 不一致'
                                    % (os.path.basename(path), k, got, month))

        def cell(block, section, label):
            rows, cnt = (rows_adv, cnt_adv) if block == 'adv' else (rows_en, cnt_en)
            key = (section, label)
            if key not in rows:
                raise TmxFetchError('%s 的 %s 表缺行 %r（小节 %r）—— 官方表结构可能已变'
                                    % (os.path.basename(path), block, label, section))
            if section is None and cnt.get(label, 0) != 1:
                raise TmxFetchError('%s 的 %s 表里 %r 出现了 %d 次，无法唯一定位'
                                    % (os.path.basename(path), block, label,
                                       cnt.get(label, 0)))
            ws = ws_adv if block == 'adv' else ws_en
            r = rows[key]
            return _num_strict(ws.cell(r, col[block]).value,
                               '%s %s 表 R%dC%d(%s)'
                               % (os.path.basename(path), block, r,
                                  col[block], label))

        raw = {}
        rec = {}
        for name, block, section, label in (spec or MX_SPEC):
            v = cell(block, section, label)
            raw[(block, section, label)] = v
            rec[name] = v
        if spec is None:
            # 反推交易日数要用到的中间量（不进 CSV）
            for cands in DAYS_BASIS.values():
                for section, label in cands:
                    for block in ('vol', 'adv'):
                        if (block, section, label) not in raw:
                            raw[(block, section, label)] = cell(block, section, label)
            for name in ('trading_days_rates', 'trading_days_equity'):
                rec[name] = _trading_days(raw, DAYS_BASIS[name],
                                          os.path.basename(path))
        return month, rec, raw
    finally:
        wb.close()


def _trading_days(raw, candidates, fname):
    """用「基准合约月度量 / 该合约 ADV」反推交易日数，见口径坑 2。

    绝不能用 GRAND TOTAL 反推（跨两套日历，2025-09 会推出 20.499）。
    停用合约会 0/0（2026 年的 BAX），所以按活跃度逐个降级；全都不够活跃就返回 None
    —— 那是 2002 年那种「ETF 期权还没上市」的正常状态，不是故障。
    """
    for section, label in candidates:
        vol = raw.get(('vol', section, label))
        adv = raw.get(('adv', section, label))
        if not vol or not adv or vol < DAYS_MIN_VOLUME:
            continue
        ratio = vol / adv
        days = round(ratio)
        if days <= 0 or abs(ratio - days) > DAYS_TOL:
            raise TmxFetchError(
                '%s 用 %s 反推交易日数得到 %.4f，不在整数附近 —— 这个基准合约多半自己就'
                '跨了两套日历，不能拿它当基准' % (fname, label, ratio))
        return float(days)
    return None


def _mx_path(cache_dir, month):
    """MX 的档案单独放 cache/tmx/ —— 一个月一份文件，全量 295 份，

    平铺在 cache/ 根下会把别家的缓存淹掉（cache/ 是给人翻的，不只是给程序读的）。
    """
    yy, mm = _yymm(month)
    d = os.path.join(cache_dir, 'tmx')
    os.makedirs(d, exist_ok=True)
    return os.path.join(d, '%s%s_stats_en.xlsx' % (yy, mm))


def _mx_download(cache_dir, month):
    """确保某月的 xlsx 在本地；官方还没发（干净 404）返回 None。"""
    path = _mx_path(cache_dir, month)
    if os.path.exists(path) and os.path.getsize(path) > 10_000:
        return path
    yy, mm = _yymm(month)
    data = _http_get_or_404(MX_TMPL.format(yy=yy, mm=mm))
    if data is None:
        return None
    with open(path, 'wb') as f:
        f.write(data)
    return path


def fetch_mx_month(cache_dir, month):
    """下载并解析某个月的 MX xlsx；官方还没发（干净 404）返回 None。

    MX_BROKEN 里那几个月走「下一份文件的上月对照列」这条备用路（见 MX_BROKEN 注释）。
    备用路不是无条件信任的：先用两份文件都有的 GRAND TOTAL（月度量 / ADV / 月末 OI）
    三个数做等值验证，全对上才采信 —— 对不上说明拿错了列或拿错了月，宁可炸。
    """
    path = _mx_download(cache_dir, month)
    if path is None:
        return None
    if month not in MX_BROKEN:
        got, rec, raw = parse_mx_workbook(path)
        if got != month:
            raise TmxFetchError('%s 里写的是 %s，与文件名要求的 %s 不符'
                                % (os.path.basename(path), got, month))
        return rec, raw

    nxt = _next_month(month)
    npath = _mx_download(cache_dir, nxt)
    if npath is None:
        raise TmxFetchError('%s 那份 xlsx 是残档（缺产品明细行），只能靠 %s 那份的'
                            '「上月」列来补，但 %s 还没发布' % (month, nxt, nxt))
    got, rec, raw = parse_mx_workbook(npath, which='prior')
    if got != month:
        raise TmxFetchError('%s 的「上月」列写的是 %s，不是 %s'
                            % (os.path.basename(npath), got, month))
    # 残档自己的三个 GRAND TOTAL 还在，拿来验证替代来源确实是同一个月的同一套数
    grand = [s for s in MX_SPEC if s[3] == 'GRAND TOTAL']
    _self_m, self_rec, _self_raw = parse_mx_workbook(path, spec=grand)
    if _self_m != month:
        raise TmxFetchError('%s 残档里写的月份是 %s' % (os.path.basename(path), _self_m))
    for k in ('mx_volume_contracts', 'mx_adv_contracts', 'mx_oi_contracts'):
        if self_rec[k] != rec[k]:
            raise TmxFetchError('%s 的 %s：残档自己写 %r，%s 的上月列写 %r —— 两者不符，'
                                '不能用后者顶替' % (month, k, self_rec[k], nxt, rec[k]))
    return rec, raw


def _next_month(month):
    y, m = int(month[:4]), int(month[5:7])
    return '%04d-%02d' % (y + 1, 1) if m == 12 else '%04d-%02d' % (y, m + 1)


def mx_latest_month(cache_dir):
    """官方当前最新的 MX 月份 'YYYY-MM'。

    先信落地页 —— 它是官方自己维护的「当前最新」指针（永远只挂最近 3 期）；
    落地页认不出来时再按模板往后探（未发布的月份是干净 404，可以放心探）。
    """
    os.makedirs(cache_dir, exist_ok=True)
    html = _http_get(MX_LANDING).decode('utf-8', 'replace')
    with open(os.path.join(cache_dir, 'tmx_mx_landing.html'), 'w',
              encoding='utf-8') as f:
        f.write(html)
    hits = re.findall(r'f_stat_en/(\d{2})(\d{2})_stats_en\.(?:xlsx|pdf)', html)
    if not hits:
        raise TmxFetchError('m-x.ca 落地页上找不到 f_stat_en 直链，源站可能改版：'
                            + MX_LANDING)
    yy, mm = max(hits)
    return '20%s-%s' % (yy, mm)


# ═════════════ 源 4：TMX Money 指数月线（一次性回补，不进 update()）═════════════
def _money_series(symbol, cache_dir, timeout=120):
    """拉一条指数的月线 -> {'YYYY-MM': 月末收盘}。

    端点没有官方契约文档（前端自用），所以每一步都当它随时会变形：
    结构对不上一律抛，绝不返回半张表让调用方以为「这个月官方没发」。
    """
    body = json.dumps({
        'operationName': 'getTimeSeriesData',
        'variables': {'symbol': symbol, 'freq': 'month', 'interval': 1,
                      'start': '1990-01-01',
                      'end': datetime.now().strftime('%Y-%m-%d')},
        'query': MONEY_QUERY,
    }).encode('utf-8')
    req = urllib.request.Request(MONEY_URL, data=body, headers={
        'Content-Type': 'application/json',
        'User-Agent': _UA,
        'Accept': '*/*',
        'Origin': 'https://money.tmx.com',
        'Referer': 'https://money.tmx.com/',
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            raw = r.read()
    except Exception as e:                                # noqa: BLE001
        raise TmxFetchError('TMX Money 取 %s 失败：%r' % (symbol, e)) from e
    if cache_dir:
        os.makedirs(cache_dir, exist_ok=True)
        fn = 'money_%s.json' % symbol.replace('^', '').lower()
        with open(os.path.join(cache_dir, fn), 'wb') as f:
            f.write(raw)                # 存原始响应：端点改版时可以事后取证
    try:
        rows = json.loads(raw)['data']['getTimeSeriesData']
    except Exception as e:                                # noqa: BLE001
        raise TmxFetchError('TMX Money 返回的不是预期结构（%s）：%s'
                            % (symbol, raw[:200])) from e
    if not rows:
        raise TmxFetchError('TMX Money 对 %s 返回 0 条 —— 代码或参数被改过' % symbol)

    # **只收已经收完的月份**：当月那根 bar 的 close 是「到今天为止」，不是月末收盘。
    # 响应里当月还会多出一条以当天为 dateTime 的重复 bar（实测 2026-08-01 与
    # 2026-08-18 同时出现），一起被这道闸挡在外面。
    cutoff = datetime.now().strftime('%Y-%m')
    out = {}
    for r in rows:
        dt = (r.get('dateTime') or '')[:7]
        if len(dt) != 7 or dt[4] != '-' or dt >= cutoff:
            continue
        v = r.get('close')
        if v is None:
            continue
        if dt in out and abs(out[dt] - float(v)) > INDEX_TOL:
            raise TmxFetchError('TMX Money 对 %s %s 给了两个不同的 close：%r / %r'
                                % (symbol, dt, out[dt], v))
        out[dt] = float(v)
    if not out:
        raise TmxFetchError('TMX Money 对 %s 一个完整月都没给出' % symbol)
    return out


def fetch_index_history(cache_dir):
    """两条指数列的全历史 -> {'YYYY-MM': {列: 点位}}；顺带做连续性自检。

    自检只查「有没有洞」，不查数值 —— 数值的把关在 backfill_index() 里靠重叠月做。
    """
    got = {sym: _money_series(sym, cache_dir) for sym, _c in MONEY_SPEC}
    out = {}
    for sym, col in MONEY_SPEC:
        s = got[sym]
        months = sorted(s)
        if months[0] > INDEX_START:
            raise TmxFetchError('%s 只回到 %s，比已知源底 %s 还晚 —— 端点被改过'
                                % (sym, months[0], INDEX_START))
        want = [months[0]] + _month_range(months[0], months[-1])
        holes = [m for m in want if m not in s]
        if holes:
            raise TmxFetchError('%s 的月线中间有洞：%s' % (sym, holes[:6]))
        for m in months:
            out.setdefault(m, {})[col] = s[m]
    return out


def backfill_index(series_dir, cache_dir):
    """把 tsx_composite_close / tsxv_composite_close 回补到 2001-12，返回新增月份。

    **update() 不调它**，理由见模块 docstring 源 4：这个端点没有官方契约，
    形状随前端改版而变，不该摆进无人值守链路。日常增量仍由 CTS 表格提供这两列。

    幂等与安全：
      · 走同一个 `_merge`，**已有值一格都不覆盖** —— CTS 那 60 个月保持原样；
      · 写之前先拿重叠月做断言：至少 `INDEX_MIN_OVERLAP` 格能对照、且每一格差
        小于 `INDEX_TOL`，否则抛异常。端点哪天换了口径（比如改成总回报指数），
        这道闸会当场拦下，而不是让一条对不上的历史悄悄接在 CTS 前面。
    """
    csv_path = os.path.join(series_dir, 'tmx.csv')
    header, body, have = _load_csv(csv_path, COLUMNS, 'month')
    idx = {n: i for i, n in enumerate(header)}
    data = fetch_index_history(cache_dir)

    n_ov, worst = 0, (0.0, None)
    for month, row in have.items():
        for _sym, col in MONEY_SPEC:
            cur = row[idx[col]].strip()
            if not cur or month not in data or col not in data[month]:
                continue
            n_ov += 1
            d = abs(float(cur) - data[month][col])
            if d > worst[0]:
                worst = (d, '%s %s（CSV %s vs 端点 %s）'
                         % (month, col, cur, data[month][col]))
    if n_ov < INDEX_MIN_OVERLAP:
        raise TmxFetchError('只有 %d 格能与 CSV 对照（要求 ≥%d），不敢写'
                            % (n_ov, INDEX_MIN_OVERLAP))
    if worst[0] > INDEX_TOL:
        raise TmxFetchError('重叠月对不上，最大差 %.4f > %.4f：%s —— '
                            'TMX Money 与 CTS 表格不是同一条序列了，拒绝回补'
                            % (worst[0], INDEX_TOL, worst[1]))

    added = []
    for month in sorted(data):
        if _merge(header, body, have, month, data[month]):
            added.append(month)
    _write_csv(csv_path, header, body)
    print('· 指数重叠 %d 格全部吻合（最大差 %.4f）；新增 %d 行，'
          '两列现自 %s 起' % (n_ov, worst[0], len(added), min(data)))
    return added


# ══════════════════════ BOX：季度 MD&A ═════════════════════════════════
def _finrep_docs(cache_dir):
    """财报 feed -> [(报告期 datetime, 报告名, MD&A/季报 PDF 直链)]，按报告期降序。

    文件名毫无规律，只能靠 feed 发现（见 docstring 源 3）。
    2024Q4 起 MD&A 单独成篇（category 'mdna'），更早的包在 'tenq' 那份大 PDF 里。
    """
    os.makedirs(cache_dir, exist_ok=True)
    raw = _http_get(FINREP_URL)
    with open(os.path.join(cache_dir, 'tmx_finrep_feed.json'), 'wb') as f:
        f.write(raw)
    try:
        items = json.loads(raw)['GetFinancialReportListResult']
    except Exception as e:              # noqa: BLE001
        raise TmxFetchError('财报 feed 不是预期的 JSON 结构：%r' % (e,)) from e
    out = []
    for it in items:
        docs = it.get('Documents') or []
        by_cat = {}
        for d in docs:
            if isinstance(d, dict) and d.get('DocumentPath'):
                by_cat.setdefault(d.get('DocumentCategory'), d['DocumentPath'])
        url = by_cat.get('mdna') or by_cat.get('tenq')
        if not url or not url.lower().endswith('.pdf'):
            continue
        try:
            d = datetime.strptime((it.get('ReportDate') or '')[:10], '%m/%d/%Y')
        except ValueError:
            continue
        out.append((d, '%s %s' % (it.get('ReportYear'), it.get('ReportSubType')), url))
    out.sort(key=lambda x: x[0], reverse=True)
    return out


def parse_box_table(text, where=''):
    """从 MD&A 文字层里抽 BOX 的「最近八个季度」表 -> {'YYYY-Qn': {列: 值}}。

    锚点只取到 "…the equity option market" 为止：YE-2024 / YE-2025 那两份少了 "share"
    一词（口径坑 15）。季度个数不写死（2023Q1 那份只有 5 个季度）。
    """
    m = _BOX_ANCHOR.search(text)
    if not m:
        raise TmxFetchError('%s 里找不到 BOX 八季度表的引子句 —— 官方措辞可能变了' % where)
    lines = [ln.strip() for ln in text[m.start():m.start() + 4000].split('\n')]
    quarters, start = [], None
    for n, ln in enumerate(lines):
        m = _QTR.match(ln)
        if m:
            if start is None:
                start = n
            quarters.append('20%s-Q%s' % (m.group(2), m.group(1)))
        elif start is not None:
            break
    if not quarters:
        raise TmxFetchError('%s：找到引子句但后面没有 Qn/YY 表头' % where)

    wanted = [
        ('box_volume_mncontracts', 'Volume (million contracts)'),
        ('box_equity_options_share_pct', 'Market Share (equity options)'),
        ('box_revenue_cadmn', 'Revenue (in millions of CAD)'),
        ('box_revenue_usdmn', 'Revenue (in millions of USD)'),
        # 2023 年前印成 CAD-USD、之后印成 USD-CAD，数字含义没变（口径坑 15）
        ('usdcad_avg', 'FX rate'),
    ]
    rest = lines[start + len(quarters):]
    out = {q: {} for q in quarters}
    for col, label in wanted:
        try:
            k = next(n for n, ln in enumerate(rest) if ln.endswith(label))
        except StopIteration:
            raise TmxFetchError('%s：BOX 表里找不到 %r 行' % (where, label)) from None
        vals = [_num(x.replace('%', '')) for x in rest[k + 1:k + 1 + len(quarters)]]
        if any(v is None for v in vals):
            raise TmxFetchError('%s：BOX 表 %r 行只读出 %r，与 %d 个季度对不上'
                                % (where, label, vals, len(quarters)))
        for q, v in zip(quarters, vals):
            out[q][col] = v
    return out


def fetch_box(cache_dir, max_reports=1):
    """取最近 max_reports 份 MD&A/季报里的 BOX 表，返回 {'YYYY-Qn': {列: 值}}。

    **从新往旧**处理，配合「只填空不覆盖」自然让最新那份的精度胜出（口径坑 14）。
    """
    os.makedirs(cache_dir, exist_ok=True)
    out, errs = {}, []
    import fitz                                  # pymupdf，只有这条路用得上
    for d, name, url in _finrep_docs(cache_dir)[:max_reports]:
        path = os.path.join(cache_dir, 'tmx_mdna_%s.pdf' % d.strftime('%Y%m%d'))
        if not (os.path.exists(path) and os.path.getsize(path) > 10_000):
            with open(path, 'wb') as f:
                f.write(_http_get(url))
        with fitz.open(path) as doc:
            text = '\n'.join(pg.get_text() for pg in doc)
        try:
            for q, rec in parse_box_table(text, name).items():
                out.setdefault(q, rec)
        except TmxFetchError as e:
            errs.append((name, str(e)))
    if not out:
        raise TmxFetchError('翻了 %d 份财报都没抽出 BOX 表：%r' % (max_reports, errs))
    return out, errs


# ══════════════════════ CSV 读写（幂等） ════════════════════════════════
def _fmt(v):
    """整数写成整数，小数保留最短往返表示。

    MX 是合约张数、现货是股/加元/笔，全都是整数；写成 '20211732.0' 只会让人怀疑
    它被谁算过一遍。指数收盘价（34856.99）才走 repr(float)。
    """
    if v is None:
        return ''
    f = float(v)
    return str(int(f)) if f == int(f) else repr(f)


def _load_csv(path, columns, key):
    """读 CSV 为 (表头, 行列表, {key: 行})。文件不存在就按 columns 建一张只有表头的。"""
    if not os.path.exists(path):
        with open(path, 'w', newline='', encoding='utf-8') as f:
            csv.writer(f, lineterminator='\n').writerow(columns)
    with open(path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    if not rows:
        raise TmxFetchError('%s 是空文件（连表头都没有）' % path)
    header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
    if header[0] != key:
        raise TmxFetchError('%s 的第一列是 %r，期望 %r' % (path, header[0], key))
    missing = [c for c in columns if c not in header]
    if missing:
        raise TmxFetchError('%s 里没有这些列：%s' % (path, missing))
    return header, body, {r[0]: r for r in body}


def _merge(header, body, have, key_value, rec):
    """把一条记录并进 CSV 行集合。返回 True 表示新建了一行。

    幂等的两条铁律都在这里：
      · 已经有值的单元格**永不覆盖** —— 官方会重述（m-x.ca 的 xlsx 是活文件，
        历史档被整批重传过），重述不由本模块自动吞进来；
      · 没被碰过的单元格是原样字符串搬运，所以「什么都没变」时文件字节级不变。
    """
    idx = {n: i for i, n in enumerate(header)}
    if key_value in have:
        row = have[key_value]
        for name, v in rec.items():
            if v is not None and not row[idx[name]].strip():
                row[idx[name]] = _fmt(v)
        return False
    row = [''] * len(header)
    row[0] = key_value
    for name, v in rec.items():
        row[idx[name]] = _fmt(v)
    have[key_value] = row
    body.append(row)
    return True


def _write_csv(path, header, body):
    body.sort(key=lambda r: r[0])
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(body)


# ── 发布日台账 ───────────────────────────────────────────────────────────
def _source_dates():
    """按路径加载仓库根的 source_dates.py（不能裸 import，见该文件的 load()）。"""
    import importlib.util
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(root, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _record_source_dates(series_dir, cts, ingested):
    """把 CTS 各期自述的发布日记进 series/source_dates.csv。

    只用 CTS 的 `PressReleaseDate` —— 那是官方自己写在 feed 里的发布日，精确到日。
    MX xlsx 一侧**一个日期都不记**：它既没有自述发布日，Last-Modified 又被整批重传
    污染过（2019-01 那份带的是 2022-07-23）。MX 比现货早到的那几天，页面抬头
    「官方发布于」那半句就该留白，而不是印一个像模像样的错日期。

    只对真的落进 series/tmx.csv 的月份作证；已有记录一律不覆盖（重述稿的日期是改稿日，
    覆盖会把首发日悄悄改晚，而页面照印不误）。
    """
    sd = _source_dates()
    for month in sorted(cts):
        _row, day, headline, _mx_ref = cts[month]
        if month not in ingested or not day:
            continue
        if sd.lookup(series_dir, 'tmx', month):
            continue
        sd.record(series_dir, 'tmx', month, day,
                  'TMX IR 新闻稿 feed 条目「%s」的 PressReleaseDate 字段' % headline)


# ══════════════════════ 对外接口 ═══════════════════════════════════════
def latest_months(cache_dir):
    """{'mx': 'YYYY-MM', 'spot': 'YYYY-MM'} —— 两条腿必须分开判。

    每个月初都会出现「MX 已有上个月、现货还没发」的正常状态（写这份代码的 2026-08-06
    就是：MX 到 2026-07，CTS 到 2026-06）。用一个标量代表「最新月」，每月都会误报一次故障。
    """
    mx = mx_latest_month(cache_dir)
    cts = fetch_cts(cache_dir, page_size=6)
    return {'mx': mx, 'spot': max(cts) if cts else None}


def latest_month(cache_dir):
    """官方源当前最新月 'YYYY-MM' —— 仓库统一签名，取两条腿里**较新**的那个。

    含义是「官方对这个月已经发出了点什么」，不是「这个月的数据齐了」。
    要区分两条腿请直接调 latest_months()。
    抓不到 / 解析不出来一律抛 TmxFetchError，不返回 None 掩盖故障。
    """
    got = latest_months(cache_dir)
    have = [v for v in got.values() if v]
    if not have:
        raise TmxFetchError('两条源都没给出任何月份')
    return max(have)


def _guard_spot_not_vanished(spot_from, cts, window):
    """已入库的**最新现货月**，这一轮必须还能从 feed 里解析出来；不见了就炸。

    这是 fetch/msci.py 的 update() 里那条「已入库月份整行不见了 -> 抛」在本模块的同款，
    也是 _crosscheck_headline_month 的补网 —— 两道网的盲区不重叠：
    那道从「标题说这一期该有什么」的正面查，这道从「我们已知库里有什么」的反面查。
    官方哪天把标题写法也一起改了（_headline_month 返回 None、那道护栏自己失效并降级成
    告警），就只剩这一道还站着；而告警改变不了总状态，改变不了就等于没拦。

    只查最新那一个月，不是偷懒，是**为了不越出 feed 窗口**：
      · window 是本轮各期标题自报的数据月，spot_from 早于窗口下沿时直接放行 ——
        窗口没伸到那儿，谈不上「不见了」；
      · 往回多查几个月看着更严，实际上只是拿误杀换覆盖：真出事的时候窗口里这几十期是
        一起解析失败的，查一个月和查六个月抓到的是同一件事，而多查一个月就多一分
        「官方撤下一期旧稿」被判成故障的机会。

    两条豁免：库里还没有任何 CTS 来源的月份（全新的 CSV），或者最新现货月早于
    SPOT_START —— 后者说明库里那段现货全部来自 CIRO 回补（口径坑 16），
    feed 窗口本来就不该包含它。
    """
    if not spot_from or spot_from < SPOT_START:
        return
    if not window or spot_from < min(window):
        return
    if spot_from in cts:
        return
    raise TmxFetchError(
        '%s 已经在 series/tmx.csv 的现货列里，这一轮 feed 窗口（%s 起，共 %d 期）却'
        '一个月度表都没为它解析出来。官方不撤稿，所以多半不是它没了，是我们没解析出来 ——'
        '本轮不写入，请对照 cache/tmx_cts_feed.json 里那一期的 Body 人工确认。'
        % (spot_from, min(window), len(window)))


def update(series_dir, cache_dir):
    """把新月份写进 series/tmx.csv（并顺带刷新季度的 tmx_box_q.csv），返回新增月份列表。

    幂等保证：
      · 已存在的月份不重复追加；
      · 已经有值的单元格**永不覆盖**（m-x.ca 的 xlsx 是活文件、会被官方重传订正，
        重述不由本模块自动吞进来；真要重刷历史请手工删行重跑）；
      · 只对既有行里**原本为空**的格子回补 —— 这条是必须的，不是可选的：
        每个月初 MX 先到、现货后到，先建的那一行现货列全空，
        不回补的话那些月份的现货数据永远补不上。回补不计入返回值（它不是新月份）。

    两条腿的取数范围：
      · MX：从 CSV 里最后一个有 MX 数的月份之后，一路补到官方最新月（CSV 空则从 2002-01）；
      · 现货：CTS feed 一次给回最近 N 期，N 按缺口自适应（见下），解析出来的月份
        全部并进去 —— 早于 2021-08 的期次正文里根本没有表格，会被跳过（口径坑 7）。

    BOX 放在最后、且**只在出现新季度时才下载 PDF**：它是季度数据，一年只有 4 次真活干；
    把 PDF 解析摆在 tmx.csv 落盘之后，是为了让「MD&A 措辞变了」这类故障至少不至于
    连累已经抓到的月度数据（下次跑 tmx.csv 那段是纯 no-op）。
    """
    csv_path = os.path.join(series_dir, 'tmx.csv')
    header, body, have = _load_csv(csv_path, COLUMNS, 'month')

    # ── MX 腿 ──
    mx_cols = [c for c, _b, _s, _l in MX_SPEC]
    idx = {n: i for i, n in enumerate(header)}
    done_mx = [r[0] for r in body if r[idx['mx_volume_contracts']].strip()]
    mx_from = max(done_mx) if done_mx else None
    mx_latest = mx_latest_month(cache_dir)
    want_mx = (_month_range(mx_from, mx_latest) if mx_from
               else [MX_START] + _month_range(MX_START, mx_latest))

    added = []
    for month in want_mx:
        got = fetch_mx_month(cache_dir, month)
        if got is None:                 # 干净 404 = 官方还没发这个月
            continue
        rec, _raw = got
        if _merge(header, body, have, month, rec):
            added.append(month)

    # ── 现货腿 ──
    done_spot = [r[0] for r in body if r[idx['tmx_all_volume_shares']].strip()]
    spot_from = max(done_spot) if done_spot else None
    gap = len(_month_range(spot_from, mx_latest)) if spot_from else 999
    # feed 一次最多回 300 条（全量 139 期也才 139 条）；缺口小的常规月只拉十来条，
    # 省掉每月几 MB 的正文下载，也少给对方站点添堵。+3 是给「官方补发/重发」留的余量。
    page = min(300, max(6, gap + 3))
    window = []
    cts = fetch_cts(cache_dir, page_size=page, window_out=window)
    # 反侧哨兵：上一轮已入库的最新现货月，这一轮必须还在（见 _guard_spot_not_vanished）。
    # 放在 _merge 之前 —— 一旦要炸就一个字节都别写。
    _guard_spot_not_vanished(spot_from, cts, window)
    for month in sorted(cts):
        row = cts[month][0]
        if _merge(header, body, have, month, row):
            added.append(month)

    # 无条件落盘：即便没有新月份，上面的现货回补也可能改了既有行。
    _write_csv(csv_path, header, body)
    # 记发布日放在落盘之后：写盘失败就不该在台账上多出一行说「这个月官方发过了」。
    _record_source_dates(series_dir, cts, set(have))

    _update_box(series_dir, cache_dir)          # 季度 BOX，多数月份是零成本 no-op
    return sorted(set(added))


def _update_box(series_dir, cache_dir, max_reports=1):
    """刷新 series/tmx_box_q.csv，返回 (新增季度列表, 解析失败的报告清单)。

    先看财报 feed 里最新那份报告对应的季度在不在表里 —— 在就直接收工，**连 PDF 都不下**。
    一年 12 次月度运行里有 8 次会走这条零成本分支，只有财报后的那一次才真去解析 PDF，
    把「MD&A 措辞变了」的故障窗口压到一年 4 次。

    max_reports=1（默认，cron 走的那条）时，那一份解析不出来就抛异常，绝不静默跳过；
    >1 只给一次性回补用，翻更早的报告，格式太老的那几份记进 errs 由调用方打出来。
    """
    path = os.path.join(series_dir, 'tmx_box_q.csv')
    header, body, have = _load_csv(path, BOX_COLUMNS, 'quarter')
    docs = _finrep_docs(cache_dir)
    if max_reports == 1 and docs:
        d = docs[0][0]
        newest_q = '%d-Q%d' % (d.year, (d.month - 1) // 3 + 1)
        if newest_q in have:
            return [], []
    box, errs = fetch_box(cache_dir, max_reports=max_reports)
    # 从新往旧并入：最新那份印的精度最高，「只填空不覆盖」自然让它胜出（口径坑 14）
    added = [q for q in sorted(box) if _merge(header, body, have, q, box[q])]
    _write_csv(path, header, body)
    return added, errs


def backfill_box(series_dir, cache_dir, max_reports=14):
    """一次性回补 BOX 季度序列：往回翻 max_reports 份财报，能解析多少算多少。

    每份 MD&A 只给最近 8 个季度，所以回补必须多翻几份。老格式（2022 年那批只有
    'news'/'webcast' 两类文档、正文里根本没有 BOX 表）解析不出来是正常的，
    返回的 errs 里会逐条列明是哪一份、为什么 —— **不静默丢弃**。
    日常 cron 不走这条路（update() 只翻最新 1 份）。
    """
    return _update_box(series_dir, cache_dir, max_reports=max_reports)


def _ensure_columns(csv_path, columns):
    """确保 CSV 表头含 columns 里的每一列，缺的按 columns 的相对次序插进去（单元格留空）。

    新列插在「它在 columns 里的前一个已有列」之后，而不是甩在行尾 —— CSV 是给人翻的，
    mx_oi_cgf/cgz 就该紧挨着 mx_oi_cgb。一列都不缺时**不碰文件**（字节级不变）。
    返回新插入的列名列表。
    """
    if not os.path.exists(csv_path):
        return []
    with open(csv_path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    if not rows:
        raise TmxFetchError('%s 是空文件（连表头都没有）' % csv_path)
    header = rows[0]
    added = []
    for n, col in enumerate(columns):
        if col in header:
            continue
        prev = [c for c in columns[:n] if c in header]
        at = header.index(prev[-1]) + 1 if prev else len(header)
        header.insert(at, col)
        for r in rows[1:]:
            if r:
                r.insert(at, '')
        added.append(col)
    if added:
        _write_csv(csv_path, header, [r for r in rows[1:] if r and r[0].strip()])
    return added


def backfill_mx_columns(series_dir, cache_dir, columns=None):
    """把 MX_SPEC 里**新登记的列**回补到全部历史月份，返回 {列: 这次填了几格}。

    用法（给 MX_SPEC 添完新行之后跑一次）：

        python3 -c "import fetch.tmx as t; print(t.backfill_mx_columns('series','cache'))"

    **为什么非要有这条路**：update() 的 MX 腿只从「CSV 里最后一个有 mx_volume_contracts
    的月份」往后走（那是它该干的事 —— 每月只解析一两份新 xlsx，不是每月重解析 295 份）。
    于是给 MX_SPEC 新增一列时，历史那几百个月它一格都不会碰，新列会只有最新月有值。
    2026-09 补 mx_oi_cgf/cgz 时就是走的这条路。

    与 backfill_index() 的分工：那条走 TMX Money 那个**没有官方契约**的端点，所以明确
    不进无人值守链路；这条不引入任何新来源 —— 同样是 cache/tmx/ 里那批 m-x.ca 的 xlsx、
    同样是 parse_mx_workbook()，只是把已经落地的档案按新的列清单再读一遍。

    安全性照旧靠 _merge：**已有值的单元格一格都不覆盖**，所以
      · 重复跑是幂等的（跑第二遍文件字节级不变）；
      · 不小心传了一个早已填满的列，也只是白读一遍 xlsx，不会改动任何数据。

    只回补**缓存里已有档案**的月份：官方档案不在本地就跳过（计进返回值旁边的打印），
    不去下载 —— 要新档案是 update() 的事，这条路的职责是「用手头的档案补新列」。
    """
    cols = list(columns or [c for c, _b, _s, _l in MX_SPEC])
    unknown = [c for c in cols if c not in COLUMNS]
    if unknown:
        raise TmxFetchError('这些列不在 COLUMNS 里，先往 MX_SPEC 添行：%s' % unknown)

    csv_path = os.path.join(series_dir, 'tmx.csv')
    new = _ensure_columns(csv_path, COLUMNS)
    if new:
        print('· 表头新增列：%s' % ', '.join(new))
    header, body, have = _load_csv(csv_path, COLUMNS, 'month')
    idx = {n: i for i, n in enumerate(header)}

    # 只认「已经有 MX 数据」的月份：现货腿单独建的行（MX 还没发那个月）不该被这条路填。
    months = [m for m in sorted(have) if have[m][idx['mx_volume_contracts']].strip()]
    filled = {c: 0 for c in cols}
    absent = []
    for month in months:
        if not os.path.exists(_mx_path(cache_dir, month)):
            absent.append(month)
            continue
        got = fetch_mx_month(cache_dir, month)
        if got is None:
            absent.append(month)
            continue
        rec, _raw = got
        before = {c: have[month][idx[c]].strip() for c in cols}
        _merge(header, body, have, month, {c: rec[c] for c in cols})
        for c in cols:
            if not before[c] and have[month][idx[c]].strip():
                filled[c] += 1
    _write_csv(csv_path, header, body)
    print('· 扫过 %d 个月（缓存缺档 %d 个%s）；新填格数：%s'
          % (len(months), len(absent), '：' + ', '.join(absent[:5]) if absent else '',
             ', '.join('%s %d' % (c, n) for c, n in filled.items() if n) or '无（已是满的）'))
    return filled


def crosscheck(cache_dir, page_size=300):
    """MX 的两个官方来源对账：m-x.ca xlsx（已入库）vs CTS 新闻稿的 MX 小节（不入库）。

    这不是装饰 —— 它是口径坑 10 的看门狗。两份文件互不相干、由 TMX 不同部门发出，
    逐位相同就说明解析没跑偏。分歧一律以 xlsx 为准（2022-09 那次官方"修订"把 CTS 的数
    改小了 6.36%，而 2210 档 xlsx 的「上月」对照列与 2209 档互证）。

    返回 (对上的月数, 总月数, [(月, 字段, xlsx 值, cts 值, 差)])。
    **不抛异常** —— 官方自己两份文件不一致，不该让 cron 停摆。
    """
    cts = fetch_cts(cache_dir, page_size=page_size)
    same, total, diffs = 0, 0, []
    for month in sorted(cts):
        path = _mx_path(cache_dir, month)
        if not os.path.exists(path):
            continue
        _m, _rec, raw = parse_mx_workbook(path)
        ref = cts[month][3]
        total += 1
        ok = True
        for field, key in (('volume', ('vol', None, 'GRAND TOTAL')),
                           ('oi', ('oi', None, 'GRAND TOTAL'))):
            x, c = raw[key], ref.get(field)
            if c is None or x is None:
                continue
            if abs(x - c) > 0.5:
                ok = False
                diffs.append((month, field, x, c, x - c))
        same += 1 if ok else 0
    return same, total, diffs


if __name__ == '__main__':
    _here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _cache, _series = os.path.join(_here, 'cache'), os.path.join(_here, 'series')
    print('latest_months:', latest_months(_cache))
    print('latest_month :', latest_month(_cache))
    print('added        :', update(_series, _cache))
