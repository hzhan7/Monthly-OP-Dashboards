# -*- coding: utf-8 -*-
"""Euronext（ENX）月度经营指标 —— 无人值守抓取。

Euronext 是本仓唯一一家「多国合并体」：一个法人实体下面挂着 8 个市场 ——
官方月度新闻稿的电头逐字写着 "Amsterdam, Athens, Brussels, Dublin, Lisbon, Milan,
Oslo and Paris"（2026-06 那期实测原文）。每一列数字都是若干个市场之和，
而**这个和的成员随时间变**。所以这份 docstring 里「每一列含哪些市场、从哪个月起含」
比数字本身更要紧 —— 数字抓错下个月就露馅，口径写错可以安静地错三年。

━━ 数据源 ━━
落地页 : https://www.euronext.com/en/investor-relations         锚点 <h2 id="monthly-volumes">
直链   : https://live.euronext.com/sites/default/files/statistics/ir/
         euronext_monthly_historical_volumes.xlsx
         5 个 sheet，Period 行自 2012-01 起逐月往下长。
         ⚠ 字节数与月数**每月都变**，下面这组是快照不是常量：
           220,465 bytes、2012-01 → 2026-07 共 175 个月（2026-08-18 对 cache 副本实测；
           上一次记录是 2026-08-06 的 219,581 bytes / 到 2026-06 / 174 个月）。
         要当前值就自己数：`openpyxl.load_workbook(...)['Equity Markets']` 的 Period 列。
伴生   : .../euronext_latest_month_volumes.xlsx
         47,476 bytes（2026-08-18 实测，同样每月变），只有最新月/上月/去年同月/本季/YTD
         五档，**并且给出官方算好的 ADV**。
         本模块只拿它做「最新月」这一列的对表自检，绝不拿它核对历史（理由见口径坑 6）。
发布日 : https://www.euronext.com/en/investor-relations/financial-information/news?page=N
         月度新闻稿列表，每行一个 <time datetime="2026-07-06T15:45:00Z">06/07/2026</time>。

**文件名固定、不带月份**，每月原地覆盖，永远指向最新一期 —— 与 CME 的 monthly-volume
别名同类。别去猜带月份的直链，那种链接这里根本不存在。

为什么还要先取落地页：不是为了拿文件名（文件名是写死的），是为了确认这个文件**还挂在
官方 IR 页上**。哪天 Euronext 换了文件名，直链多半还能下到一个再也不更新的孤儿文件 ——
那种故障不会报错，只会让序列悄悄停在某个月。落地页正则找不到这条 href 时打醒目警告
但仍然继续（直链本身也可能只是页面模板改了），真正的护栏是下面那一堆结构与恒等式校验。

抓取方式：`urllib.request` 裸奔即可。实测无 Cloudflare / Akamai 挑战、无 JS 渲染、
无登录墙、无 JA3 指纹拦截，`server: CloudFront`。**满足无人值守**。
唯一要注意的是 `www.euronext.com`（Drupal 主站，取新闻列表页用）**连续快速请求会掐连接**，
抛 RemoteDisconnected —— 所以列表页那条路径带 retry + backoff + 请求间隔；
`live.euronext.com` 的 CDN 静态文件没有这个问题。

依赖只有 openpyxl（仓里已有）。

━━ 发布节奏 ━━
每月都发，**没有季末月例外**（不像 SCHW / LPLA 要等季报）。下面这组是一次性普查的结果
（2026-08-06 那次跑）：把新闻列表页翻了 22 页（440 条稿件），
对 **2019-01 → 2026-06 共 90 个数据月逐月找它自己那期稿子，命中 90/90**：

    发布日落在次月第 3 至第 13 天，中位数第 7 天。
    最晚：2024-04 数据 → 2024-05-13（第 13 天，**90 期里仅此一次**）
    最早：第 3 天出现过 4 次（2020-03 / 2020-06 / 2020-08 / 2021-12 的数据月）

给 build/roster.py 的建议是 `LAG = (13, 13)`（两值相同）。
⚠ 闸门提前量必须写成**元组** `EARLY_BY['enx'] = (11, 11)` —— monthly_run.py 的取值处是
`EARLY_BY.get(t, (EARLY, EARLY))[1 if qe else 0]`，写成裸整数 `11` 会在下标那一步
TypeError，**崩掉的是整轮 monthly_run，不只是 enx 这一家**。
为什么是 11 不是 10：13−11 = 次月第 2 天开闸，比实测最小值（第 3 天）早一天；
第 3 天出现过 4 次，不是孤例，零余量迟早会漏。代价只是每月多一两个「还没发」的
HTTP 请求，对方是 220KB 的 CDN 静态文件。

发布日只认**新闻稿列表页的 `<time datetime>`**，记进 series/source_dates.csv。
⚠ 不要去详情页找 JSON-LD 的 `datePublished`：**那个字段在页面上根本不存在**
（详情页只有一个 ld+json 块，内容是面包屑导航）。页面上真正存在的是 <time datetime>。
详情页正文电头（"Amsterdam, Athens, Brussels, Dublin, Lisbon, Milan, Oslo and Paris
– 6 July 2026 –"）作为第二处佐证写进 evidence，取不到就不写，不猜。

━━ 口径坑（按踩坑概率排序）━━

**1. 2025-11 起并入雅典交易所（ATHEX），全表主列在这一个月同时发生断点。**
   并购时间线（本机从 IR 新闻列表页 22 页 440 条标题里逐条读出来的，不是转述）：
   2025-07-31 宣布将发起换股要约 → 2025-10-06 要约启动 → 2025-11-14 拿到希腊资本市场
   委员会批准 → **2025-11-19 宣布要约成功**。所以并表落在 2025-11，与官方脚注写死的
   "…and Euronext Athens since November 2025" 一致；月度稿电头也是从 2025-11 那期
   （2025-12-05 发）起出现 "Athens" 的。
   📌 未找到：「2026-04 完成改名 Euronext Athens」这个日期本机**没能核实**。
   检索路径：IR 新闻列表页 `?page=0..21`（440 条标题）按 rebrand / renam / becomes /
   "Euronext Athens" / integration / migrat 六个关键词过滤 2026 年稿件，**零命中**。
   本模块不依赖这个日期（口径断点只认 2025-11），所以不写进代码，也不要有人回头补上
   一个没核过的日子。
   官方同时在**每一个主指标右侧**配了一列
   表头写死为 `Athex` 的备注列，语义随月份翻转：

     · 2025-10 及以前：主列**不含** Athex，备注列 = Athex 单独数
       ⇒ 主列 + 备注列 = 官方口径的 pro-forma（可比口径）
     · 2025-11 及以后：主列**已含** Athex，备注列 = 主列里属于 Athex 的那一块
       ⇒ 主列 − 备注列 = legacy Euronext（旧口径）

   本模块把主列与 Athex 备注列**都写进 CSV**（`athex_*` 前缀），让 build 层自己决定
   画哪条。只写主列 = 把断点焊死在数据里，之后谁也修不回来。
   跨 2025-11 的同比**不可直接比**，图上必须画红色断点竖线。

   验证（本机实测，非引用）：官方 Q2 2026 业绩稿第 13 页明写 "Q2 2025 volumes are
   including Euronext Athens on a pro forma basis"，其 Q2 2025 备考值
   股指 10,796,110 / 单股 22,791,315，与 xlsx「主列+备注列」**一位不差**；
   Q2 2026 主列本身 10,212,541 / 25,104,143，与官方当期**一位不差**。两个方向都能精确重建。

**2. 单股衍生品是全表最危险的一列：Athex 占并表后的 90-98%，且有季度换月脉冲。**
   实测主列 Individual Equity Futures 月合计张数：2025-10 = 35,573 → 2025-11 = 836,511
   （其中 Athex 781,183，占 93.4%）→ 2026-06 = 5,057,868（其中 Athex 4,958,445，占 98.0%）。
   两件事同时成立：(a) 2025-11 那一格是 **20 倍以上的口径断点**，不是业务增长；
   (b) 并表后这条线在 3/6/9/12 月出现 5-7 倍脉冲（希腊单股期货被当作融券/回购替代品
   按季滚动），不是成交活跃度信号。⇒ 同比一律用 pro-forma（主+备注）口径；
   与 Cboe 的 multilist options 对比必须先取 legacy（主−备注）；year_lines 类图对它无意义。
   相比之下 Athex 单股**期权**可忽略（几百到 1 万张/月）。

**3. 三个更早的并表断点，且现货/衍生品/上市统计三套序列的断点月份各不相同。**
   官方脚注原文（从 xlsx 直接读的，不是转述）：

     现货     (Equity Markets 脚注 3)：Dublin since January 2017, Oslo since January 2018,
              Borsa Italiana since May 2021, Euronext Athens since November 2025
     衍生品   (Equity Markets 脚注 5)：Oslo since July 2019, Borsa Italiana since May 2021,
              Euronext Athens since November 2025
     上市统计 (Capital Markets 脚注 1)：Dublin and Oslo since January 2019,
              Borsa Italiana since May 2021, Euronext Athens since November 2025
     商品     (FICC 脚注 3)：Oslo Bors since July 2019
     固收现券 (FICC 脚注 2)：同现货那一套
     CSD      (Securities Services 脚注 1)：Euronext Athens since November 2025

   ⇒ 现货 ADV 图的断点竖线是 **2017-01 / 2018-01 / 2021-05 / 2025-11**；
     衍生品是 **2019-07 / 2021-05 / 2025-11**（外加电力衍生品首月 2026-03）；
     上市统计是 **2019-01 / 2021-05 / 2025-06 / 2025-11**。
     三套不要混用同一组竖线 —— 把 2019-01 当成现货断点是常见错误，现货那年没有断点。

**4. 这份 xlsx 是「按今天口径重述过的」序列，与当年新闻稿印出来的数字对不上，方向还会翻。**
   实测同一指标三个时点：2019-01 现货 ADV xlsx 7,140.4 €m vs 当年稿 6,708.1 €m（+6.4%，
   往上重述：xlsx 把 Oslo 从 2018-01 就算进去了，而 Oslo Børs 2019-06 才完成收购）；
   2020-06 现货月成交额 xlsx 234,385.7 €m vs 当年稿 244,406.7 €m（−4.1%，往下重述）。
   同一次测试里**衍生品张数一格不差**（2020-06 六个 futures/options 单元格全部完全相等）。
   ⇒ 结论不是「xlsx 错了」，而是 xlsx 内部自洽、当年新闻稿之间不自洽。本仓只认 xlsx，
   且**绝不能**把某一期新闻稿的数字手工补进序列 —— 那会插进一个 4-6% 的假台阶。
   本模块因此对已入库的值**永不覆盖**，冲突写 cache/enx_restatements.csv 供人工判断。

**5. FX 那一列的表头单位是错的。**
   FICC Markets 第 9 行写 `Volume (in M$, single counted)`，但格子里 2026-06 是
   `671602324739`。若真是百万美元，日均就成了 30 万亿美元。拿 2019-01 新闻稿的
   "$20,050 million" 反推：441,099,188,988.6 / 22 / 1e6 = 20,049.96 → **该列是绝对美元**。
   Q2 2026 再验：/65/1e9 = 28.9816 $bn vs 官方 "ADV Euronext FX 28,982 $m"，相对差 1.25e-05。
   ⇒ 除以 1e9 得 $bn。不要相信表头。（同一序列在伴生的 latest 文件里**真的是 M$**，见坑 6。）

**6. 伴生的 `euronext_latest_month_volumes.xlsx` 只能核对「最新月」这一列，不能核对历史。**
   两个官方文件对 Athens 并表用**不同基准**：latest 的脚注写
   "Includes figures from Euronext Athens since January 2025"，hist 写 "since November 2025"。
   实测 2025-06：latest 的单股衍生品 6,908,289 = hist「主列+Athex」，而 hist 主列只有
   5,630,914（差 22.7%）。拿 latest 去核对 2025-11 之前的月份会看到最高 23% 的假失配，
   然后很可能去「修」一个没坏的解析器。
   ⇒ 本模块的 `_crosscheck_latest_month()` **只比最新月那一列**，且先核对文件自报的月份。
   另外同一个 FX 序列 latest 是 M$、hist 是绝对 $，两文件绝不共用单位常量。
   latest 的第 5 张 sheet `Nord Pool` 是 2020 年的死残留（格子里写字面量 "xxx" / "xx%"），
   本模块根本不碰它。

**7. 「Commodity」是农产品，不是能源；电力是另一套，且分两层、三个不同的日数分母。**
   Euronext 的 commodity derivatives = 巴黎 MATIF 的小麦/玉米/菜籽。
   能源侧是 Nord Pool，拆成 **现货电力**（Day-ahead / Intraday，单位 TWh，**买卖双边计**，
   2020-01 起，分母是**自然日** 30/31）与 **电力衍生品**（Notional Volume/OI，单位 GWh，
   **2026-03-16 才全面上线**，分母是交易日，2026-03 只有 12 天）两块。
   ⇒ 跨家比价时 `adv_commodity_*_kcontracts` 的对手是 CME 的 `adv_ag_kcontracts`，
   **不是** `adv_energy_kcontracts`。

**8. 同一张表里混着单边计与双边计，每引用一列都要回表头看分组行。**
   现货金额 `Trading volume (single counted)`；现货笔数 `Transactions (buy and sell)`
   （**双边**，且含 reported trades，官方 Q2 稿原文 "reported trades included"）；
   股权清算 `Clearing volume (single counted)`；债券清算 `Clearing volume (double counted)`；
   Nord Pool 电力 `Volume (in TWH, buy and sell)`（**双边**）。
   本模块给每一列都在下面 COLUMN_SPEC 的注释里标了单双边，CSV 列名不带这个信息，
   查列名 → 回这里查口径。

**9. 表结构是「两层表头 + 合并单元格 + 同名标签重复出现」，必须按分组 + 标签定位。**
   `Futures` / `Options` / `Athex` / `Nb of trading days` / `Total` / `Equities`
   这些标签在同一张 sheet 里各出现 2-8 次，只有靠上面的分组行才能区分是股指还是单股、
   是成交量还是 OI。本模块的做法：用**合并单元格范围**还原每一列头上盖着的分组文字，
   拼成 (分组, 小节, 标签) 三元组去唯一定位，**绝不写死列号、也绝不全表 grep 标签**。
   分组标题带脚注编号（`Commodity derivatives (3)`、`Turnover Equities (1)`、`Period (1)`），
   `_lab()` 只剥**结尾**的 `(数字)` / `(R)` —— 不能全剥，`TA(1) MTS Repo` 的括号在中间，
   `Bonds wholesale (in EUR bln)` 的括号是单位。

**10. `Nb of trading days` 有 7 个，各管各的分母，且其中一个根本不是交易日。**
   Equity Markets 两个（现货 C3 / 股权衍生品 C16，2026-08-18 实测全部 175 个月两列相等，
   但官方保留成
   两列，本模块也存两列，不做「反正相等」的偷懒）；FICC 五个（固收 C3 / 商品 C13 /
   电力 C19 / 电力衍生品 C23 / FX C28）。其中**电力那个是自然日**（2026-04=30、
   2026-05=31，Q2 合计 91 = 30+31+30），官方 Q2 稿也印 91；FX 那个与现货不同
   （2026-04 现货 20 天、FX 22 天）。
   定位规则：某个分组的日数列 = **紧邻该分组左侧的那个 `Nb of trading days` 列**。
   这条规则也顺带解释了为什么单股衍生品没有自己的日数列 —— 它左边最近的那个就是 C16。

**11. 四张 sheet 的 Period 语义不一致，必须按 (年, 月) 归并。**
   Equity / FICC / Securities Services 的 Period 是每月 1 日；
   **Capital Markets 不是** —— 实测 2018-01-05、2018-02-02、2018-03-02…（像是每月首个周五），
   到 2026 年才变成月初。按精确日期 join 会整段对不上。

**12. `Capital Markets` 的 `Funds` 列 2018 全年是字面量字符串 `'NA'`。**
   这是整个工作簿里除死 sheet 之外**唯一**的非数值污染。`float(cell)` 会 ValueError，
   或者把 "NA" 原样写进 CSV。⇒ `listed_funds` 的起始月是 **2019-01**，不是 2018-01。

**13. 两张 sheet 是死残留，解析到会炸或写出垃圾 —— 所以白名单四张 sheet，不 for-each-sheet。**
   hist 的第 5 张 `Checkup`：唯一的数据列整列是 `#REF!` 字符串（117 行），
   表头却写着 "Euronext Cash / Turnover in millions euros"，看上去像正经数据。
   latest 的第 5 张 `Nord Pool`：字面量 "xxx" / "xx%"，停在 2020-01。

**14. 官方对上市统计做过两次口径扩大，且已回溯重述。**
   脚注 (3)：2025-06 起 `Bonds` 计入 Euronext ABM，2024 与 2025 已重述；
   脚注 (4)：2025-06 起 `Nb of Listings` 改为「所有类型的挂牌」（含私募配售、直接上市、
   市场间转板、反向并购、de-SPAC、二次上市）；脚注 (2)：2021-05 改过发行人家数的计算方法。
   三条都是「已重述」，序列内部自洽，但与 2024 年当时读到的家数/募资额对不上，
   且 `new_listings_equities` 在 2025-06 有一次口径抬升，同比要注意。

**15. 文件的 Last-Modified 不是发布日，会被重述推后。**
   实测 `Last-Modified: Thu, 16 Jul 2026 08:20:12 GMT`，而 2026-06 那期新闻稿是
   **2026-07-06** 发的 —— 文件在发布 10 天后被原地重传过。
   ⇒ source_dates 只认新闻稿的 <time datetime>，且**首次摄入某月时记一笔、事后永不覆盖**。

**16. 新闻稿标题不遵守模板，按模板拼 slug 或做全等匹配必漏。**
   2023-03 那期的标题是 "Euronext announces highest cash volumes in a year in March 2023"，
   slug 是 `euronext-announces-highest-cash-volumes-year-march`；
   拼出来的 `...-for-march-2023` 与 `...-for-march-2023-0` **双双 404**。
   市场部随时会为「创纪录」的月份改标题。⇒ 一律从**列表页**拿真 href 与真日期，
   标题用宽松匹配（announces … volume … <月> <年>），**且允许某个月取不到发布日而不抛异常**
   （仓库 source_dates.py 的原则：拿不到就让它缺席，缺席远好过印一个像模像样的错日期）。

**17. 月度新闻稿从 2020 年起正文里已经没有任何数字了。**
   2019-01 那期有完整正文数字；2020-07-03 那期正文只剩一句
   "Monthly and historical volumes table are available at this address"（统计数字在一个
   独立的附件 PDF 里）；2022 年以后连附件都没有。
   ⇒ **不要**把新闻稿当数据源去解析，它在本模块里的唯一价值就是发布日。

**18. 序列长度不齐：市值 / CSD 只到 2022-01，比主序列短十年。**
   不是断档，是官方本来就只提供这么长。逐列起始月见 COLUMN_SPEC 的 since 字段
   （那是本机实测出来的，不是抄的）。画在同一张图里会出现左半边空白，
   要么单独成图、要么图注写明起点。
   ⚠ **这些起点没有一个是抓取窗口** —— 本模块只下一份滚动全历史 xlsx 并遍历它的
   全部 Period 行，压根没有窗口这个概念。起点全部是业务史（复核过的日期）：
     · 2020-01 Nord Pool 现货电力 —— 2020-01-15 完成收购 66% 股权，自 01-16 并表，
       这是唯一一处「起点 = 交割月」；
     · 2020-01 MTS —— MTS 随 Borsa Italiana Group 进来，那笔交易 2021-04-29 才交割，
       序列却回填到 2020-01（早 15 个月）⇒ 是官方回填，不是并表日；
     · 2013-01 Euronext FX —— 前身 FastMatch，2017-08-14 完成收购约 90% 股权，回填到
       被收购方自己的历史；
     · 2021-01 athex_* 备注列 —— 雅典换股要约 2025-11-19 宣告成功（接纳期 11-17 截止、
       11-24 交割），官方把备注列回填到 2021-01；
     · 2022-01 清算 / 市值 / CSD —— **披露起点，不是事件日期**。业务前提是 2021-04-29
       Borsa Italiana Group 交割带来 Euronext Clearing（原 CC&G）与米兰 CSD Monte Titoli；
       在那之前只有波尔图 Interbolsa、奥斯陆 VPS（2019-06-18 交割）与哥本哈根
       VP Securities（2020-08-04 交割），米兰缺位 ⇒ 四家齐备最早只能到 2021-05，
       官方却从 2022-01 才按月发。CSD 的 Total 列到 2025-11 才含雅典（脚注 (1) 原文）。
     · 2026-03 电力衍生品 —— 脚注 (5) 原文写死 "fully operational on 16 March 2026"。
   ⇒ 这些起点**不需要回补**：更早的官方**月度**数据不存在。Fact Book / 年报里那些年度数
   的成员范围与定义与本表都不是一回事，硬接上去只会在 2018-01 或 2022-01 造出假台阶 ——
   要接必须先逐项证明口径相同，没证明就不接。

━━ 产出的两个文件 ━━
· `series/enx.csv` —— 72 个字段（month + 71 个数据列）× 每月一行，2012-01 起
  （2026-08-18 实测 175 行、到 2026-07；行数随每月更新增长，别把它当常量）。列序是
  交易日 → 主列（按官方表的顺序）→ 全部 `athex_*` 备注列。每一列的确切口径写在
  下面 COLUMN_SPEC 的行末注释里（张数/金额、日均/月总/月末时点、单边/双边、币种）。
· `series/enx_breaks.csv` —— 口径断点台账，**由本模块从官方脚注原文自动抽取**，
  92 行，列是 `column, break_month, footnote, athex_memo_column, official_footnote`。
  它回答的是「哪一列在哪个月与左侧不可比、以及有没有备注列能把断点消掉」。
  2025-11 那一批 26 行就是 Athens 并表影响到的全部列。
  这个文件**是 build 的输入**：`build/specs/enx.py` 的 `_read_breaks()` 每次 import 都读它，
  取出 (column, break_month) 逐列挂断点 —— 页上那些红色竖线的月份与「画在哪张图上」
  全部由这张表决定，spec 里只留一张「月份 → 中文说法」的翻译表，**月份一个都不写死**。
  ⇒ 删了它，2025-11 那条红线该画在哪几张图上就只剩人脑记忆，而人脑记不住 26 列。
  它同时也是「官方又并购了一家」的探测器 —— 官方改脚注，这张表下次跑就跟着变，
  git diff 里看得见。

━━ 没有入库的列（写清楚免得后人以为漏了）━━
· CSD 的五地分拆（Athens / Copenhagen / Milan / Oslo / Porto）：只入了 Total 与 Athens
  （Athens 是断点备注列，必须有），另外四地是地理明细，与本仓的横截面叙事无关。
· `Money Raised - Bonds`：官方季报有这一行，**月度 xlsx 里没有**，无法逐月入库。
· `Total Euronext` 这个历史概念：2020 年的稿子里它 = 三大类 + 一个已经不存在的
  `TM Derivatives` 桶（Oslo 的一个衍生品桶，2020-06 那期 254,784 张）。
  今天的官方季报附录里也没有这一行了 ⇒ 不要用它去校验本模块算出的衍生品总量。
"""

import collections
import csv
import datetime
import io
import os
import re
import time
import urllib.request

import openpyxl

# ══════════════════════════════════════════════════════════════════════
# 源地址
# ══════════════════════════════════════════════════════════════════════
LANDING_URL = 'https://www.euronext.com/en/investor-relations'
STATS_BASE = 'https://live.euronext.com/sites/default/files/statistics/ir/'
HIST_NAME = 'euronext_monthly_historical_volumes.xlsx'
LATEST_NAME = 'euronext_latest_month_volumes.xlsx'
NEWS_URL = ('https://www.euronext.com/en/investor-relations/'
            'financial-information/news?page=%d')
SITE_ROOT = 'https://www.euronext.com'

# live.euronext.com 实测连默认的 python-urllib UA 都放行；带常规 UA 是零成本的保险。
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

# 白名单四张 sheet，见口径坑 13。绝不 for-each-sheet。
SHEETS = ('Equity Markets', 'FICC Markets', 'Capital Markets',
          'Securities Services')


# ══════════════════════════════════════════════════════════════════════
# 表结构
# ══════════════════════════════════════════════════════════════════════
S_EQ, S_FICC, S_CAP, S_SEC = SHEETS

# 分组行（合并单元格）的文字，已按 _lab() 剥掉尾部脚注编号
G_CASH = 'Cash Markets (Fixed Income excluded)'
G_IDX = 'Equity Index derivatives'
G_SS = 'Individual Equity derivatives'
G_FI = 'Fixed Income Markets'
G_COM = 'Commodity derivatives'
G_PWR = 'Power trading'
G_PWRD = 'Power trading derivatives'
G_FX = 'FX trading'
G_CSD = 'Central Securities Depositary'

# 小节行（第二层表头）的文字 —— 单双边计数就写在这里，见口径坑 8
U_TRADES = 'Transactions (buy and sell)'          # 双边
U_TURNOVER = 'Trading volume (single counted)'    # 单边，Equity Markets，单位 €m
U_CLEAR1 = 'Clearing volume (single counted)'     # 单边
U_LOTS = 'Volume (in lots)'
U_OI = 'Open Interest (in lots)'
U_FI_TV = 'Trading volume (in M€, single counted)'  # 单边，FICC，单位 €m
U_CLEAR2 = 'Clearing volume (double counted)'     # 双边
U_TWH = 'Volume (in TWH, buy and sell)'           # 双边
U_GWH_V = 'Notional Volume (in GWH)'
U_GWH_OI = 'Notional Open Interest (in GWH)'
U_FX = 'Volume (in M$, single counted)'           # 表头单位是错的，见口径坑 5
U_AUC = 'AuC (in EUR bln)'
U_SETTLE = 'Nb of Settlement instructions'

# 交易日列：key -> 用来定位它的分组（该分组左邻的那个 Nb of trading days 列），见口径坑 10
DAYS_ANCHOR = {
    'cash': (S_EQ, G_CASH),
    'eqderiv': (S_EQ, G_IDX),
    'fixedincome': (S_FICC, G_FI),
    'commodity': (S_FICC, G_COM),
    'power': (S_FICC, G_PWR),
    'powerderiv': (S_FICC, G_PWRD),
    'fx': (S_FICC, G_FX),
}

# (csv 列名, days key, 起始月)。起始月是本机对当前这份 xlsx 逐列实测出来的首个有数月。
DAYS_SPEC = [
    ('trading_days_cash', 'cash', '2012-01'),
    ('trading_days_eqderiv', 'eqderiv', '2012-01'),
    ('trading_days_fixedincome', 'fixedincome', '2012-01'),
    ('trading_days_commodity', 'commodity', '2012-01'),
    ('days_power_calendar', 'power', '2020-01'),      # 自然日，不是交易日
    ('trading_days_powerderiv', 'powerderiv', '2026-03'),
    ('trading_days_fx', 'fx', '2013-01'),
]

Col = collections.namedtuple(
    'Col', 'name sheet heads label days scale since memo memo_since')


def _c(name, sheet, heads, label, days, scale, since, memo=None, memo_since=None):
    return Col(name, sheet, tuple(heads), label, days, float(scale), since,
               memo, memo_since)


# 每一列的确切口径写在行末注释里：张数还是金额、日均还是月总还是月末、单边还是双边、
# 本币还是美元、含哪些市场。下游 build/notional.py 的换算全靠这些注释，写错比数字错更难发现。
#
# 「含哪些市场」的通则（逐列不再重复）：主列 = 巴黎 + 阿姆斯特丹 + 布鲁塞尔 + 里斯本
#   + 都柏林（现货 2017-01 起 / 上市 2019-01 起）+ 奥斯陆（现货 2018-01 起 /
#   衍生品与商品 2019-07 起 / 上市 2019-01 起）+ 米兰（2021-05 起）
#   + 雅典（**2025-11 起**）。`athex_*` 备注列见口径坑 1。
COLUMN_SPEC = [
    # ── 现货（Equity Markets / Cash Markets）────────────────────────────
    # 日均成交笔数（千笔/日）。**买卖双边计**，含 reported trades。
    _c('adv_cash_trades_k', S_EQ, (G_CASH, U_TRADES), 'Total number of trades',
       'cash', 1e3, '2012-01', 'athex_adv_cash_trades_k', '2021-01'),
    # 日均成交名义额（€bn/日）。**单边计**。含股票+投资基金+ETF+结构化产品。
    _c('adv_cash_adnv_eurbn', S_EQ, (G_CASH, U_TURNOVER), 'Total Turnover',
       'cash', 1e3, '2012-01', 'athex_adv_cash_adnv_eurbn', '2021-01'),
    # 同上，只含股票与投资基金。**与 Cboe Europe 的 adv_eu_equities_adnv_eurbn 对比用这一列**
    # （Cboe 那列不含结构化产品），是全仓最干净的一对同口径可比字段。
    _c('adv_cash_equities_adnv_eurbn', S_EQ, (G_CASH, U_TURNOVER),
       'Turnover Equities', 'cash', 1e3, '2012-01',
       'athex_adv_cash_equities_adnv_eurbn', '2021-01'),
    # ETF 现货日均成交额（€bn/日，单边）。2015-01 起 ETC 从结构化产品挪进这一列。
    _c('adv_cash_etf_adnv_eurbn', S_EQ, (G_CASH, U_TURNOVER), 'Turnover ETF',
       'cash', 1e3, '2012-01', 'athex_adv_cash_etf_adnv_eurbn', '2021-01'),
    # 结构化产品现货（€bn/日，单边）。入库不是为了画图，是为了每月撞恒等式
    # Total ≡ Equities + ETF + Structured（见 _validate），撞得上说明四列一格没错行。
    _c('adv_cash_structured_adnv_eurbn', S_EQ, (G_CASH, U_TURNOVER),
       'Turnover Structured Products', 'cash', 1e3, '2012-01'),
    # Euronext Clearing 清算的股票交易笔数/手数（千/日，**单边**）。官方标签
    # "Shares (nb of contracts)"，季报里叫 "number of transactions and lots cleared"，
    # 与成交额不是同一层；值带小数（例：2026-06 月合计 26,166,723.5），不是纯计数。
    _c('adv_shares_cleared_kcontracts', S_EQ, (G_CASH, U_CLEAR1),
       'Shares (nb of contracts)', 'cash', 1e3, '2022-01',
       'athex_adv_shares_cleared_kcontracts', '2022-01'),

    # ── 股指衍生品（CAC 40 / AEX / BEL 20 / FTSE MIB / OBX / ATHEX 等）──
    # 日均张数（千张/日）。乘数各不相同（CAC 40 期货 €10/点），跨家只能指数化比。
    _c('adv_index_futures_kcontracts', S_EQ, (G_IDX, U_LOTS), 'Futures',
       'eqderiv', 1e3, '2012-01', 'athex_adv_index_futures_kcontracts', '2021-01'),
    _c('adv_index_options_kcontracts', S_EQ, (G_IDX, U_LOTS), 'Options',
       'eqderiv', 1e3, '2012-01', 'athex_adv_index_options_kcontracts', '2021-01'),
    # 月末未平仓（千张，**月末时点，不除交易日**）
    _c('oi_index_futures_kcontracts', S_EQ, (G_IDX, U_OI), 'Futures',
       None, 1e3, '2012-01', 'athex_oi_index_futures_kcontracts', '2021-01'),
    _c('oi_index_options_kcontracts', S_EQ, (G_IDX, U_OI), 'Options',
       None, 1e3, '2012-01', 'athex_oi_index_options_kcontracts', '2021-01'),

    # ── 单股衍生品 ⚠ 全表最危险的一列，见口径坑 2 ──────────────────────
    _c('adv_singlestock_futures_kcontracts', S_EQ, (G_SS, U_LOTS), 'Futures',
       'eqderiv', 1e3, '2012-01',
       'athex_adv_singlestock_futures_kcontracts', '2021-01'),
    _c('adv_singlestock_options_kcontracts', S_EQ, (G_SS, U_LOTS), 'Options',
       'eqderiv', 1e3, '2012-01',
       'athex_adv_singlestock_options_kcontracts', '2021-01'),
    _c('oi_singlestock_futures_kcontracts', S_EQ, (G_SS, U_OI), 'Futures',
       None, 1e3, '2012-01', 'athex_oi_singlestock_futures_kcontracts', '2021-01'),
    _c('oi_singlestock_options_kcontracts', S_EQ, (G_SS, U_OI), 'Options',
       None, 1e3, '2012-01', 'athex_oi_singlestock_options_kcontracts', '2021-01'),

    # ── 固收（FICC / Fixed Income Markets）─────────────────────────────
    # MTS 现券：欧洲主权债电子交易，日均成交额 €bn/日，**单边**。
    _c('adv_mts_cash_eurbn', S_FICC, (G_FI, U_FI_TV), 'MTS Cash',
       'fixedincome', 1e3, '2020-01'),
    # MTS 回购**未经期限调整**的日均量（€bn/日，单边）。官方主口径是下面那条 TAADV，
    # 两条都在表里，别混。
    _c('adv_mts_repo_eurbn', S_FICC, (G_FI, U_FI_TV), 'MTS Repo',
       'fixedincome', 1e3, '2020-01'),
    # Term Adjusted 回购日均量（€bn/日，单边）—— 官方季报印的就是这条（TAADV MTS Repo）。
    _c('taadv_mts_repo_eurbn', S_FICC, (G_FI, U_FI_TV), 'TA(1) MTS Repo',
       'fixedincome', 1e3, '2020-01'),
    # MTS 以外的债券成交（Euronext 各地债券市场，2025-11 起含 Athex），量级小，留 €m/日。
    _c('adv_other_fixed_income_eurm', S_FICC, (G_FI, U_FI_TV), 'Bonds',
       'fixedincome', 1, '2012-01', 'athex_adv_other_fixed_income_eurm', '2021-01'),
    # 债券批发清算名义额（€bn/日，**双边计**）
    _c('adv_bonds_wholesale_cleared_eurbn', S_FICC, (G_FI, U_CLEAR2),
       'Bonds wholesale (in EUR bln)', 'fixedincome', 1, '2022-01'),
    # 债券零售清算（千张/日，**双边计**）
    _c('adv_bonds_retail_cleared_kcontracts', S_FICC, (G_FI, U_CLEAR2),
       'Bonds retail (nb of contracts)', 'fixedincome', 1e3, '2022-01',
       'athex_adv_bonds_retail_cleared_kcontracts', '2022-01'),

    # ── 商品衍生品 = 巴黎 MATIF 农产品（小麦/玉米/菜籽），**不是能源**，见口径坑 7 ──
    _c('adv_commodity_futures_kcontracts', S_FICC, (G_COM, U_LOTS), 'Futures',
       'commodity', 1e3, '2012-01'),
    _c('adv_commodity_options_kcontracts', S_FICC, (G_COM, U_LOTS), 'Options',
       'commodity', 1e3, '2012-01'),
    _c('oi_commodity_futures_kcontracts', S_FICC, (G_COM, U_OI), 'Futures',
       None, 1e3, '2012-01'),
    _c('oi_commodity_options_kcontracts', S_FICC, (G_COM, U_OI), 'Options',
       None, 1e3, '2012-01'),

    # ── Nord Pool 现货电力：日均 TWh，**买卖双边计**，分母是自然日不是交易日 ──
    _c('adv_power_dayahead_twh', S_FICC, (G_PWR, U_TWH), 'Day-ahead',
       'power', 1, '2020-01'),
    _c('adv_power_intraday_twh', S_FICC, (G_PWR, U_TWH), 'Intraday',
       'power', 1, '2020-01'),

    # ── Nord Pool 电力衍生品：2026-03-16 全面上线，序列因此从 2026-03 才起 ────
    #    （官方 FICC Markets 脚注 (5) 原文："Power derivatives market became fully
    #     operational on 16 March 2026"。有几个月随每月更新增长，不写死。）
    # 日均名义量（GWh/日）
    _c('adv_power_systemprice_futures_gwh', S_FICC, (G_PWRD, U_GWH_V),
       'System price futures', 'powerderiv', 1, '2026-03'),
    _c('adv_power_epad_futures_gwh', S_FICC, (G_PWRD, U_GWH_V),
       'EPADs futures', 'powerderiv', 1, '2026-03'),
    # **月末名义未平仓（GWh，时点值，不除天数）** —— 官方 Q2 2026 季报印的就是这条
    _c('oi_power_deriv_notional_gwh', S_FICC, (G_PWRD, U_GWH_OI),
       'Total Notional Open interest', None, 1, '2026-03'),

    # ── Euronext FX（原 FastMatch）即期外汇：$bn/日，**单边**。原始单元格是绝对美元 ──
    _c('adv_fx_spot_usdbn', S_FICC, (G_FX, U_FX), 'Spot volume', 'fx', 1e9, '2013-01'),

    # ── 上市与募资（Capital Markets）───────────────────────────────────
    # 月末股票发行人家数（时点值）。2021-05 改过计算方法（脚注 2）。
    _c('issuers_equities', S_CAP, ('Nb of Issuers',), 'Equities',
       None, 1, '2018-01', 'athex_issuers_equities', '2021-01'),
    # 月末上市债券只数。2025-06 起含 Euronext ABM，官方已重述 2024-2025（脚注 3）。
    _c('listed_bonds', S_CAP, ('Nb of Listed Instruments',), 'Bonds',
       None, 1, '2018-01', 'athex_listed_bonds', '2021-01'),
    _c('listed_etfs', S_CAP, ('Nb of Listed Instruments',), 'ETFs',
       None, 1, '2018-01', 'athex_listed_etfs', '2021-01'),
    # ⚠ 2018 全年是字面量 'NA'，起始月 2019-01，见口径坑 12。官方无 Athex 备注列。
    _c('listed_funds', S_CAP, ('Nb of Listed Instruments',), 'Funds',
       None, 1, '2019-01'),
    # **当月**新增挂牌家数（月度总量，不是时点也不是日均）。
    # 2025-06 起口径扩大到「所有类型的挂牌」（脚注 4），跨那个月的同比不可直接比。
    _c('new_listings_equities', S_CAP, ('Nb of Listings',), 'Equities',
       None, 1, '2018-01', 'athex_new_listings_equities', '2023-01'),
    # **当月**新上市募资额（€m，月度总量，含超额配售）
    _c('money_raised_new_listings_eurm', S_CAP, ('Money Raised (mln of €)',),
       'Equities - New Listings', None, 1, '2018-01',
       'athex_money_raised_new_listings_eurm', '2023-01'),
    # **当月**股票再融资募资额（€m，月度总量）
    _c('money_raised_followon_eurm', S_CAP, ('Money Raised (mln of €)',),
       'Equities - Follow-ons', None, 1, '2018-01',
       'athex_money_raised_followon_eurm', '2023-01'),
    # 月末总市值（万亿欧元，时点值）。**只到 2022-01**，比主序列短十年。
    _c('mktcap_eurtn', S_CAP, ('Market cap. (trillion of €)',),
       'Total end of month', None, 1, '2022-01', 'athex_mktcap_eurtn', '2022-01'),

    # ── 结算与托管（Securities Services，五家 CSD）─────────────────────
    # 月末托管资产（€bn，时点值）。**只到 2022-01**。备注列表头是 'Athens' 不是 'Athex'。
    _c('csd_auc_eurbn', S_SEC, (G_CSD, U_AUC), 'Total',
       None, 1, '2022-01', 'athex_csd_auc_eurbn', '2022-01'),
    # **当月**结算指令笔数（百万笔，月度总量）。**只到 2022-01**。
    _c('csd_settlement_instructions_m', S_SEC, (G_CSD, U_SETTLE), 'Total',
       None, 1e6, '2022-01', 'athex_csd_settlement_instructions_m', '2022-01'),
]

# CSV 列序：交易日 → 主列（按源表顺序）→ 全部 athex_* 备注列。
# 备注列集中放在末尾而不是紧跟各自的主列，是为了让「主序列」那一段能一眼读完；
# 代价是与 xlsx 逐列对照时要跳一下，这个代价由 COLUMN_SPEC 里的成对定义抵消。
COLUMNS = ([n for n, _k, _s in DAYS_SPEC]
           + [c.name for c in COLUMN_SPEC]
           + [c.memo for c in COLUMN_SPEC if c.memo])

# 每一列的起始月（校验用）。晚于起始月还为空 = 解析出错，抛异常；早于起始月为空 = 官方就没有。
SINCE = dict([(n, s) for n, _k, s in DAYS_SPEC]
             + [(c.name, c.since) for c in COLUMN_SPEC]
             + [(c.memo, c.memo_since) for c in COLUMN_SPEC if c.memo])

# 「这个月真的有数据」的锚：现货 ADNV。官方有时会把下个月的行先开出来只填交易日
# （HKEX 教训的同类问题），所以不能用「表里最后一行」当最新月。
ANCHOR = 'adv_cash_adnv_eurbn'

# 结构性口径断点，写在这里供 build 层取用（画红色竖线），也供人查。见口径坑 1、3。
BREAKS = {
    'cash': ['2017-01', '2018-01', '2021-05', '2025-11'],
    'derivatives': ['2019-07', '2021-05', '2025-11'],
    'listing': ['2019-01', '2021-05', '2025-06', '2025-11'],
    'power_derivatives': ['2026-03'],
}


class EnxFetchError(RuntimeError):
    """源站结构变化 / 下载失败 / 解析结果不完整 / 内部恒等式不成立。

    一律炸掉。宁可整月不更新（线上留着自己的旧数据），也绝不静默写空列或 NaN。
    """


# ══════════════════════════════════════════════════════════════════════
# 网络
# ══════════════════════════════════════════════════════════════════════
def _http_get(url, timeout=90, tries=1, pause=1.5):
    """取一个 URL 的原始字节。tries>1 时带退避重试。

    www.euronext.com 是 Drupal 主站，实测连打约 20 个请求之后会
    `RemoteDisconnected: Remote end closed connection without response`
    —— 不是封禁，是连接被掐。所以走主站的调用方一律传 tries=3；
    live.euronext.com 的 CDN 静态文件没有这个毛病，用默认的 tries=1。
    """
    last = None
    for k in range(tries):
        req = urllib.request.Request(url, headers={
            'User-Agent': _UA,
            'Accept': '*/*',
            'Accept-Language': 'en-US,en;q=0.9',
        })
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read(), dict(r.headers)
        except Exception as e:                            # noqa: BLE001
            last = e
            if k + 1 < tries:
                time.sleep(pause * (k + 1))
    raise EnxFetchError('下载失败 %s: %r' % (url, last)) from last


def _write_bytes(path, data):
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    with open(path, 'wb') as f:
        f.write(data)


def _check_landing(cache_dir):
    """确认历史 xlsx 还挂在官方 IR 页上；返回 True/False，**不抛异常**。

    文件名是写死的，所以这一步不是为了「发现文件名」，而是为了发现**文件名变了**：
    直链在改名之后多半还能下到一个再也不更新的孤儿文件，那种故障不报错、
    只会让序列悄悄停在某个月。这里查不到就打醒目警告继续跑 —— 真正的护栏是
    下面那一堆结构与恒等式校验，以及 latest 文件的当月对表。
    """
    try:
        html, _h = _http_get(LANDING_URL, tries=3)
    except EnxFetchError as e:
        print('[enx] 警告：落地页取不到（%r），跳过链接存在性检查' % e)
        return False
    _write_bytes(os.path.join(cache_dir, 'enx_ir_landing.html'), html)
    txt = html.decode('utf-8', 'replace')
    # 全局正则捞 href：<h2 id="monthly-volumes"> 与下载列表分处两个 Drupal container，
    # 中间隔着约 600 字符模板标记，按锚点作用域去扫会抓空。
    hits = set(re.findall(
        r'https://live\.euronext\.com/sites/default/files/statistics/ir/'
        r'[A-Za-z0-9_.-]+\.xlsx', txt))
    if any(h.endswith(HIST_NAME) for h in hits):
        return True
    print('[enx] ⚠ 落地页 %s 上找不到 %s 的链接（页面上的 xlsx 有 %s）—— '
          '官方可能改了文件名，本次仍按写死的直链下载，请人工确认'
          % (LANDING_URL, HIST_NAME, sorted(hits) or '零个'))
    return False


def _rawkeep():
    """按路径加载 fetch/rawkeep.py。

    不能裸 import：本模块被 monthly_run 用 spec_from_file_location 加载，
    那时 sys.path 上既没有 fetch/ 也没有仓库根（同 _source_dates 的坑）。
    """
    import importlib.util
    here = os.path.dirname(os.path.abspath(__file__))
    spec = importlib.util.spec_from_file_location(
        'rawkeep', os.path.join(here, 'rawkeep.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _lm_month(last_modified):
    """把 Last-Modified 头折成 'YYYY-MM'，只给存证文件名当标签用；拿不到就 None。

    刻意用 Last-Modified 而不是数据月：本文件是**滚动全历史**，没有「属于哪个月」这回事，
    唯一有意义的标签是「这一版是什么时候挂上去的」。
    """
    if not last_modified:
        return None
    try:
        return datetime.datetime.strptime(
            last_modified, '%a, %d %b %Y %H:%M:%S %Z').strftime('%Y-%m')
    except (ValueError, TypeError):
        return None


def _download(cache_dir, name):
    os.makedirs(cache_dir, exist_ok=True)
    data, headers = _http_get(STATS_BASE + name)
    path = os.path.join(cache_dir, name)
    _write_bytes(path, data)
    last_modified = headers.get('Last-Modified')
    # ── 存证：模块 docstring 第 21 行已记「文件名固定、不带月份，每月原地覆盖」──
    # 上面这个 cache/<name> 是工作副本，下一期直接盖掉，历史版本官方一份都不留。
    # 按 <Last-Modified 月>-<sha256 前 12> 另存一份、永不覆盖，理由见 fetch/rawkeep.py。
    # 同一版重复下载只落一个文件（内容寻址），所以每天跑也不会膨胀。
    _rawkeep().keep('enx', data, 'xlsx', _lm_month(last_modified))
    return path, last_modified


# ══════════════════════════════════════════════════════════════════════
# 解析
# ══════════════════════════════════════════════════════════════════════
def _norm(v):
    return re.sub(r'\s+', ' ', str(v)).strip() if v is not None else ''


_FOOTNOTE_TAIL = re.compile(r'\s*\((?:\d+|R)\)\s*$')


def _lab(v):
    """表头文字归一化：只剥**结尾**的脚注编号 `(3)` 或修订标记 `(R)`。

    不能一律剥括号：`TA(1) MTS Repo` 的 (1) 在中间（它本身就是「Term Adjusted」的
    脚注，但剥掉之后标签会变成 'TA MTS Repo'，与表里对不上），
    `Bonds wholesale (in EUR bln)` 的括号是单位、`Cash Markets (Fixed Income excluded)`
    的括号是口径说明 —— 剥掉这些会让三元组失去区分力。
    """
    s = _norm(v)
    while True:
        t = _FOOTNOTE_TAIL.sub('', s)
        if t == s:
            return s
        s = t


def _cover_map(ws, upto_row):
    """{(row, col): 表头文字}，合并单元格里每一格都填成左上角那个值。

    两层表头 + 合并单元格是这份表的基本形态（口径坑 9）：分组标题只写在合并区左上角，
    直接读 ws.cell(8, 17) 会拿到 None。还原覆盖关系之后，每一列头上盖着哪些文字就是确定的，
    可以拼成 (分组, 小节, 标签) 三元组去唯一定位，不必写死列号。
    """
    out = {}
    for m in ws.merged_cells.ranges:
        if m.min_row > upto_row:
            continue
        v = ws.cell(m.min_row, m.min_col).value
        if v is None:
            continue
        for r in range(m.min_row, min(m.max_row, upto_row) + 1):
            for c in range(m.min_col, m.max_col + 1):
                out[(r, c)] = v
    return out


def _label_row(ws, sheet_name):
    """标签行 = A 列写着 'Period' 的那一行（Capital Markets 写成 'Period (1)'）。

    不写死行号：四张 sheet 的表头高度不一样（Equity/FICC 是第 10 行、
    Capital Markets 第 9 行、Securities Services 第 6 行），而且脚注一多就会整体下移。
    """
    for r in range(1, 25):
        if _lab(ws.cell(r, 1).value) == 'Period':
            return r
    raise EnxFetchError('sheet %s 前 25 行里找不到 A 列写 "Period" 的标签行，'
                        '官方表结构可能已变' % sheet_name)


class _Sheet(object):
    """一张 sheet 的表头索引：列 -> (分组链, 标签)，以及按三元组反查列号。"""

    def __init__(self, ws, name):
        self.ws = ws
        self.name = name
        self.lrow = _label_row(ws, name)
        cover = _cover_map(ws, self.lrow)

        def val(r, c):
            return cover.get((r, c), ws.cell(r, c).value)

        # raw_* 保留**没有剥掉脚注编号**的原文。剥过的用来定位列，没剥的用来回答
        # 「这一列挂着哪几条脚注」—— 官方的口径断点全写在脚注里，而哪些列受影响
        # 只有靠标签末尾那个 (3)/(5) 才认得出，剥早了这条信息就没了（见 breaks()）。
        self.label, self.heads = {}, {}
        self.raw_label, self.raw_heads = {}, {}
        self.raw_period = _norm(ws.cell(self.lrow, 1).value)
        for c in range(2, ws.max_column + 1):
            lab = _lab(val(self.lrow, c))
            if not lab:
                continue
            self.label[c] = lab
            self.raw_label[c] = _norm(val(self.lrow, c))
            self.heads[c] = tuple(
                _lab(val(r, c)) for r in range(1, self.lrow)
                if val(r, c) is not None)
            self.raw_heads[c] = tuple(
                _norm(val(r, c)) for r in range(1, self.lrow)
                if val(r, c) is not None)
        self._by_key = {}
        for c, lab in self.label.items():
            self._by_key.setdefault((self.heads[c], lab), []).append(c)

    def col(self, heads, label):
        """按 (分组链, 标签) 唯一定位一列；找不到或撞到多列都抛异常。

        撞到多列一定要炸，不能取第一个：`Athex` 在同一分组里就出现三次
        （Total Turnover / Turnover Equities / Turnover ETF 各配一个），
        取第一个等于把 ETF 的备注数当成整体的备注数写进 CSV，而且看上去完全正常。
        备注列不走这条路，走 memo_col()。
        """
        got = self._by_key.get((tuple(heads), label), [])
        if len(got) != 1:
            raise EnxFetchError(
                'sheet %s 里 (分组=%s, 标签=%r) 命中 %d 列（应为 1）—— '
                '官方表结构可能已变' % (self.name, list(heads), label, len(got)))
        return got[0]

    def memo_col(self, main_col):
        """Athex 备注列 = 主列**紧邻右侧**那一列，且标签是 Athex / Athens。

        为什么按位置而不按标签：备注列的 (分组, 小节, 标签) 三元组在同一分组里完全重复，
        三元组定位不了它。而「备注紧跟主列」是这张表真实的排版语义 ——
        官方在每一个配了备注的指标右边插一列，没配备注的（Structured Products、
        Bonds wholesale、Funds）右边就是下一个正经指标。所以这条位置规则同时也是
        「这个指标到底有没有备注列」的判据。
        """
        c = main_col + 1
        if self.label.get(c) in ('Athex', 'Athens'):
            return c
        return None

    def footnote_ids(self, col):
        """这一列挂着的脚注编号（含它头上分组的、以及整张表的 Period 那条）。

        `TA(1) MTS Repo` 这种括号在中间的也会被认成脚注 1 —— 无所谓：
        FICC 的脚注 1 是 "Term Adjusted"，里面没有月份，产不出断点行，自己就消化掉了。
        """
        txt = ' '.join((self.raw_period, self.raw_label.get(col, ''))
                       + self.raw_heads.get(col, ()))
        return sorted(set(re.findall(r'\((\d+)\)', txt)), key=int)

    def days_col(self, group):
        """某分组的日数列 = 紧邻该分组左侧的那个 `Nb of trading days` 列（口径坑 10）。

        `Nb of trading days` 在一张 sheet 里出现 2-5 次，标签本身没有区分力，
        而它们在表里的位置语义就是「我右边这一块用我当分母」。
        单股衍生品分组左边没有自己的日数列，按这条规则解析到的正是股权衍生品那个（C16）——
        与官方季报把两者放在同一个 "Number of trading days 62" 下面一致。
        """
        start = None
        for c, heads in self.heads.items():
            if heads and heads[0] == group:
                start = c if start is None else min(start, c)
        if start is None:
            raise EnxFetchError('sheet %s 里找不到分组 %r' % (self.name, group))
        cands = [c for c, lab in self.label.items()
                 if lab == 'Nb of trading days' and c < start]
        if not cands:
            raise EnxFetchError(
                'sheet %s 的分组 %r 左侧找不到 "Nb of trading days" 列'
                % (self.name, group))
        return max(cands)


_NULLS = {'', '-', '–', 'n/a', 'na', 'nd', 'n.a.'}


def _cell_num(ws, row, col, where):
    """单元格取数。空白与官方约定的空值记号返回 None，其余非数字一律抛异常。

    'NA' 必须进空值白名单：Capital Markets 的 `Funds` 列 2018 全年就是这个字面量字符串
    （口径坑 12），float() 会 ValueError。放进白名单不会掩盖问题 —— 该列的 SINCE 写的是
    2019-01，2019-01 之后再出现 'NA' 照样会被 _validate 当成缺值抓出来。
    """
    v = ws.cell(row, col).value
    if v is None:
        return None
    if isinstance(v, str):
        if v.strip().lower() in _NULLS:
            return None
        raise EnxFetchError('%s R%dC%d 不是数字：%r' % (where, row, col, v))
    if isinstance(v, datetime.datetime):
        raise EnxFetchError('%s R%dC%d 拿到日期而不是数字：%r' % (where, row, col, v))
    try:
        return float(v)
    except (TypeError, ValueError):
        raise EnxFetchError('%s R%dC%d 不是数字：%r' % (where, row, col, v))


def open_sheets(path):
    """打开历史 xlsx，返回 {sheet 名: _Sheet}。**白名单四张**，绝不 for-each-sheet。

    白名单不是洁癖：`Checkup` 那张 sheet 的唯一数据列是 117 个 `#REF!` 字符串，
    表头却写着 "Euronext Cash / Turnover in millions euros"（口径坑 13）—— 遍历式解析
    会把它当成正经数据去试，运气不好还真能试出一列垃圾。
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        raise EnxFetchError('%s 里缺 sheet %s（拿到 %r）'
                            % (os.path.basename(path), missing, wb.sheetnames))
    return {s: _Sheet(wb[s], s) for s in SHEETS}


def parse_workbook(path, sheets=None):
    """解析历史 xlsx，返回 {'YYYY-MM': {csv列名: float|None}}（按月升序）。

    任何一个约定的 (分组, 标签) 找不到、或撞到多列 —— 说明官方改了表结构 ——
    直接抛异常。宁可整月不更新，也不要写出一列悄悄全空的 CSV。
    """
    sh = sheets if sheets is not None else open_sheets(path)

    # 结构自检：单股衍生品那一块解析到的日数列必须与股指那一块是同一列。
    # 这不是多余的 —— 哪天官方给单股插一个独立的日数列，days_col() 会安静地改指向，
    # 而 ADV 会整体错几个百分点，没有任何东西会报错。
    if sh[S_EQ].days_col(G_SS) != sh[S_EQ].days_col(G_IDX):
        raise EnxFetchError(
            '单股衍生品与股指衍生品解析到了不同的交易日列（C%d vs C%d）—— '
            '官方可能新增了独立的日数列，需人工确认口径'
            % (sh[S_EQ].days_col(G_SS), sh[S_EQ].days_col(G_IDX)))

    days_col = {k: sh[s].days_col(g) for k, (s, g) in DAYS_ANCHOR.items()}
    days_sheet = {k: s for k, (s, _g) in DAYS_ANCHOR.items()}

    # 每张 sheet 的 (year, month) -> 行号。Capital Markets 的 Period 不是月初
    # （2018-01-05 那种），必须按 (年, 月) 归并，见口径坑 11。
    rows = {}
    for s in SHEETS:
        ws, lrow = sh[s].ws, sh[s].lrow
        rows[s] = {}
        for r in range(lrow + 1, ws.max_row + 1):
            p = ws.cell(r, 1).value
            if not isinstance(p, datetime.datetime):
                continue
            mon = '%04d-%02d' % (p.year, p.month)
            if mon in rows[s]:
                raise EnxFetchError(
                    'sheet %s 里 %s 出现两行（R%d 与 R%d）—— 按 (年,月) 归并会丢数据，'
                    '需人工确认' % (s, mon, rows[s][mon], r))
            rows[s][mon] = r

    months = sorted(set().union(*[set(v) for v in rows.values()]))
    data = {}
    for mon in months:
        rec = dict.fromkeys(COLUMNS)
        # 交易日先取，后面的 ADV 都要用它当分母
        dv = {}
        for name, key, _since in DAYS_SPEC:
            s = days_sheet[key]
            r = rows[s].get(mon)
            v = (None if r is None else
                 _cell_num(sh[s].ws, r, days_col[key], '%s/%s' % (s, mon)))
            dv[key] = v
            rec[name] = v
        for c in COLUMN_SPEC:
            r = rows[c.sheet].get(mon)
            if r is None:
                continue
            sheet = sh[c.sheet]
            where = '%s/%s' % (c.sheet, mon)
            main_col = sheet.col(c.heads, c.label)
            rec[c.name] = _scale(_cell_num(sheet.ws, r, main_col, where),
                                 dv.get(c.days), c.scale, c, mon)
            if c.memo:
                mc = sheet.memo_col(main_col)
                if mc is None:
                    raise EnxFetchError(
                        '%s 的 %r 右侧不再是 Athex/Athens 备注列（拿到 %r）—— '
                        '并表口径的可回溯性依赖这一列，拒绝静默丢弃'
                        % (c.sheet, c.label, sheet.label.get(main_col + 1)))
                rec[c.memo] = _scale(_cell_num(sheet.ws, r, mc, where),
                                     dv.get(c.days), c.scale, c, mon)
        data[mon] = rec
    if not data:
        raise EnxFetchError('%s 解析后没有任何月份' % os.path.basename(path))
    return dict(sorted(data.items()))


def _scale(raw, days, scale, col, mon):
    """月度总量 → 入库值。col.days 为 None 表示这一列是时点值（OI / 家数 / 市值），不除天数。

    ADV 的算法是「月度总量 ÷ 该月交易日数」—— 这不是我们发明的口径：官方在伴生的
    latest 文件里给出算好的 ADV，与本算法的结果**逐位相同**（2026-06 现货
    17183.521449326818 双方完全相等，见 _crosscheck_latest_month）。

    「要不要除」只看 col.days（列的定义），**绝不看 days 这个值是不是 None**。
    两者混在一起写过一版，后果是：官方哪天把某个 Nb of trading days 格子留空而照常填量，
    这一列会被当成时点值原样入库 —— 一个 20 倍偏大的数字，静悄悄，没有任何报错。
    所以日数缺失必须炸，不能退化成「那就不除了」。
    """
    if raw is None:
        return None
    if col.days is None:
        return raw / scale
    if not days:
        raise EnxFetchError('%s 的 %s 该按 %s 交易日折算，但日数是 %r —— '
                            '拒绝把月度总量当成日均写进 CSV'
                            % (mon, col.name, col.days, days))
    return raw / days / scale


def _validate(data):
    """返回最新月；任何一处不达标立刻抛异常。

    三道检查：
      1. 起始月之后不许有空格 —— 起始月是本机对当前 xlsx 逐列实测出来的，
         之后再为空只可能是解析错行或官方停发，两种都必须人来看。
      2. 恒等式 Total ≡ Equities + ETF + Structured（Athex 备注列同理，它没有
         结构化产品，所以是 Total ≡ Equities + ETF）。2026-08-18 实测 175 个月
         最大相对差 3.7e-16（量级是浮点舍入，不随月数变；月数本身会变）。
         这条撞不上，说明四列里至少有一列错行了 —— 而错行的数字全都「看上去很正常」。
      3. 交易日必须是正数。
    """
    have = [m for m in sorted(data) if data[m][ANCHOR] is not None]
    if not have:
        raise EnxFetchError('解析结果里没有任何一个月有 %s，文件疑似空壳' % ANCHOR)
    newest = have[-1]

    for mon in have:
        rec = data[mon]
        bad = [c for c in COLUMNS
               if rec[c] is None and SINCE[c] and mon >= SINCE[c]]
        if bad:
            raise EnxFetchError(
                '%s 缺列 %s（这些列自 %s 起官方就有数）—— 解析异常，拒绝写入'
                % (mon, bad, [SINCE[c] for c in bad]))
        for name, key, _s in DAYS_SPEC:
            d = rec[name]
            if d is not None and not (d > 0):
                raise EnxFetchError('%s 的 %s = %r，不是正数' % (mon, name, d))
        _identity(mon, rec, 'adv_cash_adnv_eurbn',
                  ['adv_cash_equities_adnv_eurbn', 'adv_cash_etf_adnv_eurbn',
                   'adv_cash_structured_adnv_eurbn'])
        _identity(mon, rec, 'athex_adv_cash_adnv_eurbn',
                  ['athex_adv_cash_equities_adnv_eurbn',
                   'athex_adv_cash_etf_adnv_eurbn'])
    return newest


# ══════════════════════════════════════════════════════════════════════
# 口径断点台账 series/enx_breaks.csv
# ══════════════════════════════════════════════════════════════════════
_FOOTNOTE = re.compile(r'^\((\d+)\)\s*(.+)$')
_MONTH_YEAR = re.compile(
    r'\b(January|February|March|April|May|June|July|August|September|'
    r'October|November|December)\s+(\d{4})\b')

BREAKS_NAME = 'enx_breaks.csv'
BREAKS_FIELDS = ['column', 'break_month', 'footnote', 'athex_memo_column',
                 'official_footnote']


def _sheet_footnotes(sh):
    """{脚注编号: 原文}。脚注就写在标签行以上的 A 列，形如 "(3) Includes figures from…"。"""
    ws, out = sh.ws, {}
    for r in range(1, sh.lrow):
        v = ws.cell(r, 1).value
        if not isinstance(v, str):
            continue
        m = _FOOTNOTE.match(_norm(v))
        if m:
            out[m.group(1)] = m.group(2)
    return out


def breaks(sheets):
    """从**官方脚注原文**里抽出每一列的口径断点，返回可直接写 CSV 的行列表。

    为什么要落成 series/enx_breaks.csv，而不是在代码里写死一张断点表：
    Euronext 的断点全部来自并购（Dublin 2017、Oslo 2018/2019、Borsa Italiana 2021、
    **Athens 2025-11**），而并购是会继续发生的。写死的表在下一次并购时不会报错，
    只会安静地过时 —— 图上少画一条竖线，而少画的那条恰恰是最该画的那条。
    从脚注抽就不会：官方改脚注，这张表下次跑就跟着变，git diff 里看得见。

    「哪些列受哪条脚注影响」也不是猜的：官方把脚注编号挂在列标签与分组标题的末尾
    （`Cash Markets (Fixed Income excluded) (3)`、`Equity Index derivatives (5)`、
    `Turnover ETF (2)`、`Bonds (3)`），所以列 → 脚注 → 断点月这条链全程有据。

    只给主列出行，不给 `athex_*` 备注列出行 —— 备注列本身是连续的，它恰恰是**消除**
    2025-11 断点的那把钥匙，所以它以 `athex_memo_column` 一栏的形式挂在主列那一行上。
    """
    fns = {s: _sheet_footnotes(sh) for s, sh in sheets.items()}
    rows = []
    for c in COLUMN_SPEC:
        sh = sheets[c.sheet]
        col = sh.col(c.heads, c.label)
        for fid in sh.footnote_ids(col):
            text = fns[c.sheet].get(fid)
            if not text:
                continue
            seen = set()
            for m in _MONTH_YEAR.finditer(text):
                mon = '%s-%02d' % (
                    m.group(2), datetime.datetime.strptime(m.group(1), '%B').month)
                if mon in seen:
                    continue
                seen.add(mon)
                rows.append({
                    'column': c.name,
                    'break_month': mon,
                    'footnote': '%s (%s)' % (c.sheet, fid),
                    'athex_memo_column': c.memo or '',
                    'official_footnote': text,
                })
    rows.sort(key=lambda r: (COLUMNS.index(r['column']), r['break_month'],
                             r['footnote']))

    _check_athens_month(rows)
    return rows


# 一条脚注里往往并排写着好几次并购（"…Oslo since January 2018, Borsa Italiana since
# May 2021 and Euronext Athens since November 2025"），所以只能取**紧跟在 Athens 后面**
# 的那个月份，不能拿整条脚注里的任意月份去比 —— 那样每条脚注都会误报。
_ATHENS_SINCE = re.compile(
    r'Ath(?:ens|ex)\b[^.]{0,40}?\b(January|February|March|April|May|June|July|'
    r'August|September|October|November|December)\s+(\d{4})')


def _check_athens_month(rows, expect='2025-11'):
    """雅典并表月必须是 2025-11；不是就大声警告（但不抛异常）。

    并表月是本模块所有 `athex_*` 列语义翻转的那一天（口径坑 1）：
    这个月之前「主列 + 备注列 = pro-forma」，这个月起「主列 − 备注列 = legacy」。
    它要是变了（比如官方改成按 pro-forma 重述全历史），加减方向就反了，
    而算出来的数字仍然「看上去很正常」—— 这正是没人会发现的那类错。

    只警告不抛异常：官方重述并表基准是新闻，不是故障；数据本身还是好的，
    该停下来的是**读数的人**，不是管道。
    """
    seen = {}
    for r in rows:
        m = _ATHENS_SINCE.search(r['official_footnote'])
        if not m:
            continue
        mon = '%s-%02d' % (m.group(2),
                           datetime.datetime.strptime(m.group(1), '%B').month)
        seen[r['footnote']] = mon
    bad = sorted((k, v) for k, v in seen.items() if v != expect)
    if bad:
        print('[enx] ⚠ 官方脚注里的雅典并表月不再是 %s：%s —— '
              'athex_* 备注列的加减方向必须重新确认，跨该月的同比全部作废' % (expect, bad))
    elif not seen:
        print('[enx] ⚠ 四张 sheet 的脚注里一条都没提到 Euronext Athens —— '
              '官方可能改写了脚注，2025-11 并表断点是否还成立需人工确认')
    return bad


def _write_breaks(series_dir, rows):
    """落盘 series/enx_breaks.csv；内容没变就不动文件（保持字节级幂等）。"""
    path = os.path.join(series_dir, BREAKS_NAME)
    sio = io.StringIO()
    w = csv.DictWriter(sio, BREAKS_FIELDS, lineterminator='\n')
    w.writeheader()
    w.writerows(rows)
    new = sio.getvalue()
    old = None
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            old = f.read()
    if old == new:
        return False
    tmp = path + '.tmp'
    with open(tmp, 'w', encoding='utf-8', newline='') as f:
        f.write(new)
    os.replace(tmp, path)
    return True


def _identity(mon, rec, total, parts, tol=1e-9):
    if rec.get(total) is None or any(rec.get(p) is None for p in parts):
        return
    lhs, rhs = rec[total], sum(rec[p] for p in parts)
    if lhs == 0:
        return
    rel = abs(lhs - rhs) / abs(lhs)
    if rel > tol:
        raise EnxFetchError(
            '%s 恒等式不成立：%s=%r 与 %s 之和 %r 相对差 %.3e —— '
            '多半是某一列错行了' % (mon, total, lhs, parts, rhs, rel))


# ══════════════════════════════════════════════════════════════════════
# 对表自检：官方 latest 文件（**只比最新月这一列**，见口径坑 6）
# ══════════════════════════════════════════════════════════════════════
def _crosscheck_latest_month(data, newest, cache_dir):
    """拿官方算好的 ADV 撞我们自算的 ADV，返回一句人话说明。

    这是免费的第二意见：官方在 latest 文件里直接给出 ADV Cash Market，
    而我们是从历史文件的「月总量 ÷ 交易日」自己算的。两条完全独立的路径。

    ⚠ 只比**最新月**那一列，因为 latest 的同比/上年列用的是 pro-forma 含 Athens 的基准
    （它自己的脚注写 "since January 2025"），与历史文件主列的 "since November 2025"
    不是一个口径，2025-11 之前会看到最高 23% 的假失配。当月两个基准重合，才可比。

    差得离谱（>1e-3）才抛异常：那种量级只可能是取错列或用错单位。
    取不到文件、或 latest 的月份还没跟上，一律只打印不阻断 —— 主源自己的结构校验
    与恒等式才是护栏，辅助源不该有权卡住整月发布。
    """
    try:
        path, _lm = _download(cache_dir, LATEST_NAME)
        wb = openpyxl.load_workbook(path, data_only=True)
        if 'Equity Markets' not in wb.sheetnames:
            return '对表跳过：latest 文件里没有 Equity Markets sheet'
        ws = wb['Equity Markets']
        # 第 2 列固定是「最新月」。先核对它自报的月份，对不上就不比。
        mon_cell = None
        for r in range(1, 12):
            v = ws.cell(r, 2).value
            if isinstance(v, datetime.datetime):
                mon_cell = '%04d-%02d' % (v.year, v.month)
                break
        if mon_cell != newest:
            return ('对表跳过：latest 文件的最新月是 %s，历史文件是 %s，两边不同步'
                    % (mon_cell, newest))
        # 'ADV Cash Market' 在 latest 里出现两次（笔数区一次、金额区一次），
        # 取金额区那一个：先定位 'TRANSACTION VALUE' 小节标题，再往下找。
        sec = None
        for r in range(1, ws.max_row + 1):
            a = _norm(ws.cell(r, 1).value)
            if a.upper().startswith('TRANSACTION VALUE'):
                sec = r
                break
        if sec is None:
            return '对表跳过：latest 文件里找不到 TRANSACTION VALUE 小节'
        official = None
        for r in range(sec, min(sec + 8, ws.max_row) + 1):
            if _norm(ws.cell(r, 1).value) == 'ADV Cash Market':
                official = ws.cell(r, 2).value
                break
        if not isinstance(official, (int, float)):
            return '对表跳过：latest 文件的 ADV Cash Market 取不到数（%r）' % official
        mine = data[newest]['adv_cash_adnv_eurbn'] * 1e3      # €bn/日 → €m/日
        rel = abs(mine - official) / abs(official)
    except EnxFetchError as e:
        return '对表跳过：%r' % e
    except Exception as e:                                    # noqa: BLE001
        return '对表跳过（latest 文件解析异常，不阻断主流程）：%r' % e

    if rel > 1e-3:
        raise EnxFetchError(
            '%s 现货 ADV 自算 %.6f €m/日 与官方 latest 文件的 %.6f 相对差 %.3e —— '
            '这个量级只可能是取错列或用错单位，拒绝写入' % (newest, mine, official, rel))
    verdict = '完全一致' if mine == official else '相对差 %.3e' % rel
    return ('%s 现货 ADV 自算 %r €m/日 vs 官方 latest 文件 %r，%s'
            % (newest, mine, official, verdict))


# ══════════════════════════════════════════════════════════════════════
# 发布日
# ══════════════════════════════════════════════════════════════════════
def _source_dates():
    """按路径加载仓库根的 source_dates.py。

    不能裸 import：本模块被 monthly_run 用 spec_from_file_location 加载，
    那时 sys.path 上既没有 fetch/ 也没有仓库根。
    """
    import importlib.util
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(root, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


_TR = re.compile(r'<tr>(.*?)</tr>', re.S)
_TIME = re.compile(r'<time[^>]*datetime="([^"]+)"[^>]*>([^<]*)</time>')
_A = re.compile(r'<a\s+href="([^"]+)"[^>]*>(.*?)</a>', re.S)
_TAGS = re.compile(r'<[^>]+>')
# 电头：'… Oslo and Paris – 6 July 2026 – Euronext, the …'
_DATELINE = re.compile(r'[–—-]\s*(\d{1,2}\s+[A-Z][a-z]+\s+\d{4})\s*[–—-]')


def _news_rows(cache_dir, pages=3):
    """从新闻列表页抓 [(本地日期 YYYY-MM-DD, datetime 属性原文, href, 标题)]。

    从**列表页**取而不是按模板拼 slug —— 拼 slug 的真实代价是 2023-03 那期：
    它的标题是 "Euronext announces highest cash volumes in a year in March 2023"，
    slug 是 euronext-announces-highest-cash-volumes-year-march，
    两种模板拼法（…-for-march-2023 与 …-for-march-2023-0）实测**双双 404**。

    日期以 <time> 的**渲染文本**（DD/MM/YYYY，站点自己按巴黎时间渲染的日历日）为准，
    `datetime` 属性（UTC）由 _record_source_date 拿去互证：实测发布时刻在
    07:30Z–17:06Z 之间，UTC 日期与巴黎日期恒等，但与其把这条依赖写死在代码里，
    不如每次都让两个值互相印证、不一致就写进 evidence 让人看见。
    """
    out = []
    for p in range(pages):
        try:
            html, _h = _http_get(NEWS_URL % p, tries=3)
        except EnxFetchError as e:
            print('[enx] 警告：新闻列表页第 %d 页取不到（%r），发布日可能缺席' % (p, e))
            break
        _write_bytes(os.path.join(cache_dir, 'enx_news_p%d.html' % p), html)
        txt = html.decode('utf-8', 'replace')
        for m in _TR.finditer(txt):
            seg = m.group(1)
            t, a = _TIME.search(seg), _A.search(seg)
            if not (t and a):
                continue
            shown = t.group(2).strip()
            mm = re.match(r'^(\d{2})/(\d{2})/(\d{4})$', shown)
            if not mm:
                continue
            local = '%s-%s-%s' % (mm.group(3), mm.group(2), mm.group(1))
            title = _unescape(_TAGS.sub('', a.group(2))).strip()
            out.append((local, t.group(1), a.group(1), title))
        time.sleep(0.8)             # 主站连打会掐连接，见 _http_get
    return out


def _unescape(s):
    import html as _html
    return _html.unescape(s)


def _month_end_next(month):
    y, m = int(month[:4]), int(month[5:7])
    return ('%04d-01-01' % (y + 1)) if m == 12 else '%04d-%02d-01' % (y, m + 1)


def _find_release(rows, month):
    """在列表页结果里找 month 的月度成交量稿，返回 (日期, datetime 属性, href, 标题) 或 None。

    宽松匹配：标题里同时出现 announces…volume 与「<英文月名> <年>」即可。
    这样才能同时命中标准模板与 2023-03 那种「创纪录」改写标题（口径坑 16）。
    再加一道时间窗（数据月结束后 1-60 天内），挡住把某篇回顾性文章误认成月报的情况。
    """
    y, m = int(month[:4]), int(month[5:7])
    mname = datetime.date(y, m, 1).strftime('%B')
    pat_kind = re.compile(r'announces\b.*\bvolume', re.I)
    pat_mon = re.compile(r'\b%s\s+%d\b' % (mname, y), re.I)
    lo = _month_end_next(month)
    hi = (datetime.date(*map(int, lo.split('-')))
          + datetime.timedelta(days=60)).isoformat()
    hit = [r for r in rows
           if pat_kind.search(r[3]) and pat_mon.search(r[3]) and lo <= r[0] <= hi]
    return min(hit) if hit else None


def _dateline(cache_dir, href, month):
    """从详情页正文电头取「6 July 2026」这种人可读的日期，取不到返回 None。

    只是佐证，不是主证 —— 主证是列表页的 <time datetime>。
    2023-03 那期的详情页 meta description 被联系人区块占掉，电头就抓不到；
    这种时候 evidence 少一句话，而不是整条发布日缺席。
    ⚠ 详情页**没有** JSON-LD 的 datePublished 字段（页面上唯一的 ld+json 块是面包屑），
    不要去那里找。
    """
    url = href if href.startswith('http') else SITE_ROOT + href
    try:
        html, _h = _http_get(url, tries=2)
    except EnxFetchError:
        return None
    _write_bytes(os.path.join(cache_dir, 'enx_pr_%s.html' % month), html)
    txt = html.decode('utf-8', 'replace')
    m = re.search(r'<meta name="description" content="([^"]*)"', txt)
    hay = _unescape(m.group(1)) if m else txt[:20000]
    d = _DATELINE.search(hay)
    return d.group(1) if d else None


def _record_source_date(series_dir, cache_dir, month):
    """给 month 记一条官方发布日；取不到就让它缺席，**绝不抛异常**。

    只给「本次真的新入库的那个最新月」记 —— 同一份 xlsx 里躺着全部历史月（现已一百七十多个），
    它的上线时刻只能证明最新月是这天发的，顺手给旧月份都盖上今天的日期就是造假。
    已有记录一律不覆盖：官方会重述并原地重传文件（口径坑 15），
    覆盖等于把当初那次真发布的日期改错，而页面照印不误。
    """
    sd = _source_dates()
    if sd.lookup(series_dir, 'enx', month):
        return None
    rows = _news_rows(cache_dir)
    hit = _find_release(rows, month)
    if not hit:
        print('[enx] 警告：新闻列表页里没找到 %s 的月度成交量稿，本月不记发布日'
              '（页面抬头会省掉「官方发布于」那半句）' % month)
        return None
    day, dt_attr, href, title = hit
    ev = ('新闻列表页 %s 的一行「%s」：<time datetime="%s">%s</time>'
          % (NEWS_URL % 0, title, dt_attr, _ddmmyyyy(day)))
    # UTC 属性与渲染出来的巴黎日期互证。两者不同 = 发布时刻跨了午夜，
    # 那种情况下「哪天发的」有歧义，必须让读 evidence 的人看见，而不是替他选一个。
    if dt_attr[:10] != day:
        ev += '（⚠ datetime 属性的 UTC 日期 %s 与渲染日期不同）' % dt_attr[:10]
    line = _dateline(cache_dir, href, month)
    if line:
        try:
            same = (datetime.datetime.strptime(line, '%d %B %Y').date().isoformat()
                    == day)
        except ValueError:
            same = False
        ev += '；详情页 %s%s 正文电头 "– %s –" %s' % (
            SITE_ROOT, href, line, '一致' if same else '⚠ 与列表页不一致')
    sd.record(series_dir, 'enx', month, day, ev)
    return day


def _ddmmyyyy(iso):
    return '%s/%s/%s' % (iso[8:10], iso[5:7], iso[:4])


def backfill_source_dates(series_dir, cache_dir, pages=8, months=None):
    """人工回补历史发布日（不由 update() 自动调用）。

    每个月的发布日来自**那个月自己的新闻稿**，所以批量回补并不违反「一份文件只为
    自己的最新月作证」——那条规矩管的是文件时间戳，不管逐月的新闻稿。
    但它会一次性往 series/source_dates.csv 里塞几十行，属于人工决策，
    所以留成显式入口：`python3 fetch/enx.py source-dates`。
    """
    sd = _source_dates()
    rows = _news_rows(cache_dir, pages=pages)
    if months is None:
        csv_path = os.path.join(series_dir, 'enx.csv')
        with open(csv_path, newline='', encoding='utf-8') as f:
            months = [r[0] for r in list(csv.reader(f))[1:] if r and r[0].strip()]
    done = []
    for mon in sorted(months):
        if sd.lookup(series_dir, 'enx', mon):
            continue
        hit = _find_release(rows, mon)
        if not hit:
            continue
        day, dt_attr, href, title = hit
        sd.record(series_dir, 'enx', mon, day,
                  '新闻列表页「%s」：<time datetime="%s">%s</time>（%s%s）'
                  % (title, dt_attr, _ddmmyyyy(day), SITE_ROOT, href))
        done.append((mon, day))
    return done


# ══════════════════════════════════════════════════════════════════════
# 对外接口
# ══════════════════════════════════════════════════════════════════════
def latest_month(cache_dir):
    """官方源当前最新月 'YYYY-MM'。

    以「现货 ADNV 非空的最后一个月」为准，不信文件里的最后一行 ——
    官方有时会把下个月的行先开出来只填交易日（HKEX 踩过的同类坑）。
    抓不到 / 解析不出来一律抛 EnxFetchError，不返回 None 掩盖故障。
    """
    _check_landing(cache_dir)
    path, _lm = _download(cache_dir, HIST_NAME)
    return _validate(parse_workbook(path))


def _fmt(v):
    """写回 CSV。整数写整数（交易日、家数、上市只数本来就是整数），其余用最短往返表示。

    不无脑 repr(float)：那会把 1844 写成 '1844.0'，与 cme.csv / hkex.csv 的风格不一致，
    也让人拿 CSV 与官方原表逐位对照时多一层心智负担。
    """
    if v is None:
        return ''
    f = float(v)
    return str(int(f)) if f.is_integer() and abs(f) < 1e15 else repr(f)


def update(series_dir, cache_dir):
    """把新月份写进 series/enx.csv，返回新增月份列表（升序）。

    幂等保证：
      · 已存在的月份不重复追加；
      · 已经有值的单元格**永不覆盖** —— 官方明确会回溯重述（口径坑 4，实测
        2019-01 现货 +6.4%、2020-06 现货 −4.1%），重述不由无人值守任务自动吞进来；
        官方与本仓不一致的格子写进 cache/enx_restatements.csv 供人工判断；
      · 只在既有行**原本为空**的格子上回补（正常情况下不会有：Euronext 一次给全所有列）；
      · 什么都没变时未被触碰的单元格是原样字符串搬运 ⇒ 文件字节级不变，重跑返回 []。

    首次调用时 series/enx.csv 不存在 —— 本模块会按 COLUMNS 建表并一次写满全历史
    （2012-01 起，工作簿有多少月就写多少月）。这不是「顺手做掉」：Euronext 的单一 xlsx 本来就带全序列，
    分两步（先手工 bootstrap 再增量）反而会留下「bootstrap 脚本与 fetch 解析器两套代码」
    的经典漂移，README 讲的 cost/ibkr 搬迁就是在治这个病。
    """
    csv_path = os.path.join(series_dir, 'enx.csv')
    if os.path.exists(csv_path):
        with open(csv_path, newline='', encoding='utf-8') as f:
            rows = list(csv.reader(f))
        header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
        if header != ['month'] + COLUMNS:
            raise EnxFetchError(
                'series/enx.csv 的列名与本模块不符；缺 %s，多 %s'
                % ([c for c in COLUMNS if c not in header],
                   [c for c in header[1:] if c not in COLUMNS]))
    else:
        header, body = ['month'] + COLUMNS, []
    idx = {name: i for i, name in enumerate(header)}

    _check_landing(cache_dir)
    path, last_modified = _download(cache_dir, HIST_NAME)
    sheets = open_sheets(path)
    data = parse_workbook(path, sheets)
    newest = _validate(data)
    print('[enx] %s' % _crosscheck_latest_month(data, newest, cache_dir))

    # 断点台账每次都重算：它是从官方脚注原文抽的，官方一改脚注这里就跟着变。
    # 与 enx.csv 的落盘分开、且先写 —— 断点表是「怎么读这些数」的说明书，
    # 说明书比数据更不该滞后。
    if _write_breaks(series_dir, breaks(sheets)):
        print('[enx] series/%s 已更新（口径断点来自官方脚注原文）' % BREAKS_NAME)

    have = {r[0]: r for r in body}
    added, filled, restated = [], [], []
    for mon in sorted(data):
        rec = data[mon]
        if rec[ANCHOR] is None:
            # 官方把行开出来了但还没填数（或早于该 sheet 的起始月），不建行
            continue
        if mon in have:
            row = have[mon]
            for name in COLUMNS:
                if rec[name] is None:
                    continue
                new = _fmt(rec[name])
                if not row[idx[name]].strip():
                    row[idx[name]] = new
                    filled.append((mon, name, new))
                elif row[idx[name]] != new:
                    restated.append((mon, name, row[idx[name]], new))
            continue
        row = [''] * len(header)
        row[0] = mon
        for name in COLUMNS:
            row[idx[name]] = _fmt(rec[name])
        have[mon] = row
        body.append(row)
        added.append(mon)

    # 官方与本仓不一致的格子一律落盘、绝不自动覆盖 —— 是口径重述还是解析出错，
    # 只有人能判断（照 fetch/hkex.py 与 fetch/ice.py 的做法）。
    if restated:
        os.makedirs(cache_dir, exist_ok=True)
        rp = os.path.join(cache_dir, 'enx_restatements.csv')
        with open(rp, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f, lineterminator='\n')
            w.writerow(['month', 'column', 'in_series_csv', 'in_official_xlsx'])
            w.writerows(restated)
        print('[enx] 官方源与 series 有 %d 处不一致，已写 %s（本模块不覆盖，请人工判断）'
              % (len(restated), rp))

    if not (added or filled):
        return []

    body.sort(key=lambda r: r[0])
    tmp = csv_path + '.tmp'
    with open(tmp, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(body)
    os.replace(tmp, csv_path)          # 原子替换：中途挂掉不会留下半张表

    # 记发布日放在落盘之后：写盘失败就没有「这个月官方发过了」这条断言。
    if newest in added:
        _record_source_date(series_dir, cache_dir, newest)
    if filled:
        print('[enx] 补空 %d 格：%s' % (len(filled), filled[:12]))
    print('[enx] 源文件 %s（Last-Modified %s），最新月 %s'
          % (HIST_NAME, last_modified, newest))
    return sorted(added)


if __name__ == '__main__':
    import sys
    _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _series, _cache = os.path.join(_root, 'series'), os.path.join(_root, 'cache')
    if len(sys.argv) > 1 and sys.argv[1] == 'source-dates':
        print('source_dates 回补:', backfill_source_dates(_series, _cache))
    elif len(sys.argv) > 1 and sys.argv[1] == 'latest':
        print('latest:', latest_month(_cache))
    else:
        _added = update(_series, _cache)
        print('added : %d 个月 %s'
              % (len(_added), (_added[:3] + ['…'] + _added[-3:])
                 if len(_added) > 6 else _added))
