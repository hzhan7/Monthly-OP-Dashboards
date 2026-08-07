# -*- coding: utf-8 -*-
"""Deutsche Börse Group (DB1) 月度经营指标 —— 无人值守抓取。

一家公司要缝**三个官方源、两种发布节奏**：Eurex 衍生品统计、FWB 现货统计（这两条
次月 1–5 日就出，称「快腿」），以及集团 IR 台账（次月约 10 日，称「慢腿」）。
本模块把三条腿写进同一张 series/db1.csv，**不做任何口径换算**——张数就是张数，
€bn 就是 €bn，定基名义额那一层是 build/notional.py 的事。

════════════════════════════════════════════════════════════════════════════
数据源
════════════════════════════════════════════════════════════════════════════
三个源都是 DBG 一方披露（deutsche-boerse.com / eurex.com / cashmarket.deutsche-boerse.com），
无第三方聚合商。实测普通 UA + plain urllib 直接 200，无 Cloudflare / Akamai / PerimeterX /
JS 渲染 / 登录墙 —— 满足 cron 无人值守。

A. 集团 IR 台账（慢腿，全部非成交量指标的唯一来源）
   落地页 https://www.deutsche-boerse.com/dbg-en/investor-relations/statistics
   直链   .../resource/blob/249090/{32位hex}/data/major-business-figures_en.xlsx
   格式   xlsx（openpyxl），单 sheet「Major business figures」，2026-08-06 实测
          294 行 × 23 个有数据列，**2002-01 → 2026-06 零断档**（逐月比对 missing=[]）。
   抓法   解析落地页正则取 href。blob id 249090 常驻（docProps created 2014-01-13），
          落地页改版时可作兜底；hash 段不校验（实测换成 deadbeef… 仍返回同一文件）。
   配套   同页 monthly-volume-development_en.pdf（口径脚注写在这里，只作人工核对，
          不进管道）。本模块 docstring 里所有「官方原文」都出自这份 PDF。

B. Eurex 月度统计（快腿，产品级 ADV 与 OI 的唯一来源）
   列表页 https://www.eurex.com/ex-en/data/statistics/monthly-statistics
          ⚠ 默认只挂 10 条（≈6 个月），历史必须走翻页接口：
          .../monthly-statistics/3848!search?pageNum={0..12}&hitsPerPage=50&sort=freshness%20%20desc
          实测 13 页 615 条 = 333 个月（1998-11 → 2026-07），xls 覆盖 283 个月自 2003-01。
   直链   https://www.eurex.com/resource/blob/{blobid}/{hash}/data/monthlystat_{YYYYMM}.xls
          blob id **每月都变**，必须从列表页拿，不能拼。
   格式   BIFF/OLE2 的 .xls（`Composite Document File V2`），要 xlrd；openpyxl 打不开。
          sheet = 「Cover」+「Eurex Monthly Statistics」。
   本模块只吃 2008-01 起的当前版式（223 期全部实测解析通过）。2003-01～2007-12 是
   单 sheet「Eurex Statistics」旧版式，结构完全不同，遇到直接 raise 而不是猜。

C. FWB 现货月度统计（快腿，Xetra / Frankfurt 分场所与分资产类别的唯一来源）
   列表页 https://www.cashmarket.deutsche-boerse.com/cash-en/Data-Tech/statistics/
          Turnover-Statistics/monthly-statistics-cash-market
          翻页接口同构：.../monthly-statistics-cash-market/4090756!search?pageNum=0&hitsPerPage=50&sort=sDate%20desc
   直链   .../resource/blob/{blobid}/{hash}/data/FWB_Monthly_Cash_Market_Statistics.{YYYYMMDD}.xls
          （href 是**相对路径**，Eurex 那边是绝对路径，两家写法不同）
   格式   BIFF .xls（xlrd），24 张 sheet，本模块只读「Cover」+「Total View」。
   ⚠ 官方**只挂 20 期**（2024-12 → 2026-07），这是全模块唯一的浅坑。见口径坑 9。

不进管道的两个源（只作人工核对）：现货月度新闻稿（URL 里的 id 不可预测）、
Clearstream 自己的月报页（newsroom 列表页是 Next.js 客户端渲染，无 sitemap.xml、
`?page=` 无效，**发现不了新链接**，因此不可能无人值守）。Clearstream 那份的价值是
给出 ICSD / CSD / IFS 三分拆，而 IR 台账只给合并数 —— 见口径坑 2。

════════════════════════════════════════════════════════════════════════════
发布节奏
════════════════════════════════════════════════════════════════════════════
下表的「实测节奏」是把 223 期 Eurex + 20 期 FWB **全部下下来**逐期算出来的，
不是抽样，也不是官网承诺：

| 源            | 实测节奏（月末后第几天）                    | 权威「发布日」字段              |
|---------------|--------------------------------------------|--------------------------------|
| Eurex xls     | 223 期里 208 期在第 2–5 天（中位第 2 天）；**2016-01 之后的 127 期全部在第 2–6 天** | Cover「Created on:」字符串 DD.MM.YYYY |
| FWB xls       | 20 期里 18 期在第 1–4 天                    | Cover「Created on:」**Excel 日期序列号** |
| IR 台账 xlsx  | 次月约第 10 天                              | docProps/core.xml 的 dcterms:modified |

落地页原文：*"available as of the second week after the reporting month"*；
2026-06 那期的 dcterms:modified = 2026-07-10T12:41:26Z，与配套 PDF 的
CreationDate D:20260710132703+02'00' 同日互证。

**闸门 LAG 建议 (常规月 6, 季末月 6)** —— 语义是「月末后第几天」。取 6 而不是 5：
2016 年以来两条快腿的最晚实测就是第 6 天（Eurex 2021-04 期 Cover = 2021-05-06），
按 5 设会每隔几年迟到一次。不要按慢腿的第 10 天设，那会让页面白等 4 天；
慢腿列靠「只填空回补」在后面几天自然补齐，与 fetch/cboe.py 的 RPC 完全同构。

**source_dates.csv 只记「本次运行确立的那一个月」**，日期取
max(Eurex Cover Created on, FWB Cover Created on) —— 两条快腿都到齐，该月的成交列
才写全。**绝不给回补月份补记发布日**，理由见口径坑 3（Created on 是文件生成日，
重述过的老文件会给出一个荒谬的日期）。

════════════════════════════════════════════════════════════════════════════
口径坑（按踩坑概率排序 —— 前四条是本模块每个月都会碰到的）
════════════════════════════════════════════════════════════════════════════
1. **三条腿两种节奏，新月天然缺列，这不是解析失败。**
   次月 1–5 日只能拿到 Eurex + FWB 的成交列；Clearstream / OTC / 360T / 商品 / IMS
   要等次月约 10 日的 IR 台账。所以：
   · 新月建行时慢腿格必须允许为空，不能因此抛异常；
   · 必须做「已有值永不覆盖、只填空」的回补，否则每个月的 Clearstream 列永久留白。
   慢腿列清单见下面「列口径表」的 leg=dbg 一栏，也可直接读模块常量
   `SLOW_LEG_COLUMNS`（给 build/pools.py 排除用，见口径坑 12）。

2. **「Settlement transactions」只含 ICSD，不含德国本土 CSD —— 而同一张表的 AuC 却是
   ICSD+CSD 合并。同一份文件里两行两个口径，这是本家最阴的一个坑。**
   两期独立实证（数字取自 Clearstream 自己的月报页）：
     2025-12 台账 settlement = 9.823717 m；Clearstream 稿：ICSD 9.8 / CSD 17.6 / IFS 5.9
     2026-03 台账 settlement = 11.867241 m；Clearstream 稿：ICSD 12 / CSD 25 / IFS 8
   两次都只等于 ICSD。同期 AuC 却对得上合并数：
     2025-12 auc_securities_services = 16,788.0104 €bn ≡ ICSD 9,756 + CSD 7,032
     2026-03 auc_securities_services = 17,135.8619 €bn ≡ ICSD 10,091 + CSD 7,045
   ⇒ 列名必须写成 `settle_icsd_txn_mn`，绝不能叫「Clearstream 结算笔数」；
     图注要写「不含德国本土 CSD 的约 2 倍笔数」。

3. **三条腿都会重述历史，而且比侦察阶段以为的频繁得多。本模块因此一律不覆盖。**
   本轮把 2008-01 起 223 期 Eurex 工作簿全下下来逐月比对，实测：
   · Eurex 月成交总数 vs IR 台账 `vol_fd_total_contracts`：**210 个可比月里 50 个不等**
     （不是侦察稿说的「只有 2016-01 差 3,282」）。差额从 570 到 784,771 张不等。
   · Eurex 未平仓的「上月 OI」列 vs 上一期文件的 OI 列：222 对里 **17 对不等**，
     最大 2.48%（2008-07），最近一次 2024-06。⇒ **OI 同样会被重述。**
   · FWB(Xetra+Frankfurt) vs 台账 `turnover_cash_total_eurbn`：30 个重叠月里 29 个
     **精确到分**相等，唯一例外 2025-12 差 €371.6m（0.32%）—— 而那一期 FWB 文件的
     Cover「Created on」是 2026-01-29，列表页却写 Jan 01, 2026，正是被重发过的证据。
   · Eurex 2024-02 ~ 2024-05 四期的 Cover「Created on」= **2025-05-06**，比列表页
     的原始发布日晚 340~432 天 —— 整批被重新生成过。全量比对 Cover 日 vs 列表页日：
     **223 期里 64 期不等**（老文件多半差 1–3 天，2024 那四期差一年多）。
     侦察阶段说这两个日期「逐日相同」，那只对最新一期成立。
   ⇒ 三条后果：(a) 已有单元格永不覆盖；(b) 冲突写进 `cache/db1_restatements.csv`
     供人工判断，不自动吞；(c) **Cover「Created on」不能当历史月份的发布日**，
     只能给「本次运行刚确立的那个月」作证（那时文件是新发的，实测滞后 2–5 天）。

4. **IR 台账的三分类 ≠ Eurex 工作簿的分组树，两边不要互相验算。**
   官方 PDF 两条脚注写死了这件事：总数不等于分项之和（因为含 ETC / 农产品 /
   贵金属等），且股息衍生品被摊进了股指与单股两项。本轮全量实测：
   · `vol_fd_rates_contracts` vs Eurex「Interest Rate Derivatives」组小计：
     222 个月里 **48 个不等**，2016-05～07 三个月，以及 2021-12 起几乎每月。
   · 就算按官方脚注把 Eurex 的 Dividend 组加回股指+单股，与台账两列之和相比
     **222 个月里 217 个仍然不等** —— 脚注解释不了这个缺口。
   ⇒ 本模块把两套并排入库、分别命名，**永不互相校验**：
     `vol_fd_*_contracts` = IR 台账口径（月成交张数）；
     `adv_eurex_*_contracts` / `oi_eurex_*_contracts` = Eurex 工作簿口径。
     下游画图任选其一，但**不要把两套混在同一条线里**。

5. **Eurex 工作簿是四层树，组小计必须按标签定位，而且不能只认 B 列的 `Sum`。**
   A 列=大组、B 列=产品族、C 列=子族、D 列=产品名、E 列=产品代码，F 列起是数值。
   `Sum` 字面量出现在**被它汇总的那一层的上一列**：C 列的 `Sum` 收 B 层，
   B 列的 `Sum` 收 A 层（=组小计），A 列的 `Sum` 是全表总计。
   这里有两个各自会静默出错的陷阱，本模块的定位规则是同时躲开这两个才成立的：
   (a) ⚠ 只有一个产品族的组**没有 B 列 `Sum` 行**，它的组小计落在 C 列的 `Sum` 上。
       2008-01 的 Volatility(437 张) 与 Inflation(1 张) 就是这样 —— 只认 B 列会漏掉
       这 438 张，闭合检验差 438 而查不出原因。⇒ 必须取「该组的 `Sum` 里层级最高
       （字面量所在列最靠左）的那一行」。
   (b) ⚠ **同一个组的 A 列标签会因分页而与别的组交替出现。**2015-07 的
       Capital Market Derivatives 与 Foreign Exchange Derivatives 在 R2508–R2571
       之间来回横跳，按「标签变了就切一段」划区间会切出 16 段而不是 11 个组，
       同一个小计算两遍，闭合检验多 888 张。⇒ 必须把每个 `Sum` 行挂到**它上面最近
       的那个 A 标签**上，而不是先划区间再找 Sum。
   实测 **2008-01 ~ 2026-07 全部 223 期**：每组的最高层 `Sum` 行有且只有一行，
   组小计之和与 A 列 Sum 行**逐位相等，零例外**（所以闭合检验不给容差）。
   组名逐年变（223 期里出现过 18 个不同组名，如 Property / Weather / Emission /
   Capital Market Derivatives 早已消失；ETF 组用过三个名字、股息组用过两个），
   **必须遍历取组名，不要写死清单**。
   另：A 列末尾有一行 EEX 合作脚注，靠「挂到最近 A 标签 + 只认 Sum 行」天然绕开。
   列位置（0-indexed）：5=Traded Contracts、6=Daily average、15=Capital Volume EUR、
   25=Paid Premiums、**35=Open Interest**、40=Capital OI。34 是 Change YtD，极易记错。

6. **`trading_days_eurex` 与 `trading_days_cash` 是两个不同的日历，不能互相顶替。**
   实测 222 个可比月里 **27 个不等**，而且有清清楚楚的原因：
   · 德国统一日（10/3）与圣灵降临节周一：**Eurex 开、Xetra 关**。
     201410 / 201610 / 201810 / 201910（10/3）与 201906 / 202006 / 202105（圣灵降临节
     周一）七个月全部是 Eurex 比现货多 1 天。
   · 2010-01 ~ 2011-11 另有 20 个月对不上，且方向乱跳（201104 Eurex 19 / 台账 22，
     201109 Eurex 22 / 台账 19），那一段台账的交易日列不可信。
   ⇒ Xetra 的日均成交额**必须**除以 `trading_days_cash`，不能除 `trading_days_eurex`；
     而 `trading_days_cash` 是慢腿。这就是本模块**不写现货 ADV 列**的原因，见口径坑 7。

7. **本模块不写任何现货 ADV 列 —— 现货只入库月度总额，日均由 build 层自己除。**
   官方现货工作簿只给月度总额，不给 ADV；官方新闻稿的 ADV 正是
   `月总额 ÷ 现货交易日数`（2026-07：157.51 / 23 = 6.85，与新闻稿逐位一致）。
   把这一步放进 fetcher 会有两个坏处：(a) 除数 `trading_days_cash` 是慢腿，
   算出来的 ADV 会跟着变成慢腿列，正好丢掉快腿的意义；(b) 一旦误用
   `trading_days_eurex`，27 个月会静默出错（口径坑 6）。
   ⇒ build 层请写：`adv_xetra_eurbn = turnover_xetra_eurbn / trading_days_cash`。
     Eurex 那边不受影响 —— 工作簿**自己就发 Daily average**，`adv_eurex_*` 是原始
     披露字段而非本模块算出来的（实测 = 月总张数 ÷ trading_days_eurex，逐位吻合）。

8. **EEX 三个商品列的表头写着「(in TWh)」，单元格里其实是 MWh，差 10⁶。**
   2026-06 单元格：power spot 88,034,094.5 / power deriv 960,720,291 /
   gas 679,742,442.47；同期官方 PDF：88.0 / 960.7 / 679.7 **TWh**。
   本模块按单元格原值入库、列名写 `_mwh`，**零算术**，把换算留给 build。
   （同一张表还有一个反向的表头笔误：`Total order book volume (in €m)` 实际单位是 EUR ——
   2026-06 那格 188,944,170,073.75 对应新闻稿的 €188.94 bn。别照表头换算。）

9. **FWB 分资产类别只有 20 期（2024-12 起），但场所级月度值能白捡到 2024-01。**
   每期「Total View」除了当月的 5×2 分类块，底部还有两块：本年度**逐月**的
   Xetra / Frankfurt order book turnover（EUR），以及 2001–2025 的**年度**值（Mio. EUR）。
   把 20 期全下一遍，`turnover_xetra_eurbn` / `turnover_fwb_eurbn` 因此覆盖
   **2024-01 → 2026-07 共 31 个月**（本模块已实现，跨文件重叠月逐位一致，零冲突）；
   而分资产类别列仍只有各期自己的报告月。年度块不入库（本仓只做月度）。
   再深的现货历史只能用 `turnover_cash_total_eurbn`（台账口径，Xetra+Frankfurt 合计，
   2010-01 起）—— 但那是慢腿。
   附带一个后果：「Total View」的行数每月加一行，**任何按行号写死的解析都会漂**，
   所以本模块一律按 A 列文本匹配行。

10. **order book turnover 是单边（single-counted）。**
    FWB 工作簿「Explanation Report」表自己写明成交额按单边计；IR 台账 PDF 的
    现货那行也挂着脚注 Single-counted。跨家比要注意：HKEX 的南向 ADT 是双边。
    另外 **Xetra ≠ Deutsche Börse 现货全部**：2026-07 新闻稿的 €163.37bn 是
    Xetra(XETR) + Frankfurt(XFRA) 合计，其中 Xetra €157.51bn、Frankfurt €5.86bn。

11. **Clearstream 的 AuC 不是期末时点数，是月内平均值。**
    官方 PDF 原文把这几行标成 `(average value)` / `(average outstandings)`：
    托管资产、OTC 名义未平仓、担保品在外量三项都是月内均值。
    ⇒ 下游换汇时**配月均汇率，不要配月末汇率**；做「月末跳变」类叙事会踩空。
    与 `hkex.mktcap_hkdtn`（真·月末时点）放同一张图必须注明。
    ⚠ 唯一没标的是 `aum_stoxx_dax_etf_eurbn`（STOXX/DAX 挂钩 ETF 资产）——
    同一页别的行都标了它不标，暗示是期末数，但**这是推断不是证据**，
    官方未见明文。本模块把它的 kind 标成 `stock_unknown`，
    在拿到官方定义前**不要**拿它跟 `msci.aum_eop_usdbn` 比水平值。

12. **慢腿列绝对不能进 build/exchanges.py 的横截面 panel。**
    该文件对共同窗口内的空洞直接 `raise SystemExit`，而慢腿列**天生**会在最新月留空
    （要等次月约 10 日）。进池的只能是快腿列。模块常量 `FAST_LEG_COLUMNS` /
    `SLOW_LEG_COLUMNS` 就是给 build/pools.py 做这个排除用的，别再手抄一份清单。

13. **blob URL 的 hash 段不校验，但 blob id 每月变（IR 台账那份除外）。**
    实测把 hash 换成 32 个 0 / deadbeef…，三个域名照样 200 返回同一文件。
    但 Eurex / FWB 每期一个新 blob id，**必须解析列表页拿 id**；
    只有 major-business-figures_en.xlsx 的 id `249090` 常驻，可写死作兜底。

14. **Eurex Cover 的「Reported month:」格式变过**（2008-01 是 '2008 January'，
    2010 起是 'January 2010'），别拿它当解析锚点 —— 用文件名里的 YYYYMM。
    Cover 只用来取「Created on:」，而且两家写法不同：Eurex 是字符串 '04.08.2026'
    （DD.MM.YYYY 欧洲序），FWB 是 Excel 日期序列号 46237.0（要 xldate_as_datetime）。

15. **eurex.com 偶发连接超时**（本轮 223 次下载里出现 1 次 60s timeout，不是 403、
    不是 TLS 指纹拦截）。所以下载带 3 次重试 + 退避；不需要 curl_cffi / nscurl。

════════════════════════════════════════════════════════════════════════════
列口径表 —— 下游换算全靠这张表，每一列都写死了「谁发的 / 存量还是流量 / 单位」
════════════════════════════════════════════════════════════════════════════
leg   : eurex = Eurex 工作簿（快腿）｜ fwb = FWB 现货工作簿（快腿）｜ dbg = IR 台账（慢腿）
kind  : flow_month  当月累计（配月均汇率）
        flow_daily  当月日均（配月均汇率）
        stock_eop   期末时点存量（配月末汇率）
        stock_avg   月内平均存量（**配月均汇率**，不是月末）
        stock_unknown 存量但官方未言明时点口径（见口径坑 11）
        count       计数，无汇率

| 列名                              | leg   | kind       | 单位/口径 |
|-----------------------------------|-------|------------|-----------|
| trading_days_eurex                | eurex | count      | Eurex 交易日数 |
| adv_eurex_total_contracts         | eurex | flow_daily | 全所日均成交合约数（张，非千张）。官方 Daily average 字段 |
| adv_eurex_rates_contracts         | eurex | flow_daily | Eurex「Interest Rate Derivatives」组，含 EURIBOR/欧债全系 |
| adv_eurex_index_contracts         | eurex | flow_daily | Eurex「Equity Index Derivatives」组，**不含股息衍生品** |
| adv_eurex_equity_contracts        | eurex | flow_daily | Eurex「Equity Derivatives」组＝单股期货+单股期权，**不含股息** |
| adv_eurex_dividend_contracts      | eurex | flow_daily | Eurex「Dividend Derivatives」组（2008-06 起才有此组） |
| oi_eurex_*_contracts              | eurex | stock_eop  | 月末未平仓合约数（张），分组同上 |
| adv_/oi_bund/bobl/schatz/btp/oat  | eurex | 同上       | FGBL/FGBM/FGBS/FBTP/FOAT 期货 |
| adv_/oi_euribor3m_contracts       | eurex | 同上       | FEU3 三个月 EURIBOR 期货 |
| adv_/oi_estoxx50_fut/opt          | eurex | 同上       | FESX / OESX |
| adv_/oi_dax_fut/opt               | eurex | 同上       | FDAX / ODAX |
| adv_/oi_vstoxx_fut_contracts      | eurex | 同上       | FVS（对标 cboe.adv_vix_futures_kcontracts） |
| turnover_xetra_eurbn              | fwb   | flow_month | Xetra(XETR) 电子盘 order book 月度总额，**单边**，€bn |
| turnover_fwb_eurbn                | fwb   | flow_month | Frankfurt(XFRA) 场内，单边，€bn |
| turnover_xetra_{equities,etp,structured}_eurbn | fwb | flow_month | Xetra 分资产类别，单边，€bn |
| turnover_fwb_{equities,etp,bonds,funds,structured}_eurbn | fwb | flow_month | Frankfurt 分资产类别 |
| trading_days_cash                 | dbg   | count      | **现货**交易日数（≠ Eurex，见口径坑 6） |
| vol_fd_total_contracts            | dbg   | flow_month | IR 口径全所月成交合约数（张）。含 ETC/农产品/贵金属，≠ 分项之和 |
| vol_fd_index_contracts            | dbg   | flow_month | IR 口径股指衍生品，**已把股息衍生品摊进来** |
| vol_fd_equity_contracts           | dbg   | flow_month | IR 口径单股衍生品，同样含摊入的股息 |
| vol_fd_rates_contracts            | dbg   | flow_month | IR 口径利率衍生品 |
| otc_notional_outstanding_eurbn    | dbg   | stock_avg  | EurexOTC Clear 名义未平仓，**月内平均值**，€bn |
| otc_notional_cleared_eurbn        | dbg   | flow_month | EurexOTC Clear 当月清算名义量（含压缩），€bn |
| turnover_cash_total_eurbn         | dbg   | flow_month | Xetra+Frankfurt 合计 order book，单边，€bn（2010-01 起，深史用它） |
| auc_securities_services_eurbn     | dbg   | stock_avg  | Clearstream **ICSD+CSD** 托管资产，月内平均，€bn（17,459 €bn = 17.46 €tn） |
| settle_icsd_txn_mn                | dbg   | flow_month | ⚠ **只含 ICSD** 的结算笔数（百万笔），见口径坑 2 |
| gsf_collateral_eurbn              | dbg   | stock_avg  | Global Securities Financing 担保品在外量，月内平均，€bn |
| cash_balances_eurmn               | dbg   | stock_avg  | Clearstream 日均现金余额（含受制裁冻结账户），€m |
| auc_fund_services_eurbn           | dbg   | stock_avg  | Clearstream IFS（基金）托管资产，月内平均，€bn |
| settle_ifs_txn_mn                 | dbg   | flow_month | IFS 结算笔数（百万笔） |
| aum_stoxx_dax_etf_eurbn           | dbg   | stock_unknown | 挂钩 STOXX/DAX 的 ETF 资产，€bn。**时点口径官方未言明** |
| vol_licensed_index_contracts      | dbg   | flow_month | 交易所授权的指数衍生品成交合约数（张） |
| adv_360t_fx_eurbn                 | dbg   | flow_daily | 360T 外汇日均名义额，€bn（2018-07 起含 GTX） |
| vol_power_spot_mwh                | dbg   | flow_month | EEX 电力现货，**MWh**（表头写 TWh 是笔误，见口径坑 8） |
| vol_power_deriv_mwh               | dbg   | flow_month | EEX 电力衍生品，MWh |
| vol_gas_mwh                       | dbg   | flow_month | EEX 天然气，MWh |

**各列首个非空月**（实测，不是估计）：vol_fd_index/equity/rates 与 trading_days_cash
自 2002-01；gsf_collateral 自 2007-01；Eurex 全部列自 2008-01（dividend 组自 2008-06、
FVS 自 2009-05、FBTP 自 2009-09、FOAT 自 2012-04）；vol_fd_total / turnover_cash_total /
settle_* 自 2010-01；auc_* 与 aum_stoxx_dax_etf 自 2012-01；360T 与商品与 cash_balances
自 2015-01；otc_* 与 vol_licensed 自 2016-01；turnover_xetra/fwb 自 2024-01；
分资产类别列自 2024-12。早于首月的行天然为空，不是解析失败。

━━ 依赖 ━━ openpyxl（读 IR 台账 xlsx）、xlrd（读两个 .xls；xlrd≥2.0 不再读 xlsx，
所以两个库都要）。不依赖 pandas。
"""

import csv
import datetime
import os
import re
import time
import urllib.request

# ── 站点常量 ────────────────────────────────────────────────────────────────
DBG_LANDING = 'https://www.deutsche-boerse.com/dbg-en/investor-relations/statistics'
# blob id 249090 常驻多年，落地页解析不出来时作兜底（hash 段不校验，见口径坑 13）
DBG_FALLBACK = ('https://www.deutsche-boerse.com/resource/blob/249090/'
                '00000000000000000000000000000000/data/major-business-figures_en.xlsx')
EUREX_SEARCH = ('https://www.eurex.com/ex-en/data/statistics/monthly-statistics/'
                '3848!search?pageNum=%d&hitsPerPage=50&sort=freshness%%20%%20desc')
EUREX_HOST = 'https://www.eurex.com'
FWB_SEARCH = ('https://www.cashmarket.deutsche-boerse.com/cash-en/Data-Tech/statistics/'
              'Turnover-Statistics/monthly-statistics-cash-market/'
              '4090756!search?pageNum=%d&hitsPerPage=50&sort=sDate%%20desc')
FWB_HOST = 'https://www.cashmarket.deutsche-boerse.com'

_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

# Eurex 工作簿只有 2008-01 起是当前版式；2003-2007 是单 sheet 旧版式，不猜，直接不碰。
EUREX_START = '2008-01'
# IR 台账自己的起点，也是 series/db1.csv 的起点
LEDGER_START = '2002-01'
# source_dates 的合理上界（天，自月末算起）。全量实测 223 期 Eurex + 20 期 FWB：
# 2016-01 之后的 127 期 Eurex 全部落在第 2–6 天，FWB 20 期里 18 期落在第 1–4 天；
# 超出这个带宽的都是**被重新生成过的文件**（Eurex 2024-02~05 四期写着 2025-05-06，
# 滞后 340~432 天；FWB 2025-01 与 2025-12 两期滞后 26 / 29 天）。
# 取 20：宁可让页面少印半句「官方发布于」，也不要印一个像模像样的重述日。
MAX_PUBLISH_LAG_DAYS = 20

RESTATEMENT_LOG = 'db1_restatements.csv'


class Db1FetchError(RuntimeError):
    """源站结构变化 / 下载失败 / 解析结果不完整。一律炸掉，绝不静默写空列或 NaN。"""


# ── 列定义 ──────────────────────────────────────────────────────────────────
# (列名, leg, kind, 首个应有值的月份)
# first_month 用于校验：>= first_month 的月份该列必须有值，否则说明解析错行。
_META = [
    # ── 快腿 A：Eurex 工作簿 ──
    ('trading_days_eurex',             'eurex', 'count',      '2008-01'),
    ('adv_eurex_total_contracts',      'eurex', 'flow_daily', '2008-01'),
    ('adv_eurex_rates_contracts',      'eurex', 'flow_daily', '2008-01'),
    ('adv_eurex_index_contracts',      'eurex', 'flow_daily', '2008-01'),
    ('adv_eurex_equity_contracts',     'eurex', 'flow_daily', '2008-01'),
    ('adv_eurex_dividend_contracts',   'eurex', 'flow_daily', '2008-06'),
    ('oi_eurex_total_contracts',       'eurex', 'stock_eop',  '2008-01'),
    ('oi_eurex_rates_contracts',       'eurex', 'stock_eop',  '2008-01'),
    ('oi_eurex_index_contracts',       'eurex', 'stock_eop',  '2008-01'),
    ('oi_eurex_equity_contracts',      'eurex', 'stock_eop',  '2008-01'),
    ('oi_eurex_dividend_contracts',    'eurex', 'stock_eop',  '2008-06'),
    ('adv_bund_contracts',             'eurex', 'flow_daily', '2008-01'),
    ('oi_bund_contracts',              'eurex', 'stock_eop',  '2008-01'),
    ('adv_bobl_contracts',             'eurex', 'flow_daily', '2008-01'),
    ('oi_bobl_contracts',              'eurex', 'stock_eop',  '2008-01'),
    ('adv_schatz_contracts',           'eurex', 'flow_daily', '2008-01'),
    ('oi_schatz_contracts',            'eurex', 'stock_eop',  '2008-01'),
    ('adv_btp_contracts',              'eurex', 'flow_daily', '2009-09'),
    ('oi_btp_contracts',               'eurex', 'stock_eop',  '2009-09'),
    ('adv_oat_contracts',              'eurex', 'flow_daily', '2012-04'),
    ('oi_oat_contracts',               'eurex', 'stock_eop',  '2012-04'),
    ('adv_euribor3m_contracts',        'eurex', 'flow_daily', '2008-01'),
    ('oi_euribor3m_contracts',         'eurex', 'stock_eop',  '2008-01'),
    ('adv_estoxx50_fut_contracts',     'eurex', 'flow_daily', '2008-01'),
    ('oi_estoxx50_fut_contracts',      'eurex', 'stock_eop',  '2008-01'),
    ('adv_estoxx50_opt_contracts',     'eurex', 'flow_daily', '2008-01'),
    ('oi_estoxx50_opt_contracts',      'eurex', 'stock_eop',  '2008-01'),
    ('adv_dax_fut_contracts',          'eurex', 'flow_daily', '2008-01'),
    ('oi_dax_fut_contracts',           'eurex', 'stock_eop',  '2008-01'),
    ('adv_dax_opt_contracts',          'eurex', 'flow_daily', '2008-01'),
    ('oi_dax_opt_contracts',           'eurex', 'stock_eop',  '2008-01'),
    ('adv_vstoxx_fut_contracts',       'eurex', 'flow_daily', '2009-05'),
    # FVS 2009-05 是挂牌首月，成交为 0、月末未平仓那一格官方留空（实测 223 期里
    # 唯一一个产品级空格）。所以 OI 的起点比 ADV 晚一个月，这不是解析失败。
    ('oi_vstoxx_fut_contracts',        'eurex', 'stock_eop',  '2009-06'),
    # ── 快腿 B：FWB 现货工作簿 ──
    ('turnover_xetra_eurbn',           'fwb',   'flow_month', '2024-01'),
    ('turnover_fwb_eurbn',             'fwb',   'flow_month', '2024-01'),
    ('turnover_xetra_equities_eurbn',  'fwb',   'flow_month', '2024-12'),
    ('turnover_xetra_etp_eurbn',       'fwb',   'flow_month', '2024-12'),
    # Xetra 的结构化产品格常空（20 期里 5 期有数，最大 0.0029 €bn），故不设 first_month
    ('turnover_xetra_structured_eurbn', 'fwb',  'flow_month', None),
    ('turnover_fwb_equities_eurbn',    'fwb',   'flow_month', '2024-12'),
    ('turnover_fwb_etp_eurbn',         'fwb',   'flow_month', '2024-12'),
    ('turnover_fwb_bonds_eurbn',       'fwb',   'flow_month', '2024-12'),
    ('turnover_fwb_funds_eurbn',       'fwb',   'flow_month', '2024-12'),
    ('turnover_fwb_structured_eurbn',  'fwb',   'flow_month', '2024-12'),
    # ── 慢腿：IR 台账 ──
    ('trading_days_cash',              'dbg',   'count',      '2002-01'),
    ('vol_fd_total_contracts',         'dbg',   'flow_month', '2009-01'),
    ('vol_fd_index_contracts',         'dbg',   'flow_month', '2002-01'),
    ('vol_fd_equity_contracts',        'dbg',   'flow_month', '2002-01'),
    ('vol_fd_rates_contracts',         'dbg',   'flow_month', '2002-01'),
    ('otc_notional_outstanding_eurbn', 'dbg',   'stock_avg',  '2016-01'),
    ('otc_notional_cleared_eurbn',     'dbg',   'flow_month', '2016-01'),
    ('turnover_cash_total_eurbn',      'dbg',   'flow_month', '2010-01'),
    ('auc_securities_services_eurbn',  'dbg',   'stock_avg',  '2012-01'),
    ('settle_icsd_txn_mn',             'dbg',   'flow_month', '2010-01'),
    ('gsf_collateral_eurbn',           'dbg',   'stock_avg',  '2007-01'),
    ('cash_balances_eurmn',            'dbg',   'stock_avg',  '2015-01'),
    ('auc_fund_services_eurbn',        'dbg',   'stock_avg',  '2012-01'),
    ('settle_ifs_txn_mn',              'dbg',   'flow_month', '2010-01'),
    ('aum_stoxx_dax_etf_eurbn',        'dbg',   'stock_unknown', '2012-01'),
    ('vol_licensed_index_contracts',   'dbg',   'flow_month', '2016-01'),
    ('adv_360t_fx_eurbn',              'dbg',   'flow_daily', '2015-01'),
    ('vol_power_spot_mwh',             'dbg',   'flow_month', '2015-01'),
    ('vol_power_deriv_mwh',            'dbg',   'flow_month', '2015-01'),
    ('vol_gas_mwh',                    'dbg',   'flow_month', '2015-01'),
]

COLUMNS = ['month'] + [m[0] for m in _META]
COLUMN_META = {m[0]: {'leg': m[1], 'kind': m[2], 'first_month': m[3]} for m in _META}

# 给 build/pools.py 用：慢腿列**天生**会在最新月留空，进横截面 panel 会让
# build/exchanges.py 的空值护栏 raise SystemExit（见口径坑 12）。
SLOW_LEG_COLUMNS = frozenset(c for c, meta in COLUMN_META.items() if meta['leg'] == 'dbg')
FAST_LEG_COLUMNS = frozenset(c for c, meta in COLUMN_META.items() if meta['leg'] != 'dbg')

_COLS_BY_LEG = {leg: [c for c, m in COLUMN_META.items() if m['leg'] == leg]
                for leg in ('eurex', 'fwb', 'dbg')}

# ── Eurex 工作簿的组与产品映射 ──────────────────────────────────────────────
# 组名逐年改（223 期里 18 个不同名字），所以这里给的是**别名列表**而不是唯一名字。
_EUREX_GROUPS = [
    ('rates',    ('Interest Rate Derivatives',)),
    ('index',    ('Equity Index Derivatives',)),
    ('equity',   ('Equity Derivatives',)),
    # 2013 年前后这个组叫过 'Equity Index Dividend Derivatives'
    ('dividend', ('Dividend Derivatives', 'Equity Index Dividend Derivatives')),
]
# (列名后缀, 产品代码)。产品代码是 E 列（0-indexed 4），全表唯一（实测 3365 个代码无重复）。
_EUREX_PRODUCTS = [
    ('bund', 'FGBL'), ('bobl', 'FGBM'), ('schatz', 'FGBS'),
    ('btp', 'FBTP'), ('oat', 'FOAT'), ('euribor3m', 'FEU3'),
    ('estoxx50_fut', 'FESX'), ('estoxx50_opt', 'OESX'),
    ('dax_fut', 'FDAX'), ('dax_opt', 'ODAX'),
    # 2008-01~2010-12 另有旧代码 FVSX（旧 VSTOXX 指数的期货），与 FVS 有重叠期，
    # 不入库：两者标的指数不同，接在一起会造出一条假的连续序列。
    # ⚠ 给 build/notional.py：FVS 这一行的**产品名换过**（前 93 期叫
    # 'Mini-Futures auf VSTOXX®'，后 114 期叫 'Futures on VSTOXX®'），代码始终是 FVS。
    # 本模块没有核实过换名前后的合约乘数是否相同 —— 换算名义额前请自己去查
    # Eurex 的产品规格，别默认它一直没变。
    ('vstoxx_fut', 'FVS'),
]
_EUREX_SHEET = 'Eurex Monthly Statistics'
_C_TRADED, _C_ADV, _C_OI = 5, 6, 35        # 0-indexed；34 是 Change YtD，别记错

# ── FWB「Total View」的行标签 ───────────────────────────────────────────────
# (列名后缀, A 列文本)。Total 行必须等于这 5 组之和（两个场所各自闭合）。
_FWB_GROUPS = [
    ('equities',   'Equities'),
    ('etp',        'ETFs, ETCs, ETNs'),
    ('bonds',      'Bonds'),
    ('funds',      'Funds'),
    ('structured', 'Structured Products and Other Instruments'),
]
# 只有这些 (场所, 组) 会入库。Xetra 的 bonds / funds 在 20 期里恒为空或 0
# （债券与基金只在 Frankfurt 场内成交），建列等于永久留白，故不建；
# 但解析仍要读它们，否则 Total 闭合检验对不上。
_FWB_STORE = {('xetra', 'equities'), ('xetra', 'etp'), ('xetra', 'structured'),
              ('fwb', 'equities'), ('fwb', 'etp'), ('fwb', 'bonds'),
              ('fwb', 'funds'), ('fwb', 'structured')}
_FWB_SHEET = 'Total View'

# ── IR 台账的列定位 ─────────────────────────────────────────────────────────
# 按 (第 2 行的分部标题, 第 4 行的列标题) 定位，**不写死列号** —— 官方哪天在中间插一列，
# 写死列号会静默取到隔壁的数（比如把 IFS 的 AuC 当成 Clearstream 的 AuC）。
# 实测这 23 个 (分部, 标题) 二元组全表唯一。scale 只做纯粹的量纲缩放，不改口径。
_DBG_MAP = {
    ('TRADING & CLEARING', 'Total (traded contracts)'):                      ('vol_fd_total_contracts', 1.0),
    ('TRADING & CLEARING', 'Index (traded contracts)'):                      ('vol_fd_index_contracts', 1.0),
    ('TRADING & CLEARING', 'Equity (traded contracts)'):                     ('vol_fd_equity_contracts', 1.0),
    ('TRADING & CLEARING', 'Interest rates (traded contracts)'):             ('vol_fd_rates_contracts', 1.0),
    ('TRADING & CLEARING', 'Notional outstanding volumes (in €bn)'):         ('otc_notional_outstanding_eurbn', 1.0),
    ('TRADING & CLEARING', 'Notional cleared volumes (incl. compr.) (in €bn)'): ('otc_notional_cleared_eurbn', 1.0),
    ('TRADING & CLEARING', 'Power spot volume (in TWh)'):                    ('vol_power_spot_mwh', 1.0),
    ('TRADING & CLEARING', 'Power derivatives volume (in TWh)'):             ('vol_power_deriv_mwh', 1.0),
    ('TRADING & CLEARING', 'Gas volume (in TWh)'):                           ('vol_gas_mwh', 1.0),
    ('TRADING & CLEARING', '360T (ADV) FX market (in €bn)'):                 ('adv_360t_fx_eurbn', 1.0),
    # 表头写 (in €m) 是陈年笔误，单元格其实是 EUR（口径坑 8）→ 除 1e9 得 €bn
    ('TRADING & CLEARING', 'Total order book volume (in €m)'):               ('turnover_cash_total_eurbn', 1e-9),
    ('TRADING & CLEARING', 'Trading days'):                                  ('trading_days_cash', 1.0),
    ('SECURITIES SERVICES', 'Assets under custody (in €bn)'):                ('auc_securities_services_eurbn', 1.0),
    ('SECURITIES SERVICES', 'Settlement transactions (in m)'):               ('settle_icsd_txn_mn', 1.0),
    ('SECURITIES SERVICES', 'Collateral management (average outstandings in €bn)'): ('gsf_collateral_eurbn', 1.0),
    ('SECURITIES SERVICES', 'Total (in €m)'):                                ('cash_balances_eurmn', 1.0),
    ('FUND SERVICES', 'Assets under custody IFS (in €bn)'):                  ('auc_fund_services_eurbn', 1.0),
    ('FUND SERVICES', 'Settlement transactions IFS (in m)'):                 ('settle_ifs_txn_mn', 1.0),
    ('INVESTMENT MANAGEMENT SOLUTIONS', 'Assets under management in STOXX & DAX ETFs (in €bn)'):
        ('aum_stoxx_dax_etf_eurbn', 1.0),
    ('INVESTMENT MANAGEMENT SOLUTIONS', 'Exchange licenses Index derivatives (traded contracts)'):
        ('vol_licensed_index_contracts', 1.0),
}


# ── 依赖 ────────────────────────────────────────────────────────────────────
def _openpyxl():
    try:
        import openpyxl                              # noqa: PLC0415
    except ImportError as e:                         # pragma: no cover
        raise Db1FetchError('db1 需要 openpyxl 读 IR 台账 xlsx：pip install openpyxl') from e
    return openpyxl


def _xlrd():
    """Eurex / FWB 的 .xls 是 1990 年代的 BIFF/OLE2，openpyxl 打不开，只能用 xlrd。

    注意 xlrd>=2.0 反过来不再读 .xlsx，所以两个库都得装，不能二选一。
    """
    try:
        import xlrd                                  # noqa: PLC0415
    except ImportError as e:                         # pragma: no cover
        raise Db1FetchError(
            'db1 需要 xlrd 读 Eurex/FWB 的 BIFF .xls：pip install xlrd') from e
    return xlrd


# ── 网络 ────────────────────────────────────────────────────────────────────
def _http_get(url, timeout=120, retries=3):
    """带重试的 GET。eurex.com 偶发 60s 连接超时（口径坑 15），一次性失败太脆。"""
    last = None
    for attempt in range(retries):
        req = urllib.request.Request(url, headers={
            'User-Agent': _UA,
            'Accept': '*/*',
            'Accept-Language': 'en-US,en;q=0.9',
        })
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read()
        except Exception as e:                        # noqa: BLE001
            last = e
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
    raise Db1FetchError('下载失败（%d 次重试后）%s: %r' % (retries, url, last))


def _cache(cache_dir, name):
    os.makedirs(cache_dir, exist_ok=True)
    return os.path.join(cache_dir, name)


def _download(url, path):
    data = _http_get(url)
    if len(data) < 50000:
        raise Db1FetchError('%s 只有 %d 字节，不像是正常的工作簿' % (url, len(data)))
    with open(path, 'wb') as f:
        f.write(data)
    return path


def _cached_download(url, path):
    """已经在 cache 里就不重下 —— 只用于**回补历史月份**。

    走到这里的月份要么 CSV 里那一格还空着，要么已经填过（那就不会走到这里），
    所以复用缓存不会让数据变旧；而官方重述过的老文件即使重下也不会被写进去
    （已有值永不覆盖，口径坑 3），重下 200 多个文件纯属浪费带宽。
    两条快腿的**最新一期**不走这条路，每次都用 _download 重下，见 update()。
    """
    if os.path.exists(path) and os.path.getsize(path) > 50000:
        return path
    return _download(url, path)


# ── 小工具 ──────────────────────────────────────────────────────────────────
def _norm(v):
    return re.sub(r'\s+', ' ', str(v)).strip() if v is not None else ''


def _num(v):
    """空白 / 非数字 → None；数字 → float。绝不把 '-' 之类当成 0。"""
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    if s in ('', '-', 'n/a', 'N/A'):
        return None
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return None


def _month_of(stamp):
    """'20260731' -> '2026-07'；'202607' -> '2026-07'。"""
    return '%s-%s' % (stamp[:4], stamp[4:6])


def _month_end(month):
    y, m = int(month[:4]), int(month[5:7])
    nxt = datetime.date(y + 1, 1, 1) if m == 12 else datetime.date(y, m + 1, 1)
    return nxt - datetime.timedelta(days=1)


def _months_between(start, end):
    y, m = int(start[:4]), int(start[5:7])
    out = []
    while True:
        cur = '%04d-%02d' % (y, m)
        if cur > end:
            return out
        out.append(cur)
        m += 1
        if m == 13:
            y, m = y + 1, 1


# ── 列表页发现 ──────────────────────────────────────────────────────────────
_EUREX_HIT = re.compile(
    r'<p class="search-result-date">([^<]+)</p>.*?'
    r'href="([^"]*?/data/monthlystat_(\d{6})\.xls)"', re.S)
_FWB_HIT = re.compile(
    r'<p class="search-result-date">([^<]+)</p>.*?'
    r'href="([^"]*?/data/FWB_Monthly_Cash_Market_Statistics\.(\d{8})\.xls)"', re.S)
_RESULT_BLOCK = re.compile(r'teasable-search-result-container')


def _listing(url_tmpl, host, hit_re, cache_dir, tag, all_pages):
    """翻页抓列表，返回 {'YYYY-MM': (绝对 url, 列表页自述发布日文本)}。

    all_pages=False 时只抓第 0 页 —— 排序是 freshness desc，最新一期必在首页，
    只想知道「官方最新月是几月」时不必把 13 页 615 条全抓一遍。
    列表页原样存进 cache，源站改版时可以事后取证。

    收尾判据用「这一页还有没有搜索结果块」，**不能用「这一页有没有 .xls」**：
    Eurex 归档里 1998-11~2002-12 那 50 期只有 pdf，按 freshness 排在最后一页，
    拿「没有 xls」当收尾信号，早晚会在某一页全是 pdf 时把后面的历史整段丢掉。
    """
    out, page = {}, 0
    while True:
        html = _http_get(url_tmpl % page).decode('utf-8', 'replace')
        with open(_cache(cache_dir, 'db1_%s_p%d.html' % (tag, page)), 'w',
                  encoding='utf-8') as f:
            f.write(html)
        blocks = len(_RESULT_BLOCK.findall(html))
        hits = hit_re.findall(html)
        if not hits and page == 0:
            raise Db1FetchError(
                '%s 列表页解析不到任何 .xls 直链，源站可能改版：%s' % (tag, url_tmpl % page))
        for day, href, stamp in hits:
            key = _month_of(stamp)
            if key not in out:
                out[key] = (href if href.startswith('http') else host + href, day.strip())
        page += 1
        if not all_pages or blocks == 0 or page > 20:
            break
    return out


def _dbg_url(cache_dir):
    """从落地页取 major-business-figures_en.xlsx 直链；取不到用常驻 blob id 兜底。"""
    html = _http_get(DBG_LANDING).decode('utf-8', 'replace')
    with open(_cache(cache_dir, 'db1_dbg_landing.html'), 'w', encoding='utf-8') as f:
        f.write(html)
    hit = re.search(r'/resource/blob/\d+/[0-9a-f]{32}/data/major-business-figures_en\.xlsx',
                    html)
    if hit:
        return 'https://www.deutsche-boerse.com' + hit.group(0)
    return DBG_FALLBACK


# ── 解析：IR 台账（慢腿）────────────────────────────────────────────────────
def _parse_dbg(path):
    """解析 major-business-figures_en.xlsx，返回 ({'YYYY-MM': {列: 值}}, 台账自述修改日)。

    列位置一律按 (第 2 行分部标题, 第 4 行列标题) 定位，任何一个约定的标题找不到就抛 ——
    官方在中间插一列时，写死列号的解析器会静默把隔壁的数写进来，而那种错没有任何征兆。
    """
    openpyxl = _openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheets = [s for s in wb.sheetnames if 'business figures' in s.lower()]
    if len(sheets) != 1:
        raise Db1FetchError('IR 台账里找不到唯一的「Major business figures」sheet：%r'
                            % (wb.sheetnames,))
    ws = wb[sheets[0]]

    section, colmap = '', {}
    for c in range(2, ws.max_column + 1):
        sec = _norm(ws.cell(2, c).value)
        if sec:
            section = sec
        label = _norm(ws.cell(4, c).value)
        if label and (section, label) in _DBG_MAP:
            name, scale = _DBG_MAP[(section, label)]
            if name in colmap:
                raise Db1FetchError('IR 台账里 (%s, %s) 出现两次，定位不再唯一'
                                    % (section, label))
            colmap[name] = (c, scale)
    missing = [v[0] for v in _DBG_MAP.values() if v[0] not in colmap]
    if missing:
        raise Db1FetchError('IR 台账缺列（官方表结构可能已变）：%s' % sorted(missing))

    data = {}
    for r in range(5, ws.max_row + 1):
        d = ws.cell(r, 1).value
        if not isinstance(d, datetime.datetime):
            continue
        mon = '%04d-%02d' % (d.year, d.month)
        rec = {}
        for name, (c, scale) in colmap.items():
            v = _num(ws.cell(r, c).value)
            rec[name] = None if v is None else v * scale
        if any(v is not None for v in rec.values()):
            data[mon] = rec
    if not data:
        raise Db1FetchError('IR 台账解析后一个有数的月份都没有')

    months = sorted(data)
    holes = [m for m in _months_between(months[0], months[-1]) if m not in data]
    if holes:
        raise Db1FetchError('IR 台账月份序列有断档，解析多半错了：%s' % holes[:12])
    return data, _xlsx_modified(path)


def _xlsx_modified(path):
    """取 xlsx 自己 docProps/core.xml 里的 dcterms:modified，返回 'YYYY-MM-DD' 或 None。

    这是台账**自己写下的**最后修订日（2026-06 那期 = 2026-07-10），与同页 PDF 的
    CreationDate 同日互证。目前只用于日志与人工排查：慢腿列是回补上去的，
    它不能代表任何一个具体月份的首发日，所以**不进 source_dates.csv**。
    """
    import zipfile                                    # noqa: PLC0415
    try:
        with zipfile.ZipFile(path) as z:
            xml = z.read('docProps/core.xml').decode('utf-8', 'replace')
    except Exception:                                 # noqa: BLE001
        return None
    hit = re.search(r'<dcterms:modified[^>]*>(\d{4}-\d{2}-\d{2})', xml)
    return hit.group(1) if hit else None


# ── 解析：Eurex（快腿 A）────────────────────────────────────────────────────
def _eurex_group_sums(sh):
    """定位每个大组的小计行，返回 ({组名: 行号}, 全表 Sum 行号)。

    规则：把每个写着 `Sum` 的行挂到**它上面最近的那个 A 列组名**下，再对每个组取
    「`Sum` 字面量所在列最靠左」的那一行（列越靠左，层级越高：B 列收 A 层=组小计，
    C 列收 B 层，D 列收 C 层）。

    两条都是踩出来的，缺一不可：
    · **不能按「A 列标签换了就切一段」来划组区间。**2015-07 的
      Capital Market Derivatives 与 Foreign Exchange Derivatives 两个组的标签在
      R2508–R2571 之间**交替出现**（分页重复），按区间切会切出 16 段而不是 11 个组，
      同一个小计被算两遍，闭合检验多出 888 张。挂到「最近的 A 标签」上则天然正确。
    · **不能只认 B 列的 `Sum`。**只有一个产品族的组根本不生成 B 列小计行，它的组小计
      落在 C 列的 `Sum` 上 —— 2008-01 的 Volatility(437 张) 与 Inflation(1 张) 就是这样，
      只认 B 列会把这 438 张整个漏掉。

    实测 2008-01~2026-07 全部 223 期：每个组的最高层 `Sum` 行**有且只有一行**，
    且组小计之和与 A 列 Sum 行逐位相等，零例外。所以这里对「只有一行」做硬断言 ——
    哪天官方改成一个组两段各带一个小计，宁可炸掉也不要让它悄悄少算一半。
    """
    cur, grand, hits = None, None, {}
    for r in range(11, sh.nrows):
        a = _norm(sh.cell_value(r, 0))
        if a == 'Sum':                      # A 列的 Sum = 全表总计，正文到此为止
            grand = r
            break
        if a:
            cur = a
        level = next((c for c in (1, 2, 3) if _norm(sh.cell_value(r, c)) == 'Sum'), None)
        if level is None or cur is None:
            continue
        hits.setdefault(cur, []).append((level, r))
    if grand is None:
        raise Db1FetchError('Eurex 工作簿找不到 A 列的全表 Sum 行')

    out = {}
    for name, lst in hits.items():
        top = min(lv for lv, _r in lst)
        rows = [r for lv, r in lst if lv == top]
        if len(rows) != 1:
            raise Db1FetchError('Eurex 的组「%s」有 %d 个同层小计行 %s，层级结构变了'
                                % (name, len(rows), rows))
        out[name] = rows[0]
    if not out:
        raise Db1FetchError('Eurex 工作簿一个组小计行都没定位到')
    return out, grand


def _parse_eurex(path, month):
    """解析一期 monthlystat_YYYYMM.xls，返回 ({列: 值}, Cover 自述生成日 'YYYY-MM-DD')。"""
    xlrd = _xlrd()
    wb = xlrd.open_workbook(path)
    if _EUREX_SHEET not in wb.sheet_names():
        # 2003-01~2007-12 是单 sheet 'Eurex Statistics' 旧版式，结构完全不同。
        # 宁可这一期不入库，也不要拿当前解析器去猜一张长得不一样的表。
        raise Db1FetchError('%s 是旧版式（sheets=%r），本模块只解析 2008-01 起的当前版式'
                            % (os.path.basename(path), wb.sheet_names()))
    sh = wb.sheet_by_name(_EUREX_SHEET)

    hdr = _norm(sh.cell_value(10, _C_OI))
    if hdr != 'Open Interest':
        raise Db1FetchError('Eurex 第 %d 列表头是 %r 而不是 Open Interest，列位置变了'
                            % (_C_OI, hdr))

    td = _num(sh.cell_value(7, _C_TRADED))
    if not td or td <= 0:
        raise Db1FetchError('%s 取不到交易日数' % os.path.basename(path))

    sums, grand = _eurex_group_sums(sh)
    rec = {'trading_days_eurex': td,
           'adv_eurex_total_contracts': _num(sh.cell_value(grand, _C_ADV)),
           'oi_eurex_total_contracts': _num(sh.cell_value(grand, _C_OI))}
    if rec['adv_eurex_total_contracts'] is None:
        raise Db1FetchError('%s 的全表 Sum 行没有日均成交数' % os.path.basename(path))

    # 闭合检验：组小计之和必须等于全表 Sum，逐位（不给容差）。
    # 不等 = 组定位漏了一组或把一组算了两遍，两种都会让分组列静默偏掉。
    total_traded = sum(_num(sh.cell_value(r, _C_TRADED)) or 0.0 for r in sums.values())
    grand_traded = _num(sh.cell_value(grand, _C_TRADED))
    if grand_traded is None or abs(total_traded - grand_traded) > 0.5:
        raise Db1FetchError(
            '%s 闭合检验不过：%d 个组小计合计 %.0f，全表 Sum 行 %s，差 %s'
            % (os.path.basename(path), len(sums), total_traded, grand_traded,
               None if grand_traded is None else '%.0f' % (total_traded - grand_traded)))

    for key, aliases in _EUREX_GROUPS:
        row = next((sums[a] for a in aliases if a in sums), None)
        if row is None:
            continue                        # 该组当月不存在（如 2008-01~05 没有股息组）
        rec['adv_eurex_%s_contracts' % key] = _num(sh.cell_value(row, _C_ADV))
        rec['oi_eurex_%s_contracts' % key] = _num(sh.cell_value(row, _C_OI))

    # 产品级：E 列（0-indexed 4）是产品代码，全表唯一
    wanted = {code: key for key, code in _EUREX_PRODUCTS}
    for r in range(11, grand):
        code = _norm(sh.cell_value(r, 4))
        key = wanted.pop(code, None)
        if key is None:
            continue
        rec['adv_%s_contracts' % key] = _num(sh.cell_value(r, _C_ADV))
        rec['oi_%s_contracts' % key] = _num(sh.cell_value(r, _C_OI))

    _require(rec, 'eurex', month, os.path.basename(path))
    return rec, _eurex_created_on(wb)


def _eurex_created_on(wb):
    """Cover 的「Created on:」是 DD.MM.YYYY 字符串（FWB 那边反而是 Excel 序列号）。"""
    cv = wb.sheet_by_name('Cover') if 'Cover' in wb.sheet_names() else None
    if cv is None:
        return None
    for r in range(cv.nrows):
        for c in range(cv.ncols - 1):
            if _norm(cv.cell_value(r, c)) == 'Created on:':
                raw = _norm(cv.cell_value(r, c + 1))
                try:
                    return datetime.datetime.strptime(raw, '%d.%m.%Y').strftime('%Y-%m-%d')
                except ValueError:
                    return None             # 认得出这一格、读不懂日期 → 宁缺勿猜
    return None


# ── 解析：FWB 现货（快腿 B）─────────────────────────────────────────────────
def _fwb_row_index(sh):
    """A 列文本 -> [行号]。行数每月加一行（口径坑 9），一律按文本找行，不按行号。"""
    idx = {}
    for r in range(sh.nrows):
        a = _norm(sh.cell_value(r, 0))
        if a:
            idx.setdefault(a, []).append(r)
    return idx


def _parse_fwb(path, month):
    """解析一期 FWB 现货工作簿。

    返回 ({'YYYY-MM': {列: 值}}, Cover 自述生成日)。
    除报告月自己的 5×2 分资产类别块之外，还顺手把「本年度逐月」块里的场所级总额
    一并带出来 —— 那是白捡的一年历史（口径坑 9），跨文件重叠月实测逐位一致。
    """
    xlrd = _xlrd()
    wb = xlrd.open_workbook(path)
    if _FWB_SHEET not in wb.sheet_names():
        raise Db1FetchError('%s 里没有「%s」sheet' % (os.path.basename(path), _FWB_SHEET))
    sh = wb.sheet_by_name(_FWB_SHEET)
    idx = _fwb_row_index(sh)

    # 顶部分类块：Total 行之前的那一批标签行
    total_rows = idx.get('Total', [])
    if not total_rows:
        raise Db1FetchError('%s 的 Total View 里找不到 Total 行' % os.path.basename(path))
    total_row = total_rows[0]

    venues = {'xetra': 1, 'fwb': 2}         # 1 = Xetra(XETR)，2 = Frankfurt(XFRA)
    rec, parsed = {}, {v: {} for v in venues}
    for key, label in _FWB_GROUPS:
        rows = [r for r in idx.get(label, []) if r < total_row]
        if not rows:
            raise Db1FetchError('%s 的 Total View 缺「%s」行，表结构可能变了'
                                % (os.path.basename(path), label))
        r = rows[0]
        for vname, col in venues.items():
            v = _num(sh.cell_value(r, col))
            parsed[vname][key] = v
            if (vname, key) in _FWB_STORE:
                rec['turnover_%s_%s_eurbn' % (vname, key)] = None if v is None else v / 1e9

    # 闭合检验：两个场所各自的 Total 必须等于 5 个 instrument group 之和
    for vname, col in venues.items():
        tot = _num(sh.cell_value(total_row, col))
        s = sum(v for v in parsed[vname].values() if v is not None)
        if tot is None or abs(s - tot) > 1.0:
            raise Db1FetchError('%s 的 %s 分类合计 %r 与 Total %r 对不上'
                                % (os.path.basename(path), vname, s, tot))
        rec['turnover_%s_eurbn' % vname] = tot / 1e9

    out = {month: rec}

    # 「本年度逐月」块：A 列是 'January 2026' 这种，只给场所级总额，没有分类
    for label, rows in idx.items():
        try:
            d = datetime.datetime.strptime(label, '%B %Y')
        except ValueError:
            continue
        mon = '%04d-%02d' % (d.year, d.month)
        r = rows[0]
        vals = {'turnover_xetra_eurbn': _num(sh.cell_value(r, 1)),
                'turnover_fwb_eurbn': _num(sh.cell_value(r, 2))}
        if all(v is None for v in vals.values()):
            continue
        vals = {k: (None if v is None else v / 1e9) for k, v in vals.items()}
        if mon == month:
            # 同一份文件里的两处数必须自洽，不自洽说明行定位错了
            for k, v in vals.items():
                if v is not None and abs(v - rec[k]) > 1e-6:
                    raise Db1FetchError('%s 的 %s 月度块 %r 与顶部 Total %r 不一致'
                                        % (os.path.basename(path), k, v, rec[k]))
        else:
            out.setdefault(mon, {}).update(vals)

    venue_totals = ['turnover_xetra_eurbn', 'turnover_fwb_eurbn']
    for mon, r in out.items():
        _require(r, 'fwb', mon, os.path.basename(path),
                 only=None if mon == month else venue_totals)
    return out, _fwb_created_on(wb)


def _fwb_created_on(wb):
    """FWB 的「Created on:」是 Excel 日期序列号（46237.0 → 2026-08-03），不是字符串。"""
    xlrd = _xlrd()
    if 'Cover' not in wb.sheet_names():
        return None
    cv = wb.sheet_by_name('Cover')
    for r in range(cv.nrows):
        for c in range(cv.ncols - 1):
            if _norm(cv.cell_value(r, c)) == 'Created on:':
                raw = cv.cell_value(r, c + 1)
                try:
                    return xlrd.xldate.xldate_as_datetime(
                        float(raw), wb.datemode).strftime('%Y-%m-%d')
                except Exception:                     # noqa: BLE001
                    return None
    return None


# ── 校验 ────────────────────────────────────────────────────────────────────
def _require(rec, leg, month, where, only=None):
    """该腿在这个月份应该有的列，缺任何一个就炸。

    「应该有」的判据是实测出来的 first_month（见列口径表），不是拍脑袋：
    比 first_month 早的月份天然为空（产品还没上市 / 官方还没开始披露），
    不早于 first_month 却为空，就说明解析错行了 —— 那种错会静默产出一整列空白。

    only: 限定只查这几列。FWB「本年度逐月」块只给场所级总额、没有分资产类别，
    对那些月份按整条腿去查会把「这块表本来就不含分类」误判成解析失败。
    """
    pool = only if only is not None else _COLS_BY_LEG[leg]
    bad = []
    for col in pool:
        first = COLUMN_META[col]['first_month']
        if first is None or month < first:
            continue
        if rec.get(col) is None:
            bad.append(col)
    if bad:
        raise Db1FetchError('%s（%s）缺列 %s —— 解析异常，拒绝写入' % (where, month, bad))


# ── 冲突台账 ────────────────────────────────────────────────────────────────
def _record_conflicts(cache_dir, rows):
    """已有值 vs 本次解析值不一致 → 写 cache/db1_restatements.csv，**不覆盖 CSV**。

    三条腿都会重述历史（口径坑 3），而重述有两种可能：官方修数据，或者我们解析错了。
    机器分不清，所以这里只留证据不做决定。

    ⚠ 这张表落在 cache/（gitignore），因为它是过程物不是真值。各条腿的复查密度不同，
    看这张表时要知道它**没有覆盖什么**：
      · 慢腿（IR 台账）：每次跑都全表重解析，2002-01 起每个月每次都复查；
      · 快腿最新一期：每次跑都重下重解析（官方重述最常砸在刚发过的一两个月上）；
      · 快腿的更早月份：只在首次回补时解析过一次，之后不再复查 —— 那段历史的重述
        本模块**发现不了**，要查只能手工重跑。
    """
    if not rows:
        return
    path = _cache(cache_dir, RESTATEMENT_LOG)
    old = []
    if os.path.exists(path):
        with open(path, newline='', encoding='utf-8') as f:
            old = [r for r in csv.reader(f)][1:]
    seen = {tuple(r[:3]) for r in old}
    fresh = [r for r in rows if tuple(str(x) for x in r[:3]) not in seen]
    if not fresh:
        return
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(['month', 'column', 'in_csv', 'from_source', 'source_file', 'seen_on'])
        for r in sorted(old + [[str(x) for x in r] for r in fresh]):
            w.writerow(r)


# ── 发布日 ──────────────────────────────────────────────────────────────────
def _source_dates():
    """按路径加载仓库根的 source_dates.py。

    不能裸 import：本模块被 monthly_run 用 spec_from_file_location 加载，
    那时 sys.path 上既没有 fetch/ 也没有仓库根。
    """
    import importlib.util                             # noqa: PLC0415
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(root, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _record_source_date(series_dir, month, parts):
    """给「本次运行确立的那一个月」记发布日。

    parts: [(日期 'YYYY-MM-DD', 出处文字)]，两条快腿各一条。取**较晚**的那个 ——
    两条腿都到齐，这个月的成交列才写全，早的那个不足以代表「这个月可以看了」。

    三道闸门，缺一不可：
      · 只记这一个月，绝不给回补月份补记 —— 老文件的 Created on 是重述件的生成日，
        实测最坏比首发晚 432 天（Eurex 2024-02 那期写着 2025-05-06）。
      · 已有记录不覆盖，理由同上。
      · 超过 MAX_PUBLISH_LAG_DAYS 天的直接放弃：快腿实测最晚第 5 天，
        真出现一个第 40 天的「发布日」，那一定是拿到了重述件，宁缺勿猜。
    """
    got = [p for p in parts if p and p[0]]
    if not got:
        return
    day, evidence = max(got, key=lambda p: p[0])
    if len(got) > 1:
        # 证据栏要把两条腿都写出来：将来有人怀疑这个日期时，他需要知道
        # 「取的是较晚的那一条」这件事，否则会以为我们漏掉了另一条腿。
        evidence = '%s；同月另一条腿 %s（取两者较晚者）' % (
            evidence, '、'.join(e for d, e in sorted(got) if (d, e) != (day, evidence)))
    lag = (datetime.date(*map(int, day.split('-'))) - _month_end(month)).days
    if lag > MAX_PUBLISH_LAG_DAYS:
        return
    sd = _source_dates()
    if sd.lookup(series_dir, 'db1', month):
        return
    sd.record(series_dir, 'db1', month, day, evidence)


# ── CSV 读写 ────────────────────────────────────────────────────────────────
def _read_csv(path):
    if not os.path.exists(path):
        return list(COLUMNS), []
    with open(path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    if not rows:
        return list(COLUMNS), []
    header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
    unknown = [c for c in COLUMNS if c not in header]
    if unknown:
        raise Db1FetchError('series/db1.csv 里没有这些列：%s' % unknown)
    return header, body


def _fmt(v):
    """整数写成整数、小数用最短往返表示（repr(float)），保证重跑字节级不变。

    合约张数、交易日数天生是整数，写成 236572727.0 只会让 CSV 难读；
    而 ADV / 成交额必须保留完整精度 —— 截成 6 位小数会在跨源核对时制造假差异。
    """
    if v is None:
        return ''
    v = float(v)
    if v.is_integer() and abs(v) < 1e15:
        return str(int(v))
    return repr(v)


# ── 对外接口 ────────────────────────────────────────────────────────────────
def latest_month(cache_dir):
    """官方源当前最新月 'YYYY-MM'，**只看两条快腿**。

    取 min(Eurex 最新月, FWB 最新月)：两条腿都到齐，这个月的成交列才是完整的。
    刻意不看 IR 台账 —— 它慢一周，用它定最新月会让整页白等 6 天，还会把
    build/exchanges.py 的「共同最新月」连同 CME / Cboe / HKEX 一起往回拖一个月。

    抓不到 / 解析不出来一律抛 Db1FetchError，不返回 None 掩盖故障。
    """
    eu = _listing(EUREX_SEARCH, EUREX_HOST, _EUREX_HIT, cache_dir, 'eurex', all_pages=False)
    fwb = _listing(FWB_SEARCH, FWB_HOST, _FWB_HIT, cache_dir, 'fwb', all_pages=False)
    if not eu or not fwb:
        raise Db1FetchError('列表页没解析出任何一期：eurex=%d fwb=%d' % (len(eu), len(fwb)))
    return min(max(eu), max(fwb))


def update(series_dir, cache_dir):
    """把新月份写进 series/db1.csv，返回新增月份列表（升序）。

    三段式，顺序不能换 —— 快腿在前、慢腿在后，因为「只填空不覆盖」下先写的赢，
    而 trading_days 这类两条腿都有的东西必须由各自的腿写各自的列（口径坑 6）。
      a) Eurex 最新一期（每次重下）＋ CSV 里还空着的历史月（走缓存）
         → adv_/oi_ 各列与 trading_days_eurex
      b) FWB 现货，同上 → turnover_ 各列（含「本年度逐月」块白捡的场所级历史）
      c) IR 台账全表 → 只填空地回补所有月份的集团列（Clearstream / OTC / 360T / 商品 / IMS）

    首次运行是一次性 bootstrap：要把 2008-01 起 223 期 Eurex 与 20 期 FWB 全下一遍
    （约 420 MB、本机实测解析 38 秒）。之后每个月只下 1 期 Eurex + 1 期 FWB + 1 份台账。

    幂等保证：
      · 已存在的月份不重复追加；
      · 已经有值的单元格**永不覆盖**（三条腿都会重述历史，重述不由本模块自动吞）；
      · 冲突写进 cache/db1_restatements.csv 供人工判断；
      · 回补不计入返回值（它不是新月份）；
      · 什么都没变时，未被触碰的单元格是原样字符串搬运，文件字节级不变。
    """
    csv_path = os.path.join(series_dir, 'db1.csv')
    header, body = _read_csv(csv_path)
    idx = {name: i for i, name in enumerate(header)}
    have = {r[0]: r for r in body}
    conflicts, today = [], datetime.date.today().isoformat()

    def blank(mon):
        row = [''] * len(header)
        row[idx['month']] = mon
        have[mon] = row
        body.append(row)
        return row

    def put(mon, rec, src):
        """只填空、不覆盖；冲突记账。返回这一行是不是新建的。"""
        fresh = mon not in have
        row = have.get(mon) or blank(mon)
        for name, v in rec.items():
            if v is None or name not in idx:
                continue
            cur = row[idx[name]].strip()
            if not cur:
                row[idx[name]] = _fmt(v)
                continue
            old = _num(cur)
            if old is None or abs(old - v) > max(1e-9, 1e-12 * abs(old)):
                conflicts.append([mon, name, cur, _fmt(v), src, today])
        return fresh

    added = []

    # ── a) 快腿 A：Eurex ──
    eu_idx = _listing(EUREX_SEARCH, EUREX_HOST, _EUREX_HIT, cache_dir, 'eurex',
                      all_pages=False)
    if not eu_idx:
        raise Db1FetchError('Eurex 列表页一期都没解析出来')
    eu_latest = max(eu_idx)
    need_eu = [m for m in _months_between(EUREX_START, eu_latest)
               if not (have.get(m) or [''] * len(header))[idx['adv_eurex_total_contracts']].strip()]
    # 最新一期**每次都重下**（不走缓存）：一是要取它自述的发布日，二是官方重述最常
    # 落在刚发过的那一两个月上，重下才有机会把差异捞进冲突台账。更早的月份不重下 ——
    # 反正已有值不覆盖，重下 200 多个文件只是浪费带宽。
    recheck_eu = eu_latest not in need_eu
    if recheck_eu:
        need_eu.append(eu_latest)
    if [m for m in need_eu if m not in eu_idx]:
        # 首页只挂最近 20 多个月，要回补更早的必须把 13 页翻全
        eu_idx = _listing(EUREX_SEARCH, EUREX_HOST, _EUREX_HIT, cache_dir, 'eurex',
                          all_pages=True)
    eu_pub = None
    for mon in need_eu:
        if mon not in eu_idx:
            raise Db1FetchError('Eurex 列表页上没有 %s 这一期，官方归档出现空洞' % mon)
        dst = _cache(cache_dir, 'db1_eurex_%s.xls' % mon.replace('-', ''))
        path = _download(eu_idx[mon][0], dst) if mon == eu_latest \
            else _cached_download(eu_idx[mon][0], dst)
        rec, created = _parse_eurex(path, mon)
        if put(mon, rec, os.path.basename(path)):
            added.append(mon)
        if mon == eu_latest and created:
            eu_pub = (created, 'monthlystat_%s.xls 的 Cover 表「Created on:」= %s'
                      % (mon.replace('-', ''), created))

    # ── b) 快腿 B：FWB 现货 ──
    fwb_idx = _listing(FWB_SEARCH, FWB_HOST, _FWB_HIT, cache_dir, 'fwb', all_pages=False)
    if not fwb_idx:
        raise Db1FetchError('FWB 列表页一期都没解析出来')
    fwb_latest = max(fwb_idx)
    # 分资产类别列只有各期自己的报告月才有，所以用它判断「这一期下过没有」
    need_fwb = [m for m in sorted(fwb_idx)
                if not (have.get(m) or [''] * len(header))[idx['turnover_xetra_equities_eurbn']].strip()]
    if fwb_latest not in need_fwb:                    # 理由同 Eurex 那边的 recheck
        need_fwb.append(fwb_latest)
    fwb_pub = None
    for mon in need_fwb:
        url = fwb_idx[mon][0]
        stamp = re.search(r'\.(\d{8})\.xls', url).group(1)
        dst = _cache(cache_dir, 'db1_fwb_%s.xls' % stamp)
        path = _download(url, dst) if mon == fwb_latest else _cached_download(url, dst)
        recs, created = _parse_fwb(path, mon)
        for m2 in sorted(recs):
            if put(m2, recs[m2], os.path.basename(path)) and m2 not in added:
                added.append(m2)
        if mon == fwb_latest and created:
            fwb_pub = (created,
                       'FWB_Monthly_Cash_Market_Statistics.%s.xls 的 Cover 表'
                       '「Created on:」序列号换算 = %s' % (stamp, created))

    # ── c) 慢腿：IR 台账全表回补 ──
    dbg_path = _download(_dbg_url(cache_dir),
                         _cache(cache_dir, 'db1_major_business_figures.xlsx'))
    dbg, _modified = _parse_dbg(dbg_path)
    for mon in sorted(dbg):
        if mon < LEDGER_START:
            continue
        _require(dbg[mon], 'dbg', mon, os.path.basename(dbg_path))
        if put(mon, dbg[mon], os.path.basename(dbg_path)) and mon not in added:
            added.append(mon)

    # 无条件落盘：即便没有新月份，上面的慢腿回补也可能改了既有行。
    body.sort(key=lambda r: r[0])
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(body)

    _record_conflicts(cache_dir, conflicts)
    # 记发布日放在落盘之后：写盘失败就不该留下「这个月官方发过了」这条断言。
    established = min(eu_latest, fwb_latest)
    if established in have:
        _record_source_date(series_dir, established, [eu_pub, fwb_pub])
    return sorted(added)


if __name__ == '__main__':
    _here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    print('latest:', latest_month(os.path.join(_here, 'cache')))
    print('added :', update(os.path.join(_here, 'series'),
                            os.path.join(_here, 'cache')))
