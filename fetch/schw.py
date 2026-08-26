# -*- coding: utf-8 -*-
"""Charles Schwab (SCHW) —— 月度经营指标抓取模块。

================================ 数据源 ================================
官方只有一个真源：Schwab Monthly Activity Report 随附的 Excel 附表，走 Akamai CDN 直链，
文件名完全可推导，因此不需要爬列表页、不需要登录态、不需要浏览器：

    月报附表   https://content.schwab.com/web/retail/public/about-schwab/excels/
               schw_<mon><yyyy>_table.xlsx        例：schw_may2026_table.xlsx
    月报正文   https://content.schwab.com/web/retail/public/about-schwab/
               schw_<mon><yyyy>_press_release.pdf （数值不从这里取，只取电头日期，见下）
    季报附表   https://content.schwab.com/web/retail/public/about-schwab/excels/
               schw_q<n>_<yyyy>_earnings_tables.xlsx  例：schw_q2_2026_earnings_tables.xlsx
    季报正文   https://content.schwab.com/web/retail/public/about-schwab/
               schwab_q<n>_<yyyy>_earnings_release.pdf（同上，只取电头日期）

<mon> 是小写三字母英文月份缩写（jan…dec），<yyyy> 四位年。
注意两份 PDF 的**文件名前缀不一样**：月报是 schw_、季报是 schwab_（xlsx 两个都是 schw_）。
照着 xlsx 的写法推季报 PDF 的名字必 404，这不是笔误，是官方自己就不一致。

—— 为什么不用别的源 ——
· www.aboutschwab.com（IR 站正文页）在 Akamai 后面，会按 TLS/HTTP2 指纹拦截：
  python urllib 无论带什么 UA/Sec-Fetch 头都是 403，只有真浏览器栈（curl --http2 带全套
  浏览器头也行）能过。所以**不能**把落地页解析放进无人值守主路径。
  content.schwab.com 这个 CDN 域宽松得多：urllib + 普通 Chrome UA 即可 200。
· SEC EDGAR **是可选源**（2026-08-16 更正了两遍，把过程写下来，免得第三次又绕回去）。
  实测 CIK 0000316709 的季度 8-K：**每一期**的 EX-99.1 正文里都嵌着「The Charles Schwab
  Corporation Monthly Activity Report For <月> <年>」那张 13 个月滚动表 —— 2014 年的有，
  2026 年的也有。相邻两期锚点差 3 个月、窗口重叠 10 个月，可逐期接续。
    - 原始 docstring 写「想靠 EDGAR 拿月度数据是走不通的」→ **错**。
    - 第一次更正写「2017-04 起表就不随 8-K 附了，只剩一句脚注引用」→ **也错**，
      而且错得更隐蔽：那是**解析器**的毛病不是官方的改动。两代版式放标题的位置不同 ——
      2017-03 及更早，标题在表的第一行**里**；之后，标题在表**上方的 <div> 里**。
      而正文另有一句「…please see the Monthly Activity Report.」的脚注，位置比真表靠前。
      当时的解析器锚在「monthly activity report」这个词上再往回找 <table>，于是
      新版式必然捞到上面某张不相干的表，**静默**返回空 —— 97 份里只读出 15 份，
      看上去就像「2017-04 之后官方不附表了」。
      教训：解析器读不到，和源里没有，长得一模一样；分辨这两者只能去看原文。
  现在的做法见 parse_edgar_monthly()：锚在**标题正则**上，再按标题落在表内还是表外
  决定往前还是往后找 <table>。实测 871 个与 series/schw.csv 重叠的 (月,指标) 全部相等，
  含 2019-04 那个 core NNA = -0.3（会计负号的右括号被拆进独立 <td>，见 _glue）。
  月度活动报告**本身**确实从不单独发 8-K，原文这半句没错。
  取 EDGAR 要在 User-Agent 里带联系方式，否则 403（见 _EDGAR_UA），这是它明文的使用条款。
· 新闻稿 PDF 里的数字和 xlsx 完全一致，但 PDF 要 OCR/文本抽取，没必要，xlsx 是结构化的。

================================ 发布节奏 ================================
· 非季末月：次月**第 2 周的周一到周五之间**（历史上多在 12–14 日）发月报，附表当天上线。
· 季末月（3/6/9/12）：**没有独立月报**，对应 URL 直接 404（已实测 mar2026 / jun2026 /
  dec2025 全 404）。季末月的数值有两条路可拿，本模块两条都走：
    (a) 当季**季报附表**的 "ER SMART" 页 —— 季报次月中下旬发（Q2-2026 是 7/21），最快；
    (b) 下一个月报附表的 13 个月滚动表 —— 例如 jul2026 月报的表里会带上 jun2026，
        但要等到 8 月中，比 (a) 慢一个月。
  所以「季末月漏掉」这个坑的正确解法不是特判某几个月，而是：**两个源都抓，取并集**。
· 实测发布日（月末后第几天，8 期）：月报 12 / 13 / 13 / 14，季报 16 / 16 / 21 / 21
  —— 常规月最坏 14、季末月最坏 21，就是 _LAG_DAYS 与 build/roster.py 的 LAG['schw']。
  样本来自 series/source_dates.csv 与 cache 里几份新闻稿的电头，随时可复核。
  注意实测最坏值**正好顶到**这两个数，所以任何拿它当红线的判断都必须另加余量，
  见 _crosscheck_due_month 与 _OVERDUE_MARGIN_DAYS。

========================== 官方发布日（source_dates）==========================
页面抬头「官方发布于 X」要的是**源头自己说出来的那一天**，所以只认新闻稿正文的电头：

    月报 PDF 第 1 页  "WESTLAKE, Texas, June 12, 2026 – The Charles Schwab Corporation
                      released its Monthly Activity Report today."   → 2026-05 发布于 6/12
    季报 PDF 第 1 页  "WESTLAKE, Texas, July 21, 2026 – The Charles Schwab Corporation
                      reported net income for the second quarter…"   → 2026-06 发布于 7/21
                      （季末月的数值来自季报，所以它的发布日就是季报的发布日）

**不能用 HTTP Last-Modified**：实测 q2_2026 那份 PDF 的 Last-Modified 是 7/20 17:11 GMT，
而电头写的是 7/21 —— 官方前一天晚上先把文件挂上 CDN、次日盘前才对外发布。
用 Last-Modified 会系统性地早一天，而早一天的日期看上去完全正常，没人会发现。
mtime、下载日、构建日同理，一律不用。

只有当某个月是从**它自己那期**报告里读到时才记发布日。同一个月也会出现在后续几期的
13 个月滚动表里，那些文件的电头是后来那期的日子，不是这个月首次公布的日子；
分不清就宁可让这半句话缺席（见仓库根 source_dates.py 的 docstring）。

================================ 口径坑 ================================
1. **单位在两种文件里不一样**，这是最容易静默算错的地方：
   · 月报附表：客户资产/融资余额已经是 $bn（13135.3），账户数/DATs 已经是千（461 / 11813）。
   · 季报附表：同样的行是**原始单位**（13135300000000 美元 / 490000 个账户 / 13615000 笔）。
   所以不能写死除数。本模块用「基准行反推倍率」：用 Total Client Assets 定金额倍率、用
   Active Brokerage Accounts 定计数倍率，再把倍率套到同一单位块里的其它行，并对结果做
   量级断言。硬编码 /1e9 迟早会在某次版式微调后炸掉。
2. **行标签会变**，不能按行号取数：
   · 2026-02 那期把 "Net Market Gains (Losses)" 写成 "Net Market (Losses) Gains"（词序颠倒）。
   · sheet 名月报是 'SMART'、2019 年的老文件是 'Smart'、季报是 'ER SMART'。
   所以一律按标签前缀 + 大小写无关匹配。
3. **四列是 2026-01 那期才新增的**：Client Daily Average Trades (DATs)、Margin Balances
   at month end、Transactional Sweep Cash (at month end)、Total Money Market Funds
   (at month end)。同一期把老的 "Average Margin Balances"（月均口径，单位 $mn）删掉了。
   series/schw_avg_margin.csv 就是那条已停更的老序列，它**永远停在 2025-12**，
   本模块绝不去追加它 —— 追加只会把两种口径（月均 vs 月末）混成一条假序列。
   新增的四列在 2026-01 期的 13 个月滚动表里回填到了 2025-01，所以 series 里这四列从
   2025-01 起才有值，这是数据本身的边界，不是解析漏了。
7. **「Schwab 不披露客户现金」是句流传过的假话，别再写进任何图注。** 月报的
   "Selected Balances" 块里逐月印着 Transactional Sweep Cash 与 Total Money Market
   Funds 两条月末 $bn（脚注 (7) 给了 sweep 的定义），而 "Client Activity" 块下面还有
   一行 **Client Cash as a Percentage of Client Assets**（脚注 (8)：Schwab One、若干
   现金等价物、银行存款、第三方银行存款账户与货基余额占客户总资产的比重）——
   后者自 2014-06 起每期都印。三者互相咬得住：(sweep + MMF) / Total Client Assets
   逐月复现官方印的那个占比，19 个重叠月最大偏差 0.05pp（就是 0.1pp 印刷精度）。
   本模块以前抓不到它们，只是因为 COLS 里没写这三行，不是公司没披露。
4. **core NNA ≠ NNA**。序列取的是 Core Net New Assets（剔除单笔巨额流入/流出 + 表外
   Schwab Bank Retail CD 流量）。2025 年起「巨额」的门槛从 $10bn 提到 $25bn，
   所以 2025 年前后的 core NNA 严格说不完全可比 —— 这是官方口径变更，不是数据错。
5. 月报附表**不重述历史**：apr2026 与 may2026 两期文件里 12 个重叠月份的数值逐个相同。
   但季报附表口径上是「最终版」，万一和月报打架，本模块以季报为准（见 _SOURCE_RANK）。
6. 2020-10 的 new_brokerage_accounts_k = 14718 是 TD Ameritrade 并表的一次性搬账，
   不是当月开户量。build_schw.py 已经单独处理，这里原样入库、不做清洗。

===================== series/schw_backfill.csv：留着，但不进图 =====================
那个文件是一次做了一半、从未接上的回填：表头承诺 dats_k 与 margin_balances_usdbn 两列，
dats_k 整列是空的，实际只有 2018-12…2024-12 共 7 个年末的月末融资余额。
2026-08-05 专门评估过要不要把它接进 build/schw.py，**结论是不接**，理由按分量排列：

(a) **来历不可复现，而且能证明它不可能来自本管道的源。** cache/schw_may2019_table.xlsx
    （2019 年的月报附表）里逐行查过：整张 Smart 页**没有任何融资余额行**，月均的没有、
    月末的也没有 —— Schwab 是 2020-04 才开始披露月均融资余额（见 schw_avg_margin.csv
    的起点）、2026-01 才开始披露月末融资余额（见上面第 3 条）。所以这 7 个数只可能来自
    10-K / 年报或某个外部源，而文件里没记来源、没记表名、也没有任何 fetch 代码产生它。
    本页在「口径与方法说明」里对读者写的是「无任何估算或补插」，一个连出处都指不出来的
    序列不满足这句话；§5.5「失败要响、绝不静默上线」管的也是同一件事。
(b) **边际信息只有 2 个点。** 2020-12 起的 5 个年末，schw_avg_margin.csv 已经**逐月**
    覆盖（2020-04 → 2025-12），两者逐年只差 -1.7 / -0.9 / +0.3 / +1.8 / +2.3 bn，
    符号正负都有 —— 这正是「月末 vs 月均」的基差噪音，不是新形状。真正新增的只有
    2018-12 (19.3) 与 2019-12 (19.5) 两个点。
(c) **这 2 个新点还落在 TD Ameritrade 并购的另一侧。** 并购 2020-10 完成，
    19.5 → 60.9 是资产负债表搬账不是融资需求，和第 6 条的 14,718k 是同一类假象。
    一张只有 7 个点的图，最抢眼的特征会是这个 3 倍台阶，读者读到的是并购不是杠杆周期。
(d) **年末单点混不进月度轴。** 7 个点之间是 6 段 12 个月的空档，直接塞进 Exhibit 9/12
    会造出 66 个空月并违反 CONTRACT §5.3（不可比的相邻期不得画成连续序列）。
    技术上可以另开一张明确标「year-end snapshot」的图（不与月度点连线），但那张图要用
    (a) 的无源数据、换来 (b) 的 2 个点、再给一个已经背着两条口径警告（core NNA 门槛断点、
    月末 vs 月均不可接续）的页面加上第三条 —— 不划算。

**下次做孤儿文件盘点的人：不要删它。** 它不是残留垃圾，是全仓独一份的 2018–2024 年末
月末融资余额，管道抓不回来（见 (a)），删了就永久丢失。它的正确用途是**离线核对**：
比如想验证某年年末的月末余额、或者判断月均序列与月末序列的基差量级时，手工翻它。
要让它有资格进图，得先补齐两件事 —— 在文件里写清每个数出自哪份 10-K/年报的哪张表，
并在 fetch 侧写出可复现的抓取路径；在那之前它只能停在 series/ 里当参考料。

========================== 历史回填（backfill）==========================
主流程 update() 只探最近 8 个月 / 3 个季度 —— 那是「每月往后走一格」该有的开销。
把**存量**历史一次性补齐是另一件事，走 `python3 fetch/schw.py --backfill`：
它扫 _HIST_XLSX 里那份写死的老附表清单（文件名有两处不成规律的历史变体：前缀
`schwab_`、扩展名大写 `.XLSX`，推不出来所以写死）+ 上面说的 EDGAR 那条路，
把 2013-09…2018-04 共 56 个月补进 series/schw.csv。

三条铁律：**已存在的月份一个字符都不改**；**重叠月必须逐值相等**（不等就整次失败，
要人去看是解析取错了还是官方重述了）；**回填后月份必须仍然逐月连号**（缺口会让
下游 assets.diff() 把两个月的变动记到一个月上）。首次跑通时 39 个重叠值 0 处不符，
第二次跑 166 个重叠值 0 处不符、新增 0 个月（幂等）。

回填**不改变各列自己的披露边界**：core_nna_usdbn 仍然从 2017-02 起（_CORE_NNA_FROM），
dats_k / margin_balances_usdbn 仍然从 2025-01 起。补出来的只是客户总资产与新开经纪账户。

================================ 落盘 ================================
所有下载文件只写 cache/（已 gitignore），文件名与官方一致，便于事后复核。
"""
from __future__ import annotations

import csv
import datetime as _dt
import json
import os
import re
import time as _time
import urllib.error
import urllib.request

import openpyxl

# ── 常量 ────────────────────────────────────────────────────────────────
CDN = 'https://content.schwab.com/web/retail/public/about-schwab'
MONTHLY_URL = CDN + '/excels/schw_{mon}{year}_table.xlsx'
QUARTER_URL = CDN + '/excels/schw_q{q}_{year}_earnings_tables.xlsx'
# 新闻稿正文：只为取电头日期。前缀 schw_ / schwab_ 不一致是官方的，别按 xlsx 的写法改。
PR_MONTHLY_URL = CDN + '/schw_{mon}{year}_press_release.pdf'
PR_QUARTER_URL = CDN + '/schwab_q{q}_{year}_earnings_release.pdf'
IR_PAGE = 'https://www.aboutschwab.com/financial-reports'

# CDN 只看 UA，给个普通 Chrome UA 就放行；不带 UA 或带 Python-urllib 会 403。
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

_MON = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
        'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
_MON_FULL = ['January', 'February', 'March', 'April', 'May', 'June',
             'July', 'August', 'September', 'October', 'November', 'December']

# 新闻稿电头。锚在 "WESTLAKE, Texas," 上而不是只找「月 日, 年」——正文里还有一大堆日期
# （"as of May 31, 2026"、脚注里的往期口径变更日），只匹配日期格式会抓到其中任意一个。
_DATELINE = re.compile(
    r'WESTLAKE,\s*Texas,\s*(' + '|'.join(_MON_FULL) + r')\s+(\d{1,2}),\s*(20\d{2})',
    re.IGNORECASE)

SERIES = 'schw.csv'
COLS = ['core_nna_usdbn', 'total_client_assets_usdbn',
        'new_brokerage_accounts_k', 'dats_k', 'margin_balances_usdbn',
        'client_cash_pct', 'sweep_cash_usdbn', 'mmf_usdbn']

# 标签前缀（小写、去空白后前缀匹配）。用前缀而不是全等，是因为官方在标签尾部挂脚注号，
# 脚注号每期都在变（"(1,2)" → "(1,2,3)"），全等匹配必挂。
_LABEL = {
    'core_nna_usdbn':           'core net new assets',
    'total_client_assets_usdbn': 'total client assets',
    'new_brokerage_accounts_k': 'new brokerage accounts',
    'dats_k':                   'client daily average trades',
    'margin_balances_usdbn':    'margin balances at month end',
    # 客户现金三行。前缀要写全 'total money market funds'：同一张表里还有一行
    # 'Money Market Funds'（净买卖流量，$mn），前缀写短了会取到那一行。
    'client_cash_pct':          'client cash as a percentage of client assets',
    'sweep_cash_usdbn':         'transactional sweep cash',
    'mmf_usdbn':                'total money market funds',
    # 下面两行只用来定单位倍率，不入库
    '_anchor_money':            'total client assets',
    '_anchor_count':            'active brokerage accounts',
}
# 金额类 / 计数类 / 比率类分组：决定用哪个锚点行的倍率
_MONEY = {'core_nna_usdbn', 'total_client_assets_usdbn', 'margin_balances_usdbn',
          'sweep_cash_usdbn', 'mmf_usdbn'}
_COUNT = {'new_brokerage_accounts_k', 'dats_k'}
# 比率类不跟锚点走 —— 它和金额/计数不在同一个单位块里：xlsx 里印的是**分数**
# （0.101），EDGAR 的 HTML 里印的是**百分点字符串**（'10.1%'）。两边都归一到百分点。
_RATIO = {'client_cash_pct'}

# 合理量级断言。core NNA 会是负数、也可能接近 0（2019-04 是 -0.3），无法用量级判，
# 所以它不做独立断言，只跟着锚点倍率走。
_SANE = {
    'total_client_assets_usdbn': (1_000.0, 100_000.0),    # $1tn – $100tn
    'margin_balances_usdbn':     (1.0, 5_000.0),          # $bn
    'sweep_cash_usdbn':          (1.0, 5_000.0),          # $bn
    'mmf_usdbn':                 (1.0, 5_000.0),          # $bn
    # 百分点。历史区间实测 8.6–13.0，给到 3–30 —— 这个下界同时兼任「分数还是百分点」
    # 的判据：0.086 这种分数落不进 3–30，倍率只有 100 那一档能过（见 _scale 的 cands）。
    'client_cash_pct':           (3.0, 30.0),
    'new_brokerage_accounts_k':  (10.0, 100_000.0),       # 千（含 2020-10 那个 14718）
    'dats_k':                    (100.0, 1_000_000.0),    # 千笔/日
    '_anchor_count':             (5_000.0, 500_000.0),    # 千个活跃账户
}

# DATs / 月末融资余额 / 月末 Transactional Sweep Cash / 月末 Total Money Market Funds
# 这四列自 2026-01 期起披露，那一期的 13 个月滚动表回填到 2025-01。更早的月份本就没有。
# （客户现金**占比**那一行不在此列 —— 它自 2014-06 起每期都印，见 _CASH_PCT_FROM。）
_DATS_MARGIN_FROM = (2025, 1)
# **Core** Net New Assets 这一行是 2018 年初才出现在滚动表里的：实测最早带它的一份是
# schwab_feb2018_table.XLSX（表窗 2017-02…2018-02），而 2017 年的几份季报附表（q1/q2/q3
# 2017）与 2016 年及更早的全部来源里，同一位置只有未剔除的 "Net New Assets"。
# 两者不是一条序列（2017-06 官方同时给过 37.7 与 22.1），所以**不拼接**：
# 2017-02 之前 core_nna_usdbn 一律留空，由 build 侧按各自序列起点画图。
_CORE_NNA_FROM = (2017, 2)
# 客户现金**占比**这一行的官方起点：2015-07-16 报送的 Jun-2015 月报（8-K EX-99.1）
# 第一次印它，那张 13 个月滚动表最左是 2014-06。再往前的三期（2015-04-15 / 2015-01-16 /
# 2014-10-15）实测整张表**没有这一行**，不是解析漏了 —— 三份 HTML 里连
# "Client Cash as a Percentage" 这个字符串都搜不到。
_CASH_PCT_FROM = (2014, 6)

# 同一个月同时来自月报和季报时谁说了算：季报是最终版。
_SOURCE_RANK = {'monthly': 0, 'quarterly': 1}


class FetchError(RuntimeError):
    """抓不到 / 解析不出 / 缺列，一律抛这个，绝不静默降级。"""


# ── HTTP ───────────────────────────────────────────────────────────────
def _get(url: str, timeout: int = 60) -> bytes | None:
    """404 返回 None（季末月月报本来就不存在，属正常）；其它错误抛异常。"""
    req = urllib.request.Request(url, headers={'User-Agent': _UA})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.read()
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None
        raise FetchError(f'HTTP {e.code} on {url}') from e
    except Exception as e:                      # 网络层问题必须炸出来，不能当没数据
        raise FetchError(f'{type(e).__name__} on {url}: {e}') from e


def _download(url: str, cache_dir: str, reuse: bool = True,
              magic: bytes = b'PK') -> str | None:
    """下载到 cache/，返回本地路径；404 返回 None。

    reuse=True 时复用已落盘的文件。老月份的附表官方从不重发（已核对 apr/may 两期 12 个
    重叠月逐值相同），复用是安全的；但**最近两个月的文件可能被官方重新上传更正**，
    所以调用方对新文件传 reuse=False，强制重取。

    magic 是文件头的前几个字节：xlsx 是 zip 所以 b'PK'，新闻稿 PDF 是 b'%PDF'。
    要判它是因为 CDN 偶尔用 200 返回一张「找不到页面」的 HTML，长度够大、内容全错。
    """
    os.makedirs(cache_dir, exist_ok=True)
    path = os.path.join(cache_dir, url.rsplit('/', 1)[-1])
    if reuse and os.path.exists(path) and os.path.getsize(path) > 10_000:
        return path
    blob = _get(url)
    if blob is None:
        return None
    if len(blob) < 10_000 or not blob.startswith(magic):
        raise FetchError(f'{url} 返回的不是 {magic.decode()} 文件（{len(blob)} bytes）')
    with open(path, 'wb') as f:
        f.write(blob)
    return path


def release_date(kind: str, report_ym: tuple[int, int], cache_dir: str):
    """某一期报告的官方发布日 → ('YYYY-MM-DD', 出处描述)；拿不到返回 (None, 原因)。

    发布日只从新闻稿正文第 1 页的电头读（见模块 docstring 的「官方发布日」一节）。
    PDF 一律 reuse：电头是印在正文里的，官方就算把文件重新上传一遍也不会改那一行，
    所以这里不像 xlsx 那样对新月份强制重取。
    """
    y, m = report_ym
    if kind == 'monthly':
        url = PR_MONTHLY_URL.format(mon=_MON[m - 1], year=y)
    else:
        url = PR_QUARTER_URL.format(q=(m - 1) // 3 + 1, year=y)
    name = url.rsplit('/', 1)[-1]

    path = _download(url, cache_dir, magic=b'%PDF')
    if path is None:
        return None, f'{name} 在 CDN 上是 404'
    try:
        import fitz                       # 延迟 import：同 openpyxl 的处理，缺它只该让
    except ImportError:                   # 发布日缺席，不该把整家的数据摄入一起拖挂
        return None, 'pymupdf(fitz) 未安装，读不了新闻稿 PDF'
    try:
        # 先把换行/不间断空格压成单空格：PDF 抽出来的电头经常被折行成两段，
        # 不压的话正则里的 \s+ 也救不了「Texas,\nJuly」中间夹的软连字符之类。
        txt = re.sub(r'\s+', ' ', fitz.open(path)[0].get_text().replace('\xa0', ' '))
    except Exception as e:
        return None, f'{name} 第 1 页读不出文本：{type(e).__name__}'
    hit = _DATELINE.search(txt)
    if not hit:
        return None, f'{name} 第 1 页找不到 "WESTLAKE, Texas, <月> <日>, <年>" 电头'
    mon = _MON_FULL.index(hit.group(1).capitalize()) + 1
    return (f'{int(hit.group(3)):04d}-{mon:02d}-{int(hit.group(2)):02d}',
            f'{name} 第 1 页电头 "{hit.group(0)}"')


def _ir_page_links(cache_dir: str) -> list[str]:
    """兜底诊断用：万一 CDN 命名规则变了，从 IR 落地页把真实链接捞出来。

    www.aboutschwab.com 拦 urllib（Akamai 按 TLS 指纹拦），只能借 curl 的 HTTP/2 栈，
    所以这条路**不进主流程**，只在主流程全 404 时被 update() 调用来生成有用的报错。
    """
    import subprocess
    hdr = [
        '-H', 'User-Agent: ' + _UA,
        '-H', 'Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        '-H', 'Accept-Language: en-US,en;q=0.9',
        '-H', 'Sec-Fetch-Dest: document', '-H', 'Sec-Fetch-Mode: navigate',
        '-H', 'Sec-Fetch-Site: none', '-H', 'Sec-Fetch-User: ?1',
        '-H', 'Upgrade-Insecure-Requests: 1',
    ]
    try:
        out = subprocess.run(['curl', '-sSL', '--compressed', '--http2', *hdr, IR_PAGE],
                             capture_output=True, timeout=90).stdout.decode('utf-8', 'replace')
    except Exception:
        return []
    with open(os.path.join(cache_dir, '_schw_ir_financial_reports.html'), 'w') as f:
        f.write(out)
    # 扩展名大小写都收：官方历史上真发过 .XLSX（见 _HIST_XLSX），而这个函数是
    # _crosscheck_due_month 的判官 —— 判官只认小写，官方哪天改回大写它就当作
    # 「落地页上没有更新的表」，护栏又变回静默。
    return sorted(set(re.findall(r'https://content\.schwab\.com/[^"\']+\.[xX][lL][sS][xX]', out)))


# ── 解析 ───────────────────────────────────────────────────────────────
def _sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower().endswith('smart'):     # 'SMART' / 'Smart' / 'ER SMART'
            return wb[name]
    raise FetchError(f'找不到 SMART 页，sheet 有：{wb.sheetnames}')


def _month_columns(ws, report_ym: tuple[int, int]) -> dict[tuple[int, int], int]:
    """定位 13 个月滚动表的列。

    不信任表头上方那行稀疏的年份标签（它只在每年第一列出现，一次版式微调就会错位），
    改成：找到月份缩写那一行，认定**最右一个数据列就是报告月**，然后按月倒推。
    倒推完再和表里写的月份缩写逐个核对，对不上直接抛 —— 这样版式变了会立刻炸，
    而不是悄悄把数据错位一格。
    """
    abbr = {m: i + 1 for i, m in enumerate(_MON)}
    hdr_row = hdr_cols = None
    for row in ws.iter_rows(min_row=1, max_row=15):
        cols = [(c.column, str(c.value).strip().lower()[:3]) for c in row
                if isinstance(c.value, str) and str(c.value).strip().lower()[:3] in abbr]
        if len(cols) >= 6:
            hdr_row, hdr_cols = row[0].row, cols
            break
    if hdr_cols is None:
        raise FetchError('找不到月份表头行')

    y, m = report_ym
    out: dict[tuple[int, int], int] = {}
    for k, (col, tag) in enumerate(reversed(hdr_cols)):     # 从最右（=报告月）往左
        yy, mm = y, m - k
        while mm <= 0:
            mm += 12
            yy -= 1
        if abbr[tag] != mm:
            raise FetchError(
                f'表头月份与报告月对不上：第 {hdr_row} 行第 {col} 列写的是 {tag}，'
                f'按报告月 {y}-{m:02d} 倒推应为 {mm:02d}。版式可能变了，人工核对后再跑。')
        out[(yy, mm)] = col
    return out


def _row_values(ws, prefix: str, cols: dict) -> dict | None:
    """按标签前缀找行，返回 {(y,m): 原始数值}；找不到该行返回 None。"""
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        lab = row[0].value
        if not isinstance(lab, str):
            continue
        lab = re.sub(r'\s+', ' ', lab).strip().lower()
        if not lab.startswith(prefix):
            continue
        vals = {}
        for ym, col in cols.items():
            v = ws.cell(row=row[0].row, column=col).value
            if isinstance(v, (int, float)):
                vals[ym] = float(v)
        if vals:
            return vals
    return None


def _scale(anchor: dict, lo: float, hi: float, what: str, cands=None) -> float:
    """用锚点行反推单位倍率。金额/计数的候选是 1 / 1e-3 / 1e-6 / 1e-9 四档，
    比率行传 cands=(1.0, 100.0)（分数 vs 百分点）。
    要求**所有**锚点值套上倍率后都落进合理区间，且只有一档满足 —— 有歧义就抛。"""
    ok = [s for s in (cands or (1.0, 1e-3, 1e-6, 1e-9))
          if all(lo <= abs(v) * s <= hi for v in anchor.values())]
    if len(ok) != 1:
        raise FetchError(f'{what} 单位倍率判不定（候选 {ok}，样值 {list(anchor.values())[:3]}）')
    return ok[0]


def parse_table(path: str, report_ym: tuple[int, int]) -> dict:
    """把一个 xlsx 解析成 {(year, month): {列名: 值}}。缺关键行直接抛。"""
    ws = _sheet(openpyxl.load_workbook(path, data_only=True, read_only=False))
    cols = _month_columns(ws, report_ym)

    a_money = _row_values(ws, _LABEL['_anchor_money'], cols)
    a_count = _row_values(ws, _LABEL['_anchor_count'], cols)
    if not a_money or not a_count:
        raise FetchError(f'{os.path.basename(path)}: 找不到单位锚点行')
    s_money = _scale(a_money, *_SANE['total_client_assets_usdbn'], 'money')
    s_count = _scale(a_count, *_SANE['_anchor_count'], 'count')

    raw = {c: _row_values(ws, _LABEL[c], cols) for c in COLS}
    out: dict[tuple[int, int], dict] = {ym: {} for ym in cols}
    for c in COLS:
        if raw[c] is None:
            continue
        if c in _RATIO:
            s = _scale(raw[c], *_SANE[c], f'{c}', cands=(1.0, 100.0))
        else:
            s = s_money if c in _MONEY else s_count
        for ym, v in raw[c].items():
            x = v * s
            lo_hi = _SANE.get(c)
            if lo_hi and not (lo_hi[0] <= abs(x) <= lo_hi[1]):
                raise FetchError(f'{os.path.basename(path)} {ym} {c}={x} 超出合理区间 {lo_hi}')
            out[ym][c] = round(x, 6)
    return out


def _required(ym: tuple[int, int]) -> list[str]:
    """该月**必须**解析出来的列。缺任何一列 → 抛，绝不写 NaN。

    三段披露边界，全部是官方自己的（不是抓取能力的边界）：
      · 2025-01 起才有 dats_k / margin_balances_usdbn / sweep_cash_usdbn / mmf_usdbn
        （2026-01 那期月报一次新增这四列并回填到 2025-01）
      · 2017-02 起才有 core_nna_usdbn（见 _CORE_NNA_FROM）
      · 2014-06 起才有 client_cash_pct（见 _CASH_PCT_FROM）
      · 再往前只剩客户总资产与新开经纪账户两列
    """
    out = list(COLS)
    if ym < _DATS_MARGIN_FROM:
        out = [c for c in out if c not in ('dats_k', 'margin_balances_usdbn',
                                           'sweep_cash_usdbn', 'mmf_usdbn')]
    if ym < _CORE_NNA_FROM:
        out = [c for c in out if c != 'core_nna_usdbn']
    if ym < _CASH_PCT_FROM:
        out = [c for c in out if c != 'client_cash_pct']
    return out


# ── 源枚举 ─────────────────────────────────────────────────────────────
def _today_ym() -> tuple[int, int]:
    t = _dt.date.today()
    return (t.year, t.month)


def _shift(ym: tuple[int, int], k: int) -> tuple[int, int]:
    y, m = ym
    m += k
    while m <= 0:
        m += 12
        y -= 1
    while m > 12:
        m -= 12
        y += 1
    return (y, m)


def _candidates(back: int = 8):
    """按时间倒序给出待探的 (kind, url, report_ym)。

    月报探最近 back 个月；季报探最近 3 个季度。都探是因为季末月只有季报有，
    而季报又比「下一期月报」早一个月出来。
    """
    y, m = _today_ym()
    for k in range(back):
        yy, mm = _shift((y, m), -k)
        if mm % 3 == 0:          # 季末月没有独立月报，别浪费一次请求
            continue
        yield 'monthly', MONTHLY_URL.format(mon=_MON[mm - 1], year=yy), (yy, mm)
    q = (m - 1) // 3 + 1
    for k in range(3):
        qq, yy = q - k, y
        while qq <= 0:
            qq += 4
            yy -= 1
        yield 'quarterly', QUARTER_URL.format(q=qq, year=yy), (yy, qq * 3)


def _template_url(ym: tuple[int, int]) -> str:
    """本模块**推**出来的那个 URL。和 _candidates 给的是同一套写法，只是按月点名取一个。"""
    y, m = ym
    if m % 3 == 0:                       # 季末月没有独立月报，值在当季季报附表里
        return QUARTER_URL.format(q=(m - 1) // 3 + 1, year=y)
    return MONTHLY_URL.format(mon=_MON[m - 1], year=y)


# ── 逾期对账：拦「文件名模板过期了」这一类不出声的失败 ──────────────────
# _candidates() 给的是**推**出来的文件名，不是**发现**的。官方哪天改了命名规则，
# 新一期的 URL 就 404 —— 而 404→None（_get）与 continue（_collect）都是**设计如此**：
# 当月月报还没发之前那个 URL 本来就该 404，季末月的月报更是永远 404。
# 于是坏掉的样子是这样的：最新一期永远抓不到，老月份照旧从 cache 里读得出来，
# `if not got` 那道护栏（它要求 ~11 个候选**全部** 404）永远轮不到，update() 每天
# 干干净净返回 []、报 NOCHANGE。**连续坏十天和正常十天，日志里长得一模一样。**
# 而这不是假想：本模块自己的 _HIST_XLSX 就记着官方已经改过两次名（季报附表前缀
# schwab_ / 扩展名大写 .XLSX），且改名是**只往前改**的 —— 老文件至今仍 200。
# 只往前改，恰恰是让 got 一直为 True、让上面那道护栏永远沉默的那种改法。
#
# 所以这里补 README「第四类：不出声的失败」要的那道：拿一个**独立于文件名模板**的
# 判据来对账 —— 官方 IR 落地页上真实挂着的链接（_ir_page_links）。形状照
# fetch/cboe.py 的 _crosscheck_report_month、fetch/ice.py 的 _crosscheck_workbook_month。

# (常规月, 季末月)：某个月的数据在**该月结束后第几天**发布。这两个数是第一手实测，
# build/roster.py 的 LAG['schw'] 是它的抄件（那张表的表头注释写着「数值取自各
# fetch/<t>.py docstring 里的实测发布日」）—— 哪天实测节奏变了，两处要一起改。
# 季末月单独一个数不是可有可无的讲究：季末月（3/6/9/12）没有独立月报，数值随当季
# 季报走，晚一周。拿常规月的日子去判季末月，这道护栏会**每个季度误报一次**，
# 而每季度假一次的警报，人很快就学会无视（README「新鲜度红点」讲的是同一件事）。
_LAG_DAYS = (14, 21)
# 红线再往后推的余量。红点用的是 roster.GRACE = 5，这里刻意给到它的近三倍。
# 两者方向相反：红点早响一天只是页面上多一个红圆点；这道护栏早响一天，是把一家
# 本来健康的 fetcher 变成每日 FAIL。次序必须是「红点先红、人先看见，本护栏远远
# 跟在后面才开口」。而且余量不能按 GRACE 那个量级给 —— 8 期实测发布日
# （月报第 12/13/13/14 天、季报第 16/16/21/21 天，见 series/source_dates.csv 与
# cache 里几份新闻稿的电头）**正好顶到** _LAG_DAYS，撞一次假日顺延就越线了。
_OVERDUE_MARGIN_DAYS = 14

# 从 URL 里认报告月：判官要**容错**，因为它存在的理由就是「命名规则变了」。
# 全称与三字母缩写都认（还额外认 sept 这个常见写法），分隔符可有可无，大小写无关。
# 刻意不把 'table' / 'earnings' 写进条件：那正是可能被改掉的那部分词。
# 但也刻意**不**做成「月份三字母后面随便跟什么」—— 那样 marketdata2026.xlsx 会被
# 读成 2026-03，而这道护栏一旦误判就是 FAIL，宁可少认一种写法也不能凭空造一个月份。
_URL_MONTH_RE = re.compile(
    r'(?<![a-z])(' + '|'.join([f.lower() for f in _MON_FULL] + ['sept'] + _MON)
    + r')[-_ ]?(20\d{2})(?![0-9])', re.IGNORECASE)
_URL_QUARTER_RE = re.compile(r'(?<![a-z0-9])q([1-4])[-_ ]?(20\d{2})(?![0-9])', re.IGNORECASE)
_URL_QUARTER_RE2 = re.compile(r'(?<![0-9])(20\d{2})[-_ ]?q([1-4])(?![0-9])', re.IGNORECASE)


def _month_end(ym: tuple[int, int]) -> _dt.date:
    """该月最后一天。用「下月 1 号减一天」算，省得为了闰年再引一个 calendar。"""
    ny, nm = _shift(ym, 1)
    return _dt.date(ny, nm, 1) - _dt.timedelta(days=1)


def _overdue_deadline(ym: tuple[int, int]) -> _dt.date:
    """这个报告月的红线：过了这一天还没有它，就该有人来看一眼。"""
    lag = _LAG_DAYS[1] if ym[1] % 3 == 0 else _LAG_DAYS[0]
    return _month_end(ym) + _dt.timedelta(days=lag + _OVERDUE_MARGIN_DAYS)


def _due_ym(today: _dt.date | None = None) -> tuple[int, int]:
    """截至今天，**最新一个已经越过红线**的报告月。

    从上个月往回数，第一个红线已过的就是它（当月永远不可能越线，所以从 -1 起步）。
    参数 today 只为可测：生产路径一律用系统日期。
    """
    today = today or _dt.date.today()
    ym = _shift((today.year, today.month), -1)
    for _ in range(24):                  # 正常两三步就返回；24 只是防死循环的上界
        if _overdue_deadline(ym) <= today:
            return ym
        ym = _shift(ym, -1)
    return ym


def _url_report_ym(url: str):
    """从一个 xlsx 链接里认出它是哪一期 → (y, m)；认不出返回 None。

    季报按季末月记（q2_2026 → 2026-06），和 _candidates 对季报的记法一致，
    这样它和月报可以放在同一根时间轴上比大小。
    先只看文件名，认不出再看整条 URL —— 路径里常有「/2026/」这类年份噪音。
    """
    for text in (url.rsplit('/', 1)[-1], url):
        hit = _URL_MONTH_RE.search(text)
        if hit:
            return (int(hit.group(2)), _MON.index(hit.group(1).lower()[:3]) + 1)
        hit = _URL_QUARTER_RE.search(text)
        if hit:
            return (int(hit.group(2)), int(hit.group(1)) * 3)
        hit = _URL_QUARTER_RE2.search(text)
        if hit:
            return (int(hit.group(1)), int(hit.group(2)) * 3)
    return None


def _crosscheck_due_month(merged: dict, cache_dir: str) -> None:
    """拿到的最新月 vs 按发布节奏早该有的月，落后了就去问 IR 落地页这个外部判官。

    **抛不抛，只由判官说了算，绝不由日期单独说了算。** 这是这道护栏能进无人值守
    主路径的全部理由：官方晚发几天是常有的事（8 期实测正好顶着 _LAG_DAYS），
    光凭「过了红线」就 FAIL，等于每逢一次假日顺延就把一家健康的 fetcher 判死。
    而「落地页上明明挂着比我们拿到的更新的表」不是晚发能造出来的状态 —— 那只可能
    是我们推的文件名不对了。

    落地页这条路平时**不进主流程**（www.aboutschwab.com 在 Akamai 后面按 TLS 指纹
    拦人，见模块 docstring），所以只在越线那一天才问它一次，正常日子零额外请求。

    判官自己哑了（curl 被拦、落地页改版捞不到链接）时**只喊不抛**：一道连自己的
    探针都失灵了的护栏，不能反过来把数据摄入判死。但那一声必须喊出来 —— 否则这道
    护栏就悄悄退化成它本来要消灭的那个形状（静默）。
    """
    newest, due = max(merged), _due_ym()
    if newest >= due:
        return
    due_key = f'{due[0]:04d}-{due[1]:02d}'
    newest_key = f'{newest[0]:04d}-{newest[1]:02d}'
    line = _overdue_deadline(due)

    links = _ir_page_links(cache_dir)
    if not links:
        print(f'  [overdue] {due_key} 已过红线 {line} 仍未拿到（手上最新是 {newest_key}），'
              f'而 IR 落地页这次一个 xlsx 链接都没捞到（{IR_PAGE}，curl 被 Akamai 拦？）。'
              f'判官自己哑了，本次不判失败 —— 请人工打开落地页核一眼命名规则。')
        return

    advertised = [(ym, u) for ym, u in ((_url_report_ym(u), u) for u in links) if ym]
    if not advertised:
        print(f'  [overdue] {due_key} 已过红线 {line} 仍未拿到（手上最新是 {newest_key}）；'
              f'IR 落地页捞到 {len(links)} 个 xlsx 链接，却没有一个能解析出报告月：'
              f'{links[:5]}。命名规则可能变得连判官也认不出了，请人工核对。')
        return

    # 只认「比手上新、且自己也已经越过红线」的那一档：落地页上偶尔会先挂出还没到
    # 发布期的下一期链接，拿那种未来期当证据就是自己造一个假警报。
    newer = sorted((ym, u) for ym, u in advertised if newest < ym <= due)
    if not newer:
        print(f'  [overdue] {due_key} 已过红线 {line} 仍未出现；IR 落地页上也没有比 '
              f'{newest_key} 更新的表，判定为**官方晚发**，不是抓取故障。')
        return

    ym, real = newer[-1]
    raise FetchError(
        '逾期对账失败：IR 落地页上挂着 %s 的附表，我们按文件名模板推出来的 URL 却没拿到它。\n'
        '  落地页上的真实链接：%s\n'
        '  本模块推出来的 URL：%s\n'
        '  手上最新月 %s，红线 %s（= 月末 + %s 天节奏 + %d 天余量）。\n'
        'CDN 的命名规则或路径多半变了（史上改过：季报附表前缀 schwab_、扩展名大写 .XLSX，'
        '见 _HIST_XLSX）。这时候静默返回 NOCHANGE，页面会一直挂着旧数据而日志上看不出'
        '任何异常，所以这里拒绝写入：请照落地页上的真名改 MONTHLY_URL / QUARTER_URL，'
        '再跑一次。'
        % (f'{ym[0]:04d}-{ym[1]:02d}', real, _template_url(ym), newest_key, line,
           _LAG_DAYS[1] if due[1] % 3 == 0 else _LAG_DAYS[0], _OVERDUE_MARGIN_DAYS))


# ── 历史回填源（2013-09 … 2018-04）─────────────────────────────────────
# 主流程 _candidates() 只探最近 8 个月 / 3 个季度 —— 那是「每月增量」该有的样子，
# 不该为了追历史每次多打几十个请求。但 2026-08-16 的一次历史考古发现，官方手上
# **还留着**一批更老的表，主流程碰不到它们只是因为文件名有两处变体：
#
#   (a) **前缀**：2017 年及更早的季报附表叫 `schwab_`，不是 `schw_`
#       （schw_q1_2017…=404，schwab_q1_2017…=200）；
#   (b) **扩展名大写**：2017Q3–2018Q4 的几份是 `.XLSX`
#       （schw_q2_2018_earnings_tables.xlsx=404，同名 .XLSX=200）。
#
# 这两条都不是规律，是历史遗留的手工命名，推不出来 —— 所以这里**写死一份清单**而不是
# 拼 URL。清单里的每一份都实测过 200 且解析通过；哪天官方撤下某一份，backfill() 会
# 在 stdout 上点名说它没了，而不是静默少几个月（已经落盘的历史行不受影响）。
#
# 再往前（2013-09 … 2016-12）CDN 上一份都不剩了，但 **SEC EDGAR 上有**：Schwab 把
# 「月度活动报告」原样作为季度 8-K 的 EX-99.1 附上去，正文里就是那张 13 个月滚动表。
# 这条路对**每一期**都有效（不止老年份），详见上方 docstring 里关于两代版式的那一段。
_HIST_XLSX = [
    # (报告月, URL 文件名) —— 报告月 = 该表最右一列，parse_table 按它倒推整张表
    ((2016, 9),  'schw_q3_2016_earnings_tables.xlsx'),
    ((2016, 12), 'schw_q4_2016_earnings_tables.xlsx'),
    ((2017, 3),  'schwab_q1_2017_earnings_tables.xlsx'),
    ((2017, 6),  'schwab_q2_2017_earnings_tables.xlsx'),
    ((2017, 9),  'schwab_q3_2017_earnings_tables.XLSX'),
    ((2018, 2),  'schwab_feb2018_table.XLSX'),
    ((2018, 3),  'schw_q1_2018_earnings_tables.XLSX'),
    ((2018, 6),  'schw_q2_2018_earnings_tables.XLSX'),
    ((2018, 9),  'schw_q3_2018_earnings_tables.xlsx'),
    ((2018, 12), 'schw_q4_2018_earnings_tables.XLSX'),
    ((2019, 1),  'schwab_jan2019_table.xlsx'),
    ((2019, 2),  'schwab_feb2019_table.xlsx'),
    ((2019, 4),  'schwab_apr2019_table.xlsx'),
    ((2019, 5),  'schw_may2019_table.xlsx'),
]
EDGAR_CIK = '0000316709'
EDGAR_SUB = 'https://data.sec.gov/submissions/CIK{cik}.json'
EDGAR_DIR = 'https://www.sec.gov/Archives/edgar/data/316709/{acc}'
# EDGAR 要求 UA 里带得到人的联系方式，否则 403。这不是反爬，是它明文写的使用条款。
_EDGAR_UA = 'monthly-op-dashboards research (hzhan7@gmail.com)'
# 扫到哪一期为止。**不是**因为之后就没有表了（2026 年那期 EX-99.1 里照样有），
# 而是 2018-05 起 series/schw.csv 本来就有值，再扫下去只是拿同样的数对一遍账。
# 留一年余量：往后多扫一点，重叠对账的样本就多一点，那是这条路唯一的自检。
_EDGAR_UNTIL = '2020-01-01'


def _edgar_get(url: str) -> bytes:
    req = urllib.request.Request(url, headers={'User-Agent': _EDGAR_UA,
                                               'Accept-Encoding': 'gzip'})
    with urllib.request.urlopen(req, timeout=60) as r:
        blob = r.read()
    if blob[:2] == b'\x1f\x8b':
        import gzip
        blob = gzip.decompress(blob)
    return blob


def _edgar_8k(limit_before: str = _EDGAR_UNTIL) -> list:
    """CIK 的 8-K 清单（含历史分片），只留 limit_before 之前的。→ [(date, acc, [docs])]"""
    j = json.loads(_edgar_get(EDGAR_SUB.format(cik=EDGAR_CIK)))
    rows = []
    packs = [j['filings']['recent']]
    for f in j['filings'].get('files', []):
        _time.sleep(0.15)
        packs.append(json.loads(_edgar_get(
            'https://data.sec.gov/submissions/' + f['name'])))
    for p in packs:
        rows += [(d, a) for a, fm, d in
                 zip(p['accessionNumber'], p['form'], p['filingDate'])
                 if fm.startswith('8-K') and d < limit_before]
    return sorted(set(rows))


_HTML_TAG = re.compile(r'<[^>]+>')
_TITLE_RE = re.compile(r'Monthly Activity Report\s+For\s+([A-Za-z]+)\s+(\d{4})', re.I)


def _cell_text(c: str) -> str:
    t = _HTML_TAG.sub('', c)
    t = (t.replace('&#xFEFF;', '').replace('&#x2019;', "'").replace('&amp;', '&')
          .replace('&#160;', ' ').replace('&nbsp;', ' '))
    return re.sub(r'&#\d+;', '', t).strip()


def _glue(cells: list) -> list:
    """把被拆进独立 <td> 的「零件」粘回它前面那个数，再丢掉空单元格。

    这一步**不能省，也不能放在丢空格之后**。EDGAR 的表把会计负号的右括号、百分号
    单独放进一个 <td>：`(0.3` `)` 是两格。先丢空格再配对的话，`(0.3` 过不了 float()
    变成 None，那一行就**少一个 token**，于是 zip(months, vals) 之后**整行往后错一格** ——
    错位的每个值都是「上个月的数」，看着完全正常，只有跨源对账才抓得到。
    （实测：只改锚点不加这一步，2019-07…2020-07 那批文件里 core NNA 一列错出 44 个值，
    全部等于前一个月 —— 病根就是 2019-04 的 core NNA 是 -0.3。）
    """
    out: list = []
    for c in cells:
        if c in (')', '%', ')%', 'bp', ')bp') and out:
            out[-1] += c
        elif c != '':
            out.append(c)
    return out


def _html_num(x: str):
    """'2,556.7' → 2556.7；'(0.3)' → -0.3；'10.1%' → 10.1；'-'/'' → None。

    括号是会计负号。尾部的 '%' 要剥掉：`_glue` 会把独立 <td> 里的百分号粘回数值后面，
    不剥的话 float() 失败 → 整行返回 None → 「这份文件没有这一行」，与真的没有长得一样
    （客户现金占比那一行 2013 年至今每期都印，就是这么被静默丢掉的）。
    百分点与分数的归一交给 _RATIO 那条倍率判定，这里只负责把字符串变成数。
    """
    t = x.replace(',', '').replace('$', '').strip()
    if t.endswith('%'):
        t = t[:-1].strip()
    neg = t.startswith('(') and t.endswith(')')
    if neg:
        t = t[1:-1].strip()
    if t in ('', '-', '\u2014', 'N/A', '*'):
        return None
    try:
        v = float(t)
    except ValueError:
        return None
    return -v if neg else v


def parse_edgar_monthly(html: str) -> dict:
    """8-K EX-99.1 里的月度活动报告表 → {(y, m): {col: 值}}；不是这种文件返回 {}。

    与 xlsx 那条路共用同一套规矩：**按标签前缀取行、按锚点行反推单位倍率、
    表头月份与标题月倒推逐个核对**。核对不上就整份丢掉（返回 {}）而不是猜 ——
    这批文件跨 4 年、版式改过好几次，错位一格在图上看不出来。
    """
    # ── 锚点：必须锚在**标题**上，不能锚在「monthly activity report」这个词上 ──
    # 正文里另有一句脚注「…please see the Monthly Activity Report.」，位置比真表更靠前。
    # 锚错了还不止是找不到表：老版式（2017-03 及更早）标题在表的**第一行里**，
    # 所以要往回找 <table>；新版式标题在表**上方的 <div> 里**，要往后找。
    # 一律往回找的话，新版式会捞到上面某张毫不相干的表 —— 而且**静默**返回 {}。
    # 本文件 2026-08-16 首版就踩了这个坑：97 份 EX-99.1 只读出 15 份，
    # 于是得出了「2017-04 起表就不随 8-K 附了」这个**错误**结论（docstring 已更正）。
    hits = list(_TITLE_RE.finditer(html))
    if not hits:
        return {}
    hit = hits[-1]                       # 一份文件里标题只出现一次；取最后一个最稳
    i = hit.start()
    a_back = html.lower().rfind('<table', 0, i)
    if a_back >= 0 and html.lower().find('</table>', a_back) > i:
        a = a_back                       # 老版式：标题落在这张表内部
    else:
        a = html.lower().find('<table', i)   # 新版式：标题在表上方
    b = html.find('</table>', a) if a >= 0 else -1
    if a < 0 or b < 0:
        return {}
    rows = []
    for r in re.findall(r'<tr[^>]*>(.*?)</tr>', html[a:b + 8], re.S):
        cells = [_cell_text(c) for c in re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', r, re.S)]
        rows.append(_glue(cells))
    ay, am = int(hit.group(2)), _MON.index(hit.group(1).lower()[:3]) + 1
    hdr = next(([c.lower()[:3] for c in r if c.lower()[:3] in _MON]
                for r in rows if sum(c.lower()[:3] in _MON for c in r) >= 12), None)
    if not hdr:
        return {}
    yms = []
    y, m = ay, am
    for _ in hdr:
        yms.append((y, m))
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    yms.reverse()
    if [_MON[mm - 1] for _, mm in yms] != hdr:      # 表头与标题月倒推对不上 → 版式变了
        return {}

    def row_of(prefix, skip_core=False):
        for r in rows:
            lab = re.sub(r'\s+', ' ', r[0]).strip().lower() if r else ''
            if not lab.startswith(prefix):
                continue
            if skip_core and lab.startswith('core'):
                continue
            vals = [_html_num(x) for x in r[1:1 + len(yms)]]
            if sum(v is not None for v in vals) >= len(yms) - 2:
                return {ym: v for ym, v in zip(yms, vals) if v is not None}
        return None

    a_money = row_of(_LABEL['_anchor_money'])
    a_count = row_of(_LABEL['_anchor_count'])
    if not a_money or not a_count:
        return {}
    s_money = _scale(a_money, *_SANE['total_client_assets_usdbn'], 'edgar money')
    s_count = _scale(a_count, *_SANE['_anchor_count'], 'edgar count')

    out: dict = {ym: {} for ym in yms}
    for col in COLS:
        vals = row_of(_LABEL[col])
        if vals is None:
            continue
        if col in _RATIO:
            s = _scale(vals, *_SANE[col], f'edgar {col}', cands=(1.0, 100.0))
        else:
            s = s_money if col in _MONEY else s_count
        for ym, v in vals.items():
            x = v * s
            lo_hi = _SANE.get(col)
            if lo_hi and not (lo_hi[0] <= abs(x) <= lo_hi[1]):
                raise FetchError(f'EDGAR {ym} {col}={x} 超出合理区间 {lo_hi}')
            out[ym][col] = round(x, 6)
    return {ym: v for ym, v in out.items() if v}


RESTATEMENTS: list = []      # 上一次 _collect 发现的跨源/跨期数值打架，供调用方审计
# 上一次 _collect 里每个月的数值最终由哪份文件说了算：{(y,m): (kind, 报告月)}。
# update() 靠它判断这个月是「自己那期报告」发的还是从后续滚动表里补来的 —— 只有前者
# 才有资格把那期的电头日期当成这个月的发布日。
ORIGIN: dict = {}


def _collect(cache_dir: str, back: int = 8) -> dict:
    """下载 + 解析所有能拿到的源，合并成 {(y,m): {col: val}}。

    _candidates 是「新→旧」序，所以合并规则是：**先写者胜**（越新的文件越权威），
    唯一的例外是季报——季报是该季末月的最终版，rank 更高，可以覆盖月报的同月值。
    覆盖时如果新旧值不等，说明官方重述了，记进 RESTATEMENTS 让人看得见，不静默吞掉。
    """
    merged: dict[tuple[int, int], dict] = {}
    origin = ORIGIN
    origin.clear()
    RESTATEMENTS.clear()
    got = False
    for kind, url, ym in _candidates(back):
        fresh = ym >= _shift(_today_ym(), -2)          # 最近两个月的文件不吃缓存
        path = _download(url, cache_dir, reuse=not fresh)
        if path is None:
            continue
        got = True
        for m_ym, vals in parse_table(path, ym).items():
            prev_rank = _SOURCE_RANK.get(origin.get(m_ym, (None,))[0], -1)
            slot = merged.setdefault(m_ym, {})
            for c, v in vals.items():
                if c in slot:
                    if _differs(c, slot[c], v):
                        RESTATEMENTS.append((m_ym, c, slot[c], v, os.path.basename(path)))
                    if _SOURCE_RANK[kind] <= prev_rank:
                        continue                       # 先写者胜
                slot[c] = v
            if _SOURCE_RANK[kind] > prev_rank:
                origin[m_ym] = (kind, ym)
            else:
                origin.setdefault(m_ym, (kind, ym))
    if not got:
        links = _ir_page_links(cache_dir)
        raise FetchError(
            '最近 %d 个月的月报附表和最近 3 个季度的季报附表**全部** 404 —— 整个目录或'
            '域名搬家了。（只改命名规则走不到这里：官方改名是只往前改的，老文件照旧 200，'
            '这个 got 就一直是 True。那种坏法由 _crosscheck_due_month 负责，别再指望这一条。）'
            'IR 落地页上当前的 xlsx 链接：%s'
            % (back, links[:10] or '（落地页也取不到）'))
    # 上面那道只管「一个都没拿到」。真正常见的是「老的都拿到了、最新一期没拿到」——
    # 那种没有任何 404 计数、没有任何异常，得靠外部判官对账。见该函数 docstring。
    _crosscheck_due_month(merged, cache_dir)
    return merged


# ── 对外接口 ───────────────────────────────────────────────────────────
def latest_month(cache_dir) -> str | None:
    """官方源当前最新月，"YYYY-MM"。抓不到抛 FetchError。"""
    data = _collect(cache_dir)
    ok = [ym for ym, v in data.items() if all(c in v for c in _required(ym))]
    if not ok:
        raise FetchError('源文件解析出来了，但没有任何一个月凑齐必需列')
    y, m = max(ok)
    return f'{y:04d}-{m:02d}'


def _differs(col: str, a, b) -> bool:
    """两个源给同一格的值，**按落盘精度**判是不是真的不一样。

    不用 `abs(a-b) > 1e-6`：那是原精度比较，会在 CSV 里根本体现不出来的地方报警。
    实例（2026-08-19）：client_cash_pct 的 2019-01，月报 xlsx 里是 0.1174（浮点原值），
    EDGAR 的 HTML 印的是 "11.7%" —— 两者落盘都是 11.7，CSV 一个字符都不差，
    原精度比较却每次 --backfill 都吐三行 RESTATEMENTS。这类噪音的代价不是刷屏，
    是让人学会忽略这张表，于是真的官方重述来的那天也被一起忽略掉。
    """
    return _fmt(col, a) != _fmt(col, b)


def _fmt(col: str, v) -> str:
    if v is None:
        return ''
    if col in ('new_brokerage_accounts_k', 'dats_k'):
        return str(int(round(v)))
    s = f'{round(v, 1):.1f}'
    return s


def _record_source_dates(series_dir, added: list, cache_dir) -> None:
    """给刚写进 series 的月份记下官方发布日。**只在数据确实落盘之后调**。

    这里一律不抛：发布日只影响页面抬头那半句话，抓不到就让它缺席（source_dates.py 的
    docstring 讲了为什么缺席好过瞎猜），把整月数据的摄入连坐掉是本末倒置。
    但每一次拿不到都要在 stdout 上说清楚是哪一期、卡在哪 —— 静默才是不能接受的。
    """
    import importlib.util
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(root, 'source_dates.py'))
    sd = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(sd)

    for ym in added:
        key = f'{ym[0]:04d}-{ym[1]:02d}'
        kind, report_ym = ORIGIN.get(ym, (None, None))
        if report_ym != ym:
            # 这个月是从后面某一期的 13 个月滚动表里补来的（季末月缺季报时会走到这），
            # 那期的电头是它自己的发布日，套到这个月上就是编造。
            print(f'  [source_date] {key} 跳过：数值来自 {kind} {report_ym} 那一期的滚动表，'
                  f'不是这个月自己的报告')
            continue
        try:
            date, why = release_date(kind, report_ym, cache_dir)
        except FetchError as e:
            date, why = None, str(e)
        if not date:
            print(f'  [source_date] {key} 未记录：{why}')
            continue
        try:
            sd.record(series_dir, 'schw', key, date, why)
            print(f'  [source_date] {key} → {date}（{why}）')
        except ValueError as e:      # record 的三道护栏拦下来了，说明解析取错了东西
            print(f'  [source_date] {key} 拒收 {date}：{e}')


def update(series_dir, cache_dir) -> list:
    """把新月份追加进 series/schw.csv，返回新增月份 ["YYYY-MM", ...]。

    幂等：已存在的月份一律跳过（不覆盖、不重排、不改动已有行的任何字符）。
    """
    path = os.path.join(series_dir, SERIES)
    with open(path, newline='') as f:
        rows = list(csv.reader(f))
    header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
    if header != ['month'] + COLS:
        raise FetchError(f'{SERIES} 列名与预期不符：{header}')
    have = {r[0] for r in body}

    data = _collect(cache_dir)
    added = []
    for ym in sorted(data):
        key = f'{ym[0]:04d}-{ym[1]:02d}'
        if key in have:
            continue
        vals = data[ym]
        missing = [c for c in _required(ym) if c not in vals]
        if missing:
            raise FetchError(f'{key} 解析结果缺列 {missing}，拒绝写入（不写 NaN）')
        body.append([key] + [_fmt(c, vals.get(c)) for c in COLS])
        added.append(ym)

    if not added:
        return []
    body.sort(key=lambda r: r[0])
    tmp = path + '.tmp'
    with open(tmp, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(body)
    os.replace(tmp, path)

    _record_source_dates(series_dir, added, cache_dir)
    return [f'{y:04d}-{m:02d}' for y, m in added]


def backfill(series_dir, cache_dir, verbose: bool = True) -> list:
    """把 2013-09 … 2018-04 的历史月份补进 series/schw.csv，返回新增月份。

    与 update() 的关系：update() 管**增量**（每月往后走一格，只探最近几期），
    backfill() 管**存量**（一次性把官方还留着的老表全扫一遍）。两者写同一个 CSV，
    都遵守同一条铁律：**已存在的月份一个字符都不改**。

    重叠月是这里最重要的一道自检 —— 老表与仓里现有的 99 个月有大量重叠，两边**必须**
    逐值相等。不等只有两种可能：解析取错了行/列，或者官方重述过。两种都不能静默：
    对不上的行会被打印出来并让整次回填失败（`FetchError`），要人去看。
    （2026-08-16 首次跑通时：57 个重叠 (月,指标) 全部相等，0 处不符。）

    幂等：跑第二遍不会有任何新增，也不会改动任何已有行。
    """
    def log(*a):
        if verbose:
            print(*a)

    merged: dict[tuple[int, int], dict] = {}
    origin: dict[tuple[int, int], str] = {}

    def absorb(vals: dict, src: str):
        """先写者胜。这里的「先」= 调用顺序：CDN 的表比 EDGAR 的新，先扫 CDN。"""
        for ym, cols in vals.items():
            slot = merged.setdefault(ym, {})
            for c, v in cols.items():
                if c in slot:
                    if _differs(c, slot[c], v):
                        RESTATEMENTS.append((ym, c, slot[c], v, src))
                    continue
                slot[c] = v
                origin.setdefault(ym, src)

    for ym, name in _HIST_XLSX:
        url = f'{CDN}/excels/{name}'
        try:
            path = _download(url, cache_dir)
        except FetchError as e:
            log(f'  [hist] {name} 下载失败：{e}')
            continue
        if path is None:
            log(f'  [hist] {name} 已从 CDN 撤下（404）—— 已落盘的历史行不受影响')
            continue
        absorb(parse_table(path, ym), name)
        log(f'  [hist] {name} ok')

    # EDGAR：2017-03 之前的月度表嵌在季度 8-K 的 EX-99.1 里
    try:
        fils = _edgar_8k()
    except Exception as e:
        log(f'  [edgar] 清单取不到（{type(e).__name__}: {e}），本轮只用 CDN 的历史表')
        fils = []
    for date, acc in fils:
        base = EDGAR_DIR.format(acc=acc.replace('-', ''))
        try:
            idx = json.loads(_edgar_get(base + '/index.json'))
        except Exception:
            continue
        _time.sleep(0.15)
        # 附件命名两代都有：`schw-20141015ex99157591b.htm` 与 `exhibit991093018.htm`。
        # 只筛 'ex99' 会**静默漏掉后者**（'exhibit991…' 里没有 ex99 这个连续子串），
        # 而漏掉的表现是「这期没有月度表」，与「真的没有」长得一模一样。
        # 一份 8-K 通常只挂 3–6 个文档，全试一遍的代价可以忽略，比猜命名规律稳。
        docs = [it['name'] for it in idx['directory']['item']
                if it['name'].lower().endswith(('.htm', '.html'))]
        for d in docs:
            try:
                html = _edgar_get(f'{base}/{d}').decode('utf-8', 'replace')
            except Exception:
                continue
            _time.sleep(0.15)
            got = parse_edgar_monthly(html)
            if got:
                absorb(got, f'EDGAR {acc}/{d}')
                log(f'  [edgar] {date} {d} → {min(got)}…{max(got)}')
                break

    if not merged:
        raise FetchError('历史回填一个月份都没解析出来 —— CDN 与 EDGAR 两条路都空了，'
                         '不要当成「没有历史数据」，先人工核对 _HIST_XLSX 里的链接。')

    # ── 与现有 CSV 对账 ──
    path = os.path.join(series_dir, SERIES)
    with open(path, newline='') as f:
        rows = list(csv.reader(f))
    header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
    if header != ['month'] + COLS:
        raise FetchError(f'{SERIES} 列名与预期不符：{header}')
    have = {r[0]: r for r in body}

    clash, checked = [], 0
    for ym, vals in merged.items():
        key = f'{ym[0]:04d}-{ym[1]:02d}'
        if key not in have:
            continue
        for c, v in vals.items():
            cur = have[key][1 + COLS.index(c)]
            if cur == '':
                continue
            checked += 1
            if abs(float(cur) - v) > max(0.05, abs(float(cur)) * 1e-6):
                clash.append((key, c, float(cur), v))
    log(f'  [对账] 重叠 {checked} 个 (月,指标)，不符 {len(clash)} 个')
    if clash:
        for c in clash[:20]:
            log('    MISMATCH', c)
        raise FetchError(f'历史表与仓里现有值有 {len(clash)} 处不符，拒绝写入。'
                         '要么解析取错了行，要么官方重述过 —— 两种都得人去看。')

    added = []
    for ym in sorted(merged):
        key = f'{ym[0]:04d}-{ym[1]:02d}'
        if key in have:
            continue
        vals = merged[ym]
        missing = [c for c in _required(ym) if c not in vals]
        if missing:
            log(f'  [跳过] {key} 缺 {missing}')
            continue
        body.append([key] + [_fmt(c, vals.get(c)) for c in COLS])
        added.append(key)

    if not added:
        return []
    body.sort(key=lambda r: r[0])
    # 连续性自检：回填后月份必须仍然逐月连号（build/schw.py 的 assert_monthly 也查这条，
    # 但在写盘之前就拦住，比让下游构建失败要好定位得多）。
    for i in range(1, len(body)):
        py, pm = int(body[i - 1][0][:4]), int(body[i - 1][0][5:])
        cy, cm = int(body[i][0][:4]), int(body[i][0][5:])
        if (cy, cm) != _shift((py, pm), 1):
            raise FetchError(f'回填后月份不连续：{body[i - 1][0]} → {body[i][0]}。'
                             '缺口会让 assets.diff() 把两个月的变动记到一个月上，拒绝写入。')
    tmp = path + '.tmp'
    with open(tmp, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(body)
    os.replace(tmp, path)
    log(f'  [回填] 新增 {len(added)} 个月：{added[0]} … {added[-1]}')
    return added


# ═════════════════ 列回填：给**已有的行**补上新增列（--columns）═════════════════
# 与 backfill() 的分工要说清楚，否则下一个人会把两者混成一个：
#   · update()           —— 往后走一格，**加行**（新月份）。
#   · backfill()         —— 往前补历史，**加行**（2013-09…2018-04）。
#   · backfill_columns() —— **一个行都不加**，只把新增的列填进已经存在的行里。
#
# 为什么必须单独有这一条路：前两条都只在「这个月不在 CSV 里」时才写盘，
# 于是 COLS 里新加一列时，历史行的那一格会**永远**空着 —— 加列的人跑一遍 update()
# 看见「新增 0 个月」，会以为没事。2026-08-19 加 client_cash_pct 三列时就是这样发现的。
#
# 铁律（与另两条一致）：**非空单元格一个字符都不改**；重叠值不符就整次失败。
_NEW_COLS = ('client_cash_pct', 'sweep_cash_usdbn', 'mmf_usdbn')


def _col_sources(cache_dir: str):
    """列回填要覆盖 2013-09 至今**每一个月**，所以源的排法和 backfill() 不同。

    月报的 13 个月滚动表意味着「每 12 个月取一份」就够铺满，不必逐月下载：
      · CDN 月报 schw_may<y>_table.xlsx（y=2019…今年）→ 2018-05 … 最新一期的 5 月
      · CDN 最近 8 个月 / 3 个季度（_candidates）→ 补上最新一期 5 月之后的月份
      · CDN 历史附表 _HIST_XLSX → 2015-09 … 2018-12
      · 再往前只有 SEC EDGAR（见下方 EDGAR 那一段）
    """
    y_now = _today_ym()[0]
    out = []
    for y in range(2019, y_now + 1):
        out.append(((y, 5), MONTHLY_URL.format(mon='may', year=y)))
    for _kind, url, ym in _candidates(back=8):
        out.append((ym, url))
    for ym, name in _HIST_XLSX:
        out.append((ym, CDN + '/excels/' + name))
    return out


def backfill_columns(series_dir, cache_dir, cols=_NEW_COLS,
                     dry_run: bool = False, verbose: bool = True) -> dict:
    """把 cols 里的列补进 series/schw.csv 已有的行。返回 {列: 填了几格}。"""
    log = (lambda *a: print(*a)) if verbose else (lambda *a: None)
    path = os.path.join(series_dir, SERIES)
    with open(path, newline='') as f:
        rows = list(csv.reader(f))
    header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]

    # ── 表头迁移：老表没有这几列，按 COLS 的顺序在每一行末尾补空格 ──
    old_header = ['month'] + [c for c in COLS if c not in cols]
    if header == ['month'] + COLS:
        pass
    elif header == old_header:
        log(f'  [迁移] 表头补 {len(cols)} 列：{", ".join(cols)}')
        header = ['month'] + COLS
        body = [r + [''] * len(cols) for r in body]
    else:
        raise FetchError(f'{SERIES} 列名与预期不符：{header}')
    want = {r[0] for r in body}

    # ── 采集 ──
    merged: dict[tuple[int, int], dict] = {}
    prov: dict[tuple[int, int], str] = {}

    def absorb(got: dict, src: str):
        for ym, vals in got.items():
            keep = {c: v for c, v in vals.items() if c in cols}
            if not keep:
                continue
            # 先写者胜（源按「新 → 旧」排，越新的文件越权威），与 _collect 同规矩
            if ym not in merged:
                merged[ym], prov[ym] = keep, src
            else:
                for c, v in keep.items():
                    if c not in merged[ym]:
                        merged[ym][c] = v

    def need() -> set:
        return {k for k in want
                if tuple(int(x) for x in k.split('-')) not in merged
                or 'client_cash_pct' not in merged[tuple(int(x) for x in k.split('-'))]}

    seen = set()
    for ym, url in _col_sources(cache_dir):
        name = url.rsplit('/', 1)[-1]
        if name in seen:
            continue
        seen.add(name)
        try:
            fp = _download(url, cache_dir)
        except FetchError as e:
            log(f'  [cdn] {name} 取不到：{e}')
            continue
        if fp is None:
            continue
        try:
            absorb(parse_table(fp, ym), name)
        except FetchError as e:
            log(f'  [cdn] {name} 解析失败：{e}')
            continue
        log(f'  [cdn] {name} ok')

    # ── EDGAR：CDN 上 2015-09 之前一份都不剩，只能走 8-K 的 EX-99.1 ──
    # 按**从新到旧**走，每份只解析到「还缺的月份都齐了」为止就停 —— 全量扫 160 份
    # 8-K 要几百个请求，而这里真正缺的只有 2013-09…2015-08 那二十来个月。
    if need():
        try:
            fils = sorted(_edgar_8k('2016-06-01'), reverse=True)
        except Exception as e:
            log(f'  [edgar] 清单取不到（{type(e).__name__}: {e}）')
            fils = []
        for date, acc in fils:
            if not need():
                break
            base = EDGAR_DIR.format(acc=acc.replace('-', ''))
            try:
                idx = json.loads(_edgar_get(base + '/index.json'))
            except Exception:
                continue
            _time.sleep(0.15)
            docs = [it['name'] for it in idx['directory']['item']
                    if it['name'].lower().endswith(('.htm', '.html'))]
            for d in docs:
                try:
                    html = _edgar_get(f'{base}/{d}').decode('utf-8', 'replace')
                except Exception:
                    continue
                _time.sleep(0.15)
                got = parse_edgar_monthly(html)
                if got:
                    absorb(got, f'EDGAR {acc}/{d}')
                    log(f'  [edgar] {date} {d} → {min(got)}…{max(got)}')
                    break

    # ── 对账：已有的非空格子必须逐值相等（重跑幂等；不等说明解析取错了行）──
    idx_of = {c: 1 + COLS.index(c) for c in cols}
    clash, checked, filled = [], 0, {c: 0 for c in cols}
    for r in body:
        ym = tuple(int(x) for x in r[0].split('-'))
        vals = merged.get(ym, {})
        for c in cols:
            v = vals.get(c)
            if v is None:
                continue
            cur = r[idx_of[c]]
            if cur != '':
                checked += 1
                if cur != _fmt(c, v):
                    clash.append((r[0], c, cur, _fmt(c, v)))
                continue
            r[idx_of[c]] = _fmt(c, v)
            filled[c] += 1
    log(f'  [对账] 重叠 {checked} 个 (月,指标)，不符 {len(clash)} 个')
    if clash:
        for x in clash[:20]:
            log('    MISMATCH', x)
        raise FetchError(f'列回填与仓里现有值有 {len(clash)} 处不符，拒绝写入。')

    # ── 恒等式自检：(sweep + MMF) / 客户资产 应当复现官方自己印的占比 ──
    # 三个数都是官方同一张表上印的，这一步不是「算出」占比，是**核对**我们把三行
    # 各自取对了没有。容差 0.07pp：占比印到 0.1pp、两个分量各印到 0.1bn。
    ia = 1 + COLS.index('total_client_assets_usdbn')
    bad = []
    for r in body:
        sw, mm, pc, ta = (r[idx_of.get('sweep_cash_usdbn', 0)], r[idx_of.get('mmf_usdbn', 0)],
                          r[idx_of.get('client_cash_pct', 0)], r[ia])
        if '' in (sw, mm, pc, ta):
            continue
        d = abs((float(sw) + float(mm)) / float(ta) * 100 - float(pc))
        if d > 0.07:
            bad.append((r[0], round(d, 3)))
    log(f'  [恒等式] (sweep+MMF)/资产 vs 官方占比：核了 '
        f'{sum(1 for r in body if r[idx_of.get("sweep_cash_usdbn", 0)] and r[idx_of.get("mmf_usdbn", 0)] and r[idx_of.get("client_cash_pct", 0)])} '
        f'个月，超差 {len(bad)} 个')
    if bad:
        raise FetchError(f'(sweep+MMF)/资产 与官方印的客户现金占比对不上：{bad[:10]}')

    log('  [填充] ' + '、'.join(f'{c} {n} 格' for c, n in filled.items()))
    miss = sorted(k for k in want
                  if not body[[r[0] for r in body].index(k)][idx_of['client_cash_pct']])
    if miss:
        log(f'  [仍缺] client_cash_pct 有 {len(miss)} 个月没填上：{miss[:6]}…{miss[-3:]}')
    if dry_run:
        log('  [dry-run] 不写盘')
        return filled
    if not any(filled.values()):
        return filled
    body.sort(key=lambda r: r[0])
    tmp = path + '.tmp'
    with open(tmp, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(body)
    os.replace(tmp, path)
    return filled



if __name__ == '__main__':
    import sys
    _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if '--columns' in sys.argv:
        # 给已有的行补新增列（不加行）。加完新列后跑一次，之后是幂等的。
        print('backfill_columns:',
              backfill_columns(os.path.join(_root, 'series'),
                               os.path.join(_root, 'cache'),
                               dry_run='--dry-run' in sys.argv))
        raise SystemExit(0)
    if '--backfill' in sys.argv:
        # 一次性的历史存量补齐，不进 monthly_run 的每日/每月主路径。
        print('backfill:', backfill(os.path.join(_root, 'series'),
                                    os.path.join(_root, 'cache')))
        if RESTATEMENTS:
            print('跨源数值打架（先写者胜，此处仅记录）:')
            for r in RESTATEMENTS:
                print('  ', r)
        raise SystemExit(0)
    print('latest_month:', latest_month(os.path.join(_root, 'cache')))
    print('update      :', update(os.path.join(_root, 'series'), os.path.join(_root, 'cache')))
    if RESTATEMENTS:
        print('官方重述（人工确认后再信新值）:')
        for r in RESTATEMENTS:
            print('  ', r)
