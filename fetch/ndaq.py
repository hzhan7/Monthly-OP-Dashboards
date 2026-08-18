# -*- coding: utf-8 -*-
"""Nasdaq, Inc. (NDAQ) 月度成交量与季度经营面板 —— 无人值守抓取。

本模块**只做忠实入库**：官方发张数就存张数，官方发金额就存金额并注明币种，
官方发的是「当月合计」就存当月合计。**一切换算（ADV、定基名义额、FX）都在 build/ 那一层做。**
所以下面每一列的口径写得比通常更啰嗦 —— 下游换算全靠这段文字，
口径写错比数字抓错更难被发现（数字错了图会跳，口径错了图很好看但意思是反的）。

━━━━━━━━━━━━━━━━━━━━ 数据源 ━━━━━━━━━━━━━━━━━━━━

四个域，全部是 Nasdaq 自营，没有任何第三方聚合站：

A 组 · IR「Monthly Reporting Sheet」PDF —— 四条月度序列 + 一张季度面板
    落地页 https://ir.nasdaq.com/financials/volume-statistics
    直链   https://ir.nasdaq.com/static-files/<uuid>
    2026-08-06 实测 uuid = 465d2157-c476-4546-a9f7-8d7ad0c9be77，352,432 B，2 页，
    Excel 导出（producer = "Microsoft® Excel® for Microsoft 365"）。

    **uuid 不写死**：落地页 `<div class="blueHeading">Monthly Volumes</div>` 底下唯一那条
    `<a>Monthly Metrics</a>` 就是它，本模块每次从落地页正则抽。写死能跑，但官方换 uuid 的
    那一天会静默地一直下老文件 —— 那种坏法没有任何症状。
    实测落地页上 `/static-files/<uuid>` 只有一个命中，正则不会歧义。

    ⚠ **uuid 本身不带月份，每月原地替换**。所以「最新月」只能解析 PDF 内容取
    「年份行里最后一个有数的月」，不能从 URL 推、也不能只信 `Content-Disposition`
    里的 `Monthly Reporting Sheet - July 2026 Final.pdf`（"Final" 说明存在过非 Final 版）。
    本模块把 Content-Disposition 当**交叉校验**（对不上就抛），不当唯一判据。

    历史深度：这份文件恒含「上一整年 + 本年 YTD」= 19–24 个月，**这份 PDF 本身 2025-01
    之前回不去**（IR 站不留历史副本；2016 的新闻稿 slug 实测 404、2024-10 那篇 200 但
    正文里千分位数字命中 0 个、`/static-files/` 链接 0 条；落地页全页只有 1 条 static-file、
    没有任何年份归档链接 —— 三条都在 2026-08-18 重新实测过）。跨年不掉数据：12 月那期含全年 12 个月。

    ⚠ **「PDF 回不去」≠「这四列回不去」。** 四列里有两列在别的 Nasdaq 官方档案里躺着完整历史，
    2026-08-18 已接上（见 D 组与下面的口径坑 19 / 20）：
        vol_nordic_derivs_mmcontracts  → D 组交易所公告档案，2013-01 起，19 个重叠月逐月全等
        vol_us_cash_matched_mnsh       → B 组三盘口和，2010-10 起，19 个重叠月里 18 个 |差| ≤ 0.0101%
    另外两列仍然只有 2025-01 起（**2026-08-18 复核后维持原判**）：
        vol_us_options_mmcontracts   Nasdaq 一方确实没有月度档案。同目录的 msoption{YY}.xlsx
                                     2016/2020 年份下的是 43,120 B 的 HTML 错误页；2023/2026
                                     是真 xlsx 但只有 Combined/BX(NTX)/NOM/PHLX 四行、
                                     本月/上月/去年同月三列（2023-12 Combined 160.6mm，
                                     而 IR 的六所口径是它的两倍多）—— 见口径坑 14。
        vol_nordic_cash_value_usdbn  官方月度原件是**欧元**的（D 组同一个 feed 里的
                                     'Main Market Total Equity Trading'），换算不回 IR 那条美元
                                     —— 见下面「未找到」一节，那是口径断点不是抓取难题。

B 组 · nasdaqtrader「marketshare{YY}.xlsx」—— 美股现货深度历史 + 分盘口 + 交易日
    列表页 https://www.nasdaqtrader.com/Trader.aspx?id=MarketShare
    直链   https://www.nasdaqtrader.com/content/marketstatistics/marketshare/{YYYY}/marketshare{YY}.xlsx
    **每一年的那份文件自己就含 2005-09 起的全部历史**（2026 版 250 个月），所以只需要下一份。
    nasdaqtrader.com 是 Nasdaq 自营的美国市场官方数据站（NOM/PHLX/ISE/GEMX/MRX/NTX 的规则、
    费率、统计都发在这里），不是聚合商。

    ⚠ 缺年份不返回 404：2026-08-06 实测 2019/2020/2022/2027 四个年份全部 **HTTP 200**，
    但 302 到 `Trader.aspx?id=http404` 那张 43 KB 的 HTML 错误页。所以判「这年有没有」
    必须看**内容首字节是不是 PK\\x03\\x04**，不能看状态码。当前可下的只有 2024/2025/2026。
    （侦察复核稿里那两个 43,088 B 的 "ms19.xlsx / ms20.xlsx" 就是这张 HTML 错误页。）

C 组 · 发布日证人 —— IR 新闻稿电头
    https://ir.nasdaq.com/news-releases/news-release-details/nasdaq-reports-{month}-{year}-volumes
    季末月（3/6/9/12）改成 …-volumes-and-{Q}q{yy}-statistics
    新闻稿正文一个数字都没有（只写「数据已挂在 IR 网站」），唯一用途就是作发布日的证人。

D 组 · Nasdaq Nordic 交易所公告档案 —— 北欧衍生品的长历史（**只用于回补，不用于当期**）
    检索  https://api.news.eu.nasdaq.com/news/query.action
          （globalGroup=exchangeNotice & globalName=NordicExchangeNotices，JSON、免登录）
    附件  https://attachment.news.eu.nasdaq.com/<不可推导的 hash>
    月报「Statistics report - Derivatives volumes per month January-{Month} {Year}」，
    附件名 `Derivative Volumes per Month {YYMM}.xls`（2016-12 及更早）/
           `Derivative Volumes per Month {YYYY}-{MM}.xlsx`（2017-01 起）。
    **每一册含当年 YTD**，所以每年只要下 12 月那一期就拿到整年。
    2026-08-18 实测：162 册，发布月 2013-02..2026-07 零缺口，展开成数据月 2013-01..2026-07
    也零缺口。附件 URL 里那串 hash 不可推导，必须每次从 JSON 现取（与 A 组的 uuid 同型）。

    ⚠ 这个 feed 的三条实测行为，写死在这里免得下一个人再踩：
      (a) **fromDate / toDate 被静默忽略**，dir=ASC 也无效（传 2019 的日期照样返回今天的公告）；
          start 深翻在 5000..50000 之间截断 ⇒ **不能靠时间窗或分页拿历史**。
      (b) **freeText 是全量相关性检索、不受时间窗限制**，这才是历史唯一的入口。
          检索词是打分不是短语匹配：'Statistics from Nasdaq Nordic Exchange' 命中 0，
          'Derivatives volumes per month' 一次返回 165 条（其中 162 条是要的月报）。
      (c) freeText 结果**硬上限 200 条**（start=200 返回空）。162 < 200，当前够用；
          将来册数超过 200 就必须按年分词检索，那时 `_nordic_derivs_index` 要改。

━━━━━━━━━━━━━━━━━━━━ 发布节奏 ━━━━━━━━━━━━━━━━━━━━

A 组：次月第 2–6 个日历日；**季末月晚几天**（那一期要一起更新第 2 页的季度面板）。
      逐条打开新闻稿实测（不是推算）：
        2026-01→02-04  2026-02→03-05  2026-03→**04-08**  2026-04→05-05
        2026-05→06-03  2026-06→**07-08**  2026-07→08-05
        2019-01→02-04  2019-07→08-06  2019-11→12-02   2016-01→02-04  2016-10→11-03
      ⇒ `build/roster.py` 的 LAG 建议 **(6, 9)**：常规月实测最晚第 6 天，季末月最晚第 8 天，
      季末月留 1 天余量。

B 组：慢得多。官方自述「Monthly Market Activity 约在次月第 10 个工作日可得」，
      2026-08-06 实测 marketshare26.xlsx 的 `Last-Modified` = 2026-07-13（内容到 2026-06）。
      ⇒ **B 组必然比 A 组晚一个多星期**，`latest_month()` 绝不能拿 B 组当判据。
      连带结果：每个月新建的那一行，B 组的 9 列天然为空，下一次跑才补上
      （和 `fetch/cboe.py` 的 RPC 滞后同型，`update()` 因此做「只填空、不覆盖」的回补）。

━━━━━━━━━━━━━━━ 每一列的确切口径 ━━━━━━━━━━━━━━━

series/ndaq.csv（月度）

  A 组（IR PDF 第 1 页，口径一律是**当月合计**，不是日均。段落标题里没有 "average daily" 字样）：
    vol_us_options_mmcontracts     美股期权**当月总成交合约数**，单位百万张。
                                   六家 Nasdaq 期权所（NOM / PHLX / ISE / GEMX / MRX / NTX）合计，
                                   **含指数期权**（PDF 脚注 1 原文："U.S. options revenue capture
                                   includes index options for all periods"，capture 与 volume 同口径）。
                                   ⚠ 这是**双边还是单边**：交易所口径的 contracts traded 是
                                   单边计数（一手成交记 1 张），与 Cboe / CME 的合约数口径一致。
    vol_nordic_derivs_mmcontracts  北欧（Nasdaq Nordic + Baltic）期权与期货**当月总合约数**，
                                   单位百万张，官方只印 1 位小数。PDF 标题写的是 "European"，
                                   但覆盖面只有北欧+波罗的海，**不是泛欧** —— 见下面口径坑 1。
    vol_us_cash_matched_mnsh       美股 on-exchange **matched（撮合）当月总股数**，单位百万股。
                                   = Nasdaq + NTX + PSX 三个盘口撮合量之和（已用 B 组独立核对，
                                   18 个重叠月 |差| ≤ 0.011%）。**不含**内化与 TRF 报盘。
    vol_nordic_cash_value_usdbn    北欧现货**当月成交金额**，单位**美元**十亿（不是欧元！）。
                                   官方印 1 位小数。scope 见口径坑 1。

  B 组（nasdaqtrader marketshare{YY}.xlsx，sheet "US Equities"，口径同样是**当月合计**）：
    vol_us_cash_consolidated_sh      全美股 consolidated volume，单位**股**（个位，不是百万股）。
                                     这是份额的分母，也是本仓少有的官方「全行业总量」之一。
    vol_us_cash_matched_nasdaq_sh    The Nasdaq Stock Market 盘口撮合量，单位股。2005-09 起。
    vol_us_cash_matched_ntx_sh       Nasdaq Texas 盘口撮合量，单位股。**2009-01 起**
                                     （更早官方写字符串 'n/a'，见口径坑 3）。
                                     NTX 是 2026 年起的名字，2025 及以前叫 Nasdaq BX，同一根序列。
    vol_us_cash_matched_psx_sh       Nasdaq PSX 盘口撮合量，单位股。**2010-10 起**。
    share_us_cash_matched_nasdaq     官方直接给的份额列，**小数**（0.144249… = 14.42%），
    share_us_cash_matched_ntx        分母是同月 consolidated volume。与 `series/ice.csv` 的
    share_us_cash_matched_psx        `share_nyse_*` 同为小数，横截面页可直接对比。
    share_us_cash_matched_group      **本模块唯一的合成列** = 上面三个官方份额之和，
                                     即「Nasdaq 三所合计 matched 市占」。三家齐了才算，
                                     所以**这条序列从 2010-10 起**，不是 2005-09。
                                     与 IR 季度面板那行 "Nasdaq on-exchange matched market share
                                     as a % of total industry volume" 同口径（14 个季度实测
                                     最大偏差 0.048pp）。
    trading_days_us_equities         当月美股交易日数（官方列）。列名与 `series/ice.csv` 一致，
                                     方便横截面。**本模块不用它做任何除法** ——
                                     ADV = vol_… ÷ trading_days_us_equities 是 build/ 的事。

series/ndaq_q.csv（季度，quarter 形如 2026Q2，与 series/hood_q.csv 同格式）
  IR PDF 第 2 页，14 个季度滚动窗口（当前 2023Q1..2026Q2），官方**只按季发，月度不发**。
    q_us_options_mmcontracts             季度总合约数（百万张），口径同月度那条
    q_share_us_options_matched           Nasdaq 六所在全美期权的 matched 份额（小数）
    q_capture_us_options_usd             每张合约的 estimated revenue capture（美元）
    q_nordic_derivs_mmcontracts          北欧期权+期货季度总合约数（百万张）
    q_capture_nordic_derivs_usd          每张合约 capture（美元）
    q_us_cash_matched_mnsh               美股 matched 季度总股数（百万股）
    q_share_us_cash_matched_of_total     Nasdaq matched ÷ **全行业（场内+场外）**总量（小数）
    q_share_us_onexch_of_total           全行业场内量 ÷ 全行业总量（小数，PDF 脚注 3）
    q_share_us_cash_matched_of_onexch    前两者相除（小数，PDF 脚注 4）—— 官方自己印出来的，
                                         本模块不重算
    q_capture_us_cash_usd_per1ksh        每千股 capture（美元）
    q_nordic_cash_value_usdbn            北欧现货季度成交金额（**美元**十亿）
    q_share_nordic_cash                  北欧现货份额（小数）。**分母是北欧本地市场**，见口径坑 1
    q_capture_nordic_cash_usd_per1kusd   每千美元成交额的 capture（美元）
    q_etp_aum_eop_usdbn                  跟踪 Nasdaq 指数的 ETP **期末** AUM（美元十亿）
    q_etp_aum_avg_usdbn                  同上，**季度均值**
    q_index_linked_derivs_mmcontracts    跟踪 Nasdaq 指数的期货+期货期权+指数期权成交量（百万张）。
                                         ⚠ **这不是 Nasdaq 自己撮合的量**（PDF 脚注 6），
                                         主要是 CME 的 NQ / MNQ 在 CME 成交、Nasdaq 只收授权费。
                                         **不能**和 CME 的 `adv_equity_kcontracts` 摆同一根柱子比
                                         「谁成交大」—— 那是同一批合约被两家分别记账。
    q_listed_cos_us / q_listed_cos_nordic / q_listed_cos_total   期末上市公司数（家）
    q_listed_etps_us                     The Nasdaq Stock Market 上单独挂牌的 ETP 数（只）
    q_listings_total                     上面两类之和（官方自己印的合计行）

━━━━━━━━━━━━ 口径坑（按踩坑概率排序）━━━━━━━━━━━━

 1. **PDF 里写 "European" 的那两条，实际 scope 是北欧+波罗的海，不是泛欧。**
    这是本家最容易被画成误导图的地方，而且**侦察稿给的建议是反的**（它说「比份额、别比绝对值」）。
    算术证明：2026Q2 北欧现货 $277bn、官方份额 74.5% ⇒ 隐含分母 = 277/0.745 = $372bn/季 ≈ $124bn/月；
    而 Cboe 一家同期 `adv_eu_equities_adnv_eurbn` = 14.95 ⇒ ≈ EUR 329bn/月 ≈ $382bn/月。
    **Cboe 一家的欧洲月成交额就是 Nasdaq 整个「欧洲市场」分母的 3 倍。**
    Nasdaq 的 74.5% 是北欧本地份额（NDAQ 2026Q1 8-K EX-99.1 脚注 7 逐字：
    "European cash equities markets include cash equities exchanges of Sweden, Denmark,
    Finland, and Iceland"），Cboe 的是泛欧份额（20–25%）。两个分母是两个宇宙。
    ⇒ **绝对值可比（都换成 USD 或都换成 EUR），份额不可比。** 份额若一定要画，
    必须分 panel 并写明分母不同。列名一律用 `nordic_` 而不是 `eu_`，就是为了让下一个人
    在写 build 时先撞到这条。

 2. **官方给的是当月合计，不是 ADV** —— 与 CME / Cboe / HKEX / ICE 全都不同（那四家发日均）。
    不换算直接进横截面会比同行大 ~20 倍。换算因子在本表里：美股用 `trading_days_us_equities`
    （2005-09 起 250 个月全有且全为数值）。
    **北欧的交易日 IR 一个字都不给**，本模块也没有可靠的无人值守路径去拿（见下面「未找到」一节），
    所以北欧两条只能画月度总量，不要硬转 ADV。

 3. **缺失值是字符串 `'n/a'`，不是 None。** marketshare26.xlsx 实测：
        consolidated / NASDAQ matched / Trading Days   'n/a' = 0
        NTX(BX) matched                                'n/a' = 40   （2005-09..2008-12）
        PSX matched                                    'n/a' = 61   （2005-09..2010-09）
        五个字段同时为数值的月份 = 189 / 250，最早 **2010-10**
    只判 `is None` 会让 `nasdaq + ntx + psx` 在 2010-09 及更早的行**当场抛 TypeError**
    （这是实际踩到的崩溃，不是理论风险）。本模块每个数值格都过 `isinstance(v, (int, float))`。
    ⇒ 与 IR 口径可比的合成序列 `share_us_cash_matched_group` **起点是 2010-10**。

 4. **`BX` → `NTX` 改名（2026 起），列名必须回退。**
    marketshare25.xlsx 的列头是 `BX Matched Volume`，marketshare26.xlsx 是 `NTX Matched Volume`
    （Nasdaq BX 迁址改名 Nasdaq Texas）。同一根序列换了名字，写死列名跨年必炸。
    ⚠ **绝不能拿标题文字做判别依据** —— 同目录的 msoption{YY}.xlsx 标题也从
    "…Nasdaq BX Volumes"（25 版）改成了 "…Nasdaq NTX Volumes"（26 版）。

 5. **必须硬锁 `wb['US Equities']` 这张 sheet。** 同一工作簿还有 NASDAQ / NYSE /
    Amex + Regional / US_ETF 四张分 tape 的表，**五张表头逐字节相同**，`US_ETF` 连
    `max_row` 都同样是 259。取错 sheet 会静默低估约 27% 且永远不报错：
        2026-06  US Equities  consolidated = 491,030,721,181  ← 正确
                 NASDAQ       consolidated = 253,864,108,057  ← 少 27%
    所以本模块按 sheet 名硬取，取不到就抛；**不做任何「找一张像样的表」的兜底**。

 6. **表底混着说明文字行，筛行只认「第 1 列是 datetime」。**
    "US Equities" 的 `max_row` 一路到 259，最后几行第 1 列是字符串 `'2'`/`'3'`/`'4'`
    （不是数字），第 2 列是一大段 Handled Market Share 的定义。按行号或「非空」筛必中招。

 7. **`ir.nasdaq.com` 的 static-file 路径对默认 UA 是「挂住不返回」，不是 403。**
    实测：裸 `urllib`（Python-urllib/3.x UA）落地页 35 s TimeoutError、static-file 32 s
    RemoteDisconnected；同一条 URL 换浏览器 UA 后 0.8–1.9 s 200。
    `curl` 默认 UA 则是干脆的 403 —— 同一个 TLS 栈换 UA 就从 403 变 200，
    **说明拦的是 UA 不是 JA3 指纹，不需要 curl_cffi / nscurl**（与 HOOD 那条 Akamai 记录成因不同）。
    ⇒ 两个后果：(a) UA 是硬要求；(b) **必须设 timeout**，否则 UA 哪天配错，cron 会卡 30 s 而不是快速失败。
    同域**不同路径策略不同**：新闻稿页面裸 UA 也能过，static-file 不能 ——
    「有一条 URL 能通」不能证明「这个域没问题」。`nasdaqtrader.com` 完全不挑 UA。

 8. **PDF 左缘有旋转 90° 的分区侧标**（Equity Derivatives / Cash Equities / Index / Listings）。
    按 y 聚行会把它们混进正文，第 1 页会拼出 "Derivatives January February …" 导致表头识别当场失配。
    ⚠ **不要用 bbox 坐标或「宽 < 高」判**：坐标不是恒定的（第 1 页 x0=30.9，第 2 页 x0≈28.6），
    而「宽<高 且 len>2」会误伤第 2 页脚注里的单词 `all`（宽 4.6 < 高 4.9，长度 3）。
    本模块用 PyMuPDF 的 `line['dir']`（书写方向向量）：正文恒为 (1,0)，侧标恒为 (0,-1)。
    实测第 1 页 213 行正文 / 2 行侧标、第 2 页 455 / 4，**零假阳性零漏判**。

 9. **第 2 页同一段里长标签是短标签的前缀**：
    "Nasdaq on-exchange matched market share as a % of total industry volume" 与
    "Nasdaq on-exchange matched market share" 在同一段里各是一行。
    必须**长的优先**匹配，否则两行会指到同一列。
    另外 "Volume (mm contracts)" / "Market share" / "Revenue capture per contract" 在**不同段**里
    各出现一次，所以查行必须先切段、绝不能全表 grep（与 `fetch/cboe.py` 口径坑 2 同型）。

10. **月度加总与官方季度不完全相等**（月度是四舍五入后印出来的）。实测最大偏差 0.397%
    （2026Q2 北欧现货：月和 278.1 vs 官方 277；期权第二大是 2025Q2 的 +0.104%）。
    ⇒ 自检阈值取 0.6%，取更严会每季度误报一次；也**不要**拿月度和去覆盖官方季度值。

11. **季度 revenue capture 的当期那一格是估计值，会被改。** PDF 脚注 1 原文：
    "Current period revenue capture is estimated until confirmed when final quarterly results
    are issued."。本仓的幂等约定是「已有值永不覆盖」，所以那一格进库之后**不会**被自动刷新
    —— 这是刻意的：自动覆盖会让「重跑字节级不变」这条护栏失效。
    ⇒ (a) 页面上必须把最后一个季度的 capture 标成 estimate；
       (b) 想校正就跑 `python3 fetch/ndaq.py --crosscheck`，它把 CSV 与当期官方文件逐格比
           并打印差异，由人决定要不要手工改。

12. **除 capture 外，Nasdaq 不重述美股月度成交量。** ms25(2025-12 定格) vs ms26(2026-07 版)
    重叠 244 个月、五个字段逐格比，不一致数 = 0。这与 Cboe（官方明说 subject to revisions）和
    HKEX（IPO 暂定数每月上修）性质相反 —— 所以本模块不做全量重述台账。

13. **A 组是百万股/百万张，B 组是个位股** —— 差 1e6。列名后缀 `_mnsh` vs `_sh` 就是为这个，
    横截面页上搞混会差六个数量级。

14. **不要用同目录的 `msoption{YY}.xlsx` 去填美股期权那一列。** 两个致命差别：
    (a) 它只含 NOM + PHLX + NTX/BX **三家**，不含 ISE / GEMX / MRX；
    (b) 它是**三个月的快照**（本月/上月/去年同月），不是历史序列。
    实测 2026-06：msoption26 合计 214,471,257 张，IR 同月 428 百万张 —— 差一半。

15. **发布日证人用新闻稿电头，不要用新闻稿页面的 HTTP Last-Modified。**
    实测 2016 与 2019 的老新闻稿返回的 Last-Modified 都是「今天」（CMS 渲染时间）。
    只有 static-file 的 Last-Modified 有意义，而它会被重传污染。
    本模块的顺序：新闻稿电头（首选，官方对外的正式断言）→ PDF 内嵌 creationDate（兜底，
    2026-07 那期两者同为 2026-08-05，PDF 存盘早于上线 19 分钟）。
    ⚠ 新闻稿 slug **2024 及更早还有别的拼法**，实测 `nasdaq-june-2019-volumes-and-2q19-statistics`
    与 `nasdaq-march-2020-volumes` 都 404 —— 回补老月份的发布日大概率要人工找。
    本模块只为**当期**记发布日，不做历史回补，所以不受影响。

16. **拿季度 8-K 对账时，「北欧现货」两边不是同一个东西 —— 8-K 那个是全市场分母。**
    NDAQ 季度 8-K EX-99.1「Key Drivers Detail」里的
    "Nasdaq Nordic and Nasdaq Baltic securities / Total average daily value of shares traded"
    是**整个北欧市场（含 MTF）**的日均成交额，不是 Nasdaq 自己的；
    IR Monthly Reporting Sheet 那条 `vol_nordic_cash_value_usdbn` 才是 Nasdaq 自己的。
    换算关系（两次实测，用官方北欧衍生品月报里的交易日数）：
        2026Q2  8-K $6.2bn × 60 天 × 官方份额 74.5% = 277.1  vs  IR 277   差 0.05%
        2026Q1  8-K $6.8bn × 62 天 × 官方份额 74.3% = 313.2  vs  IR 312   差 0.4%
    不做这一步换算直接把两个数并排，会得出「IR 少报一半」的错结论。

17. **北欧衍生品：IR 月度表与季度 8-K 在 2026Q1 对不上约 14%，2026Q2 对得上。** 实测：
        2026Q2  IR 13.3mm ÷ 60 天 = 221,667/日   8-K 报 221,789/日   差 0.06%   ✅
        2026Q1  IR 17.6mm ÷ 62 天 = 283,871/日   8-K 报 249,645/日   差 13.7%   ❌
    （交易日 60 / 62 来自 Nasdaq 官方「Derivative Volumes per Month 2026-07.xlsx」，
      2026 年 1-7 月分别是 20/20/22/20/19/21/23。）
    差额集中在 **2026-03 那一格**（IR 印 7.6mm，2025-03 只有 5.8mm）。原因未查明 ——
    可能是 cleared 与 traded 两种口径的差（北欧月报同时发这两套数），也可能是一次性大宗。
    ⇒ 画北欧衍生品时不要用 8-K 的 ADV 去反推月度，也不要拿这条序列做精确的 y/y；
      要用就用 IR 这一条自己的时间序列（内部一致：月度求和 vs IR 自己的季度面板 ±0.000%）。

18. **两个官方源会互相打架，而且错的那个是 nasdaqtrader。** 2026-07 实测：
        IR PDF          U.S. matched equity volume        56,161 百万股
        nasdaqtrader    NASDAQ+NTX+PSX 三盘口之和          55,998.334 百万股   差 -0.2905%
    用完全独立的第三方（Cboe 日频 tape CSV，22 个交易日逐日求和）裁决：
        Q 54,745,234,604 + B 1,039,363,729 + X 376,016,882 = 56,160.615 百万股
    ⇒ IR 与 Cboe 差 +0.0007%（纯四舍五入，**IR 是对的**）；nasdaqtrader 差 -0.2890%。
    缺量是**全市场**的、不是只少报 Nasdaq 自己：nasdaqtrader 的 consolidated
    382,591,323,502.862 vs Cboe 383,603,195,026 = -0.2638%（缺约 1,012 百万股）。
    正因分子分母同时少算，**市占率百分比完全正常**（Nasdaq Q 占比：Cboe 14.2713%
    vs nasdaqtrader 14.2678%），所以 nasdaqtrader 自己的任何内部校验都不会报警，
    错值已扩散到 marketshare26.xlsx / mscompar26.xlsx / Trader.aspx 网页至少三处。
    已排除：口径变更（2026-01..06 偏差只有 0.0003%~0.0023%，只有 7 月破裂）、
    auction 口径（auctions=n 时 Q 掉 28%，量级对不上）、BX→NTX 改名、本模块解析错。
    ⇒ **用户 2026-08-15 决定：不因两个官方源互不吻合而拒发，照官方公布值出网站。**
      因此本模块的 IR↔B 组恒等式改成两档（见 XCHECK_TOL / XCHECK_HARD_TOL）。

19. **北欧衍生品的长历史是「拼出来的」，拼法必须一字不差 —— 而且它有自检。**
    D 组月报按**产品族**分 sheet，IR 那一行是它们的和。2026-08-18 实测的重建口径：
        取每张**权益类** sheet 的 **"Cleared volumes"** 段（不是 "Traded volumes"）的
        **"No. of Contracts"** 列，逐月求和，÷1e6 后**四舍五入到 1 位小数**（IR 就印 1 位）。
    「权益类」= 2026-07 那册的 Stock Products / Index Products / OMXESG Index Futures /
    Mini Index Futures / Custom Basket Forwards and Futures 五张；固定收益的
    Mortgage / Treasury / STIBOR / Repos / Swaps 五类**不算**。
    产品族是逐年长出来的（Mini Index 2020 才有、OMXESG 2023、Custom Basket 2024），
    2016 那册只有 Stock / Index 两张 —— 所以解析器**不许写死 sheet 清单**。
    ⚠ 三个已经踩到的坑：
      (a) **不能按 sheet 首行标题分类。** derivs_2015-12.xls 的 "Stock Products" 首行是
          "Nasdaq Nordic Exchange" 而不是 "Stock Related Products 2015"，按标题筛会把
          整张股票 sheet 丢掉 —— 症状是 2015 年整年低估约 40%（4.6 vs 8.1），
          曲线还很顺眼。本模块因此按 **sheet 名**分类，并且**认不出的 sheet 直接抛**
          （见 `_NORDIC_EQ_SHEETS` / `_NORDIC_FI_SHEETS`）：官方哪天新增一族权益产品，
          宁可整月不更新，也不要静默少算一族。
      (b) **数据格可能是字符串。** OMXESG / Mini Index / Custom Basket 三张的数字格
          openpyxl 读出来是 '33362' 这样的文本，只判 isinstance(v,(int,float)) 会全部漏掉。
      (c) **每个小节的列位置都不一样**（同一册里 No. of Contracts 时而在第 4 列、时而在第 6/8 列），
          必须逐个表头行重新定位，不能整张 sheet 用一个列号。
    自检：`_nordic_derivs_backfill` 每次回补都把重建值与 CSV 里**已有的 IR 值**逐月比，
    19 个重叠月（2025-01..2026-07）2026-08-18 实测**一位小数全等**。差 0.1 只 WARN，
    差更多就抛 —— 那说明官方改了产品族或改了 cleared/traded 的分段。

20. **`vol_us_cash_matched_mnsh` 是拼接列：2025-01 起是 IR 原值，更早是 B 组三盘口和。**
    两段口径**不是同一个数字来源**，但实测是同一个东西：19 个重叠月里 18 个月
    |差| ≤ 0.0101%，只有 2026-07 差 -0.2896%（那是 nasdaqtrader 自己少计，见口径坑 18，
    第三方 Cboe tape 裁定 IR 对）。拼接点在 2024-12 / 2025-01，那一处的口径落差
    因此上界是 0.01% 量级 —— 图上看不见，但**必须在图注里写明**，否则读者会以为
    整条线都是 IR 发的。
    ⚠ 三盘口齐全从 **2010-10** 起（PSX 最后上线），所以拼接段最早只能到 2010-10。
    硬拉到 2005-09（只有 Nasdaq 一家、或 Nasdaq+BX）就是换口径，本模块不做。
    ⚠ 拼接值用 `sum/1e6` **不四舍五入**，与 `build/exchanges_na.py:305` 的
    `__ndaq_venues__` 回落列逐位相同 —— 那条回落列已经在北美页上跑了，两处必须给出同一个数。

━━━━━━━━━━━ 📌 未找到（查过，确实没有）━━━━━━━━━━━

· **IPO 募资金额（capital raised / proceeds）：Nasdaq 官方任何一处都不披露，只披露 IPO 家数。**
  检索路径（全部 2026-08-06 实测）：
    1. IR Monthly Reporting Sheet PDF 两页全文 —— Listings 段只有「公司数 / ETP 数 / 合计」三类计数，
       没有任何金额行；
    2. `ir.nasdaq.com/financials/volume-statistics` 落地页全部链接 —— 「Monthly Volumes」下
       只有 Monthly Metrics 一条，其余是每日量、加拿大、北欧统计、欧洲商品，都没有 IPO 金额；
    3. SEC EDGAR CIK 0001120193 的季度 8-K EX-99.1「Key Drivers Detail」
       （实测 accession 0001193125-26-171829，2026-04-23，1Q26）—— 有
       "Initial public offerings: The Nasdaq Stock Market 63 / Nasdaq operating company IPOs 15 /
       SPACs 48 / Nordic & Baltic 0" 与 "Total new listings"，**全是家数**；
       全文 grep `capital raised` = 0 次、`proceeds` = 1 次（与 IPO 无关的上下文）；
    4. 北欧月度统计稿 PDF 第 2 页 —— "New, this month / New, YTD / Total No of Companies"，
       同样只有计数。
  ⇒ 对照 HKEX 有 `ipo_funds_hkdbn`，Nasdaq 这一格在本仓只能永远留空。上市公司数走
  `q_listed_cos_*`（季度）。IPO **家数**若以后要，去 8-K EX-99.1 取，那是另一条抓取链。

· **北欧现货那条美元列（`vol_nordic_cash_value_usdbn`）回不到 2016：是口径断点，不是抓不到。**
  ⚠ 这一条 2026-08-18 **改写过**。原文说「历史月份不可达」，那半句被证伪了 —— 上面 D 组
  的 freeText 检索同样能把北欧现货的官方月报翻出来（`Main Market Total Equity Trading
  {YYMM}.xls(x)`，实测清单 2022-03..2026-07 连续，每册含 13 个月滚动窗；更早一代是
  2011-09..2014-07 的 `Monthly statistics_Nordics_{Month}_{YYYY}_eng.pdf`）。
  真正的阻断在**币种与 scope**：
    (a) 官方月报给的是 **on-exchange turnover EUR**（精确到分，2026-07 = 76,480,547,244.95），
        IR 那条是**美元**。拿 `series/fx.csv` 的 fx_avg_eurusd 折美元与 IR 逐月比，
        13 个重叠月**系统性低 0.77%~2.50%**；换 fx_eom 更差（-0.18%~-4.21%）。
        再把 Other Lists（First North，光 STO 2026-07 就 16.19bn）加进去又高出一大截
        ⇒ IR 那条美元列的确切 scope 与汇率口径**没能复原**。
    (b) 另有 'Trading Statistics {Month} {Year}' 公告正文（无附件）写 EUR 日均，
        实测覆盖 2016-01..2026-07 的 116/127 个月，但只有 2 位有效数字（"3.327bn"），
        写进真值 CSV 是假精度。
  ⇒ 结论不变但理由换了：**不能拿 EUR 原件首尾相接去接 IR 那条美元线**。要接就得新开一列
  `vol_nordic_cash_value_eurbn`（官方原值、不折汇），而那会连带改 `build/exchanges_eu.py`
  的 NDAQ_MON_COL 与 Exhibit 17 核对表（它现在把这列当 US$bn 直接列示、并用 EURUSD 反折
  欧元与 Cboe 对撞）—— 是一次跨页改动，不在单家回补的边界内。
  北欧现货因此仍只有 IR 的 2025-01 起 19 个月。

· **北欧交易日**：IR 一个字都不给；D 组月报里倒是有（每张 sheet 的 "Trading Days" 列，
  2026 年 1-7 月 = 20/20/22/20/19/21/23）。本模块**没有落库** —— 北欧两条列在页面上画的是
  当月总量、不换 ADV，落一列没人用的交易日只会给下一个人一个「可以除」的错误暗示。

· **Nasdaq Commodities（北欧电力/商品）月报**：`nasdaq.com/solutions/monthly-market-reports-european-commodities`
  上有 2014-01 起的逐月 PDF，但 IR 那四条序列**根本不含它**，且页面 2024-05 之后就没再更新。
  不要拿它去凑「北欧衍生品」。

━━━━━━━━━━━━━━━━ 依赖 ━━━━━━━━━━━━━━━━
PyMuPDF（读 PDF，需要 `line['dir']`）+ openpyxl（读 xlsx）+ **xlrd ≥ 2.0**（读 D 组
2016-12 及更早那批 .xls；xlrd 2.x 只支持 .xls，正好是我们要它做的事）。
xlrd 是**延迟 import** 的：只有真的要回补 2017 年以前的月份才会用到，
装不上也不影响每月的常规更新（那条路只碰 PDF 与 xlsx）。
不依赖 pandas，避免 to_csv 重排既有行的格式（幂等要求：没变的行必须字节级不变）。
"""

import csv
import json
import os
import re
import urllib.parse
import urllib.request
from datetime import datetime
from decimal import Decimal, InvalidOperation

import fitz            # PyMuPDF
import openpyxl

# ── 端点 ─────────────────────────────────────────────────────────────────
IR_LANDING = 'https://ir.nasdaq.com/financials/volume-statistics'
IR_ORIGIN = 'https://ir.nasdaq.com'
PR_BASE = 'https://ir.nasdaq.com/news-releases/news-release-details/'
MS_URL = ('https://www.nasdaqtrader.com/content/marketstatistics/marketshare/'
          '{year}/marketshare{yy}.xlsx')
# D 组（历史回补，见模块 docstring 的 D 组一节与口径坑 19）。
# 参数照抄浏览器实测那一串：fromDate/toDate/dir 全部无效，唯一起作用的是 freeText + limit。
NORDIC_NEWS_URL = (
    'https://api.news.eu.nasdaq.com/news/query.action'
    '?type=handleResponse&showAttachments=true&showCnsSpecific=true&showCompany=true'
    '&countResults=false&freeText={q}&company=&market=&cnscategory='
    '&globalGroup=exchangeNotice&globalName=NordicExchangeNotices'
    '&displayLanguage=en&language=&timeZone=CET&dateMask=yyyy-MM-dd+HH%3Amm%3Ass'
    '&limit=200&start=0&dir=DESC')
NORDIC_DERIVS_QUERY = 'Derivatives volumes per month'
# 官方档案实测可达的最早**数据**月（最早那册是 1302，含 2013-01 与 2013-02）。
NORDIC_DERIVS_START = '2013-01'

# 口径坑 7：static-file 路径对默认 UA 是「挂住 30 秒」而不是快速失败，UA 是硬要求。
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')
_TIMEOUT = 45           # 秒。宁可快速失败，也不要让 cron 卡在一条挂住的连接上。

_MS_SHEET = 'US Equities'          # 口径坑 5：硬锁，取不到就抛，不做兜底
_MS_HEADER_ROW = 2

# ── 表结构 ───────────────────────────────────────────────────────────────
# A 组：(csv 列名, PDF 第 1 页的段落标题前缀)
IR_MONTHLY_SPEC = [
    ('vol_us_options_mmcontracts',
     'U.S. equity options volume (millions of contracts)'),
    ('vol_nordic_derivs_mmcontracts',
     'European options and futures volume (millions of contracts)'),
    ('vol_us_cash_matched_mnsh',
     'U.S. matched equity volume (millions of shares)'),
    ('vol_nordic_cash_value_usdbn',
     'European equity volume (value of shares traded, $ billion)'),
]

# B 组：(csv 列名, xlsx 表头候选。多个 = 官方改过名，按顺序找第一个存在的，见口径坑 4)
MS_SPEC = [
    ('vol_us_cash_consolidated_sh',   ('Consolidated Volume',)),
    ('vol_us_cash_matched_nasdaq_sh', ('NASDAQ Matched Volume',)),
    ('vol_us_cash_matched_ntx_sh',    ('NTX Matched Volume', 'BX Matched Volume')),
    ('vol_us_cash_matched_psx_sh',    ('PSX Matched Volume',)),
    ('share_us_cash_matched_nasdaq',  ('NASDAQ Matched Market Share',)),
    ('share_us_cash_matched_ntx',     ('NTX Matched Market Share',
                                       'BX Matched Market Share')),
    ('share_us_cash_matched_psx',     ('PSX Matched Market Share',)),
    ('trading_days_us_equities',      ('Trading Days',)),
]
MS_GROUP_COL = 'share_us_cash_matched_group'      # 合成列 = 三个官方份额之和

# 口径坑 3：三个盘口的起点不同，早于这些月份官方写字符串 'n/a'。
# 数字来自 2026-08-06 对 marketshare26.xlsx 250 个月的逐格实测，不是推断。
NTX_START = '2009-01'
PSX_START = '2010-10'
GROUP_START = PSX_START            # 三家齐了才有合成份额
# 口径坑 20：`vol_us_cash_matched_mnsh` 的拼接段（B 组三盘口和）最早只能到三家齐的那个月。
VENUE_SPLICE_START = PSX_START
VENUE_SPLICE_COLS = ('vol_us_cash_matched_nasdaq_sh',
                     'vol_us_cash_matched_ntx_sh',
                     'vol_us_cash_matched_psx_sh')

# ── D 组的 sheet 分类（口径坑 19a：**按 sheet 名，不按首行标题**）────────────
# 归一化 = 折叠空白 + 小写；判前缀而不是全等（'Index  Products' 双空格、
# 'Custom Basket Forwards' 在 2024 册被截断成 'Custom Basket Forwards and Futu'）。
_NORDIC_EQ_SHEETS = ('stock products', 'index products', 'omxesg',
                     'mini index futures', 'custom basket')
_NORDIC_FI_SHEETS = ('mortgage', 'treasury', 'stibor', 'nibor', 'cibor',
                     'repos', 'repo', 'swaps')
# 这两张在 2013-01..2026-07 的每一册里都在，缺了就是下错文件/官方改版。
_NORDIC_CORE_SHEETS = ('stock products', 'index products')
_NORDIC_DERIVS_COL = 'vol_nordic_derivs_mmcontracts'
# 重建值 vs CSV 已有 IR 值的两档容差（单位：百万张，官方就印 1 位小数）。
# 实测 19 个重叠月全等 ⇒ 差 1 个末位（0.1）只 WARN，差更多就抛。
NORDIC_XCHECK_WARN = 0.1
NORDIC_XCHECK_HARD = 0.15

# C 组：(csv 列名, PDF 第 2 页的段落标题前缀, 行标签前缀)
# 段内**长标签必须排在短标签前面**（口径坑 9），本模块另外在匹配时再排一次序，双保险。
S_USOPT = 'U.S. equity options quarterly summary'
S_NDDER = 'European options and futures quarterly summary'
S_USCSH = 'U.S. equity quarterly summary'
S_NDCSH = 'European equity quarterly summary'
S_INDEX = 'ETP assets under management (AUM) and futures & options on futures'
S_LIST = 'Number of Listings'

IR_QUARTERLY_SPEC = [
    ('q_us_options_mmcontracts',           S_USOPT, 'Volume (mm contracts)'),
    ('q_share_us_options_matched',         S_USOPT, 'Market share'),
    ('q_capture_us_options_usd',           S_USOPT, 'Revenue capture per contract'),

    ('q_nordic_derivs_mmcontracts',        S_NDDER, 'Volume (mm contracts)'),
    ('q_capture_nordic_derivs_usd',        S_NDDER, 'Revenue capture per contract'),

    ('q_us_cash_matched_mnsh',             S_USCSH, 'Volume (mm matched shares)'),
    ('q_share_us_cash_matched_of_total',   S_USCSH,
     'Nasdaq on-exchange matched market share as a % of total industry volume'),
    ('q_share_us_onexch_of_total',         S_USCSH,
     'Industry on-exchange volume as a % of total industry volume'),
    ('q_share_us_cash_matched_of_onexch',  S_USCSH,
     'Nasdaq on-exchange matched market share'),
    ('q_capture_us_cash_usd_per1ksh',      S_USCSH, 'Revenue capture (per 1,000 shares)'),

    ('q_nordic_cash_value_usdbn',          S_NDCSH, 'Value of shares traded ($ billion)'),
    ('q_share_nordic_cash',                S_NDCSH, 'Market share'),
    ('q_capture_nordic_cash_usd_per1kusd', S_NDCSH, "Revenue capture (per $'000 traded)"),

    ('q_etp_aum_eop_usdbn',                S_INDEX, 'ETP AUM tracking Nasdaq indexes (in billions)'),
    ('q_etp_aum_avg_usdbn',                S_INDEX, 'Average AUM tracking Nasdaq indexes (in billions)'),
    ('q_index_linked_derivs_mmcontracts',  S_INDEX,
     'Futures & options on futures contracts volume tracking Nasdaq indexes'),

    ('q_listed_cos_us',                    S_LIST, 'Companies on The Nasdaq Stock Market'),
    ('q_listed_cos_nordic',                S_LIST,
     'Companies on exchanges that comprise Nasdaq Nordic and Nasdaq Baltic'),
    ('q_listed_cos_total',                 S_LIST, 'Total Listed companies'),
    ('q_listed_etps_us',                   S_LIST,
     'Exchange Traded Products on The Nasdaq Stock Market'),
    ('q_listings_total',                   S_LIST, 'Total Listings'),
]

IR_MONTH_COLS = [c for c, _ in IR_MONTHLY_SPEC]
MS_MONTH_COLS = [c for c, _ in MS_SPEC] + [MS_GROUP_COL]
# CSV 列序：先 IR 四条，再 B 组的量、份额（合成那条紧挨着三个分项，方便肉眼核和）、交易日。
MONTH_COLS = IR_MONTH_COLS + [
    'vol_us_cash_consolidated_sh', 'vol_us_cash_matched_nasdaq_sh',
    'vol_us_cash_matched_ntx_sh', 'vol_us_cash_matched_psx_sh',
    'share_us_cash_matched_nasdaq', 'share_us_cash_matched_ntx',
    'share_us_cash_matched_psx', MS_GROUP_COL, 'trading_days_us_equities',
]
assert sorted(MONTH_COLS) == sorted(IR_MONTH_COLS + MS_MONTH_COLS), 'MONTH_COLS 与规格表对不上'
QUARTER_COLS = [c for c, _s, _l in IR_QUARTERLY_SPEC]

# A 组与 B 组的恒等式容差，**两档**（改档理由见口径坑 18）。
#
# 软档 XCHECK_TOL = 0.15%：IR 印的是整数百万股，纯舍入误差上限约 0.0011%；
# 2026-08-06 对 18 个重叠月实测最大 0.0101%（2025-06）。0.15% 是实测最大值的约 15 倍。
# 超软档 = 两份官方文件对不上，**只 WARN，照官方公布值写入**（用户 2026-08-15 拍板）。
# 原设计是超软档即抛，2026-07 因此整月被拦下 —— 但那次实测证明错的是 nasdaqtrader
# 自己（口径坑 18），拦下来也换不回正确数字，只是让整个模块停更。
#
# 硬档 XCHECK_HARD_TOL = 1%：这一档仍然抛。它要抓的是**本模块自己解析错**那一类：
# 取错 sheet 静默低估约 27%（口径坑 5）、取错列差几倍、A/B 组单位错差 1e6。
# 1% 远在这些量级之下，而两个官方源之间的真实分歧历史最大 0.0101%、本次 0.2896%，
# 仍有 3 倍余量。**不要图省事写成 5%** —— 实测过：19 个月两两错配共 342 组，硬档 5%
# 会放过 74 组（21.6%），1% 只放过 12 组（3.5%），而 1% 对本次这类真实分歧的判定
# 一字不变。5% 唯一的好处是万一 nasdaqtrader 的缺陷继续恶化仍能照发，代价太大。
#
# ⚠ 两档制的已知盲区，必须写明（都是实测，别再乐观描述）：
#  (1) **官方新增第四个撮合盘口** —— 下面的恒等式把 B 组和硬编码成 NASDAQ+NTX+PSX 三家，
#      IR 那条却是 Nasdaq 全部撮合量，多一家就等于差那一家的占比。PSX 自己就是先例：
#      上线头六个月（2010-10 起）占三家之和 1.52% / 2.97% / 3.65% / 4.96% / 4.54% / 4.89%，
#      在 1% 硬档下第二个月就被拦下，在 5% 硬档下要到 2011-04（5.34%）才拦得住。
#      按今天的量级更糟：PSX 现在只占 0.57%~0.67%，新盘口会长期待在盲区里。
#      ⇒ 真加了第四家，要改的是 MS_SPEC 与这里的 parts，不是再放宽阈值。
#  (2) **IR 侧错抓相邻月**：19 个重叠月里有 4 个月的月环比 < 5%（2025-02 2.52%、
#      2025-06 0.24%、2025-12 1.86%、2026-02 2.17%），这种错会降级成 WARN。
#      改成 1% 硬档后只剩 2025-06(0.24%) 一个月仍在盲区。
#  (3) **B 组某个盘口那一列变成 0**（venue 停用/官方某月填 0）：_validate_marketshare 只判
#      `is None`，0.0 能通过。实测 2026-07 漏 PSX = 0.96%、漏 NTX = 2.14%，5% 档下全部静默，
#      1% 档下 NTX 那种拦得住、PSX 那种仍静默。⇒ B 组不是"结构上不可能出错"，
#      只是"取错行/列"不可能（列按表头名查、行按 A 列 datetime 键，见 _parse_marketshare）。
#
# 分母提示：下面 rel = abs(a-b)/a 以 IR 为分母，2026-07 打印 0.2896%；口径坑 18 与 commit
# message 里的 0.2905% 是以 nasdaqtrader 为分母。同一件事，别当成两个数对不上。
XCHECK_TOL = 0.0015
XCHECK_HARD_TOL = 0.01

_MONTH_NAMES = ['January', 'February', 'March', 'April', 'May', 'June', 'July',
                'August', 'September', 'October', 'November', 'December']
_QTR_HDR = re.compile(r'^([1-4])Q(\d{2})$')
_YEAR = re.compile(r'^(20\d{2})$')
# 电头形如 "NEW YORK, Aug. 05, 2026 (GLOBE NEWSWIRE)"。城市一律全大写 —— 城市那一组只收
# 大写字母/空格/点，否则剥完 HTML 标签后前面的 "PDF Version" 会被一起吞进证据文字里。
_DATELINE = re.compile(
    r'\b([A-Z][A-Z .]{2,30}?)\s*,\s*'
    r'((?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\.?\s+\d{1,2},\s+\d{4})'
    r'\s*\(GLOBE NEWSWIRE\)')
_STATIC_FILE = re.compile(
    r'href="(/static-files/[0-9a-fA-F-]{36})"[^>]*>\s*Monthly Metrics\s*<', re.I)
_STATIC_FILE_ANY = re.compile(r'/static-files/[0-9a-fA-F-]{36}')
_CD_MONTH = re.compile(r'filename="[^"]*?([A-Z][a-z]+)\s+(\d{4})[^"]*"')


class NdaqFetchError(RuntimeError):
    """源站结构变化 / 下载失败 / 解析结果不完整。一律炸掉，绝不静默写 NaN。"""


# ── 网络 ─────────────────────────────────────────────────────────────────
def _http_get(url, timeout=_TIMEOUT):
    """返回 (bytes, headers)。**必须带 UA 且必须带 timeout**，理由见口径坑 7。"""
    req = urllib.request.Request(url, headers={
        'User-Agent': _UA,
        'Accept': '*/*',
        'Accept-Language': 'en-US,en;q=0.9',
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.read(), r.headers
    except Exception as e:                                # noqa: BLE001
        raise NdaqFetchError('下载失败 %s: %r' % (url, e)) from e


def _write_bytes(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'wb') as f:
        f.write(data)


def _rawkeep():
    """按路径加载 fetch/rawkeep.py（本模块被 monthly_run 以文件路径加载，裸 import 不可用）。"""
    import importlib.util
    here = os.path.dirname(os.path.abspath(__file__))
    spec = importlib.util.spec_from_file_location(
        'rawkeep', os.path.join(here, 'rawkeep.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _fetch_ir_pdf(cache_dir):
    """下载当期 Monthly Reporting Sheet，返回 (本地路径, Content-Disposition 里的月份或 None)。

    先抓落地页再抽 uuid，而不是写死 uuid：写死能跑，但官方换 uuid 那天会一直下老文件，
    那种坏法**没有任何症状**（文件还在、解析还过、月份就是不动）。
    """
    html_bytes, _hdr = _http_get(IR_LANDING)
    html = html_bytes.decode('utf-8', 'replace')
    _write_bytes(os.path.join(cache_dir, 'ndaq_ir_landing.html'), html_bytes)

    m = _STATIC_FILE.search(html)
    if not m:
        # 退一步：页面上只有一条 static-file 链接时仍可唯一确定；有多条就不猜了。
        cands = sorted(set(_STATIC_FILE_ANY.findall(html)))
        if len(cands) != 1:
            raise NdaqFetchError(
                '落地页上定位不到 "Monthly Metrics" 的 static-file 链接（找到 %d 条候选 %s），'
                '源站可能改版：%s' % (len(cands), cands[:5], IR_LANDING))
        href = cands[0]
    else:
        href = m.group(1)

    pdf, hdr = _http_get(IR_ORIGIN + href)
    if not pdf.startswith(b'%PDF'):
        raise NdaqFetchError('%s 返回的不是 PDF（首字节 %r）' % (href, pdf[:8]))
    path = os.path.join(cache_dir, 'ndaq_monthly_reporting_sheet.pdf')
    _write_bytes(path, pdf)

    cd = hdr.get('Content-Disposition') or ''
    cdm = _CD_MONTH.search(cd)
    cd_month = None
    if cdm:
        try:
            cd_month = '%s-%02d' % (cdm.group(2),
                                    datetime.strptime(cdm.group(1), '%B').month)
        except ValueError:
            cd_month = None          # 文件名里那个词不是月名 → 放弃交叉校验，不猜

    # ── 存证：这份 PDF 是本仓最典型的「原地覆盖且历史取不回」源件 ──────────────
    # 上面那个 ndaq_monthly_reporting_sheet.pdf 是固定名工作副本，下一期直接覆盖它。
    # 存证按 <月>-<sha256 前 12> 另存一份、永不覆盖；理由与紧迫性见 fetch/rawkeep.py
    # 的模块 docstring（要点：series/ndaq.csv 的四个 IR 列现在正好 19 个月 = 当期 PDF
    # 的跨度，2027-01 版一出 2025-01 就掉出窗口，那之后这些月永久不可复核）。
    # cd_month 拿不到时照样存，文件名里落 'nomonth-'。
    _rawkeep().keep('ndaq', pdf, 'pdf', cd_month)
    return path, cd_month


def _fetch_marketshare(cache_dir, year):
    """下载 marketshare{YY}.xlsx。年份不存在时返回 None（不抛）。

    缺年份不返回 404 —— 官方 302 到一张 HTTP 200 的 HTML 错误页（口径坑 5 上方那段）。
    所以判据是内容首字节，不是状态码。
    """
    url = MS_URL.format(year=year, yy=year % 100)
    body, _hdr = _http_get(url)
    if not body.startswith(b'PK\x03\x04'):
        return None                      # 302 到 Trader.aspx?id=http404 的 HTML 页
    path = os.path.join(cache_dir, 'ndaq_marketshare%02d.xlsx' % (year % 100))
    _write_bytes(path, body)
    return path


# ══════════════════════════════════════════════════════════════════════════
# D 组 · Nasdaq Nordic 交易所公告档案 —— 北欧衍生品的历史回补
#
# 只在 series/ndaq.csv 里**真的有洞**时才发请求：没有洞就一次网络都不走。
# 这是「抓取器自己能长回去」而不是一次性脚本的原因 —— 下个月有人删了几行、
# 或者 IR 那份 PDF 某个月漏了，下一次 monthly_run 会自己把洞补上。
# ══════════════════════════════════════════════════════════════════════════
_DERIV_FILE = re.compile(
    r'^Derivative Volumes per Month\s+(?:(20\d{2})-(\d{2})|(\d{2})(\d{2}))\.xlsx?$', re.I)


def _month_seq(lo, hi):
    """['YYYY-MM', …]，闭区间。lo > hi 时返回空表。"""
    out = []
    y, m = int(lo[:4]), int(lo[5:7])
    while '%04d-%02d' % (y, m) <= hi:
        out.append('%04d-%02d' % (y, m))
        m += 1
        if m == 13:
            y, m = y + 1, 1
    return out


def _csv_col(csv_path, col):
    """{'YYYY-MM': float}，只收有值的格。文件/列不存在返回空 dict（不抛）。"""
    try:
        with open(csv_path, newline='', encoding='utf-8') as f:
            rows = list(csv.DictReader(f))
    except OSError:
        return {}
    out = {}
    for r in rows:
        v = (r.get(col) or '').strip()
        if v:
            try:
                out[r['month']] = float(v)
            except ValueError:
                pass
    return out


def _norm(v):
    """一个格 → 归一化文本。数值型的整数不要变成 '20.0'（月份表里的交易日列会撞上）。"""
    if v is None:
        return ''
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return re.sub(r'\s+', ' ', str(v).replace('\xa0', ' ')).strip()


def _nordic_derivs_index(cache_dir):
    """freeText 检索 → {'YYYY-MM'(册次): (文件名, 附件直链)}。

    册次 = 该期覆盖到的最后一个数据月（2016-12 那册含 2016-01..2016-12）。
    附件 URL 里那串 hash **不可推导**，只能每次从 JSON 现取。
    """
    url = NORDIC_NEWS_URL.format(q=urllib.parse.quote(NORDIC_DERIVS_QUERY))
    body, _hdr = _http_get(url)
    _write_bytes(os.path.join(cache_dir, 'ndaq_nordic', 'news_derivs.json'), body)
    txt = body.decode('utf-8', 'replace')
    i, j = txt.find('{'), txt.rfind('}')          # type=handleResponse 可能包 JSONP 壳
    if i < 0 or j < 0:
        raise NdaqFetchError('Nordic news 检索返回的不是 JSON（前 120 字节 %r）' % body[:120])
    try:
        items = json.loads(txt[i:j + 1])['results']['item']
    except (ValueError, KeyError, TypeError) as e:
        raise NdaqFetchError('Nordic news 检索的 JSON 结构变了：%r' % (e,)) from e
    if isinstance(items, dict):
        items = [items]

    out = {}
    for it in items:
        att = it.get('attachment') or []
        if isinstance(att, dict):
            att = [att]
        for a in att:
            m = _DERIV_FILE.match((a.get('fileName') or '').strip())
            if not m:
                continue                       # 同一次检索里混着别的公告，正常
            edition = ('%s-%s' % (m.group(1), m.group(2)) if m.group(1)
                       else '20%s-%s' % (m.group(3), m.group(4)))
            href = a.get('attachmentUrl') or a.get('url')
            if not href:
                raise NdaqFetchError('Nordic 月报 %s 的附件没有直链' % edition)
            out[edition] = ((a.get('fileName') or '').strip(), href)
    if not out:
        raise NdaqFetchError(
            'freeText=%r 一册 "Derivative Volumes per Month" 都没检索到 —— '
            '检索词是打分排序不是短语匹配，官方改了 headline 就会这样（见 D 组注释 b）'
            % NORDIC_DERIVS_QUERY)
    return out


def _fetch_nordic_derivs(cache_dir, edition, filename, href):
    """下载某一册，返回本地路径。已在 cache 里就直接用（这批文件永不改版）。"""
    ext = '.xls' if filename.lower().endswith('.xls') else '.xlsx'
    path = os.path.join(cache_dir, 'ndaq_nordic', 'derivs_%s%s' % (edition, ext))
    if os.path.exists(path) and os.path.getsize(path) > 4096:
        return path
    body, _hdr = _http_get(href)
    head = body[:4]
    ok = head == b'PK\x03\x04' if ext == '.xlsx' else head == b'\xd0\xcf\x11\xe0'
    if not ok:
        raise NdaqFetchError('%s 下回来的不是 %s（首字节 %r）' % (filename, ext, head))
    _write_bytes(path, body)
    return path


def _have_xlrd():
    """本机能不能读 .xls。xlrd **不在 requirements.txt 里**（只服务这一条回补链的
    2017 年以前那一段），所以缺它是「少回补几年」而不是「这家挂了」。"""
    try:
        import xlrd                                     # noqa: F401
    except ImportError:
        return False
    return True


def _nordic_grid(path):
    """一册 → [(sheet 名, 二维 list)]。.xls 走 xlrd、.xlsx 走 openpyxl，下游共用一套解析。"""
    if path.lower().endswith('.xlsx'):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            return [(nm, [[c.value for c in row] for row in wb[nm].iter_rows()])
                    for nm in wb.sheetnames]
        finally:
            wb.close()
    try:
        import xlrd                       # 只有 2017 年以前的册子才需要，故延迟 import
    except ImportError as e:
        raise NdaqFetchError(
            '%s 是 .xls，需要 xlrd（pip install "xlrd>=2.0"）。'
            '2017-01 起官方改发 .xlsx，不装 xlrd 只是回补不到 2017 年以前，'
            '不影响每月常规更新' % os.path.basename(path)) from e
    bk = xlrd.open_workbook(path)
    out = []
    for nm in bk.sheet_names():
        sh = bk.sheet_by_name(nm)
        out.append((nm, [[sh.cell_value(r, c) for c in range(sh.ncols)]
                         for r in range(sh.nrows)]))
    return out


def _row_label(row):
    """行里第一个非空格的文本。这批表的数据不一定从 A 列开始（口径坑 19c 的同源问题）。"""
    for cell in row:
        t = _norm(cell)
        if t:
            return t
    return ''


def _nordic_num(v, where):
    """一个张数格 → float 或 None。**必须容忍字符串**（口径坑 19b）。"""
    if isinstance(v, bool):
        raise NdaqFetchError('%s 是布尔值' % where)
    if isinstance(v, (int, float)):
        return float(v)
    t = _norm(v).replace(',', '')
    if t in ('', '-', '–', '—', 'n/a', 'N/A'):
        return None
    try:
        return float(t)
    except ValueError:
        raise NdaqFetchError('%s 不是数字：%r' % (where, v))


def _parse_nordic_derivs(path, year):
    """一册 → {'YYYY-MM': 百万张}，口径 = 权益类 sheet 的 Cleared volumes / No. of Contracts。

    口径与三个坑见模块 docstring 的口径坑 19。这里只强调一条：
    **认不出的 sheet 直接抛**。官方新增一族权益产品时，静默跳过会让整条历史低估，
    而曲线依旧顺眼 —— 2015 那册就是这么骗过第一版解析器的。
    """
    tot, used = {}, []
    base = os.path.basename(path)
    for name, rows in _nordic_grid(path):
        nm = re.sub(r'\s+', ' ', name).strip().lower()
        if nm.startswith(_NORDIC_FI_SHEETS):
            continue                                    # 固定收益，不进 IR 那一行
        if not nm.startswith(_NORDIC_EQ_SHEETS):
            raise NdaqFetchError(
                '%s 里有认不出的 sheet %r —— 官方多半新增了一族产品。'
                '**不要直接把它加进 _NORDIC_FI_SHEETS 了事**：先确认它是不是权益衍生品，'
                '是就加进 _NORDIC_EQ_SHEETS 并重跑 19 个重叠月的自检（口径坑 19）'
                % (base, name))
        used.append(nm)
        section, col = 'cleared', None       # 2016 及更早没有 Cleared/Traded 分段，默认就是 cleared
        for row in rows:
            lab = _row_label(row)
            low = lab.lower()
            if low in ('cleared volumes', 'traded volumes'):
                section, col = low.split()[0], None
                continue
            if low.startswith('nordic exchange') and 'products' in low:
                col = next((j for j, c in enumerate(row)
                            if _norm(c).lower().replace(' ', '').replace('.', '')
                            == 'noofcontracts'), None)
                if col is None:
                    raise NdaqFetchError(
                        '%s / %s 的表头行里没有 "No. of Contracts"：%r'
                        % (base, name, [_norm(c) for c in row][:10]))
                continue
            if lab not in _MONTH_NAMES or section != 'cleared':
                continue
            if col is None:
                raise NdaqFetchError('%s / %s：%s 行之前没有出现表头行' % (base, name, lab))
            v = _nordic_num(row[col] if col < len(row) else None,
                            '%s / %s / %s' % (base, name, lab))
            if v is None:
                continue
            key = '%04d-%02d' % (year, _MONTH_NAMES.index(lab) + 1)
            tot[key] = tot.get(key, 0.0) + v
    miss = [c for c in _NORDIC_CORE_SHEETS if not any(u.startswith(c) for u in used)]
    if miss:
        raise NdaqFetchError('%s 缺核心 sheet %s（实际拿到 %s）' % (base, miss, used))
    if not tot:
        raise NdaqFetchError('%s 一个月份都没解析出来' % base)
    # ÷1e6 换成百万张、四舍五入到 1 位 —— IR 那一列就印 1 位小数，两段必须同精度才接得上
    return {k: round(v / 1e6, 1) for k, v in sorted(tot.items())}


def _nordic_derivs_backfill(series_dir, cache_dir, upto, ir_months=()):
    """把 `vol_nordic_derivs_mmcontracts` 的历史洞补上，返回 {'YYYY-MM': 百万张}。

    · **CSV 里没有洞 ⇒ 一次网络请求都不发，返回 {}。** 也就是说这条链平时是静默的，
      历史一旦补齐就再也不跑 —— 定期体检不在这里，在 `python3 fetch/ndaq.py --crosscheck`
      的 [4] 段（那一段会把整条列拿去逐年重算）。
    · 一旦要跑，除了缺口那几年，**还会把 IR 覆盖到的那几年一并下下来对账**。
      这是本条链唯一的正确性证明：重建值 vs 同月 IR 印刷值，2026-08-18 实测 19 个重叠月
      一位小数全等。少了这一步，回补就成了「拿一个没人验过的公式去写 144 个月的历史」。
    · 检索/下载失败只 WARN 不抛 —— 历史档案临时取不到，不该拖垮当月的 IR 数据入库。
      解析结构变化与对账超硬阈值**照抛**：那两种是「会写进一个看上去正常的错数」。
    """
    csv_path = os.path.join(series_dir, 'ndaq.csv')
    have = _csv_col(csv_path, _NORDIC_DERIVS_COL)
    want = [m for m in _month_seq(NORDIC_DERIVS_START, upto) if m not in have]
    if not want:
        return {}

    try:
        index = _nordic_derivs_index(cache_dir)
    except NdaqFetchError as e:
        print('[ndaq] WARN 北欧衍生品历史档案检索失败，本次不回补（%d 个月仍留空）：%s'
              % (len(want), e))
        return {}

    # 缺口所在的年份 + IR 有原值的年份（后者是对账年，不写入任何东西，只用来验公式）
    years = {m[:4] for m in want} | {m[:4] for m in ir_months}
    out = {}
    for year in sorted(years):
        # 每册含当年 YTD ⇒ 每年只取该年**最后一期**（往年就是 12 月那期）
        editions = sorted(k for k in index if k[:4] == year)
        if not editions:
            print('[ndaq] WARN 北欧衍生品 %s 年没有任何一册在检索结果里，该年留空' % year)
            continue
        edition = editions[-1]
        try:
            path = _fetch_nordic_derivs(cache_dir, edition, *index[edition])
        except NdaqFetchError as e:
            print('[ndaq] WARN 北欧衍生品 %s 册下载失败，该年留空：%s' % (edition, e))
            continue
        # xlrd 不在 requirements.txt 里（它只服务这一条回补链的 2017 年以前那一段）。
        # 没装就跳过 .xls 那些年、WARN 一句 —— **不许让它拖垮整个 ndaq 的月度更新**。
        if path.lower().endswith('.xls') and not _have_xlrd():
            print('[ndaq] WARN 北欧衍生品 %s 册是 .xls，本机没有 xlrd（pip install "xlrd>=2.0"），'
                  '%s 年留空；2017-01 起官方改发 .xlsx，不影响更近的年份' % (edition, year))
            continue
        book = _parse_nordic_derivs(path, int(year))     # 结构不对就抛，见 docstring

        for mon, val in sorted(book.items()):
            old = have.get(mon)
            if old is not None:
                d = abs(val - old)
                if d > NORDIC_XCHECK_HARD:
                    raise NdaqFetchError(
                        '北欧衍生品 %s：交易所公告重建 %.1f mm 与 CSV 里的 IR 值 %.1f mm '
                        '相差 %.1f，超过硬阈值 %.2f —— 重建口径与 IR 不再是同一个东西了'
                        '（官方多半改了产品族或 cleared/traded 分段），拒绝写入。'
                        '口径坑 19 有完整的重建口径，改之前先把它读完'
                        % (mon, val, old, d, NORDIC_XCHECK_HARD))
                if d > 1e-9:
                    print('[ndaq] WARN 北欧衍生品 %s：重建 %.1f vs IR %.1f（差 %.1f，'
                          '在末位 1 个单位内，按四舍五入分歧处理，不覆盖 IR 原值）'
                          % (mon, val, old, d))
                continue                     # 已有 IR 原值 ⇒ 永不覆盖（幂等约定）
            if mon in want:
                out[mon] = val
    if out:
        print('[ndaq] 北欧衍生品历史回补：%d 个月（%s..%s），来源 = Nasdaq Nordic 交易所公告'
              ' "Derivative Volumes per Month"，口径见 fetch/ndaq.py 口径坑 19'
              % (len(out), min(out), max(out)))
    return out


def _venue_splice(ms, ir_start):
    """`vol_us_cash_matched_mnsh` 的历史段 = B 组三盘口和 ÷ 1e6（口径坑 20）。

    只覆盖 **ir_start 之前**的月份：IR 有原值的月份一律用 IR 的，不拿合成值去竞争。
    不四舍五入 —— 要与 `build/exchanges_na.py` 的 `__ndaq_venues__` 回落列逐位相同。
    """
    out = {}
    for mon in sorted(ms):
        if mon < VENUE_SPLICE_START or mon >= ir_start:
            continue
        parts = [ms[mon].get(c) for c in VENUE_SPLICE_COLS]
        if any(p is None for p in parts):
            continue                       # 三家没齐（不该发生，VENUE_SPLICE_START 之后）
        out[mon] = sum(parts) / 1e6
    return out


# ── PDF 解析 ─────────────────────────────────────────────────────────────
def _page_rows(page):
    """把一页拆成 [(y, [(x_right, text), …]), …]，按 y 升序、行内按 x 升序。

    只保留书写方向为 (1,0) 的水平行 —— 旋转 90° 的分区侧标（Equity Derivatives /
    Cash Equities / Index / Listings）方向是 (0,-1)，混进来会把表头识别整个打乱（口径坑 8）。
    用 dir 而不是 bbox 几何：坐标两页不同，而「宽<高」会误伤脚注里的 `all`。

    x 用**右边缘**：表里的数字是右对齐的，右边缘在同一列上几乎恒定
    （实测值列右缘与表头右缘恒差 1.6–2.7 pt），比左边缘稳得多。
    """
    items = []
    for blk in page.get_text('dict')['blocks']:
        for line in blk.get('lines', []):
            if tuple(round(v, 2) for v in line['dir']) != (1.0, 0.0):
                continue
            txt = ''.join(s['text'] for s in line['spans'])
            txt = re.sub(r'\s+', ' ', txt).strip()
            if txt:
                items.append((line['bbox'][1], line['bbox'][2], txt))
    items.sort(key=lambda t: (t[0], t[1]))

    rows, cur, base = [], [], None
    for y, x1, txt in items:
        if base is not None and abs(y - base) > 3.5:
            rows.append((base, sorted(cur)))
            cur, base = [], y
        elif base is None:
            base = y
        cur.append((x1, txt))
    if cur:
        rows.append((base, sorted(cur)))
    return rows


def _num(txt, where):
    """把 PDF 里印出来的一格转成数值。'$0.13'→0.13，'30.8%'→0.308，'1,018'→1018。

    百分号一律转成**小数**（与 series/ice.csv 的 share_* 列同约定），
    因为份额跨家横截面时小数不需要再记一层「这是百分点还是小数」。
    认不出来就抛 —— 宁可整月不更新，也不要写一个看上去正常的错数。

    ⚠ 除以 100 走 Decimal 而不是二进制浮点：`16.7 / 100` 在 IEEE754 下是
    0.16699999999999998，直接 repr 进 CSV 就成了一串假精度尾巴（官方只印 1 位小数）。
    `Decimal('16.7') / 100` 是精确的十进制移位，得到 0.167。
    """
    t = txt.strip()
    pct = t.endswith('%')
    t = t.rstrip('%').replace('$', '').replace(',', '').replace('−', '-').strip()
    if t in ('', '-', '–', '—', 'n/a', 'N/A'):
        return None
    try:
        v = Decimal(t)
    except InvalidOperation:
        raise NdaqFetchError('%s 这一格不是数字：%r' % (where, txt))
    return float(v / 100 if pct else v)


def _assign(values, grid, where, tol=8.0):
    """把一行里的值按右边缘 x 贴到表头的列上，返回 {列序号: 值文字}。

    不按出现顺序硬贴：中间少一格时顺序法会整体错位一列，而错位后的数看上去完全正常。
    贴不上（离最近的表头列超过 tol）就抛。
    """
    out = {}
    for x1, txt in values:
        best = min(range(len(grid)), key=lambda i: abs(grid[i] - x1))
        if abs(grid[best] - x1) > tol:
            raise NdaqFetchError(
                '%s：值 %r 的右边缘 x=%.1f 贴不到任何表头列（最近的差 %.1f pt），'
                '官方多半改版了' % (where, txt, x1, abs(grid[best] - x1)))
        if best in out:
            raise NdaqFetchError('%s：两个值 %r / %r 贴到了同一列'
                                 % (where, out[best], txt))
        out[best] = txt
    return out


def _parse_page1(page):
    """PDF 第 1 页 → {'YYYY-MM': {csv列名: float}}。

    每个段落的结构是：段标题 / 12 个月名（右对齐） / 当年那一行 / 上一年那一行。
    月份格用月名那一行当列网格（口径坑 8 已把侧标滤掉，网格才是干净的 12 列）。
    """
    rows = _page_rows(page)
    titles = {lab: col for col, lab in IR_MONTHLY_SPEC}
    data, section, grid = {}, None, None
    seen = set()

    for _y, row in rows:
        texts = [t for _x, t in row]
        if len(row) == 1:
            hit = next((lab for lab in titles if texts[0].startswith(lab)), None)
            if hit:
                section, grid = titles[hit], None
                seen.add(section)
            continue
        if section is None:
            continue
        if [t for t in texts if t in _MONTH_NAMES] == texts and len(texts) == 12:
            # 月名一行就是这一段的列网格（必须 12 个月齐，缺一个说明版式变了）
            order = {t: i for i, t in enumerate(_MONTH_NAMES)}
            if [order[t] for t in texts] != list(range(12)):
                raise NdaqFetchError('%s 段的月份表头不是 1-12 月顺序：%s' % (section, texts))
            grid = [x for x, _t in row]
            continue
        ym = _YEAR.match(texts[0])
        if not (ym and grid):
            continue
        year = int(ym.group(1))
        got = _assign(row[1:], grid, '第1页 %s %d 年那一行' % (section, year))
        for idx, txt in got.items():
            v = _num(txt, '第1页 %s %d-%02d' % (section, year, idx + 1))
            if v is None:
                continue
            data.setdefault('%04d-%02d' % (year, idx + 1), {})[section] = v

    missing = [lab for _c, lab in IR_MONTHLY_SPEC if titles[lab] not in seen]
    if missing:
        raise NdaqFetchError('PDF 第 1 页找不到这些段落，官方改版了：%s' % missing)
    if not data:
        raise NdaqFetchError('PDF 第 1 页一个月份都没解析出来')
    return dict(sorted(data.items()))


def _parse_page2(page):
    """PDF 第 2 页 → {'YYYYQn': {csv列名: float}}。

    先用段标题切段（口径坑 9：同名行标签在不同段各出现一次），
    段内按行标签**长的优先**匹配（口径坑 9：长标签是短标签的前缀）。
    """
    rows = _page_rows(page)
    sections = []
    for _c, sec, _l in IR_QUARTERLY_SPEC:
        if sec not in sections:
            sections.append(sec)
    # 段内标签长的排前面
    by_sec = {}
    for col, sec, lab in IR_QUARTERLY_SPEC:
        by_sec.setdefault(sec, []).append((lab, col))
    for sec in by_sec:
        by_sec[sec].sort(key=lambda p: -len(p[0]))

    data, section, grid, quarters = {}, None, None, None
    hit_cols = set()

    for _y, row in rows:
        texts = [t for _x, t in row]
        if len(row) == 1:
            hit = next((s for s in sections if texts[0].startswith(s)), None)
            if hit:
                section, grid, quarters = hit, None, None
            continue
        if section is None:
            continue
        qs = [_QTR_HDR.match(t) for t in texts]
        if all(qs) and len(qs) >= 4:
            grid = [x for x, _t in row]
            quarters = ['%04dQ%s' % (2000 + int(m.group(2)), m.group(1)) for m in qs]
            continue
        if grid is None:
            continue
        lab = texts[0]
        hit = next((c for l, c in by_sec[section] if lab.startswith(l)), None)
        if hit is None:
            continue
        if hit in hit_cols:
            raise NdaqFetchError('第2页 %s 段里 %r 匹配到已经用过的列 %s'
                                 % (section, lab, hit))
        hit_cols.add(hit)
        got = _assign(row[1:], grid, '第2页 %s / %s' % (section, lab))
        for idx, txt in got.items():
            v = _num(txt, '第2页 %s %s' % (hit, quarters[idx]))
            if v is None:
                continue
            data.setdefault(quarters[idx], {})[hit] = v

    miss = [c for c in QUARTER_COLS if c not in hit_cols]
    if miss:
        raise NdaqFetchError('PDF 第 2 页缺这些行，官方改版了：%s' % miss)
    if not data:
        raise NdaqFetchError('PDF 第 2 页一个季度都没解析出来')
    return dict(sorted(data.items()))


def _pdf_creation_date(path):
    """PDF 自己写的 creationDate → ('YYYY-MM-DD', 出处文字)；读不懂就 (None, None)。

    这是**兜底**证人：文件自述的存盘时刻（2026-07 那期 D:20260805140952-04'00'，
    与新闻稿电头 2026-08-05 同日、早 19 分钟）。首选仍是新闻稿电头，理由见口径坑 15。
    """
    with fitz.open(path) as doc:
        raw = (doc.metadata or {}).get('creationDate') or ''
    m = re.match(r"D:(\d{4})(\d{2})(\d{2})", raw)
    if not m:
        return None, None
    return ('%s-%s-%s' % m.groups(),
            '%s 内嵌 PDF creationDate "%s"' % (os.path.basename(path), raw))


def _press_release_date(month):
    """当期新闻稿的 GLOBE NEWSWIRE 电头 → ('YYYY-MM-DD', 出处文字)；找不到就 (None, None)。

    slug 两套（口径坑 15）：常规月 nasdaq-reports-{month}-{year}-volumes，
    季末月再挂 -and-{Q}q{yy}-statistics。2024 及更早还有别的拼法，本模块只管当期，不回补。
    任何一步失败都不抛 —— 发布日拿不到只是页面上少半句话，不该拖垮整月数据入库。
    """
    y, mo = int(month[:4]), int(month[5:7])
    name = _MONTH_NAMES[mo - 1].lower()
    slugs = ['nasdaq-reports-%s-%d-volumes' % (name, y)]
    if mo % 3 == 0:
        slugs.insert(0, 'nasdaq-reports-%s-%d-volumes-and-%dq%02d-statistics'
                     % (name, y, mo // 3, y % 100))
    for slug in slugs:
        try:
            body, _hdr = _http_get(PR_BASE + slug, timeout=30)
        except NdaqFetchError:
            continue
        text = re.sub(r'<[^>]+>', ' ', body.decode('utf-8', 'replace'))
        text = re.sub(r'\s+', ' ', text)
        m = _DATELINE.search(text)
        if not m:
            continue
        for fmt in ('%b. %d, %Y', '%B %d, %Y', '%b %d, %Y'):
            try:
                d = datetime.strptime(m.group(2), fmt)
            except ValueError:
                continue
            return d.strftime('%Y-%m-%d'), (
                '新闻稿 %s 正文电头 "%s, %s (GLOBE NEWSWIRE)"'
                % (slug, m.group(1).strip(), m.group(2)))
        return None, None            # 认得出电头、读不懂日期 → 宁缺勿猜
    return None, None


# ── xlsx 解析 ────────────────────────────────────────────────────────────
def _ms_cell(value, where):
    """marketshare 的一格 → float 或 None。

    口径坑 3：缺失是字符串 `'n/a'`，不是 None。只判 None 会让下游求和抛 TypeError，
    而且是在 2010 年那些行上才抛 —— 建库当天不一定跑得到。
    """
    if value is None:
        return None
    if isinstance(value, bool):
        raise NdaqFetchError('%s 是布尔值 %r' % (where, value))
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s.lower() in ('n/a', 'na', '', '-', '–'):
        return None
    raise NdaqFetchError('%s 既不是数字也不是已知的缺失记号：%r' % (where, value))


def _parse_marketshare(path):
    """marketshare{YY}.xlsx 的 "US Equities" 表 → {'YYYY-MM': {csv列名: float|None}}。"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        if _MS_SHEET not in wb.sheetnames:
            raise NdaqFetchError(
                '%s 里没有 "%s" 这张表（拿到 %r）。**不要退而求其次挑别的表** —— '
                '五张表的表头逐字相同，取错会静默低估约 27%%（口径坑 5）'
                % (os.path.basename(path), _MS_SHEET, wb.sheetnames))
        ws = wb[_MS_SHEET]
        header = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(_MS_HEADER_ROW, c).value
            if v is not None:
                header[re.sub(r'\s+', ' ', str(v)).strip()] = c

        colmap = {}
        for col, cands in MS_SPEC:
            hit = next((h for h in cands if h in header), None)
            if hit is None:
                raise NdaqFetchError(
                    '%s 的 "%s" 表里找不到列 %s（表头实际是 %s）'
                    % (os.path.basename(path), _MS_SHEET, list(cands),
                       sorted(header)[:16]))
            colmap[col] = header[hit]

        out = {}
        for r in range(_MS_HEADER_ROW + 1, ws.max_row + 1):
            d = ws.cell(r, 1).value
            if not isinstance(d, datetime):
                continue              # 口径坑 6：表底混着 '2'/'3'/'4' 开头的说明行
            mon = '%04d-%02d' % (d.year, d.month)
            rec = {col: _ms_cell(ws.cell(r, c).value,
                                 '%s R%dC%d(%s)' % (os.path.basename(path), r, c, col))
                   for col, c in colmap.items()}
            three = [rec['share_us_cash_matched_nasdaq'],
                     rec['share_us_cash_matched_ntx'],
                     rec['share_us_cash_matched_psx']]
            # 合成份额只在三家都有数时算；有一家是 'n/a' 就留空，绝不当 0 处理
            rec[MS_GROUP_COL] = sum(three) if all(v is not None for v in three) else None
            if mon in out:
                raise NdaqFetchError('%s 里 %s 出现了两次' % (os.path.basename(path), mon))
            out[mon] = rec
    finally:
        wb.close()
    if not out:
        raise NdaqFetchError('%s 的 "%s" 表里一行月度数据都没有'
                             % (os.path.basename(path), _MS_SHEET))
    return dict(sorted(out.items()))


# ── 校验 ─────────────────────────────────────────────────────────────────
def _validate_ir_monthly(ir):
    """IR 四条月度序列必须齐 —— 官方是一次发四行，缺一条就是解析错行。"""
    for mon in sorted(ir):
        miss = [c for c in IR_MONTH_COLS if ir[mon].get(c) is None]
        if miss:
            raise NdaqFetchError('IR PDF %s 缺列 %s —— 解析异常，拒绝写入' % (mon, miss))
    return max(ir)


def _validate_marketshare(ms):
    """B 组必须齐，早期例外只放行**实测确认**的那两段 'n/a'（口径坑 3）。"""
    for mon in sorted(ms):
        rec = ms[mon]
        allow = set()
        if mon < NTX_START:
            allow |= {'vol_us_cash_matched_ntx_sh', 'share_us_cash_matched_ntx'}
        if mon < PSX_START:
            allow |= {'vol_us_cash_matched_psx_sh', 'share_us_cash_matched_psx'}
        if mon < GROUP_START:
            allow.add(MS_GROUP_COL)
        miss = [c for c in MS_MONTH_COLS if rec.get(c) is None and c not in allow]
        if miss:
            raise NdaqFetchError(
                'marketshare %s 缺列 %s —— 这些列在该月份本该有数（NTX 自 %s、PSX 自 %s '
                '起才有），解析异常，拒绝写入' % (mon, miss, NTX_START, PSX_START))


def _crosscheck_ir_vs_ms(ir, ms):
    """两份独立文件、两个独立域名的恒等式：IR matched(百万股) ≡ (NASDAQ+NTX+PSX)/1e6。

    这是本模块最强的一道解析正确性证明 —— 它同时验证「行没抓错」和
    「IR 那条 U.S. matched equity volume 的确切口径就是三个盘口撮合量之和」。

    两档（理由见口径坑 18 与 XCHECK_TOL 上方那段）：
      · 超 XCHECK_HARD_TOL(5%)  → 抛。这一档是本模块自己解析错的量级。
      · 超 XCHECK_TOL(0.15%)    → 不抛，计入 flagged 由调用方 WARN，照官方公布值写入。
    返回 (重叠月数, 最大相对偏差, flagged)，flagged 是 [(月份, IR值, B组和, 相对偏差)]。
    """
    worst, n, flagged = 0.0, 0, []
    for mon in sorted(set(ir) & set(ms)):
        a = ir[mon].get('vol_us_cash_matched_mnsh')
        parts = [ms[mon].get('vol_us_cash_matched_nasdaq_sh'),
                 ms[mon].get('vol_us_cash_matched_ntx_sh'),
                 ms[mon].get('vol_us_cash_matched_psx_sh')]
        if a is None or any(p is None for p in parts):
            continue
        b = sum(parts) / 1e6
        rel = abs(a - b) / a
        n += 1
        if rel > worst:
            worst = rel
        if rel > XCHECK_HARD_TOL:
            raise NdaqFetchError(
                '%s：IR 的 matched %.0f 百万股 与 nasdaqtrader 三盘口之和 %.3f 百万股 '
                '相差 %.4f%%，超过硬阈值 %.4f%% —— 这个量级不是两个官方源的分歧，'
                '是本模块取错表/取错列/单位错了（口径坑 5/13），拒绝写入'
                % (mon, a, b, rel * 100, XCHECK_HARD_TOL * 100))
        if rel > XCHECK_TOL:
            flagged.append((mon, a, b, rel))
    if n == 0:
        # A 组窗口 19-24 个月、B 组 250 个月，正常一定有重叠；没有说明其中一份没解析出东西
        raise NdaqFetchError('IR 与 marketshare 没有任何重叠月份，两份文件至少有一份解析失败了')
    return n, worst, flagged


# ── 发布日台账 ───────────────────────────────────────────────────────────
def _source_dates():
    """按路径加载仓库根的 source_dates.py（本模块被 monthly_run 以文件路径加载，裸 import 不可用）。"""
    import importlib.util
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(root, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _record_source_date(series_dir, month, pdf_path):
    """只给**当期**记发布日，且已有记录一律不覆盖。

    一期文件只为它自己的最新月作证：同一份 PDF 里更早的那些月是更早那几期发出来的，
    给它们都盖上这个日期等于凭空发明「2026 年 1 月的数据是 8 月 5 日发的」。
    """
    sd = _source_dates()
    if sd.lookup(series_dir, 'ndaq', month):
        return
    day, evidence = _press_release_date(month)
    if not day:
        day, evidence = _pdf_creation_date(pdf_path)
    if day:
        sd.record(series_dir, 'ndaq', month, day, evidence)


# ── CSV 落盘 ─────────────────────────────────────────────────────────────
def _fmt(v):
    """写回 CSV 的表示。整数值写整数（交易日 21 不写成 21.0，股数不带小数尾巴），
    其余用 repr(float) 的最短往返表示。两条都是确定性的 —— 同一份输入重跑字节级不变。
    """
    if v is None:
        return ''
    f = float(v)
    if f == int(f) and abs(f) < 1e15:
        return str(int(f))
    return repr(f)


def _merge_csv(csv_path, key_name, cols, data, err):
    """把 data 并进 CSV：已有行**只填空、不覆盖**，新行追加，按 key 排序后整表重写。

    返回新增的 key 列表（升序）。「只填空不覆盖」有两个用处：
      · B 组比 A 组晚一个多星期，新月份建行时那 9 列天然为空，下次跑补上；
      · 官方哪天真重述了历史，本模块不会悄悄把新值吞进来（口径坑 11/12）。
    没有任何变化时，未被触碰的单元格是原样字符串搬运，所以文件字节级不变。
    """
    with open(csv_path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    if not rows:
        raise err('%s 是空文件，连表头都没有' % os.path.basename(csv_path))
    header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
    if header[0] != key_name:
        raise err('%s 的第一列应该是 %s，实际是 %r'
                  % (os.path.basename(csv_path), key_name, header[0]))
    idx = {name: i for i, name in enumerate(header)}
    unknown = [c for c in cols if c not in idx]
    if unknown:
        raise err('%s 里没有这些列：%s' % (os.path.basename(csv_path), unknown))

    have = {r[0]: r for r in body}
    added = []
    for key in sorted(data):
        rec = data[key]
        if key in have:
            row = have[key]
            for c in cols:
                if not row[idx[c]].strip() and rec.get(c) is not None:
                    row[idx[c]] = _fmt(rec[c])
            continue
        row = [''] * len(header)
        row[0] = key
        for c in cols:
            row[idx[c]] = _fmt(rec.get(c))
        have[key] = row
        body.append(row)
        added.append(key)

    body.sort(key=lambda r: r[0])
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(body)
    return sorted(added)


# ── 对外接口 ─────────────────────────────────────────────────────────────
def latest_month(cache_dir):
    """官方源当前最新月 'YYYY-MM'。

    以 **A 组（IR PDF 第 1 页）** 为准：uuid 每月原地替换，所以只能解析内容取
    「年份行里最后一个有数的月」。**绝不能用 B 组（nasdaqtrader）当判据** ——
    它比 A 组晚一个多星期（2026-06 的 B 组数据 07-13 才上线，而 07 月的 A 组数据 08-05 就发了）。
    Content-Disposition 里的月份只作交叉校验：对不上说明官方文件名与内容脱节，抛出来让人看一眼。

    抓不到 / 解析不出来一律抛 NdaqFetchError，不返回 None 掩盖故障。
    """
    path, cd_month = _fetch_ir_pdf(cache_dir)
    with fitz.open(path) as doc:
        if doc.page_count < 2:
            raise NdaqFetchError('Monthly Reporting Sheet 只有 %d 页，官方改版了' % doc.page_count)
        ir = _parse_page1(doc[0])
    newest = _validate_ir_monthly(ir)
    if cd_month and cd_month != newest:
        raise NdaqFetchError(
            'PDF 内容最新月是 %s，但 Content-Disposition 的文件名说是 %s —— '
            '官方文件名与内容脱节，人工看一眼再说' % (newest, cd_month))
    return newest


def update(series_dir, cache_dir):
    """把新月份写进 series/ndaq.csv（和季度面板 series/ndaq_q.csv），返回新增月份（升序）。

    幂等保证：
      · 已存在的月份/季度不重复追加；
      · 已经有值的单元格**永不覆盖** —— 官方季度 revenue capture 的当期格是估计值、
        以后会被改（口径坑 11），自动覆盖会让「重跑字节级不变」这条护栏失效。
        要校正就跑 `python3 fetch/ndaq.py --crosscheck` 看差异，再人工决定；
      · 只对既有行里**原本为空**的格子回补 —— B 组比 A 组晚一个多星期，
        不回补的话每个月的 9 个 B 组列会永久留白。回补不计入返回值（它不是新月份）。

    季度 CSV 的新增不计入返回值：monthly_run 的「有没有新数据」是按月判断的，
    季度面板每季只动一次，混进去会让 3/6/9/12 月多报一次「有更新」。
    """
    mcsv = os.path.join(series_dir, 'ndaq.csv')
    qcsv = os.path.join(series_dir, 'ndaq_q.csv')

    pdf_path, cd_month = _fetch_ir_pdf(cache_dir)
    with fitz.open(pdf_path) as doc:
        if doc.page_count < 2:
            raise NdaqFetchError('Monthly Reporting Sheet 只有 %d 页，官方改版了' % doc.page_count)
        ir = _parse_page1(doc[0])
        quarterly = _parse_page2(doc[1])
    newest = _validate_ir_monthly(ir)
    if cd_month and cd_month != newest:
        raise NdaqFetchError(
            'PDF 内容最新月是 %s，但文件名说是 %s —— 官方文件名与内容脱节' % (newest, cd_month))

    # B 组：某一年的那份文件自己就含 2005-09 起的全部历史，所以只需要一份。
    # 年初新一年的文件还没上线时回落到上一年（那份含到上一年 12 月，正是要的）。
    y = int(newest[:4])
    ms_path = _fetch_marketshare(cache_dir, y) or _fetch_marketshare(cache_dir, y - 1)
    if ms_path is None:
        raise NdaqFetchError(
            'nasdaqtrader 的 marketshare%02d.xlsx 与 marketshare%02d.xlsx 都取不到 '
            '（缺年份会 302 到 HTTP 200 的 HTML 错误页，不是 404）' % (y % 100, (y - 1) % 100))
    ms = _parse_marketshare(ms_path)
    _validate_marketshare(ms)
    _, _, _flagged = _crosscheck_ir_vs_ms(ir, ms)
    for _mon, _a, _b, _rel in _flagged:
        # 只 WARN 不抛：两份官方文件对不上时照官方公布值写入（口径坑 18）。
        # 这行必须留在 stdout 里 —— 它是「网站上这个月的 A 组与 B 组本来就对不上」的唯一痕迹。
        print('[ndaq] WARN %s：IR 的 matched %.0f 百万股 与 nasdaqtrader 三盘口之和 %.3f '
              '百万股 相差 %.4f%%（超软阈值 %.2f%%，未超硬阈值 %.2f%%）—— 两个官方源互不吻合，'
              '按口径坑 18 照官方公布值写入，A 组与 B 组这个月不自洽'
              % (_mon, _a, _b, _rel * 100, XCHECK_TOL * 100, XCHECK_HARD_TOL * 100))

    monthly = {}
    for mon in sorted(set(ir) | set(ms)):
        rec = {}
        rec.update(ir.get(mon, {}))
        rec.update(ms.get(mon, {}))
        monthly[mon] = rec

    # ── 历史回补（两条，都只填 IR 那份 PDF 够不到的月份）──────────────────────
    # 顺序要紧：**放在 _crosscheck_ir_vs_ms 之后**。那道恒等式是本模块最强的解析证明，
    # 它比的是「IR 印的 matched」对「B 组三盘口和」；先把合成值塞进 ir/ms 会让它自我对照。
    #
    # ① 美股 matched：2025-01 之前用 B 组三盘口和（口径坑 20）。
    for mon, val in _venue_splice(ms, min(ir)).items():
        monthly.setdefault(mon, {}).setdefault('vol_us_cash_matched_mnsh', val)
    # ② 北欧衍生品：2013-01 起用交易所公告档案（口径坑 19）。CSV 没有洞就不发请求。
    #    传 set(ir) 进去是为了让它顺手把「IR 覆盖到的那几年」也下下来对账 —— 那是这条
    #    重建公式唯一的正确性证明，不能只在缺口那几年上跑。
    for mon, val in _nordic_derivs_backfill(series_dir, cache_dir, newest,
                                            set(ir)).items():
        monthly.setdefault(mon, {}).setdefault(_NORDIC_DERIVS_COL, val)

    for q in sorted(quarterly):
        miss = [c for c in QUARTER_COLS if quarterly[q].get(c) is None]
        if miss:
            raise NdaqFetchError('季度面板 %s 缺列 %s —— 解析异常，拒绝写入' % (q, miss))

    added = _merge_csv(mcsv, 'month', MONTH_COLS, monthly, NdaqFetchError)
    _merge_csv(qcsv, 'quarter', QUARTER_COLS, quarterly, NdaqFetchError)

    # 记发布日放在落盘之后：写盘失败就不该在台账上多出一行说「这个月官方发过了」。
    if newest in monthly:
        _record_source_date(series_dir, newest, pdf_path)
    return added


# ── 自检（不进 monthly_run，供人手工跑）─────────────────────────────────
def _crosscheck(series_dir, cache_dir):
    """跑得起来的自检：打印三组对账，不写任何文件。

    1) IR vs nasdaqtrader 的 matched 恒等式（两份独立文件、两个独立域名）；
    2) 月度求和 vs 官方季度（口径坑 10，阈值 0.6%）；
    3) CSV 已有值 vs 当期官方文件逐格比 —— 专门用来发现季度 revenue capture 的重述
       （口径坑 11：本模块永不覆盖，差异只报不改）。
    """
    pdf_path, _cd = _fetch_ir_pdf(cache_dir)
    with fitz.open(pdf_path) as doc:
        ir, quarterly = _parse_page1(doc[0]), _parse_page2(doc[1])
    y = int(max(ir)[:4])
    ms_path = _fetch_marketshare(cache_dir, y) or _fetch_marketshare(cache_dir, y - 1)
    if ms_path is None:
        raise NdaqFetchError('marketshare%02d/%02d.xlsx 都取不到' % (y % 100, (y - 1) % 100))
    ms = _parse_marketshare(ms_path)

    n, worst, flagged = _crosscheck_ir_vs_ms(ir, ms)
    print('[1] IR vs nasdaqtrader matched 恒等式：%d 个重叠月，最大相对偏差 %.4f%%'
          '（软阈值 %.2f%% 只 WARN，硬阈值 %.2f%% 才抛）'
          % (n, worst * 100, XCHECK_TOL * 100, XCHECK_HARD_TOL * 100))
    for mon, a, b, rel in flagged:
        print('    <== %s 超软阈值：IR %.0f vs B组和 %.3f 百万股，差 %.4f%%（口径坑 18）'
              % (mon, a, b, rel * 100))

    pairs = [('vol_us_options_mmcontracts', 'q_us_options_mmcontracts'),
             ('vol_nordic_derivs_mmcontracts', 'q_nordic_derivs_mmcontracts'),
             ('vol_us_cash_matched_mnsh', 'q_us_cash_matched_mnsh'),
             ('vol_nordic_cash_value_usdbn', 'q_nordic_cash_value_usdbn')]
    print('[2] 月度求和 vs 官方季度（月度是四舍五入印出来的，阈值 0.6%）:')
    for q in sorted(quarterly):
        yy, qq = int(q[:4]), int(q[5])
        months = ['%04d-%02d' % (yy, qq * 3 - 2 + k) for k in range(3)]
        if not all(m in ir for m in months):
            continue
        for mcol, qcol in pairs:
            s = sum(ir[m][mcol] for m in months)
            o = quarterly[q][qcol]
            rel = abs(s - o) / o if o else 0.0
            flag = '  <== 超 0.6%' if rel > 0.006 else ''
            print('    %s %-30s 月和 %10.1f  官方 %10.1f  %+7.3f%%%s'
                  % (q, qcol, s, o, (s - o) / o * 100 if o else 0.0, flag))

    print('[3] CSV 已有值 vs 当期官方文件（只报不改）:')
    nbad = 0
    for name, key, cols, fresh in (('ndaq.csv', 'month', MONTH_COLS,
                                    {m: {**ir.get(m, {}), **ms.get(m, {})}
                                     for m in set(ir) | set(ms)}),
                                   ('ndaq_q.csv', 'quarter', QUARTER_COLS, quarterly)):
        p = os.path.join(series_dir, name)
        with open(p, newline='', encoding='utf-8') as f:
            rows = list(csv.DictReader(f))
        for row in rows:
            k = row[key]
            if k not in fresh:
                continue
            for c in cols:
                old, new = row.get(c, '').strip(), fresh[k].get(c)
                if not old or new is None:
                    continue
                if abs(float(old) - float(new)) > 1e-9 * max(1.0, abs(float(old))):
                    print('    %s %s %s: CSV=%s 官方=%s' % (name, k, c, old, new))
                    nbad += 1
    print('    差异格子数 =', nbad)

    # [4] 北欧衍生品：把 CSV 整条列拿去和交易所公告档案逐年重算一遍。
    # 回补链平时只在「有洞」时才跑，跑完就再也不动 —— 这一节是它唯一的定期体检：
    # 官方哪天新增一族权益产品、或改了 cleared/traded 分段，只有这里看得见。
    print('[4] 北欧衍生品 vs 交易所公告档案（逐年重算，阈值 %.2f mm）:' % NORDIC_XCHECK_HARD)
    have = _csv_col(os.path.join(series_dir, 'ndaq.csv'), _NORDIC_DERIVS_COL)
    index = _nordic_derivs_index(cache_dir)
    ncmp = nbadn = 0
    for yr in sorted({m[:4] for m in have}):
        eds = sorted(k for k in index if k[:4] == yr)
        if not eds:
            print('    %s 年档案里没有册子' % yr)
            continue
        book = _parse_nordic_derivs(
            _fetch_nordic_derivs(cache_dir, eds[-1], *index[eds[-1]]), int(yr))
        for mon, val in sorted(book.items()):
            if mon not in have:
                continue
            ncmp += 1
            d = abs(val - have[mon])
            if d > 1e-9:
                nbadn += 1
                print('    %s 重建 %.1f vs CSV %.1f（差 %.1f）%s'
                      % (mon, val, have[mon], d,
                         '  <== 超硬阈值' if d > NORDIC_XCHECK_HARD else ''))
    print('    比对 %d 个月，不一致 %d 个' % (ncmp, nbadn))

    # [5] 美股 matched 拼接列：接缝两侧各看一眼，别让口径坑 20 的落差悄悄变大。
    print('[5] vol_us_cash_matched_mnsh 拼接接缝（口径坑 20）:')
    mn = _csv_col(os.path.join(series_dir, 'ndaq.csv'), 'vol_us_cash_matched_mnsh')
    ir0 = min(ir) if ir else None
    if ir0 and ir0 in mn:
        prev = max((m for m in mn if m < ir0), default=None)
        print('    IR 段起点 %s = %.3f；拼接段末月 %s = %s'
              % (ir0, mn[ir0], prev, ('%.3f' % mn[prev]) if prev else '—'))
        if prev:
            print('    接缝环比 %+.2f%%（两段口径的落差上界在 0.01%% 量级，'
                  '这里看到的是真实的月环比）' % ((mn[ir0] / mn[prev] - 1) * 100))


if __name__ == '__main__':
    import sys
    _root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _series, _cache = os.path.join(_root, 'series'), os.path.join(_root, 'cache')
    if '--crosscheck' in sys.argv:
        _crosscheck(_series, _cache)
    else:
        print('latest:', latest_month(_cache))
        print('added :', update(_series, _cache))
