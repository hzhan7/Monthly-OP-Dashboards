# -*- coding: utf-8 -*-
"""CME Group (CME) 月度成交量抓取 —— series/cme.csv 的无人值守更新器。

────────────────────────────────────────────────────────────────────────────
源
────────────────────────────────────────────────────────────────────────────
  https://cmegroupinc.gcs-web.com/monthly-volume

  这个地址**本身就是 xlsx**（Content-Type: application/vnd.openxml…sheet），
  不是一个还要再解析出下载链接的落地页。之所以特意写死这个「无扩展名」的路径，
  是因为它是 IR 站上的**稳定别名**：每月发布时后端把它指向新文件，URL 不变，
  真实文件名只在 Content-Disposition 里给出（如
  `cme_group_adv-rpc-oi_trend_Jul26 incl cash markets.xlsx`）。
  所以千万不要去猜带月份的文件名直链 —— 那种链接每月都变，猜错就是 404，
  而这个别名永远指向最新一期。文件名里的月份反过来可以当作「官方最新月」的
  旁证，但真值仍以表内数据为准（见 latest_month 的实现）。

  站点走 Akamai，会拦掉普通 urllib / 裸 curl（不是 JS challenge / 验证码，也不需要
  登录态）—— 通道实测表见 download() 上方那一段。所以本模块仍可无人值守运行，
  前提是走对通道。

  ⚠ **别把判据写窄成「按 TLS 指纹 / JA3 拦」。** 2026-09-02 实测只能排除 UA
  （裸 curl、假 Chrome UA、urllib 三种 UA 都是 403），而能过的 curl_cffi / nscurl
  一次同时改掉了 TLS 指纹、HTTP2 SETTINGS、头集合、头顺序四类变量，**没做单变量
  隔离**，所以只能说判据落在这四类的并集里。写窄了会诱使下一个人只换 TLS 栈、
  或给回落腿塞自定义头，那会重新挂掉。（同理见 fetch/ndaq.py 的 B 组一段。）

────────────────────────────────────────────────────────────────────────────
发布节奏
────────────────────────────────────────────────────────────────────────────
  次月第 1-2 个工作日随月度成交量新闻稿一起更新（2026-07 数据的文件
  Last-Modified 为 2026-08-03，即次月第 1 个工作日）。
  建议调度：每月 3 日起每天跑一次，直到 latest_month() 前进为止。

  **工作簿里没有自述发布日**（找过：'Latest Month Summary' 的表头只写
  "Summary as of July 2026" —— 那是数据月；全部 1,936 条 sharedStrings 里唯一的
  完整日期是 "As of September 19, 2022, daily totals exclude Event Contracts" 这条
  口径脚注；批注、docProps/custom.xml 也都没有）。Cboe 那种 "Updated on August 5, 2026"
  在 CME 这里不存在，别再找了。所以发布日只能取上面那个 Last-Modified，
  再拿工作簿自己的保存时间戳对齐 —— 细节见 _publish_date()。

────────────────────────────────────────────────────────────────────────────
口径坑（这些是踩过才知道的，不要「优化」掉）
────────────────────────────────────────────────────────────────────────────
1. **ADV 与 OI 是两组完全不同的口径，不能互相推导。**
   ADV = 当月日均成交合约数（千张，已按交易日数归一），是「流量」；
   OI  = 月末未平仓合约数（张，绝对数不是千张），是「存量」。
   两者在 xlsx 里放在不同 sheet，单位差 1000 倍。build_cme.py 里
   `oi_total_contracts / 1e6` 而 `adv_total_kcontracts / 1000`，就是这个原因。

2. **月度总成交量必须自己乘交易日数，xlsx 不直接给。**
   官方只给 ADV，总量 = ADV x trading_days。Barclays 那套 day-count 调整
   之所以有意义，就是因为总量口径会被交易日数放大/掩盖真实趋势。
   trading_days 行在 'F&O ADV-RPC' 每个年份块的顶部，**没有自己的日期表头**，
   它借用同块下方 ADV 段的列位置 —— 所以解析时必须先缓存这一行，
   等拿到 ADV 段的列→月份映射之后再回填对齐，不能按固定列号硬编。

3. **资产类别 ADV 要取 'Average Daily Volume by Asset Class' 段，
   不要取 venue 段的分类小计。**
   同一 sheet 的 'F&O Asset Class ADV by Venue' 里也有 "Total Equities"，
   但那是三个 venue 四舍五入后再相加，与官方公布的分类 ADV 差 1-3 千张
   （例：2026-01 Equities 官方 7287，venue 加总 7290）。series/cme.csv
   历史上取的是前者，混用会在图上留下毛刺。

4. **ClearPort 自 2014 年起被并入 'Privately Negotiated'，官方不再单列。**
   xlsx 里 2013 及以前的年份块 venue 段有 'ClearPort' 行，2014 起整行消失。
   所以 adv_clearport_kcontracts 只有 2008-01..2013-12 有值，之后**天然为空**
   —— 这是唯一允许缺失的列，不算解析失败。别把它塞进必填校验，
   也别用 0 去填（0 会被图当成「ClearPort 归零」而不是「不再披露」）。

5. **日期表头里的「日」是脏的，只有年月可信。**
   2026 年块表头是 2026-01-01、2026-02-01…，但 2025 年块是 2025-01-20、
   2025-02-20…，2013 年块是 2013-01-01、2013-02-11…。日号是排版残留，
   解析一律只取 (year, month)。

6. **年份块的行距不是固定的**（2008-2013 有 ClearPort 行，行数多；
   2008 的 rolling 段整段为空）。所以只能靠段落标题文本扫描定位，
   不能假设「每年 N 行」按 stride 跳。

7. **官方会重述前月。** xlsx 里 2026-05 起的数值带三位小数（如 33205.291），
   而更早的月份是整数 —— CME 换过精度口径，且偶尔回改历史月。本模块的
   update() 只**追加**新月份，不改写已有行；若怀疑重述，用 verify() 对账
   后人工决定是否覆盖。

8. **venue 段的数字不能照抄，必须和分类 ADV 合计交叉验。**
   官方 xlsx 自己有错格：2021-04 的整个 venue 列被填成了「滚动 3 个月」的值
   （Globex 19485 + Pit 695 + PN 659 = 20839，而当月分类 ADV 合计只有 16484
   —— 20839 恰好是 Feb21-Apr21 的滚动合计）。series/cme.csv 里 2021-04 的
   三个 venue 列是空的，就是前人踩过这个坑手工剔掉的。
   本模块用 VENUE_TOL_REL 自动拦截：venue 分项之和与分类合计相对偏差超阈值
   就把该月 venue 置空（而不是让整月作废 —— 分类 ADV 和 OI 仍然是好的）。
   注意 2011-05 是另一种情形：官方「Total Venue」格写错了，但分项是对的，
   偏差 0.58%，在容忍范围内，照收。

────────────────────────────────────────────────────────────────────────────
不出声的失败：四道护栏，缺一道就有盲区
────────────────────────────────────────────────────────────────────────────
  本模块所有安静失败都从同一个出口走：解析结果里没有新月份 -> update() 返回 []
  -> NOCHANGE。**连续十天失败和连续十天成功，在日志里长得一模一样**
  （README「第四类：不出声的失败」）。四道护栏分别堵住四种走法，盲区不重叠：

  1. _lab()                     —— 词形容错。官方给行名挂脚注（'Metals*'）时，
                                   在标签进映射之前就把它剥干净。
  2. _parse_adv / _parse_oi     —— 行数对账（照 fetch/msci.py 那道）。
     的孤儿检查                    「认得出的行标签、却没有列可放」= 某个年份块的
                                   日期表头没读出来，整块被吞。当场炸。
  3. parse() 的全表空列检查     —— 某个 CORE 列十八年一起变空 = 行标签改名了。
                                   在 CORE 过滤器把所有行判成半拉子之前炸。
  4. update() 的两条 crosscheck —— 反侧与外部判据，都在 `if not new` 早退之前：
                                   _crosscheck_stored_tail 查「已入库的月份消失了」，
                                   _crosscheck_name_month 拿官方文件名当独立判据
                                   （同 cboe/_crosscheck_report_month、
                                   ice/_crosscheck_workbook_month）。

  四道都刻意 raise 而不是 print：warn 之后状态仍是 NOCHANGE，等于没有护栏。

────────────────────────────────────────────────────────────────────────────
接口
────────────────────────────────────────────────────────────────────────────
  latest_month(cache_dir) -> "YYYY-MM"      官方源当前最新完整月；抓不到抛异常
  update(series_dir, cache_dir) -> [str]    追加新月份，返回新增月份列表（幂等）
  verify(series_dir, cache_dir, n=3)        对账：重算最后 n 个月并逐列比对
"""
from __future__ import annotations

import datetime as _dt
import email.utils
import os
import re
import subprocess
import tempfile
import urllib.request
import zipfile
from zoneinfo import ZoneInfo

import pandas as pd

# ── 源与落盘 ────────────────────────────────────────────────────────────────
SRC_URL = 'https://cmegroupinc.gcs-web.com/monthly-volume'
CACHE_NAME = 'cme_monthly_volume.xlsx'

# 常规桌面 UA。**只给通道 3（urllib）用，别拿它去覆盖 curl_cffi / nscurl 的 UA。**
#
# 这行原来的注释写着「Akamai 对默认 Python-urllib UA 不友好，换成浏览器 UA 即可」——
# 2026-09-02 起这句已经不成立，而且它把因果说反了。实测（同一台机、同一分钟）：
#   urllib + 这个假 Chrome UA  -> 0.2 秒就 403
#   urllib + 默认 Python-urllib UA -> 挂 31 秒后 RemoteDisconnected（连响应都不给）
# 两条都进不来，所以换 UA 从来不是解法 —— UA 已被实测排除出判据（连裸 curl 用它
# 自己的 UA 也是 403）。至于判据到底是什么，**不要收窄**：见文件头那段，只能断言
# 落在 TLS 指纹 / HTTP2 指纹 / 头集合 / 头顺序四类的并集里。
# （两种拒法的耗时不同（0.2 s vs 31 s）只说明拒的方式不同，推不出「谁更刺眼」这种机制，
#  别把它写成因果。）
# 仍然不要给 curl_cffi 手工塞这个 UA：impersonate 是把 UA 连同其它三类配成一整套的，
# 手工塞一个就可能自相矛盾；给 nscurl 塞则会毁掉 Apple 自己那套（两者现在都能过）。
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36')

SHEET_ADV = 'F&O ADV-RPC'
SHEET_OI = 'F&O OI by Asset Class'

# 发布是 CME 的动作，「哪天发的」要按 CME 自己的日历算。实测两条时间证据都落在
# 芝加哥时间下午（2026-08-03 16:47 / 15:53 CDT），此刻按 UTC 取日期碰巧也对，
# 但只要哪期晚发两小时就会跨过 UTC 零点 —— 那时错的是日期本身，而且毫无迹象。
_TZ = ZoneInfo('America/Chicago')

# xlsx 行标签 → series/cme.csv 列名。
# 注意 xlsx 用 'Equities'，CSV 列名用 equity；'Pit' 对应 openoutcry。
_CLASS_MAP = {
    'interest rates': 'rates',
    'equities': 'equity',
    'energy': 'energy',
    'fx': 'fx',
    'agricultural': 'ag',
    'metals': 'metals',
    'total': 'total',
}
_VENUE_MAP = {
    'globex': 'adv_globex_electronic_kcontracts',
    'pit': 'adv_openoutcry_kcontracts',
    'privately negotiated': 'adv_privately_negotiated_kcontracts',
    'clearport': 'adv_clearport_kcontracts',
}

# series/cme.csv 的列顺序（不许改；build_cme.py 按名取用）
COLUMNS = [
    'month',
    'adv_total_kcontracts', 'adv_rates_kcontracts', 'adv_equity_kcontracts',
    'adv_energy_kcontracts', 'adv_ag_kcontracts', 'adv_fx_kcontracts',
    'adv_metals_kcontracts',
    'adv_globex_electronic_kcontracts', 'adv_openoutcry_kcontracts',
    'adv_privately_negotiated_kcontracts', 'adv_clearport_kcontracts',
    'oi_total_contracts', 'oi_rates_contracts', 'oi_equity_contracts',
    'oi_energy_contracts', 'oi_ag_contracts', 'oi_fx_contracts',
    'oi_metals_contracts',
    'trading_days',
]
# 坑 4：ClearPort 2014 起停止披露，是唯一天然允许为空的列。
OPTIONAL = {'adv_clearport_kcontracts'}

VENUE_COLS = [
    'adv_globex_electronic_kcontracts', 'adv_openoutcry_kcontracts',
    'adv_privately_negotiated_kcontracts', 'adv_clearport_kcontracts',
]
# CORE = 决定「这个月到底有没有数」的列。venue 单独判（见坑 8）。
CORE_REQUIRED = [c for c in COLUMNS
                 if c != 'month' and c not in OPTIONAL and c not in VENUE_COLS]
REQUIRED = [c for c in COLUMNS if c != 'month' and c not in OPTIONAL]

# 坑 8 的阈值：venue 三/四项之和 vs 分类 ADV 合计的相对偏差容忍度。
# 实测全 223 个月里，120 个月完全相等，次差是 2011-05 的 0.58%（官方
# 「Total Venue」格自己写错、但分项是对的，series/cme.csv 历史上采纳了分项），
# 最差是 2021-04 的 26.4%（整列被污染，必须丢弃）。2% 卡在两者之间。
VENUE_TOL_REL = 0.02

# 「已入库的月份不许从解析结果里消失」这道哨兵只查**表尾** N 个月（见
# _crosscheck_stored_tail）。取 6 是照抄 fetch/msci.py 的同款哨兵：源是月度、
# 每月只多一行，6 个月足够覆盖任何一次结构变动的波及范围；而只查表尾、不查全部
# 223 个月，是给「CME 哪天把 2008-2010 那几个年份块裁掉」留的余地 —— 那是官方
# 的正当动作，不该把一个健康的抓取器变成天天 FAIL。
SENTINEL_TAIL_MONTHS = 6

# 官方文件名自报的数据月，允许比表内最新完整月领先几期（见 _crosscheck_name_month）。
# 这个数不是随手拍的，是从本模块 docstring 记的节奏推出来的：
#   · 1 期 —— parse() 的 docstring 写着「OI 偶尔比 ADV 晚一天上线」，那几天文件
#     已经是新一期（Jul26）而我们能收全的还是上一个月（2026-06），领先 1 期是**正常**；
#     同一格余量也顺手兜住「CME 哪天改成按发布月命名」这种纯改名（那会让每个月
#     都恒定领先 1 期，一个纯粹的化妆品变化不该让抓取器停摆）。
#   · +1 期 headroom —— 上面两件事可能同时发生（按发布月命名 + 当月 OI 晚到），
#     那就是领先 2 期，仍然健康。所以容忍到 2，第 3 期才开口。
# 代价说清楚：这道判据因此只会「迟一到两个月」才发现漏月，它是**次要**护栏；
# 当场就能发现的是 _parse_adv/_parse_oi 的行数对账和 parse() 的全表空列检查。
NAME_MONTH_LEAD_TOL = 2


class FetchError(RuntimeError):
    """源不可达 / 结构变了 / 数据不完整 —— 一律显式抛，绝不静默写 NaN。"""


# ── 下载：三条通道 ──────────────────────────────────────────────────────────
# 2026-09-02 起源站开始拦 urllib（此前一直直取得到）。当天实测的通道表：
#
#   通道                         结果                              耗时
#   urllib + 假 Chrome UA        403                               0.2s
#   urllib + 默认 UA             RemoteDisconnected（连接被掐）     31s
#   裸 curl / curl + 假 UA       403                               —
#   curl_cffi(impersonate=chrome) 200 / 2,225,967B / PK 魔数        2.3s
#   nscurl（不改 UA）             200 / 2,225,967B / PK 魔数        12.9s
#
# 两条能过的通道，字节数完全一致，所以确认是**传输层被拦**，不是源站变了。
#
# 顺序为什么是 curl_cffi -> nscurl -> urllib：
#   1. curl_cffi 打头，理由有三：快 5.6 倍（2.3s vs 12.9s，无人值守链路里 41 个源
#      逐个串起来，这个差不是零）；`r.headers` 直接就是响应头字典，而 nscurl 只能
#      把头 dump 成文本再由我们解析（多一层会出错的手工活，见 _nscurl_headers）；
#      以及全仓既有范式就是它打头（fetch/hood.py:194 的 fetch_bytes 同序）。
#   2. nscurl 垫第二，正是因为它零第三方依赖 —— 那是**兜底**该有的性质，不是打头
#      该有的性质（同 hood.py 对 _via_nscurl 的定位）。curl_cffi 是 pip 包，
#      哪天升 Python 装不上、或 venv 换了，它就没了；nscurl 是 macOS 自带，跑不掉。
#      它用延迟 import，所以没装 curl_cffi 也能直接走到这条。
#   3. urllib 收尾，纯粹是「墙哪天撤了就还能用零依赖路径」的候补 —— 它今天必挂，
#      但挂得很便宜（0.2s）。放最后而不是最前，是因为放最前每一次跑都要先吐一条
#      403 当开场白，把真正的失败原因埋进噪声里。
#
# ⚠ nscurl 的一个坑（实测，不是推测）：**它对 404 也 exit 0**，并且把那张 404 的
# HTML 原样写进 -o 指定的文件。所以 `subprocess.run(check=True)` 根本不构成成功
# 判据（hood.py 那份就只判了 check=True + 非空）。这里必须从 -D dump 出来的状态行
# 里读 HTTP 状态码，否则一张拦截页会被当成 xlsx 收下 —— 那正是 PK 魔数判据存在的
# 理由，而多一道状态码判据能让报错说得出「是被拦了」而不是只说「开头不是 PK」。


def _norm_headers(pairs):
    """响应头 -> 全小写键的 dict。三条通道各自的 header 容器大小写规矩不一样
    （nscurl dump 出来的是 `Last-Modified`、`content-security-policy` 混排），
    统一压平，免得某天某条通道悄悄取不到 Last-Modified 而我们毫无察觉。"""
    return {str(k).lower(): str(v) for k, v in pairs}


def _meta(status, headers, via):
    h = _norm_headers(headers)
    return {
        'status': status,
        'via': via,
        'ctype': h.get('content-type', '').lower(),
        'disp': h.get('content-disposition', ''),
        'lastmod': h.get('last-modified', ''),
    }


def _via_curl_cffi(url):
    """通道 1：curl_cffi 用真 Chrome 的 TLS/HTTP2 指纹发包，Akamai 放行。

    只加 Accept-Language，**不加 User-Agent** —— impersonate='chrome' 已经把 UA
    连同指纹配成一套了，手工再塞一个就可能自相矛盾（见 _UA 旁注）。
    延迟 import：没装 curl_cffi 也要能落到通道 2。
    """
    from curl_cffi import requests as cr
    r = cr.get(url, impersonate='chrome', timeout=120,
               headers={'Accept-Language': 'en-US,en;q=0.9'})
    r.raise_for_status()
    return r.content, _meta(r.status_code, r.headers.items(), 'curl_cffi')


def _nscurl_headers(text):
    """解析 nscurl -D 的 dump -> (状态码, [(k, v)])。

    只认**最后一个**响应块：nscurl 默认跟随重定向（-L 是它的默认行为），跳一次就
    dump 两段头，取第一段会拿到 301 的 Last-Modified（如果有的话），那就是给发布日
    喂了一个来自错误响应的时间戳。所以每见到一行 'HTTP/' 就把已收集的清空重来。
    """
    status, pairs = None, []
    for line in text.splitlines():
        line = line.rstrip('\r')
        if line.upper().startswith('HTTP/'):
            parts = line.split(None, 2)
            status = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else None
            pairs = []                       # 新的一段响应，前一段整段作废
            continue
        if ':' in line:
            k, v = line.split(':', 1)
            pairs.append((k.strip(), v.strip()))
    return status, pairs


def _via_nscurl(url):
    """通道 2：macOS 自带 nscurl 走 NSURLSession，TLS 指纹是 Apple 的，Akamai 放行。

    零第三方依赖，是 curl_cffi 装不上/被墙时的兜底。用 -D 把响应头落到临时文件，
    这样这条通道也拿得到 Content-Disposition 和 Last-Modified —— 后者是发布日的
    唯一证据（见 _publish_date），拿不到头的通道不配当这个模块的下载通道。
    同样**不改 UA**：默认的 NSURLSession UA 与它自己的指纹是一致的，改了反而露馅。
    """
    fd_b, body = tempfile.mkstemp(suffix='.bin', prefix='cme_')
    os.close(fd_b)
    fd_h, hdr = tempfile.mkstemp(suffix='.hdr', prefix='cme_')
    os.close(fd_h)
    try:
        subprocess.run(['/usr/bin/nscurl', '-D', hdr, '--output', body, url],
                       check=True, capture_output=True, timeout=300)
        with open(hdr, encoding='utf-8', errors='replace') as f:
            status, pairs = _nscurl_headers(f.read())
        with open(body, 'rb') as f:
            blob = f.read()
    finally:
        for p in (body, hdr):
            os.path.exists(p) and os.unlink(p)
    # 见上面那条 ⚠：exit 0 什么都不代表，状态码才算数。
    if status is None:
        raise RuntimeError(f'nscurl 的 -D dump 里读不出状态行（{len(blob)}B 正文）')
    if status != 200:
        raise RuntimeError(f'nscurl 拿到 HTTP {status}（exit 0 不代表成功）')
    if not blob:
        raise RuntimeError('nscurl 返回空文件')
    return blob, _meta(status, pairs, 'nscurl')


def _via_urllib(url):
    """通道 3：零依赖候补。今天必挂（403），留着是等墙哪天撤了。

    timeout 压到 60 而不是原来的 120：这是最后一条通道，它只在前两条都挂了之后才
    跑，那时整个任务已经注定失败，没有理由再多等一分钟；而实测默认 UA 那种「掐连接」
    的挂法本来就要耗掉 31 秒。
    """
    req = urllib.request.Request(url, headers={
        'User-Agent': _UA,
        'Accept': ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,'
                   'application/vnd.ms-excel,*/*'),
    })
    with urllib.request.urlopen(req, timeout=60) as r:
        return r.read(), _meta(r.status, r.headers.items(), 'urllib')


_CHANNELS = (_via_curl_cffi, _via_nscurl, _via_urllib)


def _fetch_xlsx(url):
    """逐条降级取 xlsx，返回 (blob, meta)。全挂就抛一条把所有原因串起来的 FetchError。

    「拿到了东西但不是 xlsx」分两种，处置完全不同 —— 这就是任务书要的
    「区分被拦与源站换了格式」：

      · 像 HTML（ctype 含 html，或正文以 '<' 开头）-> **拦截页**。拦截是冲着
        「你这条通道的指纹」来的，换一条指纹再试完全可能过，所以记下原因、继续降级。
      · 不像 HTML（比如 \\xd0\\xcf\\x11\\xe0 的老 .xls、或一片纯文本 CSV）-> **源站
        换了格式**。这跟指纹无关，换通道只会把同一份新格式文件再下三遍，
        然后吐一条「三条通道全挂」的假象。所以当场抛，并且把话说明白。

    判据仍以 PK 魔数为准（xlsx 就是 zip），比看 Content-Type / 状态码硬 ——
    拦截页也可以带一个 200 和一个漂亮的 ctype。
    """
    errs = []
    for fn in _CHANNELS:
        try:
            blob, meta = fn(url)
        except Exception as e:                 # noqa: BLE001 —— 通道失败要继续试下一条
            errs.append(f'{fn.__name__}: {type(e).__name__}: {e}')
            continue

        if blob.startswith(b'PK\x03\x04'):
            if fn is not _CHANNELS[0]:
                # 主通道死了但兜底活着 —— 数据没事，可这是「下一次可能就全挂」的
                # 预警。不喊的话，等 nscurl 也挂的那天，日志里连一句铺垫都没有。
                print(f'[cme] ⚠ 主通道 {_CHANNELS[0].__name__} 没成，'
                      f'降级到 {meta["via"]} 才拿到 xlsx。前序失败：{errs}')
            return blob, meta

        head = blob[:200].decode('utf-8', 'replace')
        looks_html = ('html' in meta['ctype'] or blob.lstrip()[:1] == b'<')
        if looks_html:
            errs.append(f'{fn.__name__}: HTTP {meta["status"]} 但正文是 HTML 拦截页'
                        f'（{len(blob)}B, Content-Type={meta["ctype"]!r}）: {head!r}')
            continue
        raise FetchError(
            f'{fn.__name__} 通道 HTTP {meta["status"]} 正常返回了 {len(blob)} 字节，'
            f'但开头不是 xlsx 的 PK\\x03\\x04 魔数，也不像 HTML 拦截页'
            f'（Content-Type={meta["ctype"]!r}，开头 200 字节: {head!r}）。'
            '这不是被墙 —— 被墙给的是 HTML 或干脆不给。最可能是**源站换了文件格式**'
            '（老式 .xls 的魔数是 \\xd0\\xcf\\x11\\xe0，CSV 则是纯文本）。'
            '换通道对这种情况没有意义，所以当场停：请人工打开 SRC_URL 看一眼，'
            '确认后改 parse() 的读法，别让它继续按 xlsx 解。')

    raise FetchError(
        f'CME IR xlsx 下载失败，{len(_CHANNELS)} 条通道全挂: {url}\n  '
        + '\n  '.join(errs)
        + '\n  （源站走 Akamai，urllib 必 403；curl_cffi 与 nscurl 都进不来时'
          '多半是墙升级了，见 download() 上方通道表）')


def download(cache_dir, max_age_hours=6.0):
    """把最新 xlsx 抓到 cache_dir，返回 (路径, 官方文件名)。

    带 max_age_hours 是为了让同一次调度里 latest_month() + update() + verify()
    共用一份文件 —— 否则三次下载可能跨越官方发布瞬间，拿到不一致的两份数据。
    """
    os.makedirs(cache_dir, exist_ok=True)
    path = os.path.join(cache_dir, CACHE_NAME)
    meta = path + '.name'
    lm_meta = path + '.lastmod'

    # 缓存命中还额外要求 .lastmod 在：它是发布日的证据（见 _publish_date），
    # 和 xlsx 本体来自同一个响应。缺了它就重下一次 —— 否则命中缓存的那几次跑
    # 取不到发布日，页面抬头那半句会时有时无，而这种「有时对」最难被发现。
    if os.path.exists(path) and max_age_hours is not None and os.path.exists(lm_meta):
        age = (_dt.datetime.now().timestamp() - os.path.getmtime(path)) / 3600.0
        if age < max_age_hours:
            name = ''
            if os.path.exists(meta):
                with open(meta, encoding='utf-8') as f:
                    name = f.read().strip()
            return path, name

    # PK 魔数判据与「拦截页 vs 换格式」的分流都在 _fetch_xlsx 里，走到这里
    # blob 已经确定是 xlsx 了。
    # 注意别叫它 meta —— 上面那个 meta 是 .name 旁文件的路径，同名会把它冲掉。
    blob, resp = _fetch_xlsx(SRC_URL)
    lastmod = resp['lastmod']

    name = ''
    m = re.search(r'filename="?([^";]+)"?', resp['disp'])
    if m:
        name = m.group(1).strip()

    # Last-Modified 是发布日的唯一证据（见 _publish_date）。缺了它数据照样是好的，
    # 所以这里不抛 —— 抛了就是为了一个台账字段把整月数据挡在门外，不成比例，
    # 而且本模块早就为「发布日未记」设计了正当路径（_record_publish_date 会打印原因）。
    # 但**不能不吭声**：那条打印只在「最新月刚好这轮入库」时才走得到，NOCHANGE 的
    # 日子里一个空 .lastmod 是彻底静默的，然后某天发布日就那么缺了半句抬头。
    # 所以在写盘这一刻当场喊 —— 与 _crosscheck_name_month 丢失判据时的处置同款
    # （那里也是 print '⚠ 护栏失效' 而不是抛）。
    if not lastmod:
        print(f'[cme] ⚠ 护栏失效：{resp["via"]} 通道的响应里没有 Last-Modified 头，'
              '这一期的发布日将无从判定（.lastmod 会写成空文件）。'
              '若持续如此，说明这条通道拿不全响应头，应当换通道而不是接受它。')

    with open(path, 'wb') as f:
        f.write(blob)
    with open(meta, 'w', encoding='utf-8') as f:
        f.write(name)
    # 原样存这一行响应头（不是解析后的日期）：日后有人怀疑某个发布日，
    # 要能拿到源头当时说的原话，而不是我们转述的结果。
    with open(lm_meta, 'w', encoding='utf-8') as f:
        f.write(lastmod)
    return path, name


# ── 解析 ────────────────────────────────────────────────────────────────────
def _is_date(v):
    return isinstance(v, _dt.datetime) or isinstance(v, _dt.date)


def _mkey(v):
    """坑 5：日期表头里的「日」是排版残留，只取年月。"""
    return f'{v.year:04d}-{v.month:02d}'


def _lab(v):
    """行标签归一化：压空白 + 转小写 + 右剥脚注记号。

    用全等匹配（原来的 `a.lower()`）挂在一个假设上：官方永远不给行名加脚注。
    这个假设在别家已经被打破过 —— fetch/cboe.py 的 _lab 就是为
    'Futures - ADV (contracts, thousands)*' 写的，MSCI 更是把脚注写成裸文本，
    害得整整一行被静默丢弃（README「第四类：不出声的失败」记着这一天三连）。
    CME 今天一个脚注都没有，正因为如此才要现在剥：等它加了脚注的那个月，
    表现是某个资产类别的列**整列变空**，不是报错。

    右剥 ' *¹²³0123456789' 是安全的：本模块认的十一个标签
    （_CLASS_MAP 七个 + _VENUE_MAP 四个）没有一个以数字或星号结尾。
    唯一被剥没的是年份行 '2026' -> ''，而它本来就不在任何映射里，照样跳过。
    注意剥不掉的形态（如 'Metals (1)'）不是漏网：它进不了映射，
    于是 parse() 的全表空列检查会响 —— 两道是接力，不是重复。
    """
    if v is None:
        return ''
    return re.sub(r'\s+', ' ', str(v)).strip().lower().rstrip(' *¹²³0123456789')


def _num(v):
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '').strip())
    except ValueError:
        return None


def _colmap(row):
    """从表头行取 {列下标: 'YYYY-MM'}。"""
    return {i: _mkey(v) for i, v in enumerate(row) if _is_date(v)}


def _parse_adv(ws):
    """解析 'F&O ADV-RPC'：交易日数 + 分类 ADV + 分 venue ADV。

    坑 6：年份块行距不固定，只能顺序扫段落标题，不能按 stride 跳。
    坑 3：只认 'by Asset Class' 段，'Rolling 3 Month' 段（滚动均值/RPC）必须跳过。
    """
    out = {}
    section = None          # 'class' | 'venue' | None(=跳过)
    cmap = None
    pending_days = None     # 坑 2：Trading Days 行先于表头出现，暂存后回填
    orphans = []            # 长得像数据行、却没能落库的行；循环末尾一起抛

    for row in ws.iter_rows(values_only=True):
        a = str(row[0]).strip() if row[0] is not None else ''
        b = str(row[1]).strip() if isinstance(row[1], str) else ''
        key = _lab(a)

        if key == 'trading days':
            pending_days = row
            continue

        if b:
            low = b.lower()
            if low.startswith('rolling'):
                section, cmap = None, None       # 滚动段：整段丢弃
            elif 'average daily volume by asset class' in low:
                section, cmap = 'class', None
            elif 'average daily volume by venue' in low:
                section, cmap = 'venue', None
            elif 'rate per contract' in low:
                section, cmap = None, None

        if section and cmap is None and _is_date(row[1]):
            cmap = _colmap(row)
            if not cmap:
                raise FetchError(f'{SHEET_ADV}: 段 {section} 的日期表头解析不出月份')
            if section == 'class' and pending_days is not None:
                for i, mk in cmap.items():
                    d = _num(pending_days[i]) if i < len(pending_days) else None
                    if d is not None:
                        out.setdefault(mk, {})['trading_days'] = d
                pending_days = None
            continue

        if not section or not a:
            continue

        # 段落标题已经读到、日期表头却没读到 —— 这一段的每一行都无处安放，
        # 于是被静默丢弃。这不是假想：表头单元格是手工维护的（坑 5 记着
        # 2026-01-01 / 2025-01-20 / 2013-02-11 这种日号乱飘），哪一年新块的表头
        # 被打成文本而不是日期，这里就会整块吞掉；而新块正好在 sheet 最上方，
        # 吞掉的就是当年**全部**月份。原来的表现是 parse() 少产出十几行、
        # latest_month() 安静停在去年 12 月、update() 报 NOCHANGE ——
        # 连续十天失败和连续十天成功，在日志里长得一模一样（README 第四类）。
        # 所以凡是「我们认识这个行标签、却没地方放」的行，一律记下来最后炸。
        if cmap is None:
            if (section == 'class' and key in _CLASS_MAP) or \
                    (section == 'venue' and key in _VENUE_MAP):
                orphans.append(f'{a}（{section} 段）')
            continue

        if section == 'class' and key in _CLASS_MAP:
            col = f'adv_{_CLASS_MAP[key]}_kcontracts'
        elif section == 'venue' and key in _VENUE_MAP:
            col = _VENUE_MAP[key]
        else:
            continue                              # 'Exchange-traded Total'/'Total Venue' 等小计行不入库

        for i, mk in cmap.items():
            v = _num(row[i]) if i < len(row) else None
            if v is not None:
                out.setdefault(mk, {})[col] = v

    # 行数对账（照抄 fetch/msci.py parse() 那道）。实测：现行工作簿 38 个
    # class/venue 表头全部认得出，孤儿数 0；把 2026 年块的表头改成文本，
    # 孤儿立刻变成 10 条。该响的时候响、平时一声不吭，就是这道的意义。
    if orphans:
        raise FetchError(
            f'{SHEET_ADV}: {len(orphans)} 行长得像数据行却没能落库：{orphans[:6]!r}'
            ' —— 多半是某个年份块的日期表头不是日期单元格（被打成了文本），'
            f'整块被吞掉。宁可整次失败也不静默漏月；请对照 {CACHE_NAME} 的'
            ' 段落标题下一行确认表头写法。')
    if not out:
        raise FetchError(f'{SHEET_ADV}: 一行都没解析出来，sheet 结构可能改了')
    return out


def _parse_oi(ws):
    """解析 'F&O OI by Asset Class'：月末未平仓合约数（张，不是千张）。"""
    out = {}
    cmap = None
    orphans = []            # 同 _parse_adv：认得出的行标签却没地方放，末尾一起抛
    for row in ws.iter_rows(values_only=True):
        if _is_date(row[1]):
            cmap = _colmap(row)
            continue
        if row[0] is None:
            continue
        key = _lab(row[0])
        if key not in _CLASS_MAP:
            continue
        # 本 sheet 最新的年份块在最上方，它的日期表头若读不出来，这一块的七行
        # 就落在「还没有任何 cmap」的状态里被安静丢掉 —— 丢的正好是当年全部
        # 月份，而 OI 是 CORE 列，缺了整年的行会被 parse() 的 CORE 过滤器
        # 一并剔除，最后 update() 干干净净报 NOCHANGE。记下来最后炸。
        if cmap is None:
            orphans.append(str(row[0]).strip())
            continue
        col = f'oi_{_CLASS_MAP[key]}_contracts'
        for i, mk in cmap.items():
            v = _num(row[i]) if i < len(row) else None
            if v is not None:
                out.setdefault(mk, {})[col] = v
    if orphans:
        raise FetchError(
            f'{SHEET_OI}: {len(orphans)} 行长得像数据行却没能落库：{orphans[:6]!r}'
            ' —— 首个年份块的日期表头没解析出来（多半被打成了文本），'
            '整块被吞掉。拒绝写入，请人工看一眼该 sheet 最上面那个表头。')
    if not out:
        raise FetchError(f'{SHEET_OI}: 一行都没解析出来，sheet 结构可能改了')
    return out


def parse(xlsx_path):
    """xlsx -> DataFrame（index='YYYY-MM' 字符串，列同 series/cme.csv）。

    只保留 REQUIRED 全部齐备的月份 —— 当年剩余月份在表里是空列，
    OI 偶尔也会比 ADV 晚一天上线，半拉子月份一律不算「有」。
    """
    try:
        import openpyxl
    except ImportError as e:                       # noqa: BLE001
        raise FetchError('需要 openpyxl 才能解析 CME xlsx: pip install openpyxl') from e

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:                         # noqa: BLE001
        # 缓存里躺着半截下载 / 拦截页时，openpyxl 会抛 BadZipFile 之类的底层异常。
        # 统一成 FetchError，调度器只需要 catch 一种。
        raise FetchError(f'打不开 {xlsx_path}（缓存损坏或不是 xlsx）: {e!r}') from e
    for sn in (SHEET_ADV, SHEET_OI):
        if sn not in wb.sheetnames:
            raise FetchError(f'xlsx 缺 sheet {sn!r}；现有: {wb.sheetnames}')

    rec = _parse_adv(wb[SHEET_ADV])
    for mk, d in _parse_oi(wb[SHEET_OI]).items():
        rec.setdefault(mk, {}).update(d)
    wb.close()

    df = pd.DataFrame.from_dict(rec, orient='index')
    for c in COLUMNS[1:]:
        if c not in df.columns:
            df[c] = pd.NA
    df = df[COLUMNS[1:]].sort_index()

    # ── 全表空列检查：CORE 列一个非空值都没有，就炸 ──────────────────────
    # 下面那行 CORE 过滤器是这个模块最安静的一处失败。官方哪天把 'Metals'
    # 写成 'Metals (1)'、把 'Interest Rates' 改成单数，_CLASS_MAP 就再也匹配不上，
    # 对应的 adv_/oi_ 列**全 223 个月一起变空**（xlsx 是一份滚动覆盖全历史的
    # 文件，改模板就是改十八年），过滤器于是把每一行都判成「半拉子月份」丢掉，
    # 而 update() 见到空帧只会 `if not new: return []` —— 一次干净的 NOCHANGE，
    # 天天如此。这正是 README「第四类」的判据：连续失败十天和成功十天长得一样。
    #
    # 检查放在过滤**之前**，是为了让报错说得出是哪一列没了；放在之后只会得到
    # 一句毫无线索的「0 行」。判据故意宽到不能再宽 —— 只要全表还有**一个**
    # 非空值就放行，所以「当年剩余月份是空列」「某月 OI 晚一天上线」这些正常
    # 情形一概不会误伤（实测现行工作簿每个 CORE 列都是 223 个非空值）。
    #
    # 这里刻意 raise 而不是 print：warn 完状态仍是 NOCHANGE，等于没有护栏
    # （同 fetch/cboe.py _crosscheck_report_month 的理由）。代价要说清楚：
    # CME 若真的停发某个资产类别，这里会天天 FAIL 到有人来改 COLUMNS/_CLASS_MAP。
    # 那正是护栏 2「缺列一律失败」要的效果 —— 停发是需要人做决定的事，
    # 不该由抓取器替他悄悄决定成「这列从此为空」。
    empty_cols = [c for c in CORE_REQUIRED if not df[c].notna().any()]
    if empty_cols:
        raise FetchError(
            f'{os.path.basename(xlsx_path)}: CORE 列 {empty_cols} 在整份工作簿里'
            '一个值都没有（不是某个月缺，是十八年一起缺）—— 最可能是官方改了行标签，'
            f'_CLASS_MAP / _VENUE_MAP 匹配不上了（{SHEET_ADV} / {SHEET_OI} 的首列）。'
            '拒绝返回半张表，否则这一列会静默变成空值上线。')

    df = df[df[CORE_REQUIRED].notna().all(axis=1)]

    # 坑 8：venue 段偶尔被官方写错，必须交叉验，不能照抄。
    vsum = df[VENUE_COLS].astype(float).sum(axis=1, min_count=1)
    tot = df['adv_total_kcontracts'].astype(float)
    bad = ((vsum - tot).abs() / tot.abs() > VENUE_TOL_REL) & vsum.notna()
    df.loc[bad, VENUE_COLS] = pd.NA
    df.attrs['venue_quarantined'] = list(df.index[bad])
    return df


# ── 发布日 ──────────────────────────────────────────────────────────────────
def _source_dates():
    """按路径加载仓库根的 source_dates.py（发布日台账）。

    不能裸 import：本模块被 monthly_run 用 spec_from_file_location 加载，
    那时 sys.path 上既没有 fetch/ 也没有仓库根。
    """
    import importlib.util
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(root, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _wb_saved_at(xlsx_path):
    """工作簿自己的保存时间戳 —— docProps/core.xml 的 dcterms:modified，原文与 UTC 时间。

    这是 CME 的人（core.xml 里 lastModifiedBy 那位）按下保存时 Excel 写进文件的，
    随文件走、和我们的下载行为无关，所以能当作独立于 HTTP 头的第二条证据。
    """
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            core = z.read('docProps/core.xml').decode('utf-8', 'replace')
    except Exception:                                          # noqa: BLE001
        return None, None
    m = re.search(r'<dcterms:modified[^>]*>([^<]+)</dcterms:modified>', core)
    if not m:
        return None, None
    raw = m.group(1).strip()
    try:
        d = _dt.datetime.strptime(raw, '%Y-%m-%dT%H:%M:%SZ')
    except ValueError:
        return None, raw
    return d.replace(tzinfo=_dt.timezone.utc), raw


def _publish_date(cache_dir, xlsx_path, name):
    """这一期 xlsx 的发布日 ('YYYY-MM-DD', 出处文字)；判不准返回 (None, 原因)。

    工作簿里没有任何自述发布日（找过哪些地方见模块 docstring 的「发布节奏」节），
    所以退到本模块 docstring 从一开始就当作实测发布日的那个字段：HTTP Last-Modified。

    但单凭一个响应头不敢往页面上印 —— CDN 迁移、重新上传同一份文件都会把它推后，
    而推后之后它看上去和真发布日一模一样。所以要工作簿自己的保存时间戳来对齐：
    一条来自 CME 的 Excel、一条来自 IR 服务器，两条独立证据同一天才采信。
    实测 2026-07 那期是 20:53:17Z 保存、21:47:51Z 上线，差 54 分钟，同一天。
    对不上就返回 None 让抬头缺半句 —— 缺席远好过印一个看不出来是假的日期。
    """
    lm_meta = os.path.join(cache_dir, CACHE_NAME + '.lastmod')
    if not os.path.exists(lm_meta):
        return None, '没有 .lastmod 旁证文件（该文件由 download() 与 xlsx 一起落盘）'
    with open(lm_meta, encoding='utf-8') as f:
        lm_raw = f.read().strip()
    if not lm_raw:
        return None, '响应里没有 Last-Modified 头'
    try:
        http_at = email.utils.parsedate_to_datetime(lm_raw)
    except (TypeError, ValueError):
        return None, f'Last-Modified 读不懂: {lm_raw!r}'
    if http_at.tzinfo is None:                     # RFC 7231 要求带 GMT，缺了就按 UTC 读
        http_at = http_at.replace(tzinfo=_dt.timezone.utc)

    saved_at, saved_raw = _wb_saved_at(xlsx_path)
    if saved_at is None:
        return None, f'工作簿 docProps/core.xml 里取不到 dcterms:modified（原文 {saved_raw!r}）'

    d_http = http_at.astimezone(_TZ).date()
    d_saved = saved_at.astimezone(_TZ).date()
    if d_http != d_saved:
        return None, (f'两条证据不在同一天：Last-Modified {d_http}、'
                      f'工作簿保存 {d_saved}（芝加哥时间）')

    return d_http.strftime('%Y-%m-%d'), (
        f'{CACHE_NAME}（官方文件名 {name or "未给出"}）的 HTTP Last-Modified "{lm_raw}"，'
        f'与工作簿 docProps/core.xml 的 dcterms:modified {saved_raw} 互证，'
        f'换算到芝加哥时间同为 {d_http}')


def _name_month(name):
    """从官方文件名里的 'Jul26' 取出 'YYYY-MM'；认不出返回 None。

    只用来做发布日归属的校验（见 _record_publish_date），不参与取数 ——
    表内数据才是最新月的真值（见模块 docstring）。
    """
    m = re.search(r'_([A-Z][a-z]{2})(\d{2})(?![0-9])', name or '')
    if not m:
        return None
    try:
        mon = _dt.datetime.strptime(m.group(1), '%b').month
    except ValueError:
        return None
    return f'20{m.group(2)}-{mon:02d}'


def _month_delta(a, b):
    """a 比 b 领先几个月（'YYYY-MM' 字符串）；有一个读不懂就返回 None。"""
    try:
        ya, ma = int(a[:4]), int(a[5:7])
        yb, mb = int(b[:4]), int(b[5:7])
    except (TypeError, ValueError):
        return None
    return (ya - yb) * 12 + (ma - mb)


def _crosscheck_stored_tail(stored_months, parsed_index):
    """已入库的最近几个月，在这一轮解析结果里必须还在 —— 少一个就炸。

    这是「反侧」那道网，和 _parse_adv / _parse_oi 的行数对账盲区不重叠：
    对账从**工作簿里有什么**这一侧查（认得出的标签没地方放），这里从
    **我们已知该有什么**的反侧查。只有反侧这道才拦得住「行标签只在最近几个
    年份块被改名」那种局部走样 —— 那时旧月份照样解析得出来，全表空列检查
    看不见，可 2026 年那几行已经从解析结果里消失了，而它们明明躺在
    series/cme.csv 里。

    为什么这个源「少一行 = 我们解析丢了」而不是「官方少发了」：xlsx 是一份
    滚动覆盖全历史的文件（坑 7 说得清楚，官方只**重述**数值，不删行）。
    所以处置和 fetch/msci.py 的哨兵②一致 —— 数值变了只喊，整行没了直接抛。

    只查表尾 SENTINEL_TAIL_MONTHS 个月，理由见该常量旁注。
    """
    tail = sorted(stored_months)[-SENTINEL_TAIL_MONTHS:]
    gone = [m for m in tail if m not in parsed_index]
    if gone:
        raise FetchError(
            f'{gone} 在 series/cme.csv 里，这一轮却从 xlsx 解析结果里消失了。'
            '这个源只重述数值、不删行（坑 7），所以多半是我们解析漏了 —— '
            '行标签被改名、或某个年份块的日期表头读不出来，都会长成这样。'
            f'本次不写入，请对照 {CACHE_NAME} 人工确认。')


def _crosscheck_name_month(name, newest):
    """官方文件名自报的数据月 vs 表内解析出来的最新完整月 —— 领先太多就炸。

    本模块唯一一条**独立于工作簿解析器**的月份判据（同 fetch/cboe.py 的
    _crosscheck_report_month、fetch/ice.py 的 _crosscheck_workbook_month）：
    文件名来自 IR 服务器的 Content-Disposition，和我们怎么读 sheet 毫无关系。
    防的是行数对账和全表空列检查都看不见的那一类坏法 —— 比如最新月的数值格
    被官方填成 'n/a' 之类的文本，_num 返回 None，那一行被 CORE 过滤器判成
    半拉子月份丢掉：表头没坏、标签没改、旧月份齐全，前两道全都不响。

    只在文件名**领先**超过 NAME_MONTH_LEAD_TOL 期时抛，落后不抛：文件名滞后
    是纯化妆品问题（官方忘了改名），不该让一个数据好好的抓取器停摆。
    文件名认不出格式时也不抛，但要喊一声护栏掉了 —— 「静默地失去一道护栏」
    本身就是第四类失败。
    """
    if not newest:
        return                    # 空帧的判据在 _crosscheck_stored_tail，这里只是别炸成 IndexError
    nm = _name_month(name)
    if nm is None:
        print(f'[cme] ⚠ 护栏失效：官方文件名 {name!r} 认不出数据月（命名可能变了），'
              '这一轮没有独立于解析器的月份判据')
        return
    lead = _month_delta(nm, newest)
    if lead is None or lead <= NAME_MONTH_LEAD_TOL:
        return
    raise FetchError(
        f'官方文件名自报数据月 {nm}，表内解析出来的最新完整月却只到 {newest}，'
        f'领先 {lead} 期（容忍 {NAME_MONTH_LEAD_TOL} 期，见 NAME_MONTH_LEAD_TOL）。'
        '文件名和表内数据是同一份 artifact，不该差这么多 —— 最可能是最新那几个月'
        '被静默丢掉了（数值格变成文本、行被吞）。拒绝写入，请人工看一眼工作簿。')


def _record_publish_date(series_dir, cache_dir, xlsx_path, name, added, latest):
    """把这一期的发布日记进台账 —— 只记这个文件自己那一期（latest），且它确实刚落库。

    xlsx 是「一份文件滚动覆盖全历史」的形态：它的 Last-Modified 只证明**最新那个月**
    是这天发的。补历史时一次 update 可能追进好几个月，若给每个新月都盖上同一个日期，
    台账上就会长出「2026-05 的数据 2026-08-03 才发布」这种看着挺像真的假话 ——
    record() 的护栏只拦「发布日早于数据月」，拦不住晚三个月。

    再加一道文件名校验：官方文件名里的月份（Jul26）与表内最新完整月对不上时，
    说明这份文件已经是更新一期、而我们最新能收全的还是上一个月（例如 OI 晚上线），
    那么这个发布日属于文件里那个更新的月份，安到 latest 头上就是串月。
    """
    if latest not in added:
        return
    nm = _name_month(name)
    if nm and nm != latest:
        print(f'  发布日未记：官方文件名月份 {nm} 与表内最新完整月 {latest} 不一致')
        return
    day, evidence = _publish_date(cache_dir, xlsx_path, name)
    if not day:
        print(f'  发布日未记（{latest}）：{evidence}')
        return
    _source_dates().record(series_dir, 'cme', latest, day, evidence)


# ── 对外接口 ────────────────────────────────────────────────────────────────
def latest_month(cache_dir) -> str | None:
    """官方源当前最新完整月 "YYYY-MM"。抓不到 / 解析不出 -> 抛 FetchError。"""
    path, _ = download(cache_dir)
    df = parse(path)
    if df.empty:
        raise FetchError('xlsx 解析成功但没有任何完整月份')
    return str(df.index[-1])


def _read_series(series_dir):
    p = os.path.join(series_dir, 'cme.csv')
    if not os.path.exists(p):
        raise FetchError(f'找不到 {p}')
    cur = pd.read_csv(p, dtype={'month': str})
    if list(cur.columns) != COLUMNS:
        raise FetchError(f'series/cme.csv 列名与预期不符\n实际: {list(cur.columns)}')
    # series/cme.csv 现存是 CRLF。pandas 默认写 LF，会把整个文件变成 224 行全改的
    # git diff，把真正新增的那一两行埋掉 —— 所以沿用原文件的行尾，别硬编。
    with open(p, 'rb') as f:
        eol = '\r\n' if b'\r\n' in f.read(4096) else '\n'
    return p, cur, eol


def _fmt(v):
    """还原 CSV 现有的写法：整数不带 .0，其余用 Python 最短往返 repr。

    现存文件里同时有 `29581` 和 `35879.42523809524` 这两种形态，说明它当初就是
    「整数化 + float repr」写出来的。这里必须原样复刻 —— 若改用 f'{v:.6f}'
    之类的定长格式，新旧行的精度写法会不一致，且四舍五入会真的丢掉有效数字。
    """
    if v is None or (isinstance(v, float) and pd.isna(v)) or pd.isna(v):
        return ''
    f = float(v)
    if f.is_integer():
        return str(int(f))
    return repr(f)


def update(series_dir, cache_dir) -> list:
    """把官方源里比 series/cme.csv 更新的月份追加进去，返回新增月份列表。

    幂等：已有月份一概不动（含官方重述 —— 见坑 7，重述要人工确认后再覆盖）。
    任一必填列缺失就抛异常，绝不写 NaN。

    实现上刻意走**文本行级插入**而不是 pandas 读回再整表 to_csv：
    整表重写会把已有行的浮点精度按当前格式化规则重新渲染一遍
    （实测会把 35879.42523809524 写成 35879.425238），等于悄悄改了历史数据，
    而且 git diff 会变成整文件全改。行级插入保证已有行逐字节不动。
    """
    path, name = download(cache_dir)
    src = parse(path)
    csv_path, cur, eol = _read_series(series_dir)

    # ── 两道判据必须跑在下面那个 `if not new: return []` 之前 ──────────────
    # 那个早退是本模块所有「不出声的失败」共同的出口：只要解析结果里没有新月份，
    # 不管是因为官方今天真没发，还是因为我们把最新那一整年读丢了，update() 都
    # 返回同一个 []，monthly_run 都记同一个 NOCHANGE。判据放在早退之后
    # （比如 _record_publish_date 那一带）等于永远不执行。
    _crosscheck_stored_tail(cur['month'].astype(str).tolist(), src.index)
    _crosscheck_name_month(name, str(src.index[-1]) if len(src.index) else None)

    have = set(cur['month'].astype(str))
    new = [m for m in src.index if m not in have]
    if not new:
        return []

    quarantined = set(src.attrs.get('venue_quarantined', []))
    lines = []
    for m in new:
        r = src.loc[m]
        missing = [c for c in CORE_REQUIRED if pd.isna(r[c])]
        if missing:
            raise FetchError(f'{m}: 官方 xlsx 缺列 {missing}，拒绝写入（不写 NaN）')
        # venue 允许为空只有两种正当理由：ClearPort 已停披露（坑 4），
        # 或该月 venue 被一致性校验判为污染（坑 8）。其余情况说明解析出问题了。
        if m not in quarantined:
            vmiss = [c for c in VENUE_COLS
                     if c not in OPTIONAL and pd.isna(r[c])]
            if vmiss:
                raise FetchError(
                    f'{m}: venue 列 {vmiss} 缺失，且未被一致性校验标记为污染 —— '
                    f'多半是 {SHEET_ADV} 的 venue 段结构变了，拒绝写入')
        lines.append((m, ','.join([m] + [_fmt(r[c]) for c in COLUMNS[1:]])))

    with open(csv_path, 'r', encoding='utf-8', newline='') as f:
        raw = f.read()
    body = raw.split(eol)
    trailing = body.pop() if body and body[-1] == '' else None   # 末尾换行
    head, rest = body[0], body[1:]

    keyed = [(ln.split(',', 1)[0], ln) for ln in rest if ln]
    keyed.extend(lines)
    keyed.sort(key=lambda t: t[0])                    # 按 month 归位，稳定排序

    out = eol.join([head] + [ln for _, ln in keyed])
    if trailing is not None:
        out += eol
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        f.write(out)

    # 台账只在数据真的写进 series/cme.csv 之后才作证 —— 上面任何一步抛了，
    # 台账上就不该多出一行说「这个月官方发过了」。
    _record_publish_date(series_dir, cache_dir, path, name, set(new), str(src.index[-1]))
    return list(new)


def verify(series_dir, cache_dir, n=3):
    """对账：用本解析器重算 series/cme.csv 最后 n 个月，逐列比对。

    返回 [(month, column, csv值, 解析值, 相对偏差), ...]（只列不一致的）。
    对不上不代表解析错 —— 先看是不是官方重述（坑 7）。
    """
    path, _ = download(cache_dir)
    src = parse(path)
    _, cur, _eol = _read_series(series_dir)
    diffs = []
    for m in cur['month'].astype(str).tolist()[-n:]:
        if m not in src.index:
            diffs.append((m, '*', 'in csv', 'MISSING in xlsx', float('inf')))
            continue
        row = cur[cur['month'] == m].iloc[0]
        for c in COLUMNS[1:]:
            a, b = row[c], src.loc[m, c]
            if pd.isna(a) and pd.isna(b):
                continue
            if pd.isna(a) or pd.isna(b):
                diffs.append((m, c, a, b, float('inf')))
                continue
            a, b = float(a), float(b)
            if a == b:
                continue
            rel = abs(b - a) / abs(a) if a else float('inf')
            diffs.append((m, c, a, b, rel))
    return diffs


if __name__ == '__main__':
    import sys
    D = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    sd, cd = os.path.join(D, 'series'), os.path.join(D, 'cache')
    cmd = sys.argv[1] if len(sys.argv) > 1 else 'latest'
    if cmd == 'latest':
        print(latest_month(cd))
    elif cmd == 'update':
        print('added:', update(sd, cd))
    elif cmd == 'verify':
        d = verify(sd, cd, int(sys.argv[2]) if len(sys.argv) > 2 else 3)
        print('MISMATCHES:', len(d))
        for x in d:
            print('  ', x)
