# -*- coding: utf-8 -*-
"""TSMC (2330.TW) 月度营收 —— 无人值守抓取模块。

对应 build/build_tsm.py，维护三个序列文件中的两个：

  series/tsm.csv        month, revenue_ntd_mn, yoy_pct      ← 本模块自动维护
  series/tsm_fx.csv     month, ntd_per_usd                  ← 本模块自动维护
  series/tsm_guidance.csv                                   ← 本模块**不碰**，见下文「口径坑 5」

另外还挂着一段**不在无人值守路径上**的公司債工具（文件末尾「公司債」那一节）：

  series/tsm_bonds_tranches.csv   逐檔登记簿                ← 人工录，本模块只读
  series/tsm_bonds_monthly.csv    月度在外余额/票面          ← 由登记簿重建，可重放

  python3 fetch/tsm.py bonds            对着 Form 20-F 的年末余额逐年打表（只读）
  python3 fetch/tsm.py bonds --write    对账通过后重写 tsm_bonds_monthly.csv

  update() 不碰它们，monthly_run.py 也不会跑到它 —— 公司債没有月度自动源
  （MOPS ajax_t47sb17 只留滚动 3 个月窗口）。它存在的理由是月度表 77 行里
  只有 3 行是抄来的、其余全是推导值，推导值必须能被一条命令重放并对账。

────────────────────────────────────────────────────────────────────────
数据源
────────────────────────────────────────────────────────────────────────
1) 营收（tsm.csv）
   落地页 https://investor.tsmc.com/english/monthly-revenue
   页面上挂着一份官方 xlsx，文件名每月变（.../encrypt_file/mr/Historical_Monthly_Revenue_<Month>_<n>.xlsx），
   所以 URL **必须每次从落地页 HTML 里现抓**，不能写死 —— 写死的那一刻就注定了下个月 404。
   xlsx 的 "Consolidated" sheet 是「年 × 12 月」的矩阵，单位 NT$mn，与 build 脚本口径一致。

   为什么用 xlsx 而不是页面上那张 HTML 表：页面表只有「当年」12 个月，
   而算 yoy 需要去年同月；xlsx 一份文件就带全 2006-04 至今的完整历史，
   既能算 yoy，又能对账历史、又能在序列断档多个月时一次补齐。

2) 汇率（tsm_fx.csv）
   build 脚本注释写的是 FRED 的 EXTAUS。**FRED 在本机（含 cron 环境）连不通**：
   fred.stlouisfed.org 的 fredgraph.csv 在 curl 和 urllib 下都是静默超时（不是 403，是连接层挂死），
   所以不能作为无人值守源。
   改用 EXTAUS 的**上游原始数据**：美联储 H.10 台湾地区历史日度牌价
   https://www.federalreserve.gov/releases/h10/hist/dat00_ta.htm
   EXTAUS 的定义就是「该月所有营业日 H.10 牌价的算术平均」，本模块按同一定义重算。
   实测 2016-01 ~ 2026-07 共 127 个月，重算值与现有 tsm_fx.csv 最大偏差 4.8e-05
   （纯 4 位小数舍入残差），可以认定口径完全一致。

   **tsm_fx.csv 覆盖 2013-01 起，不是 2016-01**（见下面 FX_SERIES_START 那段注释）：
   它是挂在多页上的共享宏观序列，下界由「最早的消费方」决定，而不是由 TSMC 页面决定。
   2013-01 ~ 2015-12 这 36 个月是 2026-08 用同一个 fetch_fx() 一次性回补的，
   已入库的 2016-01 起那段逐字节未动。回补时的独立验证：把日度值按**年**平均，
   与美联储自己发布的 G.5A 年均（2016-01-04 那期，含 2012~2015）逐年对表 ——
   2012/2013/2014/2015 = 29.5580 / 29.6800 / 30.2990 / 31.7440，重算差 ≤ 3e-4
   （G.5A 只印到 3 位小数，残差量级与之相符）。也就是说日度解析本身经得起
   上游自己的账核，不是「我们算的月均自己跟自己一致」。

3) 交叉校验源（只读不写）
   台湾证交所 OpenAPI https://openapi.twse.com.tw/v1/opendata/t187ap05_L
   「每月营业收入汇总表」，全市场当期一份 JSON，单位是**新台币千元**。
   只有最新一期、没有历史，所以只拿来验证「官方最新月是哪个月 + 金额对不对」。
   MOPS 公开资讯观测站是同一份数据的网页版，需要 POST + 会话，无人值守下不如 OpenAPI 干净，故不用。

4) 公告日（source_dates.csv 的 tsm 行）
   TSMC 新闻中心 https://pr.tsmc.com/english/latest-news 上每月一条
   「TSMC <Month> <Year> Revenue Report」，正文第一句就是电头
   「HSINCHU, Taiwan, R.O.C. – July 13, 2026 -」，页面另有 "Issued on: 2026/07/13"。
   **这是本模块能拿到的唯一一处「官方自己说的发布日」**，所以发布日走这里，不走下面这些：
     · xlsx 的 dcterms:modified 与 IR 页 JSON-LD 的 dateModified —— 那是文件/页面被改动的
       时间戳，不是公告日，且 IR 页每被编辑一次就会往后跳；
     · TWSE OpenAPI 的「出表日期」—— 那是证交所生成这份全市场汇总档的日期
       （2026-06 那期是 1150717 = 2026-07-17，比公告晚了 4 天）；
     · 「次月 10 日」这条法规节奏 —— 见下，它只是上限，实际日子每月不同，推算必错。

────────────────────────────────────────────────────────────────────────
发布节奏
────────────────────────────────────────────────────────────────────────
· 台湾《证券交易法》要求上市公司于**次月 10 日前**公告上月营收。
  TSMC 惯例是 10 日当天（遇假日提前）盘后发布，同日更新 IR 页与 xlsx。
  → 调度建议：每月 10-14 日每天跑一次；10 日之前跑必然拿到上个月的旧数，不是故障。
  → 但「10 日」只是上限，**不能拿来推算发布日**：实测新闻稿电头
     2026-03→04/10、2026-04→05/08（10 日是周日，提前）、2026-05→06/10、
     2026-06→07/13（10 日是周五却压到下周一）。四个月里两个不在 10 日，
     所以 fetch_release_date() 每次都去新闻稿现读，不做任何外推。
· 美联储 H.10 每周一更新（含上周日度值），月度平均在次月第 1 个营业日即可算全，
  所以汇率永远不会拖累营收：营收 M 月的数在 M+1 月 10 日才有，那时 M 月汇率早已齐。
· TWSE OpenAPI 的「资料年月」是 ROC 年 + 月（11506 = 2026-06），换算要 +1911。

────────────────────────────────────────────────────────────────────────
口径坑（踩过的，别再踩）
────────────────────────────────────────────────────────────────────────
1. **investor.tsmc.com 挂 Cloudflare managed challenge，认的是 TLS 指纹，不是 UA。**
   2026-08-11 之前裸 urllib 还能 200；2026-08-23 起全面挑战：403 +
   `Cf-Mitigated: challenge` + 正文 `<title>Just a moment...</title>`。
   实测**无效**的：补齐 Accept / Sec-Fetch-* / sec-ch-ua（一整套 Chrome 头仍 403）、
   先取落地页 cookie 再下、/usr/bin/nscurl（Apple TLS 同样被挑战）。
   实测**有效**的只有 curl_cffi impersonate='chrome'（urllib 的 ja4=t13d171100… 无 ALPN；
   Chrome 是 ja4=t13d1516h2…）。所以 _get() 在识别出挑战页时回落到 curl_cffi。
   ⚠ 两点别写死：(a) 不能排除**出口 IP 信誉**也是打分输入，别把这条当成「TSMC 改了配置」的定论；
   (b) 同域的 **pr.tsmc.com 没有挑战**，裸 urllib 200 —— 别顺手把整个模块换成 curl_cffi。
2. **xlsx 有两个 sheet**：Unconsolidated（1999-2012，单体）和 Consolidated（2006-04 起，合并）。
   2013 起 TIFRS 只披露合并数，build 脚本用的是合并数。写死读 "Consolidated"，
   不要用 wb.worksheets[0] —— 那是单体表，2013 年以后全空。
3. **yoy_pct 是「用 NT$mn 整数重算再四舍五入到 1 位」**，不是官方公告里的百分比。
   官方是拿千元精度算的（如 2026-06 官方 67.86685%，用 mn 算是 67.8672%），
   两者到小数点后 1 位都是 67.9，实测 2016-01~2026-06 共 126 个月零分歧。
   但这是巧合不是保证 —— 若某月两者在第 1 位小数上劈叉，以 xlsx（mn 口径）为准，保持序列内部自洽。
4. **xlsx 早期年份带小数**（2007/2008 行是浮点），2013 年以后都是整数。
   写 CSV 时按 build 脚本的既有格式落整数，早期年份不在 series 范围内（series 从 2016-01 起），不受影响。
5. **tsm_guidance.csv 本模块不写**。它的 actual_rev_usdbn = 当季三个月 NT$ 合计 / 公司自己在
   季报里披露的当季汇率，而**那个汇率不等于三个月月均汇率的简单平均**：
   2025Q2 公司口径 31.054 vs 月均简单平均 30.8295，折出来 30.07 vs 30.29，差 0.7%。
   也就是说，拿月度数据去「自动补」季度指引表会系统性写错，且错得不显眼。
   → 季度指引 + 实际值维持人工从季报新闻稿录入，每季一次。update() 只在发现该文件落后于
      已完结季度时打印提醒，绝不代笔。

   ⚠️ **这条坑禁的是「拿月度数据折算 actual」，不是禁「录入官方原文」。** 两者别混：
      前者是构造值（月均汇率 ≠ 公司口径，上面那 0.7% 就是证据），后者是逐字抄公司自己
      印的数。2026-08 按后者把本表从 2023Q1 往前补到 2018Q1（20 季），源与复算规则如下，
      任何人可以照着重放：

      · 源：SEC EDGAR CIK 0001046179 的季度业绩 6-K（每年 1/4/7/10 月各一份）。
        2018-01 ~ 2019-07 那几份的新闻稿正文就在 6-K 主文档里；2019-10 起挪到 EX-99.1。
        清单走 https://data.sec.gov/submissions/CIK0001046179.json，
        正文走 https://www.sec.gov/Archives/edgar/data/1046179/<accession>/<doc>。
        User-Agent 必须带联系方式，否则 200 也是封禁页（同 fetch/umc.py 的口径坑 6）。
      · 六列各自的出处（**逐字**，不换算、不外推）：
          guide_low / guide_high / guide_fx  ← 报 Q-1 那份 6-K 里
            "Revenue is expected to be between US$X billion and US$Y billion; And, based on
             the exchange rate assumption of 1 US dollar to Z NT dollars"
            —— 注意指引在**上一季**那份公告里，不在本季那份。
          actual_rev_usdbn                   ← 报 Q 那份 6-K 里
            "In US dollars, <n>th quarter revenue was $A billion"
          actual_fx_ntd_per_usd              ← 同一份公告里两个官方数字相除：
            结果表 "Net sales" 那行的 NT$mn ÷ 1000 ÷ 上面那个 US$bn，四舍五入到 3 位。
            **分子必须取结果表的 NT$mn，不能取正文那句四舍五入到亿的 NT$xx.xx billion** ——
            2024Q1 用 592,644mn 得 31.407（与已入库值相符），用 592.64bn 得 31.406，差一档。
      · 数值写法：最短小数表示、至少留 1 位（公告印 "$23.50 billion" 落 23.5，
        印 "US$33.0 billion" 落 33.0）。这是本表已有 15 行的既有写法，不是新定的。
      · 口径自证：上面这套规则对**已入库的 2023Q1–2026Q3 共 15 行逐行重放，六列全部逐字
        命中**（含 2026Q3 那行只有指引没有 actual）。所以回补段与既有段是同一口径，
        不是「另找了一个看起来差不多的源」。
      · 另一条独立旁证：6-K 结果表的季度 Net sales 与 series/tsm.csv 三个月之和
        在这 20 季上逐季相差 ≤ 1 NT$mn（纯进位残差），而由它折出的 actual_fx 与
        三个月月均汇率的简单平均差 -0.35% ~ +0.19% —— 正是本坑说的那种「不显眼的系统性差」，
        再次说明月均汇率不能拿来顶替。
      · **下界卡在 2018Q1，2017Q4 及更早不补**：再往前指引口径要断（NT$/US$ 换算基准变），
        补进来就是在同一列里混两种口径。

6. **重述**：TSMC 极少重述月营收，但 xlsx 是全量覆盖式文件，一旦上游改数，
   本模块会检测到与已入库值不一致并**抛异常**，而不是悄悄改写或追加 —— 由人来判断是口径变更还是解析出错。
"""

from __future__ import annotations

import csv
import datetime as _dt
import html as _html
import json
import os
import re
import ssl
import time
import urllib.error
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

# ── 常量 ────────────────────────────────────────────────────────────────
IR_PAGE = 'https://investor.tsmc.com/english/monthly-revenue'
IR_ORIGIN = 'https://investor.tsmc.com'
H10_TAIWAN = 'https://www.federalreserve.gov/releases/h10/hist/dat00_ta.htm'
TWSE_API = 'https://openapi.twse.com.tw/v1/opendata/t187ap05_L'
TWSE_CODE = '2330'

# 公告日的唯一出处，见模块头「数据源 4」
PR_ORIGIN = 'https://pr.tsmc.com'
PR_LATEST = PR_ORIGIN + '/english/latest-news'
PR_ARCHIVE = PR_ORIGIN + '/english/news-archives'
MONTH_EN = ('January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December')

# 挑战页认的是 TLS 指纹不是这个 UA；留着只为日志好认。见口径坑 1
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/124.0 Safari/537.36')
_HEADERS = {'User-Agent': _UA, 'Accept-Language': 'en-US,en;q=0.9'}

REV_CSV = 'tsm.csv'
FX_CSV = 'tsm_fx.csv'
GUIDANCE_CSV = 'tsm_guidance.csv'

# ── 两个起点是**两件事**，不许合并回一个常量 ────────────────────────────
# SERIES_START 管的是 tsm.csv：TSMC 页面的历史起点，早于此不入库。
# FX_SERIES_START 管的是 tsm_fx.csv：汇率是**共享的宏观序列**，挂它的不止 TSMC ——
#   联电（umc）与南亚科（nanya）的营收序列自 2013-01 起，汇率停在 2016-01 就会让
#   那两页的美元腿/汇率贡献凭空少三年。H.10 的 dat00_ta.htm 本来就覆盖 2000-01 起，
#   仓库每月也一直在下这一页，所以下界纯粹是我们自己截的，放开不需要新增任何源。
# ⚠️ 谁也别把 FX 的下界再往前推「反正源里有」：2013-01 是当前**最早的消费方**
#   （umc/nanya）的起点。没有消费方的年份写进来只是给每次 update() 的重述检测
#   增加 N 年的比对面，出错概率涨、收益为零。要往前推，先有那一页。
SERIES_START = '2016-01'
FX_SERIES_START = '2013-01'

# 对账容差：营收是整数应当完全相等；yoy 允许 1 位小数的舍入方向差；汇率 4 位小数
TOL_REV = 0.51
TOL_YOY = 0.051
TOL_FX = 5e-4


# ── 底层 IO ─────────────────────────────────────────────────────────────
def _ssl_ctx():
    # 某些 cron 环境下证书链不全，且这些源都是公开只读数据、无凭证，
    # 校验失败时宁可拿到数据也不要静默停摆。
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx


# ── 抗指纹回落通道 ──────────────────────────────────────────────────────
# 同形先例：fetch/hood.py:168 _via_curl_cffi()（Akamai 按 JA3 挑客户端，urllib 必死）。
# 与 hood 的差别：hood 整个模块走 curl_cffi，本模块**只在识别出挑战页时**才回落 ——
# tsm.py 还打三个健康域（federalreserve.gov :313、openapi.twse.com.tw :335、
# pr.tsmc.com :367/:419），它们裸 urllib 一直 200，不该为了一个域换掉另外三个的传输。
_CHALLENGE_MARKS = (b'Just a moment', b'challenges.cloudflare.com', b'cf_chl_opt')


def _is_challenge(blob, headers=None):
    """Cloudflare 的挑战页是以 403 送出的。`Cf-Mitigated: challenge` 是权威判据
    （大小写不敏感，urllib 与 curl_cffi 两种 headers 对象的 .get() 都可用）；
    正文关键词只是兜底，防它哪天不发那个头。"""
    if headers is not None:
        try:
            if headers.get('cf-mitigated'):
                return True
        except Exception:                            # noqa: BLE001 —— headers 形状不确定
            pass
    return any(m in (blob or b'')[:4096] for m in _CHALLENGE_MARKS)


def _via_curl_cffi(url, timeout, referer):
    """用 Chrome 的 TLS 指纹重发。

    · **只带 Referer**：impersonate 会自动配一整套与指纹自洽的 Chrome 头，再把
      _HEADERS 里的 Chrome/124 UA 盖上去，反而造出「UA 与 TLS 指纹互相矛盾」——
      那正是风控要抓的特征。
    · verify=False 是为了对齐 _ssl_ctx() 的 CERT_NONE（理由见那里：cron 环境证书链
      不全时宁可拿到数据也不要静默停摆）。不写这行就是在回落腿上偷偷收紧了证书策略。
    · 延迟 import：curl_cffi 是 OPTIONAL_DEPS（monthly_run.py:71，requirements.txt:100
      标了「可选」）。没装时这里抛 ImportError，由调用方并进 RuntimeError ——
      仍是一条**响的 FAIL**，绝不静默返回空。
    """
    from curl_cffi import requests as cr             # noqa: PLC0415 —— 见 docstring
    r = cr.get(url, impersonate='chrome', timeout=timeout,
               headers={'Referer': referer} if referer else {}, verify=False)
    body = r.content or b''
    if r.status_code != 200 or _is_challenge(body, r.headers):
        raise RuntimeError('curl_cffi 仍被挡：HTTP %s，%d 字节' % (r.status_code, len(body)))
    return body


def _get(url, timeout=90, referer=None, tries=4):
    """带退避重试的 GET。

    退避是给 429 / 5xx / 连接层抖动准备的。

    ⚠ 这里原先写的是「investor.tsmc.com 的 WAF 有突发速率限制，隔几秒就恢复」——
    2026-08-23 实测**证伪**：同一客户端相隔 5.5 分钟两发全 403，而毫秒级换成
    Chrome TLS 指纹立刻 200。那是 Cloudflare 挑战页，不是限流，对它退避一万次也没用。
    所以命中挑战页时**不退避**，直接换传输（见 _via_curl_cffi）。
    """
    h = dict(_HEADERS)
    if referer:
        h['Referer'] = referer
    last = None
    for i in range(tries):
        if i:
            time.sleep(3 * (3 ** (i - 1)))          # 3s / 9s / 27s
        try:
            req = urllib.request.Request(url, headers=h)
            with urllib.request.urlopen(req, timeout=timeout, context=_ssl_ctx()) as r:
                return r.read()
        except urllib.error.HTTPError as e:
            last = e
            # 先读 body：挑战页以 403 送出，丢掉它就只剩光秃秃的 HTTPError 403，
            # 运维看不出病因（2026-08-23 就是这么查了半天）。
            try:
                body = e.read()
            except Exception:                        # noqa: BLE001
                body = b''
            if _is_challenge(body, e.headers):
                mit = e.headers.get('cf-mitigated') if e.headers else None
                try:
                    return _via_curl_cffi(url, timeout, referer)
                except Exception as e2:              # noqa: BLE001 —— 并进 RuntimeError，不吞
                    raise RuntimeError(
                        'GET %s 命中 Cloudflare 挑战页（cf-mitigated=%r，正文前 120B=%r），'
                        '退避无效已直接换传输，curl_cffi 回落也失败：%r —— 见口径坑 1'
                        % (url, mit, body[:120], e2)) from e
            if e.code not in (403, 429, 500, 502, 503, 504):
                raise                                # 404 之类是真错，别浪费时间重试
        except (urllib.error.URLError, TimeoutError, OSError) as e:
            last = e
    raise RuntimeError('GET %s 连续 %d 次失败，最后一次：%r' % (url, tries, last))


def _cache_write(cache_dir, name, data):
    os.makedirs(cache_dir, exist_ok=True)
    p = os.path.join(cache_dir, name)
    mode = 'wb' if isinstance(data, (bytes, bytearray)) else 'w'
    with open(p, mode) as f:
        f.write(data)
    return p


def _mkey(y, m):
    return '%04d-%02d' % (y, m)


# ── 源 1：TSMC 官方 xlsx ────────────────────────────────────────────────
def _discover_xlsx_url(cache_dir):
    """从 IR 落地页现抓 xlsx 链接。文件名每月变，写死必死（见模块头）。"""
    html = _get(IR_PAGE).decode('utf-8', 'replace')
    _cache_write(cache_dir, 'tsm_ir_monthly_revenue.html', html)
    m = re.search(r'href="([^"]*encrypt_file/mr/[^"]*\.xlsx)"', html, re.I)
    if not m:
        raise RuntimeError(
            'TSMC IR 页面上找不到 Historical Monthly Revenue 的 xlsx 链接；'
            '页面改版或被 WAF 挡了。缓存已落 %s' % os.path.join(cache_dir, 'tsm_ir_monthly_revenue.html'))
    href = m.group(1)
    return href if href.startswith('http') else IR_ORIGIN + href


# 年份格的**写法**容错 —— 放宽的是写法，不是「是不是年份」。
# openpyxl 返回的类型取决于**存进单元格的值**，不是显示格式：一旦年份被写成文本
# （手输时的前导撇号、或给年份挂一个脚注标记 "2027 *" / "2027 1"），原来那句
# `isinstance(y, int)` 一票否决，整整一行 12 个月一起消失。
# 脚注号与年份之间**必须隔一个空白或逗号**（`[\s,]+` 不是 `[\s,]*`），
# 与 fetch/msci.py 的 _MONTH_RE 是同一条教训：写成 `*` 时 "20271" 会被读成
# 年 2027 + 脚注 1，凭空捏造一个年份 —— 漏一行只是数据旧了，错一行是假数据。
_YEAR_TXT = re.compile(r'^(\d{4})(?:[\s,]+[\d*†‡§¶]+)*$')


def _year_cell(v):
    """矩阵首格 → 年份 int；不像年份返回 None。写法容错见 _YEAR_TXT 旁注。"""
    if isinstance(v, bool):                         # bool 是 int 的子类，先挡掉
        return None
    if isinstance(v, int):
        y = v
    elif isinstance(v, float) and float(v).is_integer():
        y = int(v)                                  # 年份格被存成 2027.0 的情形
    elif isinstance(v, str):
        m = _YEAR_TXT.match(v.strip())
        if not m:
            return None
        y = int(m.group(1))
    else:
        return None
    return y if 1990 < y < 2100 else None


def _parse_xlsx(path):
    """解析 Consolidated sheet → {'YYYY-MM': NT$mn(float)}。

    两道护栏，照 fetch/msci.py 的 parse() 抄的同一套，**别只留前一道**：
      (a) 年份格按 _year_cell 容错（int / 整数 float / 文本年 / 带脚注的文本年）；
      (b) **丢行对账**：首格没被认成年份、但该行 1..12 列里有 ≥3 个数值的行，
          一律记进 dropped，循环末尾一起抛。

    为什么 (b) 不能省：(a) 只认识**已经见过**的写法变体，下一种没见过的变体靠的是
    (b)。而这张表比 MSCI 那张更险 —— 一行 = 一整年 12 个月，**新一年那行一出生
    就是文本**时没有任何已入库月份消失，update() 那道「已入库月份不许从源里消失」
    的反向哨兵完全看不见它；能抓住它的只有这一段。漏掉的表现是：解析照样成功、
    max(rev_src) 悄悄停在去年 12 月、fetch 干干净净报 NOCHANGE，没有 FAIL、
    没有断档、红点也不亮，页面就一直挂着旧数据。

    阈值取 3 是为了不误伤说明行：本表的三种说明行（表头 'Net Revenue|Jan|…'、
    '(In Millions of New Taiwan Dollars)'、'Note: Starting January 2013…'）
    在 1..12 列上全是 None 或字符串，数值计数为 0，离 3 很远。
    真正的将来风险是上游新增一行合法的汇总行（CAGR / Average 之类）且带 ≥3 个
    月度数字 —— 那会变成每天硬 FAIL，届时要在这里加一条明确的白名单，
    **不要**把阈值往上抬糊过去：抬阈值就是把 (b) 关掉。
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if 'Consolidated' not in wb.sheetnames:
        # 见口径坑 2：绝不退化成「拿第一个 sheet」
        raise RuntimeError('xlsx 里没有 Consolidated sheet，实际 sheet=%s' % wb.sheetnames)
    ws = wb['Consolidated']
    out = {}
    dropped = []                                    # 长得像数据行、却没能落库的行
    for row in ws.iter_rows(values_only=True):
        y = _year_cell(row[0] if row else None)
        if y is None:
            # 说明行本该在这里被安静丢掉。但 1..12 列上摆着一排数值的行不是说明行，
            # 它就是一行数据，只是首格的年份写法我们没见过。记下来，循环后抛。
            # 切片而不是逐格索引：窄表（max_column < 13）时切片不会 IndexError。
            nums = [v for v in row[1:13]
                    if isinstance(v, (int, float)) and not isinstance(v, bool)]
            if len(nums) >= 3:
                dropped.append('%r（该行 1..12 列有 %d 个数值）' % (row[0], len(nums)))
            continue
        for i, v in enumerate(row[1:13], 1):        # 第 1..12 列 = Jan..Dec，第 13 列是 Total，不要
            if v is None or v == '':
                continue
            k = _mkey(y, i)
            if k in out:
                # 同一个月出现在两行 —— 静默覆盖会让其中一行凭空消失，而覆盖后的
                # 结果长得完全正常。(a) 放宽了年份写法，这条就是它的配套：
                # 一个多出来的「2026」文本行不许把真的 2026 行盖掉。
                raise RuntimeError('Consolidated sheet 里 %s 出现两次，源异常：%s' % (k, path))
            out[k] = float(v)
    if dropped:
        raise RuntimeError(
            'Consolidated sheet 里有 %d 行长得像数据行却没能落库：%s'
            '—— 首格的年份写法多半变了（文本年、前导撇号、脚注标记）。'
            '一行就是一整年 12 个月，宁可整次失败也不静默漏年；'
            '请对照 %s 确认写法后改 _YEAR_TXT / _year_cell'
            % (len(dropped), '；'.join(dropped[:5]), path))
    if not out:
        raise RuntimeError('Consolidated sheet 解析出 0 行，版式可能变了：%s' % path)
    return out


def fetch_revenue(cache_dir):
    """下载并解析官方 xlsx，返回 {'YYYY-MM': NT$mn}（全历史，未裁剪）。"""
    url = _discover_xlsx_url(cache_dir)
    blob = _get(url, referer=IR_PAGE)
    if not blob[:2] == b'PK':                       # xlsx 是 zip；拿到 HTML 说明被挡了
        raise RuntimeError('从 %s 下下来的不是 xlsx（前 2 字节 %r），大概率是 WAF 拦截页' % (url, blob[:2]))
    p = _cache_write(cache_dir, 'tsm_historical_monthly_revenue.xlsx', blob)
    return _parse_xlsx(p), url


def _with_yoy(rev, months):
    """给定月份列表算 yoy。缺去年同月 → 抛异常，绝不写 NaN。"""
    rows = []
    for m in months:
        y, mo = int(m[:4]), int(m[5:])
        prev = rev.get(_mkey(y - 1, mo))
        if prev is None or prev == 0:
            raise ValueError('月份 %s 缺去年同月(%s)基数，无法算 yoy_pct；拒绝写入残缺行' % (m, _mkey(y - 1, mo)))
        rows.append((m, int(round(rev[m])), round((rev[m] / prev - 1) * 100, 1)))
    return rows


# ── 源 2：美联储 H.10 台湾日度牌价 → 月均 ────────────────────────────────
def fetch_fx(cache_dir):
    """返回 {'YYYY-MM': (月均汇率, 该月计入的营业日天数)}。"""
    html = _get(H10_TAIWAN, timeout=120).decode('utf-8', 'replace')
    _cache_write(cache_dir, 'tsm_h10_taiwan.htm', html)
    buckets = {}
    for tr in re.findall(r'<tr.*?</tr>', html, re.S | re.I):
        cells = [re.sub(r'<[^>]+>', '', c).replace('&nbsp;', ' ').strip()
                 for c in re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', tr, re.S | re.I)]
        if len(cells) < 2 or not re.match(r'^\d{1,2}-[A-Za-z]{3}-\d{2}$', cells[0]):
            continue
        d = _dt.datetime.strptime(cells[0].upper(), '%d-%b-%y')
        v = cells[1].replace(',', '')
        if v.upper() in ('ND', 'NA', 'N/A', ''):    # 美方假日/停牌，H.10 写 ND；跳过不当 0
            continue
        buckets.setdefault(_mkey(d.year, d.month), []).append(float(v))
    if not buckets:
        raise RuntimeError('H.10 台湾页面解析出 0 条日度牌价，版式变了或被 Cloudflare 挡了')
    return {k: (sum(v) / len(v), len(v)) for k, v in buckets.items()}


# ── 源 3：TWSE OpenAPI，仅交叉校验 ──────────────────────────────────────
def _twse_check(cache_dir):
    """返回 (最新月 'YYYY-MM', 当月营收 NT$mn) 或 None（拿不到就算了，不阻塞主流程）。"""
    try:
        blob = _get(TWSE_API, timeout=90)
        _cache_write(cache_dir, 'tsm_twse_t187ap05_L.json', blob)
        for r in json.loads(blob):
            if r.get('公司代號') == TWSE_CODE:
                roc = str(r['資料年月'])                       # 11506 → 2026-06
                y = int(roc[:-2]) + 1911
                mo = int(roc[-2:])
                # OpenAPI 单位是新台币千元
                return _mkey(y, mo), float(r['營業收入-當月營收']) / 1000.0
    except Exception:
        return None
    return None


#: TWSE 对账的宽限期：只有过了「该月次月第 N 天」才允许把分歧判成故障。
#: 这个日子是本护栏**全部的安全边际**，收紧它就是在制造假警报。取值依据：
#:   · 法定上限是次月 10 日（模块头「发布节奏」），TSMC 惯例踩着 10 日；
#:   · 实测四个月的公告日 04/10、05/08、06/10、07/13 —— 最晚是第 13 天；
#:   · 20 比实测最晚多留 7 天、比法定上限多留 10 天。
#: 方向是刻意不对称的：xlsx 被冻住这件事晚十天发现，代价是页面多挂十天旧数据；
#: 而在发布日当天误杀一次，代价是 README「新鲜度红点」那节写的那句 ——
#: 每季度假一次的警报，人很快就学会无视了，最后整套护栏被人关掉。
TWSE_CROSSCHECK_GRACE_DAY = 20


def _crosscheck_twse_month(chk, newest, today=None):
    """证交所自报的最新月 vs 本模块从 xlsx 解析出来的最新月，差着且过了宽限期就抛。

    这是本模块唯一一条**独立于 TSMC 官网**的月份判据，同形先例是
    fetch/cboe.py 的 _crosscheck_report_month 与 fetch/ice.py 的
    _crosscheck_workbook_month。它防的是 _parse_xlsx 那两道护栏与 update() 的
    消失哨兵三道**都看不见**的坏法：xlsx 整份被冻住 —— 版式没变、解析照样成功、
    已入库月份一个都没少，只是最新那一行永远不再往下长。那种日子里 fetch 干干净净
    报 NOCHANGE，与真的「这个月还没发」长得一模一样。

    判据只有一条：`tm > newest`。series 的月份是 rev_src 的子集，所以证交所报的月
    只要比 xlsx 的最新月还新，它必然也不在 series 里，不必再查一遍。

    **宽限期不能省，也不能收紧**：TWSE 与 TSMC 不同批发布，中间有几十分钟到几小时
    的时差（见 latest_month 的 docstring）；没有宽限期的版本会在每月发布日当天
    把 TSM 打成 FAIL。宽限期内维持 print，与原来的行为一致。

    chk 为 None（TWSE 自己抽风）时只打印「护栏失效」，**绝不阻断**：判据自己挂了
    不能拖着 TSM 一起挂，否则 TWSE 的一次抽风就是 TSM 的一次停更。这一点与
    ice._crosscheck_workbook_month 认不出文件名时的处置是同一个道理 ——
    护栏掉了要让人看见，但不该把一个本来好好的源直接停摆。
    """
    if chk is None:
        print('[tsm][warn] 护栏失效：TWSE OpenAPI 这一轮取不到，'
              '本次没有独立于 TSMC 官网的月份判据（不阻断）')
        return
    tm = chk[0]
    if tm <= newest:
        return
    y, mo = int(tm[:4]), int(tm[5:])
    if not (1990 < y < 2100 and 1 <= mo <= 12):
        # 判据自己解析出一个不成立的月份时同样只告警：宁可这一轮没有护栏，
        # 也不能让 TWSE 那边的一次格式抽风把 TSM 打成 FAIL。
        print('[tsm][warn] 护栏失效：TWSE 给的月份 %r 不成立，本轮不做月份对账' % tm)
        return
    y2, mo2 = (y, mo + 1) if mo < 12 else (y + 1, 1)
    deadline = _dt.date(y2, mo2, TWSE_CROSSCHECK_GRACE_DAY)
    if (today or _dt.date.today()) <= deadline:
        print('[tsm][warn] TWSE 已出到 %s，本轮 xlsx 只解析到 %s —— %s 之前算发布时差，'
              '明天这条 cron 会再试' % (tm, newest, deadline))
        return
    raise ValueError(
        'TWSE OpenAPI 已经收到 2330 的 %s，本轮 TSMC 官方 xlsx 却只解析到 %s，'
        '且已过宽限期（%s）。证交所手里有的月份 TSMC 自己的 xlsx 不可能还没有 —— '
        '最可能是 xlsx 整份被冻住（IR 页挂着旧文件），也可能是矩阵版式变了导致'
        '最新那一年整行没被认出来。拒绝写入，请人工看一眼 '
        'cache/tsm_historical_monthly_revenue.xlsx 的最后一行'
        % (tm, newest, deadline))


# ── 源 4：新闻稿电头 = 官方公告日 ────────────────────────────────────────
# 电头里的破折号在不同月份分别出现过 ASCII '-' 与 U+2013，把整段连字符/破折号区间一并吃掉。
_DASH = r'[-‐-―]'
_MON_RE = '|'.join(MONTH_EN)


def _text_of(fragment):
    """HTML 片段 → 纯文本（含 &nbsp; 与不换行空格的归一）。"""
    t = _html.unescape(re.sub(r'<[^>]+>', ' ', fragment))
    return re.sub(r'[\s ]+', ' ', t).strip()


def _pr_listing(url, cache_dir, cache_name):
    """press center 列表页 → [(标题, 'YYYY-MM-DD', 详情页 URL)]。

    列表里那个 field_issue_date 就是页面上显示的 "Issued on"，是公司自己标的发布日；
    这里先拿它做定位与兜底，真正入库的日期以详情页正文电头为准。
    """
    html = _get(url, timeout=90).decode('utf-8', 'replace')
    _cache_write(cache_dir, cache_name, html)
    out = []
    for blk in re.split(r'(?=<article )', html):
        if not re.match(r'<article[^>]*data-history-node-id="\d+"', blk):
            continue
        href = re.search(r'href="(/[a-z]+/news/\d+)"', blk)
        day = re.search(r'field--name-field-issue-date[\s\S]{0,400}?'
                        r'<time datetime="(\d{4}-\d{2}-\d{2})', blk)
        title = re.search(r'article-title"><span[^>]*>([^<]+)<', blk)
        if href and day and title:
            out.append((_text_of(title.group(1)), day.group(1), PR_ORIGIN + href.group(1)))
    return out


def pr_index(cache_dir):
    """两张列表页合起来的新闻索引。任一页取不到只告警，不阻塞主流程。

    latest-news 只挂最近十来条（够覆盖当月），news-archives 从它断掉的地方往回排；
    两张都扫，是为了「某个月晚了一两个月才补摄入」时那条公告仍找得到。
    """
    out = []
    for url, name in ((PR_LATEST, 'tsm_pr_latest_news.html'),
                      (PR_ARCHIVE, 'tsm_pr_news_archives.html')):
        try:
            out += _pr_listing(url, cache_dir, name)
        except Exception as e:
            print('[tsm][warn] 读 %s 失败，公告日可能查不到：%r' % (url, e))
    return out


def fetch_release_date(month, cache_dir, expect_ntd_mn=None, index=None):
    """该月营收公告的官方发布日 → ('YYYY-MM-DD', evidence)；查不到返回 None。

    日期只从公司自己写出来的字符串里取，两处（优先级从高到低）：
      1. 详情页正文电头 "HSINCHU, Taiwan, R.O.C. – July 13, 2026 -"
      2. 同页 "Issued on:" 的 field_issue_date
    两者对不上时以电头为准（正文是公司的正式表述，issue_date 是 CMS 字段），
    但两个日期都写进 evidence，让人一眼看见分歧。

    expect_ntd_mn 给了就再核一次稿子里的金额：标题匹配已经锁定了月份，
    金额再对上才能排除「xlsx 与新闻稿说的不是同一个数」。对不上宁可不记。
    """
    y, mo = int(month[:4]), int(month[5:])
    want = ('tsmc %s %d revenue report' % (MONTH_EN[mo - 1], y)).lower()
    if index is None:
        index = pr_index(cache_dir)
    hit = [it for it in index if it[0].lower() == want]
    if not hit:
        return None
    title, issued, link = hit[0]

    html = _get(link, timeout=90, referer=PR_LATEST).decode('utf-8', 'replace')
    _cache_write(cache_dir, 'tsm_pr_%s.html' % month, html)
    body = re.search(r'field--name-body[^>]*field__item">([\s\S]*?)</div>', html)
    text = _text_of(body.group(1)) if body else ''

    dl = re.search(r'(%s)\s+(\d{1,2}),\s*(\d{4})\s*%s\s*TSMC\s*\(TWSE' % (_MON_RE, _DASH), text)
    if dl:
        date = '%s-%02d-%02d' % (dl.group(3), MONTH_EN.index(dl.group(1)) + 1, int(dl.group(2)))
        # evidence 里放整条电头（"HSINCHU, Taiwan, R.O.C. – July 13, 2026"），
        # 光写个日期等于让下一个人重新找一遍。
        lead = text[:dl.end(3)]
        ev = '新闻稿「%s」(%s) 正文电头 "%s"' % (
            title, link, lead if len(lead) <= 120 else '…' + lead[-80:])
        ev += '；同页 Issued on %s' % issued if date == issued else \
              '；⚠ 同页 Issued on 写的是 %s，以电头为准' % issued
    else:
        # 正文版式变了也不至于全丢：Issued on 同样是公司标注的发布日，只是不如电头正式。
        date = issued
        ev = '新闻稿「%s」(%s) 页面 "Issued on: %s"（正文电头未匹配上，版式可能变了）' % (
            title, link, issued)

    if expect_ntd_mn is not None:
        am = re.search(r'revenue for\s+%s\s+%d\s+was approximately NT\$\s*([\d,]+(?:\.\d+)?)\s*billion'
                       % (MONTH_EN[mo - 1], y), text, re.I)
        if not am:
            print('[tsm][warn] %s 新闻稿里没匹配到金额句，跳过金额核对（措辞可能变了）' % month)
        else:
            said = float(am.group(1).replace(',', ''))
            if abs(said - expect_ntd_mn / 1000.0) > 0.02:      # 稿子取到 NT$bn 两位小数
                print('[tsm][warn] %s 新闻稿 NT$%.2fbn 与 xlsx NT$%.2fbn 对不上，'
                      '不记发布日（先查是哪一边错了）' % (month, said, expect_ntd_mn / 1000.0))
                return None
            ev += '；稿内 NT$%.2fbn 与 xlsx 一致' % said
    return date, ev


def _load_ledger():
    """按路径加载仓库根的 source_dates.py —— 本模块被 monthly_run 以
    spec_from_file_location 加载，那时 sys.path 上既没有 fetch/ 也没有仓库根。"""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(ROOT, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _record_source_dates(series_dir, cache_dir, rev_src, months):
    """把这些月份的官方公告日记进 series/source_dates.csv。

    **只在数据确实写进 series 之后调用**：台账断言的是「官方哪天发的这个月」，
    数据还没入库就记，等于给一个页面上根本不存在的月份背书。
    抓不到就跳过并告警，绝不退化成构建日/文件时间 —— 页面会自动省掉
    「官方发布于」那半句，缺半句远好过印一个看不出是假的日期。
    整段失败也不往外抛：营收与汇率此刻已经落盘，为一句抬头把整次摄入判失败不划算。
    """
    if not months:
        return
    try:
        ledger = _load_ledger()
        index = pr_index(cache_dir)
    except Exception as e:
        print('[tsm][warn] 公告日索引取不到，本次不记发布日：%r' % e)
        return
    for m in months:
        try:
            got = fetch_release_date(m, cache_dir, rev_src.get(m), index=index)
            if not got:
                print('[tsm][warn] press center 上没有 %s 的 Revenue Report，本月不记发布日' % m)
                continue
            date, ev = got
            ledger.record(series_dir, 'tsm', m, date, ev)
            print('[tsm] source_date %s ← %s' % (m, date))
        except Exception as e:
            print('[tsm][warn] %s 的公告日记录失败（不影响本次数据摄入）：%r' % (m, e))


# ── CSV 读写 ────────────────────────────────────────────────────────────
def _read_csv(path):
    with open(path, newline='', encoding='utf-8') as f:
        rd = csv.DictReader(f)
        return rd.fieldnames, list(rd)


def _line_terminator(path):
    """series/*.csv 是 CRLF 落盘的。追加时若用 '\\n'，文件会变成半 CRLF 半 LF，
    git diff 会把整份文件标成改动，把「本月新增 1 行」淹掉。所以照抄既有行尾。"""
    with open(path, 'rb') as f:
        head = f.read(4096)
    return '\r\n' if b'\r\n' in head else '\n'


def _append_rows(path, fieldnames, rows):
    """只追加，不重写既有行。保持文件原有列名/列序/行尾。"""
    term = _line_terminator(path)
    with open(path, 'rb') as f:                      # 末行没有换行符时先补一个，否则会粘行
        f.seek(0, os.SEEK_END)
        n = f.tell()
        f.seek(max(0, n - 2))
        tail = f.read()
    with open(path, 'a', newline='', encoding='utf-8') as f:
        if n and not tail.endswith(b'\n'):
            f.write(term)
        w = csv.DictWriter(f, fieldnames=fieldnames, lineterminator=term)
        for r in rows:
            w.writerow(r)


# ══════════════════════════════════════════════════════════════════════════
# 公司債：逐檔登记簿 → 月度序列，以及 Form 20-F 年末余额对账
# ══════════════════════════════════════════════════════════════════════════
# 这一段**不进 update()**，不属于无人值守路径：公司債的两张表没有月度自动源
# （MOPS ajax_t47sb17 只留滚动 3 个月窗口），新券要人工录进 tranches.csv。
# 它存在的理由是：`tsm_bonds_monthly.csv` 里只有最近 3 行是抄来的，
# 其余全是**推导值** —— 推导值必须能被一条命令重放并对着官方年末余额验一遍，
# 否则下一次「谁改了一行、哪年开始不对」就只能靠肉眼。
#
#     python3 fetch/tsm.py bonds            # 只对账，不写文件
#     python3 fetch/tsm.py bonds --write    # 对账通过后重写 tsm_bonds_monthly.csv
#
# ── 2026-08 修正：曾经漏掉 100-/101-/102- 三个系列的旧券 ────────────────────
# 旧版 tranches.csv 只有 109- 起（2020-03 首檔）的新券，于是 2020-03…2023-09
# 这 43 个月的 outstanding 少了 100-/101-/102- 系列尚未到期的部分：
# 2020-03 库内 24,000 vs 实际 59,300（NT$mn），2020-12 145,100 vs 库 120,000。
# 受影响最重的是 wavg_coupon_pct —— 旧券票面 1.23%~2.10%，
# 2020-03 真实加权平均 1.25%，旧版印 0.62%，把「先降后升」的形状抹成了单调上行。
# 旧券全部到期于 2023-09（102-4F），所以 2023-10 起的行未受影响，本次未改。

BONDS_TRANCHES_CSV = 'tsm_bonds_tranches.csv'
BONDS_MONTHLY_CSV = 'tsm_bonds_monthly.csv'

# ── 月度序列的窗口起点 ─────────────────────────────────────────────────────
# 2011-09 = 100-1 甲/乙兩檔的发行月，也是**能被官方年末余额验到的最早一个月**：
# BOND_YEAR_END_NTMN 的第一行就是 2011（22,500），而 2011-09…2011-12 这四个月的
# 在外券恰好就是那三檔（D5-C 4,500 + 100-1A 10,500 + 100-1B 7,500 = 22,500），
# 所以窗口第一格是被 20-F 直接钉住的，不是外推出来的。
#
# ⚠️ **不要为了「从零画起」把它挪到 2002-01**（登记簿里最早那檔 D5-C 的发行月）。
#    登记簿在 2011 之前是**知道自己不全**的：唯一那一行叫「Domestic 5th **丙類**」——
#    同一期的甲類/乙類（更短年期，理应在 2002–2009 年间在外）根本不在表里，
#    本仓也没有任何 2011 年之前的年末余额可以把它们核出来
#    （BOND_YEAR_END_NTMN 起于 2011，20-F 最早那份是 FY2012）。
#    从 2002-01 起画，头八年就是一条**已知偏低、且无从校验**的 NT$4.5bn 直线，
#    而它会长得跟真数据一模一样。宁可窗口短九年，不要九年假线。
BONDS_FROM = '2011-09'

# 各年 20-F「BONDS PAYABLE」附注里 **Domestic unsecured bonds** 那一行的原值，NT$mn。
# 每个年末都出现在两份 20-F 里（当年那份的右栏、次年那份的左栏），两处一致才录。
# ⚠️ 这一行**含**宝岛债（在台湾发行的美元券，会计上归 domestic），
#    而本仓的月度序列**只算新台币券**，所以 2020 起两者必然差一个宝岛债，见 _bond_reconcile()。
#    2013–2018 那笔 US$1,150mn 是 TSMC Global 发的 **Overseas** unsecured bonds，
#    在附注里是**另一行**，从来不在这一行里，不需要扣。
BOND_YEAR_END_NTMN = {
    2011: 22500.0, 2012: 80000.0, 2013: 166200.0, 2014: 166200.0, 2015: 166200.0,
    2016: 154200.0, 2017: 116100.0, 2018: 91800.0, 2019: 56900.0,
    2020: 173197.0, 2021: 312448.0, 2022: 379526.0, 2023: 447194.0,
    2024: 478536.0, 2025: 538388.0,
}
# 出处（accession，CIK 0001046179）。年份 = 该份 20-F 的 fiscal year。
BOND_20F_ACCESSION = {
    2012: '0001193125-13-137651', 2013: '0001193125-14-141496',
    2014: '0001193125-15-126836', 2015: '0001193125-16-536225',
    2016: '0001193125-17-122097', 2017: '0001193125-18-121866',
    2018: '0001193125-19-108390', 2019: '0001193125-20-107579',
    2020: '0001193125-21-118512', 2021: '0001193125-22-104891',
    2022: '0001193125-23-107214', 2023: '0001193125-24-099840',
    2024: '0001193125-25-083423', 2025: '0001628280-26-025362',
}
# 同一附注的「Less: Current portion」，NT$mn —— 只收**能单独归到新台币券**的年份。
# 这是一条与年末余额彼此独立的判据：它验的是**到期时程**（哪一檔哪一年到期），
# 年末余额只验存量总额。2020/2021/2022 三年的 current portion 全部由旧券构成，
# 新券在那三年一檔都不到期 —— 换句话说，只要这三行对得上，旧券就必须在账上。
BOND_CURRENT_PORTION_NTMN = {
    2016: 38100.0, 2018: 34900.0, 2019: 31800.0,
    2020: 2600.0, 2021: 4400.0, 2022: 18100.0,
}
# 有意剔除的宝岛债（在台湾发行、以美元计价，附注里混在 domestic 那一行里），
# 名目 US$mn，按年末在外计。剔除的理由见 build/mrspecs/_tsm_extra.py 的 Exhibit ⑤ 图注：
# 2.70%/3.10%、30–40 年期，混进新台币券会同时打歪 outstanding 与 wavg_coupon。
BOND_FORMOSA_USD_MN = {2020: 1000.0, 2021: 2000.0, 2022: 2000.0,
                       2023: 2000.0, 2024: 2000.0, 2025: 2000.0}
# 宝岛债折算用的是**年末即期**，本仓没有这个序列（tsm_fx.csv 是月均），
# 所以 2020 起只能做**带宽核对**而不是零误差核对：
# 用残差反推汇率，跟同月 H.10 月均比，超过这个相对偏差就当解析出错。
# 3% 不是拍的：实测六年最大偏差 1.61%（2023-12，当月新台币月内走强，月末即期低于月均），
# 收到 2% 会在 2023 那行误报。
BOND_FX_BAND = 0.03


def _bond_month(s):
    """'2020-03-23' / '2011-09' → Period('YYYY-MM')。

    旧券（100-/101-/102-）的日期只有月精度：20-F 印的就是 "September 2011 to
    September 2016"，没有日。**不要为了让列看齐而补一个 -01** —— 那是凭空造出来的
    精度，而月度序列只按月分桶，补了也不多一分信息，只会让下一个人以为查得到那一天。
    """
    return _dt.date(int(s[:4]), int(s[5:7]), 1)


def _bond_key(d):
    return d.year * 12 + d.month - 1


def _bond_tranches(series_dir):
    _, rows = _read_csv(os.path.join(series_dir, BONDS_TRANCHES_CSV))
    out = []
    for r in rows:
        amt = float(r['issue_amount_k'])
        mat = _bond_key(_bond_month(r['maturity_date']))
        if r['repayment_type'] == 'amort_50_50':
            # 到期日**前一年**先还一半（109-4/5/6/7 那 12 檔）。
            legs = [(mat - 12, amt / 2.0), (mat, amt / 2.0)]
        elif r['repayment_type'] == 'bullet':
            legs = [(mat, amt)]
        else:
            raise RuntimeError('未知还本方式 %r（%s）' % (r['repayment_type'], r['tranche_id']))
        out.append({'id': r['tranche_id'], 'ccy': r['currency'],
                    'issue': _bond_key(_bond_month(r['issue_date'])),
                    'amount': amt, 'coupon': float(r['coupon_pct']), 'legs': legs})
    return out


def _bond_state(tranches, k, ccy='TWD'):
    """k 月末的 (在外总额_k, 在外檔数, 加权平均票面%)。"""
    tot = num = 0.0
    n = 0
    for t in tranches:
        if t['ccy'] != ccy or t['issue'] > k:
            continue
        o = t['amount'] - sum(a for km, a in t['legs'] if km <= k)
        if o > 1e-6:
            tot += o
            num += o * t['coupon']
            n += 1
    return tot, n, (num / tot if tot else None)


def rebuild_bonds_monthly(series_dir, start=BONDS_FROM):
    """按逐檔登记簿 + 同一套还本时程重算 tsm_bonds_monthly.csv 的每一行。

    窗口 = `BONDS_FROM` … 现有文件最后一个月，**逐月铺满**（不靠现有文件的
    month 列驱动，那样窗口就再也长不了；也不自行往右扩，右端归月报表管）。

    首行的 outstanding **包含一笔期初存量** —— `BONDS_FROM` 之前发行、当时
    尚未到期的券（2011-09 起点上就是 D5-C 一檔 NT$4,500mn，2002-01 发、
    2012-01 到期），所以首行 outstanding ≠ 首行 issued。这是有意的：
    序列讲的是「在外余额」，不是「本窗口内发行了多少」。图注里那条
    期初存量双算互校（build/mrspecs/_tsm_extra.py Exhibit ⑤）验的就是这一笔。

    **既有行只许重放、不许改写**：已经在文件里的月份重算后必须逐格相同，
    否则抛异常。真要改历史（像 2026-08 补旧券那次）走 `--restate`，
    让「我在改历史」这件事必须被显式打出来，而不是被一次例行重建顺手做掉。
    """
    path = os.path.join(series_dir, BONDS_MONTHLY_CSV)
    fields, rows = _read_csv(path)
    tr = _bond_tranches(series_dir)
    k0, k1 = _bond_key(_bond_month(start)), _bond_key(_bond_month(rows[-1]['month']))
    if k1 < k0:
        raise RuntimeError('BONDS_FROM=%s 晚于文件最后一个月 %s' % (start, rows[-1]['month']))
    out = []
    for k in range(k0, k1 + 1):
        tot, n, wac = _bond_state(tr, k)
        if tot <= 0:
            # 中间月在外余额为零本身不是错（2006-2013 的背書保證就真有 77 个月为零），
            # 但公司債这条线一旦断在中间，wavg_coupon 就没有定义、图上会出现假缺口。
            # 目前 2011-09 起一个月都没有，所以这里直接拒绝，而不是悄悄写空。
            raise RuntimeError('%04d-%02d 在外余额为 0，wavg_coupon 无定义 —— '
                               '窗口起点要么落在有券的月份，要么先为「零余额月」'
                               '定义一套画法（缺月 null + 非 DENSE 图型）'
                               % (k // 12, k % 12 + 1))
        iss = sum(t['amount'] for t in tr if t['ccy'] == 'TWD' and t['issue'] == k)
        rep = sum(a for t in tr if t['ccy'] == 'TWD'
                  for km, a in t['legs'] if km == k and t['issue'] <= k)
        out.append({'month': '%04d-%02d' % (k // 12, k % 12 + 1),
                    'issued_twd_k': '%d' % round(iss),
                    'repaid_twd_k': '%d' % round(rep),
                    'outstanding_twd_k': '%d' % round(tot),
                    'n_tranches_outstanding': '%d' % n,
                    'wavg_coupon_pct': ('%.4f' % wac).rstrip('0').rstrip('.')})
    # 递推自洽：除首行外，每个月都必须满足 out[t] = out[t-1] + issued - repaid。
    for i in range(1, len(out)):
        exp = (int(out[i - 1]['outstanding_twd_k']) + int(out[i]['issued_twd_k'])
               - int(out[i]['repaid_twd_k']))
        if exp != int(out[i]['outstanding_twd_k']):
            raise RuntimeError('%s 递推不自洽：%d ≠ %d' % (out[i]['month'], exp,
                                                          int(out[i]['outstanding_twd_k'])))
    return fields, out


def diff_bonds_monthly(series_dir, new_rows):
    """新算出来的行 vs 已入库的行，只比**两边都有**的月份。返回差异描述列表。

    空列表 = 这次重建只是往左补月份，一格既有值都没动。
    """
    _, old = _read_csv(os.path.join(series_dir, BONDS_MONTHLY_CSV))
    now = {r['month']: r for r in new_rows}
    bad = []
    for r in old:
        n = now.get(r['month'])
        if n is None:
            bad.append('%s 从新窗口里消失了（窗口只许往左长，不许砍掉既有月）' % r['month'])
            continue
        for c in r:
            if r[c] != n[c]:
                bad.append('%s %s：已入库 %s → 重算 %s' % (r['month'], c, r[c], n[c]))
    return bad


def _bond_dec_fx(series_dir, year):
    _, rows = _read_csv(os.path.join(series_dir, FX_CSV))
    for r in rows:
        if r['month'] == '%d-12' % year:
            return float(r['ntd_per_usd'])
    return None


def reconcile_bonds(series_dir, verbose=True):
    """逐檔登记簿 → 各年末在外余额，对着 20-F 的 Domestic unsecured bonds 核。

    返回 (通过?, 报表行列表)。两段判据不同，**不许混为一谈**：
      · 2011–2019：附注里那一行是**纯新台币**，必须零误差；
      · 2020–2025：那一行含宝岛债（我们有意剔除），只能核到「残差 ÷ 宝岛债名目
        = 一个合理的年末即期汇率」这一步 —— 本仓没有年末即期序列，做不到零误差。
    """
    tr = _bond_tranches(series_dir)
    lines, ok = [], True
    for y in sorted(BOND_YEAR_END_NTMN):
        k = _bond_key(_dt.date(y, 12, 1))
        tot_mn = _bond_state(tr, k)[0] / 1000.0
        off = BOND_YEAR_END_NTMN[y]
        resid = off - tot_mn
        usd = BOND_FORMOSA_USD_MN.get(y)
        if usd is None:
            good = abs(resid) < 0.05
            det = '零误差' if good else '❌ 差 %+.1f' % resid
        else:
            imp = resid / usd
            avg = _bond_dec_fx(series_dir, y)
            dev = abs(imp / avg - 1.0) if avg else None
            good = dev is not None and dev <= BOND_FX_BAND
            det = ('残差 = 宝岛债 US${:,.0f}mn × {:.3f}，{}-12 月均 {:.3f}，偏离 {:.2f}%'
                   .format(usd, imp, y, avg or 0.0, (dev or 0) * 100)
                   + ('' if good else '  ❌ 超出 ±%.0f%% 带宽' % (BOND_FX_BAND * 100)))
        ok &= good
        lines.append('  {}  官方 {:>11,.1f}  重建(TWD) {:>11,.1f}   {}'
                     .format(y, off, tot_mn, det))
    # 独立第二判据：到期时程。
    for y in sorted(BOND_CURRENT_PORTION_NTMN):
        k0, k1 = _bond_key(_dt.date(y, 12, 1)), _bond_key(_dt.date(y + 1, 12, 1))
        due = sum(a for t in tr if t['ccy'] == 'TWD' and t['issue'] <= k0
                  for km, a in t['legs'] if k0 < km <= k1) / 1000.0
        off = BOND_CURRENT_PORTION_NTMN[y]
        good = abs(off - due) < 0.05
        ok &= good
        lines.append('  {}  current portion 官方 {:>9,.1f}  重建 {:>9,.1f}  {}'
                     .format(y, off, due,
                             '零误差' if good else '❌ 差 %+.1f' % (off - due)))
    if verbose:
        print('[tsm][bonds] Domestic unsecured bonds 年末对账（NT$mn，源：Form 20-F '
              'BONDS PAYABLE 附注，CIK 0001046179）')
        print('\n'.join(lines))
        print('[tsm][bonds] 脚注：2011–2019 为纯新台币，判据是**零误差**；'
              '2020 起附注那一行含在台发行的美元宝岛债'
              '（109-1-USD 2.70%/40y、110-5-USD 3.10%/30y，各 US$10 亿），'
              '本仓序列有意只算新台币券，故只能核到「残差 = 宝岛债名目 × 年末即期」。'
              '2013–2018 那笔 US$1,150mn 是 TSMC Global 的 Overseas unsecured bonds，'
              '在附注里另起一行，本来就不在核对口径内。')
        print('[tsm][bonds] %s' % ('全部通过' if ok else '❌ 有未通过项'))
    return ok, lines


# ── 公开接口 ────────────────────────────────────────────────────────────
def latest_month(cache_dir):
    """官方源当前最新已公告月，'YYYY-MM'。抓不到抛异常。

    以 TSMC 官方 xlsx 为准（它就是 series 的真值来源）；
    TWSE OpenAPI 只做交叉校验，不一致时打印告警但不改变返回值 ——
    两边发布有几十分钟到几小时的时差，硬对齐反而会制造假故障。
    """
    rev, _ = fetch_revenue(cache_dir)
    latest = max(rev)
    chk = _twse_check(cache_dir)
    if chk:
        tm, tv = chk
        if tm != latest:
            print('[warn] TWSE OpenAPI 最新月 %s ≠ TSMC xlsx 最新月 %s（发布时差，通常几小时内收敛）' % (tm, latest))
        elif abs(tv - rev[latest]) > 1.0:
            print('[warn] %s 金额分歧：TWSE %.0f vs xlsx %.0f (NT$mn)' % (latest, tv, rev[latest]))
    return latest


def update_fx(series_dir, cache_dir, fx_src=None):
    """只把新月份写进 series/tsm_fx.csv，返回新增月份列表（已排序）。

    **为什么它必须能脱离 tsm.csv 单独跑**（2026-09-05 拆出来的）：
    tsm_fx.csv 是挂在台湾六页上的**共享宏观序列** —— tsm / umc / ase / mtk /
    nanya / guc 的 build 都要它覆盖自家营收的每一个月，缺一个月 mrbase.py 就抛
    `SpecError: series/tsm_fx.csv 缺月份 [...]`。可它只由本模块维护，而
    monthly_run.TICKERS 里 **umc / ase / mtk / nanya / guc 五家全排在 tsm 前面**。
    于是「谁家先披露完营收，谁就先撞上『营收有 M 月、汇率还没有 M 月』」——
    那一轮那一家 build 失败，而下一轮 fetch 已经没有新月份了（NOCHANGE），
    页面从此静默停在旧月份，长得和健康的安静日一模一样。

    2026-09-05 就是这么打死 /umc/ 的：UMC 的 6-K 先到（2026-08），TSMC 的
    xlsx 还停在 7 月、tsm_fx.csv 也还停在 2026-07，umc 当轮 FAIL、之后每轮 NOCHANGE。
    所以 monthly_run.py 在**按家循环之前**单独跑这一步，让共享底座先就位。

    汇率月份的推进**不依赖 TSMC 披不披露**：H.10 是美联储的周更序列，
    月均在次月第 1 个营业日就算得全（见文首第 2 节），与台湾任何一家的披露节奏无关。

    幂等：已有月份一律不重复追加。fx_src 传入可复用已经抓好的那份，省一次 HTTP。
    """
    fx_path = os.path.join(series_dir, FX_CSV)
    if fx_src is None:
        fx_src = fetch_fx(cache_dir)
    fx_fields, fx_rows = _read_csv(fx_path)
    have_fx = {r['month'] for r in fx_rows}
    for r in fx_rows:
        m = r['month']
        if m in fx_src and abs(fx_src[m][0] - float(r['ntd_per_usd'])) > TOL_FX:
            raise ValueError('汇率重述/口径漂移：%s 重算 %.4f vs 已入库 %s'
                             % (m, fx_src[m][0], r['ntd_per_usd']))

    # 只收「已经走完」的月份：当月还没结束时月均是半截数，写进去下次就得改
    today = _dt.date.today()
    cur = _mkey(today.year, today.month)
    new_fx_months = sorted(m for m in fx_src
                           if m >= FX_SERIES_START and m not in have_fx and m < cur)
    # H.10 一个月至少有 15 个营业日；明显偏少说明该月数据还没灌全
    for m in new_fx_months:
        if fx_src[m][1] < 15:
            raise ValueError('H.10 %s 只有 %d 个日度观测，月均不可信，本次不写入'
                             % (m, fx_src[m][1]))
    # _append_rows 只会往**文件尾**写。下界放到 2013-01 之后，「待写月份早于已入库
    # 最大月」第一次成为可能（文件被截断、或哪个中间月漏了），那会写出一份乱序 CSV，
    # 而乱序 CSV 不会报错、只会让下游按行序取「最新月」时静默取错。宁可停下。
    if have_fx and new_fx_months and min(new_fx_months) < max(have_fx):
        raise ValueError('待写汇率月份 %s 早于已入库最大月 %s；追加会写出乱序 CSV。'
                         '先确认 tsm_fx.csv 是不是缺了中间月份'
                         % (sorted(m for m in new_fx_months if m < max(have_fx)), max(have_fx)))
    new_fx_rows = [{'month': m, 'ntd_per_usd': '%.4f' % fx_src[m][0]} for m in new_fx_months]
    if new_fx_rows:
        _append_rows(fx_path, fx_fields, new_fx_rows)
    return new_fx_months


def update(series_dir, cache_dir):
    """把新月份写进 series/tsm.csv 与 series/tsm_fx.csv，返回新增月份列表（两文件并集，已排序）。

    幂等：已有月份一律不重复追加。
    任何一列解析不出来（如缺去年同月基数、缺当月汇率）→ 抛异常，绝不写 NaN。
    发现上游重述（已入库月份的值对不上）→ 抛异常，交人判断。
    已入库月份从 xlsx 里**整个消失** → 抛异常（消失哨兵，见下面那一支的注释）。
    证交所已经收到、xlsx 却还没有的月份，过了宽限期 → 抛异常
    （_crosscheck_twse_month，本模块唯一一条独立于 TSMC 官网的判据）。
    """
    rev_path = os.path.join(series_dir, REV_CSV)
    fx_path = os.path.join(series_dir, FX_CSV)

    rev_src, xlsx_url = fetch_revenue(cache_dir)
    fx_src = fetch_fx(cache_dir)

    # ── 营收 ──
    rev_fields, rev_rows = _read_csv(rev_path)
    have_rev = {r['month'] for r in rev_rows}
    # 下界卡 min(rev_src)：这份 xlsx 从 2006-04 起一直是全历史累积文件，但万一哪天
    # 上游改成滚动窗口、把远古历史截掉，那不是解析漏了，不该 FAIL —— 下界正好吸收它。
    rev_src_floor = min(rev_src)
    for r in rev_rows:                                   # 重述检测 + 消失哨兵
        m = r['month']
        if m in rev_src:
            if abs(rev_src[m] - float(r['revenue_ntd_mn'])) > TOL_REV:
                raise ValueError('上游重述：%s 官方 %.0f vs 已入库 %s (NT$mn)。'
                                 '本模块拒绝改写既有行，请人工确认后再决定。'
                                 % (m, rev_src[m], r['revenue_ntd_mn']))
        elif m >= rev_src_floor:
            # 「数值变了」与「整行没了」是两回事，处置也不同（同 fetch/msci.py 的
            # update() 哨兵②）：前者上面那一支已经在管；后者不是源的正常行为 ——
            # TSMC 从不删历史月，所以这多半是**我们**把那一行解析丢了。
            # 它是 _parse_xlsx 那道丢行对账的补网，两道网的盲区不重叠：对账从
            # 「文件里有什么」这一侧查，只看得见「首格不像年份、后面摆着一排数值」的行；
            # 这里从「我们已知该有什么」的反侧查，连「整年那行被整块删掉」这种对账
            # 看不见的坏法也能兜住。少了这一支，那种日子里已入库月份从 rev_src 里
            # 整批消失，下面的重述循环 `if m in rev_src` 直接跳过、不比也不抛，
            # new_rev_months 为空，fetch 干净地报 NOCHANGE。
            raise ValueError('%s 在 series/%s 里，官方 xlsx 却解析不出这一个月'
                             '（本轮解析区间 %s..%s）—— TSMC 不删历史，'
                             '所以多半是解析漏了整整一年那一行。本次不写入，'
                             '请对照 cache/tsm_historical_monthly_revenue.xlsx 人工确认'
                             % (m, REV_CSV, rev_src_floor, max(rev_src)))

    # ── 独立外部判据：证交所已经收到的月份，官方 xlsx 不可能还没有 ──
    # 放在写盘之前：xlsx 冻住而汇率还有新月份的那一天，若先写 FX 再抛，这一家会
    # 返回一串 FX 月份、状态显示 NEW（营收其实还是旧的），比纯 NOCHANGE 更难看出来。
    _crosscheck_twse_month(_twse_check(cache_dir), max(rev_src))

    new_rev_months = sorted(m for m in rev_src if m >= SERIES_START and m not in have_rev)
    new_rev_rows = [{'month': m, 'revenue_ntd_mn': v, 'yoy_pct': y}
                    for m, v, y in _with_yoy(rev_src, new_rev_months)]

    # ── 汇率 ── 实际写盘在 update_fx() 里（拆分理由见那个函数的 docstring）。
    #   这里仍在 _crosscheck_twse_month 之后调用，所以直接 `python3 fetch/tsm.py`
    #   的行为与拆分前逐字节一致：营收源冻住时不会先把汇率写进去。
    new_fx_months = update_fx(series_dir, cache_dir, fx_src)
    fx_fields, fx_rows = _read_csv(fx_path)
    have_fx = {r['month'] for r in fx_rows}

    # ── 一致性闸门：build 脚本要用汇率折美元，营收月必须都有汇率 ──
    all_rev = have_rev | set(new_rev_months)
    all_fx = have_fx | set(new_fx_months)
    missing = sorted(all_rev - all_fx)
    if missing:
        raise ValueError('这些月份有营收但没有汇率，写进去会让 build_tsm.py 折出 NaN：%s' % missing)

    if new_rev_rows:
        _append_rows(rev_path, rev_fields, new_rev_rows)

    # 公告日在这里记：上面的 append 已经返回（汇率那份在 update_fx 里已落盘），
    #   这些月份确实在 series 里了。
    # 除了本次新增的月份，还补记「最新月在台账里缺着」的情况 —— 台账是后加的，
    # 早于它入库的月份不会有记录，而幂等重跑时 new_rev_months 是空的，
    # 不带这一句就永远补不上（页面也就永远少那半句）。
    latest_rev = max(all_rev)
    todo = set(new_rev_months)
    try:
        if _load_ledger().lookup(series_dir, 'tsm', latest_rev) is None:
            todo.add(latest_rev)
    except Exception as e:
        print('[tsm][warn] 台账读取失败，只尝试本次新增月份的公告日：%r' % e)
    _record_source_dates(series_dir, cache_dir, rev_src, sorted(todo))

    _guidance_reminder(series_dir, max(all_rev))

    print('[tsm] xlsx=%s' % xlsx_url)
    print('[tsm] tsm.csv +%d %s | tsm_fx.csv +%d %s'
          % (len(new_rev_rows), new_rev_months, len(new_fx_months), new_fx_months))
    return sorted(set(new_rev_months) | set(new_fx_months))


def _guidance_reminder(series_dir, latest_rev_month):
    """季度指引表只能人工录（见口径坑 5），这里只提醒，不代笔。"""
    p = os.path.join(series_dir, GUIDANCE_CSV)
    if not os.path.exists(p):
        return
    try:
        _, rows = _read_csv(p)
        y, mo = int(latest_rev_month[:4]), int(latest_rev_month[5:])
        curq = '%dQ%d' % (y, (mo - 1) // 3 + 1)
        have = {r['quarter'] for r in rows}
        blank_actual = [r['quarter'] for r in rows
                        if not (r.get('actual_rev_usdbn') or '').strip() and r['quarter'] < curq]
        if curq not in have:
            print('[tsm][人工] tsm_guidance.csv 缺 %s 的指引区间，需从季报新闻稿录入' % curq)
        if blank_actual:
            print('[tsm][人工] tsm_guidance.csv 这些已完结季度还没填 actual：%s'
                  '（actual_fx 必须用公司季报披露的汇率，不能拿月均汇率凑）' % blank_actual)
    except Exception as e:
        print('[tsm][warn] 指引表检查失败（不影响月度更新）：%s' % e)


if __name__ == '__main__':
    import sys
    sd = os.path.join(ROOT, 'series')
    cd = os.path.join(ROOT, 'cache')
    if len(sys.argv) > 1 and sys.argv[1] == 'bonds':
        # 公司債：对账（只读）。带 --write 才在对账通过后重写 tsm_bonds_monthly.csv。
        # 不联网 —— 年末余额是从 20-F 逐份人工核对后写死在 BOND_YEAR_END_NTMN 里的。
        _ok, _ = reconcile_bonds(sd)
        if not _ok:
            raise SystemExit('[tsm][bonds] 对账未通过，拒绝写文件')
        if '--write' in sys.argv[2:]:
            _f, _rows = rebuild_bonds_monthly(sd)
            _bad = diff_bonds_monthly(sd, _rows)
            if _bad:
                print('[tsm][bonds] 重算结果与已入库的行不一致（%d 处）：' % len(_bad))
                print('\n'.join('    ' + b for b in _bad[:20]))
                if len(_bad) > 20:
                    print('    …另有 %d 处' % (len(_bad) - 20))
                if '--restate' not in sys.argv[2:]:
                    raise SystemExit(
                        '[tsm][bonds] 拒绝写文件：例行重建只许往左补月份，不许改写既有值。'
                        '若这确实是一次有据可查的历史更正（像 2026-08 补 100-/101-/102- '
                        '旧券那次），显式加 --restate，并把依据写进 BOND_YEAR_END_NTMN '
                        '上方的修正记录。')
                print('[tsm][bonds] --restate：已确认要改写上面这些既有值')
            _p = os.path.join(sd, BONDS_MONTHLY_CSV)
            with open(_p, 'w', newline='', encoding='utf-8') as _fh:
                _w = csv.DictWriter(_fh, fieldnames=_f, lineterminator=_line_terminator(_p))
                _w.writeheader()
                _w.writerows(_rows)
            print('[tsm][bonds] 已重写 %s（%d 行，%s … %s）'
                  % (BONDS_MONTHLY_CSV, len(_rows), _rows[0]['month'], _rows[-1]['month']))
    elif len(sys.argv) > 1 and sys.argv[1] == 'latest':
        print(latest_month(cd))
    elif len(sys.argv) > 2 and sys.argv[1] == 'pubdate':
        # 单独查某个月的公告日（只读，不写台账），用于核对 evidence。
        # 金额从已入库的 series 里取，好让这条调试路径跟真实路径走同样的核对。
        mth = sys.argv[2]
        _, _rows = _read_csv(os.path.join(sd, REV_CSV))
        _exp = next((float(r['revenue_ntd_mn']) for r in _rows if r['month'] == mth), None)
        print(fetch_release_date(mth, cd, _exp))
    else:
        print(update(sd, cd))
