# -*- coding: utf-8 -*-
"""Japan Exchange Group (JPX / 日本取引所グループ) 月度经营统计 —— 无人值守抓取。

东証（现货）+ 大阪取引所（衍生品）。全部来自 JPX 官方统计站，没有任何第三方聚合商。

━━ 数据源 ━━
现货成交（1985-01 起）:
  https://www.jpx.co.jp/english/markets/statistics-equities/misc/tvdivq0000000vzk-att/
    historical-genbutsu.xlsx   内国株 + 外国株，**目录 token 固定，文件每月原地覆盖**
    historical-toushin.xlsx    内国ETF / 外国ETF・ETN / 内国投資証券(REIT)
月末时价总额（1949-05 起）:
  …/misc/tvdivq0000001w3y-att/historical-jika.xlsx
上場会社資金調達額（1998-01 起，含 IPO 件数与募资额）:
  …/misc/tvdivq0000001wij-att/historical-sikin.xls
衍生品逐产品 ADV / 月末 OI（长表，2014-12 起覆盖日経225 系）:
  https://www.jpx.co.jp/automation/markets/statistics-derivatives/trading-volume/files/soukatsu_M.xlsx
  —— **稳定直链，无月份 token，可以写死**（/automation/ 目录就是给机器读的）
衍生品总量与三大类小计（宽表，1985-10 起，含已退市品种）:
  …/statistics-derivatives/trading-volume/{token}-att/tv_ts{YYYYMM}.xls
  —— token **每月都变**，必须先抓落地页
  …/english/markets/statistics-derivatives/trading-volume/index.html 取 href。
  实测把上个月的 token 拼上下个月的文件名（tv_ts202607.xls）得到 **404 而不是 403**，
  所以「还没发」与「命名规则变了」表现完全一样、无法区分 —— 猜直链这条路是死的。

**为什么宽表不能省，明明长表看起来更细**：长表只收录「当前仍在挂牌」的产品，
历史上退市的（日経300 先物、業種別先物、S&P/TOPIX150、取引所 FX 証拠金、Nifty50…）
整条不在里面。实测「长表逐产品求和 == 宽表官方 Total」**只有 2023-06 起成立**，
489 个月里 452 个月不闭合（2015-01 差 2,067,619 张）。Total 与三大类小计只能取宽表。

抓取方式：`urllib.request` + 普通 UA。2026-08-06 实测：不设 UA（默认 Python-urllib/3.x）
也全部 200，`robots.txt` 是 `User-Agent:* / Disallow:`（空 = 全站放行），HEAD 可用且返回
Last-Modified。**没有 Cloudflare / Akamai 挑战、没有 JS 渲染门、没有登录墙**，
不需要 curl_cffi / nscurl / 浏览器 —— 这是本仓无人值守条件最宽松的一家。

━━ 发布节奏 ━━
| 文件                              | 官方明文（页面 Note 原文）                    | 实测 Last-Modified |
|-----------------------------------|---------------------------------------------|--------------------|
| soukatsu_M.xlsx / tv_ts{YYYYMM}   | 每月**第 5 个营业日 10:00 前后**            | 2026-07-07 10:00 JST |
| historical-genbutsu / -toushin    | 每月**第 5 个营业日 13:00 前后**            | 2026-07-07 13:00 JST |
| historical-jika（时价总额）       | 每月**第 1 个营业日 13:00 前后**            | 2026-08-03 13:00 JST |
| historical-sikin（资金調達）      | 每月中旬（实测 07-17）                      | 2026-07-17 13:00 JST |

⇒ `build/roster.py` 的 LAG 建议 (8, 8)：第 5 个营业日最晚落在 8 号（1 号是周六 + 假期时）。
   `monthly_run.EARLY = 5` ⇒ 闸门在月末次日就开，代价是每月多打几个空请求。

**跨文件节奏错位，这是最容易建出空行的地方**：时价总额比成交额早 4 天。
2026-08-06 实测 historical-jika.xlsx 已含 **2026-07**，historical-genbutsu.xlsx 仍停在
**2026-06**。所以 `latest_month()` **不能**看 jika，否则会建出一整行只有 mktcap 的空行。
本模块取 min(现货成交额最新月, 宽表文件名里的月份)，两个都齐了才算这个月有数。

source_dates: 落地页有机器可读的自述发布日
    <div class="JPX-site-update">Update : Jul. 07, 2026</div>
本模块取「现货落地页」与「衍生品落地页」两个日期里**较晚**的那个（一行数据要两边都发齐
才成立），evidence 里把两个页面的原文和两个文件的 HTTP Last-Modified 都记下来。
⚠ **不要写成「页面日与文件日互证」**：实测 Monthly Statistics Report 落地页写
Jul. 23 而它挂的 PDF 是 Jul. 21，**不同日**。四个日期是四条独立观测，不是互证关系。
这几个文件都是原地覆盖式发布（同一 URL 每月换内容），所以只在首次摄入某月时记一笔。

━━ 口径坑（按踩坑概率排序）━━

1. **合约张数跨所不可比：mini = 1/10、micro = 1/100，而且倍率在变。**
   实测 2026-06：日経225マイクロ 单产品 ADV 1,040k 张，占全所总张数 2,366k 的 **44.0%**。
   raw/大合约当量 的倍率 2016-06 是 2.17x，2026-06 已经是 4.52x。后果是**方向性的**：
       2016-06 → 2026-06   原始张数 1,572k → 2,366k，**+50%**
                           大合约当量  723k →   524k，**−28%**
   同一个业务，两个口径十年趋势**符号相反**。JPX 自己的 IR 从不用原始张数
   （决算说明资料第 5 页脚注："the trading volumes of mini contracts and micro contracts
   are calculated using factors of 1/10 and 1/100"）。
   ⇒ 本模块把 **`adv_deriv_total_lgeq_kcontracts`（大合约当量）作为跨所对比的默认列**；
   原始张数一律带 `_raw_`，**禁止跨 2023-05（マイクロ上市）与跨所使用**。

2. **折算表必须来自官方合约规格页，逆推会错。**
   下面 `_DIVISORS` 的每一档都去 jpx.co.jp 的 Contract Specifications 页逐个核过：
       日経225先物 = 指数×¥1,000 / mini = ×¥100 / マイクロ = ×¥10        → ÷10、÷100
       日経225オプション = 権利金×¥1,000 / ミニオプション = ×¥100          → ÷10
       TOPIX先物 = TOPIX×¥10,000 / ミニTOPIX = ×¥1,000                    → ÷10
       10年国債先物 = 額面 1 億円 / 超長期(ミニ) = 額面 1,000 万円          → ÷10
       長期国債先物（現金決済型ミニ）= 10 万円 × 10年債先物価格             → ÷10
       金標準 = 1kg / 金ミニ = 金限日 = ポケットゴールド100 = 100g          → ÷10
       白金標準 = **500g** / 白金ミニ = 白金限日 = ポケットプラチナ100 = 100g → **÷5**
   **白金那一档是 ÷5 不是 ÷10**（标准合约 500g，不是 1kg）。侦察稿里逆推出来的 ÷10 是错的，
   用它算 2026Q1 商品当量得 139.4 万张，IR 印的是 141 —— 用 ÷5 得 141.48，只有 ÷5 对得上。
   校验锚点（JPX 决算説明資料 E_EM_JPX_Q1FY2026.pdf 第 5 页，与统计站是两条独立管道）：
       Financial Derivatives  IR 印 23 / 24 mn 张（季度合计），本表折算得 23.63 / 24.58
                              y/y IR 印 (3.8%)，本表得 −3.83%
       Commodity Derivatives  IR 印 141 / 349 万张，本表折算得 141.48 / 349.41
                              y/y IR 印 (59.5%)，本表得 −59.51%
       TOPIX先物(Large) 9.2/8.8 万张/日、日経225 复合 13.5/15.0、10年国債 5.2/4.3、
       日経225オプション 44.5/27.3 十亿円/日 —— 本模块的口径逐个对上（IR 只印 2-3 位有效数字）。
   ⇒ 改折算表 = 推翻上面这组对账，改之前先把 IR 的季度值重算一遍。

3. **长表每个期权品种是 プット / コール / 合計 三行，naive groupby 直接翻倍。**
   涉及 10 个期权品种。按 (产品, 年月) 求和会得到真值的两倍（成交量与 OI 都翻）。
   本模块显式 `WHERE プットコール区分 NOT IN ('プット','コール')`，
   并用 `_ANCHORS` 把「日経225オプション 2026-06 月末 OI == 765,660」钉成断言。

4. **宽表英文列名会撞车，一律按日文名定位。**
   `TOPIX Options` 在英文表头里出现 4 次、`Security Options` 5 次 —— 表尾有一段
   Flex Options 的 `【うち…】`（of which）明细，英文名与正牌产品完全同名。
   日文名带 `【うち` 前缀，天然唯一。求和时**必须排除所有 `【うち` 开头的列**，
   否则重复计数。排除后「逐产品和 == 官方 Total」在 **489/489 个月**上 diff 恰为 0
   （本模块每次运行都重算这条闭合式，不闭合就抛异常）。

5. **现货 2022-04 被劈成两张 sheet：金额要相加，立会日数**不能**相加。**
   `月間 monthly (旧 Old)` 覆盖 1985-01…2022-04（一部/二部/マザーズ/JASDAQ），
   `月間 monthly` 覆盖 2022-04…至今（Prime/Standard/Growth）。2022-04-04 市场重组。
   旧表 2022-04 那行 `立会日数 = 1`（只有 4/1），新表那行 `立会日数 = 20`。
   **新表那个 20 是「4 月的立会日数」，不是「本行数据覆盖的日数」** —— 官方注记写明
   「2022年4月は…4月1日分の値が含まれていません…1日平均は立会日数を19日として算出」。
   所以正确做法是 **金额/股数相加、立会日数取 max**：20 + 1 = 21 是错的。
   证据：官方「年間 annual」sheet 的 2022 年立会日数 = **244**，本模块按 max 规则逐月加起来
   正好 244，按相加规则得 245。侦察稿与复核稿都写成 21 天（ADT 3.4215），
   真值是 20 天、ADT 3.5926 兆円/日 —— **别把那两份稿子的证据表当回归基准**。
   （同一份稿子里 2016-06 的 `trading_days` 写成 21，文件里是 22，也是同类转写错。）

6. **单位陷阱三处。**
   (a) 现货表里 TOKYO PRO Market 段是 **株 / 千円**，其余段是 **千株 / 百万円**
       —— 混用会把 PRO 的成交额放大 1000 倍。本模块对 PRO 显式 ×0.001。
   (b) 衍生品 OI 存**张**、ADV 存**千张**，两者差 1000 倍（与 series/cme.csv 同坑）。
   (c) 宽表里 `株先50` 的取引金額单位是 **百万円**、`米国財務省証券先物` 只有取引高没有金额
       —— 本模块从不对产品级金额求和（金额只取官方 `合計` 列），所以不受影响；
       将来若要做产品级名义额，这两列必须单独处理。

7. **商品関連在宽表里被 pro-forma 回填到 1985 年，而 JPX 当时并不拥有 TOCOM。**
   宽表 `商品関連` 从 1985-10 就有数（1985-10 占总量 62.5%、2016-06 占 7.3%），
   而旧 TOCOM 品种 **2020-07-27 才迁入 OSE**（JPX 2019-10-01 完成对 TOCOM 的收购）。
   任何从 2015/2016 起画的「JPX 衍生品总量」曲线都会把 JPX 当时没有的业务算进去，
   并把 2020 年那次收购带来的台阶抹平。
   ⇒ 本模块**不删官方数**（官方发的就是这个数，删了就不叫忠实入库），而是加一列
   `cmdty_proforma`：≤2020-07 的月份写 1，表示该月 `*_cmdty_*` 与含它的 `*_total_*`
   是 pro-forma。同时提供 **`adv_deriv_fin_lgeq_kcontracts`（金融衍生品 = 株価指数関連等
   + 国債・金利関連）**，这一列全程没有 pro-forma 污染，跨 2020-07 的长历史用它。
   注意 2020-07 本身也是混的：宽表那一格是整月（含迁入前在 TOCOM 的成交），
   长表 `金標準先物` 那一格只有迁入后 5 个立会日。

8. **日経225オプション在 2015-05…2023-05 两表口径不同（宽表含週次，长表不含）。**
   官方落地页明写 "Data for Nikkei 225 Options from May 2015 to May 2023 includes data
   for Nikkei 225 Weekly Options" —— 这句**只对宽表成立**。实测 2022-10 宽表比长表多
   109,155 张，2023-06 起两表逐格相同（週次 2023-05 改称「日経225ミニオプション」独立成列）。
   ⇒ 逐产品列一律取**长表**（口径连续，不必画断点）；`*_total_*` 取宽表（週次也是真成交，
   计入总量是对的）。两者因此在 2023-06 之前对不上，这是**已知且有意的**，不是 bug。

9. **`adt_cash_*` 与 `mktcap_eom_jpytn` 的统计范围不一样，别拿来直接算换手率。**
   ADT 分子含 内国株 + 外国株 + 内国ETF + 外国ETF/ETN + 内国投資証券(REIT)
   （这是 JPX IR「Cash Equities ADT」的口径，实测 Q1FY2026 = 12.39 兆円/日与 IR 一致；
   不含 ETF/REIT 只有 11.79，差 4.9%）；而 `historical-jika.xlsx` 只有
   プライム/スタンダード/グロース/TOKYO PRO **四段上市股票，不含 ETF / REIT**。
   ETF/REIT 约占 ADT 的 5%（2026-06 = 0.72/14.08）。要算换手率得另建股票口径 ADT。

10. **官方明说资金調達額会回溯订正**（historical-sikin.xls 表头：「本統計は、計数が過去に
    遡って訂正される可能性がある」），而现货序列实测两个 vintage（2026-02-24 期 vs
    2026-07-21 期 PDF）重叠 20 个月逐字节相同。两类都按仓库惯例处理：
    **已有值永不覆盖、只填空**。要重刷历史请手工重跑全量。

11. **`ipo_*` 两列比主数据晚半个月**（资金調達額每月中旬发），所以新月份入库时这两格天然为空，
    **这不是解析失败**；下个月的文件会把它补上（本模块因此做「只填空、不覆盖」的回补）。
    另外这两列的口径是「公募のうち新規公開」，按**払込期日**归月、且 2007-04 起只统计
    「東証に直接上場する会社」的公募 —— 它**不等于**「当月新上市公司数」
    （不做公募的上市、TOKYO PRO 上市都不在里面）。列名用 `ipo_public_offerings`
    而不是 `new_listings`，就是为了不让下游把它当 HKEX 的 `new_listings` 用。

12. **JPY 是本币，跨所对比前必须换汇，而换汇口径必须先定死。** 本模块只入库円值
    （`*_jpytn` = 兆円、`*_jpybn` = 十亿円），**不做任何换算** —— 换算是 build/notional.py
    那一层的事。USDJPY 从 2016 的 ~110 走到 150+，月均还是月末、锁不锁基期，
    会让「JPX ADT 增长」差出十几个百分点，这个约定改一次就推翻所有历史图。

━━ series/jpx.csv 每一列的确切口径 ━━
（下游换算全靠这张表；口径写错比数字抓错更难被发现）

month                            YYYY-MM
trading_days                     東証**现货**立会日数（含 ToSTNeT）。见坑 5
adt_cash_total_jpytn             東証现货**日均成交金额**，兆円/日 = 月总额 ÷ trading_days。
                                 含内国株+外国株+内国ETF+外国ETF/ETN+内国投資証券(REIT)；
                                 立会市場 + ToSTNeT（场内大宗）；**单边计**（成交额本身不双算）
adt_cash_stocks_jpytn            上列里只算股票（内国 4 段 + 外国 3 段）的部分，兆円/日
adt_cash_dom_stocks_jpytn        上列里**只算内国株**（プライム/スタンダード/グロース/
                                 TOKYO PRO 四段）的部分，兆円/日。**与 adv/vol_cash_dom_shares_mn
                                 逐段同口径**（同出 dom_*、同为立会市場 + ToSTNeT、同为单边计），
                                 是本表里唯一能与股数列相除得到「加权平均成交价」的金额列。
                                 用 adt_cash_stocks（含外国株）或 adt_cash_total（还含 ETF/REIT）
                                 去除股数，得到的均价带一个方向与大小都不可知的偏差
adt_cash_etfreit_jpytn           上列里 ETF/ETN/REIT 的部分，兆円/日
val_cash_total_jpytn             同口径的**月总成交金额**（不是日均），兆円/月
val_cash_dom_stocks_jpytn        内国株的**月总成交金额**，兆円/月。与 vol_cash_dom_shares_mn
                                 配对：年度聚合时 Σ金额 ÷ Σ股数 就是年度加权平均成交价，
                                 不必再乘立会日数去还原日均列
adv_cash_dom_shares_mn           内国株**日均成交股数**，百万株/日（含 TOKYO PRO，已换算）
vol_cash_dom_shares_mn           内国株**月总成交股数**，百万株/月
mktcap_eom_jpytn                 **月末**时价总额，兆円（存量，不是流量）。
                                 プライム+スタンダード+グロース+TOKYO PRO，**不含 ETF/REIT**。见坑 9
deriv_trading_days               大阪取引所**衍生品**立会日数（与现货个别月不同，1985-10 是 11 vs 25）
adv_deriv_total_lgeq_kcontracts  **全品种大合约当量日均张数**，千张/日。跨所对比的默认列。
                                 = Σ(各产品张数 ÷ 折算系数) ÷ deriv_trading_days，系数见 `_DIVISORS`。
                                 ≤2020-07 含 pro-forma 商品（见 cmdty_proforma）
adv_deriv_fin_lgeq_kcontracts    同上，只算**金融**（株価指数関連等 + 国債・金利関連），
                                 全程无 pro-forma 污染 —— 跨 2020-07 的长历史用这一列
adv_deriv_index_lgeq_kcontracts  同上，「株価指数関連等」大类（含个股期权、FX 先物、Flex）
adv_deriv_rates_lgeq_kcontracts  同上，「国債・金利関連」大类
adv_deriv_cmdty_lgeq_kcontracts  同上，「商品関連」大类（旧 TOCOM）。≤2020-07 是 pro-forma
adv_deriv_total_raw_kcontracts   官方 `合計` 张数 ÷ 立会日数，千张/日。**原始张数，禁止跨
                                 2023-05 与跨所使用**（见坑 1）。留着是为了能复现官方公布值
adv_deriv_index_raw_kcontracts   官方「株価指数関連等」小计，同上
adv_deriv_rates_raw_kcontracts   官方「国債・金利関連」小计，同上
adv_deriv_cmdty_raw_kcontracts   官方「商品関連」小计，同上。≤2020-07 是 pro-forma
adnv_deriv_total_jpytn           官方 `合計` **取引金額** ÷ 立会日数，兆円/日。
                                 期货是名义额、期权是**权利金**（不是名义额），两者混在一列里 ——
                                 这是官方口径，不是本模块的选择
cmdty_proforma                   1 = 该月商品関連是 TOCOM 迁入 OSE 之前的 pro-forma 回填（≤2020-07）；
                                 空 = 商品业务当月确实在 OSE 上。见坑 7
adv_n225_futures_kcontracts      日経225先物（大型，指数×¥1,000）ADV，千张/日 —— 以下逐产品列
                                 全部取**长表**，立会市場 + J-NET 合计，口径见坑 8
adv_n225_mini_kcontracts         日経225mini（×¥100）ADV，千张/日 —— **原始张数，非当量**
adv_n225_micro_kcontracts        日経225マイクロ先物（×¥10）ADV，千张/日 —— 2023-05 起才有
adv_n225_lgeq_kcontracts         日経225 复合体**大合约当量** = 大型 + mini/10 + micro/100，千张/日。
                                 这是 IR「Nikkei 225 Futures (Including mini and micro)」那一行的口径
adv_topix_futures_kcontracts     TOPIX先物（TOPIX×¥10,000）ADV，千张/日
adv_minitopix_futures_kcontracts ミニTOPIX先物（×¥1,000）ADV，千张/日 —— 原始张数
adv_topix_lgeq_kcontracts        TOPIX 复合体大合约当量 = 大型 + ミニ/10，千张/日
adv_n225_options_kcontracts      日経225オプション ADV（Put+Call **合計**行），千张/日。
                                 长表口径 = **不含**週次オプション，全程连续
adv_n225_mini_options_kcontracts 日経225ミニオプション ADV，千张/日 —— 2023-05 起才有，原始张数
adnv_n225_options_jpybn          日経225オプション**日均权利金成交金额**，十亿円/日。
                                 IR 报表用的就是这个而不是张数（期权张数受行权价密度影响）
adv_jgb10y_futures_kcontracts    長期国債先物（10年 JGB，額面 1 億円）ADV，千张/日
adv_secoptions_kcontracts        有価証券オプション（个股/ETF 期权）ADV，千张/日
oi_*_contracts                   **月末**未平仓，单位是**张**（不是千张），与 series/cme.csv 同量纲。
                                 期权那条只取「合計」行，见坑 3
ipo_public_offerings             当月「公募のうち新規公開」**件数**（家）。见坑 11，**不等于新上市数**
ipo_funds_jpybn                  当月「公募のうち新規公開」**募资额**，十亿円。同上

━━ series/jpx_investors.csv（周频，日股独有的外资流向）━━
「投資部門別売買状況 / Trading by Type of Investors」，**東証プライム**一个市场，
每条记录是一个**调查周**（不是日历周也不是日历月）。列口径见 `_INV_COLUMNS` 上方注释。
关键坑：官方的「総売買代金」是**売り+買い 双边合计**，绝不能和 `adt_cash_*` 放同一张表做除法；
只有净买入（買い−売り）不受双边计数影响，所以本表只出净额与占比。

━━ 依赖 ━━
openpyxl（读 xlsx）+ **xlrd**（读 tv_ts / historical-sikin / 投資部門別的老式 .xls BIFF，
openpyxl 读不了）。不依赖 pandas。
"""

import csv
import datetime
import os
import re
import urllib.request

import openpyxl
import xlrd

# ── 源站 ────────────────────────────────────────────────────────────────
_HOST = 'https://www.jpx.co.jp'
URL_SOUKATSU = (_HOST + '/automation/markets/statistics-derivatives/'
                'trading-volume/files/soukatsu_M.xlsx')
URL_LANDING_DERIV = (_HOST + '/english/markets/statistics-derivatives/'
                     'trading-volume/index.html')
URL_LANDING_CASH = (_HOST + '/english/markets/statistics-equities/misc/index.html')
URL_GENBUTSU = (_HOST + '/english/markets/statistics-equities/misc/'
                'tvdivq0000000vzk-att/historical-genbutsu.xlsx')
URL_TOUSHIN = (_HOST + '/english/markets/statistics-equities/misc/'
               'tvdivq0000000vzk-att/historical-toushin.xlsx')
URL_JIKA = (_HOST + '/english/markets/statistics-equities/misc/'
            'tvdivq0000001w3y-att/historical-jika.xlsx')
URL_SIKIN = (_HOST + '/english/markets/statistics-equities/misc/'
             'tvdivq0000001wij-att/historical-sikin.xls')
URL_INV_LANDING = (_HOST + '/english/markets/statistics-equities/'
                   'investor-type/index.html')

# UA 可省（实测不设也 200），写上是零成本保险 —— 哪天 JPX 前面挂了 WAF，
# 带常规 UA 至少不会因为 "Python-urllib" 这四个字被第一道规则筛掉。
_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

# 序列起点：长表里日経225 系（先物 / mini / オプション）最早只到 2014-12
# —— 从这里开始每一个门面列都有值，validate 不必为「早期天然为空」开一堆例外。
# 源文件本身回溯到 1985-01（现货）/ 1985-10（衍生品），要拉全史把这个常量往前挪即可，
# 但那样 2014-12 之前所有 adv_n225_* 都会是空列，validate 的例外表要同步加。
START = '2014-12'

# 商品関連 在 OSE 挂牌之前（旧 TOCOM 品种 2020-07-27 迁入）的月份，
# 宽表把它 pro-forma 回填进 JPX 口径。见坑 7。
CMDTY_PROFORMA_THROUGH = '2020-07'


# ── 大合约当量折算表 ─────────────────────────────────────────────────────
# key = 宽表/长表里的**日文产品名**（英文名会撞车，见坑 4），value = 除数。
# 每一档都来自 jpx.co.jp 的 Contract Specifications 页（.../01.html）的
# 「Contract Unit / 取引単位」栏，2026-08-06 逐个核过；校验锚点见坑 2。
# ⚠ 加新产品前先读坑 2 —— 白金那一档标准合约是 500g 不是 1kg，逆推会得到错的 ÷10。
_DIVISORS = {
    # 株価指数関連
    '日経225mini': 10,             # 指数×¥100    vs 日経225先物 ×¥1,000
    '日経225マイクロ先物': 100,      # 指数×¥10     vs 同上
    '日経225ミニオプション': 10,     # 権利金×¥100  vs 日経225オプション ×¥1,000
    'ミニTOPIX先物取引': 10,        # TOPIX×¥1,000 vs TOPIX先物 ×¥10,000
    # 国債・金利関連
    '長期国債先物（現金決済型ミニ）': 10,   # ¥10万×10年債先物価格 vs 額面 1 億円
    '超長期国債先物（ミニ）': 10,          # 額面 1,000 万円      vs 額面 1 億円
    # 商品関連（金は標準 1kg、白金は標準 500g —— 分母不一样）
    '金ミニ先物': 10,               # 100g vs 1kg
    '金限日先物': 10,               # 100g vs 1kg
    'ポケットゴールド100先物': 10,   # 100g vs 1kg
    '白金ミニ先物': 5,              # 100g vs 500g
    '白金限日先物': 5,              # 100g vs 500g
    'ポケットプラチナ100先物': 5,    # 100g vs 500g
}
# 长表里同一批产品的名字比宽表短（宽表爱带「取引」后缀），单独给一张表。
_DIVISORS_LONG = {
    '日経225mini': 10, '日経225マイクロ先物': 100, '日経225ミニオプション': 10,
    'ミニTOPIX先物': 10, '長期国債先物（現金決済型ミニ）': 10, '超長期国債先物（ミニ）': 10,
    '金ミニ先物': 10, '金限日先物': 10, 'ポケットゴールド100先物': 10,
    '白金ミニ先物': 5, '白金限日先物': 5, 'ポケットプラチナ100先物': 5,
}
# 名字里含 ミニ/マイクロ/ポケット/mini/micro 却**不是**小合约的产品。
# 目前只有一个：アル**ミニ**ウム先物 —— 子串匹配的经典假阳性。
# 这张白名单存在的意义是让下面那道「有新的小合约上市却没进折算表」的护栏能真的开着。
_NOT_A_MINI = {'アルミニウム先物'}
_MINI_HINTS = ('ミニ', 'マイクロ', 'ポケット', 'mini', 'micro', 'Mini', 'Micro')

# 宽表三大类的产品归属。**不是抄来的**：本模块每次运行都验证
# 「按这张表分组求和 == 官方三大类小计」在每一个月上成立（489/489 实测 diff 恰为 0），
# 不成立就抛异常 —— 所以这张表是被数据证明的，不是被相信的。
_RATES_PRODUCTS = {
    '中期国債先物取引', '長期国債先物取引', '長期国債先物（現金決済型ミニ）',
    '超長期国債先物（ミニ）', 'TONA3か月金利先物', '長期国債先物オプション取引',
    '中期国債先物オプション取引', '米国財務省証券先物',
}
_CMDTY_PRODUCTS = {
    '金標準先物', 'ポケットゴールド100先物', '金ミニ先物', '金限日先物', '銀先物',
    '白金標準先物', 'ポケットプラチナ100先物', '白金ミニ先物', '白金限日先物',
    'パラジウム先物', 'CME原油等指数先物', 'ゴム（RSS3）先物', 'ゴム（TSR20）先物',
    '上海天然ゴム先物', 'とうもろこし先物', '一般大豆先物', '小豆先物', '金先物オプション取引',
    'バージガソリン先物', 'バージ灯油先物', 'バージ軽油先物', 'プラッツドバイ原油先物',
    '東エリア・ベースロード電力先物', '西エリア・ベースロード電力先物',
    '中部エリア・ベースロード電力先物', '東エリア・日中ロード電力先物',
    '西エリア・日中ロード電力先物', '中部エリア・日中ロード電力先物',
    '東エリア・週間ベースロード電力先物', '西エリア・週間ベースロード電力先物',
    '東エリア・週間日中ロード電力先物', '西エリア・週間日中ロード電力先物',
    '東エリア・年度ベースロード電力先物', '西エリア・年度ベースロード電力先物',
    '中部エリア・年度ベースロード電力先物', '東エリア・年度日中ロード電力先物',
    '西エリア・年度日中ロード電力先物', '中部エリア・年度日中ロード電力先物',
    'LNG（プラッツJKM）先物', '中京ローリーガソリン先物', '中京ローリー灯油先物',
    'アルミニウム先物', '毛糸先物', '綿糸先物', '粗糖先物',
    '指数（限日）先物', '指数（限月）先物',
    'バージガソリンスワップ', 'ローリーガソリンスワップ', 'バージ灯油スワップ',
    'ローリー灯油スワップ', 'バージ軽油スワップ', 'ローリー軽油スワップ',
}
_W_TOTAL, _W_INDEX = '合計', '株価指数関連等'
_W_RATES, _W_CMDTY = '国債・金利関連', '商品関連'
_W_SUBTOTALS = (_W_TOTAL, _W_INDEX, _W_RATES, _W_CMDTY)


# ── 列定义 ──────────────────────────────────────────────────────────────
# (列名, 小数位)。小数位不是审美：源数据是整数（百万円 / 千株 / 张），
# 定点四舍五入才能保证「同样的输入 → 同样的字符串」，进而保证重跑字节级不变。
# 由整数换算过来的列（兆円←百万円、百万株←千株、十亿円←百万円）取的位数是**无损**的；
# 带除法的列（日均）取的位数远细于官方发布精度。
_INT = -1
COLUMNS = [
    ('trading_days', _INT),
    ('adt_cash_total_jpytn', 9),
    ('adt_cash_stocks_jpytn', 9),
    ('adt_cash_dom_stocks_jpytn', 9),
    ('adt_cash_etfreit_jpytn', 9),
    ('val_cash_total_jpytn', 6),
    ('val_cash_dom_stocks_jpytn', 6),
    ('adv_cash_dom_shares_mn', 6),
    ('vol_cash_dom_shares_mn', 3),
    ('mktcap_eom_jpytn', 6),
    ('deriv_trading_days', _INT),
    ('adv_deriv_total_lgeq_kcontracts', 6),
    ('adv_deriv_fin_lgeq_kcontracts', 6),
    ('adv_deriv_index_lgeq_kcontracts', 6),
    ('adv_deriv_rates_lgeq_kcontracts', 6),
    ('adv_deriv_cmdty_lgeq_kcontracts', 6),
    ('adv_deriv_total_raw_kcontracts', 6),
    ('adv_deriv_index_raw_kcontracts', 6),
    ('adv_deriv_rates_raw_kcontracts', 6),
    ('adv_deriv_cmdty_raw_kcontracts', 6),
    ('adnv_deriv_total_jpytn', 9),
    ('cmdty_proforma', _INT),
    ('adv_n225_futures_kcontracts', 6),
    ('adv_n225_mini_kcontracts', 6),
    ('adv_n225_micro_kcontracts', 6),
    ('adv_n225_lgeq_kcontracts', 6),
    ('adv_topix_futures_kcontracts', 6),
    ('adv_minitopix_futures_kcontracts', 6),
    ('adv_topix_lgeq_kcontracts', 6),
    ('adv_n225_options_kcontracts', 6),
    ('adv_n225_mini_options_kcontracts', 6),
    ('adnv_n225_options_jpybn', 6),
    ('adv_jgb10y_futures_kcontracts', 6),
    ('adv_secoptions_kcontracts', 6),
    ('oi_n225_futures_contracts', _INT),
    ('oi_n225_mini_contracts', _INT),
    ('oi_n225_micro_contracts', _INT),
    ('oi_topix_futures_contracts', _INT),
    ('oi_n225_options_contracts', _INT),
    ('oi_jgb10y_futures_contracts', _INT),
    ('ipo_public_offerings', _INT),
    ('ipo_funds_jpybn', 3),
]
COLUMN_NAMES = [c for c, _d in COLUMNS]
HEADER = ['month'] + COLUMN_NAMES

# 某些列的产品是后来才上市的，早期月份天然为空 —— 不是解析失败。
# 日期取自长表实测的首月（2026-08-06 期）。
FIRST_MONTH = {
    'adv_n225_micro_kcontracts': '2023-05',
    'oi_n225_micro_contracts': '2023-05',
    'adv_n225_mini_options_kcontracts': '2023-05',
}
# 资金調達額文件每月中旬才发，比主数据晚半个月 —— 这两列在最新月天然为空（坑 11）。
LAGGING_COLS = ('ipo_public_offerings', 'ipo_funds_jpybn')

# 钉死的定盘星。任何一条对不上 = 要么解析错了行/列，要么官方重述了历史 ——
# 两种都必须让人看见，不能静默写进 CSV。
# 值全部来自 2026-08-06 实测（见坑 2/3/5 里的证据），改这里之前先弄清楚为什么变了。
_ANCHORS = [
    # (月份, 列, 期望值, 容差)
    ('2026-06', 'oi_n225_options_contracts', 765660, 0),          # 坑 3：只取「合計」行
    ('2026-06', 'adt_cash_total_jpytn', 14.084475877, 1e-6),      # 与 IR 12.39 兆/日 季度口径互洽
    # 内国株单列。钉它是因为它与 adt_cash_stocks 只差「外国株」那一小段（2026-06 实测
    # 相对差 7.0 ppm），量级太近 —— 哪天解析把外国株段也累进 dom_val，页面上看不出任何
    # 异常，只有这个定盘星会响。dom < stocks 这条大小关系由下面的 _validate 逐月兜底。
    ('2026-06', 'adt_cash_dom_stocks_jpytn', 13.36509165, 1e-6),
    ('2026-06', 'mktcap_eom_jpytn', 1384.750323, 1e-6),
    ('2022-04', 'trading_days', 20, 0),                           # 坑 5：立会日数取 max 不是相加
    ('2016-06', 'trading_days', 22, 0),                           # 坑 5：侦察稿写成 21
]


class JpxFetchError(RuntimeError):
    """源站结构变化 / 下载失败 / 解析结果不完整 / 自检不过。一律炸掉，绝不静默写 NaN。"""


# ── 网络 ────────────────────────────────────────────────────────────────
def _http_get(url, timeout=90):
    """返回 (bytes, headers)。headers 里的 Last-Modified 进 source_dates 的 evidence。"""
    req = urllib.request.Request(url, headers={
        'User-Agent': _UA,
        'Accept': '*/*',
        'Accept-Language': 'en-US,en;q=0.9',
    })
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.read(), dict(r.headers)
    except Exception as e:                                # noqa: BLE001
        raise JpxFetchError('下载失败 %s: %r' % (url, e)) from e


def _download(cache_dir, url, fname):
    """下载到 cache/，返回 (本地路径, Last-Modified 原文字符串或 None)。"""
    os.makedirs(cache_dir, exist_ok=True)
    data, hdr = _http_get(url)
    path = os.path.join(cache_dir, fname)
    with open(path, 'wb') as f:
        f.write(data)
    return path, hdr.get('Last-Modified')


_SITE_UPDATE = re.compile(
    r'JPX-site-update"[^>]*>\s*Update\s*[:：]\s*'
    r'([A-Z][a-z]{2}\.?\s+\d{1,2},\s*\d{4})')
_TVTS_HREF = re.compile(r'(/english/markets/statistics-derivatives/trading-volume/'
                        r'[0-9a-z]+-att/tv_ts(\d{4})(\d{2})\.xls)')


def _page_update_date(html, url):
    """落地页自述的发布日 → ('YYYY-MM-DD', 出处原文)。读不懂就 (None, None)，绝不猜。

    evidence 里放的是页面上**逐字的**那串日期（"Jul. 07, 2026"，带点），不是重新拼的 ——
    将来有人怀疑这个日期时，他要能拿这串字去页面里 grep 得到。
    月名两种写法都要试：JPX 各个统计页并不统一（有的写 "Jul."，有的写 "July"）。
    """
    m = _SITE_UPDATE.search(html)
    if not m:
        return None, None
    raw = re.sub(r'\s+', ' ', m.group(1))
    for fmt in ('%b. %d, %Y', '%b %d, %Y', '%B %d, %Y'):
        try:
            d = datetime.datetime.strptime(raw, fmt)
        except ValueError:
            continue
        return d.strftime('%Y-%m-%d'), ('%s 的 <div class="JPX-site-update"> "Update : %s"'
                                        % (url, raw))
    return None, None                    # 认得出这行、读不懂日期 → 宁缺勿猜


def _discover_deriv(cache_dir):
    """从衍生品落地页拿 (宽表 url, 宽表报告月 'YYYY-MM', 页面自述发布日, 出处)。

    宽表的 token 每月都变，且猜错与没发布都返回 404、无法区分，所以只能走落地页。
    好处是**文件名里就带 YYYYMM**：latest_month() 光靠这张 28KB 的 HTML 就知道
    衍生品出到哪个月了，不必先拖 1.65MB 的 xls 下来。
    """
    data, _hdr = _http_get(URL_LANDING_DERIV)
    html = data.decode('utf-8', 'replace')
    with open(os.path.join(cache_dir, 'jpx_landing_deriv.html'), 'wb') as f:
        f.write(data)                       # 源站改版时可以事后取证
    hits = _TVTS_HREF.findall(html)
    if not hits:
        raise JpxFetchError('衍生品落地页上找不到 tv_ts{YYYYMM}.xls 直链，源站可能改版：'
                            + URL_LANDING_DERIV)
    href, y, mm = max(hits, key=lambda h: (h[1], h[2]))
    day, ev = _page_update_date(html, URL_LANDING_DERIV)
    return _HOST + href, '%s-%s' % (y, mm), day, ev


def _discover_cash_pubdate(cache_dir):
    data, _hdr = _http_get(URL_LANDING_CASH)
    html = data.decode('utf-8', 'replace')
    with open(os.path.join(cache_dir, 'jpx_landing_cash.html'), 'wb') as f:
        f.write(data)
    return _page_update_date(html, URL_LANDING_CASH)


# ── 解析：宽表 tv_ts{YYYYMM}.xls（总量 / 三大类 / 逐产品，含已退市品种）────────
def _wide_headers(sh):
    """{日文产品名: {'vol': col, 'val': col}}。

    表头三行：第 9 行是「日文名\\n英文名」的合并单元格（合并只在第一列有值，
    所以要沿列向右延续），第 10 行是 取引高 / 取引金額，第 12 行是单位。
    产品的子列数**不是恒定 2 个** —— 取引所外国為替証拠金取引 有三个（多一个立会日数）、
    米国財務省証券先物 只有一个（没有金额）—— 所以只能按第 10 行的文字分派，不能按偏移量。
    """
    prods, cur = {}, None
    for c in range(sh.ncols):
        g = str(sh.cell_value(8, c)).strip()
        sub = str(sh.cell_value(9, c)).strip()
        if g:
            cur = g.split('\n')[0].strip()
        if c < 2 or not cur:
            continue
        slot = prods.setdefault(cur, {})
        if sub.startswith('取引高'):
            slot['vol'] = c
        elif sub.startswith('取引金額'):
            slot['val'] = c
    return prods


def parse_wide(path):
    """解析宽表，返回 (prods, {'YYYY-MM': {'days':…, (产品,'vol'|'val'): 数}})。"""
    try:
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_name('月間 monthly')
    except Exception as e:                                # noqa: BLE001
        raise JpxFetchError('%s 打不开或没有「月間 monthly」sheet: %r'
                            % (os.path.basename(path), e)) from e
    prods = _wide_headers(sh)
    for key in _W_SUBTOTALS:
        if key not in prods:
            raise JpxFetchError('宽表里找不到官方小计列 %r，表结构可能变了' % key)
    rows = {}
    for r in range(12, sh.nrows):
        v = sh.cell_value(r, 0)
        if not isinstance(v, float) or v <= 0:
            continue                        # 表尾注记行
        d = datetime.datetime(*xlrd.xldate_as_tuple(v, wb.datemode))
        rec = {'days': _as_num(sh.cell_value(r, 1))}
        for name, slot in prods.items():
            for kind, col in slot.items():
                rec[(name, kind)] = _as_num(sh.cell_value(r, col))
        rows['%04d-%02d' % (d.year, d.month)] = rec
    if not rows:
        raise JpxFetchError('%s 解析后没有任何月份' % os.path.basename(path))
    return prods, dict(sorted(rows.items()))


def _as_num(v):
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _wide_product_list(prods):
    """参与求和的产品列：剔除四个官方小计列与所有 `【うち` 明细列（坑 4）。"""
    return [p for p in prods
            if p not in _W_SUBTOTALS and not p.startswith('【うち') and 'vol' in prods[p]]


def _check_wide_closure(prods, rows):
    """三条闭合式，每次运行都重算。这是本模块最重要的一道自检。

    1) 三大类小计相加 == 官方 Total
    2) 逐产品（剔除【うち】）相加 == 官方 Total —— 实测 489/489 月 diff 恰为 0
    3) 按 _RATES_PRODUCTS / _CMDTY_PRODUCTS 分组相加 == 官方三大类小计
    第 3 条同时也是「分组表是对的」的**证明**：JPX 从不公布产品→大类的映射，
    这张表是人写的；能在 489 个月上逐格闭合，它就不再是「相信」而是「被证明」。
    另外顺手检查：有没有新的小合约上市却没进 _DIVISORS（那会让当量列悄悄虚高）。
    """
    plist = _wide_product_list(prods)
    grp = {p: ('rates' if p in _RATES_PRODUCTS else
               'cmdty' if p in _CMDTY_PRODUCTS else 'index') for p in plist}
    bad = []
    for mon, rec in rows.items():
        tot = rec.get((_W_TOTAL, 'vol')) or 0.0
        sub = sum(rec.get((k, 'vol')) or 0.0 for k in (_W_INDEX, _W_RATES, _W_CMDTY))
        if abs(sub - tot) > 1e-6:
            bad.append('%s 三大类之和 %r != 官方 Total %r' % (mon, sub, tot))
        psum = sum(rec.get((p, 'vol')) or 0.0 for p in plist)
        if abs(psum - tot) > 1e-6:
            bad.append('%s 逐产品之和 %r != 官方 Total %r' % (mon, psum, tot))
        for g, col in (('index', _W_INDEX), ('rates', _W_RATES), ('cmdty', _W_CMDTY)):
            s = sum(rec.get((p, 'vol')) or 0.0 for p in plist if grp[p] == g)
            o = rec.get((col, 'vol')) or 0.0
            if abs(s - o) > 1e-6:
                bad.append('%s %s 分组和 %r != 官方小计 %r' % (mon, col, s, o))
        if len(bad) > 6:
            break
    if bad:
        raise JpxFetchError('宽表闭合自检失败（产品分类表或表结构已变），前几条：\n  '
                            + '\n  '.join(bad))

    unknown = [p for p in _DIVISORS if p not in prods]
    if unknown:
        raise JpxFetchError('折算表 _DIVISORS 里这些产品在宽表里已经不存在（改名或退市）：%s'
                            ' —— 大合约当量会算错，先去官方合约规格页核对再改' % unknown)
    missed = [p for p in plist
              if p not in _DIVISORS and p not in _NOT_A_MINI
              and any(h in p for h in _MINI_HINTS)]
    if missed:
        raise JpxFetchError('宽表里出现了名字像小合约、却不在 _DIVISORS 里的产品：%s'
                            ' —— 去官方合约规格页查它的取引単位，要么加折算系数，'
                            '要么加进 _NOT_A_MINI 白名单' % missed)
    return plist, grp


# ── 解析：长表 soukatsu_M.xlsx（逐产品 ADV / 月末 OI）────────────────────
# 列（1-based）：1 商品等(日) 2 Products(英) 3 プットコール区分 5 年 6 月 7 立会日数
# 8 取引高合計 11 取引高一日平均 17 取引金額合計 20 取引金額一日平均 28 建玉月末
_L = {'jp': 0, 'pc': 2, 'y': 4, 'm': 5, 'days': 6,
      'vol': 7, 'advol': 10, 'val': 16, 'adval': 19, 'oi': 27}


def parse_long(path):
    """{日文产品名: {'YYYY-MM': {...}}}。期权只取「合計」行（坑 3）。"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = str(ws.cell(3, 3).value or '')
    if 'プットコール' not in hdr:
        wb.close()
        raise JpxFetchError('长表第 3 行第 3 列不是「プットコール区分」（拿到 %r）—— '
                            '列位置变了，期权 Put/Call 三行制的过滤会失效' % hdr)
    out = {}
    try:
        for row in ws.iter_rows(min_row=7, values_only=True):
            jp, y, m = row[_L['jp']], row[_L['y']], row[_L['m']]
            if jp is None or y is None or m is None:
                continue                     # 只有产品名、一格数据都没有的占位行
            if row[_L['pc']] in ('プット', 'コール'):
                continue                     # ← 坑 3：不过滤会让成交量与 OI 双双翻倍
            out.setdefault(str(jp).strip(), {})['%04d-%02d' % (int(y), int(m))] = {
                k: _as_num(row[_L[k]]) for k in ('days', 'vol', 'advol', 'val', 'adval', 'oi')}
    finally:
        wb.close()
    if not out:
        raise JpxFetchError('%s 解析后没有任何产品' % os.path.basename(path))
    return out


def _long_get(long, product, mon, field):
    if product not in long:
        raise JpxFetchError('长表里找不到产品 %r —— 官方改名或下市了，'
                            '逐产品列会整列变空，拒绝继续' % product)
    rec = long[product].get(mon)
    return None if rec is None else rec[field]


# ── 解析：现货 ───────────────────────────────────────────────────────────
# (取引高列, 取引金額列, 取引高倍率→千株, 取引金額倍率→百万円)
# TOKYO PRO 段的单位是 株 / 千円，其余段是 千株 / 百万円（坑 6a）。
_EQ_NEW_DOM = [(3, 5, 1, 1), (7, 9, 1, 1), (11, 13, 1, 1), (15, 17, 0.001, 0.001)]
_EQ_NEW_FOR = [(23, 24, 1, 1), (25, 26, 1, 1), (27, 28, 1, 1)]
_EQ_OLD_DOM = [(3, 5, 1, 1), (7, 9, 1, 1), (11, 13, 1, 1), (15, 17, 1, 1),
               (19, 21, 1, 1), (23, 25, 0.001, 0.001)]
_EQ_OLD_FOR = [(31, 32, 1, 1)]
_EQ_SHEETS = (('月間 monthly', 11, _EQ_NEW_DOM, _EQ_NEW_FOR),
              ('月間 monthly (旧 Old)', 12, _EQ_OLD_DOM, _EQ_OLD_FOR))


def _cell_month(ws, r):
    v = ws.cell(r, 1).value
    return '%04d-%02d' % (v.year, v.month) if isinstance(v, datetime.datetime) else None


def _cell_num(ws, r, c):
    v = ws.cell(r, c).value
    return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def parse_genbutsu(path):
    """内国株 + 外国株的月度成交量额。返回 {'YYYY-MM': {...}}，单位 千株 / 百万円。

    新旧两张 sheet 在 2022-04 重叠（市场重组，见坑 5）：
      · 金额与股数**相加**（旧表那行是 4/1 一天，新表那行是 4/4–4/28）
      · 立会日数取 **max**（新表那个 20 是「4 月的立会日数」，不是「本行覆盖的日数」）
    重叠月份只允许是已知的那一个；将来再出现新的重叠一定是官方又改了市场结构，
    那时相加/取 max 这两条规则都得重新验证，所以这里直接抛异常而不是默默照做。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    out, seen = {}, {}
    try:
        for sn, hdr, dom, forg in _EQ_SHEETS:
            if sn not in wb.sheetnames:
                raise JpxFetchError('%s 里没有 sheet %r（拿到 %r）'
                                    % (os.path.basename(path), sn, wb.sheetnames))
            ws = wb[sn]
            for r in range(hdr + 1, ws.max_row + 1):
                mon = _cell_month(ws, r)
                if mon is None:
                    continue
                seen.setdefault(mon, []).append(sn)
                rec = out.setdefault(mon, {'days': 0.0, 'dom_vol': 0.0, 'dom_val': 0.0,
                                           'for_vol': 0.0, 'for_val': 0.0})
                rec['days'] = max(rec['days'], _cell_num(ws, r, 2) or 0.0)
                for cv, cl, sv, sl in dom:
                    rec['dom_vol'] += (_cell_num(ws, r, cv) or 0.0) * sv
                    rec['dom_val'] += (_cell_num(ws, r, cl) or 0.0) * sl
                for cv, cl, sv, sl in forg:
                    rec['for_vol'] += (_cell_num(ws, r, cv) or 0.0) * sv
                    rec['for_val'] += (_cell_num(ws, r, cl) or 0.0) * sl
    finally:
        wb.close()
    dup = sorted(m for m, s in seen.items() if len(s) > 1)
    if dup != ['2022-04']:
        raise JpxFetchError('现货新旧两 sheet 的重叠月份是 %s，本模块只认 2022-04 —— '
                            '官方又改市场结构了，相加/取 max 两条规则要重新验证' % dup)
    if not out:
        raise JpxFetchError('%s 解析后没有任何月份' % os.path.basename(path))
    return dict(sorted(out.items()))


def parse_toushin(path):
    """内国ETF + 外国ETF・ETN + 内国投資証券(REIT)，单位 千口 / 百万円。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if '月間 monthly' not in wb.sheetnames:
            raise JpxFetchError('%s 里没有「月間 monthly」（拿到 %r）'
                                % (os.path.basename(path), wb.sheetnames))
        ws = wb['月間 monthly']
        out = {}
        for r in range(11, ws.max_row + 1):
            mon = _cell_month(ws, r)
            if mon is None:
                continue
            out[mon] = {'days': _cell_num(ws, r, 2),
                        'val': sum(_cell_num(ws, r, c) or 0.0 for c in (4, 6, 8)),
                        'vol': sum(_cell_num(ws, r, c) or 0.0 for c in (3, 5, 7))}
    finally:
        wb.close()
    if not out:
        raise JpxFetchError('%s 解析后没有任何月份' % os.path.basename(path))
    return dict(sorted(out.items()))


def parse_jika(path):
    """月末时价总额（合計列），单位 百万円。两张 sheet 无重叠（旧 …2022-03 / 新 2022-04…）。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            tcol = next((c for c in range(1, ws.max_column + 1)
                         if str(ws.cell(3, c).value or '').startswith('合計')), None)
            if tcol is None:
                raise JpxFetchError('时价总额 sheet %r 第 3 行找不到「合計」列' % sn)
            for r in range(4, ws.max_row + 1):
                mon = _cell_month(ws, r)
                v = _cell_num(ws, r, tcol)
                if mon and v is not None:
                    if mon in out and abs(out[mon] - v) > 1e-6:
                        raise JpxFetchError('时价总额 %s 在两张 sheet 里不一致（%r vs %r）'
                                            % (mon, out[mon], v))
                    out[mon] = v
    finally:
        wb.close()
    if not out:
        raise JpxFetchError('%s 解析后没有任何月份' % os.path.basename(path))
    return dict(sorted(out.items()))


# 資金調達額（月間）：col3 公募件数 / col4 うち新規公開件数 / col5 公募調達額 / col6 うち新規公開調達額
_SIKIN_IPO_CASES, _SIKIN_IPO_AMOUNT = 4, 6


def parse_sikin(path):
    """{'YYYY-MM': (IPO 公募件数, IPO 募资额 百万円)}。'－' 表示当月没有，记 0。"""
    try:
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_name('月間 monthly')
    except Exception as e:                                # noqa: BLE001
        raise JpxFetchError('%s 打不开或没有「月間 monthly」: %r'
                            % (os.path.basename(path), e)) from e
    if 'うち' not in str(sh.cell_value(8, _SIKIN_IPO_CASES)):
        raise JpxFetchError('資金調達額表第 9 行第 %d 列不是「うち新規公開」（拿到 %r）—— '
                            '列位置变了' % (_SIKIN_IPO_CASES + 1,
                                            sh.cell_value(8, _SIKIN_IPO_CASES)))
    out = {}
    for r in range(9, sh.nrows):
        v = sh.cell_value(r, 0)
        if not isinstance(v, float) or v <= 0:
            continue
        d = datetime.datetime(*xlrd.xldate_as_tuple(v, wb.datemode))
        cases = _as_num(sh.cell_value(r, _SIKIN_IPO_CASES))
        amt = _as_num(sh.cell_value(r, _SIKIN_IPO_AMOUNT))
        out['%04d-%02d' % (d.year, d.month)] = (0.0 if cases is None else cases,
                                                0.0 if amt is None else amt)
    if not out:
        raise JpxFetchError('%s 解析后没有任何月份' % os.path.basename(path))
    return dict(sorted(out.items()))


# ── 组装 ────────────────────────────────────────────────────────────────
def _adv(vol, days, div=1.0):
    """张 → 千张/日。任何一项缺就返回 None（不拿 0 冒充）。"""
    if vol is None or not days:
        return None
    return vol / div / days / 1000.0


def build_records(wide, long, genbutsu, toushin, jika, sikin, prods):
    """把五个源拼成 {'YYYY-MM': {列名: 值|None}}，只保留 START 之后、现货与衍生品都有的月份。"""
    plist, grp = _check_wide_closure(prods, wide)
    div_w = {p: _DIVISORS.get(p, 1) for p in plist}

    months = sorted(m for m in set(genbutsu) & set(wide) if m >= START)
    if not months:
        raise JpxFetchError('现货与衍生品没有任何共同月份 >= %s' % START)
    out = {}
    for mon in months:
        g, w = genbutsu[mon], wide[mon]
        t = toushin.get(mon, {'val': 0.0, 'vol': 0.0})
        days, ddays = g['days'], w['days']
        if not days or not ddays:
            raise JpxFetchError('%s 立会日数为 0 或缺失（现货 %r / 衍生品 %r）'
                                % (mon, days, ddays))
        stocks_val = g['dom_val'] + g['for_val']           # 百万円
        total_val = stocks_val + t['val']

        lgeq = {'index': 0.0, 'rates': 0.0, 'cmdty': 0.0}
        for p in plist:
            v = w.get((p, 'vol'))
            if v:
                lgeq[grp[p]] += v / div_w[p]
        lgeq_total = sum(lgeq.values())

        n225 = [_long_get(long, k, mon, 'vol')
                for k in ('日経225先物', '日経225mini', '日経225マイクロ先物')]
        n225_lgeq = None
        if n225[0] is not None and n225[1] is not None:
            n225_lgeq = n225[0] + n225[1] / 10.0 + (n225[2] or 0.0) / 100.0
        topix = [_long_get(long, k, mon, 'vol') for k in ('TOPIX先物', 'ミニTOPIX先物')]
        topix_lgeq = None
        if topix[0] is not None:
            topix_lgeq = topix[0] + (topix[1] or 0.0) / 10.0
        ldays = _long_get(long, '長期国債先物', mon, 'days') or ddays

        rec = {
            'trading_days': days,
            'adt_cash_total_jpytn': total_val / 1e6 / days,
            'adt_cash_stocks_jpytn': stocks_val / 1e6 / days,
            # 内国株**单独**一列。为什么必须有它：下游要做「成交额 = 成交股数 × 均价」的
            # 分解，而股数列（adv/vol_cash_dom_shares_mn）数的是 g['dom_vol'] —— 只有内国株。
            # 用 adt_cash_stocks（内国 + 外国）当分子，就是拿含外国株的金额去除不含外国株的
            # 股数，造出来的「均价」带一个既不知道方向也不知道大小的偏差，而且在图上完全
            # 看不出来。这一列与股数列同出 g['dom_*']、同为立会市場 + ToSTNeT、同为单边计，
            # 是本表里**唯一**与股数列逐段对齐的金额口径。
            'adt_cash_dom_stocks_jpytn': g['dom_val'] / 1e6 / days,
            'adt_cash_etfreit_jpytn': t['val'] / 1e6 / days,
            'val_cash_total_jpytn': total_val / 1e6,
            # 同上的月总量口径（忠实入库，只做百万円 → 兆円的单位换算，不除立会日数）。
            # 与 vol_cash_dom_shares_mn 配成一对，年度聚合时不必再乘立会日数去还原。
            'val_cash_dom_stocks_jpytn': g['dom_val'] / 1e6,
            'adv_cash_dom_shares_mn': g['dom_vol'] / 1e3 / days,
            'vol_cash_dom_shares_mn': g['dom_vol'] / 1e3,
            'mktcap_eom_jpytn': (jika[mon] / 1e6) if mon in jika else None,
            'deriv_trading_days': ddays,
            'adv_deriv_total_lgeq_kcontracts': lgeq_total / ddays / 1000.0,
            'adv_deriv_fin_lgeq_kcontracts': (lgeq['index'] + lgeq['rates']) / ddays / 1000.0,
            'adv_deriv_index_lgeq_kcontracts': lgeq['index'] / ddays / 1000.0,
            'adv_deriv_rates_lgeq_kcontracts': lgeq['rates'] / ddays / 1000.0,
            'adv_deriv_cmdty_lgeq_kcontracts': lgeq['cmdty'] / ddays / 1000.0,
            'adv_deriv_total_raw_kcontracts': _adv(w.get((_W_TOTAL, 'vol')), ddays),
            'adv_deriv_index_raw_kcontracts': _adv(w.get((_W_INDEX, 'vol')), ddays),
            'adv_deriv_rates_raw_kcontracts': _adv(w.get((_W_RATES, 'vol')), ddays),
            'adv_deriv_cmdty_raw_kcontracts': _adv(w.get((_W_CMDTY, 'vol')), ddays),
            'adnv_deriv_total_jpytn': (None if w.get((_W_TOTAL, 'val')) is None
                                       else w[(_W_TOTAL, 'val')] / ddays / 1e12),
            'cmdty_proforma': 1 if mon <= CMDTY_PROFORMA_THROUGH else None,
            'adv_n225_futures_kcontracts': _adv(n225[0], ldays),
            'adv_n225_mini_kcontracts': _adv(n225[1], ldays),
            'adv_n225_micro_kcontracts': _adv(n225[2], ldays),
            'adv_n225_lgeq_kcontracts': _adv(n225_lgeq, ldays),
            'adv_topix_futures_kcontracts': _adv(topix[0], ldays),
            'adv_minitopix_futures_kcontracts': _adv(topix[1], ldays),
            'adv_topix_lgeq_kcontracts': _adv(topix_lgeq, ldays),
            'adv_n225_options_kcontracts': _adv(_long_get(long, '日経225オプション', mon, 'vol'), ldays),
            'adv_n225_mini_options_kcontracts': _adv(
                _long_get(long, '日経225ミニオプション', mon, 'vol'), ldays),
            'adnv_n225_options_jpybn': _premium_adnv(long, mon, ldays),
            'adv_jgb10y_futures_kcontracts': _adv(_long_get(long, '長期国債先物', mon, 'vol'), ldays),
            'adv_secoptions_kcontracts': _adv(_long_get(long, '有価証券オプション', mon, 'vol'), ldays),
            'oi_n225_futures_contracts': _long_get(long, '日経225先物', mon, 'oi'),
            'oi_n225_mini_contracts': _long_get(long, '日経225mini', mon, 'oi'),
            'oi_n225_micro_contracts': _long_get(long, '日経225マイクロ先物', mon, 'oi'),
            'oi_topix_futures_contracts': _long_get(long, 'TOPIX先物', mon, 'oi'),
            'oi_n225_options_contracts': _long_get(long, '日経225オプション', mon, 'oi'),
            'oi_jgb10y_futures_contracts': _long_get(long, '長期国債先物', mon, 'oi'),
            'ipo_public_offerings': sikin[mon][0] if mon in sikin else None,
            'ipo_funds_jpybn': (sikin[mon][1] / 1e3) if mon in sikin else None,
        }
        out[mon] = rec
    return out


def _premium_adnv(long, mon, days):
    """日経225オプション 日均权利金成交金额 → 十亿円/日。"""
    v = _long_get(long, '日経225オプション', mon, 'val')
    return None if v is None or not days else v / days / 1e9


def _validate(data, sikin_latest):
    """缺列一律失败（README 护栏 2），只放行有据可查的空。返回最新月。"""
    months = sorted(data)
    newest = months[-1]
    for mon in months:
        rec = data[mon]
        missing = []
        for name, _dec in COLUMNS:
            if rec.get(name) is not None:
                continue
            if name == 'cmdty_proforma':
                continue                                   # 空 = 非 pro-forma，是正常取值
            if mon < FIRST_MONTH.get(name, START):
                continue                                   # 产品当时还没上市
            if name in LAGGING_COLS and mon > sikin_latest:
                continue                                   # 资金調達額每月中旬才发（坑 11）
            missing.append(name)
        if missing:
            raise JpxFetchError('%s 缺列 %s —— 解析异常或官方表结构变了，拒绝写入'
                                % (mon, missing))
        # 现货三个口径的包含关系：内国株 ⊆ 内国株+外国株 ⊆ 再加 ETF/REIT。
        # 这是分段求和的结构性后果，不是经验规律 —— 它一旦不成立，就说明某一段被累进了
        # 错误的桶（列位置漂移是本表最常见的故障，见坑 6a 的单位陷阱）。
        # 下游的量价分解正是靠「dom 那一列只含内国株」才敢与股数列相除，所以这条必须响。
        d, s, t = (rec.get('adt_cash_dom_stocks_jpytn'),
                   rec.get('adt_cash_stocks_jpytn'), rec.get('adt_cash_total_jpytn'))
        if None not in (d, s, t) and not (d <= s + 1e-9 <= t + 2e-9):
            raise JpxFetchError(
                '%s 现货口径包含关系不成立：内国株 %r > 股票合计 %r 或 股票合计 > 合计 %r '
                '—— 分段累加进错了桶' % (mon, d, s, t))
    for mon, col, want, tol in _ANCHORS:
        if mon not in data:
            continue
        got = data[mon][col]
        if got is None or abs(float(got) - want) > tol:
            raise JpxFetchError(
                '定盘星对不上：%s %s 期望 %r（±%r），实得 %r。'
                '要么解析取错了行/列，要么官方重述了历史 —— 两种都必须人来看一眼'
                % (mon, col, want, tol, got))
    return newest


# ── 发布日 ───────────────────────────────────────────────────────────────
def _source_dates():
    """按路径加载仓库根的 source_dates.py。

    不能裸 import：本模块被 monthly_run 用 spec_from_file_location 加载，
    那时 sys.path 上既没有 fetch/ 也没有仓库根。
    """
    import importlib.util
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    spec = importlib.util.spec_from_file_location(
        'source_dates', os.path.join(root, 'source_dates.py'))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _record_source_date(series_dir, month, cash, deriv):
    """把这一期官方自述的发布日记进 series/source_dates.csv。

    cash / deriv 各是 (自述日, 出处原文, 文件 Last-Modified 原文)。
    取两者里**较晚**的那个：一行 JPX 数据要现货与衍生品都发齐才成立。
    evidence 里四条观测（两个页面自述 + 两个文件 Last-Modified）**并列记录，不写「互证」**
    —— 实测 JPX 的 Monthly Statistics Report 落地页写 Jul. 23、它挂的 PDF 是 Jul. 21，
    页面日与文件日本来就可以不同日，写成互证等于凭空发明一条证据。
    已有记录一律不覆盖：这几个文件是原地覆盖式发布，下个月再读 Last-Modified 就是
    下个月的日子了，覆盖会把首发日悄悄改成某个更晚的日期，而页面照印不误。
    """
    days = [d for d, _e, _lm in (cash, deriv) if d]
    if not days:
        return
    sd = _source_dates()
    if sd.lookup(series_dir, 'jpx', month):
        return
    ev = '；'.join(
        [e for _d, e, _lm in (cash, deriv) if e]
        + ['%s 的 HTTP Last-Modified "%s"' % (u, lm)
           for u, (_d, _e, lm) in (('historical-genbutsu.xlsx', cash),
                                   ('tv_ts%s.xls' % month.replace('-', ''), deriv)) if lm])
    sd.record(series_dir, 'jpx', month, max(days), ev)


# ── CSV 读写 ────────────────────────────────────────────────────────────
def _fmt(value, dec):
    if value is None:
        return ''
    if dec == _INT:
        iv = int(round(float(value)))
        if abs(float(value) - iv) > 1e-6:
            raise JpxFetchError('本该是整数的值不是整数：%r' % value)
        return str(iv)
    return ('%.*f' % (dec, float(value))).rstrip('0').rstrip('.') or '0'


def _read_csv(path):
    if not os.path.exists(path):
        return list(HEADER), []
    with open(path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    if not rows:
        return list(HEADER), []
    return rows[0], [r for r in rows[1:] if r and r[0].strip()]


def _write_csv(path, header, body):
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    body.sort(key=lambda r: r[0])
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(body)


# ── 对外接口 ─────────────────────────────────────────────────────────────
def latest_month(cache_dir):
    """官方源当前最新月 'YYYY-MM' —— 现货与衍生品**都**发齐的那个月。

    不看 historical-jika.xlsx：它比成交额早 4 天，跟着它走会建出只有 mktcap 的空行。
    衍生品那半边只读落地页（28KB）就够 —— 文件名 tv_ts{YYYYMM}.xls 自带月份，
    不必为了问一句「出到几月了」把 1.65MB 的表拖下来。
    抓不到 / 解析不出来一律抛 JpxFetchError，不返回 None 掩盖故障。
    """
    os.makedirs(cache_dir, exist_ok=True)
    _url, deriv_month, _day, _ev = _discover_deriv(cache_dir)
    path, _lm = _download(cache_dir, URL_GENBUTSU, 'jpx_historical-genbutsu.xlsx')
    cash = parse_genbutsu(path)
    return min(max(cash), deriv_month)


def update(series_dir, cache_dir):
    """把新月份追加进 series/jpx.csv，返回新增月份列表（升序）。

    顺带把 series/jpx_investors.csv（周频外资流向）也更新一遍 —— 两张表同源同一次抓取，
    分两个文件是因为频率不同、口径也不同（周频那张是双边计数，见文件头）。

    幂等保证：
      · 已存在的月份不重复追加；
      · 已经有值的单元格**永不覆盖**（官方明说资金調達額会回溯订正，不由本模块自动吞进来）；
      · 只对既有行里**原本为空**的单元格做回补 —— 不做的话，因为资金調達額晚半个月
        而留白的 ipo_* 会永久为空（回补不计入返回值，它不是新月份）；
      · 未被触碰的单元格是原样字符串搬运，所以「什么都没变」时文件字节级不变。
    """
    os.makedirs(cache_dir, exist_ok=True)
    tv_url, deriv_month, deriv_day, deriv_ev = _discover_deriv(cache_dir)
    cash_day, cash_ev = _discover_cash_pubdate(cache_dir)

    p_wide, lm_wide = _download(cache_dir, tv_url, os.path.basename(tv_url))
    p_long, _ = _download(cache_dir, URL_SOUKATSU, 'jpx_soukatsu_M.xlsx')
    p_gen, lm_gen = _download(cache_dir, URL_GENBUTSU, 'jpx_historical-genbutsu.xlsx')
    p_tou, _ = _download(cache_dir, URL_TOUSHIN, 'jpx_historical-toushin.xlsx')
    p_jik, _ = _download(cache_dir, URL_JIKA, 'jpx_historical-jika.xlsx')
    p_sik, _ = _download(cache_dir, URL_SIKIN, 'jpx_historical-sikin.xls')

    prods, wide = parse_wide(p_wide)
    long = parse_long(p_long)
    genbutsu = parse_genbutsu(p_gen)
    toushin = parse_toushin(p_tou)
    jika = parse_jika(p_jik)
    sikin = parse_sikin(p_sik)

    if deriv_month not in wide:
        raise JpxFetchError('宽表文件名说它是 %s 期，表里却没有这个月 —— '
                            '落地页与文件对不上' % deriv_month)

    data = build_records(wide, long, genbutsu, toushin, jika, sikin, prods)
    newest = _validate(data, max(sikin))

    csv_path = os.path.join(series_dir, 'jpx.csv')
    header, body = _read_csv(csv_path)
    idx = {name: i for i, name in enumerate(header)}
    unknown = [c for c in COLUMN_NAMES if c not in idx]
    if unknown:
        raise JpxFetchError('series/jpx.csv 里没有这些列：%s' % unknown)

    have = {r[0]: r for r in body}
    added = []
    for mon in sorted(data):
        rec = data[mon]
        if mon in have:
            row = have[mon]
            for name, dec in COLUMNS:                      # 只填空，不覆盖
                if not row[idx[name]].strip() and rec[name] is not None:
                    row[idx[name]] = _fmt(rec[name], dec)
            continue
        row = [''] * len(header)
        row[0] = mon
        for name, dec in COLUMNS:
            row[idx[name]] = _fmt(rec[name], dec)
        have[mon] = row
        body.append(row)
        added.append(mon)
    _write_csv(csv_path, header, body)

    # 记发布日放在落盘之后：写盘失败就不该留下「这个月官方发过了」这条断言。
    _record_source_date(series_dir, newest,
                        (cash_day, cash_ev, lm_gen), (deriv_day, deriv_ev, lm_wide))

    update_investors(series_dir, cache_dir)
    return sorted(added)


# ══ 周频：投資部門別売買状況（東証プライム）══════════════════════════════
# 日股独有的高频资金流，其余三家交易所（CME / Cboe / HKEX）都没有对照物。
#
# 每一行是一个**调查周**，不是日历周也不是日历月：官方按完整交易周汇总，
# 月末不足一周的交易日会被推到下一份文件里。所以周期原文（week_start / week_end）
# 必须入库，别拿月份去 group by。
#
# 官方表里的「総売買代金」是 **売り + 買い 双边合计**（実測 2026-07 プライム 125.2 兆円/周
# ≈ 单边 ADT × 2 × 交易日数），**绝不能**和 series/jpx.csv 的 adt_cash_* 放一起做除法。
# 净买入（買い − 売り）不受双边计数影响，所以本表只出净额与占比，不出成交额本身。
#
# 列口径：
#   week_start / week_end        调查周的首末交易日（YYYY-MM-DD），取自表内周期原文
#   net_foreign_jpybn            海外投資家 净买入（買い−売り），十亿円。负数 = 净卖出
#   net_individual_jpybn         個人 净买入
#   net_trust_jpybn              投資信託 净买入
#   net_bizcorp_jpybn            事業法人 净买入（自社株買いの代理指标）
#   net_proprietary_jpybn        自己計（券商自营）净买入
#   foreign_share_pct            海外投資家 売買代金合計 ÷ 委託計 売買代金合計 × 100
#   src_file                     这一行是从哪个官方文件解析出来的（去重台账，见下）
#
# 为什么要 src_file：每份周文件同时印**两周**（本周 + 上周），所以同一周会出现在两份文件里；
# 而「这份文件我抓过没有」无法从周期日期反推（文件名是「2026年7月第5週」这种周序号，
# 不是日期）。落一列文件名，重跑时就能只下没抓过的那些，不必每月把整年重下一遍。
_INV_YEARS_BACK = 3        # 首次建库回溯几年。调大到 11 可以回到 2016（约 570 个文件，
                           # 一次性跑几分钟）；日常增量只会下载落地页上那 5 个新文件。
INV_HEADER = ['week_end', 'week_start', 'net_foreign_jpybn', 'net_individual_jpybn',
              'net_trust_jpybn', 'net_bizcorp_jpybn', 'net_proprietary_jpybn',
              'foreign_share_pct', 'src_file']
_INV_SHEET = 'TSE Prime'
# 行标签（列 A 的日文名，去掉全角空格后比较）→ 输出列名
_INV_ROWS = {
    '海外投資家': 'net_foreign_jpybn',
    '個人': 'net_individual_jpybn',
    '投資信託': 'net_trust_jpybn',
    '事業法人': 'net_bizcorp_jpybn',
    '自己計': 'net_proprietary_jpybn',
}
_INV_FILE = re.compile(r'href="([^"]*/stock_val_1_(\d{6})\.xls)"')
_INV_ARCHIVE = re.compile(r'"(/english/markets/statistics-equities/investor-type/'
                          r'00-00-archives-\d\d\.html)"')
_INV_PERIOD = re.compile(r'^\s*(\d{1,2})/(\d{1,2})\s*[～~-]\s*(\d{1,2})/(\d{1,2})\s*$')
_INV_HEAD = re.compile(r'(\d{4})年(\d{1,2})月')


def _inv_norm(s):
    return re.sub(r'[\s　]', '', str(s or ''))


def _inv_year(head_year, head_month, month):
    """周期里的 MM 属于哪一年：与抬头月差半年以上才算跨年。"""
    delta = month - head_month
    if delta >= 6:
        return head_year - 1
    if delta <= -6:
        return head_year + 1
    return head_year


def _inv_cell(sh, r, c):
    """表里的数字是带千分位的**字符串**，不是数值单元格。"""
    v = sh.cell_value(r, c)
    if isinstance(v, float):
        return v
    s = str(v).replace(',', '').strip()
    if s in ('', '－', '-', '―'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_investors(blob, fname):
    """解析一份周次 xls，返回 [{列: 值}]（一份文件带两周）。单位 千円 → 十亿円。"""
    try:
        wb = xlrd.open_workbook(file_contents=blob)
        sh = wb.sheet_by_name(_INV_SHEET)
    except Exception as e:                                # noqa: BLE001
        raise JpxFetchError('%s 里没有 sheet %r（2022-04 市场重组之前叫 "TSE 1st"，'
                            '要回溯到那之前得先加段别映射表）: %r' % (fname, _INV_SHEET, e)) from e
    mh = _INV_HEAD.search(str(sh.cell_value(3, 0)))
    if not mh:
        raise JpxFetchError('%s 第 4 行读不出「YYYY年M月」抬头（拿到 %r）'
                            % (fname, sh.cell_value(3, 0)))
    hy, hm = int(mh.group(1)), int(mh.group(2))
    # 行定位：每个投資部門占 売り / 買い / 合計 三行，日文部门名**只写在「売り」那一行**，
    # 「買い」那一行的列 A 放的是英文名（Foreigners / Individuals / …）。所以只能在
    # side == '売り' 时换段，见了非空标签就换段会把段名换成英文、后两行整段丢失。
    rows = {}
    cur = None
    for r in range(11, sh.nrows):
        side = _inv_norm(sh.cell_value(r, 1))
        if side == '売り':
            cur = _inv_norm(sh.cell_value(r, 0)) or None
        if cur and side in ('売り', '買い', '合計'):
            rows[(cur, side)] = r
    need = [(k, s) for k in list(_INV_ROWS) + ['委託計'] for s in ('売り', '買い', '合計')]
    miss = [k for k in need if k not in rows]
    if miss:
        raise JpxFetchError('%s 找不到这些行：%s —— 官方表结构变了' % (fname, miss))

    out = []
    for pcol, vcol in ((3, 4), (7, 8)):                    # 两周并排：周期在 pcol，金额在 vcol
        mp = _INV_PERIOD.match(str(sh.cell_value(10, pcol)))
        if not mp:
            raise JpxFetchError('%s 第 11 行第 %d 列不是「MM/DD～MM/DD」（拿到 %r）'
                                % (fname, pcol + 1, sh.cell_value(10, pcol)))
        m1, d1, m2, d2 = (int(x) for x in mp.groups())
        # 周期原文只有 MM/DD，年份得从抬头的「YYYY年M月」推。两种跨年都真实存在：
        #   · 1 月那份文件的**上一周**落在上年 12 月（实测 260101 期 = 12/29～12/30）
        #   · 1 月第 5 周的**周末**落在 2 月（实测 240105 期 = 01/29～02/02）
        # 所以不能写「月份大于抬头月就减一年」——那会把 02/02 判成上一年，
        # 生成 week_end < week_start 的行。按「与抬头月的距离」判：差半年以上才跨年。
        rec = {'week_start': '%04d-%02d-%02d' % (_inv_year(hy, hm, m1), m1, d1),
               'week_end': '%04d-%02d-%02d' % (_inv_year(hy, hm, m2), m2, d2),
               'src_file': fname}
        if rec['week_end'] < rec['week_start']:
            raise JpxFetchError('%s 解析出 week_end %s < week_start %s —— 跨年推断错了'
                                % (fname, rec['week_end'], rec['week_start']))
        for lab, col in _INV_ROWS.items():
            buy = _inv_cell(sh, rows[(lab, '買い')], vcol)
            sell = _inv_cell(sh, rows[(lab, '売り')], vcol)
            if buy is None or sell is None:
                raise JpxFetchError('%s %s 的買い/売り取不到数' % (fname, lab))
            rec[col] = (buy - sell) / 1e6                  # 千円 → 十亿円
        f_tot = _inv_cell(sh, rows[('海外投資家', '合計')], vcol)
        b_tot = _inv_cell(sh, rows[('委託計', '合計')], vcol)
        rec['foreign_share_pct'] = (None if not b_tot or f_tot is None
                                    else f_tot / b_tot * 100.0)
        out.append(rec)
    return out


def _inv_discover(cache_dir):
    """落地页 + 近 _INV_YEARS_BACK 年的档案页，返回 [(url, 文件名)]（新→旧）。"""
    data, _h = _http_get(URL_INV_LANDING)
    html = data.decode('utf-8', 'replace')
    with open(os.path.join(cache_dir, 'jpx_landing_investor.html'), 'wb') as f:
        f.write(data)
    pages = [html]
    for path in list(dict.fromkeys(_INV_ARCHIVE.findall(html)))[:_INV_YEARS_BACK]:
        d, _h = _http_get(_HOST + path)
        pages.append(d.decode('utf-8', 'replace'))
    found = {}
    for h in pages:
        for href, token in _INV_FILE.findall(h):
            found.setdefault(token, _HOST + href)
    if not found:
        raise JpxFetchError('投資部門別页面上找不到 stock_val_1_{YYMMWW}.xls，源站可能改版')
    return [(found[t], 'stock_val_1_%s.xls' % t) for t in sorted(found, reverse=True)]


def update_investors(series_dir, cache_dir):
    """更新 series/jpx_investors.csv，返回新增周数。

    只下载 src_file 台账里没有的那些文件 —— 首次建库会拉近 _INV_YEARS_BACK 年的周文件，
    之后每月只有落地页上新出现的那几个。同一周出现在两份文件里（每份印本周+上周），
    按 week_end 去重，**先到的那份留下**，与「已有值永不覆盖」同一条原则。
    """
    csv_path = os.path.join(series_dir, 'jpx_investors.csv')
    if os.path.exists(csv_path):
        with open(csv_path, newline='', encoding='utf-8') as f:
            rows = list(csv.reader(f))
        header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
    else:
        header, body = list(INV_HEADER), []
    if header != INV_HEADER:
        raise JpxFetchError('series/jpx_investors.csv 的表头与 INV_HEADER 不一致：%s' % header)
    fidx = header.index('src_file')
    done_files = {r[fidx] for r in body}
    have_weeks = {r[0] for r in body}

    added = 0
    for url, fname in _inv_discover(cache_dir):
        if fname in done_files:
            continue
        blob, _h = _http_get(url)
        with open(os.path.join(cache_dir, 'jpx_' + fname), 'wb') as f:
            f.write(blob)
        for rec in parse_investors(blob, fname):
            if rec['week_end'] in have_weeks:
                continue
            body.append([_fmt_inv(rec[c]) for c in INV_HEADER])
            have_weeks.add(rec['week_end'])
            added += 1
        done_files.add(fname)
    body.sort(key=lambda r: r[0])
    with open(csv_path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(body)
    return added


def _fmt_inv(v):
    if v is None:
        return ''
    if isinstance(v, str):
        return v
    return ('%.6f' % float(v)).rstrip('0').rstrip('.') or '0'


if __name__ == '__main__':
    _here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    print('latest:', latest_month(os.path.join(_here, 'cache')))
    print('added :', update(os.path.join(_here, 'series'),
                            os.path.join(_here, 'cache')))
