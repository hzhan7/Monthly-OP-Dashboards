# -*- coding: utf-8 -*-
"""LSEG 一级市场月度统计（Main Market + AIM）—— 无人值守抓取。

伦交所把一级市场月度数字放在两份**每月一个 xlsx 文件**的 factsheet 里：
`Main Market factsheet <Month> <Year>.xlsx` 与 `AIM factsheet <Month> <Year>.xlsx`。
本模块把两份并成一行月度记录写进 series/lseg_part_primary.csv，**不做任何加总或换算**
—— Main Market 一套列、AIM 一套列，两个市场谁都不并进谁。

════════════════════════════════════════════════════════════════════════════
数据源与「怎么找到文件」
════════════════════════════════════════════════════════════════════════════
⚠ **文件名不可拼、不可猜，这是本模块最重要的一条。**
Drupal 在同名文件重传时会自动加序号后缀，实测同一批链接里同时存在：

    AIM factsheet June 2026.xlsx          ← 无后缀
    AIM factsheet April 2026_1.xlsx       ← 第二次重传
    AIM factsheet March 2026_0.xlsx       ← 第一次重传
    AIM factsheet February 2026_0.xlsx

2026-08-07 全量统计：AIM 2018-01 起 103 期里 43 期带 `_N` 后缀，Main Market 98 期里
31 期带。**按「AIM factsheet {Month} {Year}.xlsx」硬拼，会有四成月份 404。**
后缀与月份无规律（既不是新月才有，也不是老月才有），唯一正确的做法是从索引里读链接。

索引不在 HTML 里 —— www.londonstockexchange.com 是 Angular SPA，
`/reports` 与 `/reports?tab=aim` 返回的是同一份 54,995 字节的空壳，
`curl` 拿不到任何文件链接，`sitemap.xml` 是 404。真正的索引走这条 JSON 接口：

  1) 握手（拿 tab → tabId / componentId 映射，**不要写死这些 uuid**）
     GET https://api.londonstockexchange.com/api/v1/pages?path=reports
     → components[0].content[0].value.reportsFilterToggleFilters
       里每个 subFilter 有 {label, tabId, modules:[{moduleId}]}，
       本模块要的两个：label='Main Market' 与 label='AIM'（属 'Primary markets' 组）。

  2) 取该 tab 的报告清单
     POST https://api.londonstockexchange.com/api/v1/components/refresh
     body {"path":"reports",
           "parameters":"tab%3D<slug>%26tabId%3D<tabId>",
           "components":[{"componentId":"block_content%3A<uuid>","parameters":null}]}
     ⚠ **componentId 里的冒号必须写成 %3A，parameters 必须是二次编码的查询串。**
       冒号不编码时接口照样返回 HTTP 200，但 body 是空数组 `[]` —— 静默失败，
       不是报错。本模块因此对空结果直接 raise，绝不当成「本月没有报告」。
     → content[0].value.ctaItems[*]：
         ctaButton.link          = 最新一期直链
         history.items[*].links  = 历年归档，{label:'AIM factsheet June 2026.xlsx', link:...}
       2026-08-07 实测：AIM 573 条（分 1995~2026 共 32 个年份组），
       Main Market 321 条（2009~2026 共 18 个年份组）。

  3) 月份**从 label 解析，不从 URL 解析** —— URL 里有 `_0`/`_1` 噪声，label 是干净的
     「AIM factsheet April 2026.xlsx」。解析完还要用工作簿内的标题格（第 3~4 行写着
     「April 2026」）再校验一次，不符就 raise：这是防「索引 label 与文件内容错位」的
     唯一一道闸。

下载域是 docs.londonstockexchange.com。实测普通 UA + plain urllib 直接 200，
无 Cloudflare / Akamai / PerimeterX / JS 渲染 / 登录墙；2026-08-07 一次拉 201 个
xlsx 零失败（8 线程），满足 cron 无人值守。

**不进管道的两个源**（同在 Primary markets 组下，看着像但不是月度统计）：
`New issues and IPOs_82.xlsx` 与 `Further issues_85.xlsx` —— 这两份是**逐笔明细的
滚动全表**（0.9MB / 5.2MB），没有月度汇总口径，且文件名里的 `_82`/`_85` 每次重传都
自增，属于同一类不可拼的名字。要月度数就得用 factsheet，别用这两份。

════════════════════════════════════════════════════════════════════════════
发布节奏（实测，不是官网承诺）
════════════════════════════════════════════════════════════════════════════
判据取工作簿 `docProps/core.xml` 的 `dcterms:created`（文件生成时刻），
样本 = **2018-05 起两个市场共同覆盖的全部 197 期**，逐期解出来的，不是抽样。
⚠ AIM 2016-01…2018-04 那 28 期**不进这个样本**：它们是 2026-08 分两波一次性回补
下载的，created 已经被官方重传污染 —— 2017-01 是 +97 天、2017-02 +69、2017-03 +25，
而同一批里 2017-07/10/11、2018-02/04 又都是 +1，中位 +5；2016 那 12 期更极端，
2016-11/12 给出 −6145 / −6176 天（官方把老文件另存成 xlsx 时带过来的原始创建时刻），
2016-10 及更早是 .xls，BIFF 里压根没有 `docProps/core.xml`。这种「有的 +1 有的 +97
有的 −6145」不是当年的发布节奏，是重传时刻的分布。拿它去调闸门只会得到一个假的长尾。
`publish_lags()` 里对老版式那 12 期是**硬排除**（`month <= LEGACY_LAST[tag]`），
2017-01…2018-04 那 16 期靠调用方传 `start=` 排除。

| 市场        | 样本 | 最早 | 中位 | P75 | P90 | 最晚 | ≤3 天 | ≤5 天 | ≤9 天 | ≤12 天 |
|-------------|-----|-----|-----|-----|-----|-----|-------|-------|-------|--------|
| AIM         |  99 |  +1 |  +2 |  +5 | +13 | +27 | 66 期 | 78 期 | 89 期 | 89 期 |
| Main Market |  98 |  +1 |  +2 |  +4 |  +9 | +27 | 68 期 | 79 期 | 89 期 | 91 期 |
（数字都是「月末后第几天」，+1 = 次月 1 日。）

近 24 个月（2024-08 ~ 2026-07）单独看：两个市场都是 min +1 / 中位 +3 / 24 期里 22 期 ≤+9；
AIM 最晚 +27（2026-04 那期），Main Market 最晚 +22（2026-04 那期）。

两个市场是不是一起发的？98 个共有月份里 **68 个 created 日完全相同、74 个相差 ≤1 天**，
其余 24 个最大相差 25 天。⇒ 多数月份确实是同一批生成的，但**不是每月都同步**，
所以闸门要按两条腿里更慢的那条设，不能只看 Main Market。

**闸门 LAG 建议 (常规月 9, 季末月 9)** —— 语义「月末后第几天」。
取 9 不取 5 的理由：5 只覆盖 78/99 与 79/98 期，页面会有两成月份白等；
9 覆盖到 89/99 与 89/98（≈90%）。再往上抬没有收益 —— AIM 从 +9 到 +12 一期都不增加，
剩下那 10% 长尾（+13 ~ +27）没有任何规律：既不集中在季末月，也不集中在某一年
（2023-03/05/08、2024-02/05/06/07、2026-03/04/05 都在里面），
抬闸门只会让所有月份陪最差的那一期等，性价比为负。

⚠ **`dcterms:created` 是文件生成时刻，不等于首次发布时刻。**
同一份文件被重新生成过，created 就会跟着跳到重生成那天 —— 与 fetch/db1.py 里
Eurex Cover「Created on」踩的是同一个坑。本模块因此：
· 只给「本次运行刚确立的那个月」当发布日证据（那时文件刚发，滞后落在 +1~+9）；
· **绝不给回补的历史月份补记发布日**。
另有 HTTP `Last-Modified` 头可用，但它是 CDN 上的文件 mtime，被重传污染得更狠
（AIM 103 期里 58 期与 created 不同，最大滞后 751 天），只能当旁证，不能当判据。

════════════════════════════════════════════════════════════════════════════
口径坑（按踩坑概率排序）
════════════════════════════════════════════════════════════════════════════
1. **每份 factsheet 只覆盖「本年至今」，月末存量只能从「本月这一份」里取。**
   两份工作簿的结构完全同构，都是三块：
   (a) 年度块（`Year` 行）：AIM 给 1995 起全部年份，Main Market **只给当年一行**；
   (b) 月度块（`Month` 行）：只有本年 1 月到本月，年初会重置；
   (c) 逐笔明细。
   所以 2024-03 的「月末上市公司家数 / 总市值」只存在于 2024 年 3 月那一份文件里 —
   2024 年 4 月的文件的年度块已经换成 4 月末的存量了。⇒ **一个月一个文件，
   36 个月就得下 72 个文件**，没有一份「全历史总表」可以省掉这件事。
   年度块当年那一行的存量口径 = **该 factsheet 月份的月末**（不是年末）：
   2026-07 那份写 2026 年 598 家 / £60,461.4m，正是 7 月末数，官方就是这么滚动更新的。

2. **募资额在同一份工作簿里有两处，Summary 月度块那一处会滞后，专表才是准的。**
   募资额同时出现在 Summary 月度块与 New Issues / Further Issues 两张专表上。
   把 2018-05 起 197 期**每期自己那个月**的两处逐位比对，只有 3 期不等 —— 而这 3 期
   **全都是 Summary 那一侧后来被改成专表的值**，方向一次都没反过来：

   | 期         | 字段    | Summary 月度块 | 专表          | 后续期里的 Summary   |
   |------------|---------|---------------|--------------|----------------------|
   | MM 2018-06 | new     | 237.97915935  | 7.00999995   | 7.00999995（7 月那期）|
   | AIM 2022-03| further | 207.48509751  | 214.37873751 | 214.37873751（6 月那期）|
   | MM 2026-05 | further | 833.32263722  | 969.83130022 | —（当前最新，尚未回改）|

   同一份 MM 2018-06 里还有第四个证据（不进冲突日志，因为不是它自己那个月）：
   Summary 的 3 月 new = 990.0803119，专表 3 月 = 746.9831506，
   到 7 月那期 Summary 的 3 月也被改成了 746.9831506。

   第三条还有闭合旁证：Main Market 2026-07 那期年度块的 YTD Further = 2126.76110737，
   与 Further Issues 专表 YTD **逐位相等**，而 Summary 月度块 7 个月加起来只有
   1990.25244437 —— 差的正好是 5 月那 136.5。
   ⇒ **募资额一律取专表**；取完再与 Summary 月度块比对，不等就记进
     cache/lseg_primary_conflicts.csv 供人工判断，**不自动吞、也不自动改**。
   ⚠ 两个市场的专表列名写法不同：AIM 写「New Issues (£m)」，Main Market 写
     「Money Raised - **New Shares**」。别被 Main Market 那个 New Shares 误导成
     「只算新股、不算老股减持」—— 2018-07 那期把 Summary 回改成了专表的值，
     说明官方自己就把这两个当同一个序列。

3. **年度块的 Money Raised 是第三个数，跟上面两个都不等，本模块一分钱都不从年度块取。**
   Main Market 2026-07：年度块 YTD New = 1468.62166156，
   而 New Issues 专表 YTD 与 Summary 月度块之和都是 33.74999809。差 43 倍。
   官方没给脚注解释这个缺口。
   ⇒ 本模块**只从年度块取存量（家数 / 市值）**，募资额全部走专表；
     年度块的 Money Raised 列压根不入库，也不拿它去验算别的列。

4. **`-` 与空格都表示 0，但这是被闭合检验证明出来的，不是拍脑袋。**
   月度块里没有新上市的月份写 `-`（AIM 2026-02），有时干脆留空
   （Main Market 2026-06 的 New Issues TOTAL 格）。本模块把两者都读成 0.0，
   然后**强制**跑闭合检验：月度块各月之和 必须等于 该块的 `Sum:` / `YTD Total` 行。
   2018-01 起 201 期全部通过，零例外（新上市家数、注销家数两列都查）。
   闭合不过就 raise —— 假如哪天空格真的表示「未披露」而不是 0，这道检验会立刻炸。

5. **表头两行会上下对调，按行号定位必挂。**
   2023-03 起（AIM）/ 2022-10 起（Main Market）的新版式：`Month` 在**上**一行，
   分组名（New Issues / Cancellations / Money Raised）也在上一行，子标签在下一行。
   更早的旧版式：分组名在上一行，`Month` 与子标签**一起在下一行**。
   两种版式的列号也年年变（AIM 月度块 Total 列在 2022-06 是 c5，2026-07 是 c8）。
   ⇒ 本模块一律「先找分组行（含 Cancellations 的那一行）、再向右填充分组名、
     再与下一行的子标签配对」得到 (组, 子) → 列号，**没有一个写死的行号或列号**。

6. **官方归档自己有两个洞，都不是我们抓漏的，而且都补不回来。**
   (a) Main Market **2022-12 那一期在官方索引里根本不存在** —— 2022 年组只有 1~11 月，
       跳过 12 月直接到 2023 年组（索引里搜 'December 2022' 零命中）。
   (b) AIM **2019-09 那一期的年度块 Market Value 格是空的** —— 文件在、家数在
       （753 / 129 / 882），就是市值那一格 LSEG 留白了。
   两处在别处都拿不到：年度块对**往年**只给年末数，只有当年那一行才是月末数，
   所以 2019-09 的 AIM 月末市值、2022-12 的 Main Market 全套存量，
   在后续任何一期 factsheet 里都不存在。
   ⇒ 这两个洞**各自只砸一个市场**（键是 (市场, 月份)，见模块常量 KNOWN_SOURCE_GAPS）：
     2022-12 的 AIM factsheet、2019-09 的 Main Market factsheet 都好端端地在。
     缺的那一侧整段列留空，另一侧照写 —— 按月**整行**跳过等于替官方多挖一个洞
     （2026-08-19 前的版本就是这么干的，白丢了 MM 2019-09 与 AIM 2022-12 两段）。
     除这两个白名单条目外，一个市场只要出场，它那一段列任何一格解析为空仍然直接
     raise —— 白名单是为了让模块能跑完，不是为了让它对新出现的空格睁一只眼。
   ⚠ 半空的那一侧**不许拿别处凑**：AIM 2019-09 的家数其实是有的（753/129/882），
     但市值那一格拿不到，本模块不写「只差一列」的半段 —— 一段列要么格格齐，要么整段空。
   另：AIM 2023-04 那期年度块用了另一套子标签（`Further Issues` 而不是 `Further`），
   不是洞，已在解析里兼容两种写法。

7. **官方索引里同一期常有三种格式，.xls 那一支现在读得了（AIM），但只读到 2016-01。**
   2026-08-19 实测：AIM 索引 146 个 xlsx / 214 个 xls / 214 个 pdf（xls 一直回到
   1999-01），Main Market 98 / 112 / 112（xls 回到 2009-01）。pdf 一份都不碰。
   xlsx 起点：AIM 2016-11、Main Market 2018-05；**能读的**起点见口径坑 10 / 11。
   实测（2026-08-19，含本轮回补）：AIM 2016-01 起 127 期解析通过 126 期
   （不过的就是上面 (b) 的 AIM 2019-09）、Main Market 2018-05 起 98 期全过。
   期间版式至少换过四代（老版式 → 表头对调 → 列号漂移 → AIM 的 Total 列在 2021 与
   2026 位置不同），全靠标签定位吃下来，没有一处按行号 / 列号写死。

8. **AIM 的市值没有 UK / International 拆分，Main Market 有。**
   AIM 年度块的 Market Value 组底下是空的，只有一个合计格；Main Market 那组底下
   明确写着 UK / International / Total 三个子标签。
   ⇒ 本模块**不造 `aim_marketcap_uk_eop_gbp_mn` 这种列**（宁可少一列，不写空列）。

9. **单位是「家」与「£m（百万英镑）」，不是十亿。**
   工作簿表头原文 `Market Value (£m)` / `Money Raised (£m)`。
   Main Market 2026-07 市值 4,253,398.855454 —— 这是 £4.25tn 写成百万英镑。
   列名一律带 `_gbp_mn` / `_count` 后缀，别在下游再猜一次。

10. **两个市场的可读起点差 28 个月，共用一个 START 会白扔 AIM 的 28 期。**
    之前模块只有一个全局 `START = '2018-05'`，而且 `fetch_rows()` 里写着
    「两个市场必须同时有，否则整月跳过」—— 两者叠加，AIM 单独能拿的
    2016-01…2018-04 那 28 个月一起被丢掉了。现在起点按市场分开
    （模块常量 `MARKET_START`），一行里两个市场各写各的那一段列。

    · **AIM 2017-01…2018-04（16 期）现代版式，解析器一行都不用改。**
      2026-08-19 实测全部 16 期用 `_parse_factsheet()` 直接通过（含月度块闭合检验），
      UK + International == Total 在家数与新上市两组上 16 期零违例。跨期对账也过：
      2017-12 那一行的年末存量（808/152/960，市值 106,882.266107）与 2018-05 /
      2019-01 / 2020-06 / 2026-07 四期后来的 factsheet 年度块里印的 2017 行
      **逐位相同**（后来那几期把市值印成 106882.3，是它们自己少印了小数位，
      入库取的是当期原值这个更精确的写法）；2018-01…2018-04 的新上市 / 退市家数
      与 2018-05 那期月度块也逐格相同。
    · **AIM 2016-01…2016-12（12 期）老版式，走 `_parse_*_legacy()` 那一套**
      （2016-01…2016-10 是 .xls/BIFF，2016-11/12 是 xlsx 但内容同样是老版式）。
      版式差异逐条列在「老版式解析」那一节的注释里 —— 十条差异里九条会让现代解析器
      要么 raise、要么**取到错列**，所以是另写一套，不是把 `_check_title()` 放宽。
      口径接得上，证据有三条，都是 2026-08-19 实测：
        (a) **存量恒等式跨版式连续**：`家数(t) = 家数(t-1) + 新上市(t) − 退市(t)`
            从 2016-01 一路验到 2017-06 共 18 个月零违例，其中 2016-12 → 2017-01
            那一步正好跨版式（982 + 1 − 10 = 973）。这同时证明了老版式的
            `Number of admissions` == 现代的 `New Issues`、老版式的 `Delistings`
            == 现代的 `Cancellations` —— 只要有一边多算或少算一类，恒等式立刻崩。
        (b) **年度块跨期逐位相同**：AIM 2016-12 那期年度块的「2016 to Dec」行
            （809/173/982、市值 80814.08324446996、新上市 55/9/64、
            募资 1103.6807/3661.8552）与 2017-01 那期（现代版式）年度块印的 2016 行
            逐位相同（后者把市值印成 80814.1，少印了小数位）。
        (c) **增发笔数是数出来的不是猜的**：2016-12 那期汇总块写 168 笔，同一 sheet
            里 12 月的逐笔明细正好 168 行；2017-01 那期（现代）汇总写 136，明细正好
            136 行。两边都是「当月增发笔数（含不募资的期权行权等）」。
      另外 12 期一次冲突都没记进 `cache/lseg_primary_conflicts.csv` ——
      老版式的 Summary 月度块募资额与专表汇总块逐位相等。
    · **停在 2016-01 是主动选的边界，不是版式极限。** 官方 .xls 归档一直到 1999-01，
      抽样（2015-12 / 2013-06 / 2010-06 / 2007-06 / 2004-06）版式与 2016 完全一致；
      2001-06 与 1999-01 才变成另一套（sheet 叫 `Pge 1`，连年度块都没有）。
      选 2016-01 的理由：那是同一页 orderbook 那一路的起点，宽表左端正好对齐
      （两路各 127 个月）。要往前挪就改 `MARKET_START['AIM']` 一个常量 ——
      但**先把每一期的闭合检验跑一遍再挪**，抽样通过不等于 150 期逐期通过。

11. **Main Market 的老版式：格式不是问题，口径是 —— 2026-08-19 评估后决定不做。**
    官方 .xls 一直回到 2009-01（112 期），xlrd 读得动，版式也认得出。不做的理由是
    13 列里有 4 列**根本没有对应来源**、1 列**口径明确不同**，硬拼会在图上画出一段
    假历史。逐列结论（样本 = 2018-01…2018-04 四期 .xls 与 2018-05 那期现代 xlsx
    月度块里印的同月数字逐格对照；老版式的表长得完全不一样，是「行标签 + UK/Intl
    两栏」而不是「组 + 子标签」）：

    | 现行列                           | 老版式来源                     | 结论 |
    |----------------------------------|--------------------------------|------|
    | `mm_companies_uk_eop_count`      | `T8 Co's by value` 的 `Totals*`| ✅ ① |
    | `mm_marketcap_uk_eop_gbp_mn`     | 同一张 T8 的 Equity market val | ✅ ② |
    | `mm_new_issues_uk/intl_count`    | Summary `New companies`        | ✅ ③ |
    | `mm_cancellations_count`         | Summary `Delistings/Transfers` | ✅ ④ |
    | `mm_companies_intl_eop_count`    | **无**                         | ❌ ⑤ |
    | `mm_marketcap_intl_eop_gbp_mn`   | **无**                         | ❌ ⑥ |
    | `mm_companies_eop_count`（合计） | **无**                         | ❌ ⑦ |
    | `mm_marketcap_eop_gbp_mn`（合计）| **无**                         | ❌ ⑦ |
    | `mm_further_issues_count`        | Summary `No. of co's / issues` | ❌ ⑧ |
    | `mm_money_raised_new_gbp_mn`     | Summary Equity `New Issues`    | ⚠ ⑨ |
    | `mm_money_raised_further_gbp_mn` | Summary Equity `Further issues`| ⚠ ⑩ |

    ① 接得上：2018-04 = 946 → 2018-05 = 944。
    ② 接得上：2,618,874 → 2,634,816 £m（+0.6%）。**T8 本来就是 £m，不用换算**
       —— Summary 那处写的是 £bn，从那里取会多一步单位换算，是不必要的风险。
    ③ 同口径：Jan/Feb/Mar 与现代版式逐格相同，Apr 差 1（老 2 / 新 3，是后一期重述）。
    ④ 同口径：Jan/Feb/Apr 逐格相同，Mar 差 2（同上，重述）。
    ⑤ 老版式只有 Summary 的 268（= Trading 264 + Suspended 4），现行 221，差 −17.5%；
       T8 只有 UK 与 AIM 两块，没有 Main Market International 那一段。
    ⑥ 老 £1,795bn vs 现行 £1,472bn，差 −18% —— 与 ⑤ 同源同幅，是同一处口径断裂。
    ⑦ 缺 Intl 就拼不出合计。Summary 的 `Total companies` 是 987
       （Trading 971 + Suspended 16），现行 944，取它会造 −4.4% 的假台阶。
    ⑧ **另一套数法**：2018 年 1–4 月老版式 24/30/45/86，现行专表 51/54/71/103，
       四期无一相同（老版式那一栏漏掉不募资的那些笔）。
    ⑨ 只算「募了钱的那几家」：Jan 老 4.35 vs 现行 17.13（老版式那一栏的家数写 2，
       而同一份文件的 `New companies` 写 5）。Mar 倒是与现行**重述后**的
       746.9831506 逐位相同。
    ⑩ Jan / Mar 与现行逐位相同，Feb 老 2720.10 vs 现行 999.84（2.7 倍），
       Apr 差 1.7%。四期里两期对不上，且对不上的那两期没有可解释的规律。

    ⇒ 就算只写「接得上」的那 6 列，2009-01…2018-04 这 112 期里也会出现
      「UK 家数有、合计家数没有」的一段 —— 宽表允许列各自起点不同，但那张页面上
      Exhibit 41/44（合计）与 42/45（UK）会一个从 2018-05 开始、一个从 2009-01
      开始，看图的人第一反应必然是「国际板 2018 年才存在」。收益（6 列多 112 期）
      配不上这个误读风险，所以 **Main Market 整段停在 2018-05**。
    ⚠ 老版式 Main Market 自己还会再换一次版：2013-06 / 2010-06 / 2009-01 那几期的
      Summary 是**三个市场分栏**（UK Listed / International Listed + 第三块，
      Trading 分别 65 / 91 / 101 家），列数从 6 变 9。真要重启这件事，那是第二道坎。

════════════════════════════════════════════════════════════════════════════
列口径表
════════════════════════════════════════════════════════════════════════════
| 列名                              | 含义                          | 来源 sheet / 块      |
|-----------------------------------|-------------------------------|----------------------|
| mm_companies_eop_count            | 主板月末上市公司家数（合计）   | Summary 年度块       |
| mm_companies_uk_eop_count         | 同上，UK 注册                 | Summary 年度块       |
| mm_companies_intl_eop_count       | 同上，International           | Summary 年度块       |
| mm_marketcap_eop_gbp_mn           | 主板月末总市值 £m             | Summary 年度块       |
| mm_marketcap_uk_eop_gbp_mn        | 同上，UK                      | Summary 年度块       |
| mm_marketcap_intl_eop_gbp_mn      | 同上，International           | Summary 年度块       |
| mm_new_issues_count               | 主板当月新上市家数（合计）     | Summary 月度块       |
| mm_new_issues_uk_count            | 同上，UK                      | Summary 月度块       |
| mm_new_issues_intl_count          | 同上，International           | Summary 月度块       |
| mm_cancellations_count            | 主板当月退市/注销家数          | Summary 月度块       |
| mm_further_issues_count           | 主板当月增发笔数               | Further Issues 专表  |
| mm_money_raised_new_gbp_mn        | 主板当月新上市募资 £m          | New Issues 专表      |
| mm_money_raised_further_gbp_mn    | 主板当月增发募资 £m            | Further Issues 专表  |
| aim_*                             | 同构，AIM 市场                | AIM since launch 等  |
| （AIM 无市值 UK/Intl 拆分，故只有 aim_marketcap_eop_gbp_mn）              |

⚠ **两段列的起点不同**（口径坑 10 / 11）：`aim_*` 11 列从 2016-01 起，`mm_*` 13 列从
2018-05 起。CSV 左上角 2016-01…2018-04 的 mm_* 一片空白**不是抓漏、也不是官方没发**
—— 官方 Main Market 的 .xls 回到 2009-01，是它的口径接不上（口径坑 11 逐列写了理由）。
下游画图必须按**列**自己的起点裁窗口，不能拿「表有 127 行」当每列都有 127 个观测。

「新上市」= 官方口径 New Issues，含 IPO、从另一个板转板、反向收购、introduction，
**不等于 IPO**；两份专表都给了 by Type 拆分，需要纯 IPO 时得另开列，本模块没取。
所有列来源等级 **[A] 公司/交易所原始披露** —— 全部出自 LSEG 自家 factsheet，
无第三方转述、无推算、无券商研报。

════════════════════════════════════════════════════════════════════════════
自校验（跑一次就全跑，任何一条不过直接 raise）
════════════════════════════════════════════════════════════════════════════
现代版式（AIM 2017-01 起 / Main Market 全段）：
· 工作簿标题月 == 索引给的月份（防 `_N` 重传链接与内容错位）
· 月度块逐月之和 == `Sum:` / `YTD Total` 行（新上市家数、注销家数两列）
老版式（AIM 2016-12 及更早）—— 判据换了，强度没降，有两条反而更强：
· 标题格必须有一个落在该月的**日期值**（老版式不写「January 2016」）
· 月度块逐月之和 == **年度块当年那一行**（新上市家数、募资额 New、募资额 Further
  三列）。这是**跨块**对账 —— 年度块与月度块在工作簿里是两张独立的表，
  比现代版式的块内 `Sum:` 行更难蒙混过去
· `Total Money raised == Rights + Other`（增发专表汇总块同一行内）——
  这是「笔数可以把 rights 与 other 相加」这条前提的硬证明，不成立就 raise
· `New cos` 专表 Totals 行的家数 == 月度块该月新上市家数（同一工作簿两张表互证，
  取错列会当场炸）
· 十二行月份必须连续且顺序一格不差（老版式没有 `Sum:` 行做终止判据，
  再往下的二级市场块也是十二行月份名，扫过界会把成交额行当成一级市场行加进闭合检验）
两个版式共用：
· **出场的那个市场**，它那一段列每一格都非空，否则 raise
  （某个市场整段缺席是允许的 —— 起点没到 / 官方的洞；半段缺席不允许）
另外三条是**跑完之后的事后核对**，不在模块里断言，记在这里备查
（2026-08-19 对 127 行重做过，零违例）：
· UK + International == Total（家数两组 + 主板市值一组 + 新上市家数两组）
· 存量恒等式 `家数(t) = 家数(t-1) + 新上市(t) − 退市(t)`：AIM 2016-01…2017-06
  连续 18 个月零违例，含 2016-12 → 2017-01 那一步跨版式（见口径坑 10）
· 二次运行 series/lseg_part_primary.csv 字节级相同
"""

import csv
import datetime
import email.utils
import json
import os
import re
import time
import urllib.request
import zipfile

# ── 站点常量 ────────────────────────────────────────────────────────────────
PAGES_API = 'https://api.londonstockexchange.com/api/v1/pages?path=reports'
REFRESH_API = 'https://api.londonstockexchange.com/api/v1/components/refresh'

_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36')

# 索引里的 subFilter label（在 'Primary markets' 组下）。tabId / componentId 一律
# 从握手接口现取，**不写死** —— CMS 改版时 uuid 会变，label 不会。
MARKETS = [
    # (内部 tag, 列前缀, subFilter label, ctaTitle 前缀, 汇总 sheet 名的判据)
    ('MM',  'mm',  'Main Market', 'Main Market factsheet', 'summary'),
    ('AIM', 'aim', 'AIM',         'AIM factsheet',         'since launch'),
]

# 每个市场自己的起点 —— **两个市场共用一个 START 是错的**（口径坑 10）。
# 官方索引里 AIM 的 xlsx 归档比 Main Market 深 16 个月，共用起点等于把 AIM 那 16 个月
# 白扔。值的含义：「这个市场从这个月起，用本模块现有解析器可以零改动读通」。
MARKET_START = {
    # Main Market：索引里最早的 xlsx 就是这一期。**再往前有 112 期 .xls（2009-01 起）
    # 但本模块不读** —— 不是因为读不动 BIFF（xlrd 已经在用了），是因为老版式 Main Market
    # 的口径接不上，13 列里 4 列根本没有对应来源。完整论证见口径坑 11。
    'MM':  '2018-05',
    # AIM：老版式（.xls 2016-01…2016-10 + 老版式 xlsx 2016-11/12）用 `_parse_*_legacy`
    # 这一套读，见口径坑 10。停在 2016-01 是**主动选的边界**，不是版式极限 ——
    # 官方 .xls 归档一直回到 1999-01，抽样到 2004-06 版式都还一样；
    # 2016-01 = 同一页 orderbook 那一路的起点，宽表左端对齐。
    'AIM': '2016-01',
}
# 全模块最早可能出现的月份。只用来铺循环区间，**不是任何一个市场的起点**。
START = min(MARKET_START.values())

# 官方源自己缺的月份 —— 见口径坑 6。键是 **(市场, 月份)**，不是月份：
# 两个洞各自只砸一个市场，砸不到另一个（2022-12 的 AIM、2019-09 的 Main Market
# 都好端端地在），按月整行跳过等于替官方多挖一个洞。缺的那一侧整段列留空，
# 另一侧照写 —— 宽表本来就允许列各自起点不同、各自留洞。
# 白名单之外的任何空格仍然直接 raise，别往这里加东西来「让它跑过去」。
KNOWN_SOURCE_GAPS = {
    ('MM', '2022-12'): 'Main Market factsheet 这一期在官方索引里不存在（2022 年组只到 11 月）',
    ('AIM', '2019-09'): 'AIM factsheet 这一期的年度块 Market Value 格是空的，官方留白',
}

MONTHS = ['JANUARY', 'FEBRUARY', 'MARCH', 'APRIL', 'MAY', 'JUNE', 'JULY',
          'AUGUST', 'SEPTEMBER', 'OCTOBER', 'NOVEMBER', 'DECEMBER']

CONFLICT_LOG = 'lseg_primary_conflicts.csv'

# 募资额比对容差（£m）。工作簿里同一个数在两张表里是同一个浮点常量，
# 实测差异要么是 0、要么是 136.5 这种量级的真差异，1e-6 足够分开这两种情况。
MONEY_TOL = 1e-6


class LsegPrimaryFetchError(RuntimeError):
    """源站结构变化 / 下载失败 / 解析结果不完整。一律炸掉，绝不静默写空列或 NaN。"""


# ── 列定义 ──────────────────────────────────────────────────────────────────
# (列名, 市场 tag, 解析结果里的 key)
_META = [
    ('mm_companies_eop_count',           'MM',  'companies_total'),
    ('mm_companies_uk_eop_count',        'MM',  'companies_uk'),
    ('mm_companies_intl_eop_count',      'MM',  'companies_intl'),
    ('mm_marketcap_eop_gbp_mn',          'MM',  'mktcap_total'),
    ('mm_marketcap_uk_eop_gbp_mn',       'MM',  'mktcap_uk'),
    ('mm_marketcap_intl_eop_gbp_mn',     'MM',  'mktcap_intl'),
    ('mm_new_issues_count',              'MM',  'new_total'),
    ('mm_new_issues_uk_count',           'MM',  'new_uk'),
    ('mm_new_issues_intl_count',         'MM',  'new_intl'),
    ('mm_cancellations_count',           'MM',  'canc_total'),
    ('mm_further_issues_count',          'MM',  'further_count'),
    ('mm_money_raised_new_gbp_mn',       'MM',  'money_new'),
    ('mm_money_raised_further_gbp_mn',   'MM',  'money_further'),
    ('aim_companies_eop_count',          'AIM', 'companies_total'),
    ('aim_companies_uk_eop_count',       'AIM', 'companies_uk'),
    ('aim_companies_intl_eop_count',     'AIM', 'companies_intl'),
    # AIM 年度块的 Market Value 组底下没有 UK / International 子标签，
    # 官方就是只给一个合计 —— 见口径坑 8，故不造那两列。
    ('aim_marketcap_eop_gbp_mn',         'AIM', 'mktcap_total'),
    ('aim_new_issues_count',             'AIM', 'new_total'),
    ('aim_new_issues_uk_count',          'AIM', 'new_uk'),
    ('aim_new_issues_intl_count',        'AIM', 'new_intl'),
    ('aim_cancellations_count',          'AIM', 'canc_total'),
    ('aim_further_issues_count',         'AIM', 'further_count'),
    ('aim_money_raised_new_gbp_mn',      'AIM', 'money_new'),
    ('aim_money_raised_further_gbp_mn',  'AIM', 'money_further'),
]

COLUMNS = ['month'] + [m[0] for m in _META]


def _openpyxl():
    try:
        import openpyxl                              # noqa: PLC0415
    except ImportError as e:                         # pragma: no cover
        raise LsegPrimaryFetchError(
            'lseg_primary 需要 openpyxl 读 factsheet xlsx：pip install openpyxl') from e
    return openpyxl


def _xlrd():
    """老版式的 .xls（BIFF）读取器 —— 只有 2016 年及更早的 AIM factsheet 用得到。

    延迟 import 并给明确报错，跟 `_openpyxl()` 同一个套路：xlrd 没装时是「2016 那
    12 期读不了」，不是整个模块 ImportError。**但它不是可选依赖** —— 少了它，
    `fetch_rows()` 会在 2016-01 那一期直接 raise，而不是悄悄少写 12 行
    （合流层 fetch/lseg.py 的 `_refresh_primary` 是整表重写 part CSV 的，
    静默少行会把已经入库的 12 行从 series/lseg_part_primary.csv 里抹掉）。
    加它的理由写在 requirements.txt 里。
    """
    try:
        import xlrd                                  # noqa: PLC0415
    except ImportError as e:                         # pragma: no cover
        raise LsegPrimaryFetchError(
            'lseg_primary 读 2016 年及更早的 AIM factsheet 需要 xlrd（.xls/BIFF）：'
            'pip install xlrd —— 理由见 requirements.txt') from e
    return xlrd


# ── 网络 ────────────────────────────────────────────────────────────────────
def _http(url, data=None, timeout=120, retries=3, json_body=False):
    """带重试的 GET/POST。data 非空即 POST。"""
    last = None
    headers = {'User-Agent': _UA, 'Accept': '*/*', 'Accept-Language': 'en-GB,en;q=0.9'}
    if json_body:
        headers['Content-Type'] = 'application/json'
        headers['Accept'] = 'application/json'
    for attempt in range(retries):
        req = urllib.request.Request(url, data=data, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                return r.read(), dict(r.headers)
        except Exception as e:                        # noqa: BLE001
            last = e
            if attempt < retries - 1:
                time.sleep(2 * (attempt + 1))
    raise LsegPrimaryFetchError('下载失败（%d 次重试后）%s: %r' % (retries, url, last))


def _cache(cache_dir, name):
    os.makedirs(cache_dir, exist_ok=True)
    return os.path.join(cache_dir, name)


def _download(url, path):
    data, headers = _http(url)
    if len(data) < 20000:
        raise LsegPrimaryFetchError(
            '%s 只有 %d 字节，不像是正常的 factsheet 工作簿' % (url, len(data)))
    with open(path, 'wb') as f:
        f.write(data)
    lm = headers.get('Last-Modified') or headers.get('last-modified') or ''
    with open(path + '.lm', 'w', encoding='utf-8') as f:
        f.write(lm)
    return path


def _cached_download(url, path):
    """已在 cache 里就不重下 —— 只用于回补历史月份。

    历史月份的 factsheet 内容不会因为重下而变新（官方重传时只会改 `_N` 后缀，
    那是一条新链接，索引会给出来），重下几十上百个文件纯属浪费带宽。
    **最新一期不走这条路**，每次都用 _download 重下，见 fetch_rows()。
    """
    if os.path.exists(path) and os.path.getsize(path) > 20000:
        return path
    return _download(url, path)


# ── 索引发现 ────────────────────────────────────────────────────────────────
def _slug(label):
    return re.sub(r'[^a-z0-9]+', '-', label.lower()).strip('-')


def _discover_tabs(cache_dir):
    """握手接口 → {subFilter label: (slug, tabId, componentId)}。"""
    raw, _ = _http(PAGES_API)
    with open(_cache(cache_dir, 'lseg_primary_pages.json'), 'wb') as f:
        f.write(raw)
    doc = json.loads(raw.decode('utf-8', 'replace'))
    out = {}
    for comp in doc.get('components') or []:
        for c in comp.get('content') or []:
            val = c.get('value') or {}
            for filt in val.get('reportsFilterToggleFilters') or []:
                for sub in filt.get('subFilters') or []:
                    mods = sub.get('modules') or []
                    if not sub.get('tabId') or not mods:
                        continue
                    mid = (mods[0].get('moduleId') or '').split(':', 1)[-1]
                    out[sub.get('label')] = (_slug(sub.get('label')), sub['tabId'], mid)
    if not out:
        raise LsegPrimaryFetchError(
            '握手接口没解析出任何 subFilter —— %s 的 CMS 结构变了' % PAGES_API)
    return out


def _tab_reports(slug, tab_id, component_id, cache_dir, tag):
    """POST components/refresh 拿该 tab 的报告清单 JSON。

    ⚠ componentId 里的冒号必须 %3A —— 不编码时接口返回 200 + `[]`，静默失败。
    所以空结果一律 raise，不当成「本月没有报告」。
    """
    body = json.dumps({
        'path': 'reports',
        'parameters': 'tab%%3D%s%%26tabId%%3D%s' % (slug, tab_id),
        'components': [{'componentId': 'block_content%3A' + component_id,
                        'parameters': None}],
    }).encode('utf-8')
    raw, _ = _http(REFRESH_API, data=body, json_body=True)
    with open(_cache(cache_dir, 'lseg_primary_%s.json' % tag), 'wb') as f:
        f.write(raw)
    doc = json.loads(raw.decode('utf-8', 'replace'))
    if not doc:
        raise LsegPrimaryFetchError(
            'components/refresh 对 tab=%s 返回空数组 —— 多半是 componentId 没编码，'
            '或 CMS 换了 uuid（当前用的是 %s）' % (slug, component_id))
    return doc


def _month_index(doc, cta_prefix, tag, ext='.xlsx'):
    """报告清单 JSON → {'YYYY-MM': 直链}。

    月份**从 label 解析**（'AIM factsheet April 2026.xlsx'），不从 URL 解析：
    URL 里有 `_0`/`_1` 重传后缀。同一月份出现多条时取先遇到的
    （ctaButton 的最新一期排在最前）。

    `ext` 决定收哪一种文件。同一个 tab 的清单里 .xlsx / .xls / .pdf 三种并存
    （2026-08-19 实测 AIM 146/214/214、Main Market 98/112/112），且**多数月份
    只有其中一种**，所以两种扩展名要各建一份索引，不能混在一起靠「谁先出现」取。
    """
    mon_no = {m.capitalize(): i + 1 for i, m in enumerate(MONTHS)}
    out = {}
    for comp in doc:
        for c in comp.get('content') or []:
            for cta in (c.get('value') or {}).get('ctaItems') or []:
                title = cta.get('ctaTitle') or ''
                if not title.lower().startswith(cta_prefix.lower()):
                    continue
                pairs = [(title, ((cta.get('ctaButton') or {}).get('link')))]
                for grp in (cta.get('history') or {}).get('items') or []:
                    for link in grp.get('links') or []:
                        pairs.append((link.get('label'), link.get('link')))
                for label, url in pairs:
                    if not url or not url.lower().endswith(ext):
                        continue
                    m = re.search(r'([A-Z][a-z]+)\s+(\d{4})', label or '')
                    if not m or m.group(1) not in mon_no:
                        continue
                    key = '%s-%02d' % (m.group(2), mon_no[m.group(1)])
                    out.setdefault(key, url)
    if not out:
        raise LsegPrimaryFetchError(
            '%s 的报告清单里没解析出任何 %s factsheet 链接' % (tag, ext))
    return out


# ── 小工具 ──────────────────────────────────────────────────────────────────
def _norm(v):
    return re.sub(r'\s+', ' ', str(v)).strip() if v is not None else ''


def _num(v, blank_zero=False):
    """`-` 一律 0（口径坑 4）；空格按调用点决定是 0 还是 None；其余非数字 → None。"""
    if v is None:
        return 0.0 if blank_zero else None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    if s == '':
        return 0.0 if blank_zero else None
    if s in ('-', '–', '—', 'n/a', 'N/A'):
        return 0.0
    try:
        return float(s.replace(',', '').replace('£', ''))
    except ValueError:
        return None


def _month_end(month):
    y, m = int(month[:4]), int(month[5:7])
    nxt = datetime.date(y + 1, 1, 1) if m == 12 else datetime.date(y, m + 1, 1)
    return nxt - datetime.timedelta(days=1)


def _months_between(start, end):
    y, m = int(start[:4]), int(start[5:7])
    out = []
    while '%04d-%02d' % (y, m) <= end:
        out.append('%04d-%02d' % (y, m))
        m += 1
        if m == 13:
            y, m = y + 1, 1
    return out


def _grid(sh):
    return [[sh.cell(r, c).value for c in range(1, sh.max_column + 1)]
            for r in range(1, sh.max_row + 1)]


def _find_row(g, pred, lo=0):
    for i in range(lo, len(g)):
        if pred(g[i]):
            return i
    return -1


def _colmap(g, h1, h2):
    """两行表头 → ({(组小写, 子小写): 列}, [(组, 起列, 止列)])。

    组标签向右填充到下一个组为止；组底下没有子标签时记成 (组, '')。
    表头两行会上下对调（口径坑 5），调用点负责把 h1/h2 传对。
    """
    groups = [(c, _norm(v)) for c, v in enumerate(g[h1]) if _norm(v)]
    spans = []
    for i, (c, lab) in enumerate(groups):
        end = groups[i + 1][0] if i + 1 < len(groups) else len(g[h1])
        spans.append((lab, c, end))
    out = {}
    for lab, c0, c1 in spans:
        subs = [(c, _norm(g[h2][c])) for c in range(c0, min(c1, len(g[h2])))
                if _norm(g[h2][c])]
        if subs:
            for c, s in subs:
                out[(lab.lower(), s.lower())] = c
        else:
            out[(lab.lower(), '')] = c0
    return out, spans


def _span_of(spans, prefix):
    for lab, a, b in spans:
        if lab.lower().startswith(prefix.lower()):
            return a, b
    return None


def _pick(cm, grp_prefix, sub_names):
    """(组前缀, 子标签候选) → 列号；找不到返回 None。"""
    for (grp, sub), col in cm.items():
        if grp.startswith(grp_prefix) and sub in sub_names:
            return col
    return None


# ── 解析 ────────────────────────────────────────────────────────────────────
def _sheet_named(wb, needle, path):
    for name in wb.sheetnames:
        if needle in name.lower():
            return wb[name]
    raise LsegPrimaryFetchError(
        '%s: 找不到含 %r 的 sheet，实有 %s' % (os.path.basename(path), needle, wb.sheetnames))


def _check_title(g, month, path):
    """工作簿内的标题格必须写着这个月 —— 防索引 label 与文件内容错位。"""
    want = (MONTHS[int(month[5:7]) - 1].capitalize() + ' ' + month[:4]).lower()
    seen = [_norm(v).lower() for row in g[:8] for v in row if _norm(v)]
    if want not in seen:
        raise LsegPrimaryFetchError(
            '%s: 工作簿标题月与索引不符，索引说 %s，文件里写的是 %r'
            % (os.path.basename(path), want, seen[:6]))


def _parse_summary(sh, month, path):
    """汇总 sheet（Main Market 'Summary' / AIM 'AIM since launch'）。

    只取两样东西：年度块的**存量**（家数 / 市值）与月度块的**家数流量**。
    募资额一分钱都不从这里取 —— 见口径坑 2、3。
    """
    g = _grid(sh)
    _check_title(g, month, path)
    year = month[:4]
    mname = MONTHS[int(month[5:7]) - 1]
    out = {}

    # ── 年度块 ──
    h1 = _find_row(g, lambda r: any(_norm(v).lower() == 'year' for v in r)
                   and any(_norm(v).lower().startswith('number of companies') for v in r))
    if h1 < 0:
        raise LsegPrimaryFetchError('%s: 找不到年度块表头（Year + Number of Companies）'
                                    % os.path.basename(path))
    cm, spans = _colmap(g, h1, h1 + 1)
    ycol = next(c for c, v in enumerate(g[h1]) if _norm(v).lower() == 'year')

    def _is_year(r):
        v = r[ycol]
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return str(int(v)) == year
        return _norm(v) == year
    yrow = _find_row(g, _is_year, h1)
    if yrow < 0:
        raise LsegPrimaryFetchError('%s: 年度块里没有 %s 那一行' % (os.path.basename(path), year))
    R = g[yrow]

    for key, subs in [('companies_uk', ('uk',)),
                      ('companies_intl', ('international',)),
                      ('companies_total', ('total',))]:
        col = _pick(cm, 'number of companies', subs)
        if col is None:
            raise LsegPrimaryFetchError(
                '%s: 年度块缺 Number of Companies/%s 列 —— 有 %s'
                % (os.path.basename(path), subs[0], sorted(cm)))
        out[key] = _num(R[col])

    mv = _span_of(spans, 'market value')
    if mv is None:
        raise LsegPrimaryFetchError('%s: 年度块没有 Market Value 组' % os.path.basename(path))
    tot = _pick(cm, 'market value', ('total',))
    if tot is not None:
        out['mktcap_total'] = _num(R[tot])
        out['mktcap_uk'] = _num(R[_pick(cm, 'market value', ('uk',))])
        out['mktcap_intl'] = _num(R[_pick(cm, 'market value', ('international',))])
    else:
        # AIM：组底下没有子标签，整组只有一个合计格（口径坑 8）
        vals = [_num(R[c]) for c in range(mv[0], min(mv[1], len(R)))
                if _num(R[c]) is not None]
        if len(vals) != 1:
            raise LsegPrimaryFetchError(
                '%s: Market Value 组里数值格不唯一（%d 个），版式可能变了'
                % (os.path.basename(path), len(vals)))
        out['mktcap_total'] = vals[0]
        out['mktcap_uk'] = None
        out['mktcap_intl'] = None

    # ── 月度块 ──
    # 定位判据用「含 Cancellations 的那一行」而不是「含 Month 的那一行」：
    # 新旧两种版式里 Month 一会儿在上一行、一会儿在下一行（口径坑 5），
    # 但 Cancellations 永远在分组行上。
    m1 = _find_row(g, lambda r: any(_norm(v).lower().startswith('cancellations') for v in r),
                   h1 + 1)
    if m1 < 0:
        raise LsegPrimaryFetchError('%s: 找不到月度块表头（Cancellations 组）'
                                    % os.path.basename(path))
    cm2, _spans2 = _colmap(g, m1, m1 + 1)
    mcol = None
    for row in (g[m1], g[m1 + 1]):
        for c, v in enumerate(row):
            if _norm(v).lower() == 'month':
                mcol = c
                break
        if mcol is not None:
            break
    if mcol is None:
        raise LsegPrimaryFetchError('%s: 月度块表头两行都没有 Month 格'
                                    % os.path.basename(path))

    def _mrow(nm):
        return _find_row(g, lambda r: _norm(r[mcol]).upper().rstrip(':').strip() == nm, m1)
    rr = _mrow(mname)
    if rr < 0:
        raise LsegPrimaryFetchError('%s: 月度块里没有 %s 那一行'
                                    % (os.path.basename(path), mname))
    M = g[rr]

    wanted = [('new_uk', 'new issues', ('uk',)),
              ('new_intl', 'new issues', ('international',)),
              ('new_total', 'new issues', ('total',)),
              ('canc_uk', 'cancellations', ('uk',)),
              ('canc_intl', 'cancellations', ('international',)),
              ('canc_total', 'cancellations', ('total',)),
              ('sum_money_new', 'money raised', ('new issues', 'new')),
              ('sum_money_further', 'money raised', ('further issues', 'further'))]
    cols = {}
    for key, grp, subs in wanted:
        col = _pick(cm2, grp, subs)
        if col is None:
            raise LsegPrimaryFetchError(
                '%s: 月度块缺 %s/%s 列 —— 有 %s'
                % (os.path.basename(path), grp, subs[0], sorted(cm2)))
        cols[key] = col
        out[key] = _num(M[col], blank_zero=True)
        if out[key] is None:
            raise LsegPrimaryFetchError(
                '%s: 月度块 %s 那一格解析不出数字（原值 %r）'
                % (os.path.basename(path), key, M[col]))

    # ── 闭合检验：各月之和 == Sum: / YTD Total 行（口径坑 4） ──
    trow = _find_row(g, lambda r: _norm(r[mcol]).lower().rstrip(':').strip()
                     in ('sum', 'total', 'ytd total'), m1 + 1)
    if trow < 0:
        raise LsegPrimaryFetchError('%s: 月度块没有合计行，无法做闭合检验'
                                    % os.path.basename(path))
    for key in ('new_total', 'canc_total'):
        col = cols[key]
        acc = 0.0
        for i in range(m1 + 1, trow):
            if _norm(g[i][mcol]).upper() in MONTHS:
                v = _num(g[i][col], blank_zero=True)
                if v is None:
                    raise LsegPrimaryFetchError(
                        '%s: 月度块 %s 列有不可解析的格（第 %d 行，原值 %r）'
                        % (os.path.basename(path), key, i + 1, g[i][col]))
                acc += v
        want = _num(g[trow][col], blank_zero=True)
        if want is None or abs(acc - want) > 1e-9:
            raise LsegPrimaryFetchError(
                '%s: 月度块 %s 闭合失败，逐月之和 %s ≠ 合计行 %s —— '
                '空格可能不再表示 0，别猜，去看文件' % (os.path.basename(path), key, acc, want))
    return out


def _parse_detail(wb, sheet_name, month, path, want_count):
    """New Issues / Further Issues 专表 → (募资额 £m, 笔数或 None)。

    募资额取 'Money Raised' 组跨度内该月行的最后一个数值格 —— 这个组在
    2018/2021/2026 三种版式里分别落在 c13/c14/c11~c15，列号没法写死，
    但「组跨度内唯一/最后一个数值」在 201 期实测里始终成立。
    """
    if sheet_name not in wb.sheetnames:
        raise LsegPrimaryFetchError('%s: 缺 sheet %r，实有 %s'
                                    % (os.path.basename(path), sheet_name, wb.sheetnames))
    g = _grid(wb[sheet_name])
    mname = MONTHS[int(month[5:7]) - 1]
    # 分组行 = 含 'Money Raised' 的那一行（新旧版式里它都在上面一行）
    h1 = _find_row(g, lambda r: any(_norm(v).lower().startswith('money raised') for v in r))
    if h1 < 0:
        raise LsegPrimaryFetchError('%s/%s: 找不到 Money Raised 分组行'
                                    % (os.path.basename(path), sheet_name))
    cm, spans = _colmap(g, h1, h1 + 1)
    mcol = None
    for row in (g[h1], g[h1 + 1]):
        for c, v in enumerate(row):
            if _norm(v).lower() == 'month':
                mcol = c
                break
        if mcol is not None:
            break
    if mcol is None:
        raise LsegPrimaryFetchError('%s/%s: 表头两行都没有 Month 格'
                                    % (os.path.basename(path), sheet_name))
    rr = _find_row(g, lambda r: _norm(r[mcol]).upper().rstrip(':').strip() == mname, h1)
    if rr < 0:
        raise LsegPrimaryFetchError('%s/%s: 没有 %s 那一行'
                                    % (os.path.basename(path), sheet_name, mname))
    R = g[rr]

    msp = _span_of(spans, 'money raised')
    vals = [_num(R[c]) for c in range(msp[0], min(msp[1], len(R))) if _num(R[c]) is not None]
    # 整个跨度全空 = 该月没有这一类募资，官方就是留白（口径坑 4）
    money = vals[-1] if vals else 0.0

    count = None
    if want_count:
        nsp = _span_of(spans, 'number of')
        if nsp is None:
            raise LsegPrimaryFetchError('%s/%s: 没有 Number of 组'
                                        % (os.path.basename(path), sheet_name))
        col = _pick(cm, 'number of', ('total', 'totals'))
        if col is None:
            raise LsegPrimaryFetchError(
                '%s/%s: Number of 组底下没有 Total 子标签 —— 有 %s'
                % (os.path.basename(path), sheet_name, sorted(cm)))
        count = _num(R[col], blank_zero=True)
        if count is None:
            raise LsegPrimaryFetchError('%s/%s: %s 的笔数格解析不出数字（原值 %r）'
                                        % (os.path.basename(path), sheet_name, mname, R[col]))
    return money, count


# ── 老版式解析（AIM 2016-12 及更早） ────────────────────────────────────────
# 为什么要另写一套：老版式跟现代版式**不是同一张表换了行号**，是换了一套词。
# 逐条对照（2026-08-19 实测，样本 = AIM 2016-01…2016-12 全 12 期 + 2015-12 /
# 2013-06 / 2010-06 / 2007-06 / 2004-06 五期抽样）：
#   现代                                老版式
#   ─────────────────────────────────  ────────────────────────────────────
#   标题格 'January 2017'               标题格是**日期值** 2016-01-31（序列值，
#                                       所以 `_check_title()` 那一步必炸，
#                                       放宽校验是错的，得换判据）
#   年度块有 'Year' 表头格               **没有** —— 年份就在最左列，无表头
#   年度块当年那行写 '2017'              写 '2016 to Jan' / '2016 to Dec'
#   月度块组名 'New Issues'              'Number of admissions'
#   月度块组名 'Cancellations'           'Delistings'（且**底下没有 Total 子标签**，
#                                       合计值就落在组自己那一列上）
#   月度块有 'Month' 表头格              **没有** —— 月份名在最左列，表头那格写年份
#   月度块末尾有 'Sum:' 行               **没有**（闭合只能拿年度块当年那行当靶子）
#   募资专表 'New Issues' / 'Further     'New cos' / 'Further'（或 'Further Issues'），
#   Issues' 顶上有月度汇总表              汇总块在**逐笔明细的最后**，是「本月 / YTD」
#                                       两栏而不是「一月…十二月」十二行
#   增发笔数一格 'TOTAL'                 拆成 'No. of rights issues' 与
#                                       'No. of other issues#' 两格
# ⇒ 十条里有九条会让现代解析器要么 raise、要么**取到错列**。所以是另一套函数，
#   不是给 `_parse_summary()` 加 if。共用的只有 `_colmap` / `_pick` / `_num` 这几个
#   不认识版式的小工具。
#
# 版式的**下边界没有验到底**：'AIM since launch' 这个 sheet 名到 2004-06 都还在，
# 2001-06 与 1999-01 已经变成 'Pge 1'（连年度块都没有）。本模块只把 AIM 的起点
# 放到 2016-01（= orderbook 那一路的起点，宽表左端对齐），不往下探 ——
# 抽样通过不等于 150 期逐期通过，要往前挪先把每一期的闭合检验都跑一遍。

# 用老版式读的月份上界。键是市场 tag；不在这个 dict 里的市场没有老版式路径。
# AIM：2016-11 / 2016-12 是 xlsx 但内容是老版式，2016-10 及更早是 .xls，
# 两者版式**完全一致**（同样的组名、同样的列布局），所以判据是月份不是扩展名。
LEGACY_LAST = {'AIM': '2016-12'}


def _open_book(path):
    """→ (kind, wb)。.xls 走 xlrd、.xlsx 走 openpyxl，上层不关心是哪种。"""
    if path.lower().endswith('.xls'):
        return 'xls', _xlrd().open_workbook(path, on_demand=True)
    return 'xlsx', _openpyxl().load_workbook(path, data_only=True)


def _close_book(book):
    kind, wb = book
    try:
        wb.release_resources() if kind == 'xls' else wb.close()
    except Exception:                                 # noqa: BLE001
        pass


def _book_names(book):
    kind, wb = book
    return list(wb.sheet_names()) if kind == 'xls' else list(wb.sheetnames)


def _book_grid(book, name, path):
    """按 sheet 名取二维 list。xlrd 的日期格转成 datetime，跟 openpyxl 对齐。"""
    kind, wb = book
    if name not in _book_names(book):
        raise LsegPrimaryFetchError('%s: 缺 sheet %r，实有 %s'
                                    % (os.path.basename(path), name, _book_names(book)))
    if kind == 'xlsx':
        return _grid(wb[name])
    xlrd = _xlrd()
    sh = wb.sheet_by_name(name)
    out = []
    for r in range(sh.nrows):
        row = []
        for c in range(sh.ncols):
            cell = sh.cell(r, c)
            v = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    v = datetime.datetime(*xlrd.xldate_as_tuple(v, wb.datemode))
                except (ValueError, xlrd.XLDateError):
                    # 老文件里有几格是「0 号日期」（year 0），当空格处理即可 ——
                    # 它们都在逐笔明细的填充行上，不在我们取值的任何一格。
                    v = None
            row.append(v)
        out.append(row)
    return out


def _sheet_startswith(book, prefix, path):
    """sheet 名前缀匹配。老版式里增发专表叫 'Further'（.xls）或 'Further Issues'
    （2016-11/12 的 xlsx），名字不稳，前缀是稳的。"""
    for nm in _book_names(book):
        if nm.strip().lower().startswith(prefix.lower()):
            return nm
    raise LsegPrimaryFetchError('%s: 没有以 %r 开头的 sheet，实有 %s'
                                % (os.path.basename(path), prefix, _book_names(book)))


def _is_month_date(v, month):
    """这一格是不是落在 `month` 里的日期值。"""
    if not isinstance(v, (datetime.datetime, datetime.date)):
        return False
    return '%04d-%02d' % (v.year, v.month) == month


def _check_title_legacy(g, month, path):
    """老版式的标题格是**日期序列值**，不是 'January 2016'。

    判据：前 8 行里必须有一格是落在这个月的日期。防的是同一件事 —— 索引 label
    与文件内容错位。（前 8 行里唯一另一个日期格是年度块首行的 1995-06-19
    「AIM 开市日」，只有在解析 1995-06 那一期时才可能撞上，本模块不到那么早。）
    """
    if not any(_is_month_date(v, month) for row in g[:8] for v in row):
        seen = [str(v)[:19] for row in g[:8] for v in row
                if isinstance(v, (datetime.datetime, datetime.date))]
        raise LsegPrimaryFetchError(
            '%s: 老版式标题格里找不到 %s 的日期，索引 label 与文件内容可能错位'
            '（前 8 行里的日期格：%s）' % (os.path.basename(path), month, seen or '无'))


def _vlabels(g, r0, r1):
    """把 r0..r1 这几行**竖着拼**成每一列的标签 → {列号: '拼好的小写标签'}。

    老版式的专表汇总块表头有三四行（'Market' / 'value' / '(£m)' 上下叠），
    单看任何一行都认不出这一列是什么。竖拼之后 c4 = 'market value (£m)'、
    c5 = 'money raised (£m)'，就能按名字取列，不用写死列号。
    """
    out = {}
    width = max((len(g[r]) for r in range(r0, r1 + 1) if r < len(g)), default=0)
    for c in range(width):
        parts = [_norm(g[r][c]) for r in range(r0, r1 + 1)
                 if r < len(g) and c < len(g[r]) and _norm(g[r][c])]
        if parts:
            out[c] = ' '.join(parts).lower()
    return out


def _month_rows_legacy(g, mcol, m1, path):
    """月度块的十二行月份 → 起始行号。

    老版式没有 'Sum:' 行，**也不能靠「往下扫到月份名就算」** —— 同一个 sheet
    再往下的二级市场块（Trading）也是十二行月份名，扫过界会把成交额行当成
    一级市场行加进闭合检验。所以要求：JANUARY 之后连续十二行正好是十二个月，
    顺序一格不差；对不上直接 raise。
    """
    j0 = _find_row(g, lambda r: _norm(r[mcol]).upper().rstrip(':').strip() == 'JANUARY', m1)
    if j0 < 0:
        raise LsegPrimaryFetchError('%s: 月度块里找不到 January 行' % os.path.basename(path))
    for i, nm in enumerate(MONTHS):
        if j0 + i >= len(g) or _norm(g[j0 + i][mcol]).upper().rstrip(':').strip() != nm:
            raise LsegPrimaryFetchError(
                '%s: 月度块第 %d 行应该是 %s，实为 %r —— 十二行月份不连续，版式变了'
                % (os.path.basename(path), j0 + i + 1, nm,
                   _norm(g[j0 + i][mcol]) if j0 + i < len(g) else '越界'))
    return j0


def _parse_summary_legacy(g, month, path):
    """老版式的 'AIM since launch' sheet → 存量 + 家数流量 + Summary 侧募资额。

    与现代版一样：年度块只取**存量**（家数 / 市值），月度块只取**家数流量**；
    募资额从这里读出来只是为了跟专表比对（口径坑 2），不入库。
    """
    _check_title_legacy(g, month, path)
    year = month[:4]
    out = {}

    # ── 年度块 ──（老版式没有 'Year' 表头格，判据只能是 'Number of Companies' 组）
    h1 = _find_row(g, lambda r: any(_norm(v).lower().startswith('number of companies')
                                    for v in r))
    if h1 < 0:
        raise LsegPrimaryFetchError('%s: 找不到年度块表头（Number of Companies 组）'
                                    % os.path.basename(path))
    cm, spans = _colmap(g, h1, h1 + 1)
    ycol = 0                       # 年份永远在最左列，且那一列没有表头

    def _is_year(r):
        s = _norm(r[ycol])
        if s.endswith('.0'):       # xlrd 把 2016 读成 2016.0
            s = s[:-2]
        return s.startswith(year + ' to ')
    yrow = _find_row(g, _is_year, h1)
    if yrow < 0:
        raise LsegPrimaryFetchError(
            "%s: 年度块里没有 '%s to <月>' 那一行 —— 老版式当年那行的写法变了"
            % (os.path.basename(path), year))
    R = g[yrow]

    for key, subs in [('companies_uk', ('uk',)),
                      ('companies_intl', ('international',)),
                      ('companies_total', ('total',))]:
        col = _pick(cm, 'number of companies', subs)
        if col is None:
            raise LsegPrimaryFetchError(
                '%s: 年度块缺 Number of Companies/%s 列 —— 有 %s'
                % (os.path.basename(path), subs[0], sorted(cm)))
        out[key] = _num(R[col])

    mv = _span_of(spans, 'market value')
    if mv is None:
        raise LsegPrimaryFetchError('%s: 年度块没有 Market Value 组' % os.path.basename(path))
    vals = [_num(R[c]) for c in range(mv[0], min(mv[1], len(R))) if _num(R[c]) is not None]
    if len(vals) != 1:
        raise LsegPrimaryFetchError(
            '%s: Market Value 组里数值格不唯一（%d 个），版式可能变了'
            % (os.path.basename(path), len(vals)))
    out['mktcap_total'] = vals[0]
    out['mktcap_uk'] = None        # 老版式同样没有 UK/Intl 拆分（口径坑 8）
    out['mktcap_intl'] = None

    # 年度块当年那行的三个 YTD 数 —— **只当闭合检验的靶子**，一格都不入库
    ytd = {}
    for key, grp, subs in [('new_total', 'number of admissions', ('total',)),
                           ('money_new', 'money raised', ('new',)),
                           ('money_further', 'money raised', ('further',))]:
        col = _pick(cm, grp, subs)
        if col is None:
            raise LsegPrimaryFetchError('%s: 年度块缺 %s/%s 列 —— 有 %s'
                                        % (os.path.basename(path), grp, subs[0], sorted(cm)))
        ytd[key] = _num(R[col], blank_zero=True)

    # ── 月度块 ──（判据是 Delistings 组；'Number of admissions' 在年度块里也有近似写法）
    m1 = _find_row(g, lambda r: any(_norm(v).lower().startswith('delisting') for v in r),
                   h1 + 1)
    if m1 < 0:
        raise LsegPrimaryFetchError('%s: 找不到月度块表头（Delistings 组）'
                                    % os.path.basename(path))
    cm2, spans2 = _colmap(g, m1, m1 + 1)
    mcol = 0                       # 月份名在最左列；那一列的表头格写的是年份，不是 'Month'
    j0 = _month_rows_legacy(g, mcol, m1, path)
    M = g[j0 + int(month[5:7]) - 1]

    cols = {}
    for key, grp, subs in [('new_uk', 'number of admissions', ('uk',)),
                           ('new_intl', 'number of admissions', ('international',)),
                           ('new_total', 'number of admissions', ('total',)),
                           ('sum_money_new', 'money raised', ('new',)),
                           ('sum_money_further', 'money raised', ('further',))]:
        col = _pick(cm2, grp, subs)
        if col is None:
            raise LsegPrimaryFetchError('%s: 月度块缺 %s/%s 列 —— 有 %s'
                                        % (os.path.basename(path), grp, subs[0], sorted(cm2)))
        cols[key] = col
    # Delistings 组底下没有任何子标签，合计值就在组自己那一列上（组跨度的第一列）。
    dsp = _span_of(spans2, 'delisting')
    if dsp is None:
        raise LsegPrimaryFetchError('%s: 月度块没有 Delistings 组' % os.path.basename(path))
    if _pick(cm2, 'delisting', ('total',)) is not None:
        raise LsegPrimaryFetchError(
            '%s: Delistings 组底下出现了 Total 子标签 —— 版式变了，'
            '「合计在组自己那一列」这条假设不再成立，去看文件' % os.path.basename(path))
    cols['canc_total'] = dsp[0]

    for key, col in cols.items():
        out[key] = _num(M[col], blank_zero=True)
        if out[key] is None:
            raise LsegPrimaryFetchError(
                '%s: 月度块 %s 那一格解析不出数字（原值 %r）'
                % (os.path.basename(path), key, M[col]))

    # ── 闭合检验：逐月之和 == 年度块当年那一行（口径坑 4 的老版式版本） ──
    # 现代版拿块内的 'Sum:' 行当靶子，老版式没有那一行，改拿年度块 —— 这是**跨块**
    # 对账，比块内自洽更强：年度块与月度块在工作簿里是两张独立的表。
    for key in ('new_total', 'money_new', 'money_further'):
        col = cols[key if key == 'new_total' else 'sum_' + key]
        acc = 0.0
        for i in range(12):
            v = _num(g[j0 + i][col], blank_zero=True)
            if v is None:
                raise LsegPrimaryFetchError(
                    '%s: 月度块 %s 列第 %d 行解析不出数字（原值 %r）'
                    % (os.path.basename(path), key, j0 + i + 1, g[j0 + i][col]))
            acc += v
        want = ytd[key]
        if want is None or abs(acc - want) > 1e-6:
            raise LsegPrimaryFetchError(
                '%s: 月度块 %s 闭合失败，逐月之和 %s ≠ 年度块当年那行 %s —— '
                '空格可能不再表示 0，别猜，去看文件'
                % (os.path.basename(path), key, acc, want))
    return out


def _parse_new_cos_legacy(book, month, path):
    """老版式 'New cos' 专表末尾的汇总块 → (本月募资额 £m, 本月家数)。

    块长这样（列号年年漂，所以按**竖拼**出来的列标签取，见 `_vlabels`）：

        |            | No. of    | Market value | Money raised | | No. of    | …
        |            | companies | (£m)         | (£m)         | | companies | …
        | 2016-12-30 |           |              |              | | Year to date
        | New companies |  5 | 254.94 | 45.5156 | |  44 | …
        | Transfer      |  0 |   0    |  0      | |   1 | …
        | Relisting     |  1 | 332.38 |  0      | |  19 | …
        | Totals        |  6 | 587.32 | 45.5156 | |  64 | …

    取 Totals 行、且只取「本月」那一栏（'Year to date' 那一列**左边**的那些列）。
    """
    name = _sheet_startswith(book, 'new cos', path)
    g = _book_grid(book, name, path)
    hdr = _find_row(g, lambda r: any(_norm(v).lower() == 'year to date' for v in r))
    if hdr < 0:
        raise LsegPrimaryFetchError("%s/%s: 汇总块里找不到 'Year to date' 那一格"
                                    % (os.path.basename(path), name))
    ytd_c = next(c for c, v in enumerate(g[hdr]) if _norm(v).lower() == 'year to date')
    if not any(_is_month_date(v, month) for v in g[hdr][:ytd_c]):
        raise LsegPrimaryFetchError(
            "%s/%s: 汇总块 'Year to date' 左边没有 %s 的日期格 —— 两栏认错了"
            % (os.path.basename(path), name, month))
    lab = _vlabels(g, max(0, hdr - 4), hdr - 1)
    money_c = next((c for c, t in sorted(lab.items())
                    if c < ytd_c and t.startswith('money raised')), None)
    count_c = next((c for c, t in sorted(lab.items())
                    if c < ytd_c and t.startswith('no. of companies')), None)
    if money_c is None or count_c is None:
        raise LsegPrimaryFetchError(
            '%s/%s: 汇总块本月那一栏认不出 Money raised / No. of companies 列 —— 有 %s'
            % (os.path.basename(path), name, sorted(lab.items())))
    trow = _find_row(g, lambda r: any(_norm(v).lower() == 'totals' for v in r), hdr)
    if trow < 0:
        raise LsegPrimaryFetchError('%s/%s: 汇总块没有 Totals 行'
                                    % (os.path.basename(path), name))
    money = _num(g[trow][money_c], blank_zero=True)
    count = _num(g[trow][count_c], blank_zero=True)
    if money is None or count is None:
        raise LsegPrimaryFetchError('%s/%s: Totals 行的募资额/家数解析不出数字（%r / %r）'
                                    % (os.path.basename(path), name,
                                       g[trow][money_c], g[trow][count_c]))
    return money, count


def _parse_further_legacy(book, month, path):
    """老版式 'Further' 专表末尾的汇总块 → (本月募资额 £m, 本月增发笔数)。

    块长这样：

        |               | No. of | Rights money | No. of  | Other Money | Total Money |
        |               | rights | raised       | other   | raised      | raised      |
        |               | issues | (£m)         | issues# | (£m)        | (£m)        |
        | 2016-12-30    |   0    |      0       |   168   |  392.0978   |  392.0978   |
        | Year to date  |   0    |      0       |  1923   | 3664.8552   | 3664.8552   |

    ⚠ **笔数在老版式里是拆成两格的**（rights / other），没有一格叫 Total。
    合计只能相加 —— 这是本模块唯一一处加法，所以同一行的
    `Total Money raised == Rights + Other` 被当成硬断言跑：官方自己把这两块的
    **钱**加成了 Total，就证明这两块是同一个总量的两个不相交部分，笔数照加才成立。
    加法一旦不成立（比如哪天多出第三类），断言先炸，不会静默给出一个偏小的笔数。

    笔数口径与现代版一致，**这是数出来的不是猜的**：2016-12 那期汇总块写 168，
    同一 sheet 里 12 月的逐笔明细正好 168 行；现代 2017-01 那期汇总写 136，
    明细正好 136 行。两边都是「当月增发笔数（含不募资的期权行权等）」。
    """
    name = _sheet_startswith(book, 'further', path)
    g = _book_grid(book, name, path)
    hdr = _find_row(g, lambda r: any(_norm(v).lower() == 'year to date' for v in r))
    if hdr < 0:
        raise LsegPrimaryFetchError("%s/%s: 汇总块里找不到 'Year to date' 那一行"
                                    % (os.path.basename(path), name))
    mrow = _find_row(g, lambda r: any(_is_month_date(v, month) for v in r),
                     max(0, hdr - 6))
    if mrow < 0 or mrow >= hdr:
        raise LsegPrimaryFetchError(
            "%s/%s: 汇总块里找不到 %s 那一行（'Year to date' 在第 %d 行）"
            % (os.path.basename(path), name, month, hdr + 1))
    lab = _vlabels(g, max(0, hdr - 6), mrow - 1)
    def _col(prefix):
        return next((c for c, t in sorted(lab.items()) if t.startswith(prefix)), None)
    c_rights, c_other = _col('no. of rights'), _col('no. of other')
    c_rmoney, c_omoney = _col('rights money'), _col('other money')
    c_total = _col('total money')
    if None in (c_rights, c_other, c_rmoney, c_omoney, c_total):
        raise LsegPrimaryFetchError(
            '%s/%s: 汇总块表头认不全（rights/other/total 五列）—— 有 %s'
            % (os.path.basename(path), name, sorted(lab.items())))
    R = g[mrow]
    vals = {}
    for key, c in [('rights', c_rights), ('other', c_other), ('rmoney', c_rmoney),
                   ('omoney', c_omoney), ('total', c_total)]:
        vals[key] = _num(R[c], blank_zero=True)
        if vals[key] is None:
            raise LsegPrimaryFetchError('%s/%s: 汇总块 %s 那一格解析不出数字（原值 %r）'
                                        % (os.path.basename(path), name, key, R[c]))
    if abs(vals['rmoney'] + vals['omoney'] - vals['total']) > MONEY_TOL:
        raise LsegPrimaryFetchError(
            '%s/%s: Rights %s + Other %s ≠ Total %s —— 「rights 与 other 是总量的两个'
            '不相交部分」这条前提不成立，笔数就不能相加，去看文件'
            % (os.path.basename(path), name, vals['rmoney'], vals['omoney'], vals['total']))
    return vals['total'], vals['rights'] + vals['other']


def _parse_factsheet_legacy(path, month, summary_needle):
    book = _open_book(path)
    try:
        g = _book_grid(book, _sheet_named_in(book, summary_needle, path), path)
        rec = _parse_summary_legacy(g, month, path)
        money_new, new_count = _parse_new_cos_legacy(book, month, path)
        rec['money_new'] = money_new
        rec['money_further'], rec['further_count'] = _parse_further_legacy(book, month, path)
    finally:
        _close_book(book)
    # 专表的家数与月度块的家数必须一致 —— 这两个数在同一份工作簿的两张表上，
    # 不一致说明上面那套「按标签取列」在这一期取错了列（而不是官方重述：重述发生在
    # 不同期之间，不会在同一份文件内部）。
    if abs(new_count - rec['new_total']) > 1e-9:
        raise LsegPrimaryFetchError(
            '%s: New cos 专表说本月 %s 家、月度块说 %s 家 —— 同一份工作簿内部对不上，'
            '多半是取错了列' % (os.path.basename(path), new_count, rec['new_total']))
    return rec


def _sheet_named_in(book, needle, path):
    for nm in _book_names(book):
        if needle in nm.lower():
            return nm
    raise LsegPrimaryFetchError('%s: 找不到含 %r 的 sheet，实有 %s'
                                % (os.path.basename(path), needle, _book_names(book)))


def _created_at(path):
    """docProps/core.xml 的 dcterms:created → date；拿不到返回 None。

    ⚠ 这是**文件生成时刻**，被重新生成过就会跳（见 docstring 发布节奏一节）。
    """
    try:
        with zipfile.ZipFile(path) as z:
            core = z.read('docProps/core.xml').decode('utf-8', 'replace')
    except Exception:                                 # noqa: BLE001
        return None
    m = re.search(r'<dcterms:created[^>]*>(\d{4}-\d{2}-\d{2})', core)
    return datetime.date.fromisoformat(m.group(1)) if m else None


def _last_modified(path):
    """HTTP Last-Modified（旁证，不当判据）。"""
    p = path + '.lm'
    if not os.path.exists(p):
        return None
    raw = open(p, encoding='utf-8').read().strip()
    if not raw:
        return None
    try:
        return email.utils.parsedate_to_datetime(raw).date()
    except Exception:                                 # noqa: BLE001
        return None


def _parse_factsheet(path, month, summary_needle):
    openpyxl = _openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        rec = _parse_summary(_sheet_named(wb, summary_needle, path), month, path)
        rec['money_new'], _ = _parse_detail(wb, 'New Issues', month, path, want_count=False)
        rec['money_further'], rec['further_count'] = _parse_detail(
            wb, 'Further Issues', month, path, want_count=True)
    finally:
        wb.close()
    return rec


# ── 对外接口 ────────────────────────────────────────────────────────────────
def latest_month(cache_dir):
    """两个市场都已发布 factsheet 的最新月 'YYYY-MM'。

    取 min(Main Market 最新月, AIM 最新月)：两份都到齐，这一行才写得全。
    抓不到就抛异常，不返回 None 掩盖故障。
    """
    tabs = _discover_tabs(cache_dir)
    newest = {}
    for tag, _pfx, label, cta, _needle in MARKETS:
        if label not in tabs:
            raise LsegPrimaryFetchError(
                'reports 索引里没有 %r 这个 subFilter，实有 %s' % (label, sorted(tabs)))
        slug, tab_id, comp_id = tabs[label]
        idx = _month_index(_tab_reports(slug, tab_id, comp_id, cache_dir, tag), cta, tag)
        newest[tag] = max(idx)
    return min(newest.values())


def fetch_rows(cache_dir=None, start=None, end=None):
    """返回 [{'month':'YYYY-MM', ...列...}, ...]，按月份升序。

    **一行 = 一个月，两个市场各写各的那一段列**（口径坑 10）。某个市场在这个月
    没有可读的 factsheet（还没到它的 `MARKET_START`、索引里没有、或落在
    `KNOWN_SOURCE_GAPS` 里），它那 11/13 列就整段不出现在这个 dict 里 ——
    `write_csv()` 把它们写成空格。**空格就是空格**：不填 0、不填 NaN、不用前值。

    一个市场只要出场，它那一段列就必须**格格有值**：任何一格解析为空一律抛
    LsegPrimaryFetchError（半空的一段列是「解析悄悄坏了」，不是「官方没披露」）。
    两个市场都缺席的月份整行不写。

    `start` 缺省 None = 每个市场用自己的 `MARKET_START`；给了值就当**下限**
    压在两个市场上（排障用，例如只重跑最近半年）。
    """
    if cache_dir is None:
        cache_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'cache')
    os.makedirs(cache_dir, exist_ok=True)

    tabs = _discover_tabs(cache_dir)
    index, index_legacy, newest = {}, {}, {}
    for tag, _pfx, label, cta, _needle in MARKETS:
        if label not in tabs:
            raise LsegPrimaryFetchError(
                'reports 索引里没有 %r 这个 subFilter，实有 %s' % (label, sorted(tabs)))
        slug, tab_id, comp_id = tabs[label]
        doc = _tab_reports(slug, tab_id, comp_id, cache_dir, tag)
        index[tag] = _month_index(doc, cta, tag)
        # 老版式月份的源可能是 .xls，也可能是 .xlsx（AIM 2016-11/12）。两份索引都建，
        # 取的时候 xlsx 优先 —— 同一期两种格式都在时，xlsx 是官方后来重新导出的那份。
        # `newest` 只看 xlsx 索引：最新一期永远是 xlsx，.xls 只在历史段出现。
        index_legacy[tag] = _month_index(doc, cta, tag, ext='.xls') if tag in LEGACY_LAST else {}
        newest[tag] = max(index[tag])
    common_latest = min(newest.values())
    # 右端也按市场各走各的：某个月只有一个市场发了，就先写那一段列，另一段留空等回补
    # （合流层 fetch/lseg.py 是「只填空不覆盖」的外连接，这正是它要保护的场景）。
    # `latest_month()` 仍然取 min —— 「两边都发了」是 monthly_run 闸门唯一说得清的定义。
    end = end or max(newest.values())
    # start 给了就是压在两个市场上的下限；没给则各用各的 MARKET_START。
    mstart = dict((tag, max(MARKET_START[tag], start) if start else MARKET_START[tag])
                  for tag, _p, _l, _c, _n in MARKETS)

    needle = {tag: nd for tag, _p, _l, _c, nd in MARKETS}
    rows, conflicts, skipped = [], [], []
    for month in _months_between(min(mstart.values()), end):
        rec = {'month': month}
        for tag, _pfx, _label, _cta, _nd in MARKETS:
            if month < mstart[tag]:
                continue            # 这个市场的归档还没到这里 —— 不是洞，不记 skipped
            legacy = month <= LEGACY_LAST.get(tag, '')
            why = KNOWN_SOURCE_GAPS.get((tag, month))
            url = index[tag].get(month) or (index_legacy[tag].get(month) if legacy else None)
            if why is None and url is None:
                why = ('官方索引里没有 %s 的 xlsx/xls' % tag) if legacy \
                    else ('官方索引里没有 %s 的 xlsx' % tag)
            if why:
                skipped.append((month, tag, why))
                continue
            ext = '.xls' if url.lower().endswith('.xls') else '.xlsx'
            dst = _cache(cache_dir, 'lseg_primary_%s_%s%s' % (tag, month, ext))
            # 该市场最新一期每次重下（官方会在发布后几天内补数据），历史月份复用缓存。
            path = _download(url, dst) if month == newest[tag] else _cached_download(url, dst)
            parsed = (_parse_factsheet_legacy(path, month, needle[tag]) if legacy
                      else _parse_factsheet(path, month, needle[tag]))
            # 募资额取专表，与 Summary 月度块比对（口径坑 2）：不等只记录，不吞、不改
            for key, sumkey in [('money_new', 'sum_money_new'),
                                ('money_further', 'sum_money_further')]:
                a, b = parsed[key], parsed[sumkey]
                if a is not None and b is not None and abs(a - b) > MONEY_TOL:
                    conflicts.append((month, tag, key, a, b, os.path.basename(path)))
            for col, ctag, ckey in _META:
                if ctag != tag:
                    continue
                val = parsed.get(ckey)
                if val is None:
                    raise LsegPrimaryFetchError(
                        '%s %s: 列 %s 解析结果为空 —— 缺列一律失败，不写 NaN'
                        % (month, tag, col))
                rec[col] = val
        if len(rec) > 1:            # 两个市场都缺席 → 整行不写（不写只有 month 的空行）
            rows.append(rec)

    if not rows:
        raise LsegPrimaryFetchError('一行都没解析出来（start=%s end=%s）' % (start, end))
    _write_conflicts(cache_dir, conflicts)
    fetch_rows.skipped = skipped
    fetch_rows.latest_month = common_latest
    fetch_rows.market_latest = dict(newest)
    return rows


def _write_conflicts(cache_dir, conflicts):
    if not conflicts:
        return
    path = _cache(cache_dir, CONFLICT_LOG)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(['month', 'market', 'field', 'detail_sheet_value',
                    'summary_block_value', 'file'])
        w.writerows(conflicts)


def _fmt(v):
    """整数写成整数、小数用最短往返表示，保证重跑字节级不变。"""
    if v is None:
        return ''
    v = float(v)
    if v.is_integer() and abs(v) < 1e15:
        return str(int(v))
    return repr(v)


def write_csv(series_dir, rows):
    """落到 series/lseg_part_primary.csv，首列 month，升序。

    某个市场在某个月缺席时，它那一段列写**空格**（`r` 里根本没有这些键）——
    不写 0、不写 NaN、不用前值。左上角 2017-01…2018-04 只有 aim_* 那 11 列有值，
    是官方归档深度的真实形状（AIM 的 xlsx 比 Main Market 深 16 个月）。
    """
    os.makedirs(series_dir, exist_ok=True)
    path = os.path.join(series_dir, 'lseg_part_primary.csv')
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(COLUMNS)
        for r in sorted(rows, key=lambda x: x['month']):
            w.writerow([r['month']] + [_fmt(r.get(c)) for c in COLUMNS[1:]])
    return path


def publish_lags(cache_dir=None, start=START, end=None):
    """已下载的 factsheet → [(market, month, created_lag_days, lastmod_lag_days)]。

    给 docs/CRON_WIRING.md 的 LAG 提供实测依据；本身不参与建行。

    ⚠ **老版式那 12 期整段排除**（`month <= LEGACY_LAST[tag]`），两个理由：
    2016-01…2016-10 是 `.xls`，BIFF 里根本没有 `docProps/core.xml`，给不出
    `dcterms:created`；2016-11/12 虽是 `.xlsx` 却给出 −6145 / −6176 天这种
    「文件比数据月早十七年」的值 —— 那是官方后来把老文件另存成 xlsx 时带过来的
    原始创建时刻，不是发布时刻。混进来只会污染节奏统计的样本
    （与 docstring「发布节奏」一节里排掉 AIM 2017-01…2018-04 是同一个理由）。
    """
    if cache_dir is None:
        cache_dir = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'cache')
    out = []
    for tag, _p, _l, _c, _n in MARKETS:
        for month in _months_between(start, end or '2100-01'):
            if month <= LEGACY_LAST.get(tag, ''):
                continue                  # 老版式整段排除，见 docstring
            p = _cache(cache_dir, 'lseg_primary_%s_%s.xlsx' % (tag, month))
            if not os.path.exists(p):
                continue
            cr, lm = _created_at(p), _last_modified(p)
            me = _month_end(month)
            out.append((tag, month,
                        (cr - me).days if cr else None,
                        (lm - me).days if lm else None))
    return out


if __name__ == '__main__':
    _here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    _rows = fetch_rows(os.path.join(_here, 'cache'))
    print('rows  :', len(_rows), _rows[0]['month'], '->', _rows[-1]['month'])
    print('skip  :', getattr(fetch_rows, 'skipped', []))
    print('csv   :', write_csv(os.path.join(_here, 'series'), _rows))
