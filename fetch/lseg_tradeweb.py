# -*- coding: utf-8 -*-
"""LSEG 拼图第四路：Tradeweb Markets（Nasdaq: TW）月度活动报告 —— 无人值守抓取。

这一路抓的不是 LSEG 自己发的东西，是 **LSEG 并表子公司 Tradeweb 自己发的月报**。
Tradeweb 在美股独立上市、独立按月披露成交，披露粒度比 LSEG 集团季度业绩细一个量级，
所以它是 LSEG 交易业务里**唯一一块有月频、可核对、可回溯到 2017 年**的经营数据。
并表关系与「它跟 LSEG 分部收入是什么关系」见文末「━━ 并表关系 ━━」一节 ——
**那一节的结论不进 CSV**，只用来防止有人把 Tradeweb 的 ADV 当成 LSEG 集团口径。

━━ 数据源 ━━

索引页（每次必须先抓，不许跳过）：
    https://www.tradeweb.com/newsroom/monthly-activity-reports/
    页面只挂**最新一期**的三个文件 + 一份交易日历，正文里带一句自述抬头：
    "Tradeweb Reports July 2026 Total Trading Volume of $67.5 Trillion and
     Average Daily Volume of $2.9 Trillion"（2026-08-07 实测原文）。

真正的数据文件（从索引页 href 里解析出来，**不许猜**）：
    /<6位缓存串>/globalassets/newsroom/<MM.DD.YY>-<month>-mar/
        TW Historical ADV and Day Count through <Month> <Year>.xlsx
    2026-08-07 实测 183,540 字节，13 个 sheet，本模块只用 3 个：
    `ADV - M` / `Volume - M` / `Trade Days - M`，各 50 行 × 118 列，
    **2017-01 → 2026-07 共 115 个月，一格不缺**（None=0、非数字=0，逐行实测过）。

⚠ URL 里那两段**每月都变**，所以每个月都必须重新解析索引页：
    · `/4a22c4/` 这种 6 位缓存串每次发布都换（CMS 出的 asset 版本号）；
    · 目录名 `08.06.26-july-mar` 里的 `08.06.26` 就是**发布日 MM.DD.YY**，
      与 HTTP `Last-Modified: Thu, 06 Aug 2026 10:32:21 GMT` 对得上。
      本模块拿它当 source_date，并与 Last-Modified 交叉校验，对不上就抛异常。
    实测 `/globalassets/...` 这一段去掉前面的缓存串也能 200，但**不要依赖这一点**：
    目录名里的月份仍然只能从索引页拿。

📌 走过的弯路，别再走一遍：
  1. **SEC EDGAR 上没有月报。** Tradeweb CIK 1758730，2018-11 至今一共只有 65 份 8-K，
     逐条看过 items：全是 2.02（季度业绩）+ 5.02/5.07/1.01 这些公司行为，
     **月度活动报告从来不上 8-K**。想省事去 EDGAR 找月报的，会白翻一遍再回到这里。
  2. **IR 站（investors.tradeweb.com）的新闻稿正文没有资产类别口径的合计。**
     稿子里只有产品级 ADV（美债 $258.0bn、欧债 $61.4bn、按揭 $241.4bn……）
     和一个四舍五入到 1 位小数的总 ADV（"$2.9tn"，误差可达 ±$50bn）。
     rates / credit / equities / money markets 四个**资产类别合计**只在 xlsx 里有。
     ⇒ IR 新闻稿只配当发布节奏的台账（见下），不配当数据源。

━━ 抓取方式：一个反直觉的 Cloudflare 行为 ━━

www.tradeweb.com 挂着 Cloudflare **managed challenge**，但拦的是 TLS/HTTP2 指纹，
不是 UA：

    curl（HTTP/2，任何 UA）             → 403，`cf-mitigated: challenge`
    python urllib.request（HTTP/1.1）   → **200，拿到完整 HTML**
    curl_cffi impersonate='chrome124'   → 200

⇒ 本模块首选 `urllib.request` 裸奔（纯标准库，满足无人值守），
  只在拿回来的是挑战页时才回落到 curl_cffi（requirements.txt:75 已有 `curl_cffi==0.16.0`，
  fetch/hood.py 也在用，不是新依赖）。
**不要因为 `curl` 打不开就断定这站抓不了** —— 上一轮就是这么误判的。
另外 investors.tradeweb.com（Drupal/NIR 平台）完全没有挑战，plain urllib 直接 200。

依赖只有 openpyxl（requirements.txt:45 已有）。

━━ 发布节奏（实测，不是转述）━━

本机把 investors.tradeweb.com 的新闻稿列表页**按年份翻了 2019-2026 全部 8 年、
共 279 条稿件**，逐条匹配月报标题（标题措辞这些年换过至少 6 种写法：
"Tradeweb Reports July 2026 Total Trading Volume of …" /
"…Reports Record March 2026…" / "…Reports Volume of $19.6 Trillion in May" /
"…Reports Robust April Trading" / "Record Tradeweb Volume Averages More Than
$1 Trillion Per Day in March" …），命中 **88 个数据月**：

    全样本 2019-01 → 2026-07   n=88   发布日落在次月第 2 至第 11 天，中位数第 5 天
    2021-01 起                 n=65   第 2 至第 8 天，中位数第 5 天
    最近 36 个月（2023-08 起） n=36   第 3 至第 8 天，中位数第 6 天
    星期分布：周四 31 / 周三 27 / 周二 11 / 周一 10 / 周五 9
    按天计数：2→1  3→17  4→12  5→22  6→21  7→7  8→3  9→3  10→1  11→1

    最晚：2019-03 数据 → 2019-04-11（第 11 天），其次 2019-09 → 2019-10-10（第 10 天），
          第 9 天 3 次（2019-04 / 2019-06 / 2019-08）—— 全部集中在 2019 年，
          是 2019-04 IPO 后头两年节奏未定型的产物，2021 起再没超过第 8 天。
          2021 年以后最晚的一次是 2024-06 数据 → 2024-07-08（第 8 天，65 期里仅此一次；
          另两次第 8 天是 2019-07 与 2019-12）。
    最早：2023-01 数据 → 2023-02-02（第 2 天，88 期里仅此一次）；第 3 天出现 17 次。

    ⚠ 2026-08-07 复核修正：本节此前写「最晚 2020-01 → 2020-02-12（第 12 天）」，是错的。
      2020-01 那期的实际发布日是 **2020-02-05（第 5 天）**，已由新闻稿详情页原文核对。
      错因是列表页解析用了跨 <tr> 的非贪婪正则，把某一行的日期配到了另一行的标题上；
      正确做法是**先按 `<tr>` 切块、再在块内取 `<td>` 日期与标题**（本节现有数字已按
      修正后的解析重算，并抽样回详情页验证：2019-04-11 / 2019-10-10 / 2020-02-05 均已对上）。
      全样本上限因此由第 12 天收敛到第 11 天；**LAG / EARLY_BY 不受影响**——它们只取
      2021 年以后的窗口（第 2 至第 8 天），那一段两次解析结果完全一致。

⇒ 给 build/roster.py 的建议是 `LAG = (8, 8)`（两值相同；按 2021 年以后的口径定，
  不被 2019-2020 那几个 IPO 初期的尾巴绑架）。
⚠ 闸门提前量必须写成**元组** `EARLY_BY['lseg'] = (7, 7)` —— monthly_run.py 的取值处是
  `EARLY_BY.get(t, (EARLY, EARLY))[1 if qe else 0]`，写成裸整数会在下标那一步 TypeError，
  **崩掉的是整轮 monthly_run，不只是这一家**。
  为什么是 7：8−7 = 次月第 1 天开闸，比实测最早值（第 2 天）再早一天。第 3 天出现过 17 次，
  零余量迟早会漏；代价只是每月多几个「还没发」的 HTTP 请求。

发布日只认**两处源头自述**：工作簿的 HTTP Last-Modified（主源），
与工作簿内部 docProps/core.xml 的 dcterms:modified（交叉校验）。
（2026-09 官网改版前主源是 href 目录名里的 MM.DD.YY，改版后 URL 里不再有日期，
  换成了现在这两处；替换理由与实测值见 _CORE_MODIFIED 上方那段注释。）
⚠ 发布日定不下来时**只喊不阻断**，`release_date()` 如实返回 None、数据照常入库 ——
  2026-09 那次整条腿停摆正是因为这个字段会抛。分界线与完整理由见 `_resolve_published`。
构建日、文件 mtime、下载时刻一律不算（CONTRACT.md §1 `source_date` 那一条）。
⚠ 本模块**不自己写 series/source_dates.csv** —— LSEG 四路是并行跑的，四个进程同时
  往同一个共享台账里追加必然打架。发布日由 `release_date()` 返回，交给 LSEG 那一路的
  汇总模块统一钉进去。

━━ 口径坑（按踩坑概率排序）━━

**1. 工作簿的单位是「百万美元」，而公司自己对外一律讲「万亿/十亿美元」。**
   `ADV - M` 里 2026-07 的 Grand Total 格子是 `2927614`，不是 2.9 也不是 2927.6。
   实测锚点（2026-08-06 那期新闻稿 vs 工作簿最后一列，全部对上）：
     Grand Total ADV      2,927,614 → $2.93tn   稿子写 "$2.9tn"           ✓
     Grand Total Volume  67,497,639 → $67.50tn  稿子写 "$67.5 trillion"   ✓
     US Government Bonds    258,026 → $258.0bn  稿子写 "$258.0 billion"   ✓
     Swaps/Swaptions ≥1Y    564,528 → $564.5bn  稿子写 "$564.5bn"         ✓
     Repurchase Agreements  832,868 → $832.9bn  稿子写 "$832.9bn"         ✓
   ⇒ ADV 列除以 1e3 得 $bn，月度总成交额除以 1e6 得 $tn。**列名里已经写死单位**，
     build 层不要再做二次换算。

**2. 「Trade Days」不是整数，也不是同一个分母。**
   2026-07 的 Grand Total 交易日是 **23.06** 天，而 US Government Bonds 是 22、
   European Government Bonds 是 23、Other Money Markets 是 **30.43**。
   原因写在 Disclosures 页：ICD Portal（other money markets 的主体）报的是
   「按**自然日**平均的现金余额 ADB」，Repo 报的是抵押品名义额，各产品分母天生不同；
   集团级那个 23.06 是**加权反推**出来的（= 月成交额 ÷ 月 ADV），不是日历事实。
   ⇒ 本模块只落集团级这一个 `tradeweb_trading_days_blended`，且明确叫 blended。
     **不要拿它去算任何单一产品的日均**，也不要拿它跟别家的 trading_days 直接比。

**3. 2024-12 起改过 ADV 分母口径，实测最大一格被改了 +11.5%。**
   Disclosures 原文（从 xlsx 直接读的，不是转述）：
   "Beginning with the publication of the December 2024 Monthly Activity Report,
   Tradeweb adjusted its methodology for reflecting acquisitions in its reported
   average daily volume figures. For average daily volume derived from acquisitions,
   the denominator is now the number of trading days that have elapsed from the
   acquisition date to the end date of the reporting period… Beginning in December
   2024, this methodology was applied retroactively to restate the impact of both
   2024 acquisitions; the average daily volume attributable to acquisitions
   occurring prior to 2024 was not restated."
   本机实测到了它的后果 —— 拿 **2024-02-06 那期新闻稿**与**今天的工作簿**比同一格：

       2024-01 US Government Bonds ADV   当期稿 $182.1bn → 今天 $203.073bn（**+11.5%**）

   ⚠ 因果只说到能证的那一层：这个日期与 10-K 记载的 r8fin 交割日 **2024-01-19**
   （另一笔是 ICD，2024-08-01；两笔就是官方说的 "both 2024 acquisitions"）以及
   官方自述的分母改法完全吻合 —— 2024 年 1 月只有交割后的几个交易日，换分母后
   ADV 自然跳升。但**官方没有给逐项对照表**，所以本模块只陈述实测差额，
   不断言这 11.5% 全部由该口径变更造成。
   ⇒ 工作簿**内部**是重述后的一致口径，同比可以算；真正致命的是把某一期新闻稿的
     数字手工补进序列 —— 那会插进一个 11% 的假台阶。别这么干。

**4. 历史值会被官方悄悄改。实测重述率约 6.6%，所以「已有值永不覆盖」。**
   Disclosures 原文："Volumes can reflect cancellations, corrections and settlement
   of NAV trades on ETFs that occur after prior postings; historical volumes are
   periodically updated."
   本机实测（2023-01 起每 3 个月抽 1 期新闻稿，共 15 期，与今天的工作簿逐格比，
   **121 个可比「字段×月份」**）：

       一致 113 个，被事后改过 8 个 —— 6.6%
       改动幅度：6 处在 ±0.15% 以内（四舍五入级），
                 1 处 −0.83%（2023-10 Swaps/Swaptions ≥1Y 463.4 → 459.535），
                 1 处 +11.5%（2024-01 美债，见坑 3）
       最爱被改的是 Other Money Markets（8 处里占 4 处）—— 与它按 ADB / 自然日
       口径结算、且事后才轧账的性质一致（见坑 2）。

   ⇒ `write_csv()` 只填空、不覆盖；冲突写 `cache/lseg_tradeweb_restatements.csv`
     供人工判断。跟 enx / db1 一个处理方式，理由也一样：自动吞重述会在序列里
     插一个没人知道的台阶。

**5. 「Credit」总 ADV 跟新闻稿标题里的 credit 数字对不上，因为口径不是一回事。**
   2026-07：工作簿 Credit Total ADV = 39,998（$40.0bn），
   而稿子标题只讲 "Fully electronic U.S. credit ADV … $9.4bn"。
   Credit Total = 现券（含 electronically processed）+ 信用衍生品（$21.1bn）+ 中国债 + 其他。
   ⇒ 拿 `tradeweb_adv_credit_usd_bn` 去对新闻稿里的 "U.S. credit ADV" 必然对不上，
     那不是解析错了。要对，请用 `tradeweb_adv_us_hg_fully_electronic_usd_bn`
     与 `tradeweb_adv_us_hy_fully_electronic_usd_bn` 之和。

**6. 成交额是名义本金，且部分市场双边计。**
   Disclosures："Both sides of a trade are included in volume totals where the trade
   is fully-anonymous and a Tradeweb broker dealer is the matched principal
   counterparty. In wholesale markets, U.S. Treasuries and mortgages volumes are
   single sided…"；按揭按 current face value；Repo 按抵押品名义额。
   ⇒ 跟交易所的「成交金额」不是同一种量，**跨家横比要先说明口径**，不要直接并排画柱。

**7. 非美元品种按「上一个月」的月均汇率折美元。**
   Disclosures："…the non-U.S. dollar amount for a particular month is translated
   into U.S. Dollars generally based on the monthly average foreign exchange rate
   for the prior month."
   ⇒ 欧债 / 欧洲信用这些列里含**一个月的汇率滞后**，欧元急涨急跌的月份不要把
     ADV 的变化全算成量的变化。

**8. Chinese Bonds 在 2017-01～2017-06 是 0（Bond Connect 2017-07 才通）。**
   那 6 个 0 是真实的「当时没有这个业务」，不是缺失。本模块**不落这一列**
   （它对 LSEG 拼图没用），提一句是防止有人回头加列时把 0 当成解析 bug。

━━ 落地列（27 列 + month，全部来源等级 [A]：公司原始披露）━━

    month                                        YYYY-MM，升序，首列
    tradeweb_volume_total_usd_tn                 全公司月度总成交额，万亿美元
    tradeweb_adv_total_usd_bn                    全公司月度 ADV，十亿美元
    tradeweb_trading_days_blended                集团级加权交易日数（见坑 2）
    tradeweb_adv_rates_usd_bn                    Rates 合计 ADV
    tradeweb_adv_rates_cash_usd_bn               Rates - Cash 合计
    tradeweb_adv_rates_derivatives_usd_bn        Rates - Derivatives 合计
    tradeweb_adv_us_govt_bonds_usd_bn            美国国债
    tradeweb_adv_eu_govt_bonds_usd_bn            欧洲国债
    tradeweb_adv_mortgages_usd_bn                按揭（TBA / specified pool / CMO / CMBS / ABS）
    tradeweb_adv_other_govt_bonds_usd_bn         其他政府债（日加澳新 / covered / 超主权 / 机构）
    tradeweb_adv_swaps_swaptions_ge_1y_usd_bn    利率互换/掉期期权 ≥1 年
    tradeweb_adv_swaps_swaptions_lt_1y_usd_bn    利率互换/掉期期权 <1 年
    tradeweb_adv_credit_usd_bn                   Credit 合计 ADV（见坑 5）
    tradeweb_adv_credit_cash_usd_bn              Credit - Cash 合计
    tradeweb_adv_credit_derivatives_usd_bn       Credit - Derivatives 合计（指数与单名 CDS）
    tradeweb_adv_us_hg_fully_electronic_usd_bn   美国投资级 · 全电子
    tradeweb_adv_us_hy_fully_electronic_usd_bn   美国高收益 · 全电子
    tradeweb_adv_european_credit_usd_bn          欧洲信用债
    tradeweb_adv_municipal_bonds_usd_bn          美国市政债
    tradeweb_adv_equities_usd_bn                 Equities 合计 ADV
    tradeweb_adv_equities_cash_usd_bn            Equities - Cash 合计
    tradeweb_adv_equities_derivatives_usd_bn     Equities - Derivatives 合计
    tradeweb_adv_us_etf_usd_bn                   美国 ETF
    tradeweb_adv_intl_etf_usd_bn                 国际 ETF
    tradeweb_adv_money_markets_usd_bn            Money Markets 合计 ADV
    tradeweb_adv_repo_usd_bn                     回购
    tradeweb_adv_other_money_markets_usd_bn      其他货币市场（含 ICD Portal，分母是自然日）

没落的字段与原因，写在这里省得下一个人再找一遍：
  · **各资产类别的月度成交额**（只落了集团级总额）—— 有数据，但 ADV 才是经营量，
    且分类别成交额 = 分类别 ADV × 各自不同的分母，落进来只会诱导错误的口径混用。
  · **TRACE 市占率**（`TRACE - M` sheet 有）—— 那是 Tradeweb 用 FINRA 数据自算的
    市占估计，属 [C] 推算，与本仓「只收原始披露」的铁律冲突，不落。
  · **fee per million（FPM）/ 固定费用** —— 索引页与新闻稿都写明「preliminary，
    subject to management's final review」，是会被改的预估值，不落。
  · **Chinese Bonds / Other Credit Bonds / Rates Futures / Equities Futures /
    Convertibles-Swaps-Options** —— 有数据且完整，纯粹是本次不需要，不是抓不到。

━━ 并表关系：Tradeweb 与 LSEG 是什么关系（不进 CSV）━━

以下每一条都出自 Tradeweb 10-K（FY2025，2026-02-05 报送，accession 0001758730-26-000015）
原文，等级 [A]：

  · 2021-01-29 LSEG 以全股票方式完成对 Refinitiv 的收购。10-K 原文：
    "In connection with the LSEG Transaction, the Corporation became a consolidating
    subsidiary of LSEG."（此前 Tradeweb 是 Blackstone 系 BCP York Holdings 的并表子公司）
  · 截至 2025-12-31，Refinitiv（LSEG 子公司）持有 Tradeweb **89.9% 的合并投票权**、
    **50.9% 的 TWM LLC 经济权益**；公众股东持 Class A 共 115,502,689 股，
    占 10.0% 投票权、49.0% 经济权益。
  · ⇒ LSEG **100% 并表** Tradeweb 的收入与成交量，同时在利润表下方确认约一半的
    少数股东权益。Tradeweb 仍是独立上市公司，自己按月披露。

它落在 LSEG 的哪个分部（两条都取自 LSEG 官网原文，等级 [A]）：
  · lseg.com/en/investor-relations 原文："our four business divisions – Data and
    Analytics, FTSE Russell, Risk Intelligence, and **Markets** – …"
    ⇒ LSEG 是四分部结构，其中一个就叫 **Markets**。
  · lseg.com/en/about-us/what-we-do 原文："We are home to several capital formation
    and execution venues: the London Stock Exchange, AIM, Turquoise, FXall,
    FX Matching and **Tradeweb**. **LSEG Markets** combines these flagship trading
    services with LCH, Post Trade Solutions and Regulatory Reporting Solutions…"
    ⇒ Tradeweb 与 LSE 主板 / AIM / Turquoise / FXall / LCH 同属 **LSEG Markets** 分部。
  ⚠ 以上是 LSEG 公司网站对业务结构的表述。**审计过的分部附注是否用完全相同的边界，
    本模块没有从年报核过** —— 要在交付物里引用分部收入数字，请由负责 LSEG 集团那一路
    的模块从年报/中报的分部附注里取，别从这段话推。

它跟 LSEG 集团分部收入的关系，一句话说清：
**Tradeweb 的 ADV 是「量」，LSEG 分部收入是「量 × 费率 + 固定费」，两者之间隔着
一个本模块拿不到的费率。** 具体说，(a) Tradeweb 收入里有相当一块是与成交量无关的
固定订阅费；(b) 变动费部分是 ADV × FPM，而 FPM 在产品间差一到两个数量级
（新闻稿反复强调 portfolio trading、compression 这些「低 FPM」协议占比会拉低混合费率），
所以**同样的 ADV 增速绝不等于同样的收入增速**；(c) LSEG 报表以英镑列报，
Tradeweb 以美元，中间还有汇率。
⇒ 本模块的产出只能用来解释 LSEG Markets 分部里 **Tradeweb 这一块的量**，
  **不能**用来推算该分部收入 —— 除了上面三条，Markets 分部里还装着 LSE 主板、AIM、
  Turquoise、FXall、FX Matching、LCH 与 Post Trade Solutions，Tradeweb 只是其中一个场所。
  谁要拿它做收入桥，必须自己去拿 FPM，而 FPM 是 preliminary 值（见上文「没落的字段」）。
"""

import csv
import datetime
import io
import os
import re
import urllib.request
import zipfile

# ── 常量 ────────────────────────────────────────────────────────────────────
PART = 'tradeweb'

INDEX_URL = 'https://www.tradeweb.com/newsroom/monthly-activity-reports/'
HOST = 'https://www.tradeweb.com'

_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36')

_HERE = os.path.dirname(os.path.abspath(__file__))
_REPO = os.path.dirname(_HERE)
DEFAULT_CACHE = os.path.join(_REPO, 'cache')
DEFAULT_SERIES = os.path.join(_REPO, 'series')

# build/roster.py 用；理由见文首「发布节奏」
LAG = (8, 8)
EARLY_BY = (7, 7)

ADV_SHEET = 'ADV - M'
VOLUME_SHEET = 'Volume - M'
DAYS_SHEET = 'Trade Days - M'

# 工作簿三列标签（资产类别 / Cash·Derivatives / 产品）→ 落地列名。
# 标签向下填充后三元组唯一，**按标签定位而不是按行号**：官方加一行产品就整体错位。
_LABEL_COLS = 3
_HEADER_ROW = 2
_FIRST_DATA_COL = 4

# 每项：(标签三元组, 列名, 从「百万美元」出发的除数)
_MM_TO_BN = 1e3
_MM_TO_TN = 1e6

ADV_MAP = [
    (('Grand Total', '', ''), 'tradeweb_adv_total_usd_bn', _MM_TO_BN),
    (('Rates', 'Total', ''), 'tradeweb_adv_rates_usd_bn', _MM_TO_BN),
    (('Rates', 'Cash', 'Total'), 'tradeweb_adv_rates_cash_usd_bn', _MM_TO_BN),
    (('Rates', 'Derivatives', 'Total'), 'tradeweb_adv_rates_derivatives_usd_bn', _MM_TO_BN),
    (('Rates', 'Cash', 'US Government Bonds'), 'tradeweb_adv_us_govt_bonds_usd_bn', _MM_TO_BN),
    (('Rates', 'Cash', 'European Government Bonds'), 'tradeweb_adv_eu_govt_bonds_usd_bn', _MM_TO_BN),
    (('Rates', 'Cash', 'Mortgages'), 'tradeweb_adv_mortgages_usd_bn', _MM_TO_BN),
    (('Rates', 'Cash', 'Other Government Bonds'), 'tradeweb_adv_other_govt_bonds_usd_bn', _MM_TO_BN),
    (('Rates', 'Derivatives', 'Swaps/Swaptions ≥ 1Y'),
     'tradeweb_adv_swaps_swaptions_ge_1y_usd_bn', _MM_TO_BN),
    (('Rates', 'Derivatives', 'Swaps/Swaptions < 1Y'),
     'tradeweb_adv_swaps_swaptions_lt_1y_usd_bn', _MM_TO_BN),
    (('Credit', 'Total', ''), 'tradeweb_adv_credit_usd_bn', _MM_TO_BN),
    (('Credit', 'Cash', 'Total'), 'tradeweb_adv_credit_cash_usd_bn', _MM_TO_BN),
    (('Credit', 'Derivatives', 'Total'), 'tradeweb_adv_credit_derivatives_usd_bn', _MM_TO_BN),
    (('Credit', 'Cash', 'US High Grade - Fully Electronic'),
     'tradeweb_adv_us_hg_fully_electronic_usd_bn', _MM_TO_BN),
    (('Credit', 'Cash', 'US High Yield - Fully Electronic'),
     'tradeweb_adv_us_hy_fully_electronic_usd_bn', _MM_TO_BN),
    (('Credit', 'Cash', 'European Credit'), 'tradeweb_adv_european_credit_usd_bn', _MM_TO_BN),
    (('Credit', 'Cash', 'Municipal Bonds'), 'tradeweb_adv_municipal_bonds_usd_bn', _MM_TO_BN),
    (('Equities', 'Total', ''), 'tradeweb_adv_equities_usd_bn', _MM_TO_BN),
    (('Equities', 'Cash', 'Total'), 'tradeweb_adv_equities_cash_usd_bn', _MM_TO_BN),
    (('Equities', 'Derivatives', 'Total'), 'tradeweb_adv_equities_derivatives_usd_bn', _MM_TO_BN),
    (('Equities', 'Cash', 'US ETFs'), 'tradeweb_adv_us_etf_usd_bn', _MM_TO_BN),
    (('Equities', 'Cash', 'International ETFs'), 'tradeweb_adv_intl_etf_usd_bn', _MM_TO_BN),
    (('Money Markets', 'Total', ''), 'tradeweb_adv_money_markets_usd_bn', _MM_TO_BN),
    (('Money Markets', 'Cash', 'Repurchase Agreements'), 'tradeweb_adv_repo_usd_bn', _MM_TO_BN),
    (('Money Markets', 'Cash', 'Other Money Markets'),
     'tradeweb_adv_other_money_markets_usd_bn', _MM_TO_BN),
]

VOLUME_MAP = [
    (('Grand Total', '', ''), 'tradeweb_volume_total_usd_tn', _MM_TO_TN),
]

DAYS_MAP = [
    (('Grand Total', '', ''), 'tradeweb_trading_days_blended', 1.0),
]

COLUMNS = (['month', 'tradeweb_volume_total_usd_tn']
           + [c for _, c, _ in ADV_MAP]
           + ['tradeweb_trading_days_blended'])

# 校验容差。三条恒等式的实测最大残差（115 个月全跑过）：
#   四大类 ADV 之和 vs Grand Total          2.8e-06
#   Cash + Derivatives vs 类别合计           2.0e-04
#   Grand Total ADV × 加权交易日 vs 月成交额 2.5e-04
# 残差都来自官方把数字四舍五入成整数（交易日 2 位小数），不是解析错。
TOL_CLASS_SUM = 1e-3
TOL_CASH_DERIV = 1e-3
TOL_ADV_DAYS = 2e-3

MIN_MONTHS = 36            # 历史深度红线，低于它说明解析到了残缺文件
MIN_XLSX_BYTES = 50_000


class TradewebFetchError(RuntimeError):
    """这一路自己的异常。缺列、结构变了、恒等式对不上，一律抛它，绝不静默写 NaN。"""


# ── 依赖 ────────────────────────────────────────────────────────────────────
def _openpyxl():
    try:
        import openpyxl                              # noqa: PLC0415
    except ImportError as e:                         # pragma: no cover
        raise TradewebFetchError(
            'lseg_tradeweb 需要 openpyxl 读官方 xlsx：pip install openpyxl') from e
    return openpyxl


# ── 网络 ────────────────────────────────────────────────────────────────────
_CHALLENGE = (b'Just a moment', b'_cf_chl_opt', b'cf-mitigated')


def _looks_like_challenge(data):
    head = data[:4000]
    return any(m in head for m in _CHALLENGE)


def _urllib_get(url, timeout):
    req = urllib.request.Request(url, headers={
        'User-Agent': _UA,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
    })
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read(), dict(r.headers)


def _curl_cffi_get(url, timeout):
    """回落通道。见文首：Cloudflare 拦的是指纹，chrome124 伪装能过。"""
    try:
        from curl_cffi import requests as cffi_requests   # noqa: PLC0415
    except ImportError as e:
        raise TradewebFetchError(
            'urllib 被 Cloudflare 挑战挡住，回落通道需要 curl_cffi：pip install curl_cffi'
        ) from e
    r = cffi_requests.get(url, impersonate='chrome124', timeout=timeout)
    if r.status_code != 200:
        raise TradewebFetchError('curl_cffi 回落也失败 %s -> HTTP %d' % (url, r.status_code))
    return r.content, dict(r.headers)


def _http_get(url, timeout=120, retries=3):
    """先 urllib（HTTP/1.1，能过 Cloudflare），拿到挑战页才回落 curl_cffi。

    ⚠ 不要「优化」成用 curl / requests：那两个走 HTTP/2，实测**必定** 403。
    """
    last = None
    for attempt in range(retries):
        try:
            data, headers = _urllib_get(url, timeout)
            if not _looks_like_challenge(data):
                return data, headers
            last = TradewebFetchError('urllib 拿回来的是 Cloudflare 挑战页')
        except TradewebFetchError:
            raise
        except Exception as e:                        # noqa: BLE001
            last = e
        if attempt < retries - 1:
            import time                               # noqa: PLC0415
            time.sleep(2 * (attempt + 1))
    data, headers = _curl_cffi_get(url, timeout)
    if _looks_like_challenge(data):
        raise TradewebFetchError('urllib 与 curl_cffi 都只拿到挑战页 %s（末次 %r）' % (url, last))
    return data, headers


def _cache(cache_dir, name):
    os.makedirs(cache_dir, exist_ok=True)
    return os.path.join(cache_dir, name)


# ── 小工具 ──────────────────────────────────────────────────────────────────
_MONTHS = ['January', 'February', 'March', 'April', 'May', 'June',
           'July', 'August', 'September', 'October', 'November', 'December']
_MON_NUM = {m: i + 1 for i, m in enumerate(_MONTHS)}
_MON_NUM.update({m[:3]: i + 1 for i, m in enumerate(_MONTHS)})


def _norm(v):
    return re.sub(r'\s+', ' ', str(v)).strip() if v is not None else ''


def _num(v):
    """空白 / 非数字 → None；数字 → float。绝不把 '-' 之类当成 0。"""
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    if s in ('', '-', 'n/a', 'N/A'):
        return None
    try:
        return float(s.replace(',', ''))
    except ValueError:
        return None


def _fmt(v):
    """整数写成整数、小数用最短往返表示，保证重跑字节级不变（同 db1._fmt）。"""
    if v is None:
        return ''
    v = float(v)
    if v.is_integer() and abs(v) < 1e15:
        return str(int(v))
    return repr(v)


def _next_month_first(month):
    y, m = int(month[:4]), int(month[5:7])
    return datetime.date(y + 1, 1, 1) if m == 12 else datetime.date(y, m + 1, 1)


def _months_between(start, end):
    y, m = int(start[:4]), int(start[5:7])
    out = []
    while True:
        cur = '%04d-%02d' % (y, m)
        if cur > end:
            return out
        out.append(cur)
        m += 1
        if m == 13:
            y, m = y + 1, 1


# ── 索引页发现 ──────────────────────────────────────────────────────────────
_XLSX_HREF = re.compile(r'href="([^"]*globalassets[^"]*\.xlsx)"', re.I)
_HIST_NAME = re.compile(r'historical[-_ ]?adv|adv[-_ ]?and[-_ ]?day[-_ ]?count', re.I)
# 发布日。**2026-09 官网改版后这里换过一次源**，两处都不是我们自己编的日期：
#   旧：href 里的目录名 `/globalassets/newsroom/08.06.26-july-mar/…`，MM.DD.YY 就是发布日；
#   新：目录名换成 `/<缓存串>/globalassets/newsroom/monthly-activity-reports/2026/august/`，
#       日期整个没了（实测 2026-09-05 的 href：`/4a4dd2/…/2026/august/tw-historical-adv-and-
#       day-count-through-august-2026.xlsx`；那个 4a4dd2 每份文件各不相同，是缓存串不是日期）。
# 现在改用 HTTP Last-Modified 当主源 + 工作簿自己的 docProps/core.xml 交叉校验 ——
# 仍然是**两处源头自述**，没有退化成单源，也没有退回构建日 / 文件 mtime（CONTRACT.md §1 禁止）。
_CORE_MODIFIED = re.compile(r'<dcterms:modified[^>]*>(\d{4})-(\d{2})-(\d{2})', re.I)
# 作者时间戳与上架时间的允许间隔。实测 2026-08 那期：core.xml 说 09-02、Last-Modified 说 09-04。
# 定 30 天是因为它只用来抓「两个数字根本不是一回事」，不是用来卡发布节奏。
_AUTHORED_MAX_LAG = 30
# 发布日相对数据月末的允许窗口。Tradeweb 实测在次月头几天发（2026-08 那期是月末后第 4 天）。
# 75 天是「宽到不会误伤补发，窄到接不住去年的文件」。
_PUBLISH_MAX_LAG = 75
# 文件名自己带数据月：…-through-july-2026.xlsx。这是最硬的新鲜度锚点 ——
# 它长在**我们真正下载的那个文件**上，抬头句子只是页面文案，可以先改。
_FILE_MONTH = re.compile(r'through[-_ ]([a-z]+)[-_ ](\d{4})\.xlsx$', re.I)
# 抬头句子：措辞这些年换过至少 6 种（"Reports Record March 2026…" /
# "Reports Total November 2022 Trading Volume…"），所以先框住一个窗口、再在窗口内找
# 月份与数字，不写死语序。解析不出来只降级为「不做这项交叉核对」，不整体失败。
# ⚠ 别用 `[^.]{0,N}` 当「一句话」的边界 —— 金额里就带小数点（"$67.5 Trillion"），
#   那样框出来的窗口会在 "$67" 处截断，抬头永远解析不出数字。用定长窗口 + 禁止
#   跨到下一个 "Tradeweb Reports"。
_HEADLINE_SENT = re.compile(r'Tradeweb Reports(?:(?!Tradeweb Reports).){0,400}', re.S)
_MON_YEAR = re.compile(r'\b(%s)\s+(\d{4})\b' % '|'.join(_MONTHS))
_VOL_TN = re.compile(r'Trading Volume of \$([\d.]+)\s*(?:Trillion|tn)\b', re.I)
_ADV_TN = re.compile(r'Average Daily Volume of \$([\d.]+)\s*(?:Trillion|tn)\b', re.I)


def _http_date(raw):
    """HTTP-date → datetime.date；解析不出返回 None（怎么处置由调用方定，见 _resolve_published）。"""
    if not raw:
        return None
    try:
        return datetime.datetime.strptime(raw[:25].strip(), '%a, %d %b %Y %H:%M:%S').date()
    except ValueError:
        return None


def _authored_date(data):
    """工作簿自己的 docProps/core.xml 里的 dcterms:modified → date；没有就 None。

    这是**文件自述**，与 HTTP 头是两个独立来源：头由 Tradeweb 那边的服务器写，
    core.xml 由出报表那条工具链写（实测 2026-08 那期 `dc:creator` = Workiva）。
    两个都指向同一个发布周，才敢认这个发布日。
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as z:
            core = z.read('docProps/core.xml').decode('utf-8', 'replace')
    except (zipfile.BadZipFile, KeyError):
        return None
    m = _CORE_MODIFIED.search(core)
    if not m:
        return None
    try:
        return datetime.date(*(int(x) for x in m.groups()))
    except ValueError:
        return None


def _resolve_published(url, month):
    """下工作簿，定发布日。返回 (工作簿字节, published | None)。

    主源是工作簿的 HTTP Last-Modified，两道交叉校验：
      1. 工作簿内部 docProps/core.xml 的 dcterms:modified（文件自述，独立于 HTTP 头）；
      2. 数据月自己 —— 发布日必须落在数据月末之后 _PUBLISH_MAX_LAG 天以内。
    三道都不依赖 href 长什么样。改版把日期从 URL 里拿走之后，这是剩下的全部证据。

    ══ 为什么发布日的判据一条都不抛 ══
    2026-09 那次停摆就是抛出来的：改版把目录名里的日期拿走，一条**只服务发布日**的
    正则失配，却把整条腿打死 —— 而那一期的数据完完整整躺在工作簿里，116 个月一格没少。
    发布日在本仓的实际去向只有 fetch/lseg.py 的 `release_dates()`，那边本来就接受 None
    （它自己 try/except 兜着，对不上就记 None），而且 `series/source_dates.csv` 至今
    没有 lseg 行 —— 这个字段在页面上一个字都印不出来。让一个印不出来的字段拦住
    116 个月的真数据，方向是反的。
    照 fetch/cme.py:468（同样是「没有 Last-Modified 头」）与 fetch/enx.py:1140 的既定
    写法：护栏失效只喊不阻断 ——「护栏失效的代价是漏报一阵子，护栏误报的代价是整家停更」。

    ⚠ 分界线在「护的是发布日还是护的是数据」，不在「严不严」：
      · 体积不够（多半下到了 HTML 错误页）—— 在这里**抛**；
      · 「文件名说的数据月 vs 索引页抬头说的数据月」不一致 —— 在 discover() 里**抛**。
      那两条护的是数据本身，一条都没有放松。

    ⚠ 这里做的是**整份下载**而不是 HEAD：交叉校验要读工作簿内容，而 _http_get 那条
      Cloudflare 回落链路（urllib → curl_cffi）也只支持 GET。字节顺手返回给
      download_workbook 落盘，所以一轮里只下一次，没有多花请求。
    """
    data, headers = _http_get(url)
    if len(data) < MIN_XLSX_BYTES:
        raise TradewebFetchError('%s 只有 %d 字节，不像是正常的工作簿' % (url, len(data)))

    published = _http_date(headers.get('Last-Modified') or headers.get('last-modified'))
    if published is None:
        print('[lseg_tradeweb] ⚠ 护栏失效：工作簿 %s 的响应里没有可解析的 Last-Modified，'
              '本期发布日记为「未知」（构建日 / 文件 mtime 不算，CONTRACT.md §1）。'
              '数据照常入库。' % url)
        return data, None

    y, mo = (int(x) for x in month.split('-'))
    month_end = datetime.date(y + mo // 12, mo % 12 + 1, 1) - datetime.timedelta(days=1)
    lag = (published - month_end).days
    if not 0 < lag <= _PUBLISH_MAX_LAG:
        print('[lseg_tradeweb] ⚠ 护栏失效：Last-Modified 说 %s，落在数据月 %s（月末 %s）'
              '之后第 %d 天，不在 1..%d 天的窗口里 —— 它多半不是这份文件真正的上架时间，'
              '本期发布日记为「未知」。数据照常入库。'
              % (published, month, month_end, lag, _PUBLISH_MAX_LAG))
        return data, None

    authored = _authored_date(data)
    if authored is None:
        print('[lseg_tradeweb] ⚠ 护栏失效：工作簿里没读到 docProps/core.xml 的 '
              'dcterms:modified，本期发布日 %s 只剩 HTTP Last-Modified 一处自述'
              '（另有数据月窗口兜底）。' % published)
        return data, published

    gap = (published - authored).days
    if not 0 <= gap <= _AUTHORED_MAX_LAG:
        print('[lseg_tradeweb] ⚠ 护栏失效：发布日两处对不上 —— HTTP Last-Modified 说 %s，'
              '工作簿自己的 dcterms:modified 说 %s（差 %d 天，允许 0..%d）。'
              '本期发布日记为「未知」。数据照常入库。'
              % (published, authored, gap, _AUTHORED_MAX_LAG))
        return data, None
    return data, published


def discover(cache_dir=DEFAULT_CACHE):
    """抓索引页，解析出「历史 ADV 工作簿」的绝对 URL、发布日与自述抬头。

    **只从页面 href 里取链接**。上一轮有 agent 在这类站上反复猜文件名、全是 404、
    烧掉半个上下文 —— 这个函数就是那条教训的固化：文件名与目录名每月都变
    （缓存串 + MM.DD.YY），猜是猜不出来的。

    返回 dict：
        url            工作簿绝对 URL
        data           工作簿字节（顺手带出来，download_workbook 直接落盘，不再下第二次）
        published      发布日 datetime.date，**定不下来时是 None**（不阻断数据，
                       理由见 _resolve_published 的「为什么发布日的判据一条都不抛」）
        month          数据月 'YYYY-MM'，取自**文件名**里的 through-<month>-<year>
        headline_vol_tn / headline_adv_tn  抬头自述的总成交额与总 ADV（万亿美元，
                                           已四舍五入）；抬头措辞变了就是 None，
                                           那只关掉数值交叉核对，不让整路失败
        headline_decimals                  自述值的小数位数，供比对时定容差
    """
    raw, _ = _http_get(INDEX_URL)
    html = raw.decode('utf-8', 'replace')
    with open(_cache(cache_dir, 'lseg_tradeweb_index.html'), 'w', encoding='utf-8') as f:
        f.write(html)

    hrefs = sorted(set(_XLSX_HREF.findall(html)))
    if not hrefs:
        raise TradewebFetchError(
            '索引页 %s 上一个 globalassets 的 .xlsx 都没解析出来 —— 页面多半改版了，'
            '去看 cache/lseg_tradeweb_index.html' % INDEX_URL)
    hits = [h for h in hrefs if _HIST_NAME.search(h)]
    if len(hits) != 1:
        raise TradewebFetchError(
            '索引页上「历史 ADV 工作簿」不是唯一命中（命中 %d 个）。页面全部 xlsx：%s'
            % (len(hits), hrefs))
    href = hits[0]

    fm = _FILE_MONTH.search(href)
    if not fm or fm.group(1).capitalize() not in _MON_NUM:
        raise TradewebFetchError(
            '工作簿文件名里读不出数据月（期望 …-through-<month>-<year>.xlsx）：%s' % href)
    month = '%04d-%02d' % (int(fm.group(2)), _MON_NUM[fm.group(1).capitalize()])

    vol = adv = None
    decimals = (0, 0)
    flat = re.sub(r'\s+', ' ', re.sub(r'<[^>]+>', ' ', html))
    sent = next((m.group(0) for m in _HEADLINE_SENT.finditer(flat)
                 if _VOL_TN.search(m.group(0)) and _ADV_TN.search(m.group(0))), None)
    if sent is None:
        sent = _HEADLINE_SENT.search(flat)
        sent = sent.group(0) if sent else None
    if sent:
        text = sent
        my = _MON_YEAR.search(text)
        if my:
            hm_month = '%04d-%02d' % (int(my.group(2)), _MON_NUM[my.group(1)])
            if hm_month != month:
                raise TradewebFetchError(
                    '文件名说数据月是 %s，索引页抬头却说 %s —— 页面与挂件不同步，停下来人工看'
                    % (month, hm_month))
        v, a = _VOL_TN.search(text), _ADV_TN.search(text)
        if v and a:
            vol, adv = float(v.group(1)), float(a.group(1))
            decimals = (len(v.group(1).split('.')[1]) if '.' in v.group(1) else 0,
                        len(a.group(1).split('.')[1]) if '.' in a.group(1) else 0)

    # 发布日放在最后：上面那道「文件名说的月 vs 抬头说的月」不一致时直接抛，
    # 那种日子不必先把 170KB 的工作簿下下来。
    url = href if href.startswith('http') else HOST + href
    data, published = _resolve_published(url, month)

    return {
        'url': url,
        'data': data,
        'published': published,
        'month': month,
        'headline_vol_tn': vol,
        'headline_adv_tn': adv,
        'headline_decimals': decimals,
    }


def download_workbook(found, cache_dir=DEFAULT_CACHE):
    """把 discover() 已经取回的工作簿字节落到 cache，返回路径。

    发布日的两道交叉校验在 discover() → _resolve_published() 里就做完了（改版前
    那道「目录名 vs Last-Modified」在这里做，是因为当时目录名不用下载就能读到）。
    """
    path = _cache(cache_dir, 'lseg_tradeweb_hist_adv_%s.xlsx' % found['month'])
    with open(path, 'wb') as f:
        f.write(found['data'])
    return path


# ── 工作簿解析 ──────────────────────────────────────────────────────────────
def _sheet_index(sh):
    """把三列标签向下填充成唯一三元组，返回 {(a, b, c): 行号}。

    **不用行号定位**：官方哪天在 Rates 底下加一行产品，写死的行号会整体错位一行，
    而且错得悄无声息 —— 那种故障不会报错，只会让某一列从此指向别的产品。
    """
    idx, carry = {}, ['', '', '']
    for r in range(_HEADER_ROW + 1, sh.max_row + 1):
        cells = [_norm(sh.cell(r, c).value) for c in range(1, _LABEL_COLS + 1)]
        if not any(cells):
            continue
        for i, v in enumerate(cells):
            if v:
                carry[i] = v
                for j in range(i + 1, _LABEL_COLS):
                    carry[j] = ''
        key = tuple(carry)
        if key in idx:
            raise TradewebFetchError('工作簿 %r 里标签三元组重复：%s（第 %d 行与第 %d 行）'
                                     % (sh.title, key, idx[key], r))
        idx[key] = r
    return idx


def _months_of(sh):
    """表头 'Jan 2017' … 'Jul 2026' → [('YYYY-MM', 列号), …]，并强制连续升序。"""
    out = []
    for c in range(_FIRST_DATA_COL, sh.max_column + 1):
        h = _norm(sh.cell(_HEADER_ROW, c).value)
        if not h:
            continue
        m = re.match(r'^([A-Z][a-z]{2})[a-z]* (\d{4})$', h)
        if not m:
            raise TradewebFetchError(
                '工作簿 %r 第 %d 列表头 %r 不是 "Mon YYYY" —— 表头格式变了' % (sh.title, c, h))
        out.append(('%04d-%02d' % (int(m.group(2)), _MON_NUM[m.group(1)]), c))
    if not out:
        raise TradewebFetchError('工作簿 %r 一个月份表头都没解析出来' % sh.title)
    months = [m for m, _ in out]
    expect = _months_between(months[0], months[-1])
    if months != expect:
        missing = [m for m in expect if m not in months]
        raise TradewebFetchError(
            '工作簿 %r 的月份不连续/不升序：%d 列覆盖 %s→%s，缺 %s'
            % (sh.title, len(months), months[0], months[-1], missing[:12]))
    return out


def _pull(sh, mapping, months):
    """按标签取一整张表的目标行，返回 {month: {col: value}}。

    任何一个标签三元组找不到 → 抛异常（缺列一律失败）。
    任何一个月份的格子是空 / 非数字 → 抛异常（绝不写 NaN）。
    """
    idx = _sheet_index(sh)
    missing = [key for key, _, _ in mapping if key not in idx]
    if missing:
        raise TradewebFetchError(
            '工作簿 %r 缺这些行：%s。现有标签三元组共 %d 组，样例：%s'
            % (sh.title, missing, len(idx), sorted(idx)[:8]))
    out = {mon: {} for mon, _ in months}
    for key, col, div in mapping:
        r = idx[key]
        for mon, c in months:
            v = _num(sh.cell(r, c).value)
            if v is None:
                raise TradewebFetchError(
                    '工作簿 %r 的 %s（第 %d 行）在 %s 是空/非数字：%r —— 不补 NaN，直接失败'
                    % (sh.title, key, r, mon, sh.cell(r, c).value))
            out[mon][col] = v / div
    return out


def _validate(rows_by_month, months):
    """三条恒等式 + 深度红线。对不上就抛，不许「差一点点就算了」。"""
    if len(months) < MIN_MONTHS:
        raise TradewebFetchError(
            '只解析出 %d 个月，低于历史深度红线 %d —— 多半下到了残缺文件'
            % (len(months), MIN_MONTHS))
    for mon in months:
        r = rows_by_month[mon]
        tot = r['tradeweb_adv_total_usd_bn']
        s = sum(r['tradeweb_adv_%s_usd_bn' % k]
                for k in ('rates', 'credit', 'equities', 'money_markets'))
        if abs(s - tot) > TOL_CLASS_SUM * abs(tot):
            raise TradewebFetchError(
                '%s 四大类 ADV 之和 %.3f 对不上 Grand Total %.3f（相对差 %.2e）'
                % (mon, s, tot, abs(s - tot) / abs(tot)))
        for k in ('rates', 'credit', 'equities'):
            a = r['tradeweb_adv_%s_usd_bn' % k]
            b = r['tradeweb_adv_%s_cash_usd_bn' % k] + r['tradeweb_adv_%s_derivatives_usd_bn' % k]
            if abs(b - a) > TOL_CASH_DERIV * abs(a):
                raise TradewebFetchError(
                    '%s %s 的 Cash+Derivatives %.3f 对不上类别合计 %.3f' % (mon, k, b, a))
        vol_bn = r['tradeweb_volume_total_usd_tn'] * 1e3
        implied = tot * r['tradeweb_trading_days_blended']
        if abs(implied - vol_bn) > TOL_ADV_DAYS * abs(vol_bn):
            raise TradewebFetchError(
                '%s ADV×加权交易日 %.1f 对不上月成交额 %.1f（相对差 %.2e）'
                % (mon, implied, vol_bn, abs(implied - vol_bn) / abs(vol_bn)))


def _crosscheck_headline(found, rows_by_month):
    """核对工作簿最后一个月 —— 防止下到上一期的缓存文件。

    两层：
      a) 文件名自述的数据月必须**就是**工作簿最后一个月（硬性，对不上直接抛）；
      b) 索引页抬头的两个数字对得上（软性，抬头措辞变了就跳过 —— 见 discover）。
    抬头值是四舍五入过的（"$67.5 Trillion" 只有 1 位小数），容差按它自己的小数位数定：
    1 位 → ±0.05，2 位 → ±0.005。不要因为「差了 0.03」就放宽容差，
    真下错文件时差的是整整一个月的量级。
    """
    mon = found['month']
    if mon != max(rows_by_month):
        raise TradewebFetchError(
            '文件名说最新月是 %s，工作簿最后一个月却是 %s —— 文件与文件名不同步'
            % (mon, max(rows_by_month)))
    if found['headline_vol_tn'] is None or found['headline_adv_tn'] is None:
        print('⚠ 索引页抬头没解析出数字（措辞多半改了），本次跳过数值交叉核对；'
              '文件名与月份一致性检查已通过')
        return
    r = rows_by_month[mon]
    dv, da = found['headline_decimals']
    pairs = [('月成交额', r['tradeweb_volume_total_usd_tn'], found['headline_vol_tn'], dv),
             ('总 ADV', r['tradeweb_adv_total_usd_bn'] / 1e3, found['headline_adv_tn'], da)]
    for label, got, want, dec in pairs:
        tol = 0.5 * (10 ** -dec) + 1e-9
        if abs(got - want) > tol:
            raise TradewebFetchError(
                '%s 的%s对不上：工作簿 %.4f tn，索引页自述 %s tn（容差 ±%.4f）'
                % (mon, label, got, want, tol))


# ── 对外接口 ────────────────────────────────────────────────────────────────
def fetch_rows(cache_dir=DEFAULT_CACHE):
    """抓 + 解析 + 自检，返回 [{'month': 'YYYY-MM', …}, …]，按月份升序。

    每行的键就是 COLUMNS（除 month 外全是浮点，单位写在列名里）。
    任何一步不对劲一律抛 TradewebFetchError —— 宁可这一路整体缺席，
    也不产出一份看起来完整、其实错了一列的 CSV。
    """
    openpyxl = _openpyxl()
    found = discover(cache_dir)
    path = download_workbook(found, cache_dir)

    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    for name in (ADV_SHEET, VOLUME_SHEET, DAYS_SHEET):
        if name not in wb.sheetnames:
            raise TradewebFetchError(
                '工作簿缺 sheet %r，现有：%s' % (name, wb.sheetnames))

    adv_sh, vol_sh, day_sh = wb[ADV_SHEET], wb[VOLUME_SHEET], wb[DAYS_SHEET]
    months = _months_of(adv_sh)
    for sh in (vol_sh, day_sh):
        if [m for m, _ in _months_of(sh)] != [m for m, _ in months]:
            raise TradewebFetchError('三张 sheet 的月份表头不一致（%r 与 %r）'
                                     % (adv_sh.title, sh.title))

    merged = {mon: {} for mon, _ in months}
    for sh, mapping in ((adv_sh, ADV_MAP), (vol_sh, VOLUME_MAP), (day_sh, DAYS_MAP)):
        for mon, rec in _pull(sh, mapping, months).items():
            merged[mon].update(rec)

    order = [m for m, _ in months]
    _validate(merged, order)
    _crosscheck_headline(found, merged)

    rows = []
    for mon in order:
        rec = merged[mon]
        missing = [c for c in COLUMNS if c != 'month' and c not in rec]
        if missing:
            raise TradewebFetchError('%s 缺列：%s' % (mon, missing))
        row = {'month': mon}
        row.update({c: rec[c] for c in COLUMNS if c != 'month'})
        rows.append(row)
    return rows


def release_date(cache_dir=DEFAULT_CACHE):
    """官方最新一期的 (数据月 'YYYY-MM', 发布日 'YYYY-MM-DD', 出处说明)；发布日定不下来时 None。

    发布日主源是工作簿的 HTTP Last-Modified，已与工作簿内部 docProps/core.xml 的
    dcterms:modified、以及「发布日必须落在数据月末之后的窗口」两道交叉校验过
    （见 _resolve_published）。构建日 / 文件 mtime 一律不算数（CONTRACT.md §1）。
    """
    found = discover(cache_dir)
    if found['published'] is None:
        # 如实给 None，不硬凑一个 —— fetch/lseg.py 的 release_dates() 本来就按
        # 「对不上就记 None」处理（那个 docstring 里对另外三路是同一句话）。
        return None
    return (found['month'], found['published'].isoformat(),
            'monthly-activity-reports 页工作簿 HTTP Last-Modified %s，'
            '经工作簿内部 dcterms:modified 交叉校验'
            % found['published'].strftime('%Y-%m-%d'))


def latest_month(cache_dir=DEFAULT_CACHE):
    """官方源当前最新月 'YYYY-MM'。抓不到一律抛异常，不返回 None 掩盖故障。"""
    return discover(cache_dir)['month']


# ── CSV 落盘 ────────────────────────────────────────────────────────────────
def _read_csv(path):
    if not os.path.exists(path):
        return list(COLUMNS), []
    with open(path, newline='', encoding='utf-8') as f:
        rows = list(csv.reader(f))
    if not rows:
        return list(COLUMNS), []
    header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
    unknown = [c for c in COLUMNS if c not in header]
    if unknown:
        raise TradewebFetchError('series/lseg_part_%s.csv 里没有这些列：%s' % (PART, unknown))
    return header, body


def write_csv(series_dir=DEFAULT_SERIES, cache_dir=DEFAULT_CACHE, rows=None):
    """把 fetch_rows() 的结果写进 series/lseg_part_tradeweb.csv，返回新增月份（升序）。

    幂等保证（与 enx / db1 同规矩）：
      · 已有月份不重复追加；
      · **已经有值的单元格永不覆盖** —— 官方会悄悄重述历史（口径坑 4），
        自动吞重述等于在序列里插一个没人知道的台阶；
      · 冲突写 cache/lseg_tradeweb_restatements.csv 供人工判断；
      · 什么都没变时，未被触碰的单元格是原样字符串搬运，文件字节级不变。
    """
    if rows is None:
        rows = fetch_rows(cache_dir)
    path = os.path.join(series_dir, 'lseg_part_%s.csv' % PART)
    header, body = _read_csv(path)
    idx = {name: i for i, name in enumerate(header)}
    have = {r[0]: r for r in body}
    conflicts, today = [], datetime.date.today().isoformat()
    added = []

    for rec in rows:
        mon = rec['month']
        # 缺列一律失败（README 铁律 2）。fetch_rows() 已经查过一遍，这里再查是因为
        # write_csv(rows=…) 是公开入口，调用方可能手工构造 rows；漏一列若不拦，
        # 下面的 rec.items() 循环会安静地把那一格留空 —— 正是「静默写 NaN」。
        missing = [c for c in COLUMNS if c != 'month' and c not in rec]
        if missing:
            raise TradewebFetchError('%s 缺列：%s —— 拒绝写入残缺行' % (mon, missing))
        if mon not in have:
            row = [''] * len(header)
            row[idx['month']] = mon
            have[mon] = row
            body.append(row)
            added.append(mon)
        row = have[mon]
        for name, v in rec.items():
            if name == 'month' or name not in idx:
                continue
            cur = row[idx[name]].strip()
            if not cur:
                row[idx[name]] = _fmt(v)
                continue
            old = _num(cur)
            if old is None or abs(old - v) > max(1e-9, 1e-9 * abs(old)):
                conflicts.append([mon, name, cur, _fmt(v), 'tw-historical-adv', today])

    body.sort(key=lambda r: r[idx['month']])
    os.makedirs(series_dir, exist_ok=True)
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(body)

    if conflicts:
        cp = _cache(cache_dir, 'lseg_tradeweb_restatements.csv')
        new = not os.path.exists(cp)
        with open(cp, 'a', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            if new:
                w.writerow(['month', 'column', 'in_csv', 'from_source', 'source', 'seen_on'])
            w.writerows(conflicts)
        print('⚠ %d 处与已入库值冲突（官方重述），已记 %s，CSV 未覆盖' % (len(conflicts), cp))
    return sorted(added)


def main():
    rows = fetch_rows()
    added = write_csv(rows=rows)
    rd = release_date()
    print('lseg_tradeweb: %d 个月 %s → %s，新增 %d 个月'
          % (len(rows), rows[0]['month'], rows[-1]['month'], len(added)))
    if rd is None:
        print('最新一期：数据月 %s，发布日**未知**（上面的 ⚠ 护栏失效说明了原因）'
              % rows[-1]['month'])
    else:
        print('最新一期：数据月 %s，发布日 %s（%s）' % rd)
    last = rows[-1]
    print('自检 %s：总成交额 %.3f tn，总 ADV %.1f bn，加权交易日 %.2f'
          % (last['month'], last['tradeweb_volume_total_usd_tn'],
             last['tradeweb_adv_total_usd_bn'], last['tradeweb_trading_days_blended']))


if __name__ == '__main__':
    main()
