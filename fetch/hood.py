# -*- coding: utf-8 -*-
"""Robinhood Markets (HOOD) 月度经营指标 —— 无人值守抓取。

═══ 源 ═══
索引页  https://investors.robinhood.com/financials/monthly-metrics/default.aspx
文件    https://investors.robinhood.com/static-files/<uuid>   （uuid 每月新生成，
        没有可推导的命名规律，所以**必须先抓索引页再解析出链接**，不能硬编码 URL）

HOOD 的月度经营数据**不走 8-K**，是按 Reg FD 挂在 IR 站上的 —— 盯 EDGAR 抓不到，
这也是本模块不去 EDGAR 找替代源的原因（那里根本没有）。

═══ 发布节奏 ═══
平常月：次月 9–13 日。一次发四种格式（Press Release PDF / Metrics PDF / Excel /
Dashboard PDF），本模块只吃 Excel。实测 Last-Modified：
    2025-10 → 11/12 · 2025-11 → 12/10 · 2026-02 → 03/12 · 2026-04 → 05/13 · 2026-05 → 06/09

**季末月（3/6/9/12 月）没有独立的 Monthly Metrics Excel**，那一格挂的是当季
「Qx'YY Earnings Supplement.xlsx」，随财报一起发，所以晚半个月到一个月：
    2026-03 → 04/28（Q1 财报日） · 2026-06 → 07/29（Q2 财报日）
连带效应：**季末月之后那一个月也会被拖后**（2026-01 的 Excel 迟至 02/19 才挂，因为
排在 Q4'25 财报之后）。所以「本月 15 号还没有上月数据」不一定是抓取坏了。
update() 对此不做特殊处理 —— 拿不到就是还没发，返回空列表，不报错。

═══ 反爬：这是本模块最大的坑 ═══
investors.robinhood.com → robinhood.gcs-web.com → Akamai。Akamai 按 **TLS 指纹**（JA3）
挑客户端，不是按 UA：
  · curl / python urllib / requests / httpx  → **连上但永远不返回**（30s 超时，不是 403，
    不是 429，所以看日志会误以为是网络问题）。换 UA、换 HTTP/1.1、加全套 header 都没用。
  · 换成 Chrome 的 TLS 指纹就 200。
所以通道按这个顺序试（都不需要浏览器登录态、不需要点验证码，可无人值守）：
  1. curl_cffi（impersonate='chrome'），pip 装得到，首选；
  2. /usr/bin/nscurl —— macOS 自带，走 NSURLSession/Apple TLS，零依赖，实测同样 200，
     且下下来的 xlsx 与通道 1 逐字节一致（md5 相同）；
  3. node 的原生 fetch（undici）—— 实测也能过。
静态文件不经 CDN 跳转（Content-Type 直接是 xlsx，无 302），所以**没有 q4cdn / S3 直链
可绕**，别再去找了。

═══ 口径坑 ═══
1. **两种 Excel 布局，不能用同一套行号。**
   · Earnings Supplement：多 sheet（Quarterly GAAP P&L / Quarterly Balance Sheet /
     Quarterly KPIs / Monthly KPIs / Definitions），月度表在 'Monthly KPIs'，
     窗口是**滚动 39 个月**（= 本季起始月往前推三年）：Q2'26 那份从 2023-04 起、
     Q1'26 那份从 2023-01 起。**不是全历史** —— 早年月份要去 Quarterly Results 页翻
     当年的 Supplement（Q1'23 那份就覆盖 2021-01~2023-03），那是历史回填脚本
     build/basefill/hood_2021.py 的活，本模块不碰。
   · 独立 Monthly Metrics：单 sheet 'Monthly Metrics'，只有**滚动 13 个月**，
     且指标行整体下移 3 行。
   build/extract_hood.py 是按行号硬取的（因为标签重复：'Equity ($B)' 在总量和日均各出现
   一次，'Robinhood App'/'Bitstamp' 也各两次），只对 Earnings Supplement 有效。
   本模块改成 **(章节标题, 行标签) 两级定位**，两种布局通吃，官方插一行也不会错位。
2. 标签带脚注角标，两种写法都有：'Cash Sweep6, 7'（普通数字）和 'Cash Sweep⁶ ⁷'
   （Unicode 上标）。匹配前必须两种都剥掉。
3. 季度收入（hood_q.csv）**只有 Earnings Supplement 才有**。独立月度文件没有 P&L 页，
   所以平常月只能更新 hood.csv，季度那份要等财报。
4. 官方会**重述**历史月：口径变更（Bitstamp 2025-06 并入、High-Yield Cash 2026-02 改版、
   WonderFi 2026-06 并入）之后，早期月份的数字会在新文件里被改写。本模块**只追加、
   不改写已有行**（幂等要求），重述会被 verify_tail() 抓出来报差异，由人决定是否回填。
5. 季末月的数据源是 Earnings Supplement，独立月度文件根本没有那一格 —— 因此
   latest_month() 判断「最新月」时不能只认文件名里带 'Monthly Metrics' 的。
"""
import email.utils
import importlib.util
import os
import re
import subprocess
import sys
import unicodedata

import openpyxl
import pandas as pd

INDEX_URL = 'https://investors.robinhood.com/financials/monthly-metrics/default.aspx'
BASE = 'https://investors.robinhood.com'
HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
MON = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# ── 月度表：(章节, 行标签) → 列名。章节是必须的，光凭标签会拿错行。──
MONTHLY_SPEC = {
    ('funded customer growth', 'funded customers'): 'funded_customers_mn',
    ('asset growth', 'total platform assets'): 'total_platform_assets_usdbn',
    ('asset growth', 'net deposits'): 'net_deposits_usdbn',
    ('trading', 'equities and options trading days'): 'eqopt_trading_days',
    ('trading', 'crypto and prediction markets trading days'): 'crypto_trading_days',
    ('total trading volumes', 'equity ($b)'): 'vol_equity_usdbn',
    ('total trading volumes', 'options contracts (m)'): 'vol_options_mn',
    ('total trading volumes', 'crypto ($b)'): 'vol_crypto_usdbn',
    ('total trading volumes', 'robinhood app'): 'vol_crypto_app_usdbn',
    ('total trading volumes', 'bitstamp'): 'vol_crypto_bitstamp_usdbn',
    ('total trading volumes', 'event contracts (b)'): 'vol_event_bn',
    ('average daily trading volumes', 'equity ($b)'): 'adv_equity_usdbn',
    ('average daily trading volumes', 'options contracts (m)'): 'adv_options_mn',
    ('average daily trading volumes', 'crypto ($m)'): 'adv_crypto_usdmn',
    ('average daily trading volumes', 'robinhood app'): 'adv_crypto_app_usdmn',
    ('average daily trading volumes', 'bitstamp'): 'adv_crypto_bitstamp_usdmn',
    ('average daily trading volumes', 'event contracts (m)'): 'adv_event_mn',
    ('daily average trades (dats) (m)', 'equity'): 'dats_equity_mn',
    ('daily average trades (dats) (m)', 'options'): 'dats_options_mn',
    ('daily average trades (dats) (m)', 'crypto'): 'dats_crypto_mn',
    ('interest earning assets ($b)', 'margin book'): 'margin_book_usdbn',
    ('interest earning assets ($b)', 'cash and deposits'): 'cash_and_deposits_usdbn',
    ('interest earning assets ($b)', 'cash sweep'): 'cash_sweep_usdbn',
    ('securities lending ($m)', 'total securities lending revenue'): 'seclend_total_usdmn',
    ('securities lending ($m)', 'securities lending, net'): 'seclend_net_usdmn',
}

# ── 纯改名的别名。官方对**同一行**用过两种措辞，新旧文件混用：
#   · 章节「Interest Earning Assets ($B)」在 Earnings Supplement 里叫
#     「Customer Margin and Cash Sweep ($B)」（2024 年前的还带 Balances 后缀，
#     故按前缀匹配）—— 两边都是 Margin Book / Cash and Deposits / Cash Sweep 三行；
#   · 「Total Trading Volumes」下的两条加密拆分行在 Supplement 里带单位后缀
#     （'Robinhood App ($B)' / 'Bitstamp ($B)'），日均那一节则不带。
# 别名与本名指向同一个列名。parse_sheet 按**列名**判缺，不按 key 判，所以两套写法
# 同时挂在 spec 里不会互相判成「缺列」。
MONTHLY_ALIASES = {
    ('total trading volumes', 'robinhood app ($b)'): 'vol_crypto_app_usdbn',
    ('total trading volumes', 'bitstamp ($b)'): 'vol_crypto_bitstamp_usdbn',
    ('customer margin and cash sweep', 'margin book'): 'margin_book_usdbn',
    ('customer margin and cash sweep', 'cash and deposits'): 'cash_and_deposits_usdbn',
    ('customer margin and cash sweep', 'cash sweep'): 'cash_sweep_usdbn',
}
MONTHLY_SPEC.update(MONTHLY_ALIASES)

# ⚠ **DARTs 故意不做别名**，虽然它看着只是改名。Q2'26 Earnings Supplement（2026-07-29）
# 之前，那一节叫「Daily Average Revenue Trades (DARTs) (M)」，之后改叫
# 「Daily Average Trades (DATs) (M)」，而且**数字跟着变了**：同一个 2025-01，
# Q1'26 文件印 2.6，Q2'26 文件印 3.3（+27%）；crypto 同样（0.7 → 1.1 之类）。
# 只有 2025-01 起被重述，2023-04~2024-12 两种口径逐月完全相同 —— 也就是说
# DATs = DARTs + 不产生收入的交易，而那部分 2025 年才变得可观。
# 若把 DARTs 挂成 dats_* 的别名，某个月官方万一又发一份老版式文件，update() 就会把
# 窄口径的数悄悄写进宽口径的列，一列里混两把尺子且没人会发现。
# 历史回填要用老文件的 DARTs 是另一回事：那走 build/basefill/hood_2021.py，
# 由人一次性判断并写在图注里，不进无人值守链路。

# 季度 P&L：这张表标签不重复，'revenues:' 一个章节盖全部
PL_SPEC = {
    ('revenues:', 'options'): 'rev_options_usdmn',
    ('revenues:', 'cryptocurrencies'): 'rev_crypto_usdmn',
    ('revenues:', 'equities'): 'rev_equities_usdmn',
    ('revenues:', 'event contracts'): 'rev_event_usdmn',
    ('revenues:', 'other'): 'rev_other_txn_usdmn',
    ('revenues:', 'transaction-based revenues'): 'rev_transaction_usdmn',
    ('revenues:', 'net interest revenues'): 'rev_net_interest_usdmn',
    ('revenues:', 'other revenues'): 'rev_other_usdmn',
    ('revenues:', 'total net revenues'): 'rev_total_usdmn',
}

# 季度 KPI：只取成交量。余额类在季度表里是「季末月」值不是季度合计，混用会错。
QVOL_SPEC = {
    ('total trading volumes', 'equity ($b)'): 'q_vol_equity_usdbn',
    ('total trading volumes', 'options contracts (m)'): 'q_vol_options_mn',
    ('total trading volumes', 'crypto ($b)'): 'q_vol_crypto_usdbn',
    ('total trading volumes', 'event contracts (b)'): 'q_vol_event_bn',
}

SECTIONS = sorted({s for s, _ in MONTHLY_SPEC} | {s for s, _ in PL_SPEC}
                  | {s for s, _ in QVOL_SPEC}, key=len, reverse=True)


def _section_of(lab):
    """章节标题带单位后缀且各文件不完全一致（'Funded Customer Growth (M)' /
    'Daily Average Trades (DATs) (M)'），所以用前缀匹配、长的优先。"""
    return next((s for s in SECTIONS if lab.startswith(s)), None)


# ═══════════════════════ 下载：三条无人值守通道 ═══════════════════════

def _via_curl_cffi(url):
    from curl_cffi import requests as cr          # 延迟 import：没装也能走通道 2
    r = cr.get(url, impersonate='chrome', timeout=120,
               headers={'Accept-Language': 'en-US,en;q=0.9'})
    r.raise_for_status()
    return r.content


def _via_nscurl(url):
    """macOS 自带 nscurl 走 NSURLSession，TLS 指纹是 Apple 的，Akamai 放行。
    零第三方依赖，是 curl_cffi 装不上时的兜底。"""
    import tempfile
    fd, tmp = tempfile.mkstemp(suffix='.bin')
    os.close(fd)
    try:
        subprocess.run(['/usr/bin/nscurl', '--output', tmp, url],
                       check=True, capture_output=True, timeout=180)
        with open(tmp, 'rb') as f:
            b = f.read()
        if not b:
            raise RuntimeError('nscurl 返回空文件')
        return b
    finally:
        os.path.exists(tmp) and os.unlink(tmp)


def fetch_bytes(url):
    errs = []
    for fn in (_via_curl_cffi, _via_nscurl):
        try:
            b = fn(url)
            if b:
                return b
        except Exception as e:                     # noqa: BLE001 —— 通道失败要继续试下一条
            errs.append(f'{fn.__name__}: {type(e).__name__}: {e}')
    raise RuntimeError(
        'HOOD IR 三条通道全部失败（Akamai 按 TLS 指纹拦截，普通 urllib/curl 必定超时）。\n  '
        + '\n  '.join(errs))


# ═══════════════════════ 发布日：HTTP Last-Modified ═══════════════════════
# 本文件开头「发布节奏」里那一串实测日期就是这么量出来的。HOOD 的官方文件**没有任何
# 自述发布日**：xlsx 内嵌的 dcterms 是 Workiva 的作者时间戳（作者是公司财务，不是挂网
# 动作），索引页上只有 Jan…Dec 的锚文本，没有日期；月度数据又不走 8-K，EDGAR 上没有
# 申报日可用。静态文件的 Last-Modified 是唯一能观测到的「官方把它挂上去」的时刻。
# 日历日一律取 header 原样的 GMT 日，不折算美东 —— 与上面那串实测值同口径。
# （两者只在 00:00–05:00 GMT 之间上传时差一天，2026-01 那期就是这种情形。）

def _lm_via_curl_cffi(url):
    from curl_cffi import requests as cr
    r = cr.head(url, impersonate='chrome', timeout=60,
                headers={'Accept-Language': 'en-US,en;q=0.9'})
    r.raise_for_status()
    return r.headers.get('last-modified')


def _lm_via_nscurl(url):
    out = subprocess.run(['/usr/bin/nscurl', '-I', '-D', '-', '-o', os.devnull, url],
                         check=True, capture_output=True, timeout=120, text=True).stdout
    m = re.search(r'(?im)^last-modified:\s*(.+)$', out)
    return m.group(1).strip() if m else None


def last_modified(url):
    """→ (原始 header 字符串, 'YYYY-MM-DD')；任一通道都拿不到就 (None, None)。

    通道顺序与 fetch_bytes 一致（HEAD 一样过 Akamai）。取不到不抛异常 —— 这条信息
    只决定页面抬头上的半句话，不该把已经落盘的月度数据连累成一次失败的摄入。
    """
    for fn in (_lm_via_curl_cffi, _lm_via_nscurl):
        try:
            raw = fn(url)
        except Exception:                          # noqa: BLE001 —— 换下一条通道
            continue
        if not raw:
            continue
        try:
            return raw, email.utils.parsedate_to_datetime(raw).strftime('%Y-%m-%d')
        except (TypeError, ValueError):
            continue
    return None, None


_SD = None


def _source_dates():
    """加载仓库根的 source_dates.py。不能裸 import：本模块被 monthly_run.py 用
    spec_from_file_location 加载，那时 sys.path 上既没有 fetch/ 也没有仓库根。
    与 fetch/cost.py 的 _release() 同规则。"""
    global _SD
    if _SD is None:
        p = os.path.join(ROOT, 'source_dates.py')
        spec = importlib.util.spec_from_file_location('source_dates', p)
        _SD = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(_SD)
    return _SD


def _record_source_dates(series_dir, files, added):
    """把新增月份的发布日记进 series/source_dates.csv。

    只对**索引页上属于那个月的那一格**取 Last-Modified：一个月的数据不可能在该月结束前
    挂出来，所以那一格的文件就是这个月数据第一次公开的载体。不能拿 update() 实际下载的
    文件顶替 —— Earnings Supplement 带全历史，用它的日期去盖前面几个月，会把 07/29
    安到 04 月的数据上。
    """
    by_period = {str(e['period']): e for e in files}
    for mon in added:
        e = by_period.get(str(mon))
        if not e:
            continue
        raw, day = last_modified(e['url'])
        if not day:
            print(f'  [source_date] {mon}: 拿不到 Last-Modified，跳过（页面抬头会省掉那半句）')
            continue
        try:
            _source_dates().record(
                series_dir, 'hood', mon, day,
                f'IR 静态文件「{e["title"]}」的 HTTP Last-Modified: {raw}（取 GMT 日历日）')
        except Exception as ex:                    # noqa: BLE001 —— 数据已落盘，不能回滚
            print(f'  [source_date] {mon}: 写台账失败 {type(ex).__name__}: {ex}')


# ═══════════════════════ 索引页解析 ═══════════════════════

# 索引页快照。**只留一份、每次覆盖**（不像 msci 那样按日期各存一份）：这个页面
# 125KB 且天天抓，按日期存一年就是四十几 MB 的散文件，会把 tools/prune_cache.py
# 报告末尾「未登记的散文件」那个阈值顶穿。留这一份只有一个用途 —— 下面三道护栏
# 抛异常之后，人能立刻打开当时那份 HTML 看官方到底改成了什么样，而不是对着一句
# 报错去猜。（这道护栏写出来之前，仓里一份索引页快照都没有。）
INDEX_SNAPSHOT = 'hood_index_latest.html'

# 锚点按「先切出整个 <a>，再分别取属性」两步来，不把 href 与 title 写死在同一条
# 正则里 —— 写死就意味着官方把两个属性调个个儿、或者某一格漏了 title，整条链接
# 直接消失，而消失是没有声音的。
_A_RE = re.compile(r'<a\s([^>]*?)>(.*?)</a>', re.S | re.I)
_ATTR_HREF = re.compile(r'href\s*=\s*"([^"]*)"', re.I)
_ATTR_TITLE = re.compile(r'title\s*=\s*"([^"]*)"', re.I)
# 容错版年份标题：只要 <h2> 的**可读文本**是四位年就算数，容得下
# <h2><span>2027</span></h2> 这种嵌套改法。生产解析仍走下面那条按位置切分的
# re.split（不动它），这条只服务护栏③。
_H2_RE = re.compile(r'<h2[^>]*>(.*?)</h2>', re.S | re.I)
_MON_IDX = {m.lower(): i + 1 for i, m in enumerate(MON)}


def _text_of(frag):
    """标签内文本 → 可读字符串：先剥嵌套标签，再解实体，再压空白。

    三步缺一不可，分别对应索引页三种改法：锚文本被包成 <b>Jul</b>（原来的
    `[^<]*` 整条匹配失败，那一格丢掉）、写成 `&nbsp;Jul`（strip() 看到的是原样
    实体串，label[:3] 会变成 '&nb'）、以及换行缩进混进锚文本。
    """
    return re.sub(r'\s+', ' ', _unescape(re.sub(r'<[^>]+>', ' ', frag))).strip()


def _month_of_label(label):
    """锚文本 → 月序号 1..12，认不出返回 None。

    大小写不敏感，也认全称（'January'[:3] == 'Jan'）。官方今天写的是 'Jan'，
    但这十二格是人工逐月维护的，容错的成本是零、漏一格的代价是一个月的静默停更。
    """
    return _MON_IDX.get(label[:3].lower())


def _year_groups(html):
    """→ [(年份, 该年份组的原始 HTML)]，供护栏③ 用的**容错**版年份切分。

    边界取「下一个 <h2> 开始处」而不是「下一个年份 <h2>」，免得最后一个年份组
    把页脚（'Sign up for email alerts' 那一段）也吞进来。
    """
    marks = []
    for m in _H2_RE.finditer(html):
        y = re.fullmatch(r'(20\d\d)', _text_of(m.group(1)))
        marks.append((int(y.group(1)) if y else None, m.start(), m.end()))
    out = []
    for i, (year, _s, e) in enumerate(marks):
        if year is None:
            continue
        end = marks[i + 1][1] if i + 1 < len(marks) else len(html)
        out.append((year, html[e:end]))
    return out


def _save_index_snapshot(cache_dir, html):
    """存一份索引页原文。写失败只打印、不抛 —— 与 last_modified() 同一条规矩：
    快照是给事后取证用的，不该把一次本来能成功的摄入拖成失败。"""
    if not cache_dir:
        return
    try:
        os.makedirs(cache_dir, exist_ok=True)
        with open(os.path.join(cache_dir, INDEX_SNAPSHOT), 'w', encoding='utf-8') as f:
            f.write(html)
    except Exception as ex:                        # noqa: BLE001 —— 快照失败不算故障
        print(f'  [hood] 索引页快照没写成 {type(ex).__name__}: {ex}（不影响本次抓取）')


def _snap_hint(cache_dir):
    return f'（本次快照 cache/{INDEX_SNAPSHOT}）' if cache_dir else ''


def list_excel_files(cache_dir=None):
    """→ [{'period': Period('2026-06','M'), 'title': ..., 'url': ...}]，按月份升序。

    页面结构：<h2 class="h4">2026</h2> 起一个年份组，组里四个 <h3> 小节
    （Press Release / PDF / Excel / Dashboard），每个小节 12 个 <li> 月份格，
    锚文本是 Jan…Dec；**还没发的月份是空格子，压根没有 <a>**。
    年份只出现在组标题上，文件名里不一定有（'Q2'26 Earnings Supplement' 只有两位年），
    所以**年份必须从组标题取，不能从文件名猜**。

    ═══ 三道护栏，防的是同一件事：某个月的链接在这里被静默丢掉 ═══
    这十二格是官方**每月人工新增**的（一次加四条：PR / PDF / Excel / Dashboard），
    所以最可能坏的恰恰是最新那一格，而且坏了不会有任何声音：out 里还留着几十条旧
    链接，末尾 `if not out` 那道地板判据够不着；回到 update() 之后 want 为空、干净
    返回 []，monthly_run 报 NOCHANGE。**连续十天这样，和连续十天「官方还没发」，
    在日志里一模一样。** 三道分别是：
      ① 容错匹配（_A_RE + _text_of + _month_of_label）—— 属性顺序变了、缺 title、
         锚文本里套了标签、前面粘个 &nbsp;、月份写成大写或全称，原来这五种写法
         每一种都会让整条链接凭空消失，现在都认。它只是把匹配面放宽，永远只会比
         从前多认、不会少认。
      ② 未落库锚点对账 —— Excel 小节里凡是指向 /static-files/ 却解析不出月份的
         锚点，一律抛。这是**唯一够得着最新那一格**的网：update() 那道反向哨兵
         只检查已入库的月份，按定义碰不到还没入库的新月（msci.parse() 里 `dropped`
         那一段讲的是同一件事）。
      ③ 年份组覆盖 —— 页面登记了某个年份、组里明明挂着 .xlsx，却一条都没解析
         出来就抛。它管的是 ①② 都够不着的坏法：<h2>/<h3> 结构变了导致整个年份组
         根本没进循环（比如年份被包成 <h2><span>2027</span></h2>）。
    三道的盲区不重叠，这也是 msci 那边坚持解析侧与入库侧两张网都留着的理由。
    """
    html = fetch_bytes(INDEX_URL).decode('utf-8', 'replace')
    _save_index_snapshot(cache_dir, html)
    out = []
    unread = []            # Excel 小节里「是静态文件链接、却没解析出月份」的锚点
    seen_years = set()     # 真解析出条目的年份，供护栏③ 比对
    # 按年份组切分；每组内再切出各小节
    for grp in re.split(r'<h2[^>]*>(?=\s*20\d\d\s*<)', html)[1:]:
        ym = re.match(r'\s*(20\d\d)\s*<', grp)
        if not ym:
            continue
        year = int(ym.group(1))
        for sec in re.split(r'<h3 class="result-title">', grp)[1:]:
            head = re.sub(r'<[^>]+>', '', sec.split('</h3>')[0]).strip()
            if 'excel' not in head.lower():
                continue
            for attrs, inner in _A_RE.findall(sec):
                h = _ATTR_HREF.search(attrs)
                # 不指向静态文件的锚点（小节里的导航链接之类）本来就不是数据，
                # 安静跳过 —— 护栏② 只对「确实是个文件、却读不出是哪个月」发作。
                if not h or not h.group(1).startswith('/static-files/'):
                    continue
                href = h.group(1)
                label = _text_of(inner)
                mon = _month_of_label(label)
                if mon is None:
                    unread.append((year, href, label))
                    continue
                t = _ATTR_TITLE.search(attrs)
                # 缺 title 不再让整条链接消失：title 只拿去起 cache 文件名和记
                # source_dates 台账，退化成锚文本或 uuid 一样唯一，不值得为它丢一个月。
                out.append({'period': pd.Period(f'{year}-{mon:02d}', 'M'),
                            'title': _unescape(t.group(1)) if t
                                     else (label or href.rsplit('/', 1)[-1]),
                            'url': BASE + href})
                seen_years.add(year)

    # ── 护栏②：未落库锚点对账 ──
    # 先按 href 去重。同一个文件在同一格里被渲染两遍（文字链 + 图标链，IR 模板换
    # 个皮肤就可能这样）时，图标那条剥完标签是空字符串、解析不出月份 —— 那不是漏
    # 月，是同一个月的重复渲染。不去这个重，一次纯样式改动就能让 HOOD 天天 FAIL，
    # 而这道网真正要抓的坏法（最新那一格的 href 谁也没认领）一个都不会被它放过。
    known = {e['url'][len(BASE):] for e in out}
    unread = [u for u in unread if u[1] not in known]
    if unread:
        raise RuntimeError(
            f'Excel 小节里有 {len(unread)} 个静态文件链接解析不出月份：'
            f'{[(y, lab) for y, _h, lab in unread[:5]]!r} —— 锚文本写法多半改了'
            f'（原来是纯 "Jan"…"Dec"）。宁可整次失败也不静默漏月：这十二格是人工'
            f'维护的，坏的往往正是最新那个月，而漏掉它不产生任何 FAIL。'
            f'请人工打开 {INDEX_URL}{_snap_hint(cache_dir)} 核对写法，再改 _A_RE '
            f'或 _month_of_label；官方若真在这一节挂了非月份的 Excel（Definitions、'
            f'全年汇总、打包下载），在这里加白名单，别把这道对账删了')

    # ── 护栏③：年份组覆盖 ──
    # 判据刻意带「该组里确实挂着 .xlsx」这个前提，两处误伤都是它挡下的：每年年初
    # 页面先建出空的新年份组（十二个格子全空）时不许 FAIL；四种格式虽说同批发布，
    # PR/PDF/Dashboard 先挂上、Excel 晚几小时的窗口也是真实存在的，所以不能只看
    # 组里有没有 /static-files/。**别把这个前提简化掉。**
    for year, body in _year_groups(html):
        if year in seen_years or '.xlsx' not in body.lower():
            continue
        raise RuntimeError(
            f'索引页上有 {year} 年份组、组里挂着 .xlsx，却一条 Excel 链接都没解析'
            f'出来 —— <h2>/<h3> 的结构多半改了，整个年份组没进循环。别的年份照常'
            f'解析，所以这一次不会有任何其它症状：请人工打开 '
            f'{INDEX_URL}{_snap_hint(cache_dir)} 核对年份组与小节标题')

    if not out:
        raise RuntimeError('索引页解析不出任何 Excel 链接 —— 页面结构可能改了，'
                           '先人工打开 ' + INDEX_URL + ' 核对')
    out.sort(key=lambda d: d['period'])
    return out


def _unescape(s):
    import html as _h
    return _h.unescape(s)


def latest_month(cache_dir=None):
    """官方源当前最新月 "YYYY-MM"。抓不到直接抛异常（不返回 None 假装没更新）。"""
    return str(list_excel_files(cache_dir)[-1]['period'])


# ═══════════════════════ Excel 解析 ═══════════════════════

def _norm(v):
    """标签归一化：剥脚注角标（'Cash Sweep6, 7' 与 'Cash Sweep⁶ ⁷' 两种写法都有）、
    换行、多余空白，再转小写。"""
    if v is None:
        return ''
    s = unicodedata.normalize('NFKC', str(v))     # ⁶ → 6
    s = s.replace('\n', ' ').replace('\xa0', ' ')
    s = re.sub(r'[\d,\s*]+$', '', s.strip())      # 尾部脚注数字/逗号/空格
    return re.sub(r'\s+', ' ', s).strip().lower()


def _grid(ws):
    W = ws.max_column
    return [list(r) + [None] * (W - len(r)) for r in ws.iter_rows(values_only=True)], W


def _periods(rows, W, kind):
    """扫出期次列。表头行号在各 sheet 不一致（月度 3 或 6，季度 P&L 4），所以扫不写死。"""
    want = MON if kind == 'M' else ['Q1', 'Q2', 'Q3', 'Q4']
    pr = None
    for ri in range(min(10, len(rows))):
        hits = sum(1 for i in range(W)
                   if rows[ri][i] and str(rows[ri][i]).strip()[:3] in want)
        if hits >= 4:
            pr = ri
            break
    if pr is None:
        raise RuntimeError(f'找不到 {kind} 表头行')
    yr, out, cur = rows[pr - 1], [], None
    for i in range(W):
        v = yr[i]
        if v is not None and str(v).replace(',', '').strip().isdigit():
            cur = int(str(v).replace(',', '').strip())
        s = str(rows[pr][i]).strip()[:3] if rows[pr][i] else ''
        if kind == 'M' and s in MON:
            out.append((i, pd.Period(f'{cur}-{MON.index(s) + 1:02d}', 'M')))
        elif kind == 'Q' and s in ('Q1', 'Q2', 'Q3', 'Q4'):
            out.append((i, pd.Period(f'{cur}Q{s[1]}', 'Q')))
    if not out:
        raise RuntimeError(f'{kind} 表头行找到了但一个期次都没解析出来')
    return out


def parse_sheet(ws, spec, kind):
    """按 (章节, 标签) 两级定位取数 —— 不用行号，两种布局通吃。"""
    rows, W = _grid(ws)
    cols = _periods(rows, W, kind)
    data_cols = [i for i, _ in cols]
    hit, section = {}, None
    for r in rows:
        lab = _norm(next((c for c in r[:3] if c is not None and str(c).strip()), None))
        if not lab:
            continue
        empty = all(r[i] is None or not isinstance(r[i], (int, float)) for i in data_cols)
        if empty:
            section = _section_of(lab) or section   # 章节标题行本身没有数值
            continue
        col = spec.get((section, lab))
        # 按**列名**去重（不是按 key）：同一个列可以有多个 (章节, 标签) 写法挂在 spec 上
        # （见 MONTHLY_ALIASES），命中哪一个都算这一列有了。仍然只认第一次，防脚注区重名。
        if col and col not in hit:
            hit[col] = [r[i] if isinstance(r[i], (int, float)) else None for i in data_cols]
    want = list(dict.fromkeys(spec.values()))      # 列序固定，跟 CSV 一致
    missing = [c for c in want if c not in hit]
    if missing:
        raise RuntimeError(f'{ws.title}: 这些行没找到 → {missing}；'
                           f'官方大概率改了标签或章节名，先人工看一眼 Excel')
    return pd.DataFrame({c: hit[c] for c in want}, index=[p for _, p in cols])


def parse_workbook(path):
    """→ (月度 DataFrame, 季度 DataFrame 或 None)。季度那份只有 Earnings Supplement 才有。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    msheet = next((s for s in wb.sheetnames
                   if s.strip().lower() in ('monthly kpis', 'monthly metrics')), None)
    if msheet is None:
        raise RuntimeError(f'{os.path.basename(path)}: 找不到月度表，sheet 有 {wb.sheetnames}')
    m = parse_sheet(wb[msheet], MONTHLY_SPEC, 'M')

    pl = next((s for s in wb.sheetnames if s.startswith('Quarterly GAAP')), None)
    if pl is None:
        return m, None
    q = parse_sheet(wb[pl], PL_SPEC, 'Q')
    k = next(s for s in wb.sheetnames if s.strip().lower() == 'quarterly kpis')
    q = q.join(parse_sheet(wb[k], QVOL_SPEC, 'Q'), how='outer')
    return m, q


# ═══════════════════════ 落盘 / 追加 ═══════════════════════

def _safe_name(title, period):
    slug = re.sub(r'[^A-Za-z0-9]+', '_', title).strip('_')[:60]
    return f'hood_{period}_{slug}.xlsx'


def download(entry, cache_dir):
    os.makedirs(cache_dir, exist_ok=True)
    path = os.path.join(cache_dir, _safe_name(entry['title'], entry['period']))
    if not os.path.exists(path) or os.path.getsize(path) < 10000:
        with open(path, 'wb') as f:
            f.write(fetch_bytes(entry['url']))
    return path


def _fmt_like(existing, val):
    """按该列已有单元格的写法格式化新值 —— 保证 CSV 里新旧行长得一样。
    （直接 pandas to_csv 会把整数列写成 30.0，跟历史行不一致，diff 全是噪声。）

    但**格式一致优先级低于不丢精度**：整数列若哪天官方给了小数（比如 adv_crypto_usdmn
    一直是整数，某月报 543.5），就照实写，不四舍五入 —— 宁可格式不齐也不改数。
    """
    v = float(val)
    nonempty = [s for s in existing if s]
    if all(re.fullmatch(r'-?\d+', s) for s in nonempty) and v.is_integer():
        return str(int(v))
    seen = max([len(s.split('.')[1]) for s in existing if '.' in s] or [1])
    txt = repr(v)
    own = len(txt.split('.')[1]) if '.' in txt and 'e' not in txt else 0
    return f'{v:.{max(seen, own)}f}'


def _append(csv_path, new_rows):
    """把 new_rows（{期次: {列: 值}}）追加进 CSV。已有的期次一律跳过 —— 幂等靠这里。
    解析结果缺任何一个已有列就抛异常，绝不静默写 NaN。"""
    old = pd.read_csv(csv_path, dtype=str).fillna('')
    cols = list(old.columns)
    have = set(old[cols[0]])
    added = []
    lines = []
    for per in sorted(new_rows):
        if str(per) in have:
            continue
        rec = new_rows[per]
        miss = [c for c in cols[1:] if c not in rec or rec[c] is None
                or (isinstance(rec[c], float) and pd.isna(rec[c]))]
        if miss:
            raise RuntimeError(f'{os.path.basename(csv_path)} {per}: 解析结果缺列 {miss}；'
                               f'宁可不写也不写 NaN —— 先人工看一眼官方 Excel 是不是改版了')
        cells = [str(per)] + [_fmt_like(list(old[c]), rec[c]) for c in cols[1:]]
        lines.append(','.join(cells))
        added.append(str(per))
    if lines:
        with open(csv_path, 'a', encoding='utf-8') as f:
            f.write('\n'.join(lines) + '\n')
    return added


def update(series_dir, cache_dir):
    """把新月份写进 series/hood.csv（和有财报时的 hood_q.csv），返回新增月份列表。

    只处理**比 CSV 末期更新**的期次，不回填中间空洞。
    series 现在从 2021-01 起（2026-08-19 由 build/basefill/hood_2021.py 回填），
    而回填段有一半的列是空的 —— 老版式的官方表根本没有 ADV / Cash and Deposits /
    Bitstamp / Event contracts 那几行。若把「CSV 里有空格」当成缺口让 update() 去补，
    它会一路往回翻 2023 年的老 Excel，而那些文件里这些行不存在，解析必然报缺列；
    真要它别报错就得放宽「缺列即失败」，那等于把无人值守链路唯一的安全网拆了。
    历史回填是一次性的、有人看着的活，走 build/basefill/hood_2021.py，别让日常任务去碰。
    """
    files = list_excel_files(cache_dir)
    mcsv = os.path.join(series_dir, 'hood.csv')
    qcsv = os.path.join(series_dir, 'hood_q.csv')
    last_m = pd.Period(pd.read_csv(mcsv, dtype=str).iloc[-1, 0], 'M')
    last_q = pd.Period(pd.read_csv(qcsv, dtype=str).iloc[-1, 0], 'Q')

    # ── 反向哨兵：已入库的最后一个月，必须在索引页上还找得到 ──
    # 和 list_excel_files() 那三道方向相反，补的正是它们的盲区：那三道从「页面上
    # 有什么」查，这一道从「我们已知该有什么」反查，连整页换模板、链接都在但一个
    # 月份都对不上号这种坏法也兜得住。它不看 want，所以「官方还没发」时绝不会响。
    #
    # **只查 last_m 这一个月，别扩成 msci 那样查最近六个月。** 索引页是按年份组
    # 累加的档案，但它只回溯到 2023 年（2026-08-26 实测：页面只登记 2023/2024/
    # 2025/2026 四组），而 series/hood.csv 从 2021-01 起（build/basefill/hood_2021.py
    # 一次性回填的）。往前多查几个月，就可能查到官方本来就没挂的年份上去，把一个
    # 健康的抓取变成天天 FAIL。last_m 永远只有一两个月大，任何合理的滚动窗口都还
    # 留着它 —— 这个「窄」是设计，不是偷懒。
    if last_m not in {e['period'] for e in files}:
        raise RuntimeError(
            f'{last_m} 在 series/hood.csv 里，索引页却解析不出这一格 —— 官方不删'
            f'历史月（那页是按年份累加的档案），所以多半是我们把整段解析丢了。'
            f'本次不写入，请人工打开 {INDEX_URL}{_snap_hint(cache_dir)} 确认')

    want = [e['period'] for e in files if e['period'] > last_m]
    if not want:
        # **空 want 是这个源大部分日子的正常状态，这里绝不能抛。**
        # 平常月次月 9–13 日才发，季末月要等财报（2026-06 的数据 07/29 才挂），
        # 之后那个月还被连带拖后 —— 一个月里有二十几天在这里干净返回 [] 是设计，
        # 不是故障（见文件头「发布节奏」）。「悄悄不更新」那一类由上面的反向哨兵和
        # list_excel_files() 的三道护栏拦，它们都不看 want，所以既拦得住漏月，
        # 也不会误伤「还没发」。
        return []

    # 取**能覆盖缺口的最少文件**：Earnings Supplement 带全历史，独立月度带滚动 13 个月，
    # 通常最新那一个就够。从新往旧翻，凑齐就停；翻到比 last_m 还老的文件就没必要继续。
    rows_m, rows_q = {}, {}
    for e in reversed(files):
        if set(want) <= set(rows_m):
            break
        if e['period'] <= last_m:
            break
        path = download(e, cache_dir)
        m, q = parse_workbook(path)
        for per, rec in m.iterrows():
            if per > last_m:
                rows_m.setdefault(per, rec.to_dict())
        if q is not None:
            for per, rec in q.iterrows():
                if per > last_q:
                    rows_q.setdefault(per, rec.to_dict())

    missing = sorted(set(want) - set(rows_m))
    if missing:
        raise RuntimeError(f'索引页说有 {missing}，但翻遍相关 Excel 都没解析出这些月 —— '
                           f'多半是官方那一格挂错了文件，人工看一眼 {INDEX_URL}')

    added = _append(mcsv, rows_m)
    if rows_q:
        _append(qcsv, rows_q)
    # 记发布日放在这里而不是下载之后：_append 跳过已有期次，added 才是「这一趟真的写进
    # series 的月份」。挂在下载后会给那些根本没入库的月份也记上一笔。
    _record_source_dates(series_dir, files, added)
    return added


# ═══════════════════════ 对账（跑得起来的自检，不是装饰）═══════════════════════

def verify_tail(series_dir, cache_dir, n=3):
    """用解析器重算 series CSV 最后 n 个期次，逐列比对，打印最大偏差。"""
    files = list_excel_files(cache_dir)
    path = download(files[-1], cache_dir)
    m, q = parse_workbook(path)
    for csv_name, df, freq in [('hood.csv', m, 'M'), ('hood_q.csv', q, 'Q')]:
        if df is None:
            print(f'{csv_name}: 最新文件不含该表，跳过')
            continue
        old = pd.read_csv(os.path.join(series_dir, csv_name), index_col=0)
        old.index = pd.PeriodIndex(old.index, freq=freq)
        worst, ncell, nbad = (None, None, 0.0), 0, 0
        print(f'\n=== {csv_name} 最后 {n} 期逐列对账（源：{os.path.basename(path)}）===')
        for per in old.index[-n:]:
            if per not in df.index:
                print(f'  {per}: 官方文件里没有这一期'); nbad += 1; continue
            for c in old.columns:
                a, b = float(old.loc[per, c]), df.loc[per, c]
                ncell += 1
                if b is None or pd.isna(b):
                    print(f'  {per} {c}: CSV={a} 官方=空'); nbad += 1; continue
                d = abs(a - float(b))
                rel = d / abs(a) * 100 if a else (0.0 if d == 0 else float('inf'))
                if d > 1e-9:
                    print(f'  {per} {c}: CSV={a} 官方={b} 差 {d:g}({rel:.2f}%)')
                    nbad += 1
                if rel > worst[2]:
                    worst = (per, c, rel)
        print(f'  比对 {len(old.index[-n:])} 期 x {len(old.columns)} 列 = {ncell} 个单元格，'
              f'不一致 {nbad} 个，最大相对偏差 {worst[2]:.4f}%'
              + (f' @ {worst[0]} {worst[1]}' if worst[0] else ''))


if __name__ == '__main__':
    R = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    S, C = os.path.join(R, 'series'), os.path.join(R, 'cache')
    if len(sys.argv) > 1 and sys.argv[1] == 'verify':
        verify_tail(S, C)
    elif len(sys.argv) > 1 and sys.argv[1] == 'update':
        print('新增月份:', update(S, C))
    else:
        print('官方最新月:', latest_month(C))
