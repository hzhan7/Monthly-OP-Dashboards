# -*- coding: utf-8 -*-
r"""创意电子（GUC，3443.TW）月度营收 —— 无人值守抓取模块。

对应 build/specs/guc.py（页面由通用底座 build/single.py 生成），维护一个序列文件：

  series/guc.csv    month, revenue_ntd_mn, revenue_turnkey_ntd_mn, revenue_nre_other_ntd_mn

────────────────────────────────────────────────────────────────────────
数据源
────────────────────────────────────────────────────────────────────────
1) 营收（guc.csv）—— 官方 xlsx，一份文件带全 2017-01 至今
   落地页 https://www.guc-asic.com/en/investor/financial?financial-tab=option2
   页面上挂一个「Historical Monthly revenue」按钮，指向
   /upload/<上传日>/8_<随机串>.xlsx。单 sheet `revenue breakdown`，单位 NT$K，
   逐年一块，每块 = 月份表头行（YYYYMM）+ Turnkey / NRE / Others / Total 四行。

   **URL 每月随新数据换一次**（路径含上传日，与当月营收公告同日），所以必须每次从
   落地页现抓，写死的那一刻就注定了下个月拿到的是上个月的文件。

   ⚠️ 这份 xlsx 的**最老年份块就是 2017**（实测 2026-08-18 那份：`A1:O84`，
      2017 块在 row 63，其下 row 68-84 全空）。所以它是**分部拆分**的硬边界，
      不是「解析漏了」。合并营收本身早得多 —— 见源 1b。

1b) 合并营收的历史回补（只补 `revenue_ntd_mn` 一列）—— MOPS 月营收月档
   https://mopsov.twse.com.tw/server-java/FileDownLoad
     POST step=9 & functionName=show_file2 & filePath=/t21/sii/
        & fileName=t21sc03_<民国年>_<月>.csv
   全市场一份 CSV（~160KB），单位新台币千元，`公司代號` = 3443 那一行的
   `營業收入-當月營收` 就是当期法定申报的**合并**月营收原值。

   **为什么走这条路而不是手工贴数**：xlsx 到不了 2017 以前，而窗口要拉到 2016-01。
   MOPS 是同一个数的另一个官方通道，可复算、可无人值守重跑；手工贴的数下个月没人
   能复算（同 fetch/mtk.py 那一轮的处置）。

   **口径不是「差不多」，是逐格对齐的**（2026-08-18 实测）：
     · 2017-01 … 2017-12 十二个月，MOPS `當月營收` 与本序列（出自 xlsx）**逐位相等**，
       所以这两条通道量的是同一个东西，拼在一起不会在 2016/2017 交界处产生台阶；
     · 2016 年十二个月加总 = 9,290,421 千元 = MOPS `ajax_t163sb04`（TYPEK=sii,
       year=105, season=04）里 3443 的**年度查核营业收入**，逐位相等；
       2017 年同法 = 12,160,606，与年报（build/mrspecs/guc.py `_REF` 的 AR17）相同。
     · 每个 106_M 档的 `營業收入-去年當月營收` = 我们从 105_M 档取到的当月值，12/12 相等
       —— 这是**两份不同文件**互证，能抓到单月错位（年度加总抓不到互换两个月这种错）。
   这三条都写成了 `update()` 里的护栏，不是一次性的人工核对（见 `_mops_backfill`）。

   MOPS 的档案地板实测在 **2013-01**（`t21sc03_102_1.csv` 有档、`101_12` 只有 320 字节
   的「查無資料」壳页）。本模块的 `START_MONTH` 停在 **2016-01** 而不是停在源的边界：
   全站页面窗口统一自 2016-01 起，再往前接只会让「合并列比分部列长」的那一段
   （见口径坑 8）白白多出三年，收益是零。要往前接改 `START_MONTH` 即可，
   `_mops_backfill()` 会自己把缺的月份补齐 —— 但先把 2013-2015 的年度查核数补进
   `_ANNUAL_AUDITED`，否则那三年就没有加总护栏了。

2) 交叉校验源（只读不写）
   TWSE OpenAPI https://openapi.twse.com.tw/v1/opendata/t187ap05_L
   全市场当期一份 JSON，单位新台币千元。只有最新一期、没有历史，所以只用来验证
   「官方最新月是哪个月 + 金额对不对」。3443 是**本国公司**，在 t187ap05_L 里；
   世芯-KY 3661、矽力-KY 6415 那类外国发行人不在这张表、走 TPEx 的 _O 端点。

3) 公告日（source_dates.csv 的 guc 行）
   新闻稿列表页 https://www.guc-asic.com/en/news/PressRelease
   正文首句是电头：「Hsinchu, Taiwan, Aug 5, 2026 - GUC (TAIEX: 3443) today
   announced its net sales for July 2026 were NT5,769 million...」

────────────────────────────────────────────────────────────────────────
发布节奏
────────────────────────────────────────────────────────────────────────
· 台湾《证券交易法》要求次月 10 日前公告，但 GUC 的实际惯例是**次月 5 日 14:00**，
  而且公司在 IR 财务日历（同一份 HTML，见口径坑 5）里**逐条预告**下一次的日期。
  实测预告日与新闻稿电头日逐条相等：
    2025-12→01/05  2026-01→02/05  2026-03→04/07  2026-04→05/05
    2026-05→06/05  2026-06→07/06  2026-07→08/05
  → 调度：roster LAG 取 (7, 7)，覆盖撞假日顺延的最坏情形（2026-03 那期撞清明落到 4/7）。
  → 但**不要拿「次月 5 日」外推公告日**，照样每次去新闻稿现读 —— 顺延不规律。

────────────────────────────────────────────────────────────────────────
口径坑（踩过的，别再踩）
────────────────────────────────────────────────────────────────────────
1. **Total 行 115 格里 61 格是活公式，不许读它**。xlsx 的 Total 行在 2018~2022
   整年是 `=SUM(B23:B25)` 这种公式串；openpyxl 默认 `data_only=False` 会静默读到
   字符串，`data_only=True` 读到的是 **Excel 写入的缓存值** —— 而缓存是发布者保存
   工具的副产品，上游某月改用 LibreOffice 或脚本另存就没了，整块变 None 静默漏年。
   → 本模块一律**对分项求和**（Turnkey + NRE/Others）。实测 115 个月分项和与 Total
     无一例外相等，所以这不是近似，是绕开一个会静默失效的依赖。

2. **2026-01 起官方把 NRE 与 Others 合并**：年份块从
   `Turnkey / NRE / Others / Total` 四行变成 `Turnkey / NRE & Others / Total` 三行
   （IR deck 里同一口径又叫 'NRE & IP'）。Total 不受影响，但拆分列有断点。
   → 落库统一成两列，2017-2025 的 NRE + Others 在这里就合并掉，序列口径逐月连续。

3. **xlsx 链接是相对路径**。href 写作 `/upload/2026_08_05/8_...xlsx`，不是绝对 URL。
   用 `https?://[^"']+\.xlsx` 这种正则会一个都抓不到，然后误判成「页面是 JS 渲染的」。
   → 必须 urljoin。（实测踩过。）

4. **新闻稿 slug 不可拼，且软 404 是 302 不是 404**。
   · slug 命名法至少 7 种：`202607revenue` / `202604revenueE` / `202601revenuE`
     （官网自己拼错）/ `202505` / `202502E-20250305` / `GUCMonthlySalesReportinJune2025`…
   · 不存在的 slug **302 跳到 /en**，跟随重定向后是 172,215 字节的通用壳页，
     HTTP 状态码 200 —— 按状态码校验完全无效，会把壳页当文章解析出错日期。
   · 壳页里**也有 'Hsinchu'**（页脚地址），所以 'Hsinchu' 没有判别力，
     真正有判别力的是正文的 'net sales'。
   → 本模块从列表页抓 href，且 `_no_redirect_opener` 关掉重定向：3xx 直接判失败。

5. **`financial-tab=option2` 与 `option7` 返回完全同一份 HTML**（353,794 字节，
   tab 是纯前端切换）。一次请求同时拿到最新 xlsx 链接与下次公告日历，
   不必两次请求，也不存在两个页面不同步的问题。

6. **重述**：xlsx 是全量覆盖式文件，而且 URL 每月一换（比 TSMC 的替换频率还高）。
   一旦上游改历史，本模块会检测到与已入库值不一致并**抛异常**，
   而不是悄悄改写或追加 —— 由人判断是口径变更还是解析出错。同 fetch/tsm.py 口径坑 6。

7. guc-asic.com **没有 WAF**，裸 urllib 直接 200，不像 investor.tsmc.com 认 UA 指纹。

8. **2016 年那 12 行只有合并列，两个分部列是空的 —— 这是事实不是缺陷，别去填它。**
   月度 Turnkey / NRE & Others 拆分只有 IR 那份 xlsx 有，而它自 2017-01 起（源 1 的 ⚠️）；
   MOPS 只申报一个合并总额，没有分部。所以 2016-01…2016-12 的
   `revenue_turnkey_ntd_mn` / `revenue_nre_other_ntd_mn` 写**空串**：
   · 不许补 0 —— 0 会被下游当成「那个月这块业务真的是零」，堆叠柱与占比线都会照画；
   · 不许按 2017 的比例摊 —— 那是造数；
   · 不许拿 Total 减一个猜出来的数 —— 同上。
   下游（build/mrspecs/guc.py + build/mrbase.py）认得空格：占比图由 `mrwin.resolve()`
   自动右移到 2017-01，全历史图的两条分部线在 2016 那段是断笔。

9. **MOPS 有两种「失败」长得不一样，别混：**
   · **限流**：HTTP 200 + 564 字节的 `Overrun` 页 —— 退避重试（同 fetch/nanya.py 口径坑 1）；
   · **无档**：HTTP 200 + 320 字节的「查無資料」壳页 —— 那个月本来就没有，重试一万次也没有。
   另外 mopsov.twse.com.tw 会**在 TLS 握手阶段直接 reset**（实测第一次连接常见
   `ConnectionResetError(54)`，重试就好）。三者都在 `_mops_get()` 里分开处理。

10. **回补分支平时一次网络请求都不发。** `update()` 只对「`START_MONTH` 起、xlsx 首月
   之前、且库里没有」的月份去取 MOPS —— 回补做完之后这个集合恒为空。
   所以每月 cron 的耗时与从前一样，回补能力只在真的缺月时才醒过来。
"""

from __future__ import annotations

import csv
import datetime
import io
import json
import os
import re
import ssl
import time
import urllib.error
import urllib.parse
import urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)

ORIGIN = 'https://www.guc-asic.com'
IR_PAGE = ORIGIN + '/en/investor/financial?financial-tab=option2'
PR_LIST = ORIGIN + '/en/news/PressRelease'
TWSE_API = 'https://openapi.twse.com.tw/v1/opendata/t187ap05_L'
TWSE_CODE = '3443'
MOPS_DL = 'https://mopsov.twse.com.tw/server-java/FileDownLoad'

# 序列起点。**改了会生效**：低于 xlsx 首月的那一段由 `_mops_backfill()` 从 MOPS 月档
# 自动补齐（源 1b）。往前挪之前先补 `_ANNUAL_AUDITED`，否则新那几年没有加总护栏。
START_MONTH = '2016-01'
COLUMNS = ['month', 'revenue_ntd_mn', 'revenue_turnkey_ntd_mn', 'revenue_nre_other_ntd_mn']

_UA = ('Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 '
       '(KHTML, like Gecko) Chrome/126.0 Safari/537.36')
_CTX = ssl.create_default_context()

# 真页 >100KB；壳页 172KB 但没有 'net sales'；WAF/错误页远小于此
_MIN_HTML = 50_000
_MIN_XLSX = 5_000
# MOPS 月档实测 129KB ~ 204KB；限流页 564 字节、无档壳页 320 字节（口径坑 9）
_MIN_CSV = 100_000
_MOPS_BACKOFF = 20          # 秒；命中限流页后的退避
_MOPS_GAP = 2.0             # 秒；顺序取档之间的间隔，**别并发**（并发必被限流）

# 年度**查核**合并营业收入（新台币千元）。用途只有一个：给回补进来的整年做加总护栏。
# 出处：MOPS `ajax_t163sb04`（TYPEK=sii, season=04, year=民国年）里 3443 的营业收入栏，
#   2016（民 105）9,290,421 · 2017（民 106）12,160,606。
#   2017 那个数与 2017 年度年报的经营结果表相同（build/mrspecs/guc.py `_REF` 的 AR17
#   逐位相等），两条独立通道互校过。
# **这不是入库数据**，一格都不写进 CSV，只在 `update()` 末尾拿来核账。
# ⚠️ START_MONTH 往前挪时**必须**同步补行，否则新那几年只有逐月互证、没有加总护栏。
_ANNUAL_AUDITED = {
    2016: 9_290_421,
    2017: 12_160_606,
}


class GucFetchError(RuntimeError):
    """本模块的故障出口。抓不到 / 认不出来一律抛它，不返回 None 掩盖故障。"""


def _get(url, *, follow=True, tries=3, min_bytes=0):
    """取 URL。follow=False 时 3xx 直接判失败（见口径坑 4）。"""
    if follow:
        opener = urllib.request.build_opener()
    else:
        class _NoRedirect(urllib.request.HTTPRedirectHandler):
            def redirect_request(self, req, fp, code, msg, headers, newurl):
                raise GucFetchError(f'{url} 重定向到 {newurl}（{code}）—— 该资源不存在')
        opener = urllib.request.build_opener(_NoRedirect)
    last = None
    for attempt in range(tries):
        try:
            req = urllib.request.Request(url, headers={'User-Agent': _UA})
            body = opener.open(req, timeout=120).read()
            if len(body) < min_bytes:
                raise GucFetchError(f'{url} 只有 {len(body)} 字节（<{min_bytes}），疑似壳页')
            return body
        except GucFetchError:
            raise
        except Exception as exc:                                  # noqa: BLE001
            last = exc
    raise GucFetchError(f'{url} 取不到：{last!r}')


def _ir_html():
    return _get(IR_PAGE, min_bytes=_MIN_HTML).decode('utf-8', 'ignore')


def _xlsx_url(html):
    """从落地页 HTML 取最新 xlsx 的绝对 URL。相对路径必须 urljoin（口径坑 3）。"""
    hits = re.findall(r'["\']([^"\']*\.xlsx)["\']', html)
    if not hits:
        raise GucFetchError('落地页里找不到 .xlsx 链接（页面改版？）')
    return urllib.parse.urljoin(ORIGIN, hits[0])


# 年份块表头格 → YYYYMM。**判据是形态不是单元格类型**：原来写成
# `isinstance(row[1], (int, float))`，于是发布方的导出工具哪天把某一格存成文本
# （手工在新列里敲 '202609' 就够）或存成日期，那一格就从 months 里掉出去；
# 数据行仍按 `[1:1+len(months)]` **从左切片**，前面逐格对齐、只丢表尾一格 ——
# 最新月干干净净地消失，`if not out: raise` 因为历史块还在而永远不触发。
# 口径坑 1 早就在警告这份文件对导出工具的类型敏感，这里把三种写法都认下来。
_YM_RE = re.compile(r'^(20\d{2})\s*[-/.年]?\s*(0?[1-9]|1[0-2])\s*月?$')
# 第一格被写上标签时的表头门槛（见 _parse_xlsx 里的说明）。
_HDR_MIN_LABELLED = 2


def _ym(cell):
    """单元格 → YYYYMM 整数；不像月份表头就返回 None。int/float、文本、日期都认。

    年份上界卡在 2099、月份卡在 1–12，为的是别把营收数字当表头：
    2026-08 那份实测全表零个数据格落进这个形态（Turnkey 2019-02 的 294,205「年」
    是 2942，NRE 2018-01 的 209,149「月」是 49，都出局）。
    """
    if cell is None or isinstance(cell, bool):
        return None
    if isinstance(cell, (datetime.datetime, datetime.date)):
        v = cell.year * 100 + cell.month
    elif isinstance(cell, (int, float)):
        if float(cell) != int(cell):
            return None
        v = int(cell)
    else:
        m = _YM_RE.match(str(cell).strip())
        if not m:
            return None
        v = int(m.group(1)) * 100 + int(m.group(2))
    return v if 200001 <= v <= 209912 and 1 <= v % 100 <= 12 else None


def _parse_xlsx(blob):
    """解析 revenue breakdown sheet，返回 {'YYYY-MM': (turnkey, nre_other)}，单位 NT$K。

    只对分项求和，永不读 Total 行（口径坑 1）。

    表头识别与落库对账是**两件事，缺一不可**（照 fetch/msci.py parse() 的两件套）：
    `_ym()` 只认识已经见过的三种写法，挡不住下一种没见过的；函数末尾那道
    `dropped` 对账才是挡住它的那一道 —— 凡「这一行已经被认定是年份块表头，
    却有 YYYYMM 形态的格子没被收进 months」，一律抛，宁可整次失败也不静默错位。
    """
    try:
        import openpyxl
    except ImportError as exc:                                     # pragma: no cover
        raise GucFetchError('需要 openpyxl 才能解析 GUC 的 xlsx') from exc
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
    if 'revenue breakdown' not in wb.sheetnames:
        raise GucFetchError(f'xlsx 里没有 revenue breakdown sheet：{wb.sheetnames}')
    rows = [list(r) for r in wb['revenue breakdown'].iter_rows(values_only=True)]

    out = {}
    dropped = []              # 已认定是表头、却没读全的行；循环末尾一起抛
    for i, row in enumerate(rows):
        ym_cells = [_ym(c) for c in row[1:]]
        n_ym = sum(1 for v in ym_cells if v is not None)
        # 年份块表头行的判据：从 B 列起是一串 YYYYMM。两条门槛不一样，理由不同：
        # · 第一格空（历来的形态）→ **1 格就够**。每年 1 月新起的那一块只有一列
        #   （2 月跑的时候当年块就是 `空 | 202601`），门槛写成「≥3 格」会在每年
        #   2、3 月把整个当年块丢掉 —— 那正是这道护栏要防的病，不能自己再造一次。
        # · 第一格被写上标签（2026-01 那次把当年块从四行改成三行，证明当年块
        #   确实会被单独重排）→ 门槛提到 2 格，免得某个恰好落在 20xxxx 形态的
        #   营收数字把一行数据当成表头。
        if n_ym < (1 if row[0] is None else _HDR_MIN_LABELLED):
            continue
        # 表头是从 B 列起**左对齐、连续**的一段。左边有洞（B 列那格认不出）或
        # 右边还有漏网的月份（中间某格认不出），都会让数据行整体错位一格 ——
        # 实测这不止能静默漏最新月，还能把上个月的数值当新月写进 CSV。
        lead = 0
        while lead < len(ym_cells) and ym_cells[lead] is None:
            lead += 1
        months, j = [], lead
        while j < len(ym_cells) and ym_cells[j] is not None:
            months.append(ym_cells[j])
            j += 1
        tail = [v for v in ym_cells[j:] if v is not None]
        if lead or tail:
            dropped.append(
                f'sheet 第 {i + 1} 行：认出 {months}，'
                f'但左边空了 {lead} 格 / 右边还剩 {tail}')
            continue
        block = {}
        for j in range(i + 1, min(i + 6, len(rows))):
            label = str(rows[j][0] or '').strip()
            if not label:
                continue
            block[label] = list(rows[j][1:1 + len(months)])
            if label.lower().startswith('total'):
                break
        turnkey = block.get('Turnkey')
        if turnkey is None:
            raise GucFetchError(f'年份块 {months[0]} 里没有 Turnkey 行')
        merged = block.get('NRE & Others')        # 2026 起
        nre, others = block.get('NRE'), block.get('Others')        # 2025 及以前
        if merged is None and nre is None:
            raise GucFetchError(f'年份块 {months[0]} 里既没有 NRE 也没有 NRE & Others')
        for k, ym in enumerate(months):
            t = turnkey[k]
            if merged is not None:
                n = merged[k]
            else:
                n = (nre[k] if nre else 0) + (others[k] if others else 0)
            if t is None or n is None:
                continue
            out[f'{ym // 100}-{ym % 100:02d}'] = (float(t), float(n))

    # ── 落库对账：挡住「表头写法又变了 → 整块错位 / 静默漏月」的那道闸 ──
    # 为什么不能只留上面那个 `_ym()`：正则只认识已经见过的变体（数字 / 文本 /
    # 日期）。下一次上游换个写法，认不出来的那一格照样会让右边整体错位，而错位
    # 不产生 FAIL、不产生断档（断档只看已入库月份之间的洞，管不到表尾少一行），
    # 红点与连续失败计数全都抓不到。**下一次靠的是这一段，不是那条正则。**
    if dropped:
        raise GucFetchError(
            f'年份块表头有 {len(dropped)} 行没读全：{dropped[:3]!r} —— '
            '多半是发布方的导出工具把某一格的 YYYYMM 存成了别的写法（口径坑 1）。'
            '数据行是从左切片的，少认一格就整体错位，宁可整次失败也不静默写错数；'
            '本次不写入，请人工打开 revenue breakdown sheet 确认表头写法。')
    if not out:
        raise GucFetchError('xlsx 解析出 0 个月（版式变了？）')
    return out


# ── 独立于解析器的月份判据（照 fetch/ice.py 的 _crosscheck_workbook_month）──
_UPLOAD_DATE_RE = re.compile(r'/upload/(\d{4})_(\d{2})_(\d{2})/')


def _crosscheck_upload_month(url, newest):
    """xlsx 的**上传日**反推出来的最新数据月 vs 实际解析出来的最新月，前者更新就炸。

    这是本模块唯一一条**独立于解析器**的月份判据。防的是这一类：表头写法变了、
    最新那一列被静默丢掉，`out` 里历史块还在所以 `if not out: raise` 不触发，
    update() 找不到新月份于是干干净净报 NOCHANGE —— 没有 FAIL、没有红点、
    连续失败计数也不动，页面就一直挂着旧数据（README「第四类：不出声的失败」）。

    为什么这一条能 raise，而末尾那条 TWSE 对账只能 warn：URL 与文件内容是
    **同一个 artifact 同批发布**的（口径坑：URL 每月随新数据换一次，路径里的
    上传日就是当月营收公告日），不存在时间差；TWSE opendata 与 IR xlsx 是两份
    分别发布的东西，公告日任一边领先几小时都正常。同 fetch/ice.py 里
    _crosscheck_workbook_month 与 _crosscheck_release_month 的分工。

    方向是**单向**的：解析结果比上传日反推的还旧 → 抛（这就是漏月的样子）；
    反过来比它还新 → 只喊一声，那多半是官方换了上传节奏或命名，不是漏月。
    命名认不出来时同样只喊不抛 —— 官方改个命名不该让一个本来好好的源停摆，
    但必须让人看见护栏掉了。

    已知的唯一误伤形态，写在这里免得下一个人重新推一遍：GUC 若在**次月**重传一份
    数据还停在上上个月的旧文件（上传日跨了月、内容没跟着走），这条会开火。
    源 1 记的节奏是「URL 每月随新数据换一次、与当月营收公告同日」，两者同批走，
    所以这个形态至今没出现过；真出现了也确实该有人看一眼。
    """
    m = _UPLOAD_DATE_RE.search(url)
    if not m:
        print('[guc] ⚠ 护栏失效：xlsx URL 里认不出上传日（/upload/YYYY_MM_DD/ 命名变了？），'
              f'这一轮没有独立于解析器的月份判据。URL={url}')
        return
    y, mo = int(m.group(1)), int(m.group(2))
    expect = f'{y - 1}-12' if mo == 1 else f'{y}-{mo - 1:02d}'
    if newest < expect:
        raise GucFetchError(
            f'xlsx 上传日 {m.group(1)}-{m.group(2)}-{m.group(3)} 反推最新数据月应是 '
            f'{expect}，但解析出来的最新月是 {newest}（URL={url}）。两者同批发布、'
            '不该不一致 —— 最可能是年份块表头写法变了，最新那一列被静默丢弃。'
            '拒绝写入，请人工打开 revenue breakdown sheet 看一眼当年块的表头行。')
    if newest > expect:
        print(f'[guc] ⚠ 解析出的最新月 {newest} 比上传日反推的 {expect} 还新 —— '
              '不是漏月，但上传节奏或命名可能变了，值得看一眼。')


# ── 源 1b：MOPS 月档（只回补 revenue_ntd_mn 一列，见文件头源 1b 与口径坑 8/9）────
def _mops_get(year, month, cache_dir):
    """取某个月的全市场月营收 CSV 正文；该月**无档**返回 None。

    三种「失败」分开处理（口径坑 9）：TLS reset 重连、限流页退避、无档返回 None。
    抓到的原始档存 cache（gitignore 的目录），重跑不再打 MOPS。
    """
    roc = year - 1911
    fname = f't21sc03_{roc}_{month}.csv'
    body = None
    cpath = os.path.join(cache_dir, 'guc_' + fname) if cache_dir else None
    if cpath and os.path.exists(cpath) and os.path.getsize(cpath) >= _MIN_CSV:
        body = open(cpath, 'rb').read()
    if body is None:
        payload = urllib.parse.urlencode({
            'step': '9', 'functionName': 'show_file2',
            'filePath': '/t21/sii/', 'fileName': fname}).encode()
        for attempt in range(6):
            try:
                req = urllib.request.Request(
                    MOPS_DL, data=payload,
                    headers={'User-Agent': _UA, 'Accept': '*/*'})
                body = urllib.request.urlopen(req, timeout=180).read()
            except Exception as exc:                               # noqa: BLE001
                # mopsov 常在首次 TLS 握手时 reset（口径坑 9）。重连，不当成没档。
                print(f'[guc][warn] MOPS {fname} 连接失败（{exc!r}），重试 {attempt + 1}/6')
                time.sleep(3 + 3 * attempt)
                continue
            if len(body) >= _MIN_CSV:
                break
            head = body[:400].decode('utf-8', 'ignore')
            if 'Overrun' in head or '過於頻繁' in head:
                print(f'[guc][warn] MOPS 限流（{len(body)} 字节），'
                      f'退避 {_MOPS_BACKOFF}s 后重试 {fname}')
                time.sleep(_MOPS_BACKOFF)
                continue
            # 320 字节的「查無資料」壳页 = 这个月 MOPS 上真的没有档（地板在 2013-01）
            return None
        else:
            raise GucFetchError(f'MOPS {fname} 连续 6 次没取到正文，本次放弃')
        if cpath:
            os.makedirs(cache_dir, exist_ok=True)
            open(cpath, 'wb').write(body)
        time.sleep(_MOPS_GAP)
    try:
        text = body.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = body.decode('big5', 'ignore')
    if '公司代號' not in text:
        raise GucFetchError(f'MOPS {fname} 里没有「公司代號」表头，版式变了？')
    return text


def _mops_row(year, month, cache_dir):
    """3443 在该月档里的整行 dict；该月无档返回 None。找到了但对不上一律抛异常。"""
    text = _mops_get(year, month, cache_dir)
    if text is None:
        return None
    for row in csv.DictReader(io.StringIO(text)):
        if (row.get('公司代號') or '').strip() == TWSE_CODE:
            row = {(k or '').strip(): (v or '').strip() for k, v in row.items()}
            ym = row.get('資料年月', '')
            want = (f'{year - 1911}/{month}', f'{year - 1911}{month:02d}')
            if ym not in want:
                raise GucFetchError(
                    f'MOPS {year}-{month:02d} 档里 3443 的資料年月是 {ym!r}，'
                    f'不是 {want[0]!r} —— 取到了别的月份的档')
            return row
    raise GucFetchError(f'MOPS {year}-{month:02d} 档里没有 {TWSE_CODE} 那一行')


def _mops_backfill(have, floor_month, cache_dir):
    """回补 `START_MONTH` 起、xlsx 首月之前、库里还没有的那些月的**合并**营收。

    返回 {'YYYY-MM': 當月營收 千元}。缺月集合为空时一次网络请求都不发（口径坑 10）。

    每个月两道护栏，都是**两份不同文件**互证（单靠一份档没法发现自己取错了年份）：
      ① 次年同月档的 `營業收入-去年當月營收` 必须与本月取到的 `當月營收` 逐位相等；
      ② 次年同月如果已经在库里（出自 xlsx），那份档的 `當月營收` 也必须与库内值相等
         —— 这条同时钉死了「MOPS 与 xlsx 是同一个口径」，也就是拼接点不会有台阶。
    对不上一律抛异常、整轮不写入：本模块的规矩是让人来判，不是猜一个更像的值。
    """
    want, y, m = [], int(START_MONTH[:4]), int(START_MONTH[5:7])
    while f'{y}-{m:02d}' < floor_month:
        if f'{y}-{m:02d}' not in have:
            want.append((y, m))
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)
    if not want:
        return {}
    print(f'[guc] MOPS 回补 {len(want)} 个月（{want[0][0]}-{want[0][1]:02d} … '
          f'{want[-1][0]}-{want[-1][1]:02d}），顺序取档，每档间隔 {_MOPS_GAP}s')

    out = {}
    for y, m in want:
        row = _mops_row(y, m, cache_dir)
        if row is None:
            raise GucFetchError(
                f'MOPS 没有 {y}-{m:02d} 的月档（档案地板实测在 2013-01）—— '
                f'START_MONTH={START_MONTH} 要到这个月，但源上没有。'
                f'要么把 START_MONTH 往后挪，要么换一个能到那么早的源。')
        cur = float(row['營業收入-當月營收'])

        nxt = _mops_row(y + 1, m, cache_dir)
        if nxt is None:
            raise GucFetchError(
                f'取不到 {y + 1}-{m:02d} 的月档，无法为回补的 {y}-{m:02d} 做互证；本次不写入')
        base = (nxt.get('營業收入-去年當月營收') or '').strip()
        if not base:
            raise GucFetchError(
                f'{y + 1}-{m:02d} 档里没有「營業收入-去年當月營收」栏 —— '
                f'MOPS 版式变了，回补的互证护栏失效，本次不写入')
        if abs(float(base) - cur) > 0.5:
            raise GucFetchError(
                f'{y}-{m:02d} 的 MOPS 當月營收 {cur:,.0f} 与 {y + 1}-{m:02d} 档里的'
                f'去年當月 {float(base):,.0f} 不符（千元）—— 疑似重述或取错档，本次不写入')
        known = have.get(f'{y + 1}-{m:02d}')
        if known:
            lib = float(known[1]) * 1000.0
            nxt_cur = float(nxt['營業收入-當月營收'])
            if abs(nxt_cur - lib) > 0.5:
                raise GucFetchError(
                    f'{y + 1}-{m:02d} MOPS 當月營收 {nxt_cur:,.0f} 与库内（出自 IR xlsx）'
                    f'{lib:,.0f} 千元不符 —— 两条通道口径已经分叉，'
                    f'回补上去会在拼接点造出台阶，本次不写入')
        out[f'{y}-{m:02d}'] = cur
    return out


def _annual_check(body):
    """整年加总 vs 年度查核营业收入。对不上抛异常。只查 `_ANNUAL_AUDITED` 里列出的年。

    为什么这条护栏值得每次跑：逐月互证抓得到「取错档」，抓不到「两个月互换」；
    加总对审计数抓得到后者。两条一起才封得住。
    """
    by = {}
    for r in body:
        by.setdefault(r[0][:4], {})[r[0][5:7]] = float(r[1])
    for year, ref in sorted(_ANNUAL_AUDITED.items()):
        got = by.get(str(year)) or {}
        if len(got) != 12:                       # 整年没配齐就不报半个结论
            continue
        s = sum(got.values()) * 1000.0
        if abs(s - ref) > 0.5:
            raise GucFetchError(
                f'FY{year} 逐月加总 {s:,.0f} 与年度查核营业收入 {ref:,.0f} 千元不符'
                f'（差 {s - ref:+,.0f}）—— 出处 MOPS ajax_t163sb04（TYPEK=sii, '
                f'year={year - 1911}, season=04）。数据侧出事了，本次不写入。')


def _twse_latest():
    """TWSE OpenAPI 里 3443 的 (月份, 当月营收 NT$K)。交叉校验用，失败不阻断。"""
    try:
        blob = _get(TWSE_API, tries=2, min_bytes=1000)
        for rec in json.loads(blob.decode('utf-8', 'ignore')):
            if rec.get('公司代號') == TWSE_CODE:
                roc = str(rec['資料年月'])           # 11507 = 2026-07
                month = f'{int(roc[:-2]) + 1911}-{int(roc[-2:]):02d}'
                return month, float(rec['營業收入-當月營收'])
    except Exception as exc:                                       # noqa: BLE001
        print(f'[guc][warn] TWSE OpenAPI 交叉校验跳过：{exc!r}')
    return None, None


def latest_month(cache_dir):                                       # noqa: ARG001
    """官方源当前最新月 'YYYY-MM'。抓不到一律抛 GucFetchError。"""
    url = _xlsx_url(_ir_html())
    newest = max(_parse_xlsx(_get(url, min_bytes=_MIN_XLSX)))
    _crosscheck_upload_month(url, newest)
    return newest


def _fmt(v):
    return f'{v / 1000:.3f}'


def update(series_dir, cache_dir):
    """把新月份追加进 series/guc.csv，返回新增月份列表（升序）。

    幂等：已入库的月份不重写；既有行原样搬运，没有新月份时文件字节级不变。
    已有值永不覆盖 —— 与官方值不一致时**抛异常**（口径坑 6），由人判断。

    两条入库通道，各自管一段、互不覆盖：
      · **xlsx**（源 1）管 2017-01 起，三列都给；
      · **MOPS 月档**（源 1b）管 `START_MONTH` 起、xlsx 首月之前那一段，
        只给合并列，两个分部列留**空**（口径坑 8：不补 0、不摊、不猜）。
    xlsx 的重述体检对「xlsx 里没有这个月」的跳过是**有边界**的：只跳过
    `min(official)` 以下那一段（MOPS 补进来的 2016，源 1b / 口径坑 8 —— xlsx 里
    本来就没有，是唯一合法的缺席）。落在 xlsx 自身覆盖区之内还消失的月份一律抛，
    见下面那道哨兵。别把这个 skip 再放宽回无条件。
    """
    csv_path = os.path.join(series_dir, 'guc.csv')
    with open(csv_path, newline='', encoding='utf-8') as fh:
        rows = list(csv.reader(fh))
    header, body = rows[0], [r for r in rows[1:] if r and r[0].strip()]
    if header != COLUMNS:
        raise GucFetchError(f'series/guc.csv 列不对：{header} != {COLUMNS}')
    have = {r[0]: r for r in body}

    xlsx_url = _xlsx_url(_ir_html())
    official = _parse_xlsx(_get(xlsx_url, min_bytes=_MIN_XLSX))
    # 独立于解析器的月份判据，放在写盘之前（说明见 _crosscheck_upload_month）
    _crosscheck_upload_month(xlsx_url, max(official))

    # 重述体检：已入库月份逐格比对，不一致抛异常而不是改写
    lo, hi = min(official), max(official)
    drift, vanished = [], []
    for month, row in have.items():
        if month not in official:
            # ── 哨兵：已入库月份在 xlsx 里消失就抛（照 fetch/msci.py update() 哨兵②）──
            # 判据不是「差得多不多」，而是「这件事是源的正常行为，还是我们把行
            # 解析丢了」：这份 xlsx 是**全量覆盖式、从不删历史月**（口径坑 6），
            # 所以覆盖区之内的月份读不到，只能是解析侧出事了 —— 多半是年份块表头
            # 行没被认出来，整块（当年）静默丢失，而 `if not out: raise` 因为
            # 2017-2025 那些块还在而永远不触发。
            # 边界是 [min(official), max(official)]：低于下界的那一段是
            # _mops_backfill() 从 MOPS 补进来的 2016（源 1b / 口径坑 8），xlsx 里
            # 本来就没有，是唯一合法的缺席。**别把这个 skip 放宽回无条件。**
            # 这道网**够不着两头**，两头各有自己的网，别以为这一条包圆了：
            #   · 表尾（最新那一块整块丢了）→ min/max 会跟着缩，哨兵看不见；
            #     接它的是 _crosscheck_upload_month（上传日反推的独立判据）。
            #   · 表头（最老那一块 2017 整块丢了）→ 下界跟着上移，哨兵也看不见。
            #     这是**刻意**的：GUC 从头部裁短历史是合法输入，为它每天 FAIL 不值当；
            #     而丢掉最老一块也不会让页面变旧（那些月早已入库），代价只是少一层复核。
            if lo <= month <= hi:
                vanished.append(month)
            continue
        t, n = official[month]
        want = [month, _fmt(t + n), _fmt(t), _fmt(n)]
        if row != want:
            drift.append((month, row, want))
    if vanished:
        raise GucFetchError(
            f'{sorted(vanished)[:5]} 在 series/guc.csv 里，却在官方 xlsx 的覆盖区'
            f'（{lo} ~ {hi}）里解析不出来（共 {len(vanished)} 个月）—— '
            '这份 xlsx 是全量覆盖式、从不删历史月，所以多半是年份块表头行没被 '
            '_parse_xlsx 认出来（第一格被写上标签 / YYYYMM 变成文本或日期），'
            '整块被静默丢掉。本次不写入，请人工打开 revenue breakdown sheet 确认。')
    if drift:
        raise GucFetchError(
            'GUC 官方 xlsx 与已入库值不一致（疑似重述或解析变形），本次不写入：\n  '
            + '\n  '.join(f'{m}: 库内 {old} vs 官方 {new}' for m, old, new in drift[:5]))

    added = []

    # ── 历史回补（源 1b）：xlsx 到不了的那一段走 MOPS，只补合并列 ──────────────
    # 缺月集合为空时这一整段零请求（口径坑 10），所以每月 cron 的耗时不变。
    for month, cur in sorted(_mops_backfill(have, min(official), cache_dir).items()):
        body.append([month, _fmt(cur), '', ''])
        have[month] = body[-1]
        added.append(month)

    # ── 增量（源 1）：xlsx 里有、库里没有的月份，三列齐全 ─────────────────────
    for month in sorted(official):
        if month < START_MONTH or month in have:
            continue
        t, n = official[month]
        body.append([month, _fmt(t + n), _fmt(t), _fmt(n)])
        have[month] = body[-1]
        added.append(month)

    if added:
        body.sort(key=lambda r: r[0])
        # 写盘前最后一道：整年加总 vs 年度查核营业收入（`_ANNUAL_AUDITED`）。
        # 逐月互证抓「取错档」，这条抓「两个月互换」，两条一起才封得住。
        _annual_check(body)
        with open(csv_path, 'w', newline='', encoding='utf-8') as fh:
            w = csv.writer(fh)
            w.writerow(header)
            w.writerows(body)
        added.sort()

    # 交叉校验：官方最新月与 TWSE OpenAPI 对账（只告警，不阻断）
    tw_month, tw_val = _twse_latest()
    if tw_month:
        newest = max(official)
        if tw_month != newest:
            print(f'[guc][warn] TWSE 最新月 {tw_month} 与 IR xlsx 最新月 {newest} 不一致')
        elif tw_val is not None:
            t, n = official[newest]
            if abs((t + n) - tw_val) > 1.0:
                print(f'[guc][warn] {newest} IR {t + n:,.0f} vs TWSE {tw_val:,.0f} NT$K 不符')

    return added
